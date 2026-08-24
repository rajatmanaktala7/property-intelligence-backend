import os, re, io, json, uuid, hashlib, html
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Form, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/whatsapp-capture/database", tags=["Alliance WhatsApp Group Master Database"])

DATABASE_URL = (
    os.getenv("WHATSAPP_DATABASE_URL")
    or os.getenv("DATABASE_URL")
    or ""
).strip()

def _db_url(url: str) -> str:
    if url.startswith("postgres://"):
        return url.replace("postgres://", "postgresql+psycopg://", 1)
    if url.startswith("postgresql://"):
        return url.replace("postgresql://", "postgresql+psycopg://", 1)
    return url

engine = create_engine(_db_url(DATABASE_URL), pool_pre_ping=True, pool_recycle=300) if DATABASE_URL else None

SCHEMA = """
CREATE TABLE IF NOT EXISTS alliance_master_listings(
    id BIGSERIAL PRIMARY KEY,
    listing_id TEXT UNIQUE NOT NULL,
    source_message_id TEXT,
    source_group TEXT,
    source_account TEXT,
    message_date DATE,
    message_time TEXT,
    sender_phone TEXT,
    transaction TEXT,
    property_type TEXT,
    location TEXT,
    building_project TEXT,
    budget_inr NUMERIC(16,2),
    budget_text TEXT,
    area_sqft NUMERIC(14,2),
    area_text TEXT,
    contact_numbers TEXT,
    poster_name TEXT,
    verified TEXT DEFAULT 'Unverified',
    raw_summary TEXT NOT NULL,
    fingerprint TEXT UNIQUE,
    record_status TEXT DEFAULT 'ACTIVE',
    created_at TIMESTAMPTZ DEFAULT NOW(),
    updated_at TIMESTAMPTZ DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS alliance_master_requirements(
    id BIGSERIAL PRIMARY KEY,
    req_id TEXT UNIQUE NOT NULL,
    source_group TEXT,
    region TEXT,
    location TEXT,
    property_type TEXT,
    transaction TEXT,
    budget_text TEXT,
    budget_min_inr NUMERIC(16,2),
    budget_max_inr NUMERIC(16,2),
    area_text TEXT,
    minimum_area_sqft NUMERIC(14,2),
    maximum_area_sqft NUMERIC(14,2),
    contact_no TEXT,
    client_broker TEXT,
    ai_confidence NUMERIC(5,2),
    raw_summary TEXT NOT NULL,
    status TEXT DEFAULT 'New',
    team_member TEXT,
    source_type TEXT DEFAULT 'WHATSAPP',
    source_requirement_id TEXT,
    fingerprint TEXT UNIQUE,
    created_at TIMESTAMPTZ DEFAULT NOW(),
    updated_at TIMESTAMPTZ DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS alliance_master_contacts(
    id BIGSERIAL PRIMARY KEY,
    contact_no TEXT UNIQUE NOT NULL,
    name TEXT,
    firm_brokerage TEXT,
    region TEXT,
    notes TEXT,
    message_count INTEGER DEFAULT 0,
    first_seen TIMESTAMPTZ,
    last_seen TIMESTAMPTZ,
    created_at TIMESTAMPTZ DEFAULT NOW(),
    updated_at TIMESTAMPTZ DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS alliance_master_matches(
    id BIGSERIAL PRIMARY KEY,
    req_id TEXT NOT NULL,
    listing_id TEXT NOT NULL,
    match_percentage NUMERIC(5,2),
    match_grade TEXT,
    reasons JSONB DEFAULT '[]'::jsonb,
    created_at TIMESTAMPTZ DEFAULT NOW(),
    UNIQUE(req_id, listing_id)
);

CREATE TABLE IF NOT EXISTS alliance_master_sync_runs(
    id BIGSERIAL PRIMARY KEY,
    run_id TEXT UNIQUE NOT NULL,
    started_at TIMESTAMPTZ DEFAULT NOW(),
    completed_at TIMESTAMPTZ,
    listings_synced INTEGER DEFAULT 0,
    requirements_synced INTEGER DEFAULT 0,
    contacts_synced INTEGER DEFAULT 0,
    duplicates_skipped INTEGER DEFAULT 0,
    error_message TEXT
);

CREATE INDEX IF NOT EXISTS idx_master_listing_location ON alliance_master_listings(location);
CREATE INDEX IF NOT EXISTS idx_master_listing_type ON alliance_master_listings(property_type);
CREATE INDEX IF NOT EXISTS idx_master_req_location ON alliance_master_requirements(location);
CREATE INDEX IF NOT EXISTS idx_master_req_status ON alliance_master_requirements(status);
CREATE INDEX IF NOT EXISTS idx_master_match_req ON alliance_master_matches(req_id);
"""

PROPERTY_TYPES = [
    "Apartment","Independent House / Villa","Commercial Shop","Commercial Showroom",
    "Commercial Space","Office","Plot / Land","Hotel","Guest House","Restaurant","Cafe",
    "Club","Banquet","Warehouse / Industrial","Farmhouse","Other","UNKNOWN"
]
TRANSACTIONS = ["SALE","RENT","SALE/RENT","REQUIREMENT","UNKNOWN"]
STATUSES = ["New","Matching","Shared with Client","Site Visit Scheduled","Site Visit Done","Closed - Won","Closed - Lost","Dead / Stale"]

PHONE_RE = re.compile(r"(?<!\d)(?:(?:\+?91)[\s-]?|0)?[6-9]\d(?:[\s-]?\d){8}(?!\d)")

def require_db():
    if engine is None:
        raise HTTPException(503, "DATABASE_URL / WHATSAPP_DATABASE_URL is not configured.")

def init_db():
    require_db()
    with engine.begin() as c:
        for stmt in [x.strip() for x in SCHEMA.split(";") if x.strip()]:
            c.execute(text(stmt))

@router.on_event("startup")
def startup():
    if engine is not None:
        try:
            init_db()
        except Exception as e:
            print("Alliance Master Property DB init warning:", e)

def esc(v):
    return html.escape("" if v is None else str(v))

def money(v):
    if v in (None, "", "UNKNOWN"):
        return "—"
    try:
        n=float(v)
        if n>=10_000_000:
            return f"₹{n/10_000_000:.2f} Cr"
        if n>=100_000:
            return f"₹{n/100_000:.2f} L"
        return f"₹{n:,.0f}"
    except Exception:
        return esc(v)

def norm(v):
    return re.sub(r"\s+"," ",str(v or "").replace("\u00a0"," ")).strip()

def phone_numbers(txt):
    out=[]
    for m in PHONE_RE.finditer(txt or ""):
        d=re.sub(r"\D","",m.group(0))
        if len(d)==11 and d.startswith("0"):
            d=d[1:]
        if len(d)==12 and d.startswith("91"):
            d=d[2:]
        if len(d)==10:
            p="+91"+d
            if p not in out:
                out.append(p)
    return out

def fingerprint(values):
    return hashlib.sha256("|".join(str(x or "").lower().strip() for x in values).encode()).hexdigest()

def shell(title, body, active="Dashboard"):
    nav = [
        ("Dashboard","/whatsapp-capture/database"),
        ("Listings","/whatsapp-capture/database/listings"),
        ("Requirements","/whatsapp-capture/database/requirements"),
        ("Contacts","/whatsapp-capture/database/contacts"),
        ("AI Matches","/whatsapp-capture/database/matches"),
        ("System Health","/whatsapp-capture/database/system-health"),
    ]
    links="".join(
        f'<a class="{"active" if n==active else ""}" href="{u}">{esc(n)}</a>'
        for n,u in nav
    )
    return f"""<!doctype html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{esc(title)}</title>
<style>
*{{box-sizing:border-box}}body{{margin:0;font-family:Arial,sans-serif;background:#f5f7fa;color:#101828}}
header{{background:#101828;color:#fff;padding:18px 24px}}header h1{{margin:0;font-size:23px}}header small{{color:#98a2b3}}
nav{{display:flex;gap:6px;flex-wrap:wrap;background:#fff;padding:10px 18px;border-bottom:1px solid #e4e7ec;position:sticky;top:0;z-index:10}}
nav a{{text-decoration:none;color:#344054;padding:9px 12px;border-radius:8px}}
nav a.active,nav a:hover{{background:#101828;color:#fff}}
main{{max-width:1600px;margin:22px auto;padding:0 18px}}
.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(190px,1fr));gap:14px}}
.card{{background:#fff;border:1px solid #e4e7ec;border-radius:12px;padding:17px}}
.num{{font-size:30px;font-weight:750;margin-top:4px}}
.muted{{color:#667085}}.good{{color:#027a48}}.bad{{color:#b42318}}
.btn{{display:inline-block;border:0;background:#101828;color:#fff;padding:10px 14px;border-radius:8px;text-decoration:none;cursor:pointer}}
.btn2{{background:#175cd3}}.btn3{{background:#039855}}.btnwarn{{background:#b54708}}
table{{width:100%;border-collapse:collapse;background:#fff;font-size:13px}}
th,td{{padding:10px;border-bottom:1px solid #eaecf0;text-align:left;vertical-align:top}}
th{{background:#f9fafb;position:sticky;top:58px;z-index:3}}
.scroll{{overflow:auto;max-height:72vh;border:1px solid #e4e7ec;border-radius:12px}}
input,select,textarea{{width:100%;padding:10px;border:1px solid #d0d5dd;border-radius:8px;background:#fff}}
form.gridform{{display:grid;grid-template-columns:repeat(auto-fit,minmax(190px,1fr));gap:10px}}
.full{{grid-column:1/-1}}
.pill{{padding:4px 8px;border-radius:999px;background:#eef4ff;color:#3538cd;display:inline-block}}
</style></head>
<body>
<header><h1>Alliance WhatsApp Group Property Database</h1><small>WhatsApp Groups · Listings · Requirements · Contacts · AI Matching</small></header>
<nav>{links}<a href="/workspace">← Main Workspace</a></nav>
<main>{body}</main></body></html>"""

def score_match(req, listing):
    score=0.0
    reasons=[]

    req_loc=(req.get("location") or "").lower().strip()
    prop_loc=(listing.get("location") or "").lower().strip()
    if req_loc and prop_loc:
        if req_loc in prop_loc or prop_loc in req_loc:
            score += 30
            reasons.append("Location")
        else:
            from difflib import SequenceMatcher
            s=SequenceMatcher(None,req_loc,prop_loc).ratio()
            if s>=0.72:
                score += 22
                reasons.append("Similar location")

    req_type=(req.get("property_type") or "").lower()
    prop_type=(listing.get("property_type") or "").lower()
    if req_type and prop_type and req_type not in ("unknown","other"):
        if req_type==prop_type:
            score += 15
            reasons.append("Property type")
        elif req_type in prop_type or prop_type in req_type:
            score += 10
            reasons.append("Related property type")

    req_tx=(req.get("transaction") or "").upper()
    prop_tx=(listing.get("transaction") or "").upper()
    if req_tx and prop_tx:
        if req_tx==prop_tx:
            score += 10
            reasons.append("Transaction")
        elif prop_tx=="SALE/RENT":
            score += 8
            reasons.append("Transaction available")

    mn=req.get("minimum_area_sqft")
    mx=req.get("maximum_area_sqft")
    area=listing.get("area_sqft")
    if area and (mn or mx):
        area=float(area)
        low=float(mn or mx)
        high=float(mx or mn)
        if low<=area<=high:
            score += 20
            reasons.append("Area")
        else:
            target=(low+high)/2
            if target>0:
                ratio=min(area,target)/max(area,target)
                if ratio>=0.8:
                    score += 12
                    reasons.append("Area near requirement")

    bmax=req.get("budget_max_inr")
    price=listing.get("budget_inr")
    if bmax and price:
        if float(price)<=float(bmax):
            score += 15
            reasons.append("Budget")
        else:
            over=(float(price)-float(bmax))/float(bmax)
            if over<=0.15:
                score += 8
                reasons.append("Budget near requirement")

    if listing.get("verified")=="Verified":
        score += 5
        reasons.append("Verified")

    if listing.get("contact_numbers"):
        score += 5
        reasons.append("Contact available")

    grade="EXCELLENT" if score>=90 else "STRONG" if score>=80 else "POSSIBLE" if score>=70 else "WEAK"
    return min(round(score,1),100),grade,reasons

def upsert_contact(c, phone, name="", firm="", region="", notes=""):
    if not phone:
        return
    c.execute(text("""
        INSERT INTO alliance_master_contacts(
            contact_no,name,firm_brokerage,region,notes,message_count,first_seen,last_seen
        ) VALUES(:p,:n,:f,:r,:notes,1,NOW(),NOW())
        ON CONFLICT(contact_no) DO UPDATE SET
            name=COALESCE(NULLIF(EXCLUDED.name,''),alliance_master_contacts.name),
            firm_brokerage=COALESCE(NULLIF(EXCLUDED.firm_brokerage,''),alliance_master_contacts.firm_brokerage),
            region=COALESCE(NULLIF(EXCLUDED.region,''),alliance_master_contacts.region),
            notes=COALESCE(NULLIF(EXCLUDED.notes,''),alliance_master_contacts.notes),
            message_count=alliance_master_contacts.message_count+1,
            last_seen=NOW(),
            updated_at=NOW()
    """),{"p":phone,"n":name,"f":firm,"r":region,"notes":notes})

def sync_from_whatsapp():
    """
    Read-only from current wa_* source tables.
    Master tables are refreshed/upserted. Existing wa_* tables are never altered.
    """
    require_db(); init_db()
    run_id="MASTER-"+uuid.uuid4().hex[:10].upper()
    stats={"listings":0,"requirements":0,"contacts":0,"duplicates":0}

    with engine.begin() as c:
        c.execute(text("INSERT INTO alliance_master_sync_runs(run_id) VALUES(:r)"),{"r":run_id})

        # Listings
        props=c.execute(text("""
            SELECT p.*,s.group_name,s.source_name
            FROM wa_properties p
            LEFT JOIN wa_sources s ON s.source_id=p.source_id
            WHERE COALESCE(p.record_status,'ACTIVE')='ACTIVE'
              AND COALESCE(p.duplicate_status,'UNIQUE')<>'DUPLICATE'
            ORDER BY p.id
        """)).mappings().all()

        for p in props:
            raw=norm(p.get("raw_text"))
            if not raw:
                continue
            contact=p.get("broker_phone") or p.get("owner_phone") or p.get("sender_phone") or ""
            contacts=" / ".join(phone_numbers(raw)) or contact
            tx=p.get("transaction_type") or "UNKNOWN"
            if tx=="SALE_RENT":
                tx="SALE/RENT"
            price=p.get("rent_inr") if tx=="RENT" else p.get("sale_price_inr")
            area=p.get("area_sqft")
            fp=fingerprint([
                p.get("wa_property_id"),p.get("location"),p.get("property_type"),tx,area,price,raw
            ])
            lid="LST-"+hashlib.sha256(fp.encode()).hexdigest()[:20].upper()
            c.execute(text("""
                INSERT INTO alliance_master_listings(
                    listing_id,source_message_id,source_group,source_account,sender_phone,transaction,
                    property_type,location,building_project,budget_inr,budget_text,area_sqft,area_text,
                    contact_numbers,poster_name,verified,raw_summary,fingerprint
                ) VALUES(
                    :id,:mid,:grp,:acct,:sender,:tx,:ptype,:loc,:proj,:budget,:btxt,:area,:atxt,:contacts,
                    :poster,:verified,:raw,:fp
                )
                ON CONFLICT(fingerprint) DO UPDATE SET
                    source_group=EXCLUDED.source_group,
                    source_account=EXCLUDED.source_account,
                    sender_phone=EXCLUDED.sender_phone,
                    transaction=EXCLUDED.transaction,
                    property_type=EXCLUDED.property_type,
                    location=EXCLUDED.location,
                    budget_inr=EXCLUDED.budget_inr,
                    budget_text=EXCLUDED.budget_text,
                    area_sqft=EXCLUDED.area_sqft,
                    area_text=EXCLUDED.area_text,
                    contact_numbers=EXCLUDED.contact_numbers,
                    poster_name=EXCLUDED.poster_name,
                    verified=EXCLUDED.verified,
                    raw_summary=EXCLUDED.raw_summary,
                    updated_at=NOW()
            """),{
                "id":lid,"mid":str(p.get("message_id") or ""),"grp":p.get("group_name") or "",
                "acct":p.get("source_name") or "","sender":p.get("sender_phone") or "",
                "tx":tx,"ptype":p.get("property_type") or "UNKNOWN","loc":p.get("location") or "UNKNOWN",
                "proj":p.get("property_name") or "","budget":price,"btxt":money(price) if price else "",
                "area":area,"atxt":f"{area} Sq Ft" if area else "","contacts":contacts,
                "poster":p.get("broker_name") or p.get("owner_name") or p.get("sender_name") or "",
                "verified":"Verified" if p.get("verification_status")=="VERIFIED_AVAILABLE" else "Unverified",
                "raw":raw,"fp":fp
            })
            stats["listings"]+=1
            for ph in [x.strip() for x in contacts.split("/") if x.strip()]:
                upsert_contact(
                    c, ph,
                    name=p.get("broker_name") or p.get("owner_name") or p.get("sender_name") or "",
                    firm="",
                    region=p.get("city") or "",
                    notes=f"Source group: {p.get('group_name') or ''}"
                )
                stats["contacts"]+=1

        # Requirements
        reqs=c.execute(text("""
            SELECT r.*,s.group_name,s.source_name
            FROM wa_requirements r
            LEFT JOIN wa_sources s ON s.source_id=r.source_id
            WHERE COALESCE(r.status,'ACTIVE')='ACTIVE'
            ORDER BY r.id
        """)).mappings().all()

        for r in reqs:
            raw=norm(r.get("raw_text"))
            if not raw:
                continue
            tx=r.get("transaction_type") or "UNKNOWN"
            if tx=="SALE_RENT":
                tx="SALE/RENT"
            fp=fingerprint([
                r.get("wa_requirement_id"),r.get("preferred_locations"),r.get("property_type"),tx,
                r.get("minimum_area_sqft"),r.get("maximum_area_sqft"),r.get("budget_max_inr"),
                r.get("contact_phone"),raw
            ])
            rid="REQ-"+hashlib.sha256(fp.encode()).hexdigest()[:20].upper()
            c.execute(text("""
                INSERT INTO alliance_master_requirements(
                    req_id,source_group,region,location,property_type,transaction,budget_text,budget_min_inr,
                    budget_max_inr,area_text,minimum_area_sqft,maximum_area_sqft,contact_no,client_broker,
                    ai_confidence,raw_summary,status,source_type,source_requirement_id,fingerprint
                ) VALUES(
                    :id,:grp,:region,:loc,:ptype,:tx,:btxt,:bmin,:bmax,:atxt,:amin,:amax,:phone,:client,
                    :conf,:raw,'New','WHATSAPP',:srid,:fp
                )
                ON CONFLICT(fingerprint) DO UPDATE SET
                    source_group=EXCLUDED.source_group,
                    region=EXCLUDED.region,
                    location=EXCLUDED.location,
                    property_type=EXCLUDED.property_type,
                    transaction=EXCLUDED.transaction,
                    budget_text=EXCLUDED.budget_text,
                    budget_min_inr=EXCLUDED.budget_min_inr,
                    budget_max_inr=EXCLUDED.budget_max_inr,
                    area_text=EXCLUDED.area_text,
                    minimum_area_sqft=EXCLUDED.minimum_area_sqft,
                    maximum_area_sqft=EXCLUDED.maximum_area_sqft,
                    contact_no=EXCLUDED.contact_no,
                    client_broker=EXCLUDED.client_broker,
                    ai_confidence=EXCLUDED.ai_confidence,
                    raw_summary=EXCLUDED.raw_summary,
                    updated_at=NOW()
            """),{
                "id":rid,"grp":r.get("group_name") or "","region":r.get("city") or "",
                "loc":r.get("preferred_locations") or "UNKNOWN","ptype":r.get("property_type") or "UNKNOWN",
                "tx":tx,"btxt":money(r.get("budget_max_inr")) if r.get("budget_max_inr") else "",
                "bmin":r.get("budget_min_inr"),"bmax":r.get("budget_max_inr"),
                "atxt":f"{r.get('minimum_area_sqft') or ''} - {r.get('maximum_area_sqft') or ''} Sq Ft",
                "amin":r.get("minimum_area_sqft"),"amax":r.get("maximum_area_sqft"),
                "phone":r.get("contact_phone") or "","client":r.get("contact_name") or r.get("client_name") or "",
                "conf":r.get("confidence") or 0,"raw":raw,"srid":r.get("wa_requirement_id") or "","fp":fp
            })
            stats["requirements"]+=1
            if r.get("contact_phone"):
                upsert_contact(
                    c,r.get("contact_phone"),
                    name=r.get("contact_name") or r.get("client_name") or "",
                    firm=r.get("company_name") or "",
                    region=r.get("city") or "",
                    notes=f"Requirement source group: {r.get('group_name') or ''}"
                )
                stats["contacts"]+=1

        c.execute(text("""
            UPDATE alliance_master_sync_runs
            SET completed_at=NOW(),listings_synced=:l,requirements_synced=:r,contacts_synced=:c
            WHERE run_id=:id
        """),{"l":stats["listings"],"r":stats["requirements"],"c":stats["contacts"],"id":run_id})

    return stats

def recompute_matches(req_id):
    require_db()
    with engine.begin() as c:
        req=c.execute(text("SELECT * FROM alliance_master_requirements WHERE req_id=:r"),{"r":req_id}).mappings().first()
        if not req:
            raise HTTPException(404,"Requirement not found")
        listings=c.execute(text("""
            SELECT * FROM alliance_master_listings
            WHERE COALESCE(record_status,'ACTIVE')='ACTIVE'
            ORDER BY id DESC
        """)).mappings().all()

        scored=[]
        for listing in listings:
            score,grade,reasons=score_match(req,listing)
            if score>=40:
                scored.append((score,grade,reasons,listing))
        scored.sort(key=lambda x:x[0], reverse=True)

        c.execute(text("DELETE FROM alliance_master_matches WHERE req_id=:r"),{"r":req_id})
        for score,grade,reasons,listing in scored[:300]:
            c.execute(text("""
                INSERT INTO alliance_master_matches(req_id,listing_id,match_percentage,match_grade,reasons)
                VALUES(:r,:l,:s,:g,CAST(:reasons AS JSONB))
                ON CONFLICT(req_id,listing_id) DO UPDATE SET
                    match_percentage=EXCLUDED.match_percentage,
                    match_grade=EXCLUDED.match_grade,
                    reasons=EXCLUDED.reasons,
                    created_at=NOW()
            """),{"r":req_id,"l":listing["listing_id"],"s":score,"g":grade,"reasons":json.dumps(reasons)})

        c.execute(text("UPDATE alliance_master_requirements SET status='Matching',updated_at=NOW() WHERE req_id=:r"),{"r":req_id})
        return scored[:300]

@router.get("", response_class=HTMLResponse)
def dashboard():
    require_db(); init_db()
    with engine.begin() as c:
        stats={
            "Listings":c.execute(text("SELECT COUNT(*) FROM alliance_master_listings WHERE COALESCE(record_status,'ACTIVE')='ACTIVE'")).scalar() or 0,
            "Requirements":c.execute(text("SELECT COUNT(*) FROM alliance_master_requirements")).scalar() or 0,
            "Contacts":c.execute(text("SELECT COUNT(*) FROM alliance_master_contacts")).scalar() or 0,
            "Verified Listings":c.execute(text("SELECT COUNT(*) FROM alliance_master_listings WHERE verified='Verified'")).scalar() or 0,
            "New Requirements":c.execute(text("SELECT COUNT(*) FROM alliance_master_requirements WHERE status='New'")).scalar() or 0,
        }
        recent=c.execute(text("""
            SELECT req_id,raw_summary,contact_no,budget_text,area_text,source_group,location,property_type,transaction,status
            FROM alliance_master_requirements ORDER BY id DESC LIMIT 10
        """)).mappings().all()

    cards="".join(f'<div class=card><div class=muted>{esc(k)}</div><div class=num>{v}</div></div>' for k,v in stats.items())
    rows="".join(
        f"<tr><td style='min-width:360px'>{esc(r['raw_summary'])}</td><td>{esc(r['contact_no'])}</td><td>{esc(r['budget_text'])}</td>"
        f"<td>{esc(r['area_text'])}</td><td>{esc(r['source_group'])}</td><td>{esc(r['location'])}</td>"
        f"<td>{esc(r['property_type'])}</td><td>{esc(r['transaction'])}</td><td>{esc(r['status'])}</td>"
        f"<td><a class=btn href='/whatsapp-capture/database/requirements/{esc(r['req_id'])}/matches'>Match</a></td></tr>"
        for r in recent
    )
    body=f"""
    <div style='display:flex;justify-content:space-between;align-items:center;gap:10px;flex-wrap:wrap'>
      <div><h2>Command Centre</h2><p class=muted>This is the final master structure. Existing wa_* source tables remain untouched.</p></div>
      <div>
        <a class='btn btn3' href='/whatsapp-capture/database/sync-whatsapp'>SYNC WHATSAPP DATABASE</a>
        <a class='btn btn2' href='/whatsapp-capture/database/requirements/new'>+ ADD OFFLINE REQUIREMENT</a>
        <a class='btn' href='/whatsapp-capture/database/export.xlsx'>EXPORT EXCEL</a>
      </div>
    </div>
    <div class=grid>{cards}</div>
    <h3>Recent Requirements</h3>
    <div class=scroll><table>
    <tr><th>Requirement Details</th><th>Contact No.</th><th>Budget</th><th>Area</th><th>Source Group</th><th>Location</th>
    <th>Property Type</th><th>Transaction</th><th>Status</th><th></th></tr>{rows}</table></div>
    """
    return HTMLResponse(shell("Alliance WhatsApp Group Property Database",body,"Dashboard"))

@router.get("/sync-whatsapp")
def sync_whatsapp():
    sync_from_whatsapp()
    return RedirectResponse("/whatsapp-capture/database",303)

@router.get("/listings", response_class=HTMLResponse)
def listings():
    require_db(); init_db()
    with engine.begin() as c:
        rows=c.execute(text("""
            SELECT * FROM alliance_master_listings
            WHERE COALESCE(record_status,'ACTIVE')='ACTIVE'
            ORDER BY id DESC LIMIT 3000
        """)).mappings().all()
    trs="".join(
        f"<tr><td style='min-width:420px;white-space:pre-wrap'>{esc(r['raw_summary'])}</td>"
        f"<td>{esc(r['contact_numbers'])}</td><td>{esc(r['budget_text'])}</td><td>{esc(r['area_text'])}</td>"
        f"<td>{esc(r['source_group'])}</td><td>{esc(r['location'])}</td><td>{esc(r['property_type'])}</td>"
        f"<td>{esc(r['transaction'])}</td><td>{esc(r['building_project'])}</td><td>{esc(r['poster_name'])}</td>"
        f"<td>{esc(r['verified'])}</td></tr>"
        for r in rows
    )
    body=f"""<h2>Listings</h2>
    <p class=muted>One property / distinct property offer = one row.</p>
    <div class=scroll><table>
    <tr><th>Raw Details</th><th>Contact No.</th><th>Price / Rent</th><th>Area</th><th>Source Group</th>
    <th>Location</th><th>Property Type</th><th>Transaction</th><th>Building / Project</th><th>Poster / Broker</th><th>Verified</th></tr>
    {trs}</table></div>"""
    return HTMLResponse(shell("Listings",body,"Listings"))

@router.get("/requirements", response_class=HTMLResponse)
def requirements():
    require_db(); init_db()
    with engine.begin() as c:
        rows=c.execute(text("""
            SELECT * FROM alliance_master_requirements
            ORDER BY id DESC LIMIT 3000
        """)).mappings().all()
    trs="".join(
        f"<tr><td style='min-width:420px;white-space:pre-wrap'>{esc(r['raw_summary'])}</td>"
        f"<td>{esc(r['contact_no'])}</td><td>{esc(r['budget_text'])}</td><td>{esc(r['area_text'])}</td>"
        f"<td>{esc(r['source_group'])}</td><td>{esc(r['location'])}</td><td>{esc(r['property_type'])}</td>"
        f"<td>{esc(r['transaction'])}</td><td>{esc(r['client_broker'])}</td><td>{esc(r['ai_confidence'])}</td>"
        f"<td>{esc(r['status'])}</td><td><a class=btn href='/whatsapp-capture/database/requirements/{esc(r['req_id'])}/matches'>Find Matches</a></td></tr>"
        for r in rows
    )
    body=f"""<div style='display:flex;justify-content:space-between;align-items:center;gap:10px;flex-wrap:wrap'>
      <div><h2>Requirements</h2><p class=muted>Full requirement details are visible on the front page.</p></div>
      <a class='btn btn3' href='/whatsapp-capture/database/requirements/new'>+ ADD OFFLINE REQUIREMENT</a>
    </div>
    <div class=scroll><table>
    <tr><th>Raw Requirement Details</th><th>Contact No.</th><th>Budget / Rent</th><th>Area</th><th>Source Group</th>
    <th>Location</th><th>Property Type</th><th>Transaction</th><th>Client / Broker</th><th>AI Confidence %</th><th>Status</th><th></th></tr>
    {trs}</table></div>"""
    return HTMLResponse(shell("Requirements",body,"Requirements"))

@router.get("/requirements/new", response_class=HTMLResponse)
def new_requirement():
    body=f"""
    <h2>Add Offline / Manual Requirement</h2>
    <div class=card><form class=gridform method=post>
      <div><label>Client / Broker</label><input name=client_broker required></div>
      <div><label>Contact No.</label><input name=contact_no></div>
      <div><label>Source Group / Source</label><input name=source_group value="Alliance Manual / Offline Requirement"></div>
      <div><label>Region</label><input name=region placeholder="Delhi NCR"></div>
      <div><label>Location</label><input name=location required></div>
      <div><label>Property Type</label><select name=property_type>{''.join(f'<option>{x}</option>' for x in PROPERTY_TYPES)}</select></div>
      <div><label>Transaction</label><select name=transaction>{''.join(f'<option>{x}</option>' for x in TRANSACTIONS)}</select></div>
      <div><label>Budget Min INR</label><input type=number step=0.01 name=budget_min_inr></div>
      <div><label>Budget Max INR</label><input type=number step=0.01 name=budget_max_inr></div>
      <div><label>Minimum Area Sq Ft</label><input type=number step=0.01 name=minimum_area_sqft></div>
      <div><label>Maximum Area Sq Ft</label><input type=number step=0.01 name=maximum_area_sqft></div>
      <div><label>Team Member</label><input name=team_member></div>
      <div class=full><label>Raw Requirement Details</label><textarea rows=5 name=raw_summary required></textarea></div>
      <div class=full><button class='btn btn3' type=submit>SAVE & FIND MATCHES</button></div>
    </form></div>
    """
    return HTMLResponse(shell("Add Requirement",body,"Requirements"))

@router.post("/requirements/new")
def create_requirement(
    client_broker: str=Form(...),
    contact_no: str=Form(""),
    source_group: str=Form("Alliance Manual / Offline Requirement"),
    region: str=Form(""),
    location: str=Form(...),
    property_type: str=Form("UNKNOWN"),
    transaction: str=Form("UNKNOWN"),
    budget_min_inr: Optional[float]=Form(None),
    budget_max_inr: Optional[float]=Form(None),
    minimum_area_sqft: Optional[float]=Form(None),
    maximum_area_sqft: Optional[float]=Form(None),
    team_member: str=Form(""),
    raw_summary: str=Form(...)
):
    require_db(); init_db()
    fp=fingerprint([
        source_group,region,location,property_type,transaction,budget_min_inr,budget_max_inr,
        minimum_area_sqft,maximum_area_sqft,contact_no,client_broker,raw_summary
    ])
    rid="REQ-"+hashlib.sha256(fp.encode()).hexdigest()[:20].upper()
    budget_text=money(budget_max_inr) if budget_max_inr else ""
    area_text=f"{minimum_area_sqft or ''} - {maximum_area_sqft or ''} Sq Ft"
    with engine.begin() as c:
        c.execute(text("""
            INSERT INTO alliance_master_requirements(
                req_id,source_group,region,location,property_type,transaction,budget_text,budget_min_inr,budget_max_inr,
                area_text,minimum_area_sqft,maximum_area_sqft,contact_no,client_broker,ai_confidence,raw_summary,
                status,team_member,source_type,fingerprint
            ) VALUES(
                :id,:grp,:region,:loc,:ptype,:tx,:btxt,:bmin,:bmax,:atxt,:amin,:amax,:phone,:client,100,:raw,
                'New',:team,'MANUAL',:fp
            )
            ON CONFLICT(fingerprint) DO UPDATE SET
                updated_at=NOW(),status='New',team_member=EXCLUDED.team_member
        """),{
            "id":rid,"grp":source_group,"region":region,"loc":location,"ptype":property_type,"tx":transaction,
            "btxt":budget_text,"bmin":budget_min_inr,"bmax":budget_max_inr,"atxt":area_text,
            "amin":minimum_area_sqft,"amax":maximum_area_sqft,"phone":contact_no,"client":client_broker,
            "raw":raw_summary,"team":team_member,"fp":fp
        })
        if contact_no:
            upsert_contact(c,contact_no,name=client_broker,region=region,notes="Manual/offline requirement")
    return RedirectResponse(f"/whatsapp-capture/database/requirements/{rid}/matches",303)

@router.get("/requirements/{req_id}/matches", response_class=HTMLResponse)
def matches_for_requirement(req_id: str):
    scored=recompute_matches(req_id)
    with engine.begin() as c:
        req=c.execute(text("SELECT * FROM alliance_master_requirements WHERE req_id=:r"),{"r":req_id}).mappings().first()
    trs=""
    for score,grade,reasons,l in scored:
        trs += (
            f"<tr><td><span class=pill><b>{score:.0f}%</b> {esc(grade)}</span></td>"
            f"<td style='min-width:380px;white-space:pre-wrap'>{esc(l['raw_summary'])}</td>"
            f"<td>{esc(l['contact_numbers'])}</td><td>{esc(l['budget_text'])}</td><td>{esc(l['area_text'])}</td>"
            f"<td>{esc(l['source_group'])}</td><td>{esc(l['location'])}</td><td>{esc(l['property_type'])}</td>"
            f"<td>{esc(l['verified'])}</td><td>{esc(', '.join(reasons))}</td></tr>"
        )
    body=f"""<h2>Matches</h2>
    <div class=card><b>Requirement:</b><p>{esc(req['raw_summary'])}</p></div>
    <div class=scroll><table>
    <tr><th>Match %</th><th>Raw Property Details</th><th>Contact No.</th><th>Price / Rent</th><th>Area</th>
    <th>Source Group</th><th>Location</th><th>Property Type</th><th>Verified</th><th>Why Matched</th></tr>
    {trs}</table></div>"""
    return HTMLResponse(shell("AI Matches",body,"AI Matches"))

@router.get("/matches", response_class=HTMLResponse)
def matches_index():
    require_db(); init_db()
    with engine.begin() as c:
        rows=c.execute(text("""
            SELECT r.req_id,r.raw_summary,r.location,r.status,
                   COUNT(m.id) AS match_count,MAX(m.match_percentage) AS best_match
            FROM alliance_master_requirements r
            LEFT JOIN alliance_master_matches m ON m.req_id=r.req_id
            GROUP BY r.req_id,r.raw_summary,r.location,r.status
            ORDER BY MAX(r.updated_at) DESC LIMIT 1000
        """)).mappings().all()
    trs="".join(
        f"<tr><td style='min-width:360px'>{esc(r['raw_summary'])}</td><td>{esc(r['location'])}</td>"
        f"<td>{esc(r['status'])}</td><td>{r['match_count']}</td><td>{float(r['best_match'] or 0):.0f}%</td>"
        f"<td><a class=btn href='/whatsapp-capture/database/requirements/{esc(r['req_id'])}/matches'>Open</a></td></tr>"
        for r in rows
    )
    return HTMLResponse(shell("AI Matches",f"<h2>AI Matches</h2><div class=scroll><table><tr><th>Requirement</th><th>Location</th><th>Status</th><th>Matches</th><th>Best Match</th><th></th></tr>{trs}</table></div>","AI Matches"))

@router.get("/contacts", response_class=HTMLResponse)
def contacts():
    require_db(); init_db()
    with engine.begin() as c:
        rows=c.execute(text("""
            SELECT * FROM alliance_master_contacts
            ORDER BY last_seen DESC NULLS LAST,id DESC LIMIT 3000
        """)).mappings().all()
    trs="".join(
        f"<tr><td>{esc(r['contact_no'])}</td><td>{esc(r['name'])}</td><td>{esc(r['firm_brokerage'])}</td>"
        f"<td>{esc(r['region'])}</td><td>{r['message_count']}</td><td>{esc(r['first_seen'])}</td>"
        f"<td>{esc(r['last_seen'])}</td><td>{esc(r['notes'])}</td></tr>"
        for r in rows
    )
    return HTMLResponse(shell("Contacts",f"<h2>Contacts</h2><div class=scroll><table><tr><th>Contact No.</th><th>Name</th><th>Firm / Brokerage</th><th>Region</th><th>Message Count</th><th>First Seen</th><th>Last Seen</th><th>Notes</th></tr>{trs}</table></div>","Contacts"))

@router.get("/system-health", response_class=HTMLResponse)
def system_health():
    require_db(); init_db()
    checks=[]
    with engine.begin() as c:
        for name,q in [
            ("Master Listings","SELECT COUNT(*) FROM alliance_master_listings"),
            ("Master Requirements","SELECT COUNT(*) FROM alliance_master_requirements"),
            ("Master Contacts","SELECT COUNT(*) FROM alliance_master_contacts"),
            ("Master Matches","SELECT COUNT(*) FROM alliance_master_matches"),
            ("WhatsApp Source Properties","SELECT COUNT(*) FROM wa_properties"),
            ("WhatsApp Source Requirements","SELECT COUNT(*) FROM wa_requirements"),
        ]:
            try:
                v=c.execute(text(q)).scalar() or 0
                checks.append((name,"OK",v))
            except Exception as e:
                checks.append((name,"ERROR",str(e)[:150]))
    trs="".join(f"<tr><td>{esc(n)}</td><td class='{'good' if s=='OK' else 'bad'}'>{esc(s)}</td><td>{esc(v)}</td></tr>" for n,s,v in checks)
    body=f"""<h2>System Health</h2>
    <div class=card><b>Safety:</b> The master database reads from existing wa_* source tables. It does not ALTER or DROP them.</div>
    <div class=scroll><table><tr><th>Component</th><th>Status</th><th>Records / Error</th></tr>{trs}</table></div>"""
    return HTMLResponse(shell("System Health",body,"System Health"))

@router.get("/export.xlsx")
def export_excel():
    require_db(); init_db()
    with engine.begin() as c:
        listings=c.execute(text("SELECT * FROM alliance_master_listings ORDER BY id")).mappings().all()
        requirements=c.execute(text("SELECT * FROM alliance_master_requirements ORDER BY id")).mappings().all()
        contacts=c.execute(text("SELECT * FROM alliance_master_contacts ORDER BY id")).mappings().all()

    wb=Workbook()
    wb.remove(wb.active)

    def add_sheet(name,headers,rows):
        ws=wb.create_sheet(name)
        for col,h in enumerate(headers,1):
            cell=ws.cell(1,col,h)
            cell.font=Font(bold=True,color="FFFFFF")
            cell.fill=PatternFill("solid",fgColor="0F766E")
            cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        for ri,row in enumerate(rows,2):
            for ci,h in enumerate(headers,1):
                ws.cell(ri,ci,row.get(h))
        for ci,h in enumerate(headers,1):
            width=min(max(len(h)+3,14),45)
            ws.column_dimensions[get_column_letter(ci)].width=width
        ws.freeze_panes="A2"
        ws.auto_filter.ref=ws.dimensions

    add_sheet("Listings",
        ["listing_id","message_date","message_time","sender_phone","transaction","property_type","location",
         "building_project","budget_text","area_text","contact_numbers","poster_name","verified","raw_summary"],
        listings
    )
    add_sheet("Requirements",
        ["req_id","source_group","region","location","property_type","transaction","budget_text","area_text",
         "contact_no","client_broker","ai_confidence","raw_summary","status"],
        requirements
    )
    add_sheet("Contacts",
        ["contact_no","name","firm_brokerage","region","notes","message_count","first_seen","last_seen"],
        contacts
    )

    bio=io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    headers={"Content-Disposition":"attachment; filename=Alliance_Master_Property_Database.xlsx"}
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )
