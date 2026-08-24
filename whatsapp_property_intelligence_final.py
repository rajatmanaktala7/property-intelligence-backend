import os, re, io, json, uuid, hashlib, html
from datetime import datetime, timezone
from typing import Optional, List, Dict, Any

from fastapi import APIRouter, Form, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse
from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/whatsapp-capture/intelligence", tags=["WhatsApp Property Intelligence Final"])

DATABASE_URL = (os.getenv("WHATSAPP_DATABASE_URL") or os.getenv("DATABASE_URL") or "").strip()
LLM_PROVIDER = os.getenv("WA_EXTRACTION_PROVIDER","").strip().lower()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY","").strip()
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY","").strip()
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY","").strip()

def _db_url(url: str) -> str:
    if url.startswith("postgres://"):
        return url.replace("postgres://","postgresql+psycopg://",1)
    if url.startswith("postgresql://"):
        return url.replace("postgresql://","postgresql+psycopg://",1)
    return url

engine = create_engine(_db_url(DATABASE_URL), pool_pre_ping=True, pool_recycle=300) if DATABASE_URL else None

SCHEMA = """
CREATE TABLE IF NOT EXISTS wai_groups (
    id UUID PRIMARY KEY,
    name TEXT NOT NULL,
    region TEXT,
    is_active BOOLEAN DEFAULT TRUE,
    created_at TIMESTAMPTZ DEFAULT NOW(),
    UNIQUE(name)
);

CREATE TABLE IF NOT EXISTS wai_raw_messages (
    id UUID PRIMARY KEY,
    group_id UUID REFERENCES wai_groups(id),
    sender_phone TEXT,
    sender_display_name TEXT,
    message_text TEXT NOT NULL,
    sent_at TIMESTAMPTZ,
    ingested_at TIMESTAMPTZ DEFAULT NOW(),
    extraction_status TEXT DEFAULT 'pending',
    source_message_key TEXT UNIQUE
);

CREATE TABLE IF NOT EXISTS wai_contacts (
    id UUID PRIMARY KEY,
    phone TEXT UNIQUE NOT NULL,
    display_name TEXT,
    firm_name TEXT,
    message_count INT DEFAULT 0,
    first_seen_at TIMESTAMPTZ,
    last_seen_at TIMESTAMPTZ,
    trust_score NUMERIC DEFAULT 0
);

CREATE TABLE IF NOT EXISTS wai_listings (
    id UUID PRIMARY KEY,
    source_message_id UUID REFERENCES wai_raw_messages(id),
    contact_id UUID REFERENCES wai_contacts(id),
    transaction TEXT,
    property_type TEXT,
    location TEXT,
    region TEXT,
    budget_text TEXT,
    budget_numeric NUMERIC,
    area_text TEXT,
    area_sqft_numeric NUMERIC,
    summary TEXT,
    extraction_method TEXT,
    confidence_score NUMERIC,
    status TEXT DEFAULT 'unverified',
    is_public BOOLEAN DEFAULT FALSE,
    duplicate_of UUID REFERENCES wai_listings(id) NULL,
    created_at TIMESTAMPTZ DEFAULT NOW(),
    verified_at TIMESTAMPTZ,
    verified_by TEXT,
    listing_fingerprint TEXT UNIQUE,
    source_group_name TEXT,
    poster_name TEXT,
    raw_listing_text TEXT
);

CREATE TABLE IF NOT EXISTS wai_verification_log (
    id UUID PRIMARY KEY,
    listing_id UUID REFERENCES wai_listings(id),
    action TEXT,
    field_changed TEXT,
    old_value TEXT,
    new_value TEXT,
    actor TEXT,
    created_at TIMESTAMPTZ DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS wai_matches (
    id UUID PRIMARY KEY,
    requirement_listing_id UUID REFERENCES wai_listings(id),
    supply_listing_id UUID REFERENCES wai_listings(id),
    match_score NUMERIC,
    match_reason TEXT,
    status TEXT DEFAULT 'suggested',
    created_at TIMESTAMPTZ DEFAULT NOW(),
    UNIQUE(requirement_listing_id,supply_listing_id)
);

CREATE TABLE IF NOT EXISTS wai_public_leads (
    id UUID PRIMARY KEY,
    listing_id UUID REFERENCES wai_listings(id) NULL,
    name TEXT,
    email TEXT,
    phone TEXT,
    message TEXT,
    source TEXT,
    created_at TIMESTAMPTZ DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS wai_pipeline_runs (
    id UUID PRIMARY KEY,
    started_at TIMESTAMPTZ DEFAULT NOW(),
    completed_at TIMESTAMPTZ,
    messages_ingested INT DEFAULT 0,
    messages_extracted INT DEFAULT 0,
    listing_rows_created INT DEFAULT 0,
    requirement_rows_created INT DEFAULT 0,
    skipped INT DEFAULT 0,
    failed INT DEFAULT 0,
    notes TEXT
);

CREATE INDEX IF NOT EXISTS idx_wai_listing_status ON wai_listings(status);
CREATE INDEX IF NOT EXISTS idx_wai_listing_tx ON wai_listings(transaction);
CREATE INDEX IF NOT EXISTS idx_wai_listing_location ON wai_listings(location);
CREATE INDEX IF NOT EXISTS idx_wai_raw_group ON wai_raw_messages(group_id);
"""

PROPERTY_TYPES = [
    "Apartment","Independent House / Villa","Plot / Land","Office","Commercial Shop",
    "Commercial Showroom","Warehouse / Industrial","Hotel","Guest House","Restaurant",
    "Banquet","Farmhouse","Cafe","Club","Commercial Space","Other"
]
TRANSACTIONS = ["SALE","RENT","REQUIREMENT"]
KNOWN_LOCATIONS = [
    "Siolim","Assagao","Anjuna","Vagator","Morjim","Parra","Porvorim","Panjim","Panaji","Miramar",
    "Caranzalem","Taleigao","Dona Paula","Bambolim","Saligao","Sangolda","Guirim","Old Goa","Campal",
    "Delhi","South Delhi","Defence Colony","Greater Kailash","GK 1","GK 2","Vasant Kunj","Vasant Vihar",
    "Saket","Green Park","Hauz Khas","Janakpuri","Dwarka","Karol Bagh","Rohini","Gurugram","Gurgaon",
    "Noida","Faridabad","Ghaziabad","Vaishali","Indirapuram","Hapur","Sohna Road","DLF Phase 1",
    "DLF Phase 2","DLF Phase 3","DLF Phase 4","Sector 50","Sector 51","Sector 57","Sector 61","Sector 65"
]

PHONE_RE = re.compile(r"(?<!\d)(?:(?:\+?91)[\s-]?|0)?[6-9]\d(?:[\s-]?\d){8}(?!\d)")
AREA_RE = re.compile(r"(\d[\d,]*(?:\.\d+)?)\s*(sq\.?\s*ft|sqft|sft|sq\.?\s*yd|sqyd|gaj|sq\.?\s*m|sqm|acre|acres|bigha)", re.I)
MONEY_RE = re.compile(r"(?:₹|rs\.?|inr)?\s*(\d[\d,]*(?:\.\d+)?)\s*(cr|crore|crores|l|lac|lakh|lakhs|k|psf)?", re.I)
REQ_WORDS = ["require","required","requirement","wanted","looking for","need","buyer looking","tenant looking","client looking"]
SALE_WORDS = ["for sale","sale","selling","sell","auction","reserve price"]
RENT_WORDS = ["for rent","rent","lease","leasing","available on rent"]
PROPERTY_WORDS = [
    "apartment","flat","bhk","villa","kothi","house","plot","land","office","shop","showroom",
    "warehouse","industrial","factory","hotel","guesthouse","guest house","restaurant","banquet",
    "farmhouse","farm house","commercial","retail","cafe","club"
]

def require_db():
    if engine is None:
        raise HTTPException(503,"DATABASE_URL / WHATSAPP_DATABASE_URL is not configured.")

def init_db():
    require_db()
    with engine.begin() as c:
        for stmt in [x.strip() for x in SCHEMA.split(";") if x.strip()]:
            c.execute(text(stmt))

@router.on_event("startup")
def _startup():
    if engine is not None:
        try:init_db()
        except Exception as e:print("WAI init warning:",e)

def esc(v): return html.escape("" if v is None else str(v))
def norm(v): return re.sub(r"\s+"," ",str(v or "").replace("\u00a0"," ")).strip()

def phones(txt):
    out=[]
    for m in PHONE_RE.finditer(txt or ""):
        d=re.sub(r"\D","",m.group(0))
        if len(d)==12 and d.startswith("91"):d=d[2:]
        if len(d)==11 and d.startswith("0"):d=d[1:]
        if len(d)==10 and d[0] in "6789":
            p="+91"+d
            if p not in out:out.append(p)
    return out

def detect_transaction(txt):
    low=(txt or "").lower()
    if any(x in low for x in REQ_WORDS): return "REQUIREMENT"
    if any(x in low for x in SALE_WORDS) and any(x in low for x in RENT_WORDS): return "SALE"
    if any(x in low for x in SALE_WORDS): return "SALE"
    if any(x in low for x in RENT_WORDS): return "RENT"
    return None

def detect_type(txt):
    low=(txt or "").lower()
    mapping=[
        ("Warehouse / Industrial",["warehouse","industrial","factory","godown"]),
        ("Commercial Showroom",["showroom"]),
        ("Commercial Shop",["shop"]),
        ("Office",["office"]),
        ("Plot / Land",["plot","land"]),
        ("Farmhouse",["farmhouse","farm house"]),
        ("Banquet",["banquet"]),
        ("Hotel",["hotel","resort"]),
        ("Guest House",["guesthouse","guest house"]),
        ("Restaurant",["restaurant","restro"]),
        ("Cafe",["cafe"]),
        ("Club",["club"]),
        ("Independent House / Villa",["villa","kothi","independent house","row house"]),
        ("Apartment",["apartment","flat","bhk","builder floor"]),
        ("Commercial Space",["commercial","retail space"]),
    ]
    for out,keys in mapping:
        if any(k in low for k in keys): return out
    return None

def detect_location(txt):
    low=(txt or "").lower()
    found=[]
    for loc in sorted(KNOWN_LOCATIONS,key=len,reverse=True):
        if loc.lower() in low and loc not in found:
            found.append(loc)
    sec=re.findall(r"\b(?:sector|sec)[\s\-]*([0-9]{1,3}[a-z]?)\b",txt or "",re.I)
    for s in sec:
        v="Sector "+s.upper()
        if v not in found:found.append(v)
    return ", ".join(found[:4]) if found else None

def extract_area(txt):
    m=AREA_RE.search(txt or "")
    if not m:return None,None
    raw=m.group(0)
    try:v=float(m.group(1).replace(",",""))
    except:return raw,None
    u=m.group(2).lower().replace(".","").replace(" ","")
    if u in ("sqyd","gaj"):sqft=v*9
    elif u in ("sqm","sqmtr","sqmtrs"):sqft=v*10.7639
    elif u.startswith("acre"):sqft=v*43560
    else:sqft=v
    return raw,round(sqft,2)

def extract_budget(txt):
    low=(txt or "").lower()
    label = re.search(r"(?:asking|demand|price|rent|reserve price)\s*[:=@-]?\s*(?:₹|rs\.?|inr)?\s*(\d[\d,]*(?:\.\d+)?)\s*(cr|crore|crores|l|lac|lakh|lakhs|k|psf)?",txt or "",re.I)
    m=label
    if not m:return None,None
    raw=m.group(0)
    num=m.group(1).replace(",","")
    suf=(m.group(2) or "").lower()
    try:v=float(num)
    except:return raw,None
    if suf in ("cr","crore","crores"):v*=10000000
    elif suf in ("l","lac","lakh","lakhs"):v*=100000
    elif suf=="k":v*=1000
    return raw,v

def region_for(txt,loc):
    low=(txt or "").lower()
    if any(x in low for x in ["goa","siolim","assagao","anjuna","vagator","morjim","parra","porvorim","panjim","panaji","miramar","caranzalem","dona paula"]):
        return "Goa - North"
    if any(x in low for x in ["delhi","gurgaon","gurugram","noida","faridabad","ghaziabad","vaishali","indirapuram","hapur"]):
        return "Delhi NCR"
    if "mumbai" in low or "juhu" in low:return "Mumbai"
    return "Other"

def confidence_score(d,method):
    score=0
    loc=d.get("location")
    if loc and any(k.lower() in loc.lower() for k in KNOWN_LOCATIONS):score+=30
    if d.get("property_type"):score+=20
    if d.get("transaction"):score+=20
    if d.get("budget_text") or d.get("area_text"):score+=15
    if d.get("contact_numbers"):score+=10
    if method=="llm":score+=5
    return min(score,100)

def entity_fingerprint(group_name, raw_piece, tx,ptype,loc,area,budget):
    key="|".join([norm(group_name).lower(),norm(raw_piece).lower(),str(tx or ""),str(ptype or ""),str(loc or ""),str(area or ""),str(budget or "")])
    return hashlib.sha256(key.encode()).hexdigest()

def looks_like_property_piece(txt):
    low=(txt or "").lower()
    if any(x in low for x in ["join group","group chat invite","view channel"]) and not any(x in low for x in PROPERTY_WORDS):
        return False
    return any(x in low for x in PROPERTY_WORDS) and (detect_transaction(txt) is not None or extract_area(txt)[0] or extract_budget(txt)[0])

def split_multi_listing_message(raw):
    """
    Critical rule: one WhatsApp message may produce N listings.
    Split on strong list boundaries, then keep only full property-like entities.
    """
    raw=str(raw or "").replace("\r\n","\n").replace("\r","\n")
    # Numbered list: 1) ... 2) ... / 1. ... 2. ...
    p = re.compile(r"(?m)(?=^\s*(?:option\s*)?\d{1,3}\s*[\)\.\-:]\s*)", re.I)
    parts=[norm(x) for x in p.split(raw) if norm(x)]
    good=[x for x in parts if looks_like_property_piece(x)]
    if len(good)>=2:return good

    # Emoji/keycap numbered lists: 1️⃣, 2️⃣, 3️⃣ ...
    parts=[norm(x) for x in re.split(r"(?m)(?=^\s*[0-9]\ufe0f?\u20e3\s*)",raw) if norm(x)]
    good=[x for x in parts if looks_like_property_piece(x)]
    if len(good)>=2:return good

    # Circled-number lists: ① ② ③ ...
    parts=[norm(x) for x in re.split(r"(?m)(?=^\s*[①②③④⑤⑥⑦⑧⑨⑩]\s*)",raw) if norm(x)]
    good=[x for x in parts if looks_like_property_piece(x)]
    if len(good)>=2:return good

    # OPTION markers even when inline
    parts=[norm(x) for x in re.split(r"(?i)(?=(?:OPTION|OPT)\s*\d+\b)",raw) if norm(x)]
    good=[x for x in parts if looks_like_property_piece(x)]
    if len(good)>=2:return good

    # Repeated BUILDING blocks are distinct commercial listings
    parts=[norm(x) for x in re.split(r"(?i)(?=\bBUILDING\s*[-:])",raw) if norm(x)]
    good=[x for x in parts if looks_like_property_piece(x)]
    if len(good)>=2:return good

    # Lines that each independently contain type + area/price
    lines=[norm(x) for x in raw.splitlines() if norm(x)]
    good=[x for x in lines if looks_like_property_piece(x)]
    if len(good)>=2:
        return good

    return [norm(raw)] if looks_like_property_piece(raw) else []

def rule_extract(piece, whole_message=None):
    tx=detect_transaction(piece)
    ptype=detect_type(piece)
    loc=detect_location(piece)
    area_text,area_num=extract_area(piece)
    budget_text,budget_num=extract_budget(piece)
    ph=phones(piece)
    if not ph and whole_message:
        ph=phones(whole_message)
    d={
        "transaction":tx,
        "property_type":ptype,
        "location":loc,
        "region":region_for(piece,loc),
        "budget_text":budget_text,
        "budget_numeric":budget_num,
        "area_text":area_text,
        "area_sqft_numeric":area_num,
        "contact_numbers":ph,
        "poster_name":None,
        "summary":norm(piece),
        "raw_listing_text":norm(piece)
    }
    return d

def should_llm(d):
    return not d.get("location") or not d.get("property_type") or not d.get("transaction")

def llm_extract(raw):
    """
    Optional Stage 2. Returns [] if no configured provider.
    The system remains fully operational with Stage 1 only.
    """
    prompt = """You are extracting structured real estate data from a raw WhatsApp broker message.
Return ONLY JSON array. If a field is absent use null.
Each distinct property/listing/requirement must be a separate object.
Never merge multiple buildings/units/options into one object.
Schema per object:
transaction: SALE|RENT|REQUIREMENT|null
property_type: Apartment|Independent House / Villa|Plot / Land|Office|Commercial Shop|Commercial Showroom|Warehouse / Industrial|Hotel|Guest House|Restaurant|Banquet|Farmhouse|Cafe|Club|Commercial Space|null
location: string|null
region: Goa - North|Goa - South|Delhi NCR|Mumbai|Other|null
budget_text: original string|null
area_text: original string|null
contact_numbers: array
poster_name: string|null
summary: concise but complete single-listing description
raw_listing_text: exact relevant portion of the source message
"""
    try:
        if LLM_PROVIDER=="openai" and OPENAI_API_KEY:
            from openai import OpenAI
            client=OpenAI(api_key=OPENAI_API_KEY)
            r=client.responses.create(model=os.getenv("WA_OPENAI_MODEL","gpt-5-mini"),input=prompt+"\n\n"+raw)
            data=json.loads(r.output_text)
            return data if isinstance(data,list) else []
        if LLM_PROVIDER=="anthropic" and ANTHROPIC_API_KEY:
            import anthropic
            client=anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
            r=client.messages.create(model=os.getenv("WA_ANTHROPIC_MODEL","claude-sonnet-4-20250514"),max_tokens=3000,messages=[{"role":"user","content":prompt+"\n\n"+raw}])
            data=json.loads(r.content[0].text)
            return data if isinstance(data,list) else []
        if LLM_PROVIDER=="gemini" and GEMINI_API_KEY:
            from google import genai
            client=genai.Client(api_key=GEMINI_API_KEY)
            r=client.models.generate_content(model=os.getenv("WA_GEMINI_MODEL","gemini-2.5-flash"),contents=prompt+"\n\n"+raw,config={"response_mime_type":"application/json"})
            data=json.loads(r.text)
            return data if isinstance(data,list) else []
    except Exception as e:
        print("WAI LLM extraction warning:",e)
    return []

def upsert_contact(c,phone,name="",firm=""):
    if not phone:return None
    cid=uuid.uuid5(uuid.NAMESPACE_URL,"contact:"+phone)
    now=datetime.now(timezone.utc)
    c.execute(text("""
        INSERT INTO wai_contacts(id,phone,display_name,firm_name,message_count,first_seen_at,last_seen_at)
        VALUES(:id,:p,:n,:f,1,:now,:now)
        ON CONFLICT(phone) DO UPDATE SET
          display_name=COALESCE(NULLIF(EXCLUDED.display_name,''),wai_contacts.display_name),
          firm_name=COALESCE(NULLIF(EXCLUDED.firm_name,''),wai_contacts.firm_name),
          message_count=wai_contacts.message_count+1,
          last_seen_at=EXCLUDED.last_seen_at
    """),{"id":cid,"p":phone,"n":name or "","f":firm or "","now":now})
    return cid

def ingest_current_whatsapp_source():
    """
    Adapter from existing production wa_* tables into the spec's raw audit tables.
    It never alters or deletes wa_* source data.
    """
    require_db();init_db()
    run_id=uuid.uuid4()
    stats={"messages":0,"extracted":0,"listings":0,"requirements":0,"skipped":0,"failed":0}
    with engine.begin() as c:
        c.execute(text("INSERT INTO wai_pipeline_runs(id) VALUES(:id)"),{"id":run_id})
        rows=c.execute(text("""
            SELECT m.message_id,m.source_id,m.raw_text,m.sender_name,m.sender_phone,
                   m.message_timestamp,m.created_at AS source_created_at,
                   s.group_name,s.source_name
            FROM wa_messages m
            LEFT JOIN wa_sources s ON s.source_id=m.source_id
            WHERE COALESCE(m.raw_text,'')<>''
            ORDER BY m.id
        """)).mappings().all()

        for r in rows:
            stats["messages"]+=1
            group_name=r.get("group_name") or r.get("source_name") or "Unknown WhatsApp Group"
            gid=uuid.uuid5(uuid.NAMESPACE_URL,"group:"+group_name)
            c.execute(text("""
                INSERT INTO wai_groups(id,name,region,is_active) VALUES(:id,:n,:r,TRUE)
                ON CONFLICT(name) DO UPDATE SET is_active=TRUE
            """),{"id":gid,"n":group_name,"r":region_for(group_name,None)})

            source_key=str(r.get("message_id") or hashlib.sha256((group_name+"|"+norm(r.get("raw_text"))).encode()).hexdigest())
            mid=uuid.uuid5(uuid.NAMESPACE_URL,"msg:"+source_key)
            c.execute(text("""
                INSERT INTO wai_raw_messages(id,group_id,sender_phone,sender_display_name,message_text,sent_at,source_message_key)
                VALUES(:id,:gid,:p,:n,:txt,:sent,:key)
                ON CONFLICT(source_message_key) DO NOTHING
            """),{"id":mid,"gid":gid,"p":r.get("sender_phone") or "","n":r.get("sender_name") or "",
                  "txt":r.get("raw_text") or "","sent":r.get("source_created_at"),"key":source_key})

            already=c.execute(text("SELECT extraction_status FROM wai_raw_messages WHERE id=:id"),{"id":mid}).scalar()
            if already in ("extracted","skipped"):
                stats["skipped"]+=1
                continue

            raw=r.get("raw_text") or ""
            pieces=split_multi_listing_message(raw)
            if not pieces:
                c.execute(text("UPDATE wai_raw_messages SET extraction_status='skipped' WHERE id=:id"),{"id":mid})
                stats["skipped"]+=1
                continue

            extracted=[]
            method="rules"
            for p in pieces:
                d=rule_extract(p,raw)
                if looks_like_property_piece(p):
                    extracted.append(d)

            if (not extracted or any(should_llm(x) for x in extracted)) and LLM_PROVIDER:
                llm_rows=llm_extract(raw)
                if llm_rows:
                    extracted=[]
                    method="llm"
                    for x in llm_rows:
                        piece=x.get("raw_listing_text") or x.get("summary") or raw
                        base=rule_extract(piece,raw)
                        for k in ["transaction","property_type","location","region","budget_text","area_text","poster_name","summary","raw_listing_text","contact_numbers"]:
                            if x.get(k) not in (None,"",[]):
                                base[k]=x.get(k)
                        # retain numeric normalizations from rules where possible
                        extracted.append(base)

            if not extracted:
                c.execute(text("UPDATE wai_raw_messages SET extraction_status='failed' WHERE id=:id"),{"id":mid})
                stats["failed"]+=1
                continue

            for idx,d in enumerate(extracted):
                conf=confidence_score(d,method)
                status="rejected" if conf<40 else "unverified"
                ph=(d.get("contact_numbers") or [r.get("sender_phone") or ""])[0] if (d.get("contact_numbers") or r.get("sender_phone")) else ""
                contact_id=upsert_contact(c,ph,d.get("poster_name") or r.get("sender_name") or "")
                fp=entity_fingerprint(group_name,d.get("raw_listing_text") or d.get("summary"),d.get("transaction"),
                                      d.get("property_type"),d.get("location"),d.get("area_text"),d.get("budget_text"))
                lid=uuid.uuid5(uuid.NAMESPACE_URL,"listing:"+fp)
                c.execute(text("""
                    INSERT INTO wai_listings(
                        id,source_message_id,contact_id,transaction,property_type,location,region,budget_text,budget_numeric,
                        area_text,area_sqft_numeric,summary,extraction_method,confidence_score,status,is_public,
                        listing_fingerprint,source_group_name,poster_name,raw_listing_text
                    ) VALUES(
                        :id,:mid,:cid,:tx,:ptype,:loc,:region,:btxt,:bnum,:atxt,:anum,:summary,:method,:conf,:status,FALSE,
                        :fp,:grp,:poster,:raw
                    )
                    ON CONFLICT(listing_fingerprint) DO UPDATE SET
                        transaction=EXCLUDED.transaction,
                        property_type=EXCLUDED.property_type,
                        location=EXCLUDED.location,
                        region=EXCLUDED.region,
                        budget_text=EXCLUDED.budget_text,
                        budget_numeric=EXCLUDED.budget_numeric,
                        area_text=EXCLUDED.area_text,
                        area_sqft_numeric=EXCLUDED.area_sqft_numeric,
                        summary=EXCLUDED.summary,
                        extraction_method=EXCLUDED.extraction_method,
                        confidence_score=EXCLUDED.confidence_score,
                        source_group_name=EXCLUDED.source_group_name,
                        poster_name=EXCLUDED.poster_name,
                        raw_listing_text=EXCLUDED.raw_listing_text
                """),{"id":lid,"mid":mid,"cid":contact_id,"tx":d.get("transaction"),"ptype":d.get("property_type"),
                      "loc":d.get("location"),"region":d.get("region"),"btxt":d.get("budget_text"),
                      "bnum":d.get("budget_numeric"),"atxt":d.get("area_text"),"anum":d.get("area_sqft_numeric"),
                      "summary":d.get("summary"),"method":method,"conf":conf,"status":status,"fp":fp,"grp":group_name,
                      "poster":d.get("poster_name") or r.get("sender_name") or "","raw":d.get("raw_listing_text") or d.get("summary")})
                if d.get("transaction")=="REQUIREMENT":stats["requirements"]+=1
                else:stats["listings"]+=1

            c.execute(text("UPDATE wai_raw_messages SET extraction_status='extracted' WHERE id=:id"),{"id":mid})
            stats["extracted"]+=1

        c.execute(text("""
            UPDATE wai_pipeline_runs SET completed_at=NOW(),messages_ingested=:m,messages_extracted=:e,
            listing_rows_created=:l,requirement_rows_created=:r,skipped=:s,failed=:f WHERE id=:id
        """),{"m":stats["messages"],"e":stats["extracted"],"l":stats["listings"],"r":stats["requirements"],
              "s":stats["skipped"],"f":stats["failed"],"id":run_id})
    return stats

def score_match(req,supply):
    score=0;reasons=[]
    rl=(req.get("location") or "").lower();sl=(supply.get("location") or "").lower()
    if rl and sl and (rl in sl or sl in rl):score+=35;reasons.append("location")
    rt=(req.get("property_type") or "").lower();st=(supply.get("property_type") or "").lower()
    if rt and st and rt==st:score+=20;reasons.append("property type")
    ra=req.get("area_sqft_numeric");sa=supply.get("area_sqft_numeric")
    if ra and sa:
        ratio=min(float(ra),float(sa))/max(float(ra),float(sa))
        if ratio>=0.85:score+=20;reasons.append("area")
        elif ratio>=0.70:score+=12;reasons.append("area near")
    rb=req.get("budget_numeric");sb=supply.get("budget_numeric")
    if rb and sb:
        if float(sb)<=float(rb):score+=20;reasons.append("budget")
        elif (float(sb)-float(rb))/float(rb)<=0.15:score+=10;reasons.append("budget near")
    if supply.get("status")=="verified":score+=5;reasons.append("verified")
    return min(score,100),", ".join(reasons)

def rebuild_matches():
    require_db();init_db()
    with engine.begin() as c:
        reqs=c.execute(text("SELECT * FROM wai_listings WHERE transaction='REQUIREMENT' AND status<>'rejected'")).mappings().all()
        supply=c.execute(text("SELECT * FROM wai_listings WHERE transaction IN ('SALE','RENT') AND status<>'rejected' AND duplicate_of IS NULL")).mappings().all()
        c.execute(text("DELETE FROM wai_matches"))
        for r in reqs:
            rows=[]
            for s in supply:
                score,reason=score_match(r,s)
                if score>=40:rows.append((score,reason,s))
            rows.sort(key=lambda x:x[0],reverse=True)
            for score,reason,s in rows[:100]:
                c.execute(text("""
                    INSERT INTO wai_matches(id,requirement_listing_id,supply_listing_id,match_score,match_reason)
                    VALUES(:id,:r,:s,:score,:reason)
                    ON CONFLICT(requirement_listing_id,supply_listing_id) DO UPDATE SET
                      match_score=EXCLUDED.match_score,match_reason=EXCLUDED.match_reason
                """),{"id":uuid.uuid4(),"r":r["id"],"s":s["id"],"score":score,"reason":reason})
    return True

def shell(title,body,active="Dashboard"):
    nav=[
      ("Dashboard","/whatsapp-capture/intelligence"),
      ("WhatsApp Sources","/whatsapp-capture/intelligence/sources"),
      ("Property Database","/whatsapp-capture/intelligence/properties"),
      ("Requirements","/whatsapp-capture/intelligence/requirements"),
      ("AI Matches","/whatsapp-capture/intelligence/matches"),
      ("Verification","/whatsapp-capture/intelligence/verification"),
      ("System Health","/whatsapp-capture/intelligence/health"),
    ]
    links="".join(f'<a class="{"active" if n==active else ""}" href="{u}">{esc(n)}</a>' for n,u in nav)
    return f"""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1">
<title>{esc(title)}</title><style>
*{{box-sizing:border-box}}body{{margin:0;background:#f7f8fa;color:#101828;font-family:Arial,sans-serif}}
header{{background:#111827;color:#fff;padding:18px 22px}}header h1{{margin:0;font-size:22px}}
nav{{display:flex;gap:6px;flex-wrap:wrap;background:#fff;padding:10px 16px;border-bottom:1px solid #e5e7eb;position:sticky;top:0;z-index:10}}
nav a{{text-decoration:none;color:#374151;padding:9px 11px;border-radius:7px}}nav a.active,nav a:hover{{background:#111827;color:#fff}}
main{{max-width:1800px;margin:18px auto;padding:0 16px}}
.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:12px}}
.card{{background:white;border:1px solid #e5e7eb;border-radius:10px;padding:14px}}
.num{{font-size:28px;font-weight:700}}.muted{{color:#667085}}
.btn{{display:inline-block;background:#111827;color:white;padding:9px 12px;border-radius:7px;text-decoration:none;border:0;cursor:pointer}}
.green{{background:#047857}}.blue{{background:#1d4ed8}}.red{{background:#b91c1c}}
.scroll{{overflow:auto;max-height:74vh;border:1px solid #e5e7eb;border-radius:8px;background:#fff}}
table{{border-collapse:collapse;width:100%;font-size:13px;background:white}}th,td{{padding:9px;border-bottom:1px solid #e5e7eb;vertical-align:top;text-align:left}}
th{{position:sticky;top:55px;background:#f3f4f6;z-index:2;white-space:nowrap}}
.raw{{min-width:420px;white-space:pre-wrap;line-height:1.35}}.pill{{padding:3px 7px;background:#eef2ff;border-radius:999px}}
</style></head><body>
<header><h1>WhatsApp Property Intelligence System</h1><div style="color:#cbd5e1">Structured · Verified · Searchable · One property per row</div></header>
<nav>{links}<a href="/whatsapp-capture">← WhatsApp Group</a></nav><main>{body}</main></body></html>"""

@router.get("",response_class=HTMLResponse)
def dashboard():
    require_db();init_db()
    with engine.begin() as c:
        vals={
          "Raw Messages":c.execute(text("SELECT COUNT(*) FROM wai_raw_messages")).scalar() or 0,
          "Listings":c.execute(text("SELECT COUNT(*) FROM wai_listings WHERE transaction IN ('SALE','RENT') AND duplicate_of IS NULL")).scalar() or 0,
          "Requirements":c.execute(text("SELECT COUNT(*) FROM wai_listings WHERE transaction='REQUIREMENT' AND duplicate_of IS NULL")).scalar() or 0,
          "Pending Verification":c.execute(text("SELECT COUNT(*) FROM wai_listings WHERE status='unverified' AND confidence_score>=70")).scalar() or 0,
          "Verified":c.execute(text("SELECT COUNT(*) FROM wai_listings WHERE status='verified'")).scalar() or 0,
          "Matches":c.execute(text("SELECT COUNT(*) FROM wai_matches")).scalar() or 0,
        }
        run=c.execute(text("SELECT * FROM wai_pipeline_runs ORDER BY started_at DESC LIMIT 1")).mappings().first()
    cards="".join(f'<div class=card><div class=muted>{esc(k)}</div><div class=num>{v}</div></div>' for k,v in vals.items())
    run_html=""
    if run:
        run_html=f"<div class=card><b>Last pipeline run</b><br>Messages: {run['messages_ingested']} · Extracted: {run['messages_extracted']} · Listing rows: {run['listing_rows_created']} · Requirement rows: {run['requirement_rows_created']} · Failed: {run['failed']}</div>"
    body=f"""<div style='display:flex;justify-content:space-between;gap:10px;flex-wrap:wrap;align-items:center'>
      <div><h2>Dashboard</h2><p class=muted>Raw message → rule extraction → optional LLM repair → verification → matching.</p></div>
      <div><a class='btn blue' href='/whatsapp-capture/intelligence/accounts'>WHATSAPP NUMBERS + AI STATUS</a>
      <a class='btn green' href='/whatsapp-capture/intelligence/run-pipeline'>RUN / REBUILD PIPELINE</a>
      <a class='btn blue' href='/whatsapp-capture/intelligence/export.xlsx'>EXPORT EXCEL</a></div></div>
      <div class=grid>{cards}</div><br>{run_html}"""
    return HTMLResponse(shell("Dashboard",body,"Dashboard"))

@router.get("/run-pipeline")
def run_pipeline():
    with engine.connect() as lock_conn:
        locked=bool(lock_conn.execute(text("SELECT pg_try_advisory_lock(:k)"),{"k":918811955}).scalar())
        if not locked:
            return RedirectResponse("/whatsapp-capture/intelligence/accounts",303)
        try:
            ingest_current_whatsapp_source()
            rebuild_matches()
        finally:
            try:
                lock_conn.execute(text("SELECT pg_advisory_unlock(:k)"),{"k":918811955})
                lock_conn.commit()
            except Exception:
                pass
    return RedirectResponse("/whatsapp-capture/intelligence",303)

@router.get("/sources",response_class=HTMLResponse)
def sources():
    require_db();init_db()
    with engine.begin() as c:
        rows=c.execute(text("""
          SELECT g.name,g.region,g.is_active,COUNT(r.id) raw_count,MAX(r.ingested_at) last_ingested
          FROM wai_groups g LEFT JOIN wai_raw_messages r ON r.group_id=g.id
          GROUP BY g.id,g.name,g.region,g.is_active ORDER BY raw_count DESC,g.name
        """)).mappings().all()
    trs="".join(f"<tr><td>{esc(r['name'])}</td><td>{esc(r['region'])}</td><td>{'ACTIVE' if r['is_active'] else 'PAUSED'}</td><td>{r['raw_count']}</td><td>{esc(r['last_ingested'])}</td></tr>" for r in rows)
    return HTMLResponse(shell("Sources",f"<h2>WhatsApp Sources</h2><div class=scroll><table><tr><th>Group</th><th>Region</th><th>Status</th><th>Raw Messages</th><th>Last Ingested</th></tr>{trs}</table></div>","WhatsApp Sources"))

@router.get("/properties",response_class=HTMLResponse)
def properties():
    require_db();init_db()
    with engine.begin() as c:
        rows=c.execute(text("""
          SELECT l.*,c.phone,c.display_name,c.firm_name
          FROM wai_listings l LEFT JOIN wai_contacts c ON c.id=l.contact_id
          WHERE l.transaction IN ('SALE','RENT') AND l.duplicate_of IS NULL
          ORDER BY l.created_at DESC LIMIT 5000
        """)).mappings().all()
    trs="".join(
      f"<tr><td class=raw>{esc(r['raw_listing_text'] or r['summary'])}</td><td>{esc(r['phone'])}</td><td>{esc(r['budget_text'])}</td>"
      f"<td>{esc(r['area_text'])}</td><td>{esc(r['source_group_name'])}</td><td>{esc(r['location'])}</td><td>{esc(r['property_type'])}</td>"
      f"<td>{esc(r['transaction'])}</td><td>{esc(r['poster_name'] or r['display_name'])}</td><td>{float(r['confidence_score'] or 0):.0f}%</td><td>{esc(r['status'])}</td></tr>"
      for r in rows
    )
    body=f"""<h2>Property Database</h2><p class=muted>Excel-style grid. Every distinct property/unit/option is a separate row.</p>
    <div class=scroll><table><tr><th>Raw Property Details</th><th>Contact No.</th><th>Budget / Rent</th><th>Area</th><th>Source Group</th>
    <th>Location</th><th>Property Type</th><th>Transaction</th><th>Poster / Broker</th><th>Confidence</th><th>Status</th></tr>{trs}</table></div>"""
    return HTMLResponse(shell("Property Database",body,"Property Database"))

@router.get("/requirements",response_class=HTMLResponse)
def requirements():
    require_db();init_db()
    with engine.begin() as c:
        rows=c.execute(text("""
          SELECT l.*,c.phone,c.display_name
          FROM wai_listings l LEFT JOIN wai_contacts c ON c.id=l.contact_id
          WHERE l.transaction='REQUIREMENT' AND l.duplicate_of IS NULL
          ORDER BY l.created_at DESC LIMIT 3000
        """)).mappings().all()
    trs="".join(
      f"<tr><td class=raw>{esc(r['raw_listing_text'] or r['summary'])}</td><td>{esc(r['phone'])}</td><td>{esc(r['budget_text'])}</td>"
      f"<td>{esc(r['area_text'])}</td><td>{esc(r['source_group_name'])}</td><td>{esc(r['location'])}</td><td>{esc(r['property_type'])}</td>"
      f"<td>{esc(r['display_name'] or r['poster_name'])}</td><td>{float(r['confidence_score'] or 0):.0f}%</td><td>{esc(r['status'])}</td>"
      f"<td><a class=btn href='/whatsapp-capture/intelligence/requirements/{r['id']}/matches'>Matches</a></td></tr>"
      for r in rows
    )
    body=f"""<h2>Requirements</h2><p class=muted>WhatsApp buyer/tenant requirements in the same Excel-style structure.</p>
    <div class=scroll><table><tr><th>Raw Requirement Details</th><th>Contact No.</th><th>Budget</th><th>Area</th><th>Source Group</th>
    <th>Location</th><th>Property Type</th><th>Client / Broker</th><th>Confidence</th><th>Status</th><th></th></tr>{trs}</table></div>"""
    return HTMLResponse(shell("Requirements",body,"Requirements"))

@router.get("/verification",response_class=HTMLResponse)
def verification():
    require_db();init_db()
    with engine.begin() as c:
        rows=c.execute(text("""
          SELECT l.*,c.phone FROM wai_listings l LEFT JOIN wai_contacts c ON c.id=l.contact_id
          WHERE l.status='unverified' AND l.confidence_score>=70
          ORDER BY l.confidence_score DESC,l.created_at DESC LIMIT 1500
        """)).mappings().all()
    trs="".join(
      f"<tr><td class=raw>{esc(r['raw_listing_text'] or r['summary'])}</td><td>{esc(r['phone'])}</td><td>{esc(r['location'])}</td>"
      f"<td>{esc(r['property_type'])}</td><td>{esc(r['transaction'])}</td><td>{float(r['confidence_score'] or 0):.0f}%</td>"
      f"<td><a class='btn green' href='/whatsapp-capture/intelligence/verify/{r['id']}/approve'>Approve</a> "
      f"<a class='btn red' href='/whatsapp-capture/intelligence/verify/{r['id']}/reject'>Reject</a></td></tr>"
      for r in rows
    )
    return HTMLResponse(shell("Verification",f"<h2>Verification Queue</h2><div class=scroll><table><tr><th>Raw Details</th><th>Contact</th><th>Location</th><th>Type</th><th>Transaction</th><th>Confidence</th><th>Action</th></tr>{trs}</table></div>","Verification"))

@router.get("/verify/{listing_id}/{action}")
def verify(listing_id:str,action:str):
    require_db();init_db()
    if action not in ("approve","reject"):raise HTTPException(400,"Invalid action")
    status="verified" if action=="approve" else "rejected"
    with engine.begin() as c:
        c.execute(text("UPDATE wai_listings SET status=:s,verified_at=CASE WHEN :s='verified' THEN NOW() ELSE verified_at END,verified_by='team' WHERE id=:id"),
                  {"s":status,"id":listing_id})
        c.execute(text("""
          INSERT INTO wai_verification_log(id,listing_id,action,actor) VALUES(:id,:lid,:a,'team')
        """),{"id":uuid.uuid4(),"lid":listing_id,"a":"approved" if action=="approve" else "rejected"})
    return RedirectResponse("/whatsapp-capture/intelligence/verification",303)

@router.get("/matches",response_class=HTMLResponse)
def matches():
    require_db();init_db();rebuild_matches()
    with engine.begin() as c:
        rows=c.execute(text("""
          SELECT m.match_score,m.match_reason,
                 r.raw_listing_text req_text,r.location req_loc,
                 s.raw_listing_text supply_text,s.location supply_loc,s.budget_text,s.area_text,s.source_group_name,
                 c.phone
          FROM wai_matches m
          JOIN wai_listings r ON r.id=m.requirement_listing_id
          JOIN wai_listings s ON s.id=m.supply_listing_id
          LEFT JOIN wai_contacts c ON c.id=s.contact_id
          ORDER BY m.match_score DESC LIMIT 3000
        """)).mappings().all()
    trs="".join(
      f"<tr><td><span class=pill><b>{float(r['match_score'] or 0):.0f}%</b></span></td><td class=raw>{esc(r['req_text'])}</td>"
      f"<td class=raw>{esc(r['supply_text'])}</td><td>{esc(r['phone'])}</td><td>{esc(r['budget_text'])}</td><td>{esc(r['area_text'])}</td>"
      f"<td>{esc(r['source_group_name'])}</td><td>{esc(r['supply_loc'])}</td><td>{esc(r['match_reason'])}</td></tr>"
      for r in rows
    )
    body=f"<h2>AI Matches</h2><div class=scroll><table><tr><th>Match %</th><th>Requirement</th><th>Property</th><th>Contact No.</th><th>Budget/Rent</th><th>Area</th><th>Source Group</th><th>Location</th><th>Why Matched</th></tr>{trs}</table></div>"
    return HTMLResponse(shell("AI Matches",body,"AI Matches"))

@router.get("/requirements/{req_id}/matches",response_class=HTMLResponse)
def req_matches(req_id:str):
    require_db();init_db();rebuild_matches()
    with engine.begin() as c:
        req=c.execute(text("SELECT * FROM wai_listings WHERE id=:id AND transaction='REQUIREMENT'"),{"id":req_id}).mappings().first()
        if not req:raise HTTPException(404,"Requirement not found")
        rows=c.execute(text("""
          SELECT m.*,s.*,c.phone
          FROM wai_matches m JOIN wai_listings s ON s.id=m.supply_listing_id
          LEFT JOIN wai_contacts c ON c.id=s.contact_id
          WHERE m.requirement_listing_id=:id ORDER BY m.match_score DESC
        """),{"id":req_id}).mappings().all()
    trs="".join(
      f"<tr><td><b>{float(r['match_score'] or 0):.0f}%</b></td><td class=raw>{esc(r['raw_listing_text'])}</td><td>{esc(r['phone'])}</td>"
      f"<td>{esc(r['budget_text'])}</td><td>{esc(r['area_text'])}</td><td>{esc(r['source_group_name'])}</td><td>{esc(r['location'])}</td><td>{esc(r['match_reason'])}</td></tr>"
      for r in rows
    )
    body=f"<div class=card><b>Requirement</b><div class=raw>{esc(req['raw_listing_text'])}</div></div><br><div class=scroll><table><tr><th>Match %</th><th>Raw Property Details</th><th>Contact</th><th>Budget/Rent</th><th>Area</th><th>Source Group</th><th>Location</th><th>Why Matched</th></tr>{trs}</table></div>"
    return HTMLResponse(shell("Requirement Matches",body,"AI Matches"))

@router.get("/health",response_class=HTMLResponse)
def health():
    require_db();init_db()
    checks=[]
    with engine.begin() as c:
        for n,q in [
          ("Raw messages","SELECT COUNT(*) FROM wai_raw_messages"),
          ("Structured listings","SELECT COUNT(*) FROM wai_listings"),
          ("Contacts","SELECT COUNT(*) FROM wai_contacts"),
          ("Verification log","SELECT COUNT(*) FROM wai_verification_log"),
          ("Matches","SELECT COUNT(*) FROM wai_matches"),
          ("Failed extraction","SELECT COUNT(*) FROM wai_raw_messages WHERE extraction_status='failed'"),
        ]:
            try:checks.append((n,"OK",c.execute(text(q)).scalar() or 0))
            except Exception as e:checks.append((n,"ERROR",str(e)[:120]))
    trs="".join(f"<tr><td>{esc(a)}</td><td>{esc(b)}</td><td>{esc(c)}</td></tr>" for a,b,c in checks)
    body=f"<h2>System Health</h2><div class=card>Five-layer pipeline is separated: ingestion → extraction → database → verification → team app/matching.</div><br><div class=scroll><table><tr><th>Component</th><th>Status</th><th>Value</th></tr>{trs}</table></div>"
    return HTMLResponse(shell("Health",body,"System Health"))

@router.get("/export.xlsx")
def export_excel():
    require_db();init_db()
    with engine.begin() as c:
        props=c.execute(text("""
          SELECT l.raw_listing_text,c.phone,l.budget_text,l.area_text,l.source_group_name,l.location,l.property_type,
                 l.transaction,l.poster_name,l.confidence_score,l.status
          FROM wai_listings l LEFT JOIN wai_contacts c ON c.id=l.contact_id
          WHERE l.transaction IN ('SALE','RENT') AND l.duplicate_of IS NULL ORDER BY l.created_at DESC
        """)).mappings().all()
        reqs=c.execute(text("""
          SELECT l.raw_listing_text,c.phone,l.budget_text,l.area_text,l.source_group_name,l.location,l.property_type,
                 l.poster_name,l.confidence_score,l.status
          FROM wai_listings l LEFT JOIN wai_contacts c ON c.id=l.contact_id
          WHERE l.transaction='REQUIREMENT' AND l.duplicate_of IS NULL ORDER BY l.created_at DESC
        """)).mappings().all()
        contacts=c.execute(text("SELECT phone,display_name,firm_name,message_count,first_seen_at,last_seen_at,trust_score FROM wai_contacts ORDER BY last_seen_at DESC")).mappings().all()

    wb=Workbook();wb.remove(wb.active)
    def add(name,headers,rows):
        ws=wb.create_sheet(name)
        for i,h in enumerate(headers,1):
            cell=ws.cell(1,i,h);cell.font=Font(bold=True,color="FFFFFF");cell.fill=PatternFill("solid",fgColor="0F766E");cell.alignment=Alignment(horizontal="center")
        for ri,row in enumerate(rows,2):
            for ci,h in enumerate(headers,1):ws.cell(ri,ci,row.get(h))
        ws.freeze_panes="A2";ws.auto_filter.ref=ws.dimensions
        for i,h in enumerate(headers,1):ws.column_dimensions[get_column_letter(i)].width=55 if "raw" in h else min(max(len(h)+4,14),28)
    add("Property Database",["raw_listing_text","phone","budget_text","area_text","source_group_name","location","property_type","transaction","poster_name","confidence_score","status"],props)
    add("Requirements",["raw_listing_text","phone","budget_text","area_text","source_group_name","location","property_type","poster_name","confidence_score","status"],reqs)
    add("Contacts",["phone","display_name","firm_name","message_count","first_seen_at","last_seen_at","trust_score"],contacts)
    bio=io.BytesIO();wb.save(bio);bio.seek(0)
    return StreamingResponse(bio,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition":"attachment; filename=Alliance_WhatsApp_Property_Intelligence.xlsx"})

# Database Improvement V2: triage + source health
try:
    from whatsapp_database_improvement_v2 import install as _install_db_improvement_v2
    _install_db_improvement_v2(router, engine, require_db, init_db, shell, esc)
except Exception as _dbv2_error:
    print('WAI Database Improvement V2 warning:', _dbv2_error)
# WhatsApp AI Source Control: accounts + mapping + automatic segregation
try:
    from whatsapp_ai_source_control import install as _install_ai_source_control
    _install_ai_source_control(
        router, engine, require_db, init_db, shell, esc,
        ingest_current_whatsapp_source, rebuild_matches
    )
except Exception as _source_control_error:
    print("WAI AI Source Control warning:", _source_control_error)
