
from __future__ import annotations
import csv, io, re, tempfile, os
from fastapi import Request, UploadFile, File, Form
from fastapi.responses import HTMLResponse
from openpyxl import load_workbook
from alliance_v33_contact_vault import ensure_schema, upsert

MODULE_VERSION="3.3A-PHONE-IMPORT-TEAM-OPERATIONS"
BUCKETS=["PHONE_IMPORT","WHATSAPP_GROUP","MAGAZINE","NEWSPAPER","HOSPITALITY_BOT","RETAIL_EXPANSION","MANUAL","OTHER_DATABASE"]
CATEGORIES=["CAFE","LOUNGE","RESTAURANT","BANQUET","CLUB","GUEST_HOUSE","HOTEL","BAR","CLOUD_KITCHEN","JEWELLERY","FASHION","FOOTWEAR","BEAUTY","ELECTRONICS","GROCERY","QSR","FITNESS","HOME_DECOR","SPECIALITY_RETAIL","OTHER"]

def _n(v): return re.sub(r"\s+"," ",str(v or "")).strip()
def _pick(row,*names):
    low={str(k).strip().lower():v for k,v in row.items()}
    for name in names:
        v=low.get(name.lower())
        if v not in (None,""): return v
    return None

def parse_vcf(text_data):
    rows=[];current={}
    for raw in text_data.replace("\r\n","\n").split("\n"):
        line=raw.strip()
        if line.upper()=="BEGIN:VCARD": current={}
        elif line.upper()=="END:VCARD":
            name=current.get("FN") or current.get("N") or "Phone Contact"
            phones=current.get("TEL",[]); emails=current.get("EMAIL",[])
            if not isinstance(phones,list):phones=[phones]
            if not isinstance(emails,list):emails=[emails]
            if not phones and not emails:continue
            if phones:
                for p in phones:
                    rows.append({"contact_name":name,"business_name":current.get("ORG") or name,"contact_phone":p,"email":emails[0] if emails else None,"role_title":current.get("TITLE")})
            else:
                rows.append({"contact_name":name,"business_name":current.get("ORG") or name,"email":emails[0],"role_title":current.get("TITLE")})
        elif ":" in line:
            left,val=line.split(":",1);key=left.split(";",1)[0].upper();val=val.replace("\\n"," ")
            if key in {"TEL","EMAIL"}:current.setdefault(key,[]).append(val)
            elif key=="N":
                parts=[x for x in val.split(";") if x];current[key]=" ".join(reversed(parts[:2])) if parts else val
            elif key in {"FN","ORG","TITLE"}:current[key]=val
    return rows

def parse_csv_bytes(data):
    text_data=data.decode("utf-8-sig",errors="replace")
    if not text_data.strip():return []
    try:dialect=csv.Sniffer().sniff(text_data[:4096],delimiters=",;\t")
    except Exception:dialect=csv.excel
    reader=csv.DictReader(io.StringIO(text_data),dialect=dialect)
    out=[]
    for r in reader:
        out.append({
            "contact_name":_pick(r,"name","contact name","contact_name","full name"),
            "business_name":_pick(r,"business","business name","company","company name","brand","brand name","organization","organisation"),
            "contact_phone":_pick(r,"phone","mobile","contact","contact no","contact number","telephone","tel","whatsapp"),
            "whatsapp_phone":_pick(r,"whatsapp","whatsapp phone","whatsapp_phone"),
            "email":_pick(r,"email","email id","email address"),
            "role_title":_pick(r,"role","designation","title","job title"),
            "location":_pick(r,"location","address","area"),"city":_pick(r,"city"),
            "website":_pick(r,"website","url"),"category":_pick(r,"category","business type","type"),
        })
    return out

def parse_xlsx_bytes(data):
    with tempfile.NamedTemporaryFile(suffix=".xlsx",delete=False) as f:
        f.write(data);tmp=f.name
    try:
        wb=load_workbook(tmp,read_only=True,data_only=True);ws=wb[wb.sheetnames[0]]
        values=list(ws.iter_rows(values_only=True))
        if not values:return []
        headers=[_n(x).lower() for x in values[0]];out=[]
        for vals in values[1:]:
            r={headers[i]:vals[i] if i<len(vals) else None for i in range(len(headers))}
            out.append({
                "contact_name":_pick(r,"name","contact name","contact_name","full name"),
                "business_name":_pick(r,"business","business name","company","company name","brand","brand name","organization","organisation"),
                "contact_phone":_pick(r,"phone","mobile","contact","contact no","contact number","telephone","tel","whatsapp"),
                "whatsapp_phone":_pick(r,"whatsapp","whatsapp phone","whatsapp_phone"),
                "email":_pick(r,"email","email id","email address"),
                "role_title":_pick(r,"role","designation","title","job title"),
                "location":_pick(r,"location","address","area"),"city":_pick(r,"city"),
                "website":_pick(r,"website","url"),"category":_pick(r,"category","business type","type"),
            })
        return out
    finally:
        try:os.unlink(tmp)
        except Exception:pass

def register(core):
    app,engine=core.app,core.engine
    @app.get("/api/v3/contacts/import/status")
    def status(req:Request):
        if hasattr(core,"need_login"):core.need_login(req)
        return {"version":MODULE_VERSION,"status":"OK","formats":["VCF","CSV","XLSX"],"permanent_contact_vault":True,"duplicate_merge":True,"automatic_marketing_approval":False}
    @app.get("/v3/contact-import",response_class=HTMLResponse)
    def import_page(req:Request):
        if hasattr(core,"need_login"):core.need_login(req)
        options="".join(f"<option>{x}</option>" for x in BUCKETS);cats="".join(f"<option>{x}</option>" for x in CATEGORIES)
        return HTMLResponse(f"""<!doctype html><html><head><meta name="viewport" content="width=device-width,initial-scale=1"><style>
body{{font-family:Arial;background:#f5f7fa;margin:0;color:#172033}}.wrap{{max-width:760px;margin:24px auto;padding:16px}}.card{{background:white;border-radius:16px;padding:24px;box-shadow:0 2px 12px #0001}}
label{{display:block;font-weight:700;margin-top:16px}}select,input,button{{width:100%;padding:12px;margin-top:7px;border:1px solid #ccd3dd;border-radius:9px;box-sizing:border-box}}
button{{background:#172033;color:white;font-weight:700;border:0;margin-top:22px}}.note{{background:#fff7df;padding:12px;border-radius:9px;margin-top:18px}}</style></head><body><div class="wrap"><div class="card">
<h1>Import Contacts from Phone</h1><p>Export contacts from your phone as VCF/vCard, CSV or Excel, then upload here.</p>
<form action="/api/v3/contacts/import-file" method="post" enctype="multipart/form-data">
<label>Contact file</label><input type="file" name="file" accept=".vcf,.vcard,.csv,.xlsx" required>
<label>Source bucket</label><select name="source_bucket">{options}</select>
<label>Category</label><select name="category">{cats}</select>
<label>Source name / Team note</label><input name="source_name" placeholder="Example: Team iPhone Aug 2026">
<button type="submit">Import Contacts</button></form><div class="note"><b>Important:</b> Imported contacts are stored for review. They are not automatically approved for WhatsApp marketing.</div>
</div></div></body></html>""")
    @app.post("/api/v3/contacts/import-file")
    async def import_file(req:Request,file:UploadFile=File(...),source_bucket:str=Form("PHONE_IMPORT"),category:str=Form("OTHER"),source_name:str=Form("PHONE_CONTACT_IMPORT")):
        if hasattr(core,"need_login"):core.need_login(req)
        ensure_schema(engine);data=await file.read();name=(file.filename or "").lower()
        if name.endswith((".vcf",".vcard")):rows=parse_vcf(data.decode("utf-8-sig",errors="replace"));fmt="VCF"
        elif name.endswith(".csv"):rows=parse_csv_bytes(data);fmt="CSV"
        elif name.endswith(".xlsx"):rows=parse_xlsx_bytes(data);fmt="XLSX"
        else:return {"version":MODULE_VERSION,"status":"UNSUPPORTED_FILE","allowed":[".vcf",".vcard",".csv",".xlsx"]}
        bucket=_n(source_bucket).upper();bucket=bucket if bucket in BUCKETS else "PHONE_IMPORT"
        default_cat=_n(category).upper();default_cat=default_cat if default_cat in CATEGORIES else "OTHER"
        saved=skipped=0
        for i,r in enumerate(rows,1):
            cat=_n(r.get("category")).upper();cat=cat if cat in CATEGORIES else default_cat
            cid=upsert(engine,{**r,"category":cat,"verification_status":"UNVERIFIED","whatsapp_status":"NOT_VERIFIED","marketing_status":"REVIEW_REQUIRED"},
                {"source_type":bucket,"source_name":source_name or "PHONE_CONTACT_IMPORT","source_record_id":f"{file.filename}:{i}","evidence_text":f"Manual phone import from {file.filename}"})
            if cid:saved+=1
            else:skipped+=1
        return {"version":MODULE_VERSION,"status":"OK","file":file.filename,"format":fmt,"source_bucket":bucket,"default_category":default_cat,
                "rows_detected":len(rows),"contacts_saved_or_merged":saved,"rows_skipped_no_phone_or_email":skipped,
                "marketing_status":"REVIEW_REQUIRED","auto_whatsapp_approval":False}
