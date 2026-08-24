# ENTITY-PURITY-V3: one WhatsApp entity per row; combined-row safety quarantine
import re, io, json, uuid, hashlib, html
from datetime import datetime
from fastapi import APIRouter, Form
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from whatsapp_property_intelligence_final import engine, require_db, esc

router = APIRouter(prefix="/clean", tags=["WhatsApp Clean Database Final"])

PROPERTY_TYPES = [
    ("Warehouse / Industrial", ["warehouse","industrial","factory","godown","udyog"]),
    ("Commercial Showroom", ["showroom"]),("Commercial Shop", ["shop","retail outlet"]),
    ("Office", ["office","workspace"]),("Farmhouse", ["farmhouse","farm house","farm land"]),
    ("Banquet", ["banquet"]),("Hotel", ["hotel","resort"]),("Guest House", ["guest house","guesthouse"]),
    ("Restaurant", ["restaurant","restro","resto bar","restobar"]),("Cafe", ["cafe","café"]),
    ("Club", ["club","lounge"]),("Independent House / Villa", ["villa","kothi","independent house","bungalow","bunglow","row house"]),
    ("Apartment", ["apartment","flat","bhk","builder floor","floor with terrace"]),
    ("Plot / Land", ["plot","land","acre","sqyd","gaj"]),("Commercial Space", ["commercial space","commercial building","commercial","retail space"])
]
REQ_WORDS=["requirement","required","require ","wanted","looking for","need ","buyer looking","tenant looking","client looking","want ","wanted on","required for","need on"]
SUPPLY_SALE=["for sale","available for sale","on sale","sale option","mandate sale","sale mandate","selling","sell ","auction","preleased","pre-leased","sale -","sale:"]
SUPPLY_RENT=["for rent","available on rent","available for rent","for lease","available on lease","lease option","rent -","rent:","rent @","rent@"]
NOISE_WORDS=["high court","supreme court","breaking news","we're hiring","we are hiring","join our team","send us your cv","job opening","vacancy","follow our","facebook.com","instagram.com","group chat invite","view channel","subscribe","likes and follow"]
PHONE_RE=re.compile(r"(?<!\d)(?:(?:\+?91)[\s-]?)?([6-9](?:[\s-]?\d){9})(?!\d)")
AREA_RE=re.compile(r"(?P<a>\d[\d,]*(?:\.\d+)?)\s*(?:-|to|–)?\s*(?P<b>\d[\d,]*(?:\.\d+)?)?\s*(?P<u>sq\.?\s*ft|sqft|sft|sq\.?\s*yds?|sqyds?|sqyrd|sq\.?\s*m(?:tr|trs)?|sqm(?:tr|trs)?|gaj|yards?|yds?|acre?s?|bigha|mtr?s?)",re.I)
BUDGET_RE=re.compile(r"(?:(?:budget|price|demand|asking|rent|reserve price|amount|range)\s*(?:is|@|:|-|=)?\s*)(?:₹|rs\.?|inr)?\s*(?P<a>\d[\d,]*(?:\.\d+)?)(?:\s*(?:-|to|–)\s*(?P<b>\d[\d,]*(?:\.\d+)?))?\s*(?P<u>cr|crore?s?|lac?s?|lakh?s?|l|k|thousand)?",re.I)

KNOWN_LOCATIONS=["Siolim","Assagao","Anjuna","Vagator","Morjim","Mandrem","Parra","Arpora","Calangute","Candolim","Baga","Porvorim","Panjim","Panaji","Miramar","Caranzalem","Taleigao","Dona Paula","Bambolim","Saligao","Sangolda","Guirim","Old Goa","Campal","Mapusa","Margao","Colva","Nerul","Reis Magos","Pilerne","Moira","Ribandar","Kadamba Plateau","St. Inez","Fontainhas","Socorro","Merces","Chimbel","Betim","Corlim","Carambolim","Delhi","South Delhi","Defence Colony","Greater Kailash","GK-1","GK-2","Vasant Kunj","Vasant Vihar","Saket","Green Park","Hauz Khas","Janakpuri","Dwarka","Karol Bagh","Rohini","Rajouri Garden","Punjabi Bagh","Paschim Vihar","Pitampura","Vikaspuri","Tilak Nagar","Subhash Nagar","Moti Nagar","Friends Colony","Maharani Bagh","South Extension","Lajpat Nagar","Jor Bagh","Anand Lok","Niti Bagh","Panchsheel Park","Shanti Niketan","Sundar Nagar","Golf Links","Gulmohar Park","Connaught Place","East of Kailash","SDA","Safdarjung Enclave","Gurugram","Golf Course Road","Sohna Road","DLF Phase-1","DLF Phase-2","DLF Phase-4","Palam Vihar","Udyog Vihar","Noida","Faridabad","Ghaziabad","Vaishali","Indirapuram","Hapur"]

SCHEMA="""
CREATE TABLE IF NOT EXISTS wai_clean_records(
 id UUID PRIMARY KEY, source_message_id UUID, source_group TEXT, record_type TEXT NOT NULL, transaction TEXT,
 raw_details TEXT NOT NULL, contact_no TEXT, all_contacts JSONB DEFAULT '[]'::jsonb,
 budget_text TEXT, budget_min NUMERIC, budget_max NUMERIC, budget_period TEXT,
 area_text TEXT, area_min NUMERIC, area_max NUMERIC, area_unit TEXT, area_sqft_min NUMERIC, area_sqft_max NUMERIC,
 location TEXT, all_locations JSONB DEFAULT '[]'::jsonb, property_type TEXT, person_name TEXT, firm_name TEXT,
 confidence NUMERIC, status TEXT DEFAULT 'unverified', rejection_reason TEXT, source_fingerprint TEXT UNIQUE,
 source_created_at TIMESTAMPTZ, processed_at TIMESTAMPTZ DEFAULT NOW(), edited_at TIMESTAMPTZ,
 source_entity_index INTEGER, source_entity_count INTEGER
);
CREATE INDEX IF NOT EXISTS idx_wcr_type ON wai_clean_records(record_type);
CREATE INDEX IF NOT EXISTS idx_wcr_location ON wai_clean_records(location);
CREATE INDEX IF NOT EXISTS idx_wcr_phone ON wai_clean_records(contact_no);
CREATE TABLE IF NOT EXISTS wai_clean_runs(
 id UUID PRIMARY KEY, started_at TIMESTAMPTZ DEFAULT NOW(), completed_at TIMESTAMPTZ,
 source_messages INT DEFAULT 0, inventory_rows INT DEFAULT 0, requirement_rows INT DEFAULT 0,
 rejected_rows INT DEFAULT 0, failed INT DEFAULT 0, notes TEXT
);
"""

MIGRATIONS=[
    "ALTER TABLE wai_clean_records ADD COLUMN IF NOT EXISTS edited_at TIMESTAMPTZ",
    "ALTER TABLE wai_clean_records ADD COLUMN IF NOT EXISTS source_entity_index INTEGER",
    "ALTER TABLE wai_clean_records ADD COLUMN IF NOT EXISTS source_entity_count INTEGER",
]

def init_clean_db():
    require_db()
    with engine.begin() as c:
        for stmt in [x.strip() for x in SCHEMA.split(";") if x.strip()]:
            c.execute(text(stmt))
        for stmt in MIGRATIONS:
            c.execute(text(stmt))

def norm(v): return re.sub(r"\s+"," ",str(v or "").replace("\u00a0"," ")).strip()

def clean_phone(v):
    d=re.sub(r"\D","",str(v or ""))
    if len(d)==12 and d.startswith("91"): d=d[2:]
    if len(d)==11 and d.startswith("0"): d=d[1:]
    return "+91 "+d[:5]+" "+d[5:] if len(d)==10 and d[0] in "6789" else ""

def phones(txt):
    out=[]
    for m in PHONE_RE.finditer(txt or ""):
        p=clean_phone(m.group(0))
        if p and p not in out: out.append(p)
    return out

def merge_phones(*values):
    out=[]
    for value in values:
        if isinstance(value,(list,tuple)):
            candidates=value
        else:
            candidates=phones(str(value or ""))
            if not candidates:
                p=clean_phone(value)
                candidates=[p] if p else []
        for p in candidates:
            if p and p not in out: out.append(p)
    return out

def phone_line(row):
    vals=row.get("all_contacts") or []
    if isinstance(vals,str):
        try: vals=json.loads(vals)
        except Exception: vals=[]
    vals=merge_phones(vals,row.get("contact_no"))
    return " | ".join(vals)

def fmt_date(v):
    if not v:return ""
    try:return v.astimezone().strftime("%d-%m-%Y %I:%M %p")
    except Exception:return str(v)

def detect_type(txt):
    """Classify the actual asset, not incidental amenity/project words."""
    low=(" "+str(txt or "").lower()+" ")
    # Strong object evidence first. This prevents "Club Membership" => Club
    # and "Central Park Resorts 4BHK Apartment" => Hotel.
    if re.search(r"\b(?:studio\s+apartment|apartment|flat|builder\s+floor|floor\s+with\s+terrace|\d(?:\.\d)?\s*bhk|[1-9]\s*bhk)\b",low):
        return "Apartment"
    if re.search(r"\b(?:villa|kothi|independent\s+house|bungalow|bunglow|row\s+house)\b",low):
        return "Independent House / Villa"
    if re.search(r"\b(?:warehouse|industrial|factory|godown|industrial\s+shed)\b",low):
        return "Warehouse / Industrial"
    if re.search(r"\b(?:showroom)\b",low):
        return "Commercial Showroom"
    if re.search(r"\b(?:office|workspace|business\s+centre|business\s+center)\b",low):
        return "Office"
    if re.search(r"\b(?:farmhouse|farm\s+house)\b",low):
        return "Farmhouse"
    if re.search(r"\b(?:banquet|banquet\s+hall)\b",low):
        return "Banquet"
    if re.search(r"\b(?:guest\s*house)\b",low):
        return "Guest House"
    if re.search(r"\b(?:hotel|boutique\s+hotel)\b",low):
        return "Hotel"
    if re.search(r"\b(?:restaurant|restro|resto\s*bar|restobar)\b",low):
        return "Restaurant"
    if re.search(r"\b(?:cafe|café)\b",low):
        return "Cafe"
    # Club/lounges are classified only when they are the asset, not an amenity.
    if re.search(r"\b(?:club|lounge)\b",low) and not re.search(r"\bclub\s+(?:membership|house)\b",low):
        return "Club"
    if re.search(r"\b(?:commercial\s+shop|retail\s+shop|shop|retail\s+outlet)\b",low):
        return "Commercial Shop"
    if re.search(r"\b(?:plot|commercial\s+land|residential\s+land|farm\s+land|land\s+for|acre|acres|bigha)\b",low):
        return "Plot / Land"
    if re.search(r"\b(?:commercial\s+space|commercial\s+building|retail\s+space)\b",low):
        return "Commercial Space"
    return None

def detect_record_type(txt):
    low=(" "+str(txt or "").lower()+" ")
    # Requirement intent wins only when it is actually asking/looking.
    if any(x in low for x in REQ_WORDS):
        return "REQUIREMENT"
    if any(x in low for x in SUPPLY_SALE+SUPPLY_RENT):
        return "INVENTORY"
    # Common WhatsApp variants not covered by exact phrases.
    if re.search(r"\b(?:sale|rent|lease)\b",low) and detect_type(txt):
        return "INVENTORY"
    return None

def detect_transaction(txt,rtype):
    low=(txt or "").lower()
    if rtype=="REQUIREMENT":
        if any(x in low for x in ["rent","lease","tenant","on rent"]): return "RENT"
        if any(x in low for x in ["purchase","buy","buyer","outright","sale"]): return "SALE"
        return "REQUIREMENT"
    if any(x in low for x in SUPPLY_RENT): return "RENT"
    if any(x in low for x in SUPPLY_SALE): return "SALE"
    return None

def locations(txt):
    low=(txt or "").lower();out=[]
    aliases={"gurgaon":"Gurugram","gk1":"GK-1","gk2":"GK-2","donapaula":"Dona Paula","donapula":"Dona Paula","provorim":"Porvorim","caranzhalem":"Caranzalem","caranzalim":"Caranzalem","taligao":"Taleigao","stinez":"St. Inez","kadamba platue":"Kadamba Plateau"}
    for a,c in aliases.items():
        if re.search(r"(?<!\w)"+re.escape(a)+r"(?!\w)",low) and c not in out: out.append(c)
    for loc in sorted(KNOWN_LOCATIONS,key=len,reverse=True):
        if re.search(r"(?<!\w)"+re.escape(loc.lower())+r"(?!\w)",low) and loc not in out: out.append(loc)
    for sec in re.findall(r"\b(?:sector|sec)[\s\-]*([0-9]{1,3}[a-z]?)\b",txt or "",re.I):
        s="Sector "+sec.upper()
        if s not in out: out.append(s)
    return out[:12]

def area_to_sqft(v,u):
    if v is None:return None
    u=(u or "").lower().replace(".","").replace(" ","")
    if u in ("sqyd","sqyds","sqyrd","yards","yard","yds","gaj"): return v*9
    if u.startswith("sqm") or u.startswith("sqmt") or u in ("mtr","mtrs"): return v*10.7639
    if u.startswith("acre"): return v*43560
    if u=="bigha": return v*27000
    return v

def extract_area(txt):
    m=AREA_RE.search(txt or "")
    if not m:return (None,None,None,None,None,None)
    a=float(m.group("a").replace(",",""));b=float((m.group("b") or m.group("a")).replace(",",""));u=m.group("u")
    return (m.group(0),a,b,u,round(area_to_sqft(a,u),2),round(area_to_sqft(b,u),2))

def money_value(n,u):
    v=float(str(n).replace(",",""));u=(u or "").lower()
    if u in ("cr","crore","crores"):v*=10000000
    elif u in ("lac","lacs","lakh","lakhs","l"):v*=100000
    elif u in ("k","thousand"):v*=1000
    return v

def extract_budget(txt):
    textv=str(txt or "")
    # Only monetary labels. Do not treat "range 1437 sqft" as a budget.
    patterns=[
        r"(?P<label>budget|price|demand|asking|rent|reserve\s+price|sale\s+price|amount)\s*(?:is|@|:|-|=)?\s*(?:₹|rs\.?|inr)?\s*(?P<a>\d[\d,]*(?:\.\d+)?)\s*(?:-|to|–)?\s*(?P<b>\d[\d,]*(?:\.\d+)?)?\s*(?P<u>cr|crore?s?|lac?s?|lakh?s?|l|k|thousand)?",
        r"(?:₹|rs\.?|inr)\s*(?P<a>\d[\d,]*(?:\.\d+)?)\s*(?:-|to|–)?\s*(?P<b>\d[\d,]*(?:\.\d+)?)?\s*(?P<u>cr|crore?s?|lac?s?|lakh?s?|l|k|thousand)?"
    ]
    m=None
    for pat in patterns:
        m=re.search(pat,textv,re.I)
        if m: break
    if not m:return (None,None,None,None)
    raw=m.group(0)
    a=money_value(m.group("a"),m.group("u"))
    b=money_value(m.group("b") or m.group("a"),m.group("u"))
    low=textv.lower()
    period="per_month" if any(x in low for x in ["per month","/month","monthly","rent","lease"]) else "total"
    if any(x in low for x in ["psf","per sq ft","/sqft","per sqft"]):period="per_sqft"
    return (raw,a,b,period)

PROPERTY_OBJECT_RE=re.compile(
    r"\b(?:apartment|flat|\d(?:\.\d)?\s*bhk|villa|kothi|row\s+house|plot|land|office|shop|showroom|warehouse|factory|hotel|guest\s*house|restaurant|banquet|farm\s*house|farmhouse|commercial\s+space|builder\s+floor)\b",
    re.I
)
AREA_TOKEN_RE=re.compile(r"\b\d[\d,]*(?:\.\d+)?\s*(?:sq\.?\s*ft|sqft|sft|sq\.?\s*yds?|sqyds?|sqyrd|sq\.?\s*m(?:tr|trs)?|sqm(?:tr|trs)?|gaj|yards?|yds?|acre?s?|bigha)\b",re.I)
PRICE_TOKEN_RE=re.compile(r"(?:₹|rs\.?|inr|\b(?:rent|price|asking|demand|reserve\s+price)\b).{0,18}\d",re.I)

def _looks_like_entity(piece):
    low=str(piece or "").lower()
    return bool(PROPERTY_OBJECT_RE.search(piece or "")) and (
        bool(AREA_TOKEN_RE.search(piece or "")) or bool(PRICE_TOKEN_RE.search(piece or "")) or
        any(x in low for x in SUPPLY_SALE+SUPPLY_RENT+REQ_WORDS)
    )

def _context_prefix(prefix):
    p=norm(prefix)
    if not p:return ""
    # Keep only useful parent context; avoid copying a giant previous entity.
    hints=[]
    low=p.lower()
    if any(x in low for x in REQ_WORDS): hints.append("REQUIREMENT")
    elif any(x in low for x in SUPPLY_RENT) or re.search(r"\brent\b",low): hints.append("FOR RENT")
    elif any(x in low for x in SUPPLY_SALE) or re.search(r"\bsale\b",low): hints.append("FOR SALE")
    pt=detect_type(p)
    if pt: hints.append(pt)
    loc=locations(p)
    if loc: hints.append(", ".join(loc[:3]))
    return " | ".join(hints)

def _inline_numbered_split(textv):
    # Splits "1. Property A ... 2. Property B ..." even when WhatsApp export flattened it to one line.
    marked=re.sub(r"(?<![\d.])\s+(?=(?:\d{1,2})[\.)]\s+(?=[A-Za-z*]))","\n",textv)
    mfirst=re.search(r"(?m)^\s*(?:\d{1,2})[\.)]\s+",marked)
    prefix=marked[:mfirst.start()] if mfirst else ""
    ctx=_context_prefix(prefix)
    parts=[norm(x.strip("* \t")) for x in re.split(r"(?m)(?=^\s*(?:\d{1,2})[\.)]\s+)",marked) if norm(x.strip("* \t"))]
    parts=[x for x in parts if re.match(r"^\s*\d{1,2}[\.)]\s+",x)]
    good=[]
    for x in parts:
        enriched=(ctx+" | "+x) if ctx else x
        if _looks_like_entity(enriched):good.append(enriched)
    return good if len(good)>=2 else []

def _building_split(textv):
    # Common broker format: "FOR SALE OFFICE SPACE. BUILDING - A ... BUILDING - B ..."
    ms=list(re.finditer(r"(?i)\b(?:BUILDING|PROJECT|PROPERTY|OPTION)\s*[-:#]\s*",textv))
    if len(ms)<2:return []
    prefix=textv[:ms[0].start()]
    ctx=_context_prefix(prefix)
    parts=[]
    for i,m in enumerate(ms):
        end=ms[i+1].start() if i+1<len(ms) else len(textv)
        piece=norm(textv[m.start():end])
        enriched=(ctx+" | "+piece) if ctx else piece
        if _looks_like_entity(enriched):parts.append(enriched)
    return parts if len(parts)>=2 else []

def _bullet_project_split(textv):
    # Handles compact project lists:
    # "Heritage One • 1996 Sq.Ft • 3BHK • ₹4.26 Cr Adani Samsara • 1725 Sq.Ft ..."
    marked=re.sub(
        r"(?i)(\b(?:cr|crore|crores|lac|lakh|lakhs)\b)\s+(?=[A-Z][A-Za-z0-9&'()./-]*(?:\s+[A-Z][A-Za-z0-9&'()./-]*){0,5}\s*•\s*\d[\d,]*(?:\.\d+)?\s*(?:sq\.?\s*ft|sqft|sq\.?\s*yds?|sqyds?|sqm|sq\.?\s*m))",
        r"\1\n",textv
    )
    parts=[norm(x) for x in marked.splitlines() if norm(x)]
    if len(parts)<2:return []
    # The first line contains the parent intent; pass that intent/type to later rows.
    ctx=_context_prefix(parts[0])
    good=[]
    for i,x in enumerate(parts):
        enriched=x if i==0 else ((ctx+" | "+x) if ctx else x)
        if _looks_like_entity(enriched):good.append(enriched)
    return good if len(good)>=2 else []

def _section_split(textv):
    # Preserve SALE/RENT/REQUIREMENT sections before entity splitting.
    pat=r"(?i)(?=(?:^|\s)(?:EXCLUSIVE\s+DEALS?\s+ON\s+SALE|AVAILABLE\s+FOR\s+SALE|AVAILABLE\s+FOR\s+RENT|FOR\s+SALE|FOR\s+RENT|FOR\s+LEASE|MANDATE\s+SALE|URGENT(?:LY)?\s+REQUIRE(?:MENT|D)?|IMMEDIATELY\s+REQUIRE|REQUIREMENT\s*[:\-]))"
    parts=[norm(x) for x in re.split(pat,textv) if norm(x)]
    good=[x for x in parts if _looks_like_entity(x)]
    return good if len(good)>=2 else [norm(textv)]

def entity_complexity(piece):
    """Count evidence that a supposedly single row still contains multiple assets."""
    s=str(piece or "")
    numbered=len(re.findall(r"(?<!\d)(?:^|\s)\d{1,2}[\.)]\s+(?=[A-Za-z*])",s))
    building=len(re.findall(r"(?i)\b(?:BUILDING|PROJECT|OPTION)\s*[-:#]",s))
    # Repeated distinct area+BHK/property anchors are a strong multi-entity signal.
    areas=len(AREA_TOKEN_RE.findall(s))
    prices=len(re.findall(r"(?i)\b(?:rent|price|asking|demand|reserve\s+price)\s*[:=@-]?\s*(?:₹|rs\.?|inr)?\s*\d",s))
    return max(numbered,building, min(areas,prices) if areas and prices else 0)

def atomic_segments(raw):
    """Deterministic entity normalizer: one property/requirement per segment."""
    textv=str(raw or "").replace("\r\n","\n").replace("\r","\n")
    textv=re.sub(r"[ \t]+"," ",textv).strip()
    if not textv:return []

    final=[]
    for section in _section_split(textv):
        candidates=(_inline_numbered_split(section) or _building_split(section) or _bullet_project_split(section) or [norm(section)])
        final.extend(candidates)

    # Second pass: if a section still contains inline numbered entities, split again.
    out=[]
    for piece in final:
        sub=_inline_numbered_split(piece)
        out.extend(sub if sub else [piece])

    # De-duplicate while preserving order.
    seen=set();clean=[]
    for x in out:
        x=norm(x)
        if len(x)<12:continue
        key=x.lower()
        if key in seen:continue
        seen.add(key);clean.append(x)
    return clean

def extract_person(seg, fallback=""):
    s=str(seg or "")
    patterns=[
        r"(?i)\b(?:call|contact|whatsapp|connect\s+with|regards?)\s*[:\-]?\s*([A-Z][A-Za-z .&'-]{2,45}?)(?=\s*(?:\+?91[\s-]?)?[6-9]\d)",
        r"(?i)\b([A-Z][A-Za-z .'-]{2,35})\s*[-–]\s*(?:\+?91[\s-]?)?[6-9]\d",
    ]
    for pat in patterns:
        m=re.search(pat,s)
        if m:
            name=norm(m.group(1)).strip(" -:")
            if 2<=len(name)<=60:return name
    fb=norm(fallback)
    return "" if fb.lower() in {"unknown","none","na","n/a"} else fb

def parse_message(row):
    full=row["message_text"] or ""
    inherited=phones(full) or merge_phones(row.get("sender_phone"))
    segments=atomic_segments(full)
    out=[]
    count=len(segments)
    for idx,seg in enumerate(segments,1):
        low=seg.lower()
        if any(x in low for x in NOISE_WORDS):
            out.append({"rtype":"REJECTED","raw":seg,"phone":inherited[0] if inherited else "","phones":inherited,"tx":None,"ptype":None,"locs":[],"area":(None,)*6,"budget":(None,)*4,"confidence":0,"status":"rejected","reason":"News / jobs / social promotion","idx":idx,"entity_count":count});continue

        ptype=detect_type(seg);rtype=detect_record_type(seg);ar=extract_area(seg);bu=extract_budget(seg)
        if not rtype and ptype and (ar[0] or bu[0]):rtype="INVENTORY"
        if not rtype or not ptype: continue

        local_phones=phones(seg)
        ph=local_phones if local_phones else inherited
        locs=locations(seg)
        complexity=entity_complexity(seg)
        conf=25+20+(15 if locs else 0)+(15 if ph else 0)+(10 if ar[0] else 0)+(10 if bu[0] else 0)+5
        status="unverified" if ph else "CONTACT_MISSING"
        reason=None if ph else "No phone number in message or WhatsApp sender metadata"
        # Never pollute the active database with a row we still believe contains multiple assets.
        if complexity>=2:
            status="NEEDS_SPLIT"
            reason=f"Parser safety hold: possible {complexity} properties still combined"
        person=extract_person(seg,row.get("sender_display_name") or "")
        out.append({"rtype":rtype,"raw":seg,"phone":ph[0] if ph else "","phones":ph,"tx":detect_transaction(seg,rtype),
                    "ptype":ptype,"locs":locs,"area":ar,"budget":bu,"confidence":min(conf,100),"status":status,
                    "reason":reason,"idx":idx,"entity_count":count,"person":person})
    return out

def backfill_missing_contacts(c):
    rows=c.execute(text("""SELECT w.id,w.contact_no,w.all_contacts,r.sender_phone,w.raw_details
        FROM wai_clean_records w LEFT JOIN wai_raw_messages r ON r.id=w.source_message_id
        WHERE w.record_type IN ('INVENTORY','REQUIREMENT') AND COALESCE(w.contact_no,'')=''""")).mappings().all()
    fixed=0
    for r in rows:
        ph=merge_phones(phones(r["raw_details"] or ""),r["sender_phone"])
        if not ph: continue
        c.execute(text("""UPDATE wai_clean_records
            SET contact_no=:p, all_contacts=CAST(:allp AS jsonb),
                status=CASE WHEN status='CONTACT_MISSING' THEN 'unverified' ELSE status END,
                rejection_reason=CASE WHEN rejection_reason LIKE 'No phone number%' THEN NULL ELSE rejection_reason END,
                processed_at=NOW()
            WHERE id=:id"""),{"p":ph[0],"allp":json.dumps(ph),"id":r["id"]})
        fixed+=1
    return fixed

def refresh_clean_database(full_rebuild=False):
    init_clean_db();run_id=uuid.uuid4();stats={"source_messages":0,"inventory_rows":0,"requirement_rows":0,"rejected_rows":0,"failed":0,"contacts_backfilled":0}
    with engine.begin() as c:
        c.execute(text("INSERT INTO wai_clean_runs(id,notes) VALUES(:id,:n)"),{"id":run_id,"n":"full rebuild" if full_rebuild else "incremental refresh"})
        if full_rebuild:c.execute(text("DELETE FROM wai_clean_records"))
        existing=set() if full_rebuild else {r[0] for r in c.execute(text("SELECT DISTINCT source_message_id FROM wai_clean_records")).all()}
        src=c.execute(text("""SELECT r.id source_message_id,r.message_text,r.sender_phone,r.sender_display_name,r.sent_at,g.name source_group
            FROM wai_raw_messages r LEFT JOIN wai_groups g ON g.id=r.group_id
            WHERE COALESCE(r.message_text,'')<>'' ORDER BY r.ingested_at,r.id""")).mappings().all()
        for row in src:
            stats["source_messages"]+=1
            if row["source_message_id"] in existing:continue
            try:
                for rec in parse_message(row):
                    fp=hashlib.sha256(f"{row['source_message_id']}|{rec['idx']}|{norm(rec['raw']).lower()}".encode()).hexdigest()
                    ar=rec["area"];bu=rec["budget"];locs=rec["locs"]
                    c.execute(text("""INSERT INTO wai_clean_records(
                        id,source_message_id,source_group,record_type,transaction,raw_details,contact_no,all_contacts,
                        budget_text,budget_min,budget_max,budget_period,area_text,area_min,area_max,area_unit,
                        area_sqft_min,area_sqft_max,location,all_locations,property_type,person_name,confidence,status,
                        rejection_reason,source_fingerprint,source_created_at,source_entity_index,source_entity_count)
                        VALUES(:id,:mid,:grp,:rtype,:tx,:raw,:phone,CAST(:phones AS jsonb),:btxt,:bmin,:bmax,:bperiod,
                        :atxt,:amin,:amax,:aunit,:asmin,:asmax,:loc,CAST(:locs AS jsonb),:ptype,:person,:conf,:status,:reason,:fp,:sent,:eidx,:ecount)
                        ON CONFLICT(source_fingerprint) DO NOTHING"""),
                    {"id":uuid.uuid4(),"mid":row["source_message_id"],"grp":row["source_group"] or "","rtype":rec["rtype"],"tx":rec["tx"],"raw":rec["raw"],
                     "phone":rec["phone"],"phones":json.dumps(rec["phones"]),"btxt":bu[0],"bmin":bu[1],"bmax":bu[2],"bperiod":bu[3],
                     "atxt":ar[0],"amin":ar[1],"amax":ar[2],"aunit":ar[3],"asmin":ar[4],"asmax":ar[5],"loc":locs[0] if locs else None,
                     "locs":json.dumps(locs),"ptype":rec["ptype"],"person":rec.get("person") or row["sender_display_name"] or "","conf":rec["confidence"],
                     "status":rec["status"],"reason":rec["reason"],"fp":fp,"sent":row["sent_at"],"eidx":rec.get("idx"),"ecount":rec.get("entity_count")})
                    if rec["rtype"]=="INVENTORY":stats["inventory_rows"]+=1
                    elif rec["rtype"]=="REQUIREMENT":stats["requirement_rows"]+=1
                    else:stats["rejected_rows"]+=1
            except Exception:
                stats["failed"]+=1
        stats["contacts_backfilled"]=backfill_missing_contacts(c)
        c.execute(text("""UPDATE wai_clean_runs SET completed_at=NOW(),source_messages=:s,inventory_rows=:i,
            requirement_rows=:r,rejected_rows=:x,failed=:f,notes=:n WHERE id=:id"""),
            {"s":stats["source_messages"],"i":stats["inventory_rows"],"r":stats["requirement_rows"],"x":stats["rejected_rows"],
             "f":stats["failed"],"n":f"contacts_backfilled={stats['contacts_backfilled']}","id":run_id})
    return stats

def shell(title,body):
    links=[("Dashboard","/whatsapp-capture/intelligence/clean"),("Inventory","/whatsapp-capture/intelligence/clean/properties"),
           ("Requirements","/whatsapp-capture/intelligence/clean/requirements"),("AI Match Database","/whatsapp-capture/intelligence/clean/matches"),
           ("Contacts","/whatsapp-capture/intelligence/clean/contacts"),("Needs Contact","/whatsapp-capture/intelligence/clean/needs-contact"),
           ("Data Quality","/whatsapp-capture/intelligence/clean/data-quality"),
           ("Rejected","/whatsapp-capture/intelligence/clean/rejected"),("Excel","/whatsapp-capture/intelligence/clean/export"),
           ("← Sources","/whatsapp-capture/intelligence/accounts")]
    nav=" ".join(f"<a href='{u}'>{html.escape(n)}</a>" for n,u in links)
    return f"""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1">
    <style>body{{font-family:Arial;margin:0;background:#f5f7fb;color:#172437}}header{{background:#fff;padding:18px 22px;border-bottom:1px solid #ddd}}
    nav{{display:flex;gap:12px;flex-wrap:wrap}}nav a{{text-decoration:none;font-weight:700}}main{{padding:20px}}.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(170px,1fr));gap:12px}}
    .card{{background:#fff;padding:14px;border:1px solid #ddd;border-radius:9px}}.num{{font-size:28px;font-weight:800}}.scroll{{overflow:auto;background:#fff;border:1px solid #ddd}}
    table{{border-collapse:collapse;width:100%;min-width:1550px}}th,td{{padding:9px;border-bottom:1px solid #ddd;vertical-align:top;text-align:left}}th{{background:#fafafa;position:sticky;top:0}}
    td.raw{{min-width:420px}}.btn{{display:inline-block;padding:8px 11px;background:#155eef;color:#fff;text-decoration:none;border-radius:7px;font-weight:700;border:0;cursor:pointer}}
    .green{{background:#067647}}.edit{{background:#7a5af8}}.muted{{color:#667085}}.phone{{white-space:nowrap;font-weight:800}}.date{{white-space:nowrap}}.warn{{color:#b42318;font-weight:800}}
    input,select,textarea{{width:100%;box-sizing:border-box;padding:9px;border:1px solid #ccd6e2;border-radius:7px}}textarea{{min-height:150px}}label{{font-weight:700}}.formgrid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:12px}}
    </style></head><body><header><h2>{html.escape(title)}</h2><div class=muted>Structured · one entity per row · contacts and dates preserved</div><nav>{nav}</nav></header><main>{body}</main></body></html>"""

@router.get("",response_class=HTMLResponse)
def dashboard():
    init_clean_db()
    with engine.begin() as c:
        backfill_missing_contacts(c)
        r=c.execute(text("""SELECT COUNT(*) FILTER(WHERE record_type='INVENTORY' AND COALESCE(contact_no,'')<>'') inv,
            COUNT(*) FILTER(WHERE record_type='REQUIREMENT' AND COALESCE(contact_no,'')<>'') req,
            COUNT(*) FILTER(WHERE record_type='REJECTED') rej,
            COUNT(*) FILTER(WHERE record_type IN ('INVENTORY','REQUIREMENT') AND COALESCE(contact_no,'')='') missing,
            COUNT(*) FILTER(WHERE status='NEEDS_SPLIT') needs_split,
            COUNT(DISTINCT NULLIF(contact_no,'')) contacts,MAX(processed_at) last_processed FROM wai_clean_records""")).mappings().first()
    body=f"""<div class=grid><div class=card><div>Inventory with Contact</div><div class=num>{r['inv']}</div></div>
    <div class=card><div>Requirements with Contact</div><div class=num>{r['req']}</div></div>
    <div class=card><div>Unique Contacts</div><div class=num>{r['contacts']}</div></div>
    <div class=card><div>Needs Contact</div><div class=num>{r['missing']}</div></div>
    <div class=card><div>Needs Split Review</div><div class=num>{r['needs_split']}</div></div>
    <div class=card><div>Rejected Noise</div><div class=num>{r['rej']}</div></div></div><br>
    <div class=card><b>Data purity rule:</b> one property/requirement = one row. Active Inventory, Requirements and AI Match exclude combined rows and records without contact details.
    Missing numbers are recovered first from the message and then from the WhatsApp sender number. Records with no recoverable number stay under <b>Needs Contact</b>.</div><br>
    <a class='btn green' href='/whatsapp-capture/intelligence/clean/rebuild'>FULL REBUILD FROM RAW WHATSAPP</a>
    <a class=btn href='/whatsapp-capture/intelligence/clean/refresh'>PROCESS NEW + BACKFILL CONTACTS</a>
    <p class=muted>Last processed: {fmt_date(r['last_processed'])}</p>"""
    return HTMLResponse(shell("WhatsApp Clean Property Database",body))

@router.get("/refresh")
def refresh():
    refresh_clean_database(False);return RedirectResponse("/whatsapp-capture/intelligence/clean",303)

@router.get("/rebuild")
def rebuild():
    refresh_clean_database(True);return RedirectResponse("/whatsapp-capture/intelligence/clean",303)

def get_rows(rtype, require_contact=True):
    where="record_type=:r"
    if require_contact and rtype in ("INVENTORY","REQUIREMENT"):
        where+=" AND COALESCE(contact_no,'')<>'' AND status NOT IN ('inactive','NEEDS_SPLIT')"
    with engine.begin() as c:
        backfill_missing_contacts(c)
        return c.execute(text(f"SELECT * FROM wai_clean_records WHERE {where} ORDER BY source_created_at DESC NULLS LAST LIMIT 5000"),{"r":rtype}).mappings().all()

def records_table(rows, label):
    trs="".join(f"""<tr>
      <td class=date>{esc(fmt_date(r['source_created_at']))}</td>
      <td class=raw>{esc(r['raw_details'])}</td>
      <td class=phone>{esc(phone_line(r))}</td>
      <td>{esc(r['budget_text'])}</td><td>{esc(r['area_text'])}</td><td>{esc(r['source_group'])}</td>
      <td>{esc(', '.join(r['all_locations'] or []))}</td><td>{esc(r['property_type'])}</td><td>{esc(r['transaction'])}</td>
      <td>{esc(r['person_name'])}</td><td>{float(r['confidence'] or 0):.0f}%</td><td>{esc(r['status'])}</td>
      <td><a class='btn edit' href='/whatsapp-capture/intelligence/clean/edit/{r["id"]}'>Edit</a></td></tr>""" for r in rows)
    return f"""<h2>{html.escape(label)}</h2><div class=scroll><table><tr><th>Date</th><th>Raw Details</th><th>Contact No.</th><th>Price / Budget</th>
    <th>Area</th><th>Source Group</th><th>Location</th><th>Property Type</th><th>Transaction</th><th>Person / Broker</th>
    <th>Confidence</th><th>Status</th><th>Edit</th></tr>{trs}</table></div>"""

@router.get("/properties",response_class=HTMLResponse)
def properties():
    init_clean_db();return HTMLResponse(shell("Property Database",records_table(get_rows("INVENTORY"),"Property Database")))

@router.get("/requirements",response_class=HTMLResponse)
def requirements():
    init_clean_db();return HTMLResponse(shell("Requirements",records_table(get_rows("REQUIREMENT"),"Buyer / Tenant Requirements")))

@router.get("/needs-contact",response_class=HTMLResponse)
def needs_contact():
    init_clean_db()
    with engine.begin() as c:
        backfill_missing_contacts(c)
        rows=c.execute(text("""SELECT * FROM wai_clean_records WHERE record_type IN ('INVENTORY','REQUIREMENT')
            AND COALESCE(contact_no,'')='' ORDER BY source_created_at DESC NULLS LAST LIMIT 5000""")).mappings().all()
    trs="".join(f"""<tr><td class=date>{esc(fmt_date(r['source_created_at']))}</td><td>{esc(r['record_type'])}</td>
      <td class=raw>{esc(r['raw_details'])}</td><td>{esc(r['source_group'])}</td><td>{esc(r['person_name'])}</td>
      <td class=warn>CONTACT MISSING</td><td><a class='btn edit' href='/whatsapp-capture/intelligence/clean/edit/{r["id"]}'>Add Contact</a></td></tr>""" for r in rows)
    body=f"""<h2>Needs Contact</h2><p>These records are preserved but are not treated as active leads or used in AI matching until a contact number is added.</p>
    <div class=scroll><table><tr><th>Date</th><th>Type</th><th>Raw Details</th><th>Source Group</th><th>Name</th><th>Status</th><th>Edit</th></tr>{trs}</table></div>"""
    return HTMLResponse(shell("Needs Contact",body))

@router.get("/edit/{record_id}",response_class=HTMLResponse)
def edit_record(record_id:str):
    init_clean_db()
    with engine.begin() as c:
        r=c.execute(text("SELECT * FROM wai_clean_records WHERE id=:id"),{"id":record_id}).mappings().first()
    if not r:return HTMLResponse(shell("Edit Lead","<h3>Record not found.</h3>"),404)
    locs=", ".join(r["all_locations"] or [])
    body=f"""<h2>Edit Lead</h2><form method=post action='/whatsapp-capture/intelligence/clean/edit/{record_id}' class=card>
    <div class=formgrid>
      <div><label>Contact No(s) - one line</label><input name=contact_numbers value="{html.escape(phone_line(r),quote=True)}" placeholder="+91 98765 43210 | +91 98118 95500"></div>
      <div><label>Person / Broker / Client</label><input name=person_name value="{html.escape(str(r['person_name'] or ''),quote=True)}"></div>
      <div><label>Property Type</label><input name=property_type value="{html.escape(str(r['property_type'] or ''),quote=True)}"></div>
      <div><label>Transaction</label><select name=transaction><option>{esc(r['transaction'] or '')}</option><option>SALE</option><option>RENT</option><option>REQUIREMENT</option></select></div>
      <div><label>Location(s)</label><input name=location_text value="{html.escape(locs,quote=True)}"></div>
      <div><label>Price / Budget</label><input name=budget_text value="{html.escape(str(r['budget_text'] or ''),quote=True)}"></div>
      <div><label>Area</label><input name=area_text value="{html.escape(str(r['area_text'] or ''),quote=True)}"></div>
      <div><label>Status</label><select name=status><option>{esc(r['status'] or '')}</option><option>unverified</option><option>verified</option><option>CONTACT_MISSING</option><option>inactive</option></select></div>
    </div><br><label>Raw Details</label><textarea name=raw_details>{esc(r['raw_details'])}</textarea><br><br>
    <button class='btn green' type=submit>Save Changes</button> <a class=btn href="javascript:history.back()">Cancel</a></form>
    <div class=card><b>Source Date:</b> {esc(fmt_date(r['source_created_at']))}<br><b>Source Group:</b> {esc(r['source_group'])}</div>"""
    return HTMLResponse(shell("Edit Lead",body))

@router.post("/edit/{record_id}")
def save_edit(record_id:str,contact_numbers:str=Form(""),person_name:str=Form(""),property_type:str=Form(""),
              transaction:str=Form(""),location_text:str=Form(""),budget_text:str=Form(""),area_text:str=Form(""),
              status:str=Form("unverified"),raw_details:str=Form("")):
    init_clean_db()
    ph=merge_phones(contact_numbers)
    locs=[norm(x) for x in re.split(r"[,|]",location_text or "") if norm(x)]
    if ph and status=="CONTACT_MISSING":status="unverified"
    if not ph and status not in ("inactive",):status="CONTACT_MISSING"
    ar=extract_area(area_text or raw_details);bu=extract_budget(budget_text or raw_details)
    with engine.begin() as c:
        c.execute(text("""UPDATE wai_clean_records SET contact_no=:phone,all_contacts=CAST(:phones AS jsonb),person_name=:person,
            property_type=:ptype,transaction=:tx,location=:loc,all_locations=CAST(:locs AS jsonb),budget_text=:btxt,
            budget_min=:bmin,budget_max=:bmax,budget_period=:bperiod,area_text=:atxt,area_min=:amin,area_max=:amax,
            area_unit=:aunit,area_sqft_min=:asmin,area_sqft_max=:asmax,status=:status,raw_details=:raw,edited_at=NOW(),
            processed_at=NOW() WHERE id=:id"""),
            {"phone":ph[0] if ph else "","phones":json.dumps(ph),"person":person_name,"ptype":property_type,"tx":transaction,
             "loc":locs[0] if locs else None,"locs":json.dumps(locs),"btxt":budget_text or bu[0],"bmin":bu[1],"bmax":bu[2],"bperiod":bu[3],
             "atxt":area_text or ar[0],"amin":ar[1],"amax":ar[2],"aunit":ar[3],"asmin":ar[4],"asmax":ar[5],
             "status":status,"raw":raw_details,"id":record_id})
    return RedirectResponse("/whatsapp-capture/intelligence/clean/properties" if transaction!="REQUIREMENT" else "/whatsapp-capture/intelligence/clean/requirements",303)

def overlap(a1,a2,b1,b2):
    if None in (a1,a2,b1,b2):return False
    return max(float(a1),float(b1)) <= min(float(a2),float(b2))

def match_score(req,supply):
    score=0;reasons=[]
    if req["property_type"] and supply["property_type"] and req["property_type"].lower()==supply["property_type"].lower():
        score+=35;reasons.append("Same property type")
    req_l={str(x).lower() for x in (req["all_locations"] or [])};sup_l={str(x).lower() for x in (supply["all_locations"] or [])}
    if req_l and sup_l and req_l.intersection(sup_l):
        score+=30;reasons.append("Location matched")
    if req["transaction"] in ("SALE","RENT") and supply["transaction"]==req["transaction"]:
        score+=15;reasons.append("Same sale/rent intent")
    if overlap(req["area_sqft_min"],req["area_sqft_max"],supply["area_sqft_min"],supply["area_sqft_max"]):
        score+=10;reasons.append("Area overlaps")
    if overlap(req["budget_min"],req["budget_max"],supply["budget_min"],supply["budget_max"]):
        score+=5;reasons.append("Budget/rent overlaps")
    if float(req["confidence"] or 0)>=70 and float(supply["confidence"] or 0)>=70:
        score+=5;reasons.append("High data confidence")
    return score,reasons

@router.get("/matches",response_class=HTMLResponse)
def clean_matches():
    init_clean_db()
    with engine.begin() as c:
        backfill_missing_contacts(c)
        reqs=c.execute(text("""SELECT * FROM wai_clean_records WHERE record_type='REQUIREMENT'
            AND COALESCE(contact_no,'')<>'' AND status NOT IN ('inactive','NEEDS_SPLIT') ORDER BY source_created_at DESC NULLS LAST LIMIT 1000""")).mappings().all()
        supplies=c.execute(text("""SELECT * FROM wai_clean_records WHERE record_type='INVENTORY'
            AND COALESCE(contact_no,'')<>'' AND status NOT IN ('inactive','NEEDS_SPLIT') ORDER BY source_created_at DESC NULLS LAST LIMIT 3000""")).mappings().all()
    rows=[]
    for q in reqs:
        for p in supplies:
            s,reasons=match_score(q,p)
            if s>=35:
                rows.append((s,q,p,reasons))
    def _sort_date(v):
        try:return v.timestamp()
        except Exception:return 0
    rows.sort(key=lambda x:(x[0],_sort_date(x[1]["source_created_at"])),reverse=True)
    trs="".join(f"""<tr><td><b>{s}%</b></td>
      <td class=date>{esc(fmt_date(q['source_created_at']))}</td><td class=raw>{esc(q['raw_details'])}</td>
      <td class=phone><b>BUYER / TENANT:</b><br>{esc(phone_line(q))}<br>{esc(q['person_name'])}</td>
      <td class=date>{esc(fmt_date(p['source_created_at']))}</td><td class=raw>{esc(p['raw_details'])}</td>
      <td class=phone><b>SELLER / OWNER / BROKER:</b><br>{esc(phone_line(p))}<br>{esc(p['person_name'])}</td>
      <td>{esc(p['source_group'])}</td><td>{esc(' | '.join(reasons))}</td></tr>""" for s,q,p,reasons in rows[:5000])
    body=f"""<h2>AI Match Database</h2><p><b>Buyer and seller contact numbers are shown separately and clearly.</b>
    Only records with contact details are matched.</p><div class=scroll><table><tr><th>Match %</th><th>Buyer Date</th>
    <th>Buyer / Tenant Requirement</th><th>BUYER / TENANT CONTACT</th><th>Seller Date</th><th>Matched Property</th>
    <th>SELLER / OWNER / BROKER CONTACT</th><th>Source Group</th><th>Why Matched</th></tr>{trs}</table></div>"""
    return HTMLResponse(shell("AI Match Database",body))

@router.get("/contacts",response_class=HTMLResponse)
def contacts():
    init_clean_db()
    with engine.begin() as c:
        backfill_missing_contacts(c)
        rows=c.execute(text("""SELECT contact_no,MAX(NULLIF(person_name,'')) person_name,
            STRING_AGG(DISTINCT source_group,' | ') source_groups,
            COUNT(*) FILTER(WHERE record_type='INVENTORY') inventory_count,
            COUNT(*) FILTER(WHERE record_type='REQUIREMENT') requirement_count,
            MAX(source_created_at) last_seen,ROUND(AVG(confidence),0) confidence
            FROM wai_clean_records WHERE COALESCE(contact_no,'')<>'' GROUP BY contact_no ORDER BY last_seen DESC""")).mappings().all()
    trs="".join(f"<tr><td class=phone>{esc(r['contact_no'])}</td><td>{esc(r['person_name'])}</td><td>{esc(r['source_groups'])}</td><td>{r['inventory_count']}</td><td>{r['requirement_count']}</td><td class=date>{esc(fmt_date(r['last_seen']))}</td><td>{float(r['confidence'] or 0):.0f}%</td></tr>" for r in rows)
    return HTMLResponse(shell("Contacts Database",f"<h2>Contacts Database</h2><div class=scroll><table><tr><th>Contact No.</th><th>Name</th><th>Source Groups</th><th>Inventory</th><th>Requirements</th><th>Last Seen</th><th>Confidence</th></tr>{trs}</table></div>"))

@router.get("/data-quality",response_class=HTMLResponse)
def data_quality():
    init_clean_db()
    with engine.begin() as c:
        rows=c.execute(text("""SELECT * FROM wai_clean_records
            WHERE status='NEEDS_SPLIT'
            ORDER BY source_created_at DESC NULLS LAST LIMIT 5000""")).mappings().all()
    trs="".join(f"""<tr><td class=date>{esc(fmt_date(r['source_created_at']))}</td>
      <td>{esc(r['record_type'])}</td><td class=raw>{esc(r['raw_details'])}</td>
      <td class=phone>{esc(phone_line(r))}</td><td>{esc(r['source_group'])}</td>
      <td class=warn>{esc(r['rejection_reason'])}</td>
      <td><a class='btn edit' href='/whatsapp-capture/intelligence/clean/edit/{r["id"]}'>Review / Edit</a></td></tr>""" for r in rows)
    body=f"""<h2>Data Quality · Needs Split Review</h2>
    <p>Safety queue. These rows are <b>not</b> used as inventory, requirements or AI matches until they are normalized.</p>
    <div class=scroll><table><tr><th>Date</th><th>Type</th><th>Combined Details</th><th>Contact</th>
    <th>Source Group</th><th>Reason</th><th>Action</th></tr>{trs}</table></div>"""
    return HTMLResponse(shell("Data Quality",body))

@router.get("/rejected",response_class=HTMLResponse)
def rejected():
    init_clean_db();rows=get_rows("REJECTED",False)
    trs="".join(f"<tr><td class=date>{esc(fmt_date(r['source_created_at']))}</td><td class=raw>{esc(r['raw_details'])}</td><td>{esc(r['rejection_reason'])}</td><td>{esc(r['source_group'])}</td><td class=phone>{esc(phone_line(r))}</td></tr>" for r in rows)
    return HTMLResponse(shell("Rejected / Noise",f"<h2>Rejected / Noise</h2><div class=scroll><table><tr><th>Date</th><th>Raw Message</th><th>Reason</th><th>Source Group</th><th>Contact</th></tr>{trs}</table></div>"))

@router.get("/export")
def export():
    init_clean_db();inv=get_rows("INVENTORY");req=get_rows("REQUIREMENT")
    wb=Workbook();ws=wb.active;ws.title="Inventory";ws.append(["Date","Raw Property Details","Contact No(s)","Price / Rent","Area","Source Group","Location","Property Type","Transaction","Poster / Broker","Confidence %","Status"])
    for r in inv:ws.append([fmt_date(r["source_created_at"]),r["raw_details"],phone_line(r),r["budget_text"],r["area_text"],r["source_group"],", ".join(r["all_locations"] or []),r["property_type"],r["transaction"],r["person_name"],r["confidence"],r["status"]])
    wr=wb.create_sheet("Requirements");wr.append(["Date","Raw Requirement Details","Contact No(s)","Budget","Area","Source Group","Location","Property Type","Transaction","Client / Broker","Confidence %","Status"])
    for r in req:wr.append([fmt_date(r["source_created_at"]),r["raw_details"],phone_line(r),r["budget_text"],r["area_text"],r["source_group"],", ".join(r["all_locations"] or []),r["property_type"],r["transaction"],r["person_name"],r["confidence"],r["status"]])
    for sh in wb.worksheets:
        for c in sh[1]:c.font=Font(bold=True);c.alignment=Alignment(wrap_text=True)
        for col in range(1,sh.max_column+1):sh.column_dimensions[get_column_letter(col)].width=24
    bio=io.BytesIO();wb.save(bio);bio.seek(0)
    return StreamingResponse(bio,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":"attachment; filename=Alliance_WhatsApp_Clean_Database.xlsx"})

try:init_clean_db()
except Exception as e:print("WhatsApp clean DB init warning:",e)
