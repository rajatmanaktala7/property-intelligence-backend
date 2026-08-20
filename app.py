
import re
import os, io, csv, json, uuid, hmac, hashlib, base64, tempfile
from html import escape
from urllib.parse import quote_plus
from datetime import datetime, timezone, date
from decimal import Decimal
from typing import Optional, Literal

from fastapi import FastAPI, Request, UploadFile, File, Query, BackgroundTasks, Form, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response
from pydantic import BaseModel, Field
from sqlalchemy import create_engine, text
from google import genai
from google.genai import types
from PIL import Image
import fitz
from pypdf import PdfReader, PdfWriter

VERSION="13.2.1-REFINED-MAGAZINE-IMPORT-FIX"
DATABASE_URL=os.getenv("DATABASE_URL","")
GEMINI_API_KEY=os.getenv("GEMINI_API_KEY","")
GEMINI_MODEL=os.getenv("GEMINI_MODEL","gemini-3.1-flash-lite")
MAX_UPLOAD_MB=int(os.getenv("MAX_UPLOAD_MB","100"))
PDF_PAGES_PER_BATCH=int(os.getenv("PDF_PAGES_PER_BATCH","2"))
MAX_PROPERTY_IMAGES=int(os.getenv("MAX_PROPERTY_IMAGES","12"))
MAX_IMAGE_MB=int(os.getenv("MAX_IMAGE_MB","10"))
VERIFICATION_DUE_DAYS=int(os.getenv("VERIFICATION_DUE_DAYS","30"))
SCAN_TILE_COLS=int(os.getenv("SCAN_TILE_COLS","3"))
SCAN_TILE_ROWS=int(os.getenv("SCAN_TILE_ROWS","3"))
SCAN_TILE_OVERLAP=float(os.getenv("SCAN_TILE_OVERLAP","0.12"))
PDF_RENDER_DPI=int(os.getenv("PDF_RENDER_DPI","220"))
ADMIN_CODE=os.getenv("ADMIN_CODE","admin-change-me")
TEAM_CODE=os.getenv("TEAM_CODE","team-change-me")
SESSION_SECRET=os.getenv("SESSION_SECRET","change-this-secret")

if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL is required")
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL=DATABASE_URL.replace("postgres://","postgresql+psycopg://",1)
elif DATABASE_URL.startswith("postgresql://"):
    DATABASE_URL=DATABASE_URL.replace("postgresql://","postgresql+psycopg://",1)

engine=create_engine(DATABASE_URL,pool_pre_ping=True,pool_recycle=300)
client=genai.Client(api_key=GEMINI_API_KEY) if GEMINI_API_KEY else None
app=FastAPI(title="Property Intelligence Unified Workspace",version=VERSION)

SCHEMA='''
CREATE TABLE IF NOT EXISTS pi_properties(
 id BIGSERIAL PRIMARY KEY, property_id VARCHAR(50) UNIQUE NOT NULL,
 fingerprint VARCHAR(64), property_name VARCHAR(255), entry_status VARCHAR(50) DEFAULT 'Active',
 availability_status VARCHAR(50) DEFAULT 'Available', property_type VARCHAR(100) DEFAULT 'NA',
 city VARCHAR(100) DEFAULT 'NA', location VARCHAR(255) DEFAULT 'NA',
 available_area_sqft NUMERIC(14,2), minimum_area_sqft NUMERIC(14,2),
 maximum_area_sqft NUMERIC(14,2), floor VARCHAR(100), rent_or_sale VARCHAR(30),
 possession VARCHAR(100), nearby_brands TEXT, suitable_category TEXT, parking TEXT,
 google_maps_pin TEXT, owner_name VARCHAR(255), owner_contact VARCHAR(100),
 broker_name VARCHAR(255), broker_contact VARCHAR(100), remarks TEXT,
 image_urls TEXT, video_urls TEXT, brochure_url TEXT,
 verification_status VARCHAR(50) DEFAULT 'UNVERIFIED', verified_by VARCHAR(255),
 verified_date DATE, source VARCHAR(255), source_id BIGINT, extraction_confidence NUMERIC(5,2),
 created_at TIMESTAMPTZ DEFAULT NOW(), updated_at TIMESTAMPTZ DEFAULT NOW()
);
CREATE TABLE IF NOT EXISTS pi_requirements(
 id BIGSERIAL PRIMARY KEY, requirement_id VARCHAR(50) UNIQUE NOT NULL,
 fingerprint VARCHAR(64), client_name VARCHAR(255), company_name VARCHAR(255),
 contact_phone VARCHAR(100), contact_email VARCHAR(255), requirement_type VARCHAR(100),
 property_type VARCHAR(100), city VARCHAR(100), preferred_locations TEXT,
 minimum_area_sqft NUMERIC(14,2), maximum_area_sqft NUMERIC(14,2), rent_or_sale VARCHAR(30),
 nearby_brands TEXT, suitable_category TEXT, additional_points TEXT, source VARCHAR(255),
 source_id BIGINT, status VARCHAR(50) DEFAULT 'New', extraction_confidence NUMERIC(5,2),
 created_at TIMESTAMPTZ DEFAULT NOW(), updated_at TIMESTAMPTZ DEFAULT NOW()
);
CREATE TABLE IF NOT EXISTS pi_sources(
 id BIGSERIAL PRIMARY KEY, source_type VARCHAR(50) NOT NULL, source_name VARCHAR(255),
 source_reference TEXT, original_filename VARCHAR(500), mime_type VARCHAR(150),
 ingestion_status VARCHAR(50) DEFAULT 'RECEIVED', extracted_record_type VARCHAR(50),
 processed_records INTEGER DEFAULT 0, duplicate_records INTEGER DEFAULT 0,
 error_message TEXT, ai_provider VARCHAR(50), ai_model VARCHAR(100),
 uploaded_at TIMESTAMPTZ DEFAULT NOW(), processed_at TIMESTAMPTZ
);
CREATE TABLE IF NOT EXISTS pi_matches(
 id BIGSERIAL PRIMARY KEY, requirement_id VARCHAR(50) NOT NULL, property_id VARCHAR(50) NOT NULL,
 match_score NUMERIC(5,2) DEFAULT 0, rank INTEGER, match_reasons JSONB DEFAULT '[]'::jsonb,
 status VARCHAR(50) DEFAULT 'READY_FOR_REVIEW', created_at TIMESTAMPTZ DEFAULT NOW()
);
CREATE TABLE IF NOT EXISTS pi_verification_log(
 id BIGSERIAL PRIMARY KEY, property_id VARCHAR(50), requirement_id VARCHAR(50),
 action VARCHAR(100) NOT NULL, performed_by VARCHAR(255), notes TEXT,
 old_value JSONB, new_value JSONB,
 created_at TIMESTAMPTZ DEFAULT NOW()
);
CREATE TABLE IF NOT EXISTS pi_ai_jobs(
 id BIGSERIAL PRIMARY KEY, source_id BIGINT, job_type VARCHAR(50) NOT NULL,
 status VARCHAR(50) DEFAULT 'PENDING', provider VARCHAR(50), model VARCHAR(100),
 input_summary TEXT, output_summary TEXT, error_message TEXT,
 started_at TIMESTAMPTZ, completed_at TIMESTAMPTZ, created_at TIMESTAMPTZ DEFAULT NOW()

);

CREATE SEQUENCE IF NOT EXISTS pi_property_code_seq START 1;
CREATE SEQUENCE IF NOT EXISTS pi_requirement_code_seq START 1;

CREATE TABLE IF NOT EXISTS pi_extraction_batches(
 id BIGSERIAL PRIMARY KEY,
 source_id BIGINT NOT NULL,
 start_page INTEGER NOT NULL,
 end_page INTEGER NOT NULL,
 status VARCHAR(50) DEFAULT 'PENDING',
 attempts INTEGER DEFAULT 0,
 records_created INTEGER DEFAULT 0,
 duplicates INTEGER DEFAULT 0,
 error_message TEXT,
 updated_at TIMESTAMPTZ DEFAULT NOW(),
 UNIQUE(source_id,start_page,end_page)
);

CREATE TABLE IF NOT EXISTS pi_message_drafts(
 id BIGSERIAL PRIMARY KEY,
 requirement_id TEXT NOT NULL,
 recipient_name TEXT,
 recipient_phone TEXT,
 channel VARCHAR(30) DEFAULT 'WHATSAPP',
 message_text TEXT NOT NULL,
 status VARCHAR(50) DEFAULT 'READY_FOR_REVIEW',
 ai_provider VARCHAR(50),
 ai_model VARCHAR(100),
 created_at TIMESTAMPTZ DEFAULT NOW()
);


CREATE TABLE IF NOT EXISTS pi_property_media(
 id BIGSERIAL PRIMARY KEY,
 media_id UUID UNIQUE NOT NULL,
 property_id TEXT NOT NULL,
 media_type VARCHAR(30) DEFAULT 'IMAGE',
 filename TEXT,
 mime_type TEXT,
 file_size BIGINT,
 content BYTEA NOT NULL,
 created_at TIMESTAMPTZ DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_pi_property_media_property_id ON pi_property_media(property_id);


CREATE TABLE IF NOT EXISTS pi_scan_tiles(
 id BIGSERIAL PRIMARY KEY,
 source_id BIGINT NOT NULL,
 page_number INTEGER,
 tile_label TEXT NOT NULL,
 status VARCHAR(50) DEFAULT 'PENDING',
 attempts INTEGER DEFAULT 0,
 records_created INTEGER DEFAULT 0,
 duplicates INTEGER DEFAULT 0,
 error_message TEXT,
 created_at TIMESTAMPTZ DEFAULT NOW(),
 updated_at TIMESTAMPTZ DEFAULT NOW(),
 UNIQUE(source_id,page_number,tile_label)
);
CREATE INDEX IF NOT EXISTS idx_pi_scan_tiles_source ON pi_scan_tiles(source_id);

'''

MIGRATIONS=[
"ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS fingerprint VARCHAR(64)",
"ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS source_id BIGINT",
"ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS extraction_confidence NUMERIC(5,2)",
"ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS verification_status VARCHAR(50) DEFAULT 'UNVERIFIED'",
"ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS updated_at TIMESTAMPTZ DEFAULT NOW()",
"ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS image_urls TEXT",
"ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS video_urls TEXT",
"ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS brochure_url TEXT",
"ALTER TABLE pi_verification_log ADD COLUMN IF NOT EXISTS old_value JSONB",
"ALTER TABLE pi_verification_log ADD COLUMN IF NOT EXISTS new_value JSONB",
"ALTER TABLE pi_requirements ADD COLUMN IF NOT EXISTS fingerprint VARCHAR(64)",
"ALTER TABLE pi_requirements ADD COLUMN IF NOT EXISTS source_id BIGINT",
"ALTER TABLE pi_requirements ADD COLUMN IF NOT EXISTS extraction_confidence NUMERIC(5,2)",
"ALTER TABLE pi_requirements ADD COLUMN IF NOT EXISTS updated_at TIMESTAMPTZ DEFAULT NOW()",
"ALTER TABLE pi_sources ADD COLUMN IF NOT EXISTS mime_type VARCHAR(150)",
"ALTER TABLE pi_sources ADD COLUMN IF NOT EXISTS extracted_record_type VARCHAR(50)",
"ALTER TABLE pi_sources ADD COLUMN IF NOT EXISTS duplicate_records INTEGER DEFAULT 0",
"ALTER TABLE pi_sources ADD COLUMN IF NOT EXISTS ai_provider VARCHAR(50)",
"ALTER TABLE pi_properties ALTER COLUMN property_name TYPE TEXT",
"ALTER TABLE pi_properties ALTER COLUMN property_type TYPE TEXT",
"ALTER TABLE pi_properties ALTER COLUMN city TYPE TEXT",
"ALTER TABLE pi_properties ALTER COLUMN location TYPE TEXT",
"ALTER TABLE pi_properties ALTER COLUMN floor TYPE TEXT",
"ALTER TABLE pi_properties ALTER COLUMN rent_or_sale TYPE TEXT",
"ALTER TABLE pi_properties ALTER COLUMN possession TYPE TEXT",
"ALTER TABLE pi_properties ALTER COLUMN owner_name TYPE TEXT",
"ALTER TABLE pi_properties ALTER COLUMN owner_contact TYPE TEXT",
"ALTER TABLE pi_properties ALTER COLUMN broker_name TYPE TEXT",
"ALTER TABLE pi_properties ALTER COLUMN broker_contact TYPE TEXT",
"ALTER TABLE pi_properties ALTER COLUMN source TYPE TEXT",
"ALTER TABLE pi_requirements ALTER COLUMN client_name TYPE TEXT",
"ALTER TABLE pi_requirements ALTER COLUMN company_name TYPE TEXT",
"ALTER TABLE pi_requirements ALTER COLUMN contact_phone TYPE TEXT",
"ALTER TABLE pi_requirements ALTER COLUMN contact_email TYPE TEXT",
"ALTER TABLE pi_requirements ALTER COLUMN requirement_type TYPE TEXT",
"ALTER TABLE pi_requirements ALTER COLUMN property_type TYPE TEXT",
"ALTER TABLE pi_requirements ALTER COLUMN city TYPE TEXT",
"ALTER TABLE pi_requirements ALTER COLUMN rent_or_sale TYPE TEXT",
"ALTER TABLE pi_requirements ALTER COLUMN source TYPE TEXT"
]

@app.on_event("startup")
def startup():
    with engine.begin() as c:
        for stmt in [x.strip() for x in SCHEMA.split(";") if x.strip()]:
            c.execute(text(stmt))
        for stmt in MIGRATIONS:
            c.execute(text(stmt))

def signed(role):
    sig=hmac.new(SESSION_SECRET.encode(),role.encode(),hashlib.sha256).hexdigest()
    return base64.urlsafe_b64encode((role+"|"+sig).encode()).decode()

def get_role(req):
    token=req.cookies.get("pi_session")
    if not token:return None
    try:
        raw=base64.urlsafe_b64decode(token).decode()
        role,sig=raw.rsplit("|",1)
        expected=hmac.new(SESSION_SECRET.encode(),role.encode(),hashlib.sha256).hexdigest()
        return role if role in {"team","admin"} and hmac.compare_digest(sig,expected) else None
    except Exception:
        return None

def need_login(req):
    role=get_role(req)
    if not role:
        raise HTTPException(401,"Login required")
    return role

def actor_name(req):
    return (req.headers.get("x-user-name") or get_role(req) or "unknown").strip()[:255]

def page_role_or_redirect(req):
    role=get_role(req)
    if role:
        return role
    return None

@app.exception_handler(Exception)
async def all_errors(req,exc):
    return JSONResponse(status_code=500,content={"status":"error","message":str(exc),"path":req.url.path})

class Property(BaseModel):
    property_name:Optional[str]=None
    property_type:str="NA"
    city:str="NA"
    location:str="NA"
    available_area_sqft:Optional[float]=None
    minimum_area_sqft:Optional[float]=None
    maximum_area_sqft:Optional[float]=None
    floor:Optional[str]=None
    rent_or_sale:Optional[str]=None
    nearby_brands:Optional[str]=None
    suitable_category:Optional[str]=None
    parking:Optional[str]=None
    owner_name:Optional[str]=None
    owner_contact:Optional[str]=None
    broker_name:Optional[str]=None
    broker_contact:Optional[str]=None
    remarks:Optional[str]=None
    image_urls:Optional[str]=None
    video_urls:Optional[str]=None
    brochure_url:Optional[str]=None
    source:Optional[str]="Manual"
    extraction_confidence:Optional[float]=None

class Requirement(BaseModel):
    client_name:Optional[str]=None
    company_name:Optional[str]=None
    contact_phone:Optional[str]=None
    contact_email:Optional[str]=None
    requirement_type:Optional[str]="Store Opening"
    property_type:Optional[str]="Retail"
    city:Optional[str]=None
    preferred_locations:Optional[str]=None
    minimum_area_sqft:Optional[float]=None
    maximum_area_sqft:Optional[float]=None
    rent_or_sale:Optional[str]=None
    nearby_brands:Optional[str]=None
    suitable_category:Optional[str]=None
    additional_points:Optional[str]=None
    source:Optional[str]="Manual"
    extraction_confidence:Optional[float]=None

class EP(Property):
    record_type:Literal["property"]="property"

class ER(Requirement):
    record_type:Literal["requirement"]="requirement"

class Envelope(BaseModel):
    properties:list[EP]=Field(default_factory=list)
    requirements:list[ER]=Field(default_factory=list)

class TextInput(BaseModel):
    source_type:str="WHATSAPP"
    source_name:Optional[str]=None
    text_content:str=Field(min_length=1)


class WhatsAppDraftResult(BaseModel):
    message: str


def make_id(prefix,conn=None):
    sequence="pi_property_code_seq" if prefix=="PROP" else "pi_requirement_code_seq"
    def get_value(c):
        return c.execute(text("SELECT nextval(:seq::regclass)"),{"seq":sequence}).scalar_one()
    # PostgreSQL does not allow a bind parameter directly in nextval(regclass) on all drivers,
    # so use a fixed allow-listed sequence name.
    sql=f"SELECT nextval('{sequence}')"
    if conn is not None:
        number=conn.execute(text(sql)).scalar_one()
    else:
        with engine.begin() as c:
            number=c.execute(text(sql)).scalar_one()
    return f"{prefix}-{datetime.utcnow().strftime('%Y%m%d')}-{int(number):010d}"

def fingerprint(values, minimum_identity_fields=3):
    cleaned=[str(v or "").lower().strip() for v in values]
    meaningful=[v for v in cleaned if v and v not in {"na","n/a","none","null","unknown","0"}]
    if len(meaningful) < minimum_identity_fields:
        return hashlib.sha256(("sparse|"+uuid.uuid4().hex).encode()).hexdigest()
    return hashlib.sha256("|".join(cleaned).encode()).hexdigest()

def source_row(source_type,name=None,filename=None,mime=None,reference=None):
    sql='INSERT INTO pi_sources(source_type,source_name,original_filename,mime_type,source_reference,ingestion_status) VALUES(:t,:n,:f,:m,:r,\'RECEIVED\') RETURNING id'
    with engine.begin() as c:
        return c.execute(text(sql),{"t":source_type,"n":name,"f":filename,"m":mime,"r":reference}).scalar_one()

def save_property(data,sid=None):
    d=dict(data)
    fp=fingerprint([d.get("property_name"),d.get("city"),d.get("location"),d.get("property_type"),d.get("available_area_sqft"),d.get("floor"),d.get("rent_or_sale"),d.get("owner_contact"),d.get("broker_contact")])
    with engine.begin() as c:
        old=c.execute(text("SELECT property_id FROM pi_properties WHERE fingerprint=:f LIMIT 1"),{"f":fp}).first()
        if old:return {"status":"duplicate","property_id":old[0]}
        pid=make_id("PROP",c)
        p={"pid":pid,"fp":fp,"sid":sid,**{k:d.get(k) for k in Property.model_fields}}
        sql='''INSERT INTO pi_properties(property_id,fingerprint,property_name,property_type,city,location,available_area_sqft,minimum_area_sqft,maximum_area_sqft,floor,rent_or_sale,nearby_brands,suitable_category,parking,owner_name,owner_contact,broker_name,broker_contact,remarks,image_urls,video_urls,brochure_url,source,source_id,extraction_confidence)
        VALUES(:pid,:fp,:property_name,:property_type,:city,:location,:available_area_sqft,:minimum_area_sqft,:maximum_area_sqft,:floor,:rent_or_sale,:nearby_brands,:suitable_category,:parking,:owner_name,:owner_contact,:broker_name,:broker_contact,:remarks,:image_urls,:video_urls,:brochure_url,:source,:sid,:extraction_confidence)'''
        c.execute(text(sql),p)
        c.execute(text("INSERT INTO pi_verification_log(property_id,action,performed_by,notes) VALUES(:p,'CREATED','SYSTEM','Queued for review')"),{"p":pid})
    return {"status":"created","property_id":pid}

def save_requirement(data,sid=None):
    d=dict(data)
    fp=fingerprint([d.get("company_name") or d.get("client_name"),d.get("contact_phone"),d.get("contact_email"),d.get("city"),d.get("preferred_locations"),d.get("minimum_area_sqft"),d.get("maximum_area_sqft")])
    with engine.begin() as c:
        old=c.execute(text("SELECT requirement_id FROM pi_requirements WHERE fingerprint=:f LIMIT 1"),{"f":fp}).first()
        if old:return {"status":"duplicate","requirement_id":old[0]}
        rid=make_id("REQ",c)
        p={"rid":rid,"fp":fp,"sid":sid,**{k:d.get(k) for k in Requirement.model_fields}}
        sql='''INSERT INTO pi_requirements(requirement_id,fingerprint,client_name,company_name,contact_phone,contact_email,requirement_type,property_type,city,preferred_locations,minimum_area_sqft,maximum_area_sqft,rent_or_sale,nearby_brands,suitable_category,additional_points,source,source_id,extraction_confidence)
        VALUES(:rid,:fp,:client_name,:company_name,:contact_phone,:contact_email,:requirement_type,:property_type,:city,:preferred_locations,:minimum_area_sqft,:maximum_area_sqft,:rent_or_sale,:nearby_brands,:suitable_category,:additional_points,:source,:sid,:extraction_confidence)'''
        c.execute(text(sql),p)
    return {"status":"created","requirement_id":rid}

PROMPT="""You are a high-recall real-estate magazine data extractor.
Extract EVERY distinct property listing and EVERY client/retailer requirement visible in the supplied pages.
Treat each classified advertisement/listing as a separate record.
Never summarize, sample, merge, or return only the best listings.
Do not invent facts. Use null for unknown fields. Preserve visible phone/contact details.
If a page contains 40 listings, return approximately 40 separate records.
Return all records in the required schema with extraction_confidence from 0 to 100.
"""

def complete_source(sid,envelope):
    props=[save_property({**p.model_dump(exclude={"record_type"}),"source":"AI_SOURCE_"+str(sid)},sid) for p in envelope.properties]
    reqs=[save_requirement({**r.model_dump(exclude={"record_type"}),"source":"AI_SOURCE_"+str(sid)},sid) for r in envelope.requirements]
    with engine.begin() as c:
        c.execute(text("UPDATE pi_sources SET ingestion_status='PROCESSED',processed_records=:n,duplicate_records=:d,ai_provider='gemini',ai_model=:m,processed_at=NOW() WHERE id=:id"),
                  {"n":len(props)+len(reqs),"d":sum(x["status"]=="duplicate" for x in props+reqs),"m":GEMINI_MODEL,"id":sid})
    return props,reqs

def run_text_job(sid,jid,content):
    try:
        if not client:raise RuntimeError("GEMINI_API_KEY missing")
        resp=client.models.generate_content(model=GEMINI_MODEL,contents=[PROMPT,"\nSOURCE:\n",content],
            config=types.GenerateContentConfig(response_mime_type="application/json",response_schema=Envelope,temperature=0.1))
        env=Envelope.model_validate(resp.parsed) if getattr(resp,"parsed",None) is not None else Envelope.model_validate_json(resp.text)
        p,r=complete_source(sid,env)
        with engine.begin() as c:c.execute(text("UPDATE pi_ai_jobs SET status='COMPLETED',output_summary=:o,completed_at=NOW() WHERE id=:id"),{"o":f"{len(p)} properties, {len(r)} requirements","id":jid})
    except Exception as ex:
        with engine.begin() as c:
            c.execute(text("UPDATE pi_sources SET ingestion_status='FAILED',error_message=:e WHERE id=:id"),{"e":str(ex),"id":sid})
            c.execute(text("UPDATE pi_ai_jobs SET status='FAILED',error_message=:e,completed_at=NOW() WHERE id=:id"),{"e":str(ex),"id":jid})


def parse_envelope_response(resp):
    if getattr(resp,"parsed",None) is not None:
        return Envelope.model_validate(resp.parsed)

    raw=(resp.text or "").strip()

    try:
        return Envelope.model_validate_json(raw)
    except Exception as first_error:
        # Gemini may wrap JSON in markdown fences.
        cleaned=raw
        if cleaned.startswith("```json"):
            cleaned=cleaned[7:]
        elif cleaned.startswith("```"):
            cleaned=cleaned[3:]
        if cleaned.endswith("```"):
            cleaned=cleaned[:-3]
        cleaned=cleaned.strip()

        try:
            return Envelope.model_validate_json(cleaned)
        except Exception:
            raise RuntimeError(
                "GEMINI_JSON_TRUNCATED_OR_INVALID: "
                + str(first_error)
                + f" | response_chars={len(raw)}"
            )

def extract_gemini_batch(path,mime,label):
    if not client:
        raise RuntimeError("GEMINI_API_KEY missing")

    uploaded=client.files.upload(
        file=path,
        config={"mime_type":mime}
    )

    resp=client.models.generate_content(
        model=GEMINI_MODEL,
        contents=[
            PROMPT
            + "\n"
            + label
            + "\nReturn compact JSON. Do not repeat source text. Do not add explanations."
            ,
            uploaded
        ],
        config=types.GenerateContentConfig(
            response_mime_type="application/json",
            response_schema=Envelope,
            temperature=0.0,
            max_output_tokens=16384
        )
    )

    return parse_envelope_response(resp)

def split_pdf(path):
    reader=PdfReader(path)
    total=len(reader.pages)
    batches=[]
    for start in range(0,total,PDF_PAGES_PER_BATCH):
        end=min(start+PDF_PAGES_PER_BATCH,total)
        writer=PdfWriter()
        for i in range(start,end): writer.add_page(reader.pages[i])
        fd,bp=tempfile.mkstemp(suffix=".pdf"); os.close(fd)
        with open(bp,"wb") as f: writer.write(f)
        batches.append((bp,start+1,end,total))
    return batches


def write_pdf_range(reader,start_idx,end_idx):
    writer=PdfWriter()
    for i in range(start_idx,end_idx):
        writer.add_page(reader.pages[i])
    fd,temp_path=tempfile.mkstemp(suffix=".pdf")
    os.close(fd)
    with open(temp_path,"wb") as f:
        writer.write(f)
    return temp_path


def batch_state(source_id,start_page,end_page):
    with engine.connect() as c:
        row=c.execute(
            text("""SELECT status,attempts,records_created,duplicates,error_message
                    FROM pi_extraction_batches
                    WHERE source_id=:sid AND start_page=:sp AND end_page=:ep"""),
            {"sid":source_id,"sp":start_page,"ep":end_page}
        ).first()
    return dict(row._mapping) if row else None

def mark_batch(source_id,start_page,end_page,status,created=0,duplicates=0,error=None):
    with engine.begin() as c:
        c.execute(
            text("""INSERT INTO pi_extraction_batches(
                        source_id,start_page,end_page,status,attempts,records_created,duplicates,error_message,updated_at
                    ) VALUES(:sid,:sp,:ep,:status,1,:created,:dup,:err,NOW())
                    ON CONFLICT(source_id,start_page,end_page)
                    DO UPDATE SET
                        status=EXCLUDED.status,
                        attempts=pi_extraction_batches.attempts+1,
                        records_created=EXCLUDED.records_created,
                        duplicates=EXCLUDED.duplicates,
                        error_message=EXCLUDED.error_message,
                        updated_at=NOW()"""),
            {"sid":source_id,"sp":start_page,"ep":end_page,"status":status,
             "created":created,"dup":duplicates,"err":error}
        )

def extract_pdf_range_recursive(reader,start_idx,end_idx,total_pages,sid,jid,progress):
    """
    Extract a page range. If Gemini returns truncated/invalid JSON,
    split the range in half and retry automatically.
    """
    start_page=start_idx+1
    end_page=end_idx
    previous=batch_state(sid,start_page,end_page)
    if previous and previous.get("status")=="COMPLETED":
        return {
            "created":int(previous.get("records_created") or 0),
            "duplicates":int(previous.get("duplicates") or 0),
            "property_outputs":0,
            "requirement_outputs":0
        }

    mark_batch(sid,start_page,end_page,"RUNNING")
    temp_path=write_pdf_range(reader,start_idx,end_idx)
    try:
        label=f"PDF pages {start_page}-{end_page} of {total_pages}. Extract every distinct listing."

        try:
            env=extract_gemini_batch(
                temp_path,
                "application/pdf",
                label
            )
        except Exception as exc:
            msg=str(exc)

            # If a multi-page range is too large, split and retry.
            if "GEMINI_JSON_TRUNCATED_OR_INVALID" in msg and (end_idx-start_idx)>1:
                mid=start_idx + (end_idx-start_idx)//2

                left=extract_pdf_range_recursive(
                    reader,start_idx,mid,total_pages,sid,jid,progress
                )
                right=extract_pdf_range_recursive(
                    reader,mid,end_idx,total_pages,sid,jid,progress
                )
                combined_created=left["created"]+right["created"]
                combined_duplicates=left["duplicates"]+right["duplicates"]
                mark_batch(sid,start_page,end_page,"COMPLETED",combined_created,combined_duplicates,None)
                return {
                    "created":combined_created,
                    "duplicates":combined_duplicates,
                    "property_outputs":left["property_outputs"]+right["property_outputs"],
                    "requirement_outputs":left["requirement_outputs"]+right["requirement_outputs"]
                }

            # A single page can still contain a huge classifieds grid.
            # Retry once with an even stricter compact-output instruction.
            if "GEMINI_JSON_TRUNCATED_OR_INVALID" in msg and (end_idx-start_idx)==1:
                strict_label=(
                    label
                    + "\nThis single page contains many ads. "
                    + "Return ONLY the schema fields. "
                    + "Use null for unknowns. "
                    + "Keep remarks concise. "
                    + "Do not copy long advertisement text."
                )
                env=extract_gemini_batch(
                    temp_path,
                    "application/pdf",
                    strict_label
                )
            else:
                raise

        props=[
            save_property(
                {
                    **x.model_dump(exclude={"record_type"}),
                    "source":f"AI_SOURCE_{sid}_PAGES_{start_idx+1}_{end_idx}"
                },
                sid
            )
            for x in env.properties
        ]

        reqs=[
            save_requirement(
                {
                    **x.model_dump(exclude={"record_type"}),
                    "source":f"AI_SOURCE_{sid}_PAGES_{start_idx+1}_{end_idx}"
                },
                sid
            )
            for x in env.requirements
        ]

        created=sum(x["status"]=="created" for x in props+reqs)
        duplicates=sum(x["status"]=="duplicate" for x in props+reqs)

        progress["created"]+=created
        progress["duplicates"]+=duplicates
        progress["ranges_done"]+=1

        with engine.begin() as c:
            c.execute(
                text("""UPDATE pi_sources
                        SET ingestion_status='PROCESSING',
                            processed_records=:n,
                            duplicate_records=:d
                        WHERE id=:id"""),
                {
                    "n":progress["created"],
                    "d":progress["duplicates"],
                    "id":sid
                }
            )

            c.execute(
                text("""UPDATE pi_ai_jobs
                        SET output_summary=:o
                        WHERE id=:id"""),
                {
                    "o":(
                        f"{progress['created']} records stored; "
                        f"processed through page {end_idx} of {total_pages}"
                    ),
                    "id":jid
                }
            )

        mark_batch(sid,start_page,end_page,"COMPLETED",created,duplicates,None)
        return {
            "created":created,
            "duplicates":duplicates,
            "property_outputs":len(props),
            "requirement_outputs":len(reqs)
        }

    except Exception as exc:
        mark_batch(sid,start_page,end_page,"FAILED",0,0,str(exc))
        raise

    finally:
        try:
            os.unlink(temp_path)
        except Exception:
            pass

def scan_tile_state(source_id,page_number,tile_label):
    with engine.connect() as c:
        row=c.execute(text("""SELECT status,attempts,records_created,duplicates,error_message
                              FROM pi_scan_tiles
                              WHERE source_id=:sid AND page_number IS NOT DISTINCT FROM :pg AND tile_label=:tile"""),
                      {"sid":source_id,"pg":page_number,"tile":tile_label}).first()
    return dict(row._mapping) if row else None

def mark_scan_tile(source_id,page_number,tile_label,status,created=0,duplicates=0,error=None):
    with engine.begin() as c:
        c.execute(text("""INSERT INTO pi_scan_tiles(source_id,page_number,tile_label,status,attempts,records_created,duplicates,error_message,updated_at)
                          VALUES(:sid,:pg,:tile,:status,1,:created,:dup,:err,NOW())
                          ON CONFLICT(source_id,page_number,tile_label)
                          DO UPDATE SET status=EXCLUDED.status,attempts=pi_scan_tiles.attempts+1,
                                        records_created=EXCLUDED.records_created,duplicates=EXCLUDED.duplicates,
                                        error_message=EXCLUDED.error_message,updated_at=NOW()"""),
                  {"sid":source_id,"pg":page_number,"tile":tile_label,"status":status,
                   "created":created,"dup":duplicates,"err":error})

def save_scanned_envelope(env,sid,suffix):
    props=[save_property({**x.model_dump(exclude={"record_type"}),"source":f"AI_SOURCE_{sid}_{suffix}"},sid) for x in env.properties]
    reqs=[save_requirement({**x.model_dump(exclude={"record_type"}),"source":f"AI_SOURCE_{sid}_{suffix}"},sid) for x in env.requirements]
    return (sum(x["status"]=="created" for x in props+reqs),
            sum(x["status"]=="duplicate" for x in props+reqs),len(props),len(reqs))

def crop_overlapping_tiles(image_path):
    image=Image.open(image_path).convert("RGB")
    w,h=image.size
    cols=max(1,SCAN_TILE_COLS); rows=max(1,SCAN_TILE_ROWS)
    ov=max(0.0,min(SCAN_TILE_OVERLAP,0.35))
    cw=w/cols; ch=h/rows; out=[]
    for r in range(rows):
        for c in range(cols):
            x0=max(0,int(c*cw-cw*ov)); y0=max(0,int(r*ch-ch*ov))
            x1=min(w,int((c+1)*cw+cw*ov)); y1=min(h,int((r+1)*ch+ch*ov))
            tile=image.crop((x0,y0,x1,y1))
            fd,tp=tempfile.mkstemp(suffix=".jpg"); os.close(fd)
            tile.save(tp,"JPEG",quality=95)
            out.append((f"R{r+1}C{c+1}",tp))
    return out

def scan_image_exhaustive(image_path,sid,jid,page_number=1):
    total={"created":0,"duplicates":0,"property_outputs":0,"requirement_outputs":0,"failed":0}
    units=[("FULL_PAGE",image_path,False)]+[(label,tp,True) for label,tp in crop_overlapping_tiles(image_path)]
    for label,path,is_temp in units:
        try:
            state=scan_tile_state(sid,page_number,label)
            if state and state.get("status")=="COMPLETED":
                continue
            mark_scan_tile(sid,page_number,label,"RUNNING")
            prompt=(PROMPT+"\nHigh-recall classified scanner. Extract EVERY distinct property advertisement visible in this image region. "
                    "Do not summarize, sample or merge neighboring ads. Preserve all legible phone numbers and names. "
                    "Overlapping crops are intentional; duplicates will be handled later. Region: "+label)
            env=extract_gemini_batch(path,"image/jpeg",prompt)
            cr,du,po,ro=save_scanned_envelope(env,sid,f"PAGE_{page_number}_{label}")
            mark_scan_tile(sid,page_number,label,"COMPLETED",cr,du,None)
            total["created"]+=cr; total["duplicates"]+=du; total["property_outputs"]+=po; total["requirement_outputs"]+=ro
            with engine.begin() as c:
                c.execute(text("UPDATE pi_ai_jobs SET output_summary=:o WHERE id=:id"),
                          {"o":f"Page {page_number} {label}: {total['created']} new records in this page scan","id":jid})
        except Exception as exc:
            mark_scan_tile(sid,page_number,label,"FAILED",0,0,str(exc)); total["failed"]+=1
        finally:
            if is_temp:
                try: os.unlink(path)
                except Exception: pass
    return total

def render_pdf_page(doc,page_index):
    page=doc.load_page(page_index)
    pix=page.get_pixmap(matrix=fitz.Matrix(PDF_RENDER_DPI/72.0,PDF_RENDER_DPI/72.0),alpha=False)
    fd,p=tempfile.mkstemp(suffix=".jpg"); os.close(fd); pix.save(p); return p

def run_file_job(sid,jid,path,mime):
    try:
        created=duplicates=property_outputs=requirement_outputs=failed=0
        is_pdf=(mime=="application/pdf" or path.lower().endswith(".pdf"))
        is_image=(mime or "").startswith("image/") or path.lower().endswith((".jpg",".jpeg",".png",".webp"))
        if is_pdf:
            doc=fitz.open(path)
            with engine.begin() as c:
                c.execute(text("UPDATE pi_ai_jobs SET input_summary=:x WHERE id=:id"),
                          {"x":f"Exhaustive scanner: {doc.page_count} PDF pages, full-page plus overlapping tiles","id":jid})
            for i in range(doc.page_count):
                page_img=None
                try:
                    page_img=render_pdf_page(doc,i)
                    r=scan_image_exhaustive(page_img,sid,jid,i+1)
                    created+=r["created"]; duplicates+=r["duplicates"]; property_outputs+=r["property_outputs"]; requirement_outputs+=r["requirement_outputs"]; failed+=r["failed"]
                except Exception as exc:
                    failed+=1
                    with engine.begin() as c:
                        c.execute(text("UPDATE pi_ai_jobs SET output_summary=:o WHERE id=:id"),{"o":f"Continuing after page {i+1} error: {exc}","id":jid})
                finally:
                    if page_img:
                        try: os.unlink(page_img)
                        except Exception: pass
            doc.close()
        elif is_image:
            fd,jpg=tempfile.mkstemp(suffix=".jpg"); os.close(fd)
            try:
                Image.open(path).convert("RGB").save(jpg,"JPEG",quality=95)
                r=scan_image_exhaustive(jpg,sid,jid,1)
                created=r["created"]; duplicates=r["duplicates"]; property_outputs=r["property_outputs"]; requirement_outputs=r["requirement_outputs"]; failed=r["failed"]
            finally:
                try: os.unlink(jpg)
                except Exception: pass
        else:
            env=extract_gemini_batch(path,mime,"Extract every distinct property and requirement from this source.")
            created,duplicates,property_outputs,requirement_outputs=save_scanned_envelope(env,sid,"SOURCE")
        status="PROCESSED_WITH_ERRORS" if failed else "PROCESSED"
        with engine.begin() as c:
            c.execute(text("""UPDATE pi_sources SET ingestion_status=:st,processed_records=:n,duplicate_records=:d,
                              ai_provider='gemini',ai_model=:m,processed_at=NOW() WHERE id=:id"""),
                      {"st":status,"n":created,"d":duplicates,"m":GEMINI_MODEL,"id":sid})
            c.execute(text("""UPDATE pi_ai_jobs SET status='COMPLETED',output_summary=:o,completed_at=NOW() WHERE id=:id"""),
                      {"o":f"Exhaustive scan complete: {created} new records, {duplicates} duplicates/overlap, {failed} failed units","id":jid})
    except Exception as ex:
        with engine.begin() as c:
            c.execute(text("UPDATE pi_sources SET ingestion_status='FAILED',error_message=:e WHERE id=:id"),{"e":str(ex),"id":sid})
            c.execute(text("UPDATE pi_ai_jobs SET status='FAILED',error_message=:e,completed_at=NOW() WHERE id=:id"),{"e":str(ex),"id":jid})
    finally:
        try: os.unlink(path)
        except Exception: pass

def create_job(sid,kind,summary):
    sql="INSERT INTO pi_ai_jobs(source_id,job_type,status,provider,model,input_summary,started_at) VALUES(:s,:k,'RUNNING','gemini',:m,:x,NOW()) RETURNING id"
    with engine.begin() as c:
        return c.execute(text(sql),{"s":sid,"k":kind,"m":GEMINI_MODEL,"x":summary}).scalar_one()

@app.get("/login",response_class=HTMLResponse)
def login_page(req:Request):
    existing=get_role(req)
    if existing:
        return RedirectResponse("/workspace",status_code=303)

    html="""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1">
<title>Property Intelligence Login</title>
<style>
*{box-sizing:border-box}
body{font-family:Arial,sans-serif;background:#f4f6f8;margin:0;min-height:100vh;display:grid;place-items:center;padding:18px}
.card{width:min(430px,100%);background:white;padding:26px;border-radius:14px;border:1px solid #e5e7eb;box-shadow:0 10px 30px rgba(0,0,0,.06)}
h2{margin:0 0 6px}.muted{color:#6b7280;margin:0 0 18px}
label{display:block;font-size:13px;font-weight:700;margin-top:10px}
select,input,button{width:100%;padding:13px;margin:6px 0;border-radius:9px;border:1px solid #d1d5db;font-size:16px}
button{background:#111827;color:white;border:0;font-weight:700;cursor:pointer;margin-top:14px}
.note{font-size:12px;color:#6b7280;margin-top:14px;line-height:1.5}
</style>
</head>
<body>
<form class="card" method="post" action="/login">
<h2>Property Intelligence</h2>
<p class="muted">Unified Team & Admin Workspace</p>
<label>Login as</label>
<select name="role">
<option value="team">Team</option>
<option value="admin">Admin</option>
</select>
<label>Access code</label>
<input name="code" type="password" placeholder="Enter access code" autocomplete="current-password" required>
<button type="submit">Open Workspace</button>
<p class="note">Use this same website on Android, iPhone, Windows or Mac. Each device logs in separately.</p>
</form>
</body>
</html>"""
    return HTMLResponse(html)

@app.post("/login")
def login_post(role:str=Form(...),code:str=Form(...)):
    ok=(role=="team" and hmac.compare_digest(code,TEAM_CODE)) or (role=="admin" and hmac.compare_digest(code,ADMIN_CODE))
    if not ok:return HTMLResponse("Invalid code",401)
    resp=RedirectResponse("/workspace",303)
    resp.set_cookie(
        "pi_session",
        signed(role),
        httponly=True,
        secure=True,
        samesite="lax",
        max_age=60*60*24*7,
        path="/"
    )
    return resp

@app.get("/logout")
def logout():
    resp=RedirectResponse("/login",303);resp.delete_cookie("pi_session",path="/");return resp

@app.get("/health")
def health():
    with engine.connect() as c:c.execute(text("SELECT 1"))
    return {"status":"ok","service":"property-intelligence-unified","version":VERSION,"database":"connected","gemini_configured":bool(GEMINI_API_KEY)}


@app.get("/api/upload-status/{job_id}")
def upload_status(job_id:int,req:Request):
    need_login(req)
    with engine.connect() as c:
        row=c.execute(
            text("""SELECT j.id,j.status,j.output_summary,j.error_message,
                           j.created_at,j.completed_at,
                           s.id AS source_id,s.ingestion_status,s.processed_records,
                           s.duplicate_records,s.original_filename
                    FROM pi_ai_jobs j
                    LEFT JOIN pi_sources s ON s.id=j.source_id
                    WHERE j.id=:id"""),
            {"id":job_id}
        ).first()
    if not row:
        raise HTTPException(404,"Upload job not found")
    d={}
    for k,v in dict(row._mapping).items():
        if isinstance(v,(date,datetime)):
            d[k]=v.isoformat()
        elif isinstance(v,Decimal):
            d[k]=float(v)
        else:
            d[k]=v
    return {"status":"ok","job":d}

@app.get("/api/status")
def status(req:Request):
    need_login(req)
    with engine.connect() as c:
        tables={"properties":"pi_properties","requirements":"pi_requirements","sources":"pi_sources","matches":"pi_matches","ai_jobs":"pi_ai_jobs"}
        counts={k:c.execute(text("SELECT COUNT(*) FROM "+v)).scalar_one() for k,v in tables.items()}
    return {"status":"ok","role":get_role(req),"records":counts,"gemini_configured":bool(GEMINI_API_KEY)}


def serialize_db_value(value):
    if isinstance(value,(date,datetime)):
        return value.isoformat()
    if isinstance(value,Decimal):
        return float(value)
    return value

def property_with_verification_status(row):
    d={k:serialize_db_value(v) for k,v in dict(row._mapping).items()}
    verified=d.get("verified_date")
    if not verified:
        d["last_verified_label"]="Never Verified"
        d["verification_due"]=True
        d["verification_age_days"]=None
        return d

    verified_date=date.fromisoformat(verified) if isinstance(verified,str) else verified
    age=(date.today()-verified_date).days
    d["verification_age_days"]=age
    d["verification_due"]=age>=VERIFICATION_DUE_DAYS
    d["last_verified_label"]=(
        f"Verification Due · Last verified {verified}"
        if d["verification_due"]
        else f"Last verified {verified}"
    )
    return d

@app.post("/api/properties")
def add_property(p:Property,req:Request):
    need_login(req);return save_property(p.model_dump())


@app.get("/api/properties/{property_id}")
def get_property(property_id:str,req:Request):
    need_login(req)
    with engine.connect() as c:
        row=c.execute(
            text("SELECT * FROM pi_properties WHERE property_id=:pid"),
            {"pid":property_id}
        ).first()
    if not row:
        raise HTTPException(404,"Property not found")
    return {"status":"ok","property":property_with_verification_status(row)}

@app.put("/api/properties/{property_id}")
def update_property(property_id:str,p:Property,req:Request):
    need_login(req)
    actor=actor_name(req)
    data=p.model_dump()

    with engine.begin() as c:
        row=c.execute(
            text("SELECT * FROM pi_properties WHERE property_id=:pid FOR UPDATE"),
            {"pid":property_id}
        ).first()
        if not row:
            raise HTTPException(404,"Property not found")

        old={k:serialize_db_value(v) for k,v in dict(row._mapping).items()}

        allowed=[
            "property_name","property_type","city","location","available_area_sqft",
            "minimum_area_sqft","maximum_area_sqft","floor","rent_or_sale",
            "nearby_brands","suitable_category","parking","owner_name","owner_contact",
            "broker_name","broker_contact","remarks","image_urls","video_urls","brochure_url"
        ]

        params={k:data.get(k) for k in allowed}
        params["pid"]=property_id

        sets=",".join(f"{k}=:{k}" for k in allowed)
        c.execute(
            text(f"UPDATE pi_properties SET {sets},updated_at=NOW() WHERE property_id=:pid"),
            params
        )

        newrow=c.execute(
            text("SELECT * FROM pi_properties WHERE property_id=:pid"),
            {"pid":property_id}
        ).first()
        newd={k:serialize_db_value(v) for k,v in dict(newrow._mapping).items()}

        c.execute(
            text("""INSERT INTO pi_verification_log(
                        property_id,action,performed_by,notes,old_value,new_value
                    ) VALUES(
                        :pid,'PROPERTY_EDIT',:actor,'Property details edited',
                        CAST(:old AS JSONB),CAST(:new AS JSONB)
                    )"""),
            {
                "pid":property_id,
                "actor":actor,
                "old":json.dumps(old,default=str),
                "new":json.dumps(newd,default=str)
            }
        )

    return {"status":"updated","property_id":property_id,"property":property_with_verification_status(newrow)}

@app.post("/api/properties/{property_id}/verify")
def verify_property(property_id:str,req:Request):
    need_login(req)
    actor=actor_name(req)

    with engine.begin() as c:
        row=c.execute(
            text("SELECT verified_date,verified_by FROM pi_properties WHERE property_id=:pid FOR UPDATE"),
            {"pid":property_id}
        ).first()
        if not row:
            raise HTTPException(404,"Property not found")

        old={
            "verified_date":serialize_db_value(row._mapping["verified_date"]),
            "verified_by":row._mapping["verified_by"]
        }

        c.execute(
            text("""UPDATE pi_properties
                    SET verified_date=CURRENT_DATE,
                        verified_by=:actor,
                        verification_status='VERIFIED',
                        updated_at=NOW()
                    WHERE property_id=:pid"""),
            {"actor":actor,"pid":property_id}
        )

        c.execute(
            text("""INSERT INTO pi_verification_log(
                        property_id,action,performed_by,notes,old_value,new_value
                    ) VALUES(
                        :pid,'VERIFIED',:actor,'Property verified',
                        CAST(:old AS JSONB),CAST(:new AS JSONB)
                    )"""),
            {
                "pid":property_id,
                "actor":actor,
                "old":json.dumps(old,default=str),
                "new":json.dumps({
                    "verified_date":date.today().isoformat(),
                    "verified_by":actor,
                    "verification_status":"VERIFIED"
                })
            }
        )

    return {
        "status":"verified",
        "property_id":property_id,
        "verified_date":date.today().isoformat(),
        "verified_by":actor,
        "next_verification_due_days":VERIFICATION_DUE_DAYS
    }


@app.post("/api/properties/{property_id}/media")
async def upload_property_media(property_id:str,req:Request,file:UploadFile=File(...)):
    need_login(req)

    filename=file.filename or "property-image.jpg"
    mime=(file.content_type or "").lower()
    allowed={"image/jpeg","image/png","image/webp","image/gif"}

    if mime not in allowed:
        ext=os.path.splitext(filename)[1].lower()
        extmap={".jpg":"image/jpeg",".jpeg":"image/jpeg",".png":"image/png",".webp":"image/webp",".gif":"image/gif"}
        mime=extmap.get(ext,mime)

    if mime not in allowed:
        raise HTTPException(400,"Only JPG, JPEG, PNG, WEBP and GIF images are allowed.")

    data=await file.read()
    if not data:
        raise HTTPException(400,"Image file is empty.")

    if len(data) > MAX_IMAGE_MB*1024*1024:
        raise HTTPException(413,f"Each image must be {MAX_IMAGE_MB} MB or smaller.")

    with engine.begin() as c:
        exists=c.execute(
            text("SELECT 1 FROM pi_properties WHERE property_id=:pid"),
            {"pid":property_id}
        ).first()
        if not exists:
            raise HTTPException(404,"Property not found.")

        count=c.execute(
            text("SELECT COUNT(*) FROM pi_property_media WHERE property_id=:pid"),
            {"pid":property_id}
        ).scalar_one()

        if count >= MAX_PROPERTY_IMAGES:
            raise HTTPException(413,f"Maximum {MAX_PROPERTY_IMAGES} images per property.")

        media_id=str(uuid.uuid4())
        c.execute(
            text("""INSERT INTO pi_property_media(
                        media_id,property_id,media_type,filename,mime_type,file_size,content
                    ) VALUES(
                        CAST(:mid AS UUID),:pid,'IMAGE',:fn,:mime,:size,:content
                    )"""),
            {
                "mid":media_id,
                "pid":property_id,
                "fn":filename,
                "mime":mime,
                "size":len(data),
                "content":data
            }
        )

    base=str(req.base_url).rstrip("/")
    public_url=f"{base}/media/{media_id}"

    with engine.begin() as c:
        old=c.execute(
            text("SELECT image_urls FROM pi_properties WHERE property_id=:pid"),
            {"pid":property_id}
        ).scalar_one_or_none()

        urls=[u.strip() for u in str(old or "").splitlines() if u.strip()]
        if public_url not in urls:
            urls.append(public_url)

        c.execute(
            text("UPDATE pi_properties SET image_urls=:urls,updated_at=NOW() WHERE property_id=:pid"),
            {"urls":"\n".join(urls),"pid":property_id}
        )

    return {
        "status":"uploaded",
        "property_id":property_id,
        "media_id":media_id,
        "url":public_url,
        "filename":filename
    }

@app.get("/media/{media_id}")
def public_property_media(media_id:str):
    try:
        uuid.UUID(media_id)
    except Exception:
        raise HTTPException(404,"Media not found.")

    with engine.connect() as c:
        row=c.execute(
            text("""SELECT mime_type,content,filename
                    FROM pi_property_media
                    WHERE media_id=CAST(:mid AS UUID)"""),
            {"mid":media_id}
        ).first()

    if not row:
        raise HTTPException(404,"Media not found.")

    return Response(
        content=bytes(row._mapping["content"]),
        media_type=row._mapping["mime_type"] or "application/octet-stream",
        headers={
            "Cache-Control":"public, max-age=31536000, immutable",
            "Content-Disposition":f'inline; filename="{row._mapping["filename"] or "property-image"}"'
        }
    )

@app.post("/api/requirements")
def add_requirement(r:Requirement,req:Request):
    need_login(req);return save_requirement(r.model_dump())

@app.post("/api/ingest/text")
def ingest_text(p:TextInput,bg:BackgroundTasks,req:Request):
    need_login(req)
    sid=source_row(p.source_type,p.source_name,reference=p.text_content)
    jid=create_job(sid,"TEXT_EXTRACTION",p.source_name or p.source_type)
    bg.add_task(run_text_job,sid,jid,p.text_content)
    return {"status":"ACCEPTED","source_id":sid,"job_id":jid}

@app.post("/api/ingest/file")
async def ingest_file(
    bg:BackgroundTasks,
    req:Request,
    file:UploadFile=File(...),
    source_type:str=Query("DOCUMENT"),
    source_name:Optional[str]=Query(None)
):
    need_login(req)

    filename=file.filename or "upload.bin"
    ext=os.path.splitext(filename)[1].lower()
    mime=file.content_type or "application/octet-stream"

    # Browser/device MIME types can be missing or inconsistent.
    mime_map={
        ".jpg":"image/jpeg",
        ".jpeg":"image/jpeg",
        ".png":"image/png",
        ".webp":"image/webp",
        ".pdf":"application/pdf",
        ".csv":"text/csv",
        ".txt":"text/plain"
    }
    if mime in {"application/octet-stream",""} or not mime:
        mime=mime_map.get(ext,"application/octet-stream")

    suffix=ext or ".bin"

    # Gemini PDF processing limit is 50 MB.
    pdf_limit=50*1024*1024
    app_limit=MAX_UPLOAD_MB*1024*1024

    fd,path=tempfile.mkstemp(suffix=suffix)
    os.close(fd)

    total=0
    try:
        with open(path,"wb") as out:
            while True:
                chunk=await file.read(1024*1024)  # 1 MB chunks
                if not chunk:
                    break
                total += len(chunk)

                if total > app_limit:
                    raise HTTPException(
                        413,
                        f"File too large. Maximum workspace upload is {MAX_UPLOAD_MB} MB."
                    )

                if (mime=="application/pdf" or filename.lower().endswith(".pdf")) and total > pdf_limit:
                    raise HTTPException(
                        413,
                        "PDF too large for Gemini processing. Maximum PDF size is 50 MB. Please split/compress the PDF."
                    )

                out.write(chunk)

        sid=source_row(
            source_type.upper(),
            source_name or filename,
            filename,
            mime
        )

        # CSV is processed directly after streamed upload.
        if filename.lower().endswith(".csv"):
            inserted=0
            duplicates=0
            errors=[]

            with open(path,"r",encoding="utf-8-sig",errors="replace",newline="") as csvfile:
                reader=csv.DictReader(csvfile)
                for line,row in enumerate(reader,start=2):
                    item={
                        "property_name":row.get("Property name") or row.get("Property Name"),
                        "property_type":row.get("Property type") or row.get("Property Type") or "NA",
                        "city":row.get("City") or "NA",
                        "location":row.get("Location") or "NA",
                        "available_area_sqft":row.get("Available area") or row.get("Available Area") or None,
                        "minimum_area_sqft":row.get("Minimum area") or row.get("Minimum Area") or None,
                        "maximum_area_sqft":row.get("Maximum area") or row.get("Maximum Area") or None,
                        "floor":row.get("Floor"),
                        "rent_or_sale":row.get("Rent/Sale"),
                        "nearby_brands":row.get("Nearby brand") or row.get("Nearby brands"),
                        "suitable_category":row.get("Suitable category"),
                        "parking":row.get("Parking"),
                        "owner_name":row.get("Owner name"),
                        "owner_contact":row.get("Owner contact"),
                        "broker_name":row.get("Broker name"),
                        "broker_contact":row.get("Broker contact"),
                        "remarks":row.get("Remarks"),
                        "source":"CSV:"+filename
                    }

                    for numkey in ["available_area_sqft","minimum_area_sqft","maximum_area_sqft"]:
                        try:
                            item[numkey]=float(str(item[numkey]).replace(",","")) if item[numkey] not in (None,"") else None
                        except Exception:
                            item[numkey]=None

                    try:
                        result=save_property(item,sid)
                        if result["status"]=="created":
                            inserted+=1
                        else:
                            duplicates+=1
                    except Exception as exc:
                        errors.append({"row":line,"error":str(exc)})

            with engine.begin() as c:
                c.execute(
                    text("""UPDATE pi_sources
                            SET ingestion_status=:status,
                                processed_records=:n,
                                duplicate_records=:d,
                                error_message=:e,
                                processed_at=NOW()
                            WHERE id=:id"""),
                    {
                        "status":"PROCESSED" if not errors else "PROCESSED_WITH_ERRORS",
                        "n":inserted,
                        "d":duplicates,
                        "e":json.dumps(errors[:20]) if errors else None,
                        "id":sid
                    }
                )

            try: os.unlink(path)
            except: pass

            return {
                "status":"PROCESSED" if not errors else "PROCESSED_WITH_ERRORS",
                "source_id":sid,
                "inserted":inserted,
                "duplicates":duplicates,
                "errors":errors[:20],
                "file_size_mb":round(total/1024/1024,2)
            }

        jid=create_job(sid,"FILE_EXTRACTION",filename)
        bg.add_task(run_file_job,sid,jid,path,mime)

        return {
            "status":"ACCEPTED",
            "source_id":sid,
            "job_id":jid,
            "file_size_mb":round(total/1024/1024,2),
            "message":"Upload received. AI processing continues in background."
        }

    except Exception:
        try: os.unlink(path)
        except: pass
        raise

TABLES={"properties":"pi_properties","requirements":"pi_requirements","matches":"pi_matches","whatsapp_drafts":"pi_message_drafts","sources":"pi_sources","ai_jobs":"pi_ai_jobs","verification":"pi_verification_log","batches":"pi_extraction_batches","media":"pi_property_media","scan_tiles":"pi_scan_tiles"}
PRIVATE={"fingerprint","owner_name","owner_contact","broker_name","broker_contact","remarks","verified_by","verified_date","source"}

@app.get("/api/database/{name}")
def database(name:str,req:Request,limit:int=Query(500,ge=1,le=2000)):
    role=need_login(req)
    if name not in TABLES:raise HTTPException(404,"Unknown table")
    if role!="admin" and name in {"sources","ai_jobs","verification","batches","media"}:raise HTTPException(403,"Admin only")
    with engine.connect() as c:
        if name=="media":
            result=c.execute(text("""SELECT id,media_id,property_id,media_type,filename,mime_type,file_size,created_at
                                     FROM pi_property_media ORDER BY id DESC LIMIT :n"""),{"n":limit})
        else:
            result=c.execute(text("SELECT * FROM "+TABLES[name]+" ORDER BY id DESC LIMIT :n"),{"n":limit})
        rows=[]
        for row in result:
            d={}
            for k,v in dict(row._mapping).items():
                d[k]=v.isoformat() if isinstance(v,(date,datetime)) else float(v) if isinstance(v,Decimal) else v
            rows.append(d)
    if name=="properties":
        enhanced=[]
        for x in rows:
            verified=x.get("verified_date")
            if not verified:
                x["last_verified"]="Never Verified"
                x["verification_due"]=True
            else:
                vd=date.fromisoformat(verified) if isinstance(verified,str) else verified
                age=(date.today()-vd).days
                x["last_verified"]=f"{verified} ({age} days ago)"
                x["verification_due"]=age>=VERIFICATION_DUE_DAYS
            enhanced.append(x)
        rows=enhanced

    if name=="properties" and role!="admin":
        # Team can see verification date/status but not private owner/broker/internal fields.
        team_private=PRIVATE-{"verified_date","verified_by"}
        rows=[{k:v for k,v in x.items() if k not in team_private} for x in rows]

    return {"status":"ok","table":name,"count":len(rows),"rows":rows}


def fallback_whatsapp_message(requirement,top_properties):
    top_properties=[p for p in top_properties if _norm(p.get("verification_status"))=="verified"]
    name=requirement.get("client_name") or requirement.get("company_name") or "there"
    if not top_properties:
        return "No WhatsApp draft created because no shortlisted property has been verified for current availability."
    lines=[f"Hi {name}, we found a few verified property options matching your requirement:"]
    for i,p in enumerate(top_properties[:3],1):
        bits=[f"{i}. {p.get('property_name') or p.get('property_id')}"]
        if p.get("location"): bits.append(str(p["location"]))
        if p.get("available_area_sqft") is not None: bits.append(f"{p['available_area_sqft']} sq ft")
        if p.get("rent_or_sale"): bits.append(str(p["rent_or_sale"]))
        lines.append(" | ".join(bits))
        if p.get("image_urls"): lines.append("Photos: "+str(p["image_urls"]))
        if p.get("video_urls"): lines.append("Video: "+str(p["video_urls"]))
        if p.get("brochure_url"): lines.append("Brochure: "+str(p["brochure_url"]))
    lines.append("Please let me know which option you would like to review in detail or schedule a site visit for.")
    return "\\n".join(lines)

def generate_whatsapp_message(requirement,top_properties):
    fallback=fallback_whatsapp_message(requirement,top_properties)
    if not client:
        return fallback,"fallback"

    safe_properties=[]
    for p in [x for x in top_properties if _norm(x.get('verification_status'))=='verified'][:5]:
        safe_properties.append({
            "property_id":p.get("property_id"),
            "property_name":p.get("property_name"),
            "city":p.get("city"),
            "location":p.get("location"),
            "available_area_sqft":float(p["available_area_sqft"]) if p.get("available_area_sqft") is not None else None,
            "floor":p.get("floor"),
            "rent_or_sale":p.get("rent_or_sale"),
            "nearby_brands":p.get("nearby_brands"),
            "suitable_category":p.get("suitable_category"),
            "image_urls":p.get("image_urls"),
            "video_urls":p.get("video_urls"),
            "brochure_url":p.get("brochure_url"),
        })

    prompt=f"""Write one concise professional WhatsApp message for a real-estate client.
Use ONLY the supplied requirement and matched property facts.
Do not reveal owner names, broker names, owner contacts, broker contacts, internal remarks or source data.
Do not invent rent, price, amenities or availability details.
Mention no more than the best 3 options.
If image_urls, video_urls or brochure_url are available for a selected property, include the relevant links directly under that property.
Never invent, rewrite or shorten a media URL. Copy supplied URLs exactly.
End with a simple call to action for details or a site visit.

Requirement:
{json.dumps({k:v for k,v in requirement.items() if k not in {'fingerprint','source_id'}},default=str)}

Matched properties:
{json.dumps(safe_properties,default=str)}
"""

    try:
        response=client.models.generate_content(
            model=GEMINI_MODEL,
            contents=prompt,
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
                response_schema=WhatsAppDraftResult,
                temperature=0.3
            )
        )
        parsed=WhatsAppDraftResult.model_validate(response.parsed) if getattr(response,"parsed",None) is not None else WhatsAppDraftResult.model_validate_json(response.text)
        message=(parsed.message or "").strip()
        return (message if message else fallback),"gemini"
    except Exception:
        return fallback,"fallback"

def store_whatsapp_draft(requirement,message,provider):
    with engine.begin() as c:
        did=c.execute(
            text("""INSERT INTO pi_message_drafts(
                        requirement_id,recipient_name,recipient_phone,channel,message_text,status,ai_provider,ai_model
                    ) VALUES(
                        :rid,:name,:phone,'WHATSAPP',:msg,'READY_FOR_REVIEW',:provider,:model
                    ) RETURNING id"""),
            {
                "rid":requirement.get("requirement_id"),
                "name":requirement.get("client_name") or requirement.get("company_name"),
                "phone":requirement.get("contact_phone"),
                "msg":message,
                "provider":provider,
                "model":GEMINI_MODEL if provider=="gemini" else None
            }
        ).scalar_one()
    return did

@app.post("/api/match/{rid}")
def match(rid:str,req:Request):
    need_login(req)
    return robust_match_requirement(rid, create_whatsapp=True)


WORKSPACE='''<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Property Intelligence</title>
<style>*{box-sizing:border-box}body{font-family:Arial;margin:0;background:#f5f7fb}header{background:#111827;color:white;padding:18px 24px;display:flex;justify-content:space-between}nav{background:white;padding:10px 20px;border-bottom:1px solid #ddd}button{padding:9px 12px;margin:3px;border:0;border-radius:7px;background:#111827;color:white}.wrap{padding:20px}.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(330px,1fr));gap:15px}.card{background:white;padding:16px;border:1px solid #ddd;border-radius:10px}input,select,textarea{width:100%;padding:9px;margin:5px 0}.hidden{display:none}pre{background:#f8fafc;padding:9px;max-height:250px;overflow:auto}table{border-collapse:collapse;width:100%;font-size:12px}th,td{padding:8px;border-bottom:1px solid #eee;text-align:left;white-space:nowrap}.tablebox{overflow:auto;max-height:65vh}</style></head>
<body><header><div><b>Property Intelligence Unified Workspace</b><br><small>Team + Admin on one domain</small></div><div>__ROLE__ | <a href="/logout" style="color:white">Logout</a></div></header>
<nav>
<button type="button" onclick="sec('ops')">Operations</button>
<button type="button" onclick="sec('db')">Database</button>
<button type="button" onclick="sec('status')">Status</button>
__ADMINBTN__
</nav>
<div id="appErrorBanner" style="display:none;margin:12px 20px 0;padding:10px;border-radius:8px;background:#fef2f2;color:#991b1b;border:1px solid #fecaca"></div>
<div class="wrap"><section id="ops"><div class="grid">
<div class="card">
<h3>Upload Photo / Magazine / PDF / CSV</h3>
<input id="sn" placeholder="Source name">
<select id="st">
<option>MAGAZINE</option><option>NEWSPAPER</option><option>PHOTO</option>
<option>PDF</option><option>CSV</option>
</select>
<input id="fu" type="file" accept=".jpg,.jpeg,.png,.webp,.pdf,.csv,.txt">
<button id="uploadBtn" onclick="up()">Upload</button>
<div id="progressWrap" style="display:none;margin-top:14px">
  <div style="display:flex;justify-content:space-between;font-size:13px;margin-bottom:6px">
    <span id="progressText">Preparing upload...</span>
    <b id="progressPct">0%</b>
  </div>
  <div style="height:14px;background:#e5e7eb;border-radius:999px;overflow:hidden">
    <div id="progressBar" style="height:100%;width:0%;background:#111827;transition:width .15s"></div>
  </div>
</div>
<div id="uploadResult" style="margin-top:12px;font-size:14px"></div>
</div>
<div class="card"><h3>Paste WhatsApp / Email</h3><input id="tn" placeholder="Source name"><textarea id="tc" rows="7"></textarea><button onclick="txt()">Process</button><pre id="to"></pre></div>
<div class="card"><h3>Add Property Manually</h3>
<p>The manual property form has moved to one master page so the fields stay consistent.</p>
<a href="/property-manual" style="display:inline-block;padding:10px 14px;border-radius:8px;background:#111827;color:white;text-decoration:none;font-weight:700">Open Add Property Form</a>
</div>
<div class="card"><h3>Edit / Verify Existing Property</h3>
<input id="epid" placeholder="Property ID e.g. PROP-20260810-0000000123">
<button onclick="loadEditProperty()">Load Property</button>
<div id="verifyBadge" style="display:none;padding:10px;border-radius:8px;margin:10px 0;font-weight:700"></div>
<div id="editFields" style="display:none">
<input id="epn" placeholder="Property name">
<input id="ept" placeholder="Property type">
<input id="epc" placeholder="City">
<input id="epl" placeholder="Location">
<input id="epa" type="number" placeholder="Available sqft">
<input id="epf" placeholder="Floor">
<select id="epx"><option>Rent</option><option>Sale</option></select>
<input id="epnb" placeholder="Nearby brands">
<input id="epsc" placeholder="Suitable category">
<input id="eppk" placeholder="Parking">
<input id="epimg" placeholder="Photo links">
<input id="epvid" placeholder="Video links">
<input id="epbro" placeholder="Brochure link">
<textarea id="eprem" rows="3" placeholder="Remarks"></textarea>
<button onclick="saveEditProperty()">Save Changes</button>
<button onclick="verifyNow()">Verify Now</button>
</div>
<pre id="epo"></pre>
</div>
<div class="card"><h3>Add Requirement</h3><input id="rc" placeholder="Client"><input id="rci" placeholder="City"><input id="rl" placeholder="Preferred locations"><input id="rmin" type="number" placeholder="Min sqft"><input id="rmax" type="number" placeholder="Max sqft"><select id="rx"><option>Rent</option><option>Sale</option></select><button onclick="req()">Save</button><pre id="ro"></pre></div>
<div class="card"><h3>Run Matcher + WhatsApp Draft</h3><input id="rid" placeholder="Requirement ID"><button onclick="mt()">Match + Create WhatsApp</button><pre id="mo"></pre><div id="waBox" style="display:none;margin-top:10px"><b>WhatsApp Draft</b><textarea id="waText" rows="9" readonly></textarea><small>READY_FOR_REVIEW. Nothing is sent automatically.</small></div></div>
</div></section>
<section id="db" class="hidden"><div class="card"><h3>Database</h3><div id="tabs"></div><p id="meta"></p><div class="tablebox"><table id="grid"></table></div></div></section>
<section id="status" class="hidden"><div class="card"><button onclick="stat()">Refresh</button><pre id="so"></pre></div></section>
<section id="admin" class="hidden"><div class="card"><h3>Admin</h3><div id="atabs"></div><p id="ameta"></p><div class="tablebox"><table id="agrid"></table></div></div></section></div>
<script>
window.addEventListener("error",function(ev){
  const banner=document.getElementById("appErrorBanner");
  if(banner){
    banner.style.display="block";
    banner.textContent="Workspace error: "+(ev.message||"Unknown browser error");
  }
});
const ROLE="__ROLELOW__";const e=i=>document.getElementById(i),v=i=>e(i).value,s=(i,d)=>e(i).textContent=JSON.stringify(d,null,2);
async function jf(u,o={}){
  const r=await fetch(u,o);
  const t=await r.text();
  let d;
  try{ d=JSON.parse(t); }
  catch(err){ d={message:t||("HTTP "+r.status)}; }
  if(!r.ok)throw d;
  return d;
}
function sec(i){
  ["ops","db","status","admin"].forEach(x=>{
    const el=e(x);
    if(el)el.classList.add("hidden");
  });
  const target=e(i);
  if(!target)return;
  target.classList.remove("hidden");
  if(i==="db")load("properties","grid","meta");
  if(i==="status")stat();
}
function setProgress(pct,text){
  e("progressWrap").style.display="block";
  e("progressPct").innerText=pct+"%";
  e("progressBar").style.width=pct+"%";
  if(text)e("progressText").innerText=text;
}

function showUploadMessage(text,ok=true){
  e("uploadResult").innerHTML='<div style="padding:10px;border-radius:8px;background:'+
    (ok?'#ecfdf5;color:#065f46':'#fef2f2;color:#991b1b')+'">'+text+'</div>';
}

function watchJob(jobId){
  let tries=0;
  const timer=setInterval(async()=>{
    tries++;
    try{
      const d=await jf("/api/upload-status/"+jobId);
      const j=d.job||{};
      if(j.status==="COMPLETED"){
        clearInterval(timer);
        setProgress(100,"AI processing completed");
        showUploadMessage("✓ Upload and AI extraction completed successfully. Processed records: "+
          (j.processed_records??0)+(j.duplicate_records?(" | Duplicates: "+j.duplicate_records):""));
      }else if(j.status==="FAILED"){
        clearInterval(timer);
        setProgress(100,"Upload completed, AI processing failed");
        showUploadMessage("Upload succeeded, but AI extraction failed: "+(j.error_message||"Unknown error"),false);
      }else{
        e("progressText").innerText="AI is processing magazine batches... "+(j.processed_records??0)+" records stored so far";
      }
    }catch(err){}
    if(tries>120){
      clearInterval(timer);
      showUploadMessage("Upload completed. AI processing is continuing in the background.");
    }
  },2500);
}

function up(){
  const file=e("fu").files[0];
  if(!file){
    showUploadMessage("Please choose a file first.",false);
    return;
  }

  const fd=new FormData();
  fd.append("file",file);

  const q=new URLSearchParams({
    source_type:v("st"),
    source_name:v("sn")
  });

  e("uploadBtn").disabled=true;
  e("uploadResult").innerHTML="";
  setProgress(0,"Starting upload...");

  const xhr=new XMLHttpRequest();
  xhr.open("POST","/api/ingest/file?"+q.toString(),true);

  xhr.upload.onprogress=function(ev){
    if(ev.lengthComputable){
      const pct=Math.min(99,Math.round((ev.loaded/ev.total)*100));
      setProgress(pct,"Uploading "+file.name);
    }
  };

  xhr.onerror=function(){
    e("uploadBtn").disabled=false;
    showUploadMessage("Upload failed because the network connection was interrupted. Please try again.",false);
  };

  xhr.onload=function(){
    e("uploadBtn").disabled=false;

    let data={};
    try{data=JSON.parse(xhr.responseText||"{}")}catch(err){
      data={detail:xhr.responseText||"Unknown response"};
    }

    if(xhr.status<200 || xhr.status>=300){
      setProgress(0,"Upload failed");
      showUploadMessage(data.detail||data.message||"Upload failed.",false);
      return;
    }

    setProgress(100,"Upload completed");

    if(data.status==="PROCESSED"){
      showUploadMessage("✓ File uploaded and imported successfully. Records added: "+(data.inserted??0));
      return;
    }

    if(data.status==="ACCEPTED"){
      showUploadMessage("✓ File uploaded successfully. AI is now reading and organizing it.");
      if(data.job_id)watchJob(data.job_id);
      return;
    }

    showUploadMessage("✓ Upload completed successfully.");
  };

  xhr.send(fd);
}

async function txt(){try{s("to",await jf("/api/ingest/text",{method:"POST",headers:{"Content-Type":"application/json"},body:JSON.stringify({source_type:"WHATSAPP",source_name:v("tn"),text_content:v("tc")})}))}catch(x){s("to",x)}}

let selectedPropertyFiles=[];

function renderPropertyPreviews(){
  const grid=e("previewGrid");
  grid.innerHTML="";
  selectedPropertyFiles.forEach((file,index)=>{
    const box=document.createElement("div");
    box.style.position="relative";
    box.style.border="1px solid #e5e7eb";
    box.style.borderRadius="8px";
    box.style.overflow="hidden";
    box.style.background="#fff";

    const img=document.createElement("img");
    img.style.width="100%";
    img.style.height="90px";
    img.style.objectFit="cover";
    img.src=URL.createObjectURL(file);

    const btn=document.createElement("button");
    btn.type="button";
    btn.innerText="×";
    btn.style.position="absolute";
    btn.style.top="3px";
    btn.style.right="3px";
    btn.style.width="26px";
    btn.style.height="26px";
    btn.style.padding="0";
    btn.style.borderRadius="50%";
    btn.onclick=()=>{
      selectedPropertyFiles.splice(index,1);
      renderPropertyPreviews();
    };

    box.appendChild(img);
    box.appendChild(btn);
    grid.appendChild(box);
  });
  e("photoCount").innerText=selectedPropertyFiles.length+" photos selected";
}

function addPropertyFiles(fileList){
  const incoming=Array.from(fileList||[]).filter(f=>f.type.startsWith("image/"));
  for(const f of incoming){
    if(selectedPropertyFiles.length>=12) break;
    const duplicate=selectedPropertyFiles.some(x=>x.name===f.name && x.size===f.size);
    if(!duplicate) selectedPropertyFiles.push(f);
  }
  renderPropertyPreviews();
}

window.addEventListener("DOMContentLoaded",()=>{
  const dz=e("dropZone");
  const input=e("pfiles");

  dz.onclick=()=>input.click();
  input.onchange=()=>addPropertyFiles(input.files);

  ["dragenter","dragover"].forEach(evt=>{
    dz.addEventListener(evt,ev=>{
      ev.preventDefault();
      dz.style.background="#eef2ff";
      dz.style.borderColor="#4f46e5";
    });
  });

  ["dragleave","drop"].forEach(evt=>{
    dz.addEventListener(evt,ev=>{
      ev.preventDefault();
      dz.style.background="#f9fafb";
      dz.style.borderColor="#9ca3af";
    });
  });

  dz.addEventListener("drop",ev=>{
    addPropertyFiles(ev.dataTransfer.files);
  });
});

async function uploadPropertyImages(propertyId){
  const uploaded=[];
  for(let i=0;i<selectedPropertyFiles.length;i++){
    const file=selectedPropertyFiles[i];
    e("propertySaveStatus").innerText="Uploading photo "+(i+1)+" of "+selectedPropertyFiles.length+"...";
    const fd=new FormData();
    fd.append("file",file);
    const d=await jf("/api/properties/"+encodeURIComponent(propertyId)+"/media",{
      method:"POST",
      body:fd
    });
    uploaded.push(d);
  }
  return uploaded;
}

async function prop(){
  try{
    e("savePropertyBtn").disabled=true;
    e("propertySaveStatus").innerText="Saving property...";

    const d=await jf("/api/properties",{
      method:"POST",
      headers:{"Content-Type":"application/json"},
      body:JSON.stringify({
        property_name:v("pn"),
        property_type:v("pt")||"NA",
        city:v("pc")||"NA",
        location:v("pl")||"NA",
        available_area_sqft:Number(v("pa"))||null,
        floor:v("pf")||null,
        rent_or_sale:v("px"),
        nearby_brands:v("pnb")||null,
        suitable_category:v("psc")||null,
        parking:v("ppk")||null,
        image_urls:v("pimg")||null,
        video_urls:v("pvid")||null,
        brochure_url:v("pbro")||null,
        remarks:v("prem")||null,
        source:"Manual"
      })
    });

    if(d.status==="duplicate"){
      s("po",d);
      e("propertySaveStatus").innerText="This property appears to already exist. Photos were not uploaded.";
      return;
    }

    let uploaded=[];
    if(d.property_id && selectedPropertyFiles.length){
      uploaded=await uploadPropertyImages(d.property_id);
    }

    s("po",{
      status:"saved",
      property_id:d.property_id,
      photos_uploaded:uploaded.length,
      photo_urls:uploaded.map(x=>x.url)
    });

    e("propertySaveStatus").innerText="✓ Property saved successfully with "+uploaded.length+" photos.";
    selectedPropertyFiles=[];
    renderPropertyPreviews();

  }catch(x){
    s("po",x);
    e("propertySaveStatus").innerText="Property/photo upload failed: "+(x.detail||x.message||"Unknown error");
  }finally{
    e("savePropertyBtn").disabled=false;
  }
}


async function loadEditProperty(){
  try{
    const d=await jf("/api/properties/"+encodeURIComponent(v("epid")));
    const p=d.property||{};
    e("editFields").style.display="block";
    e("epn").value=p.property_name||"";
    e("ept").value=p.property_type||"";
    e("epc").value=p.city||"";
    e("epl").value=p.location||"";
    e("epa").value=p.available_area_sqft||"";
    e("epf").value=p.floor||"";
    e("epx").value=p.rent_or_sale==="Sale"?"Sale":"Rent";
    e("epnb").value=p.nearby_brands||"";
    e("epsc").value=p.suitable_category||"";
    e("eppk").value=p.parking||"";
    e("epimg").value=p.image_urls||"";
    e("epvid").value=p.video_urls||"";
    e("epbro").value=p.brochure_url||"";
    e("eprem").value=p.remarks||"";

    const badge=e("verifyBadge");
    badge.style.display="block";
    badge.innerText=p.last_verified_label||"Never Verified";
    if(p.verification_due){
      badge.style.background="#fef2f2";
      badge.style.color="#991b1b";
      badge.style.border="1px solid #fecaca";
    }else{
      badge.style.background="#ecfdf5";
      badge.style.color="#065f46";
      badge.style.border="1px solid #a7f3d0";
    }
    s("epo",{status:"loaded",property_id:p.property_id});
  }catch(x){s("epo",x)}
}

async function saveEditProperty(){
  try{
    const pid=v("epid");
    const d=await jf("/api/properties/"+encodeURIComponent(pid),{
      method:"PUT",
      headers:{"Content-Type":"application/json","x-user-name":ROLE},
      body:JSON.stringify({
        property_name:v("epn"),property_type:v("ept")||"NA",city:v("epc")||"NA",location:v("epl")||"NA",
        available_area_sqft:Number(v("epa"))||null,floor:v("epf")||null,rent_or_sale:v("epx"),
        nearby_brands:v("epnb")||null,suitable_category:v("epsc")||null,parking:v("eppk")||null,
        image_urls:v("epimg")||null,video_urls:v("epvid")||null,brochure_url:v("epbro")||null,
        remarks:v("eprem")||null,source:"Manual"
      })
    });
    s("epo",d);
    await loadEditProperty();
  }catch(x){s("epo",x)}
}

async function verifyNow(){
  try{
    const d=await jf("/api/properties/"+encodeURIComponent(v("epid"))+"/verify",{
      method:"POST",
      headers:{"x-user-name":ROLE}
    });
    s("epo",d);
    await loadEditProperty();
  }catch(x){s("epo",x)}
}

async function req(){try{const d=await jf("/api/requirements",{method:"POST",headers:{"Content-Type":"application/json"},body:JSON.stringify({client_name:v("rc"),city:v("rci"),preferred_locations:v("rl"),minimum_area_sqft:Number(v("rmin"))||null,maximum_area_sqft:Number(v("rmax"))||null,rent_or_sale:v("rx")})});s("ro",d);if(d.requirement_id)e("rid").value=d.requirement_id}catch(x){s("ro",x)}}
async function mt(){try{const d=await jf("/api/match/"+encodeURIComponent(v("rid")),{method:"POST"});s("mo",{status:d.status,matches:d.matches});if(d.whatsapp_draft){e("waBox").style.display="block";e("waText").value=d.whatsapp_draft.message||""}}catch(x){s("mo",x)}}
async function stat(){try{s("so",await jf("/api/status"))}catch(x){s("so",x)}}
function render(rows,target){const g=e(target);if(!rows.length){g.innerHTML="<tr><td>No records yet</td></tr>";return}const c=Object.keys(rows[0]);g.innerHTML="<thead><tr>"+c.map(x=>"<th>"+x+"</th>").join("")+"</tr></thead><tbody>"+rows.map(r=>"<tr>"+c.map(x=>{let val=String(r[x]??"");if(x==="last_verified"){const due=!!r.verification_due;return "<td><span style=\"padding:5px 8px;border-radius:999px;font-weight:700;background:"+(due?"#fef2f2;color:#991b1b":"#ecfdf5;color:#065f46")+"\">"+val+"</span></td>"}return "<td>"+val+"</td>"}).join("")+"</tr>").join("")+"</tbody>"}
async function load(n,target="grid",meta="meta"){try{const d=await jf("/api/database/"+n);e(meta).innerText=d.count+" records";render(d.rows,target)}catch(x){e(meta).innerText=x.detail||x.message||"Error"}}
e("tabs").innerHTML=["properties","requirements","matches","whatsapp_drafts"]
  .map(name=>`<button type="button" data-table="${name}" class="dbTab">${name}</button>`)
  .join("");

document.querySelectorAll(".dbTab").forEach(btn=>{
  btn.addEventListener("click",()=>load(btn.dataset.table,"grid","meta"));
});
if(ROLE=="admin"){
  e("atabs").innerHTML=["sources","ai_jobs","verification","batches","media"]
    .map(name=>`<button type="button" data-admin-table="${name}" class="adminDbTab">${name}</button>`)
    .join("");

  document.querySelectorAll(".adminDbTab").forEach(btn=>{
    btn.addEventListener("click",()=>load(btn.dataset.adminTable,"agrid","ameta"));
  });
}
</script></body></html>'''


def ui_shell(role,title,body):
    admin_link='<a class="navbtn" href="/admin-page">Admin</a>' if role=="admin" else ""
    return f"""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{escape(title)}</title>
<style>
*{{box-sizing:border-box}}
body{{font-family:Arial,sans-serif;margin:0;background:#f5f7fb;color:#172033}}
header{{background:#111827;color:#fff;padding:18px 22px;display:flex;justify-content:space-between;gap:15px;align-items:center}}
nav{{background:#fff;border-bottom:1px solid #ddd;padding:10px 16px;display:flex;gap:8px;flex-wrap:wrap}}
.navbtn,.btn{{display:inline-block;padding:10px 14px;border-radius:8px;background:#111827;color:#fff;text-decoration:none;border:0;cursor:pointer;font-size:14px}}
.navbtn.secondary{{background:#374151}}
.wrap{{padding:18px;max-width:1500px;margin:auto}}
.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(320px,1fr));gap:15px}}
.card{{background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:16px;margin-bottom:15px}}
input,select,textarea{{width:100%;padding:10px;border:1px solid #d1d5db;border-radius:8px;margin:5px 0 10px;font-size:16px}}
table{{border-collapse:collapse;width:100%;font-size:12px}}
th,td{{border-bottom:1px solid #eee;padding:8px;text-align:left;white-space:nowrap}}
.tablebox{{overflow:auto;max-height:70vh}}
.good{{background:#ecfdf5;color:#065f46;padding:8px;border-radius:8px}}
.warn{{background:#fef2f2;color:#991b1b;padding:8px;border-radius:8px}}
.info{{background:#eff6ff;color:#1e40af;padding:8px;border-radius:8px}}
small{{color:#6b7280}}
</style>
</head>
<body>
<header>
<div><b>Property Intelligence</b><br><small style="color:#d1d5db">Reliable Team + Admin Workspace</small></div>
<div>{escape(role.upper())} · <a href="/logout" style="color:white">Logout</a></div>
</header>
<nav>
<a class="navbtn" href="/workspace">Operations</a>
<a class="navbtn" href="/property-discovery">Find Property</a>
<a class="navbtn" href="/database-page">Database</a>
<a class="navbtn" href="/status-page">Status</a>
{admin_link}
</nav>
<div class="wrap">
<h2>{escape(title)}</h2>
{body}
</div>
</body>
</html>"""

def ui_message(value,kind="info"):
    return f'<div class="{kind}">{escape(str(value))}</div>'

@app.get("/database-page",response_class=HTMLResponse)
def database_page(req:Request,table_name:str=Query("properties"),q:str=Query("")):
    role=need_login(req)
    allowed=["properties","requirements","matches","whatsapp_drafts"]
    if table_name not in allowed:
        table_name="properties"

    table=TABLES[table_name]
    search_term=(q or "").strip()

    with engine.connect() as c:
        if table_name=="properties" and search_term:
            like="%"+search_term+"%"
            result=c.execute(
                text("""SELECT * FROM pi_properties
                        WHERE
                            property_id ILIKE :q OR
                            COALESCE(property_name,'') ILIKE :q OR
                            COALESCE(property_type,'') ILIKE :q OR
                            COALESCE(city,'') ILIKE :q OR
                            COALESCE(location,'') ILIKE :q OR
                            COALESCE(floor,'') ILIKE :q OR
                            COALESCE(rent_or_sale,'') ILIKE :q OR
                            COALESCE(nearby_brands,'') ILIKE :q OR
                            COALESCE(suitable_category,'') ILIKE :q OR
                            COALESCE(owner_name,'') ILIKE :q OR
                            COALESCE(owner_contact,'') ILIKE :q OR
                            COALESCE(broker_name,'') ILIKE :q OR
                            COALESCE(broker_contact,'') ILIKE :q OR
                            COALESCE(remarks,'') ILIKE :q
                        ORDER BY id DESC
                        LIMIT 500"""),
                {"q":like}
            )
        elif table_name=="requirements" and search_term:
            like="%"+search_term+"%"
            result=c.execute(
                text("""SELECT * FROM pi_requirements
                        WHERE
                            requirement_id ILIKE :q OR
                            COALESCE(client_name,'') ILIKE :q OR
                            COALESCE(company_name,'') ILIKE :q OR
                            COALESCE(contact_phone,'') ILIKE :q OR
                            COALESCE(contact_email,'') ILIKE :q OR
                            COALESCE(city,'') ILIKE :q OR
                            COALESCE(preferred_locations,'') ILIKE :q OR
                            COALESCE(rent_or_sale,'') ILIKE :q OR
                            COALESCE(status,'') ILIKE :q
                        ORDER BY id DESC
                        LIMIT 500"""),
                {"q":like}
            )
        else:
            result=c.execute(text("SELECT * FROM "+table+" ORDER BY id DESC LIMIT 500"))

        rows=[]
        for row in result:
            d={}
            for k,v in dict(row._mapping).items():
                if isinstance(v,(date,datetime)):
                    d[k]=v.isoformat()
                elif isinstance(v,Decimal):
                    d[k]=float(v)
                else:
                    d[k]=v
            rows.append(d)

    if table_name=="properties":
        for row in rows:
            verified=row.get("verified_date")
            if not verified:
                row["last_verified"]="Never Verified"
                row["verification_due"]=True
            else:
                vd=date.fromisoformat(verified) if isinstance(verified,str) else verified
                age=(date.today()-vd).days
                row["last_verified"]=f"{verified} ({age} days ago)"
                row["verification_due"]=age>=VERIFICATION_DUE_DAYS

        if role!="admin":
            team_private=PRIVATE-{"verified_date","verified_by"}
            rows=[{k:v for k,v in row.items() if k not in team_private} for row in rows]

    if table_name=="properties":
        for row in rows:
            row_id=row.get("id")
            row["action"]=f'<a class="btn" href="/property/edit-row/{int(row_id)}">Edit</a>' if row_id is not None else ""

    search_form=f"""
    <div class="card">
      <form method="get" action="/database-page">
        <input type="hidden" name="table_name" value="{escape(table_name,quote=True)}">
        <label><b>Search Database</b></label>
        <input name="q" value="{escape(search_term,quote=True)}"
               placeholder="Property ID, name, location, city, broker, phone, nearby brand...">
        <button class="btn" type="submit">Search</button>
        <a class="navbtn secondary" href="/database-page?table_name={escape(table_name)}">Clear</a>
      </form>
      <small>Search is not case-sensitive. Enter any part of a Property ID, name, phone number, location or broker.</small>
    </div>
    """

    tabs=" ".join(
        f'<a class="navbtn secondary" href="/database-page?table_name={escape(name)}">{escape(name.replace("_"," ").title())}</a>'
        for name in allowed
    )

    if not rows:
        table_html="<p>No records yet.</p>"
    else:
        cols=list(rows[0].keys())
        header="".join(f"<th>{escape(str(col))}</th>" for col in cols)
        rendered=[]
        for row in rows:
            cells=[]
            for col in cols:
                val=row.get(col,"")
                if col=="last_verified":
                    cls="warn" if row.get("verification_due") else "good"
                    cells.append(f'<td><span class="{cls}">{escape(str(val))}</span></td>')
                elif col=="action":
                    cells.append(f"<td>{val}</td>")
                else:
                    cells.append(f"<td>{escape(str(val if val is not None else ''))}</td>")
            rendered.append("<tr>"+"".join(cells)+"</tr>")
        table_html=f'<div class="tablebox"><table><thead><tr>{header}</tr></thead><tbody>{"".join(rendered)}</tbody></table></div>'

    result_label=(
        f'<p><b>{len(rows)} results</b> for “{escape(search_term)}”</p>'
        if search_term else
        f'<p><b>{len(rows)} records shown</b></p>'
    )
    return HTMLResponse(ui_shell(role,"Database",search_form+tabs+result_label+table_html))


@app.get("/property/edit-row/{row_id}",response_class=HTMLResponse)
def edit_property_by_row(req:Request,row_id:int):
    role=need_login(req)
    with engine.connect() as c:
        row=c.execute(text("SELECT property_id FROM pi_properties WHERE id=:id"),{"id":row_id}).first()
    if not row:
        return HTMLResponse(ui_shell(role,"Edit Property",ui_message("Property database row not found.","warn")),status_code=404)
    property_id=row._mapping["property_id"]
    if not property_id:
        with engine.begin() as c:
            property_id=make_id("PROP",c)
            c.execute(text("UPDATE pi_properties SET property_id=:pid,updated_at=NOW() WHERE id=:id"),{"pid":property_id,"id":row_id})
    return RedirectResponse("/property/edit/"+quote_plus(str(property_id)),status_code=303)

@app.get("/property/edit/{property_id}",response_class=HTMLResponse)
def edit_property_page(req:Request,property_id:str,message:str=Query("")):
    role=need_login(req)
    with engine.connect() as c:
        row=c.execute(
            text("SELECT * FROM pi_properties WHERE property_id=:pid"),
            {"pid":property_id}
        ).first()

    if not row:
        return HTMLResponse(
            ui_shell(role,"Edit Property",ui_message("Property not found.","warn")),
            status_code=404
        )

    d=dict(row._mapping)

    def fv(name):
        value=d.get(name)
        return escape(str(value if value is not None else ""), quote=True)

    notice=ui_message(message,"good") if message else ""
    body=notice+f"""
    <div class="card">
      <div class="info">
        <b>Property ID:</b> {escape(property_id)}
        <br><small>The Property ID never changes when details are edited.</small>
      </div>

      <form action="/property/edit/{quote_plus(property_id)}" method="post">
        <label>Property Name</label>
        <input name="property_name" value="{fv('property_name')}">

        <label>Property Type</label>
        <input name="property_type" value="{fv('property_type')}">

        <label>City</label>
        <input name="city" value="{fv('city')}">

        <label>Location</label>
        <input name="location" value="{fv('location')}">

        <label>Available Area (sqft)</label>
        <input name="available_area_sqft" type="number" step="any" value="{fv('available_area_sqft')}">

        <label>Minimum Area (sqft)</label>
        <input name="minimum_area_sqft" type="number" step="any" value="{fv('minimum_area_sqft')}">

        <label>Maximum Area (sqft)</label>
        <input name="maximum_area_sqft" type="number" step="any" value="{fv('maximum_area_sqft')}">

        <label>Floor</label>
        <input name="floor" value="{fv('floor')}">

        <label>Rent / Sale</label>
        <input name="rent_or_sale" value="{fv('rent_or_sale')}">

        <label>Nearby Brands</label>
        <textarea name="nearby_brands" rows="2">{escape(str(d.get('nearby_brands') or ''))}</textarea>

        <label>Suitable Category</label>
        <input name="suitable_category" value="{fv('suitable_category')}">

        <label>Parking</label>
        <input name="parking" value="{fv('parking')}">

        <label>Image URLs</label>
        <textarea name="image_urls" rows="3" placeholder="One or more image links">{escape(str(d.get('image_urls') or ''))}</textarea>

        <label>Video URLs</label>
        <textarea name="video_urls" rows="3" placeholder="One or more video links">{escape(str(d.get('video_urls') or ''))}</textarea>

        <label>Brochure URL</label>
        <input name="brochure_url" value="{fv('brochure_url')}">

        <label>Remarks</label>
        <textarea name="remarks" rows="4">{escape(str(d.get('remarks') or ''))}</textarea>

        <button class="btn" type="submit">Save Changes</button>
        <a class="navbtn secondary" href="/database-page?table_name=properties">Back to Database</a>
      </form>
    </div>

    <div class="card">
      <h3>Verification</h3>
      <p><b>Last Verified:</b> {escape(str(d.get('verified_date') or 'Never Verified'))}</p>
      <p><b>Verified By:</b> {escape(str(d.get('verified_by') or ''))}</p>
      <form action="/ui/verify" method="post">
        <input type="hidden" name="property_id" value="{escape(property_id, quote=True)}">
        <button class="btn" type="submit">Verify Today</button>
      </form>
      <small>Saving an edit does not change Last Verified. Use Verify Today only after actual confirmation.</small>
    </div>
    """
    return HTMLResponse(ui_shell(role,"Edit Property",body))


@app.post("/property/edit/{property_id}")
def save_property_edit(
    req:Request,
    property_id:str,
    property_name:str=Form(""),
    property_type:str=Form(""),
    city:str=Form(""),
    location:str=Form(""),
    available_area_sqft:Optional[float]=Form(None),
    minimum_area_sqft:Optional[float]=Form(None),
    maximum_area_sqft:Optional[float]=Form(None),
    floor:str=Form(""),
    rent_or_sale:str=Form(""),
    nearby_brands:str=Form(""),
    suitable_category:str=Form(""),
    parking:str=Form(""),
    image_urls:str=Form(""),
    video_urls:str=Form(""),
    brochure_url:str=Form(""),
    remarks:str=Form("")
):
    role=need_login(req)
    actor=actor_name(req)

    editable={
        "property_name":property_name or None,
        "property_type":property_type or None,
        "city":city or None,
        "location":location or None,
        "available_area_sqft":available_area_sqft,
        "minimum_area_sqft":minimum_area_sqft,
        "maximum_area_sqft":maximum_area_sqft,
        "floor":floor or None,
        "rent_or_sale":rent_or_sale or None,
        "nearby_brands":nearby_brands or None,
        "suitable_category":suitable_category or None,
        "parking":parking or None,
        "image_urls":image_urls or None,
        "video_urls":video_urls or None,
        "brochure_url":brochure_url or None,
        "remarks":remarks or None
    }

    with engine.begin() as c:
        existing=c.execute(
            text("SELECT * FROM pi_properties WHERE property_id=:pid"),
            {"pid":property_id}
        ).first()

        if not existing:
            return HTMLResponse(
                ui_shell(role,"Edit Property",ui_message("Property not found.","warn")),
                status_code=404
            )

        old=dict(existing._mapping)
        changes={}
        for key,new_value in editable.items():
            old_value=old.get(key)
            if isinstance(old_value,Decimal):
                old_value=float(old_value)
            if old_value != new_value:
                changes[key]={"old":old_value,"new":new_value}

        c.execute(
            text("""UPDATE pi_properties SET
                property_name=:property_name,
                property_type=:property_type,
                city=:city,
                location=:location,
                available_area_sqft=:available_area_sqft,
                minimum_area_sqft=:minimum_area_sqft,
                maximum_area_sqft=:maximum_area_sqft,
                floor=:floor,
                rent_or_sale=:rent_or_sale,
                nearby_brands=:nearby_brands,
                suitable_category=:suitable_category,
                parking=:parking,
                image_urls=:image_urls,
                video_urls=:video_urls,
                brochure_url=:brochure_url,
                remarks=:remarks,
                updated_at=NOW()
                WHERE property_id=:property_id"""),
            {**editable,"property_id":property_id}
        )

        if changes:
            c.execute(
                text("""INSERT INTO pi_verification_log
                        (property_id,action,performed_by,notes,old_values,new_values)
                        VALUES(:pid,'EDITED',:actor,:notes,CAST(:oldv AS JSONB),CAST(:newv AS JSONB))"""),
                {
                    "pid":property_id,
                    "actor":actor,
                    "notes":"Property edited directly from database",
                    "oldv":json.dumps({k:v["old"] for k,v in changes.items()},default=str),
                    "newv":json.dumps({k:v["new"] for k,v in changes.items()},default=str)
                }
            )

    result_message="Changes saved successfully." if changes else "No changes were made."
    return RedirectResponse(
        "/property/edit/"+quote_plus(property_id)+"?message="+quote_plus(result_message),
        status_code=303
    )

@app.get("/status-page",response_class=HTMLResponse)
def status_page(req:Request):
    role=need_login(req)
    with engine.connect() as c:
        counts={}
        for label,table in {
            "Properties":"pi_properties",
            "Requirements":"pi_requirements",
            "Sources":"pi_sources",
            "Matches":"pi_matches",
            "AI Jobs":"pi_ai_jobs"
        }.items():
            counts[label]=c.execute(text(f"SELECT COUNT(*) FROM {table}")).scalar_one()

    cards="".join(
        f'<div class="card"><h3>{escape(label)}</h3><div style="font-size:28px;font-weight:700">{count}</div></div>'
        for label,count in counts.items()
    )
    body=f'<div class="grid">{cards}</div><div class="card"><b>Version:</b> {VERSION}<br><b>Gemini configured:</b> {bool(GEMINI_API_KEY)}<br><b>Verification due:</b> {VERIFICATION_DUE_DAYS} days</div>'
    return HTMLResponse(ui_shell(role,"System Status",body))

@app.get("/admin-page",response_class=HTMLResponse)
def admin_page(req:Request,table_name:str=Query("sources")):
    role=need_login(req)
    if role!="admin":
        return HTMLResponse(ui_shell(role,"Admin",ui_message("Admin access required.","warn")),status_code=403)

    allowed=["sources","ai_jobs","verification","batches","media","scan_tiles"]
    if table_name not in allowed:
        table_name="sources"

    with engine.connect() as c:
        if table_name=="media":
            result=c.execute(text("""SELECT id,media_id,property_id,media_type,filename,mime_type,file_size,created_at
                                     FROM pi_property_media ORDER BY id DESC LIMIT 500"""))
        else:
            result=c.execute(text("SELECT * FROM "+TABLES[table_name]+" ORDER BY id DESC LIMIT 500"))
        rows=[]
        for row in result:
            d={}
            for k,v in dict(row._mapping).items():
                if isinstance(v,(date,datetime)):
                    d[k]=v.isoformat()
                elif isinstance(v,Decimal):
                    d[k]=float(v)
                else:
                    d[k]=v
            rows.append(d)

    tabs=" ".join(
        f'<a class="navbtn secondary" href="/admin-page?table_name={escape(name)}">{escape(name.replace("_"," ").title())}</a>'
        for name in allowed
    )

    if not rows:
        table_html="<p>No records yet.</p>"
    else:
        cols=list(rows[0].keys())
        header="".join(f"<th>{escape(str(col))}</th>" for col in cols)
        body="".join(
            "<tr>"+"".join(f"<td>{escape(str(row.get(col,'') if row.get(col,'') is not None else ''))}</td>" for col in cols)+"</tr>"
            for row in rows
        )
        table_html=f'<div class="tablebox"><table><thead><tr>{header}</tr></thead><tbody>{body}</tbody></table></div>'

    return HTMLResponse(ui_shell(role,"Admin",tabs+f"<p>{len(rows)} records</p>"+table_html))

@app.post("/ui/upload")
async def ui_upload(
    req:Request,
    bg:BackgroundTasks,
    source_type:str=Form("DOCUMENT"),
    source_name:str=Form(""),
    file:UploadFile=File(...)
):
    need_login(req)
    filename=file.filename or "upload.bin"
    ext=os.path.splitext(filename)[1].lower()
    mime=file.content_type or "application/octet-stream"

    mime_map={
        ".jpg":"image/jpeg",".jpeg":"image/jpeg",".png":"image/png",".webp":"image/webp",
        ".pdf":"application/pdf",".csv":"text/csv",".txt":"text/plain"
    }
    if mime in {"application/octet-stream",""} or not mime:
        mime=mime_map.get(ext,"application/octet-stream")

    fd,path=tempfile.mkstemp(suffix=ext or ".bin")
    os.close(fd)
    total=0

    try:
        with open(path,"wb") as output:
            while True:
                chunk=await file.read(1024*1024)
                if not chunk:
                    break
                total+=len(chunk)
                if total>MAX_UPLOAD_MB*1024*1024:
                    raise HTTPException(413,f"File too large. Maximum {MAX_UPLOAD_MB} MB.")
                if (mime=="application/pdf" or ext==".pdf") and total>50*1024*1024:
                    raise HTTPException(413,"PDF too large. Maximum PDF size is 50 MB.")
                output.write(chunk)

        sid=source_row(source_type.upper(),source_name or filename,filename,mime)

        if ext==".csv":
            inserted=0
            with open(path,"r",encoding="utf-8-sig",errors="replace",newline="") as csvfile:
                reader=csv.DictReader(csvfile)
                for row in reader:
                    item={
                        "property_name":row.get("Property name") or row.get("Property Name"),
                        "property_type":row.get("Property type") or row.get("Property Type") or "NA",
                        "city":row.get("City") or "NA",
                        "location":row.get("Location") or "NA",
                        "available_area_sqft":row.get("Available area") or row.get("Available Area") or None,
                        "floor":row.get("Floor"),
                        "rent_or_sale":row.get("Rent/Sale"),
                        "nearby_brands":row.get("Nearby brand") or row.get("Nearby brands"),
                        "suitable_category":row.get("Suitable category"),
                        "parking":row.get("Parking"),
                        "source":"CSV:"+filename
                    }
                    try:
                        if item["available_area_sqft"]:
                            item["available_area_sqft"]=float(str(item["available_area_sqft"]).replace(",",""))
                    except Exception:
                        item["available_area_sqft"]=None

                    if save_property(item,sid)["status"]=="created":
                        inserted+=1

            with engine.begin() as c:
                c.execute(
                    text("UPDATE pi_sources SET ingestion_status='PROCESSED',processed_records=:n,processed_at=NOW() WHERE id=:id"),
                    {"n":inserted,"id":sid}
                )
            try: os.unlink(path)
            except Exception: pass
            message=f"CSV uploaded successfully. {inserted} properties inserted."
        else:
            jid=create_job(sid,"FILE_EXTRACTION",filename)
            bg.add_task(run_file_job,sid,jid,path,mime)
            message=f"Upload accepted. AI processing started in background. Job ID: {jid}"

        return RedirectResponse("/workspace?message="+quote_plus(message),status_code=303)

    except Exception:
        try: os.unlink(path)
        except Exception: pass
        raise

@app.post("/ui/add-property")
def ui_add_property(
    req:Request,
    property_name:str=Form(""),
    property_type:str=Form("NA"),
    city:str=Form("NA"),
    location:str=Form("NA"),
    available_area_sqft:Optional[float]=Form(None),
    floor:str=Form(""),
    rent_or_sale:str=Form("Rent"),
    nearby_brands:str=Form(""),
    suitable_category:str=Form(""),
    parking:str=Form(""),
    image_urls:str=Form(""),
    video_urls:str=Form(""),
    brochure_url:str=Form(""),
    remarks:str=Form("")
):
    need_login(req)
    result=save_property({
        "property_name":property_name or None,
        "property_type":property_type or "NA",
        "city":city or "NA",
        "location":location or "NA",
        "available_area_sqft":available_area_sqft,
        "floor":floor or None,
        "rent_or_sale":rent_or_sale or None,
        "nearby_brands":nearby_brands or None,
        "suitable_category":suitable_category or None,
        "parking":parking or None,
        "image_urls":image_urls or None,
        "video_urls":video_urls or None,
        "brochure_url":brochure_url or None,
        "remarks":remarks or None,
        "source":"Manual"
    })
    message=f"Property {result.get('status')}: {result.get('property_id','')}"
    return RedirectResponse("/workspace?message="+quote_plus(message),status_code=303)

@app.post("/ui/add-requirement")
def ui_add_requirement(
    req:Request,
    client_name:str=Form(""),
    company_name:str=Form(""),
    contact_phone:str=Form(""),
    city:str=Form(""),
    preferred_locations:str=Form(""),
    minimum_area_sqft:Optional[float]=Form(None),
    maximum_area_sqft:Optional[float]=Form(None),
    rent_or_sale:str=Form("Rent")
):
    need_login(req)
    result=save_requirement({
        "client_name":client_name or None,
        "company_name":company_name or None,
        "contact_phone":contact_phone or None,
        "city":city or None,
        "preferred_locations":preferred_locations or None,
        "minimum_area_sqft":minimum_area_sqft,
        "maximum_area_sqft":maximum_area_sqft,
        "rent_or_sale":rent_or_sale or None,
        "source":"Manual"
    })
    message=f"Requirement {result.get('status')}: {result.get('requirement_id','')}"
    return RedirectResponse("/workspace?message="+quote_plus(message),status_code=303)

@app.post("/ui/verify")
def ui_verify(req:Request,property_id:str=Form(...)):
    need_login(req)
    actor=actor_name(req)

    with engine.begin() as c:
        exists=c.execute(text("SELECT 1 FROM pi_properties WHERE property_id=:pid"),{"pid":property_id}).first()
        if not exists:
            return RedirectResponse("/workspace?message="+quote_plus("Property not found"),status_code=303)

        c.execute(
            text("""UPDATE pi_properties
                    SET verified_date=CURRENT_DATE,
                        verified_by=:actor,
                        verification_status='VERIFIED',
                        updated_at=NOW()
                    WHERE property_id=:pid"""),
            {"actor":actor,"pid":property_id}
        )
        c.execute(
            text("""INSERT INTO pi_verification_log(property_id,action,performed_by,notes)
                    VALUES(:pid,'VERIFIED',:actor,'Property verified from server UI')"""),
            {"pid":property_id,"actor":actor}
        )

    return RedirectResponse("/workspace?message="+quote_plus(f"{property_id} verified today"),status_code=303)

@app.post("/ui/match")
def ui_match(req:Request,requirement_id:str=Form(...)):
    need_login(req)
    try:
        result=robust_match_requirement(requirement_id, create_whatsapp=True)
        message=f"Matcher complete. {len(result.get('matches',[]))} ranked matches. " + result.get("diagnostic",{}).get("message","")
    except HTTPException as exc:
        message=f"Matcher error: {exc.detail}"
    except Exception as exc:
        message=f"Matcher error: {exc}"
    return RedirectResponse("/workspace?message="+quote_plus(message),status_code=303)


@app.get("/legacy-workspace",response_class=HTMLResponse)
def legacy_workspace(req:Request,message:str=Query("")):
    role=page_role_or_redirect(req)
    if not role:
        return RedirectResponse("/login",status_code=303)

    notice=ui_message(message,"good") if message else ""

    upload_card="""
    <div class="card">
      <h3>Upload Photo / Magazine / PDF / CSV</h3>
      <form action="/ui/upload" method="post" enctype="multipart/form-data">
        <input name="source_name" placeholder="Source name">
        <select name="source_type">
          <option>MAGAZINE</option>
          <option>NEWSPAPER</option>
          <option>PHOTO</option>
          <option>PDF</option>
          <option>CSV</option>
        </select>
        <input name="file" type="file" required>
        <button class="btn" type="submit">Upload</button>
      </form>
      <small>High-recall scanner runs full-page + overlapping tiles in the background for photos and PDFs.</small>
    </div>"""

    property_card="""
    <div class="card">
      <h3>Add Property Manually</h3>
      <p>Use the new master property form for Owner, Broker, contact numbers, rent in figures,
      assigned team member, verification status and direct image/video upload.</p>
      <a class="btn" href="/property-manual">Open Add Property Form</a>
    </div>"""

    requirement_card="""
    <div class="card">
      <h3>Add Requirement</h3>
      <form action="/ui/add-requirement" method="post">
        <input name="client_name" placeholder="Client name">
        <input name="company_name" placeholder="Company">
        <input name="contact_phone" placeholder="Contact phone">
        <input name="city" placeholder="City">
        <input name="preferred_locations" placeholder="Preferred locations">
        <input name="minimum_area_sqft" type="number" step="any" placeholder="Min sqft">
        <input name="maximum_area_sqft" type="number" step="any" placeholder="Max sqft">
        <select name="rent_or_sale"><option>Rent</option><option>Sale</option></select>
        <button class="btn" type="submit">Save Requirement</button>
      </form>
    </div>"""

    verify_card="""
    <div class="card">
      <h3>Verify Existing Property</h3>
      <form action="/ui/verify" method="post">
        <input name="property_id" placeholder="Property ID" required>
        <button class="btn" type="submit">Verify Today</button>
      </form>
      <small>Verification date and verifier are recorded.</small>
    </div>"""

    match_card="""
    <div class="card">
      <h3>Run Matcher + WhatsApp Draft</h3>
      <form action="/ui/match" method="post">
        <input name="requirement_id" placeholder="Requirement ID" required>
        <button class="btn" type="submit">Match + Create WhatsApp</button>
      </form>
      <small>Draft remains READY_FOR_REVIEW. Nothing is sent automatically.</small>
    </div>"""

    quick_search_card="""
    <div class="card">
      <h3>Find a Property</h3>
      <form action="/database-page" method="get">
        <input type="hidden" name="table_name" value="properties">
        <input name="q" placeholder="Property ID / Location / Name / Phone / Broker">
        <button class="btn" type="submit">Search Property</button>
      </form>
    </div>"""

    body=notice+'<div class="grid">'+quick_search_card+upload_card+property_card+requirement_card+verify_card+match_card+'</div>'
    return HTMLResponse(ui_shell(role,"Operations",body))

@app.get("/")
def root(req:Request):
    return RedirectResponse("/workspace" if get_role(req) else "/login")

# ============================================================================
# AI DEAL INTELLIGENCE OS V4
# Organized data + Property + Hospitality + Retail + Demand Discovery
# ============================================================================

import re as _re
import math as _math
import socket as _socket
import ipaddress as _ipaddress
from urllib.parse import urlparse as _urlparse
try:
    import httpx as _httpx
except Exception:
    _httpx=None

DEAL_OS_VERSION="4.0.0"
SERPER_API_KEY=os.getenv("SERPER_API_KEY","").strip()
GOOGLE_PLACES_API_KEY=os.getenv("GOOGLE_PLACES_API_KEY","").strip()
APOLLO_API_KEY=os.getenv("APOLLO_API_KEY","").strip()
FLOWCONNECT_WEBHOOK_URL=os.getenv("FLOWCONNECT_WEBHOOK_URL","").strip()
FLOWCONNECT_API_URL=os.getenv("FLOWCONNECT_API_URL","").strip()
FLOWCONNECT_API_KEY=os.getenv("FLOWCONNECT_API_KEY","").strip()
MAX_VIDEO_MB=int(os.getenv("MAX_VIDEO_MB","80"))

V4_SCHEMA = """
CREATE TABLE IF NOT EXISTS pi_owners(
 id BIGSERIAL PRIMARY KEY,
 owner_id VARCHAR(50) UNIQUE NOT NULL,
 owner_name TEXT NOT NULL,
 contact_number TEXT,
 email TEXT,
 city TEXT,
 notes TEXT,
 created_at TIMESTAMPTZ DEFAULT NOW(),
 updated_at TIMESTAMPTZ DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS pi_brokers(
 id BIGSERIAL PRIMARY KEY,
 broker_id VARCHAR(50) UNIQUE NOT NULL,
 broker_name TEXT NOT NULL,
 contact_number TEXT,
 email TEXT,
 company_name TEXT,
 city TEXT,
 notes TEXT,
 created_at TIMESTAMPTZ DEFAULT NOW(),
 updated_at TIMESTAMPTZ DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS ai_companies(
 id BIGSERIAL PRIMARY KEY,
 company_id VARCHAR(50) UNIQUE NOT NULL,
 division VARCHAR(30) NOT NULL,
 company_name TEXT NOT NULL,
 category TEXT,
 primary_contact_name TEXT,
 primary_contact_phone TEXT,
 primary_contact_email TEXT,
 website TEXT,
 linkedin_url TEXT,
 city TEXT,
 target_markets TEXT,
 expansion_status VARCHAR(50) DEFAULT 'DISCOVERED',
 expansion_score NUMERIC(5,2) DEFAULT 0,
 source_name TEXT,
 source_url TEXT,
 source_excerpt TEXT,
 assigned_to TEXT,
 crm_status VARCHAR(50) DEFAULT 'NOT_SYNCED',
 created_at TIMESTAMPTZ DEFAULT NOW(),
 updated_at TIMESTAMPTZ DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS ai_contacts(
 id BIGSERIAL PRIMARY KEY,
 contact_id VARCHAR(50) UNIQUE NOT NULL,
 company_id VARCHAR(50),
 full_name TEXT,
 designation TEXT,
 business_email TEXT,
 business_phone TEXT,
 linkedin_url TEXT,
 verification_status VARCHAR(40) DEFAULT 'UNVERIFIED',
 source_url TEXT,
 created_at TIMESTAMPTZ DEFAULT NOW(),
 updated_at TIMESTAMPTZ DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS ai_expansion_signals(
 id BIGSERIAL PRIMARY KEY,
 signal_id VARCHAR(50) UNIQUE NOT NULL,
 company_id VARCHAR(50),
 division VARCHAR(30) DEFAULT 'RETAIL',
 title TEXT,
 signal_type VARCHAR(80),
 market TEXT,
 source_name TEXT,
 source_url TEXT,
 published_at TIMESTAMPTZ,
 excerpt TEXT,
 signal_score NUMERIC(5,2) DEFAULT 0,
 status VARCHAR(40) DEFAULT 'NEW',
 created_at TIMESTAMPTZ DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS ai_marketing_contacts(
 id BIGSERIAL PRIMARY KEY,
 contact_id VARCHAR(50) UNIQUE NOT NULL,
 fingerprint VARCHAR(64) UNIQUE,
 business_type TEXT,
 brand_name TEXT,
 contact_name TEXT,
 phone TEXT,
 email TEXT,
 website TEXT,
 location TEXT,
 city TEXT,
 source_name TEXT,
 source_url TEXT,
 consent_status VARCHAR(40) DEFAULT 'UNKNOWN',
 verification_status VARCHAR(40) DEFAULT 'UNVERIFIED',
 assigned_to TEXT,
 created_at TIMESTAMPTZ DEFAULT NOW(),
 updated_at TIMESTAMPTZ DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS ai_requirement_campaigns(
 id BIGSERIAL PRIMARY KEY,
 campaign_id VARCHAR(50) UNIQUE NOT NULL,
 property_id VARCHAR(50),
 campaign_name TEXT,
 property_type TEXT,
 city TEXT,
 location TEXT,
 area_sqft NUMERIC(14,2),
 monthly_rent NUMERIC(14,2),
 rent_or_sale TEXT,
 suitable_category TEXT,
 nearby_brands TEXT,
 additional_points TEXT,
 post_draft TEXT,
 assigned_to TEXT,
 status VARCHAR(40) DEFAULT 'DRAFT',
 created_at TIMESTAMPTZ DEFAULT NOW(),
 updated_at TIMESTAMPTZ DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS ai_demand_signals(
 id BIGSERIAL PRIMARY KEY,
 signal_id VARCHAR(50) UNIQUE NOT NULL,
 campaign_id VARCHAR(50),
 source_type VARCHAR(60),
 source_name TEXT,
 source_url TEXT,
 title TEXT,
 excerpt TEXT,
 contact_name TEXT,
 contact_phone TEXT,
 contact_email TEXT,
 company_name TEXT,
 location TEXT,
 intent_score NUMERIC(5,2) DEFAULT 0,
 status VARCHAR(40) DEFAULT 'NEW',
 created_at TIMESTAMPTZ DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS ai_followups(
 id BIGSERIAL PRIMARY KEY,
 followup_id VARCHAR(50) UNIQUE NOT NULL,
 entity_type VARCHAR(40),
 entity_id VARCHAR(50),
 division VARCHAR(30),
 channel VARCHAR(30),
 due_at TIMESTAMPTZ,
 status VARCHAR(40) DEFAULT 'DUE',
 assigned_to TEXT,
 notes TEXT,
 created_at TIMESTAMPTZ DEFAULT NOW(),
 completed_at TIMESTAMPTZ
);

CREATE TABLE IF NOT EXISTS ai_meetings(
 id BIGSERIAL PRIMARY KEY,
 meeting_id VARCHAR(50) UNIQUE NOT NULL,
 entity_type VARCHAR(40),
 entity_id VARCHAR(50),
 division VARCHAR(30),
 meeting_at TIMESTAMPTZ,
 status VARCHAR(40) DEFAULT 'SCHEDULED',
 owner TEXT,
 notes TEXT,
 created_at TIMESTAMPTZ DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS ai_bot_runs(
 id BIGSERIAL PRIMARY KEY,
 run_id VARCHAR(50) UNIQUE NOT NULL,
 bot_name TEXT NOT NULL,
 division VARCHAR(30),
 status VARCHAR(40) DEFAULT 'RUNNING',
 summary TEXT,
 records_found INTEGER DEFAULT 0,
 records_created INTEGER DEFAULT 0,
 error_message TEXT,
 started_at TIMESTAMPTZ DEFAULT NOW(),
 completed_at TIMESTAMPTZ
);

CREATE TABLE IF NOT EXISTS ai_activity_ledger(
 id BIGSERIAL PRIMARY KEY,
 event_id VARCHAR(50) UNIQUE NOT NULL,
 actor_type VARCHAR(30) DEFAULT 'AI',
 actor_name TEXT,
 division VARCHAR(30),
 action TEXT NOT NULL,
 entity_type VARCHAR(40),
 entity_id VARCHAR(50),
 summary TEXT,
 status VARCHAR(40) DEFAULT 'SUCCESS',
 created_at TIMESTAMPTZ DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS ai_crm_sync(
 id BIGSERIAL PRIMARY KEY,
 sync_id VARCHAR(50) UNIQUE NOT NULL,
 entity_type VARCHAR(40),
 entity_id VARCHAR(50),
 direction VARCHAR(30),
 target VARCHAR(50) DEFAULT 'FLOWCONNECT',
 status VARCHAR(40) DEFAULT 'PENDING',
 request_payload JSONB,
 response_payload JSONB,
 error_message TEXT,
 created_at TIMESTAMPTZ DEFAULT NOW(),
 completed_at TIMESTAMPTZ
);

CREATE INDEX IF NOT EXISTS idx_ai_companies_division ON ai_companies(division);
CREATE INDEX IF NOT EXISTS idx_ai_marketing_phone ON ai_marketing_contacts(phone);
CREATE INDEX IF NOT EXISTS idx_ai_demand_campaign ON ai_demand_signals(campaign_id);
CREATE INDEX IF NOT EXISTS idx_ai_activity_created ON ai_activity_ledger(created_at DESC);
"""

V4_MIGRATIONS = [
    "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS monthly_rent NUMERIC(14,2)",
    "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS assigned_to TEXT",
    "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS owner_id VARCHAR(50)",
    "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS broker_id VARCHAR(50)",
    "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS contact_number TEXT",
    "ALTER TABLE ai_companies ADD COLUMN IF NOT EXISTS primary_contact_name TEXT",
    "ALTER TABLE ai_companies ADD COLUMN IF NOT EXISTS primary_contact_phone TEXT",
    "ALTER TABLE ai_companies ADD COLUMN IF NOT EXISTS primary_contact_email TEXT",
    "ALTER TABLE ai_demand_signals ADD COLUMN IF NOT EXISTS match_breakdown JSONB",
    "ALTER TABLE ai_demand_signals ADD COLUMN IF NOT EXISTS contact_verification_status VARCHAR(40) DEFAULT 'UNVERIFIED'",
    "ALTER TABLE ai_demand_signals ADD COLUMN IF NOT EXISTS source_contact_text TEXT"
]

@app.on_event("startup")
def v4_startup():
    with engine.begin() as c:
        for stmt in [x.strip() for x in V4_SCHEMA.split(";") if x.strip()]:
            c.execute(text(stmt))
        for stmt in V4_MIGRATIONS:
            try: c.execute(text(stmt))
            except Exception: pass

def _new_code(prefix):
    return f"{prefix}-{datetime.utcnow().strftime('%Y%m%d')}-{uuid.uuid4().hex[:8].upper()}"

def _norm(v):
    return _re.sub(r"\s+"," ",str(v or "").strip().lower())

def _tokens(v):
    stop={"the","and","or","in","at","of","for","road","rd","sector","sec","phase","near"}
    return {x for x in _re.findall(r"[a-z0-9]+",_norm(v)) if len(x)>1 and x not in stop}

def _float(v):
    try:
        if v is None or v=="": return None
        return float(str(v).replace(",","").replace("₹","").strip())
    except Exception:
        return None

def _json_rows(rows):
    return [{k:serialize_db_value(v) for k,v in dict(r._mapping).items()} for r in rows]

def _log_activity(actor_name,division,action,entity_type=None,entity_id=None,summary=None,status="SUCCESS"):
    eid=_new_code("EVT")
    with engine.begin() as c:
        c.execute(text("""INSERT INTO ai_activity_ledger
          (event_id,actor_type,actor_name,division,action,entity_type,entity_id,summary,status)
          VALUES(:e,'AI',:a,:d,:ac,:et,:ei,:s,:st)"""),
          {"e":eid,"a":actor_name,"d":division,"ac":action,"et":entity_type,"ei":entity_id,"s":summary,"st":status})
    return eid

def _upsert_owner(c,name,phone=None,email=None,city=None,notes=None):
    if not (name or phone): return None
    row=None
    if phone:
        row=c.execute(text("SELECT owner_id FROM pi_owners WHERE contact_number=:p LIMIT 1"),{"p":phone}).first()
    if not row and name:
        row=c.execute(text("SELECT owner_id FROM pi_owners WHERE LOWER(owner_name)=LOWER(:n) LIMIT 1"),{"n":name}).first()
    if row:
        oid=row[0]
        c.execute(text("""UPDATE pi_owners SET owner_name=COALESCE(NULLIF(:n,''),owner_name),
                         contact_number=COALESCE(NULLIF(:p,''),contact_number),
                         email=COALESCE(NULLIF(:e,''),email),city=COALESCE(NULLIF(:city,''),city),
                         notes=COALESCE(NULLIF(:notes,''),notes),updated_at=NOW() WHERE owner_id=:id"""),
                  {"n":name or "","p":phone or "","e":email or "","city":city or "","notes":notes or "","id":oid})
        return oid
    oid=_new_code("OWN")
    c.execute(text("""INSERT INTO pi_owners(owner_id,owner_name,contact_number,email,city,notes)
                     VALUES(:id,:n,:p,:e,:city,:notes)"""),
              {"id":oid,"n":name or "Unknown Owner","p":phone,"e":email,"city":city,"notes":notes})
    return oid

def _upsert_broker(c,name,phone=None,email=None,company=None,city=None,notes=None):
    if not (name or phone): return None
    row=None
    if phone:
        row=c.execute(text("SELECT broker_id FROM pi_brokers WHERE contact_number=:p LIMIT 1"),{"p":phone}).first()
    if not row and name:
        row=c.execute(text("SELECT broker_id FROM pi_brokers WHERE LOWER(broker_name)=LOWER(:n) LIMIT 1"),{"n":name}).first()
    if row:
        bid=row[0]
        c.execute(text("""UPDATE pi_brokers SET broker_name=COALESCE(NULLIF(:n,''),broker_name),
                         contact_number=COALESCE(NULLIF(:p,''),contact_number),
                         email=COALESCE(NULLIF(:e,''),email),company_name=COALESCE(NULLIF(:co,''),company_name),
                         city=COALESCE(NULLIF(:city,''),city),notes=COALESCE(NULLIF(:notes,''),notes),
                         updated_at=NOW() WHERE broker_id=:id"""),
                  {"n":name or "","p":phone or "","e":email or "","co":company or "","city":city or "","notes":notes or "","id":bid})
        return bid
    bid=_new_code("BRK")
    c.execute(text("""INSERT INTO pi_brokers(broker_id,broker_name,contact_number,email,company_name,city,notes)
                     VALUES(:id,:n,:p,:e,:co,:city,:notes)"""),
              {"id":bid,"n":name or "Unknown Broker","p":phone,"e":email,"co":company,"city":city,"notes":notes})
    return bid

async def _store_media_files(property_id,files):
    saved=[]
    for file in files or []:
        if not file or not file.filename: continue
        mime=(file.content_type or "").lower()
        is_img=mime in {"image/jpeg","image/png","image/webp","image/gif"}
        is_vid=mime in {"video/mp4","video/webm","video/quicktime","video/x-m4v"}
        if not (is_img or is_vid): continue
        data=await file.read()
        maxb=(MAX_IMAGE_MB if is_img else MAX_VIDEO_MB)*1024*1024
        if len(data)>maxb:
            raise HTTPException(413,f"{file.filename} is too large.")
        mid=str(uuid.uuid4())
        with engine.begin() as c:
            c.execute(text("""INSERT INTO pi_property_media(media_id,property_id,media_type,filename,mime_type,file_size,content)
                VALUES(CAST(:mid AS UUID),:pid,:mt,:fn,:mime,:sz,:data)"""),
                {"mid":mid,"pid":property_id,"mt":"IMAGE" if is_img else "VIDEO","fn":file.filename,
                 "mime":mime,"sz":len(data),"data":data})
        saved.append({"media_id":mid,"type":"IMAGE" if is_img else "VIDEO","filename":file.filename})
    return saved

@app.post("/api/v4/properties/manual")
async def v4_add_property_manual(
    req:Request,
    property_name:str=Form(""),
    property_type:str=Form("NA"),
    city:str=Form("NA"),
    location:str=Form("NA"),
    available_area_sqft:Optional[str]=Form(None),
    minimum_area_sqft:Optional[str]=Form(None),
    maximum_area_sqft:Optional[str]=Form(None),
    floor:Optional[str]=Form(None),
    rent_or_sale:Optional[str]=Form(None),
    monthly_rent:Optional[str]=Form(None),
    nearby_brands:Optional[str]=Form(None),
    suitable_category:Optional[str]=Form(None),
    parking:Optional[str]=Form(None),
    verification_status:str=Form("UNVERIFIED"),
    assigned_to:Optional[str]=Form(None),
    contact_number:Optional[str]=Form(None),
    owner_name:Optional[str]=Form(None),
    owner_contact:Optional[str]=Form(None),
    owner_email:Optional[str]=Form(None),
    broker_name:Optional[str]=Form(None),
    broker_contact:Optional[str]=Form(None),
    broker_email:Optional[str]=Form(None),
    broker_company:Optional[str]=Form(None),
    remarks:Optional[str]=Form(None),
    media:list[UploadFile]=File(default=[])
):
    need_login(req)
    verification_status="VERIFIED" if verification_status.upper()=="VERIFIED" else "UNVERIFIED"
    with engine.begin() as c:
        oid=_upsert_owner(c,owner_name,owner_contact,owner_email,city,remarks)
        bid=_upsert_broker(c,broker_name,broker_contact,broker_email,broker_company,city,remarks)
        pid=make_id("PROP",c)
        fp=fingerprint([property_name,city,location,property_type,available_area_sqft,floor,rent_or_sale,owner_contact,broker_contact])
        c.execute(text("""INSERT INTO pi_properties(
            property_id,fingerprint,property_name,property_type,city,location,available_area_sqft,
            minimum_area_sqft,maximum_area_sqft,floor,rent_or_sale,monthly_rent,nearby_brands,
            suitable_category,parking,owner_id,broker_id,owner_name,owner_contact,broker_name,broker_contact,
            remarks,source,verification_status,verified_date,verified_by,assigned_to,contact_number)
            VALUES(:pid,:fp,:pn,:pt,:city,:loc,:aa,:mn,:mx,:floor,:rs,:rent,:nb,:cat,:park,:oid,:bid,
            :on,:oc,:bn,:bc,:rem,'Manual V4',:vs,
            CASE WHEN :vs='VERIFIED' THEN CURRENT_DATE ELSE NULL END,
            CASE WHEN :vs='VERIFIED' THEN :actor ELSE NULL END,:assigned,:contact)"""),
            {"pid":pid,"fp":fp,"pn":property_name,"pt":property_type,"city":city,"loc":location,
             "aa":_float(available_area_sqft),"mn":_float(minimum_area_sqft),"mx":_float(maximum_area_sqft),
             "floor":floor,"rs":rent_or_sale,"rent":_float(monthly_rent),"nb":nearby_brands,"cat":suitable_category,
             "park":parking,"oid":oid,"bid":bid,"on":owner_name,"oc":owner_contact,"bn":broker_name,"bc":broker_contact,
             "rem":remarks,"vs":verification_status,"actor":actor_name(req),"assigned":assigned_to,
             "contact":contact_number or owner_contact or broker_contact})
        c.execute(text("""INSERT INTO pi_verification_log(property_id,action,performed_by,notes)
                         VALUES(:pid,'CREATED_V4',:actor,:notes)"""),
                  {"pid":pid,"actor":actor_name(req),"notes":f"Manual property; verification={verification_status}"})
    saved=await _store_media_files(pid,media)
    _log_activity("Dashboard","PROPERTY","PROPERTY_CREATED","property",pid,
                  f"{property_name or pid}; assigned to {assigned_to or 'Unassigned'}")
    return {"status":"created","property_id":pid,"owner_id":oid,"broker_id":bid,"media":saved}

@app.get("/api/v4/owners")
def v4_owners(req:Request,limit:int=Query(300,ge=1,le=1000)):
    need_login(req)
    with engine.connect() as c:
        rows=c.execute(text("SELECT * FROM pi_owners ORDER BY updated_at DESC LIMIT :n"),{"n":limit}).fetchall()
    return {"status":"ok","rows":_json_rows(rows)}

@app.get("/api/v4/brokers")
def v4_brokers(req:Request,limit:int=Query(300,ge=1,le=1000)):
    need_login(req)
    with engine.connect() as c:
        rows=c.execute(text("SELECT * FROM pi_brokers ORDER BY updated_at DESC LIMIT :n"),{"n":limit}).fetchall()
    return {"status":"ok","rows":_json_rows(rows)}

class V4CompanyInput(BaseModel):
    division: Literal["HOSPITALITY","RETAIL"]
    company_name:str
    category:Optional[str]=None
    primary_contact_name:Optional[str]=None
    primary_contact_phone:Optional[str]=None
    primary_contact_email:Optional[str]=None
    website:Optional[str]=None
    linkedin_url:Optional[str]=None
    city:Optional[str]=None
    target_markets:Optional[str]=None
    expansion_status:Optional[str]="DISCOVERED"
    expansion_score:Optional[float]=0
    source_name:Optional[str]="Manual"
    source_url:Optional[str]=None
    source_excerpt:Optional[str]=None
    assigned_to:Optional[str]=None

@app.post("/api/v4/companies")
def v4_add_company(p:V4CompanyInput,req:Request):
    need_login(req)
    cid=_new_code("CMP")
    with engine.begin() as c:
        c.execute(text("""INSERT INTO ai_companies(company_id,division,company_name,category,primary_contact_name,
            primary_contact_phone,primary_contact_email,website,linkedin_url,city,target_markets,expansion_status,
            expansion_score,source_name,source_url,source_excerpt,assigned_to)
            VALUES(:id,:division,:company_name,:category,:primary_contact_name,:primary_contact_phone,
            :primary_contact_email,:website,:linkedin_url,:city,:target_markets,:expansion_status,:expansion_score,
            :source_name,:source_url,:source_excerpt,:assigned_to)"""),{"id":cid,**p.model_dump()})
        if p.primary_contact_name or p.primary_contact_phone or p.primary_contact_email:
            c.execute(text("""INSERT INTO ai_contacts(contact_id,company_id,full_name,business_email,business_phone,
                linkedin_url,verification_status,source_url)
                VALUES(:id,:cid,:n,:e,:p,:li,'UNVERIFIED',:src)"""),
                {"id":_new_code("CON"),"cid":cid,"n":p.primary_contact_name,"e":p.primary_contact_email,
                 "p":p.primary_contact_phone,"li":p.linkedin_url,"src":p.source_url})
    _log_activity("Dashboard",p.division,"PROSPECT_CREATED","company",cid,p.company_name)
    return {"status":"created","company_id":cid}

@app.get("/api/v4/companies")
def v4_companies(req:Request,division:str=Query("RETAIL"),limit:int=Query(200,ge=1,le=500)):
    need_login(req)
    with engine.connect() as c:
        rows=c.execute(text("SELECT * FROM ai_companies WHERE division=:d ORDER BY expansion_score DESC,created_at DESC LIMIT :n"),
                       {"d":division.upper(),"n":limit}).fetchall()
    return {"status":"ok","rows":_json_rows(rows)}

def _marketing_fp(brand,phone,email,location):
    return hashlib.sha256("|".join([_norm(brand),_norm(phone),_norm(email),_norm(location)]).encode()).hexdigest()

def _save_marketing_contact(data):
    fp=_marketing_fp(data.get("brand_name"),data.get("phone"),data.get("email"),data.get("location"))
    with engine.begin() as c:
        old=c.execute(text("SELECT contact_id FROM ai_marketing_contacts WHERE fingerprint=:f"),{"f":fp}).first()
        if old: return {"status":"duplicate","contact_id":old[0]}
        cid=_new_code("MKT")
        c.execute(text("""INSERT INTO ai_marketing_contacts(contact_id,fingerprint,business_type,brand_name,contact_name,
            phone,email,website,location,city,source_name,source_url,consent_status,verification_status,assigned_to)
            VALUES(:id,:fp,:bt,:brand,:name,:phone,:email,:web,:loc,:city,:source,:url,:consent,:verify,:assigned)"""),
            {"id":cid,"fp":fp,"bt":data.get("business_type"),"brand":data.get("brand_name"),
             "name":data.get("contact_name"),"phone":data.get("phone"),"email":data.get("email"),
             "web":data.get("website"),"loc":data.get("location"),"city":data.get("city"),
             "source":data.get("source_name"),"url":data.get("source_url"),
             "consent":data.get("consent_status") or "UNKNOWN","verify":data.get("verification_status") or "UNVERIFIED",
             "assigned":data.get("assigned_to")})
    return {"status":"created","contact_id":cid}

@app.get("/api/v4/marketing-contacts")
def v4_marketing_contacts(req:Request,limit:int=Query(500,ge=1,le=2000)):
    need_login(req)
    with engine.connect() as c:
        rows=c.execute(text("SELECT * FROM ai_marketing_contacts ORDER BY created_at DESC LIMIT :n"),{"n":limit}).fetchall()
    return {"status":"ok","rows":_json_rows(rows)}

@app.post("/api/v4/marketing-contacts/upload")
async def v4_upload_contacts(req:Request,file:UploadFile=File(...)):
    need_login(req)
    raw=await file.read()
    textdata=raw.decode("utf-8-sig",errors="replace")
    reader=csv.DictReader(io.StringIO(textdata))
    created=duplicates=0
    for row in reader:
        data={
            "business_type":row.get("business_type") or row.get("Business Type") or row.get("category") or row.get("Category"),
            "brand_name":row.get("brand_name") or row.get("Brand Name") or row.get("company") or row.get("Company"),
            "contact_name":row.get("contact_name") or row.get("Contact Name") or row.get("name") or row.get("Name"),
            "phone":row.get("phone") or row.get("Phone") or row.get("contact_no") or row.get("Contact No"),
            "email":row.get("email") or row.get("Email"),
            "website":row.get("website") or row.get("Website"),
            "location":row.get("location") or row.get("Location"),
            "city":row.get("city") or row.get("City") or "Delhi NCR",
            "source_name":"Uploaded CSV: "+(file.filename or "contacts.csv"),
            "consent_status":row.get("consent_status") or row.get("Consent Status") or "UNKNOWN",
            "verification_status":row.get("verification_status") or row.get("Verification Status") or "UNVERIFIED",
            "assigned_to":row.get("assigned_to") or row.get("Team Member")
        }
        r=_save_marketing_contact(data)
        if r["status"]=="created": created+=1
        else: duplicates+=1
    _log_activity("Contact Upload","HOSPITALITY","CONTACT_DATABASE_UPLOADED","contact_database",None,
                  f"{created} new; {duplicates} duplicates")
    return {"status":"processed","created":created,"duplicates":duplicates}

@app.get("/api/v4/marketing-contacts/export.csv")
def v4_export_contacts(req:Request):
    need_login(req)
    with engine.connect() as c:
        rows=c.execute(text("""SELECT business_type,brand_name,contact_name,phone,email,website,location,city,
                              consent_status,verification_status,assigned_to,source_name,source_url
                              FROM ai_marketing_contacts ORDER BY created_at DESC""")).fetchall()
    out=io.StringIO()
    w=csv.writer(out)
    w.writerow(["business_type","brand_name","contact_name","phone","email","website","location","city",
                "consent_status","verification_status","assigned_to","source_name","source_url"])
    for r in rows: w.writerow(list(r))
    return Response(content=out.getvalue(),media_type="text/csv",
                    headers={"Content-Disposition":"attachment; filename=hospitality_marketing_contacts.csv"})

def _safe_public_url(url):
    try:
        u=_urlparse(url)
        if u.scheme not in {"http","https"} or not u.hostname: return False
        ip=_socket.gethostbyname(u.hostname)
        obj=_ipaddress.ip_address(ip)
        return not (obj.is_private or obj.is_loopback or obj.is_link_local or obj.is_reserved)
    except Exception:
        return False

def _public_email_from_website(url):
    if not _httpx or not url or not _safe_public_url(url): return None
    try:
        r=_httpx.get(url,timeout=8.0,follow_redirects=True,headers={"User-Agent":"Mozilla/5.0"})
        if r.status_code>=400: return None
        emails=_re.findall(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}",r.text,_re.I)
        bad=("example.com","wixpress.com","sentry.io")
        emails=[e for e in emails if not any(b in e.lower() for b in bad)]
        return emails[0][:255] if emails else None
    except Exception:
        return None

def _google_places_search(query):
    if not GOOGLE_PLACES_API_KEY:
        raise RuntimeError("GOOGLE_PLACES_API_KEY is not configured.")
    if not _httpx: raise RuntimeError("httpx dependency missing.")
    fields="places.id,places.displayName,places.formattedAddress,places.nationalPhoneNumber,places.internationalPhoneNumber,places.websiteUri,places.googleMapsUri,places.types"
    r=_httpx.post("https://places.googleapis.com/v1/places:searchText",
        headers={"X-Goog-Api-Key":GOOGLE_PLACES_API_KEY,"X-Goog-FieldMask":fields,"Content-Type":"application/json"},
        json={"textQuery":query,"maxResultCount":20,"languageCode":"en"},timeout=30.0)
    r.raise_for_status()
    return r.json().get("places",[])

def _start_bot(name,division,summary):
    rid=_new_code("RUN")
    with engine.begin() as c:
        c.execute(text("""INSERT INTO ai_bot_runs(run_id,bot_name,division,status,summary)
                         VALUES(:id,:n,:d,'RUNNING',:s)"""),{"id":rid,"n":name,"d":division,"s":summary})
    return rid

def _finish_bot(run_id,status,found,created,summary,error=None):
    with engine.begin() as c:
        c.execute(text("""UPDATE ai_bot_runs SET status=:st,records_found=:f,records_created=:cr,
            summary=:s,error_message=:e,completed_at=NOW() WHERE run_id=:id"""),
            {"st":status,"f":found,"cr":created,"s":summary,"e":error,"id":run_id})

def _hospitality_worker(run_id):
    categories=["restaurant","cafe","lounge","club","nightclub","banquet hall","hotel","guest house","guesthouse","wedding venue","commercial farmhouse"]
    zones=["South Delhi","West Delhi","Central Delhi","Gurgaon","Noida","Faridabad","Ghaziabad"]
    found=created=0; errors=[]
    try:
        for category in categories:
            for zone in zones:
                try:
                    for p in _google_places_search(f"{category} in {zone}"):
                        found+=1
                        name=((p.get("displayName") or {}).get("text") or "").strip()
                        if not name: continue
                        phone=p.get("nationalPhoneNumber") or p.get("internationalPhoneNumber")
                        website=p.get("websiteUri")
                        address=p.get("formattedAddress")
                        email=_public_email_from_website(website) if website else None
                        r=_save_marketing_contact({
                            "business_type":category.title(),"brand_name":name,"contact_name":None,
                            "phone":phone,"email":email,"website":website,"location":address,"city":"Delhi NCR",
                            "source_name":"Google Places","source_url":p.get("googleMapsUri"),
                            "consent_status":"UNKNOWN","verification_status":"PUBLIC_SOURCE"
                        })
                        if r["status"]=="created":
                            created+=1
                            _log_activity("Hospitality Data Bot","HOSPITALITY","CONTACT_FOUND","marketing_contact",
                                          r["contact_id"],f"{name} | {phone or 'no phone'}")
                except Exception as ex:
                    errors.append(f"{category}/{zone}: {ex}")
        _finish_bot(run_id,"COMPLETED" if created or not errors else "FAILED",found,created,
                    f"Hospitality scan: {found} places reviewed; {created} new contacts", " | ".join(errors[:5]) or None)
    except Exception as ex:
        _finish_bot(run_id,"FAILED",found,created,"Hospitality scan failed",str(ex))

@app.post("/api/v4/hospitality-bot/start")
def v4_hospitality_bot(bg:BackgroundTasks,req:Request):
    need_login(req)
    run=_start_bot("Delhi NCR Hospitality Data Bot","HOSPITALITY","Fetching public business contact data")
    bg.add_task(_hospitality_worker,run)
    return {"status":"ACCEPTED","run_id":run}

def _serper_search(query,num=10):
    if not SERPER_API_KEY: raise RuntimeError("SERPER_API_KEY is not configured.")
    if not _httpx: raise RuntimeError("httpx dependency missing.")
    r=_httpx.post("https://google.serper.dev/search",
        headers={"X-API-KEY":SERPER_API_KEY,"Content-Type":"application/json"},
        json={"q":query,"num":num},timeout=30.0)
    r.raise_for_status()
    return r.json()

def _retail_score(title,snippet):
    s=_norm((title or "")+" "+(snippet or ""))
    score=10
    groups=[
        (["delhi ncr","delhi","gurgaon","gurugram","noida","faridabad","ghaziabad"],25),
        (["expansion","expand","rollout","growth plan"],20),
        (["new store","new stores","outlet","flagship","store opening"],20),
        (["lease","leasing","retail space","mall","high street"],15),
        (["india entry","north india"],10)
    ]
    for terms,pts in groups:
        if any(t in s for t in terms): score+=pts
    return min(100,score)

def _company_guess(title):
    parts=_re.split(r"\s+(?:plans|plan|to|opens|open|launches|launch|expands|expand|eyes|targets|set to)\s+",
                    str(title or "").strip(),1,flags=_re.I)
    return (parts[0] if parts else str(title or "")).strip(" -:|")[:180] or "Unknown Retail Company"

def _retail_worker(run_id):
    queries=[
        '"Delhi NCR" retail expansion new stores brand',
        'Gurugram retail brand expansion store opening',
        'Noida retail brand expansion new store',
        'India retail expansion plans stores 2026',
        'site:linkedin.com/posts retail expansion Delhi NCR',
        'retail news India brand opening stores Delhi Gurgaon Noida'
    ]
    found=created=0; errors=[]
    for q in queries:
        try:
            for item in _serper_search(q,10).get("organic",[]):
                found+=1
                title=item.get("title") or ""; link=item.get("link") or ""; snippet=item.get("snippet") or ""
                score=_retail_score(title,snippet); company=_company_guess(title)
                with engine.begin() as c:
                    exists=c.execute(text("SELECT 1 FROM ai_expansion_signals WHERE source_url=:u LIMIT 1"),{"u":link}).first()
                    if exists: continue
                    row=c.execute(text("""SELECT company_id FROM ai_companies
                        WHERE division='RETAIL' AND LOWER(company_name)=LOWER(:n) LIMIT 1"""),{"n":company}).first()
                    if row: cid=row[0]
                    else:
                        cid=_new_code("CMP")
                        c.execute(text("""INSERT INTO ai_companies(company_id,division,company_name,category,target_markets,
                            expansion_status,expansion_score,source_name,source_url,source_excerpt)
                            VALUES(:id,'RETAIL',:n,'Retail','Delhi NCR','SIGNAL_DETECTED',:s,'Public Web / Retail News',:u,:x)"""),
                            {"id":cid,"n":company,"s":score,"u":link,"x":snippet})
                    sid=_new_code("SIG")
                    c.execute(text("""INSERT INTO ai_expansion_signals(signal_id,company_id,division,title,signal_type,market,
                        source_name,source_url,excerpt,signal_score)
                        VALUES(:sid,:cid,'RETAIL',:t,'EXPANSION_SIGNAL','Delhi NCR','Public Web / Retail News',:u,:x,:s)"""),
                        {"sid":sid,"cid":cid,"t":title,"u":link,"x":snippet,"s":score})
                    c.execute(text("UPDATE ai_companies SET expansion_score=GREATEST(expansion_score,:s),updated_at=NOW() WHERE company_id=:id"),
                              {"s":score,"id":cid})
                    created+=1
                _log_activity("Retail Expansion Bot","RETAIL","EXPANSION_SIGNAL_FOUND","company",cid,f"{company} | {score}")
        except Exception as ex: errors.append(str(ex))
    _finish_bot(run_id,"COMPLETED" if created or not errors else "FAILED",found,created,
                f"Retail scan: {found} results; {created} new signals"," | ".join(errors[:5]) or None)

@app.post("/api/v4/retail-bot/start")
def v4_retail_bot(bg:BackgroundTasks,req:Request):
    need_login(req)
    run=_start_bot("Retail Expansion Bot","RETAIL","Scanning public retail expansion signals")
    bg.add_task(_retail_worker,run)
    req_run=_start_bot("Retail LinkedIn Requirement Bot","DEMAND","Scanning LinkedIn/public-indexed retail leasing requirements")
    bg.add_task(_retail_linkedin_requirement_worker,req_run)
    return {"status":"ACCEPTED","run_id":run,"requirement_run_id":req_run}

class CampaignInput(BaseModel):
    campaign_name:str
    property_id:Optional[str]=None
    property_type:Optional[str]=None
    city:Optional[str]="Delhi NCR"
    location:Optional[str]=None
    area_sqft:Optional[float]=None
    monthly_rent:Optional[float]=None
    rent_or_sale:Optional[str]="Rent"
    suitable_category:Optional[str]=None
    nearby_brands:Optional[str]=None
    additional_points:Optional[str]=None
    assigned_to:Optional[str]=None

def _post_draft(p):
    bits=[p.property_type or "commercial property",p.location or p.city or "Delhi NCR"]
    if p.area_sqft: bits.append(f"{p.area_sqft:,.0f} sqft")
    if p.monthly_rent: bits.append(f"₹{p.monthly_rent:,.0f}/month")
    if p.suitable_category: bits.append(f"suitable for {p.suitable_category}")
    return "Available: "+", ".join(bits)+". Genuine business requirements may connect with our leasing team."

@app.post("/api/v4/requirement-campaigns")
def v4_create_campaign(p:CampaignInput,req:Request):
    need_login(req)
    cid=_new_code("CAM")
    draft=_post_draft(p)
    with engine.begin() as c:
        c.execute(text("""INSERT INTO ai_requirement_campaigns(campaign_id,property_id,campaign_name,property_type,city,
            location,area_sqft,monthly_rent,rent_or_sale,suitable_category,nearby_brands,additional_points,post_draft,assigned_to)
            VALUES(:id,:property_id,:campaign_name,:property_type,:city,:location,:area_sqft,:monthly_rent,:rent_or_sale,
            :suitable_category,:nearby_brands,:additional_points,:draft,:assigned_to)"""),
            {"id":cid,"draft":draft,**p.model_dump()})
    _log_activity("Dashboard","DEMAND","DEMAND_CAMPAIGN_CREATED","campaign",cid,p.campaign_name)
    return {"status":"created","campaign_id":cid,"post_draft":draft}

def _extract_public_contacts(text_value):
    raw=str(text_value or "")
    phones=[]
    for m in _re.findall(r'(?:(?:\+?91[\s-]?)?[6-9]\d{9})', raw):
        digits=_re.sub(r'\D','',m)
        if len(digits)==12 and digits.startswith("91"):
            digits=digits[2:]
        if len(digits)==10 and digits[0] in "6789" and digits not in phones:
            phones.append(digits)
    emails=[]
    for e in _re.findall(r'[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}', raw, _re.I):
        e=e.lower()
        if e not in emails:
            emails.append(e)
    return phones[:5], emails[:5]

def _campaign_match_score(title,snippet,camp):
    textv=_norm((title or "")+" "+(snippet or ""))
    score=0
    breakdown={}

    intent_terms=[
        "looking for","requirement","required","seeking","need space","needs space",
        "want to lease","wants to lease","looking to lease","looking to rent",
        "expansion requirement","space requirement","require commercial space",
        "looking for property","looking for shop","looking for showroom",
        "looking for restaurant space","looking for retail space"
    ]
    hits=[x for x in intent_terms if x in textv]
    s=30 if hits else 0
    breakdown["intent"]={"score":s,"max":30,"hits":hits[:4]}
    score+=s

    target_loc=camp.get("location") or camp.get("city") or ""
    loc_tokens=_tokens(target_loc); text_tokens=_tokens(textv); overlap=loc_tokens & text_tokens
    s=0
    if loc_tokens:
        ratio=len(overlap)/max(1,len(loc_tokens))
        if ratio>=0.75:s=25
        elif ratio>=0.5:s=20
        elif ratio>=0.25:s=12
        elif any(x in textv for x in ["delhi ncr","delhi","gurgaon","gurugram","noida"]) and any(x in _norm(target_loc) for x in ["delhi ncr","delhi","gurgaon","gurugram","noida"]):
            s=10
    else:s=10
    breakdown["location"]={"score":s,"max":25,"hits":sorted(overlap)[:6]}
    score+=s

    target_cat=camp.get("suitable_category") or camp.get("property_type") or ""
    cat_tokens=_type_family(target_cat); cat_overlap=cat_tokens & text_tokens
    s=0
    if cat_tokens:
        ratio=len(cat_overlap)/max(1,min(len(cat_tokens),4))
        if ratio>=0.75:s=20
        elif ratio>=0.5:s=16
        elif cat_overlap:s=10
    else:s=8
    breakdown["category"]={"score":s,"max":20,"hits":sorted(cat_overlap)[:6]}
    score+=s

    tx=_transaction_family(camp.get("rent_or_sale"))
    s=0
    if tx=="rent" and any(x in textv for x in ["lease","rent","rental","to let"]):s=10
    elif tx=="sale" and any(x in textv for x in ["buy","purchase","sale","acquire"]):s=10
    elif not tx:s=5
    breakdown["transaction"]={"score":s,"max":10,"target":tx}
    score+=s

    area_target=_float(camp.get("area_sqft")); s=0; area_hits=[]
    if area_target:
        vals=[]
        for m in _re.findall(r'(\d[\d,]{2,})\s*(?:sq\s*ft|sqft|square\s*feet|sft)', textv, _re.I):
            try: vals.append(float(m.replace(",","")))
            except: pass
        if vals:
            best=min(abs(x-area_target)/max(area_target,1) for x in vals)
            if best<=0.10:s=10
            elif best<=0.20:s=8
            elif best<=0.35:s=5
            area_hits=[int(x) for x in vals[:4]]
    else:s=5
    breakdown["area"]={"score":s,"max":10,"hits":area_hits}
    score+=s

    context_tokens=_tokens((camp.get("nearby_brands") or "")+" "+(camp.get("additional_points") or ""))
    context_overlap=context_tokens & text_tokens
    s=5 if context_overlap else 0
    breakdown["context"]={"score":s,"max":5,"hits":sorted(context_overlap)[:5]}
    score+=s

    return min(100,score),breakdown

def _fetch_public_source_text(url):
    if not _httpx or not url or not _safe_public_url(url):
        return ""
    try:
        r=_httpx.get(url,timeout=8.0,follow_redirects=True,headers={"User-Agent":"Mozilla/5.0 (compatible; DealIntelligenceBot/1.0)"})
        if r.status_code>=400:return ""
        ctype=(r.headers.get("content-type") or "").lower()
        if "text/html" not in ctype and "text/plain" not in ctype:return ""
        return (r.text or "")[:250000]
    except Exception:
        return ""

def _requirement_worker(run_id,campaign_id):
    found=created=rejected=0
    errors=[]
    with engine.connect() as c:
        row=c.execute(text("SELECT * FROM ai_requirement_campaigns WHERE campaign_id=:id"),{"id":campaign_id}).first()
    if not row:
        _finish_bot(run_id,"FAILED",0,0,"Campaign not found","Campaign not found")
        return

    camp=dict(row._mapping)
    loc=camp.get("location") or camp.get("city") or "Delhi NCR"
    cat=camp.get("suitable_category") or camp.get("property_type") or "commercial"
    tx=camp.get("rent_or_sale") or "Rent"

    queries=[
        f'"looking for" "{cat}" "{loc}" {tx}',
        f'"space requirement" "{cat}" "{loc}"',
        f'"requirement" "{cat}" "{loc}" lease',
        f'"seeking" "{cat}" space "{loc}"',
        f'site:linkedin.com/posts "looking for" "{cat}" "{loc}"',
        f'site:linkedin.com/posts "requirement" "{cat}" "{loc}"',
        f'site:facebook.com "looking for" "{cat}" "{loc}"',
        f'site:instagram.com "looking for" "{cat}" "{loc}"',
        f'site:99acres.com "{cat}" "{loc}" requirement',
        f'site:magicbricks.com "{cat}" "{loc}" requirement'
    ]

    for q in queries:
        try:
            for item in _serper_search(q,10).get("organic",[]):
                found+=1
                link=item.get("link") or ""
                title=item.get("title") or ""
                snippet=item.get("snippet") or ""
                score,breakdown=_campaign_match_score(title,snippet,camp)

                if score<90:
                    rejected+=1
                    continue

                with engine.connect() as c:
                    if c.execute(text("SELECT 1 FROM ai_demand_signals WHERE source_url=:u AND campaign_id=:cid LIMIT 1"),
                                 {"u":link,"cid":campaign_id}).first():
                        continue

                source="Public Web"
                if "linkedin.com" in link:source="LinkedIn public/indexed"
                elif "facebook.com" in link:source="Facebook public/indexed"
                elif "instagram.com" in link:source="Instagram public/indexed"
                elif "99acres.com" in link:source="99acres public/indexed"
                elif "magicbricks.com" in link:source="Magicbricks public/indexed"

                phones,emails=_extract_public_contacts(title+" "+snippet)
                if not phones or not emails:
                    page=_fetch_public_source_text(link)
                    p2,e2=_extract_public_contacts(page)
                    for p in p2:
                        if p not in phones:phones.append(p)
                    for e in e2:
                        if e not in emails:emails.append(e)

                phone=phones[0] if phones else None
                email=emails[0] if emails else None
                verification="PUBLIC_SOURCE" if (phone or email) else "NOT_FOUND"
                sid=_new_code("DEM")

                with engine.begin() as c:
                    c.execute(text("""INSERT INTO ai_demand_signals(
                        signal_id,campaign_id,source_type,source_name,source_url,title,excerpt,
                        contact_phone,contact_email,location,intent_score,status,match_breakdown,
                        contact_verification_status,source_contact_text)
                        VALUES(:sid,:cid,'WEB_DEMAND',:source,:url,:title,:excerpt,:phone,:email,:loc,:score,
                        '90_PLUS_MATCH',CAST(:breakdown AS JSONB),:verification,:contact_text)"""),
                        {"sid":sid,"cid":campaign_id,"source":source,"url":link,"title":title,"excerpt":snippet,
                         "phone":phone,"email":email,"loc":loc,"score":score,
                         "breakdown":json.dumps(breakdown),"verification":verification,
                         "contact_text":(title+" | "+snippet)[:2500]})
                    created+=1

                _log_activity("Requirement Discovery Bot","DEMAND","90_PLUS_DEMAND_MATCH","demand_signal",sid,
                              f"{source} | score {score} | contact {'yes' if phone else 'no'}")
        except Exception as ex:
            errors.append(str(ex))

    with engine.begin() as c:
        c.execute(text("UPDATE ai_requirement_campaigns SET status='SCANNED_90_PLUS',updated_at=NOW() WHERE campaign_id=:id"),
                  {"id":campaign_id})

    _finish_bot(run_id,"COMPLETED" if created or not errors else "FAILED",found,created,
                f"Strict demand scan: {found} reviewed; {created} saved at 90%+; {rejected} filtered out",
                " | ".join(errors[:5]) or None)

@app.post("/api/v4/requirement-campaigns/{campaign_id}/start")
def v4_start_requirement_campaign(campaign_id:str,bg:BackgroundTasks,req:Request):
    need_login(req)
    run=_start_bot("Requirement Discovery Bot","DEMAND",f"Scanning demand for {campaign_id}")
    bg.add_task(_requirement_worker,run,campaign_id)
    return {"status":"ACCEPTED","run_id":run}

@app.get("/api/v4/requirement-campaigns")
def v4_campaigns(req:Request):
    need_login(req)
    with engine.connect() as c:
        rows=c.execute(text("SELECT * FROM ai_requirement_campaigns ORDER BY created_at DESC LIMIT 200")).fetchall()
    return {"status":"ok","rows":_json_rows(rows)}

@app.get("/api/v4/demand-signals")
def v4_demand_signals(req:Request,campaign_id:Optional[str]=Query(None)):
    need_login(req)
    with engine.connect() as c:
        if campaign_id:
            rows=c.execute(text("SELECT * FROM ai_demand_signals WHERE campaign_id=:id ORDER BY intent_score DESC,created_at DESC"),
                           {"id":campaign_id}).fetchall()
        else:
            rows=c.execute(text("SELECT * FROM ai_demand_signals ORDER BY intent_score DESC,created_at DESC LIMIT 300")).fetchall()
    return {"status":"ok","rows":_json_rows(rows)}

def _city_family(value):
    v=_norm(value)
    if not v: return set()
    families=set(_tokens(v))
    if any(x in v for x in ["gurgaon","gurugram"]):
        families.update({"gurgaon","gurugram","delhi","ncr"})
    if "new delhi" in v or v=="delhi" or " delhi" in (" "+v):
        families.update({"delhi","newdelhi","ncr"})
    if "noida" in v:
        families.update({"noida","delhi","ncr"})
        if "greater" in v: families.add("greaternoida")
    if "faridabad" in v:
        families.update({"faridabad","delhi","ncr"})
    if "ghaziabad" in v:
        families.update({"ghaziabad","delhi","ncr"})
    if "delhi ncr" in v or v=="ncr":
        families.update({"delhi","ncr","gurgaon","gurugram","noida","faridabad","ghaziabad"})
    return families

def _transaction_family(value):
    v=_norm(value)
    if any(x in v for x in ["rent","lease","leasing","rental","to let"]): return "rent"
    if any(x in v for x in ["sale","sell","purchase","buy"]): return "sale"
    return v

def _type_family(value):
    toks=_tokens(value)
    v=_norm(value)
    if any(x in v for x in ["retail","shop","showroom","store"]):
        toks.update({"retail","shop","showroom","store","commercial"})
    if any(x in v for x in ["restaurant","cafe","café","qsr","food","f&b","fnb","lounge","club"]):
        toks.update({"restaurant","cafe","qsr","food","f&b","fnb","lounge","club","commercial","retail"})
    if any(x in v for x in ["banquet","wedding","farmhouse","hotel","hospitality"]):
        toks.update({"banquet","wedding","farmhouse","hotel","hospitality","commercial"})
    if any(x in v for x in ["office","commercial","business centre","cowork"]):
        toks.update({"office","commercial"})
    if any(x in v for x in ["warehouse","industrial","factory","logistics"]):
        toks.update({"warehouse","industrial","commercial"})
    if any(x in v for x in ["residential","apartment","flat","villa","house","builder floor","residence"]):
        toks.update({"residential"})
    return toks


def _property_class(*values):
    v=_norm(" ".join(str(x or "") for x in values))
    commercial=[
        "commercial","retail","shop","showroom","office","restaurant","cafe","café",
        "qsr","food","f&b","fnb","lounge","club","banquet","hotel","hospitality",
        "warehouse","industrial","factory","business centre","cowork"
    ]
    residential=[
        "residential","apartment","flat","villa","house","builder floor","residence",
        "residential floor","independent floor"
    ]
    if any(x in v for x in commercial):
        return "COMMERCIAL"
    if any(x in v for x in residential):
        return "RESIDENTIAL"
    return "UNKNOWN"


def _canonical_city(value):
    v=_norm(value)
    if not v:
        return None
    if "delhi ncr" in v or v=="ncr" or "national capital region" in v:
        return "NCR"
    if "gurgaon" in v or "gurugram" in v:
        return "GURUGRAM"
    if "greater noida" in v:
        return "GREATER_NOIDA"
    if "noida" in v:
        return "NOIDA"
    if "faridabad" in v:
        return "FARIDABAD"
    if "ghaziabad" in v:
        return "GHAZIABAD"
    if "new delhi" in v or v=="delhi" or v.endswith(" delhi") or v.startswith("delhi "):
        return "DELHI"
    return v.upper().replace(" ","_")


_SOUTH_DELHI_ALIASES={
    "south delhi","greater kailash","gk","gk 1","gk1","gk 2","gk2",
    "defence colony","defense colony","south extension","south ex",
    "hauz khas","green park","saket","vasant kunj","lajpat nagar",
    "new friends colony","nfc","kalkaji","nehru place","malviya nagar",
    "cr park","chittaranjan park","panchsheel","panchsheel park",
    "safdarjung","safdarjung enclave","vasant vihar","east of kailash",
    "greater kailash 1","greater kailash 2"
}


def _south_delhi_match(requirement_location, property_location):
    rq=_norm(requirement_location)
    pl=_norm(property_location)
    if "south delhi" not in rq:
        return False
    return any(alias in pl for alias in _SOUTH_DELHI_ALIASES if alias!="south delhi") or "south delhi" in pl


def _location_similarity(a,b):
    ta=_tokens(a); tb=_tokens(b)
    if not ta or not tb:
        return 0.0,[]
    overlap=ta & tb
    union=ta | tb
    score=len(overlap)/max(1,len(union))
    na=_norm(a); nb=_norm(b)
    if na and nb and (na in nb or nb in na):
        score=max(score,0.90)
    if _south_delhi_match(a,b):
        score=max(score,0.92)
    return score,sorted(overlap)


def _phone_key(value):
    digits="".join(ch for ch in str(value or "") if ch.isdigit())
    return digits[-10:] if len(digits)>=10 else digits


def _is_restaurant_requirement(q):
    text_value=" ".join(str(q.get(k) or "") for k in [
        "property_type","suitable_category","additional_points","requirement_type"
    ])
    v=_norm(text_value)
    return any(x in v for x in [
        "restaurant","food","f&b","fnb","qsr","cafe","café","lounge","club"
    ])


def _restaurant_compatible(p):
    text_value=" ".join(str(p.get(k) or "") for k in [
        "property_type","suitable_category","remarks"
    ])
    v=_norm(text_value)
    if _property_class(text_value)=="RESIDENTIAL":
        return False
    return any(x in v for x in [
        "restaurant","food","f&b","fnb","qsr","cafe","café","lounge","club",
        "retail","shop","showroom","commercial"
    ])


def _requirement_area_window(q):
    mn=_float(q.get("minimum_area_sqft"))
    mx=_float(q.get("maximum_area_sqft"))
    if mn is None and mx is None:
        return None,None,None
    if mn is not None and mx is not None:
        base_lo=min(mn,mx); base_hi=max(mn,mx); target=(base_lo+base_hi)/2.0
    else:
        target=mn if mn is not None else mx
        base_lo=target; base_hi=target
    return max(0.0,base_lo*0.80),base_hi*1.20,target


def _property_area(p):
    for key in ["available_area_sqft","area_sqft","maximum_area_sqft","minimum_area_sqft"]:
        value=_float(p.get(key))
        if value is not None and value>0:
            return value
    return None


def _budget_max(q):
    for key in ["budget_max","monthly_rent","maximum_rent","rent_budget"]:
        value=_float(q.get(key))
        if value is not None and value>0:
            return value
    return None


def _property_monthly_rent(p):
    value=_float(p.get("monthly_rent"))
    if value is not None and value>0:
        return value
    psf=_float(p.get("asking_rent_per_sqft"))
    area=_property_area(p)
    if psf is not None and area is not None and psf>0 and area>0:
        return psf*area
    return None


def _match_band(score):
    if score>=90: return "EXCELLENT"
    if score>=80: return "STRONG"
    if score>=70: return "GOOD"
    return "POSSIBLE"


def _area_match_score(target,area):
    if target is None or area is None or target<=0:
        return 10
    ratio=abs(area-target)/target
    if ratio<=0.05: return 25
    if ratio<=0.10: return 23
    if ratio<=0.15: return 21
    if ratio<=0.20: return 18
    return 0


def _location_match_score(q,p):
    q_city=_canonical_city(q.get("city"))
    p_city=_canonical_city(p.get("city"))
    q_loc=q.get("preferred_locations") or ""
    p_loc=" ".join(str(x or "") for x in [p.get("location"),p.get("micro_market")])
    city_points=0
    location_points=0
    if not q_city:
        city_points=5
    elif q_city=="NCR":
        if p_city in {"DELHI","GURUGRAM","NOIDA","GREATER_NOIDA","FARIDABAD","GHAZIABAD","NCR"}:
            city_points=10
    elif p_city==q_city:
        city_points=10
    if not q_loc:
        location_points=10
    elif _south_delhi_match(q_loc,p_loc):
        location_points=20
    else:
        sim,_overlap=_location_similarity(q_loc,p_loc)
        if sim>=0.85: location_points=20
        elif sim>=0.60: location_points=17
        elif sim>=0.30: location_points=12
        elif sim>0: location_points=6
    return min(30,city_points+location_points)


def _property_type_score(q,p):
    q_text=" ".join(str(q.get(k) or "") for k in ["property_type","suitable_category","additional_points"])
    p_text=" ".join(str(p.get(k) or "") for k in ["property_type","suitable_category","remarks"])
    q_class=_property_class(q_text)
    p_class=_property_class(p_text)
    qt=_type_family(q_text)
    pt=_type_family(p_text)
    if q_class!="UNKNOWN" and q_class==p_class:
        if qt and pt and qt & pt: return 15
        return 12
    if qt and pt and qt & pt: return 12
    if q_class=="UNKNOWN" or p_class=="UNKNOWN": return 6
    return 0


def _suitable_use_score(q,p):
    if _is_restaurant_requirement(q):
        return 15 if _restaurant_compatible(p) else 0
    q_cat=_type_family(q.get("suitable_category"))
    p_cat=_type_family(" ".join(str(x or "") for x in [
        p.get("suitable_category"),p.get("property_type"),p.get("remarks")
    ]))
    if not q_cat: return 8
    if q_cat and p_cat and q_cat & p_cat: return 15
    return 5 if not p_cat else 0


def _budget_score(q,p):
    maximum=_budget_max(q)
    rent=_property_monthly_rent(p)
    if maximum is None: return 5
    if rent is None: return 3
    ratio=rent/maximum if maximum else 999
    if ratio<=1.00: return 10
    if ratio<=1.05: return 9
    if ratio<=1.10: return 7
    if ratio<=1.20: return 4
    return 0


def _verification_score(p):
    return 5 if _norm(p.get("verification_status"))=="verified" else 1


def _hard_filter_property(q,p):
    exclusions=[]

    quality=_organize_property_v4(p)
    if not quality["match_eligible"]:
        exclusions.append("MATCH_DATA_INCOMPLETE")
    if p.get("v9_final_bucket"):
        if p.get("v9_final_bucket")!="MATCH_READY" or p.get("v9_match_eligible") is False:
            exclusions.append("NOT_MATCH_READY_V9")
    elif p.get("v8_bucket"):
        if p.get("v8_bucket")!="MATCH_READY" or p.get("v8_match_eligible") is False:
            exclusions.append("NOT_MATCH_READY_V8")
    elif p.get("matching_bucket") and p.get("matching_bucket")!="MATCH_READY":
        exclusions.append("NOT_MATCH_READY_V7")
    request_phones=_contact_number_set(q.get("contact_phone"))
    property_phones=_contact_number_set(
        p.get("owner_contact"),p.get("broker_contact"),p.get("contact_number"),
        p.get("owner_contact_normalized"),p.get("broker_contact_normalized"),p.get("general_contact_normalized")
    )
    if request_phones and property_phones and (request_phones & property_phones):
        exclusions.append("SELF_INVENTORY")

    availability=_norm(p.get("availability_status") or "available")
    if any(x in availability for x in [
        "unavailable","sold","leased","not available","inactive","removed","closed"
    ]):
        exclusions.append("NOT_AVAILABLE")

    q_tx=_transaction_family(q.get("rent_or_sale"))
    p_tx=_transaction_family(p.get("rent_or_sale"))
    p_tx_raw=_norm(p.get("rent_or_sale"))
    if q_tx and p_tx:
        sale_or_rent=any(x in p_tx_raw for x in ["sale or rent","rent or sale","sale/rent","rent/sale"])
        if q_tx!=p_tx and not sale_or_rent:
            exclusions.append("TRANSACTION_MISMATCH")

    q_class=_property_class(q.get("property_type"),q.get("suitable_category"),q.get("additional_points"),q.get("requirement_type"))
    p_class=_property_class(p.get("property_type"),p.get("suitable_category"),p.get("remarks"))
    if q_class!="UNKNOWN" and p_class!="UNKNOWN" and q_class!=p_class:
        exclusions.append("PROPERTY_CLASS_MISMATCH")

    if _is_restaurant_requirement(q) and not _restaurant_compatible(p):
        exclusions.append("USE_MISMATCH")

    q_city=_canonical_city(q.get("city"))
    p_city=_canonical_city(p.get("city"))
    if q_city and q_city!="NCR" and p_city and p_city!="NCR" and q_city!=p_city:
        exclusions.append("CITY_MISMATCH")

    q_loc=q.get("preferred_locations") or ""
    p_loc=" ".join(str(x or "") for x in [p.get("location"),p.get("micro_market")])
    if "south delhi" in _norm(q_loc) and p_loc and not _south_delhi_match(q_loc,p_loc):
        exclusions.append("LOCATION_MISMATCH")

    hard_lo,hard_hi,_target=_requirement_area_window(q)
    area=_property_area(p)
    if hard_lo is not None and hard_hi is not None and area is not None:
        if area<hard_lo or area>hard_hi:
            exclusions.append("AREA_OUTSIDE_20_PERCENT_RANGE")

    return exclusions


def robust_match_requirement(rid,create_whatsapp=False):
    with engine.begin() as c:
        qrow=c.execute(text("SELECT * FROM pi_requirements WHERE requirement_id=:id"),{"id":rid}).first()
        if not qrow:
            raise HTTPException(404,"Requirement not found")
        q=dict(qrow._mapping)
        all_rows=c.execute(text("SELECT * FROM pi_properties ORDER BY created_at DESC")).fetchall()
        all_count=len(all_rows)
        c.execute(text("DELETE FROM pi_matches WHERE requirement_id=:id"),{"id":rid})

        eligible=[]
        excluded=[]
        exclusion_counts={}
        _hard_lo,_hard_hi,target_area=_requirement_area_window(q)

        for row in all_rows:
            p=dict(row._mapping)
            exclusion_reasons=_hard_filter_property(q,p)
            if exclusion_reasons:
                for reason in exclusion_reasons:
                    exclusion_counts[reason]=exclusion_counts.get(reason,0)+1
                excluded.append({
                    "property_id":p.get("property_id"),
                    "property_name":p.get("property_name"),
                    "city":p.get("city"),
                    "location":p.get("location"),
                    "property_type":p.get("property_type"),
                    "available_area_sqft":_property_area(p),
                    "source":p.get("source"),
                    "reasons":exclusion_reasons
                })
                continue

            area=_property_area(p)
            breakdown={
                "location":_location_match_score(q,p),
                "area":_area_match_score(target_area,area),
                "property_type":_property_type_score(q,p),
                "suitable_use":_suitable_use_score(q,p),
                "budget":_budget_score(q,p),
                "verification":_verification_score(p)
            }
            score=round(sum(breakdown.values()),2)
            reasons=[]
            if breakdown["location"]>=22: reasons.append("Strong location")
            elif breakdown["location"]>=15: reasons.append("Location compatible")
            if breakdown["area"]>=21: reasons.append("Strong area fit")
            elif breakdown["area"]>=18: reasons.append("Area within 20% tolerance")
            if breakdown["property_type"]>=12: reasons.append("Property type compatible")
            if breakdown["suitable_use"]>=12: reasons.append("Suitable use compatible")
            if breakdown["budget"]>=7: reasons.append("Budget/rent compatible")
            if breakdown["verification"]==5: reasons.append("Verified")

            gaps=[]
            if not p.get("location"): gaps.append("Property location missing")
            if area is None: gaps.append("Property area missing")
            if not p.get("rent_or_sale"): gaps.append("Property Rent/Sale missing")
            if _property_class(p.get("property_type"),p.get("suitable_category"),p.get("remarks"))=="UNKNOWN":
                gaps.append("Property class unclear")
            if _norm(p.get("verification_status"))!="verified":
                gaps.append("Availability not verified")

            eligible.append({
                "property_id":p.get("property_id"),
                "property_name":p.get("property_name"),
                "city":p.get("city"),
                "location":p.get("location"),
                "micro_market":p.get("micro_market"),
                "available_area_sqft":area,
                "monthly_rent":_property_monthly_rent(p),
                "rent_or_sale":p.get("rent_or_sale"),
                "property_type":p.get("property_type"),
                "suitable_category":p.get("suitable_category"),
                "verification_status":p.get("verification_status"),
                "owner_name":p.get("owner_name"),
                "owner_contact":p.get("owner_contact_normalized") or p.get("owner_contact"),
                "broker_name":p.get("broker_name"),
                "broker_contact":p.get("broker_contact_normalized") or p.get("broker_contact"),
                "contact_number":p.get("general_contact_normalized") or p.get("contact_number"),
                "score":score,
                "match_band":_match_band(score),
                "score_breakdown":breakdown,
                "reasons":reasons,
                "gaps":gaps
            })

        eligible.sort(key=lambda x:(x["score"],1 if _norm(x.get("verification_status"))=="verified" else 0),reverse=True)
        matches=eligible[:50]

        for rank,x in enumerate(matches,1):
            sql_insert="INSERT INTO pi_matches(requirement_id,property_id,match_score,rank,match_reasons,status) VALUES(:r,:p,:s,:rank,CAST(:reason AS JSONB),:status)"
            c.execute(
                text(sql_insert),
                {
                    "r":rid,
                    "p":x["property_id"],
                    "s":x["score"],
                    "rank":rank,
                    "reason":json.dumps({
                        "reasons":x["reasons"],
                        "score_breakdown":x["score_breakdown"],
                        "match_band":x["match_band"]
                    }),
                    "status":"READY_FOR_REVIEW"
                }
            )

    if all_count==0:
        msg="No properties exist in the database. Add/upload inventory first."
    elif not matches:
        msg="No eligible property passed the mandatory hard filters. Review Excluded Inventory below."
    else:
        msg=f"{len(matches)} eligible properties found after hard filtering. Highest scores are shown first."

    diagnostic={
        "engine":"MATCHING_V2",
        "database_properties":int(all_count),
        "active_properties_checked":int(all_count),
        "eligible_count":len(eligible),
        "excluded_count":len(excluded),
        "matches_returned":len(matches),
        "fallback_low_confidence":False,
        "area_rule":"Default hard tolerance is 80%-120% of the requirement",
        "weights":{"location":30,"area":25,"property_type":15,"suitable_use":15,"budget":10,"verification":5},
        "exclusion_counts":exclusion_counts,
        "requirement":{
            "requirement_id":rid,
            "city":q.get("city"),
            "preferred_locations":q.get("preferred_locations"),
            "minimum_area_sqft":_float(q.get("minimum_area_sqft")),
            "maximum_area_sqft":_float(q.get("maximum_area_sqft")),
            "rent_or_sale":q.get("rent_or_sale"),
            "property_type":q.get("property_type"),
            "suitable_category":q.get("suitable_category")
        },
        "message":msg
    }

    draft=None
    if create_whatsapp and matches:
        with engine.connect() as c:
            property_map={}
            for x in matches[:10]:
                if _norm(x.get("verification_status"))!="verified":
                    continue
                rr=c.execute(text("SELECT * FROM pi_properties WHERE property_id=:id"),{"id":x["property_id"]}).first()
                if rr:
                    property_map[x["property_id"]]=dict(rr._mapping)
        tops=[property_map[x["property_id"]] for x in matches[:10] if x["property_id"] in property_map][:5]
        if tops:
            message,provider=generate_whatsapp_message(q,tops)
            did=store_whatsapp_draft(q,message,provider)
            draft={"id":did,"status":"READY_FOR_REVIEW","message":message,"generated_by":provider}

    return {
        "status":"READY_FOR_REVIEW",
        "engine":"MATCHING_V2",
        "matches":matches,
        "excluded":excluded[:200],
        "diagnostic":diagnostic,
        "whatsapp_draft":draft
    }

@app.post("/api/v4/properties/{property_id}/availability-verification")
def set_property_availability_verification(property_id:str, req:Request, status:str=Form(...)):
    need_login(req)
    status=str(status or "").strip().upper()
    if status not in {"VERIFIED","UNVERIFIED"}: raise HTTPException(400,"Invalid verification status")
    with engine.begin() as c:
        if not c.execute(text("SELECT 1 FROM pi_properties WHERE property_id=:id"),{"id":property_id}).first():
            raise HTTPException(404,"Property not found")
        c.execute(text("UPDATE pi_properties SET verification_status=:s, updated_at=NOW() WHERE property_id=:id"),
                  {"s":status,"id":property_id})
    return {"status":"OK","property_id":property_id,"verification_status":status}

def _property_db_html_value(value):
    if value is None:
        return ""
    if isinstance(value,(dict,list)):
        try:
            return json.dumps(value,ensure_ascii=False,default=str)
        except Exception:
            return str(value)
    return str(value)


@app.get("/property-database",response_class=HTMLResponse)
def property_database_page(req:Request):
    role=page_role_or_redirect(req)
    if not role:
        return RedirectResponse("/login",status_code=303)

    with engine.connect() as c:
        total=int(c.execute(text("SELECT COUNT(*) FROM pi_properties")).scalar_one())
        verified=int(c.execute(text("SELECT COUNT(*) FROM pi_properties WHERE UPPER(COALESCE(verification_status,''))='VERIFIED'")).scalar_one())
        available=int(c.execute(text("""SELECT COUNT(*) FROM pi_properties
            WHERE UPPER(TRIM(COALESCE(availability_status,'AVAILABLE')))
            NOT IN ('UNAVAILABLE','LEASED','SOLD','INACTIVE','REMOVED','NOT AVAILABLE')""")).scalar_one())
        rows=c.execute(text("""SELECT *
            FROM pi_properties
            ORDER BY created_at DESC
            LIMIT 5000""")).fetchall()

    records=[dict(r._mapping) for r in rows]
    row_html=[]
    for p in records:
        pid=_property_db_html_value(p.get("property_id"))
        pname=_property_db_html_value(p.get("property_name")) or pid
        city=_property_db_html_value(p.get("city"))
        loc=_property_db_html_value(p.get("location"))
        ptype=_property_db_html_value(p.get("property_type"))
        area=_property_db_html_value(p.get("available_area_sqft"))
        rent=_property_db_html_value(p.get("monthly_rent") or p.get("asking_rent_per_sqft"))
        availability=_property_db_html_value(p.get("availability_status"))
        verification=_property_db_html_value(p.get("verification_status"))
        owner=_property_db_html_value(p.get("owner_name"))
        owner_phone=_property_db_html_value(p.get("owner_contact"))
        broker=_property_db_html_value(p.get("broker_name"))
        broker_phone=_property_db_html_value(p.get("broker_contact"))
        source=_property_db_html_value(p.get("source"))
        created=_property_db_html_value(p.get("created_at"))
        search_blob=" ".join([pid,pname,city,loc,ptype,availability,verification,owner,owner_phone,broker,broker_phone,source]).lower()
        row_html.append(
            f"""<tr data-search="{escape(search_blob)}">
            <td><a class="recordlink" href="/property-record/{quote_plus(pid)}"><b>{escape(pname)}</b><br><small>{escape(pid)}</small></a></td>
            <td>{escape(city)}</td><td>{escape(loc)}</td><td>{escape(ptype)}</td>
            <td>{escape(area)}</td><td>{escape(rent)}</td>
            <td>{escape(availability)}</td><td>{escape(verification)}</td>
            <td>{escape(owner)}<br><small>{escape(owner_phone)}</small></td>
            <td>{escape(broker)}<br><small>{escape(broker_phone)}</small></td>
            <td>{escape(source)}</td><td>{escape(created)}</td>
            <td><a class="btn" href="/property-record/{quote_plus(pid)}">View Full Property</a></td>
            </tr>"""
        )

    return HTMLResponse(f"""<!doctype html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Full Property Database</title>
<style>
*{{box-sizing:border-box}}body{{margin:0;background:#f4f7fb;font-family:Arial,sans-serif;color:#142033}}
header{{background:#0d1d2d;color:white;padding:16px 22px;display:flex;justify-content:space-between;align-items:center;gap:12px;flex-wrap:wrap}}
.wrap{{padding:20px;max-width:1800px;margin:auto}}.kpis{{display:grid;grid-template-columns:repeat(3,minmax(160px,1fr));gap:12px;margin:16px 0}}
.kpi,.card{{background:white;border:1px solid #e4eaf1;border-radius:12px;padding:15px}}.kpi b{{font-size:28px;display:block;margin-top:6px}}
.toolbar{{display:flex;gap:9px;flex-wrap:wrap;align-items:center;margin:12px 0}}input{{padding:11px;border:1px solid #ccd7e4;border-radius:8px;min-width:300px;flex:1}}
.btn{{display:inline-block;background:#1677ff;color:white;padding:8px 11px;border-radius:7px;text-decoration:none;font-weight:700;font-size:12px}}
.btn.gray{{background:#edf2f7;color:#24364b}}.tablewrap{{overflow:auto;max-height:70vh;border:1px solid #e4eaf1;border-radius:10px}}
table{{width:100%;border-collapse:collapse;background:white;font-size:12px}}th,td{{padding:9px;border-bottom:1px solid #edf1f5;text-align:left;white-space:nowrap;vertical-align:top}}
th{{position:sticky;top:0;background:#f8fafc;z-index:1}}.recordlink{{color:#135fb8;text-decoration:none}}small{{color:#6d7b90}}
@media(max-width:760px){{.kpis{{grid-template-columns:1fr}}input{{min-width:100%}}}}
</style></head>
<body>
<header><div><b>Full Property Database</b><br><small style="color:#b9c8d8">Master inventory saved in Property Intelligence</small></div>
<div><a class="btn gray" href="/workspace">Back to Workspace</a> <a class="btn gray" href="/data-quality">Data Quality</a> <a class="btn" href="/property-manual">Add Property</a></div></header>
<div class="wrap">
<div class="kpis">
<div class="kpi"><span>TOTAL PROPERTIES</span><b id="totalCount">{total}</b></div>
<div class="kpi"><span>AVAILABLE</span><b>{available}</b></div>
<div class="kpi"><span>VERIFIED</span><b>{verified}</b></div>
</div>
<div class="card">
<div class="toolbar"><input id="propertySearch" placeholder="Search Property ID, name, city, location, type, owner, broker, phone or source">
<span id="visibleCount"><b>{len(records)}</b> records shown</span></div>
<div class="tablewrap"><table>
<thead><tr><th>Property / ID</th><th>City</th><th>Location</th><th>Type</th><th>Area</th><th>Rent</th><th>Availability</th><th>Verification</th><th>Owner</th><th>Broker</th><th>Source</th><th>Saved</th><th>Action</th></tr></thead>
<tbody id="propertyRows">{''.join(row_html)}</tbody>
</table></div></div></div>
<script>
const s=document.getElementById('propertySearch'),rows=[...document.querySelectorAll('#propertyRows tr')],count=document.getElementById('visibleCount');
function filterRows(){{const q=(s.value||'').toLowerCase().trim();let n=0;rows.forEach(r=>{{let show=!q||(r.dataset.search||'').includes(q);r.style.display=show?'':'none';if(show)n++}});count.innerHTML='<b>'+n+'</b> records shown'}}
s.addEventListener('input',filterRows);
</script></body></html>""")


@app.get("/property-record/{property_id}",response_class=HTMLResponse)
def property_record_page(property_id:str,req:Request):
    role=page_role_or_redirect(req)
    if not role:
        return RedirectResponse("/login",status_code=303)

    with engine.connect() as c:
        row=c.execute(text("SELECT * FROM pi_properties WHERE property_id=:id"),{"id":property_id}).first()
        if not row:
            raise HTTPException(404,"Property not found")
        p=dict(row._mapping)

        media=[]
        for table_name in ["pi_property_media","pi_media"]:
            try:
                rr=c.execute(text(f"SELECT * FROM {table_name} WHERE property_id=:id ORDER BY created_at DESC"),{"id":property_id}).fetchall()
                media.extend([dict(x._mapping) for x in rr])
            except Exception:
                pass

        try:
            verification=[dict(x._mapping) for x in c.execute(
                text("SELECT * FROM pi_verification_log WHERE property_id=:id ORDER BY created_at DESC LIMIT 100"),
                {"id":property_id}
            ).fetchall()]
        except Exception:
            verification=[]

        try:
            matches=[dict(x._mapping) for x in c.execute(
                text("""SELECT m.*,r.client_name,r.company_name,r.city AS requirement_city,
                    r.preferred_locations,r.minimum_area_sqft,r.maximum_area_sqft
                    FROM pi_matches m
                    LEFT JOIN pi_requirements r ON r.requirement_id=m.requirement_id
                    WHERE m.property_id=:id ORDER BY m.created_at DESC LIMIT 100"""),
                {"id":property_id}
            ).fetchall()]
        except Exception:
            matches=[]

    preferred_order=[
        "property_id","property_name","property_type","entry_status","availability_status",
        "verification_status","city","location","micro_market","address","google_maps_pin",
        "area_sqft","available_area_sqft","minimum_area_sqft","maximum_area_sqft","floor",
        "rent_or_sale","monthly_rent","asking_rent_per_sqft","asking_sale_price","possession",
        "nearby_brands","suitable_category","parking","ceiling_height","power_load",
        "cam_per_sqft","security_deposit","frontage","assigned_to","verified_date","verified_by",
        "remarks","source","source_id","extraction_confidence","created_at","updated_at"
    ]
    private_order=[
        "owner_name","owner_contact","owner_email","broker_name","broker_contact",
        "broker_email","broker_company","contact_number"
    ]

    def field_rows(keys):
        out=[]
        for key in keys:
            if key in p:
                val=_property_db_html_value(p.get(key))
                out.append(f"<tr><th>{escape(key.replace('_',' ').title())}</th><td>{escape(val)}</td></tr>")
        return "".join(out)

    used=set(preferred_order+private_order)
    other_keys=[k for k in p.keys() if k not in used and k not in {"fingerprint"}]
    general_html=field_rows(preferred_order+other_keys)
    private_html=field_rows(private_order)

    media_html=[]
    for m in media:
        mid=_property_db_html_value(m.get("media_id") or m.get("id"))
        filename=_property_db_html_value(m.get("filename") or m.get("title"))
        mtype=_property_db_html_value(m.get("media_type"))
        url=_property_db_html_value(m.get("url"))
        if url:
            media_html.append(f'<div class="media"><b>{escape(mtype)}</b><br>{escape(filename)}<br><a target="_blank" href="{escape(url)}">Open Media</a></div>')
        else:
            media_html.append(f'<div class="media"><b>{escape(mtype)}</b><br>{escape(filename)}<br><small>Stored media ID: {escape(mid)}</small></div>')

    verification_html="".join(
        f"<tr><td>{escape(_property_db_html_value(v.get('created_at')))}</td><td>{escape(_property_db_html_value(v.get('action')))}</td><td>{escape(_property_db_html_value(v.get('performed_by')))}</td><td>{escape(_property_db_html_value(v.get('notes')))}</td></tr>"
        for v in verification
    ) or '<tr><td colspan="4">No verification history saved.</td></tr>'

    matches_html="".join(
        f"<tr><td>{escape(_property_db_html_value(m.get('requirement_id')))}</td><td>{escape(_property_db_html_value(m.get('company_name') or m.get('client_name')))}</td><td>{escape(_property_db_html_value(m.get('requirement_city')))}</td><td>{escape(_property_db_html_value(m.get('preferred_locations')))}</td><td>{escape(_property_db_html_value(m.get('match_score')))}</td><td>{escape(_property_db_html_value(m.get('status')))}</td></tr>"
        for m in matches
    ) or '<tr><td colspan="6">This property has no saved match history.</td></tr>'

    title=_property_db_html_value(p.get("property_name")) or property_id
    return HTMLResponse(f"""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{escape(title)} - Property Record</title>
<style>
*{{box-sizing:border-box}}body{{margin:0;background:#f4f7fb;font-family:Arial,sans-serif;color:#142033}}header{{background:#0d1d2d;color:white;padding:16px 22px;display:flex;justify-content:space-between;gap:12px;align-items:center;flex-wrap:wrap}}
.wrap{{max-width:1500px;margin:auto;padding:20px}}.card{{background:white;border:1px solid #e4eaf1;border-radius:12px;padding:16px;margin:14px 0}}.grid2{{display:grid;grid-template-columns:1fr 1fr;gap:14px}}
table{{width:100%;border-collapse:collapse;font-size:13px}}th,td{{padding:9px;border-bottom:1px solid #edf1f5;text-align:left;vertical-align:top}}th{{width:240px;background:#fafbfd}}
.btn{{display:inline-block;background:#1677ff;color:white;padding:9px 12px;border-radius:7px;text-decoration:none;font-weight:700}}.btn.gray{{background:#edf2f7;color:#24364b}}.private{{border-left:5px solid #df8b13}}.mediaGrid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(180px,1fr));gap:10px}}.media{{border:1px solid #e4eaf1;border-radius:9px;padding:10px}}
.tablewrap{{overflow:auto}}.badge{{display:inline-block;background:#eaf8f2;color:#086d49;padding:5px 8px;border-radius:20px;font-weight:700}}
@media(max-width:900px){{.grid2{{grid-template-columns:1fr}}}}
</style></head><body>
<header><div><b>{escape(title)}</b><br><small>{escape(property_id)}</small></div>
<div><a class="btn gray" href="/workspace">Workspace</a> <a class="btn" href="/property-database">Full Property Database</a></div></header>
<div class="wrap">
<div class="card"><h2>Property Record</h2><span class="badge">{escape(_property_db_html_value(p.get("verification_status") or "UNVERIFIED"))}</span>
<div class="tablewrap"><table>{general_html}</table></div></div>
<div class="card private"><h2>Private / Internal Contact Details</h2><p>For team verification and follow-up. Do not automatically include these contacts in client-facing WhatsApp drafts.</p>
<div class="tablewrap"><table>{private_html or '<tr><td>No owner/broker contact details saved.</td></tr>'}</table></div></div>
<div class="card"><h2>Property Media</h2><div class="mediaGrid">{''.join(media_html) or 'No media saved for this property.'}</div></div>
<div class="grid2">
<div class="card"><h2>Verification History</h2><div class="tablewrap"><table><thead><tr><th>Date</th><th>Action</th><th>By</th><th>Notes</th></tr></thead><tbody>{verification_html}</tbody></table></div></div>
<div class="card"><h2>Match History</h2><div class="tablewrap"><table><thead><tr><th>Requirement</th><th>Client</th><th>City</th><th>Location</th><th>Score</th><th>Status</th></tr></thead><tbody>{matches_html}</tbody></table></div></div>
</div></div></body></html>""")


# ============================================================
# DATA ORGANIZER V4
# Deterministic cleanup only. Never invents missing phone digits.
# ============================================================

_PHONE_INVALID_MARKERS={"","na","n/a","none","null","unknown","not available","not specified","-"}

def _dq_text(value):
    return str(value or "").strip()

def _dq_unknown(value):
    return _dq_text(value).lower() in _PHONE_INVALID_MARKERS

def _extract_contact_numbers(raw):
    s=_dq_text(raw)
    if not s or _dq_unknown(s):
        return {"mobiles":[],"landlines":[],"invalid_raw":s}

    mobiles=[]
    landlines=[]

    mobile_pat=re.compile(r'(?<!\d)(?:\+?91[\s\-]*)?([6-9](?:[\s\-]*\d){9})(?!\d)')
    for m in mobile_pat.finditer(s):
        digits=re.sub(r'\D','',m.group(1))
        if len(digits)==10 and digits[0] in "6789" and digits not in mobiles:
            mobiles.append(digits)

    land_pat=re.compile(r'(?<!\d)(0\d(?:[\s\-]*\d){8,9})(?!\d)')
    for m in land_pat.finditer(s):
        digits=re.sub(r'\D','',m.group(1))
        if len(digits) in (10,11) and digits.startswith("0") and digits not in landlines:
            landlines.append(digits)

    return {"mobiles":mobiles,"landlines":landlines,"invalid_raw":s}

def _contact_display(info):
    values=list(info.get("mobiles") or [])+list(info.get("landlines") or [])
    return ", ".join(values) if values else None

def _contact_number_set(*values):
    out=set()
    for value in values:
        info=_extract_contact_numbers(value)
        out.update(info["mobiles"])
        out.update(info["landlines"])
    return out

def _canonical_city_v4(value):
    v=_norm(value)
    if not v or v in {"na","n a","unknown","not specified","none"}:
        return "UNKNOWN"
    if "new delhi" in v or v=="delhi" or v.endswith(" delhi") or v.startswith("delhi "):
        return "DELHI"
    if "gurgaon" in v or "gurugram" in v:
        return "GURUGRAM"
    if "greater noida" in v:
        return "GREATER NOIDA"
    if "noida" in v:
        return "NOIDA"
    if "faridabad" in v:
        return "FARIDABAD"
    if "ghaziabad" in v:
        return "GHAZIABAD"
    return _dq_text(value).upper()

def _canonical_property_type_v4(value, suitable=None, remarks=None):
    raw=" ".join(_dq_text(x) for x in [value,suitable,remarks]).lower()
    raw=re.sub(r'\s+',' ',raw).strip()
    if not raw or raw in {"na","n/a","unknown","not specified"}:
        return "UNKNOWN"
    if any(x in raw for x in ["residential","3bhk","4bhk","2bhk","1bhk","apartment","flat","villa","builder floor","house"]):
        return "RESIDENTIAL"
    if any(x in raw for x in ["industrial","factory","warehouse","mohan co-operative","industrial area"]):
        return "INDUSTRIAL"
    if any(x in raw for x in ["office","cowork","business centre"]):
        return "OFFICE"
    if any(x in raw for x in ["shop","retail","showroom","commercial center","commercial centre"]):
        return "RETAIL/COMMERCIAL"
    if any(x in raw for x in ["restaurant","cafe","café","qsr","food","f&b","fnb","lounge","club"]):
        return "F&B/COMMERCIAL"
    if any(x in raw for x in ["banquet","hotel","guest house","hospitality","farmhouse"]):
        return "HOSPITALITY"
    if any(x in raw for x in ["mix land","mixland","mixed use","commercial/residential","commercial residential"]):
        return "MIXED USE"
    if "commercial" in raw or "comercial" in raw:
        return "COMMERCIAL"
    if re.fullmatch(r'[\s+/&\-]*(bmt|basement|gf|ff|sf|tf|terr|floor|unit)[\s+/&\-\w]*',raw,re.I):
        return "UNKNOWN"
    return _dq_text(value).upper()

def _canonical_transaction_v4(value):
    v=_norm(value)
    if not v or v in {"na","n a","unknown","not specified"}:
        return "UNKNOWN"
    has_rent=any(x in v for x in ["rent","lease","leasing"])
    has_sale=any(x in v for x in ["sale","sell","selling"])
    if has_rent and has_sale:
        return "RENT/SALE"
    if has_rent:
        return "RENT"
    if has_sale:
        return "SALE"
    return "UNKNOWN"

def _organize_property_v4(p):
    owner_info=_extract_contact_numbers(p.get("owner_contact"))
    broker_info=_extract_contact_numbers(p.get("broker_contact"))
    general_info=_extract_contact_numbers(p.get("contact_number"))

    valid_contacts=[]
    for info in [owner_info,broker_info,general_info]:
        valid_contacts.extend(info["mobiles"])
        valid_contacts.extend(info["landlines"])
    valid_contacts=list(dict.fromkeys(valid_contacts))

    city=_canonical_city_v4(p.get("city"))
    ptype=_canonical_property_type_v4(p.get("property_type"),p.get("suitable_category"),p.get("remarks"))
    transaction=_canonical_transaction_v4(p.get("rent_or_sale"))
    location=_dq_text(p.get("location") or p.get("micro_market"))
    area=_property_area(p)

    issues=[]
    if city=="UNKNOWN":
        issues.append("MISSING_OR_UNKNOWN_CITY")
    if not location or _dq_unknown(location):
        issues.append("MISSING_OR_UNKNOWN_LOCATION")
    if ptype=="UNKNOWN":
        issues.append("MISSING_OR_AMBIGUOUS_PROPERTY_TYPE")
    if area is None or area<=0:
        issues.append("MISSING_AREA")
    if transaction=="UNKNOWN":
        issues.append("MISSING_TRANSACTION")

    raw_contacts=[
        _dq_text(p.get("owner_contact")),
        _dq_text(p.get("broker_contact")),
        _dq_text(p.get("contact_number"))
    ]
    raw_contacts=[x for x in raw_contacts if x and not _dq_unknown(x)]
    if raw_contacts and not valid_contacts:
        issues.append("INVALID_OR_TRUNCATED_CONTACT")
    elif not raw_contacts:
        issues.append("MISSING_CONTACT")

    for raw in raw_contacts:
        digits=re.sub(r'\D','',raw)
        parsed=_extract_contact_numbers(raw)
        if digits and not parsed["mobiles"] and not parsed["landlines"]:
            if "INVALID_OR_TRUNCATED_CONTACT" not in issues:
                issues.append("INVALID_OR_TRUNCATED_CONTACT")

    match_eligible=not any(x in issues for x in [
        "MISSING_OR_UNKNOWN_CITY","MISSING_OR_UNKNOWN_LOCATION",
        "MISSING_OR_AMBIGUOUS_PROPERTY_TYPE","MISSING_AREA","MISSING_TRANSACTION"
    ])
    contact_ready=bool(valid_contacts)

    if match_eligible and contact_ready:
        status="READY"
    elif match_eligible:
        status="MATCH_READY_CONTACT_REVIEW"
    else:
        status="NEEDS_REVIEW"

    return {
        "canonical_city":city,
        "canonical_property_type":ptype,
        "canonical_transaction":transaction,
        "owner_contact_normalized":_contact_display(owner_info),
        "broker_contact_normalized":_contact_display(broker_info),
        "general_contact_normalized":_contact_display(general_info),
        "normalized_contacts":valid_contacts,
        "quality_issues":issues,
        "data_quality_status":status,
        "match_eligible":match_eligible,
        "contact_ready":contact_ready
    }

def _ensure_data_quality_columns():
    statements=[
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS raw_owner_contact TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS raw_broker_contact TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS raw_contact_number TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS owner_contact_normalized TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS broker_contact_normalized TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS general_contact_normalized TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS normalized_contacts JSONB DEFAULT '[]'::jsonb",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS canonical_city TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS canonical_property_type TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS canonical_transaction TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS data_quality_status TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS match_eligible BOOLEAN",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS contact_ready BOOLEAN",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS quality_issues JSONB DEFAULT '[]'::jsonb"
    ]
    with engine.begin() as c:
        for stmt in statements:
            c.execute(text(stmt))

def _audit_property_database_v4(apply_changes=False):
    _ensure_data_quality_columns()
    summary={
        "total":0,"ready":0,"match_ready_contact_review":0,"needs_review":0,
        "match_eligible":0,"contact_ready":0,"invalid_or_truncated_contact":0,
        "missing_contact":0,"missing_location":0,"missing_area":0,
        "missing_type":0,"missing_city":0,"missing_transaction":0
    }
    reviewed=[]
    with engine.begin() as c:
        rows=c.execute(text("SELECT * FROM pi_properties ORDER BY created_at DESC")).fetchall()
        for row in rows:
            p=dict(row._mapping)
            q=_organize_property_v4(p)
            summary["total"]+=1
            if q["data_quality_status"]=="READY": summary["ready"]+=1
            elif q["data_quality_status"]=="MATCH_READY_CONTACT_REVIEW": summary["match_ready_contact_review"]+=1
            else: summary["needs_review"]+=1
            if q["match_eligible"]: summary["match_eligible"]+=1
            if q["contact_ready"]: summary["contact_ready"]+=1
            issues=q["quality_issues"]
            if "INVALID_OR_TRUNCATED_CONTACT" in issues: summary["invalid_or_truncated_contact"]+=1
            if "MISSING_CONTACT" in issues: summary["missing_contact"]+=1
            if "MISSING_OR_UNKNOWN_LOCATION" in issues: summary["missing_location"]+=1
            if "MISSING_AREA" in issues: summary["missing_area"]+=1
            if "MISSING_OR_AMBIGUOUS_PROPERTY_TYPE" in issues: summary["missing_type"]+=1
            if "MISSING_OR_UNKNOWN_CITY" in issues: summary["missing_city"]+=1
            if "MISSING_TRANSACTION" in issues: summary["missing_transaction"]+=1

            if len(reviewed)<1000 and issues:
                reviewed.append({
                    "property_id":p.get("property_id"),
                    "property_name":p.get("property_name"),
                    "city":p.get("city"),
                    "location":p.get("location"),
                    "property_type":p.get("property_type"),
                    "area":_property_area(p),
                    "owner_name":p.get("owner_name"),
                    "owner_contact_raw":p.get("owner_contact"),
                    "broker_name":p.get("broker_name"),
                    "broker_contact_raw":p.get("broker_contact"),
                    "normalized_contacts":q["normalized_contacts"],
                    "status":q["data_quality_status"],
                    "issues":issues
                })

            if apply_changes:
                c.execute(text("""UPDATE pi_properties SET
                    raw_owner_contact=COALESCE(raw_owner_contact,owner_contact),
                    raw_broker_contact=COALESCE(raw_broker_contact,broker_contact),
                    raw_contact_number=COALESCE(raw_contact_number,contact_number),
                    owner_contact_normalized=:ocn,
                    broker_contact_normalized=:bcn,
                    general_contact_normalized=:gcn,
                    normalized_contacts=CAST(:contacts AS JSONB),
                    canonical_city=:city,
                    canonical_property_type=:ptype,
                    canonical_transaction=:tx,
                    data_quality_status=:status,
                    match_eligible=:me,
                    contact_ready=:cr,
                    quality_issues=CAST(:issues AS JSONB),
                    updated_at=NOW()
                    WHERE property_id=:id"""),{
                        "ocn":q["owner_contact_normalized"],
                        "bcn":q["broker_contact_normalized"],
                        "gcn":q["general_contact_normalized"],
                        "contacts":json.dumps(q["normalized_contacts"]),
                        "city":q["canonical_city"],
                        "ptype":q["canonical_property_type"],
                        "tx":q["canonical_transaction"],
                        "status":q["data_quality_status"],
                        "me":q["match_eligible"],
                        "cr":q["contact_ready"],
                        "issues":json.dumps(q["quality_issues"]),
                        "id":p.get("property_id")
                    })
    return {"summary":summary,"reviewed":reviewed}

@app.get("/api/v4/data-quality/summary")
def data_quality_summary_v4(req:Request):
    need_login(req)
    return {"status":"ok",**_audit_property_database_v4(False)}

@app.post("/api/v4/data-quality/organize")
def data_quality_organize_v4(req:Request):
    need_login(req)
    result=_audit_property_database_v4(True)
    return {"status":"ok","message":"Database organized safely. Original contacts preserved; no missing digits were guessed.",**result}

@app.get("/data-quality",response_class=HTMLResponse)
def data_quality_page_v4(req:Request):
    role=page_role_or_redirect(req)
    if not role:
        return RedirectResponse("/login",status_code=303)
    return HTMLResponse("""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Property Data Quality</title><style>
*{box-sizing:border-box}body{margin:0;background:#f4f7fb;font-family:Arial,sans-serif;color:#142033}
header{background:#0d1d2d;color:#fff;padding:16px 22px;display:flex;justify-content:space-between;align-items:center;gap:12px;flex-wrap:wrap}
.wrap{max-width:1700px;margin:auto;padding:20px}.kpis{display:grid;grid-template-columns:repeat(4,minmax(150px,1fr));gap:10px}.kpi,.card{background:#fff;border:1px solid #e4eaf1;border-radius:12px;padding:14px;margin-bottom:12px}.kpi b{font-size:25px;display:block;margin-top:5px}
.btn{border:0;border-radius:8px;padding:10px 13px;background:#1677ff;color:white;font-weight:700;cursor:pointer;text-decoration:none;display:inline-block}.btn.orange{background:#df8b13}.btn.gray{background:#edf2f7;color:#24364b}
.tablewrap{overflow:auto;max-height:65vh;border:1px solid #e4eaf1;border-radius:10px}table{width:100%;border-collapse:collapse;background:#fff;font-size:12px}th,td{padding:9px;border-bottom:1px solid #edf1f5;text-align:left;white-space:nowrap;vertical-align:top}th{position:sticky;top:0;background:#f8fafc}.warn{color:#a35c00;font-weight:700}.good{color:#08734b;font-weight:700}
@media(max-width:900px){.kpis{grid-template-columns:repeat(2,1fr)}}
</style></head><body><header><div><b>Property Data Organizer V4</b><br><small>Safe cleanup · multiple contacts · no guessed phone digits</small></div><div><a class="btn gray" href="/workspace">Workspace</a> <a class="btn" href="/property-database">Property Database</a></div></header>
<div class="wrap"><div class="card"><b>Rule:</b> 10-digit Indian mobiles are normalized. Multiple valid numbers are preserved separately. Full landlines are preserved. Truncated numbers such as 98105 remain in the RAW field and are marked for review. We never invent missing digits.</div>
<div class="card"><button class="btn orange" onclick="organize()">Organize Entire Database</button> <button class="btn" onclick="load()">Refresh Audit</button><span id="msg"></span></div>
<div class="kpis" id="kpis"></div>
<div class="card"><h3>Records Needing Review</h3><div class="tablewrap"><table><thead><tr><th>Property</th><th>City</th><th>Location</th><th>Type</th><th>Area</th><th>Owner / Raw Contact</th><th>Broker / Raw Contact</th><th>Valid Contacts Found</th><th>Status</th><th>Issues</th><th>Open</th></tr></thead><tbody id="rows"></tbody></table></div></div></div>
<script>
const esc=x=>String(x??'').replace(/[&<>"']/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[m]));
async function call(u,o={}){let r=await fetch(u,o),d=await r.json();if(!r.ok)throw new Error(d.detail||d.message||'Request failed');return d}
async function load(){let d=await call('/api/v4/data-quality/summary'),s=d.summary;document.querySelector('#kpis').innerHTML=[
['TOTAL',s.total],['READY',s.ready],['MATCH READY / CONTACT REVIEW',s.match_ready_contact_review],['NEEDS REVIEW',s.needs_review],
['MATCH ELIGIBLE',s.match_eligible],['CONTACT READY',s.contact_ready],['INVALID / TRUNCATED CONTACT',s.invalid_or_truncated_contact],['MISSING CONTACT',s.missing_contact],
['MISSING LOCATION',s.missing_location],['MISSING AREA',s.missing_area],['MISSING TYPE',s.missing_type],['MISSING TRANSACTION',s.missing_transaction]
].map(x=>`<div class="kpi"><span>${x[0]}</span><b>${Number(x[1]||0).toLocaleString()}</b></div>`).join('');
document.querySelector('#rows').innerHTML=(d.reviewed||[]).map(x=>`<tr><td><b>${esc(x.property_name||x.property_id)}</b><br>${esc(x.property_id)}</td><td>${esc(x.city||'')}</td><td>${esc(x.location||'')}</td><td>${esc(x.property_type||'')}</td><td>${esc(x.area||'')}</td><td>${esc(x.owner_name||'')}<br>${esc(x.owner_contact_raw||'')}</td><td>${esc(x.broker_name||'')}<br>${esc(x.broker_contact_raw||'')}</td><td>${esc((x.normalized_contacts||[]).join(', '))}</td><td class="${x.status==='READY'?'good':'warn'}">${esc(x.status)}</td><td>${esc((x.issues||[]).join(', '))}</td><td><a href="/property-record/${encodeURIComponent(x.property_id)}" target="_blank">View</a></td></tr>`).join('')}
async function organize(){if(!confirm('Organize all properties now? Original contacts will be preserved. Missing digits will NOT be guessed.'))return;document.querySelector('#msg').textContent=' Organizing...';try{let d=await call('/api/v4/data-quality/organize',{method:'POST'});document.querySelector('#msg').textContent=' '+d.message;await load()}catch(e){document.querySelector('#msg').textContent=' '+e.message}}
load();
</script></body></html>""")


# ============================================================
# DATA ORGANIZER V5 - INTELLIGENT RECOVERY + RETAIL LINKEDIN DEMAND
# Conservative recovery. Never invents phone digits or unsupported facts.
# ============================================================

_V5_LOCALITY_MAP = {
    "greater kailash":"Greater Kailash","gk 1":"Greater Kailash 1","gk1":"Greater Kailash 1",
    "gk 2":"Greater Kailash 2","gk2":"Greater Kailash 2","kailash colony":"Kailash Colony",
    "east of kailash":"East of Kailash","new friends colony":"New Friends Colony",
    "nfc":"New Friends Colony","safdarjung enclave":"Safdarjung Enclave",
    "safdurjung enclave":"Safdarjung Enclave","green park":"Green Park",
    "vasant vihar":"Vasant Vihar","vasant kunj":"Vasant Kunj","hauz khas":"Hauz Khas",
    "panchsheel park":"Panchsheel Park","panchsheel enclave":"Panchsheel Enclave",
    "defence colony":"Defence Colony","defense colony":"Defence Colony",
    "lajpat nagar":"Lajpat Nagar","jangpura":"Jangpura","nizamuddin":"Nizamuddin",
    "saket":"Saket","malviya nagar":"Malviya Nagar","chittaranjan park":"Chittaranjan Park",
    "cr park":"Chittaranjan Park","nehru place":"Nehru Place","okhla":"Okhla",
    "mathura road":"Mathura Road","mohan co-operative":"Mohan Co-operative",
    "mayapuri":"Mayapuri","pitampura":"Pitampura","sainik farm":"Sainik Farm",
    "gurgaon":"Gurugram","gurugram":"Gurugram","noida":"Noida","faridabad":"Faridabad","ghaziabad":"Ghaziabad"
}

def _v5_source_text(p):
    vals=[
        p.get("property_name"),p.get("property_type"),p.get("city"),p.get("location"),
        p.get("micro_market"),p.get("address"),p.get("remarks"),p.get("suitable_category"),
        p.get("nearby_brands"),p.get("source"),p.get("raw_owner_contact"),
        p.get("raw_broker_contact"),p.get("raw_contact_number")
    ]
    return " | ".join(str(x or "") for x in vals if str(x or "").strip())

def _v5_infer_location(text_value):
    v=_norm(text_value)
    hits=[]
    for key,label in _V5_LOCALITY_MAP.items():
        if key in v:
            hits.append((len(key),label))
    if not hits:
        return None,0
    hits.sort(reverse=True)
    return hits[0][1],95

def _v5_infer_city(text_value, location=None):
    v=_norm((text_value or "")+" "+(location or ""))
    if any(x in v for x in ["gurgaon","gurugram"]): return "Gurugram",98
    if "noida" in v: return "Noida",98
    if "faridabad" in v: return "Faridabad",98
    if "ghaziabad" in v: return "Ghaziabad",98
    if "delhi" in v: return "New Delhi",98
    known_delhi_localities={"Greater Kailash","Greater Kailash 1","Greater Kailash 2","Kailash Colony","East of Kailash","New Friends Colony","Safdarjung Enclave","Green Park","Vasant Vihar","Vasant Kunj","Hauz Khas","Panchsheel Park","Panchsheel Enclave","Defence Colony","Lajpat Nagar","Jangpura","Nizamuddin","Saket","Malviya Nagar","Chittaranjan Park","Nehru Place","Okhla","Mathura Road","Mohan Co-operative","Mayapuri","Pitampura","Sainik Farm"}
    if location in known_delhi_localities: return "New Delhi",96
    return None,0

def _v5_infer_type(text_value):
    v=_norm(text_value)
    rules=[
        (["restaurant","cafe","café","qsr","f&b","fnb","food outlet"],"F&B/COMMERCIAL",98),
        (["shop","showroom","retail","dda mkt","market shop"],"RETAIL/COMMERCIAL",96),
        (["office","business centre","cowork"],"OFFICE",96),
        (["industrial","factory","warehouse"],"INDUSTRIAL",96),
        (["banquet","hotel","guest house","hospitality","farmhouse"],"HOSPITALITY",94),
        (["1bhk","2bhk","3bhk","4bhk","5bhk","apartment","flat","villa","builder floor"],"RESIDENTIAL",96),
        (["commercial"],"COMMERCIAL",92),
    ]
    for terms,label,conf in rules:
        if any(t in v for t in terms):
            return label,conf
    return None,0

def _v5_infer_transaction(text_value):
    v=_norm(text_value)
    rent_terms=["for rent","on rent","to let","available for lease","available on lease","space for lease","offered for lease"]
    sale_terms=["for sale","on sale","available for sale","sale @","sale rs","asking sale","resale","sell","selling"]
    rent=any(x in v for x in rent_terms)
    sale=any(x in v for x in sale_terms)
    if rent and sale:return "RENT/SALE",98
    if rent:return "RENT",98
    if sale:return "SALE",98
    return None,0

def _v5_infer_area(text_value):
    raw=str(text_value or "")
    patterns=[
        (r'(?i)(\d[\d,]{1,6})\s*(?:sq\.?\s*ft|sqft|sft|square\s*feet|ft2)\b',1.0,"SQFT"),
        (r'(?i)(\d[\d,]{1,5})\s*(?:sq\.?\s*yd|sqyd|sq\.?\s*yard|yards?|yds?)\b',9.0,"SQYD"),
        (r'(?i)\b(\d[\d,]{2,6})\s*ft\b',1.0,"FT")
    ]
    for pat,mult,unit in patterns:
        m=_re.search(pat,raw)
        if m:
            try:
                val=float(m.group(1).replace(",",""))*mult
                if 50 <= val <= 500000:
                    return round(val,2),95,unit
            except Exception:
                pass
    return None,0,None

def _v5_recovery_suggestion(p):
    base=_organize_property_v4(p)
    txt=_v5_source_text(p)
    suggestions={}
    reasons=[]

    current_location=_dq_text(p.get("location") or p.get("micro_market"))
    if not current_location or _dq_unknown(current_location):
        loc,conf=_v5_infer_location(txt)
        if loc:
            suggestions["location"]={"value":loc,"confidence":conf,"reason":"Locality found in stored property/source text"}
            reasons.append("LOCATION_RECOVERABLE")

    current_city=_canonical_city_v4(p.get("city"))
    if current_city=="UNKNOWN":
        loc=(suggestions.get("location") or {}).get("value") or current_location
        city,conf=_v5_infer_city(txt,loc)
        if city:
            suggestions["city"]={"value":city,"confidence":conf,"reason":"City inferred from explicit locality/city text"}
            reasons.append("CITY_RECOVERABLE")

    current_type=_canonical_property_type_v4(p.get("property_type"),p.get("suitable_category"),p.get("remarks"))
    if current_type=="UNKNOWN":
        pt,conf=_v5_infer_type(txt)
        if pt:
            suggestions["property_type"]={"value":pt,"confidence":conf,"reason":"Explicit use/type keyword in stored text"}
            reasons.append("TYPE_RECOVERABLE")

    if _property_area(p) is None:
        area,conf,unit=_v5_infer_area(txt)
        if area and unit in {"SQFT","FT"}:
            suggestions["available_area_sqft"]={"value":area,"confidence":conf,"reason":f"Explicit {unit} usable/built-up area found in stored text"}
            reasons.append("AREA_RECOVERABLE")
        elif area and unit=="SQYD":
            reasons.append("PLOT_AREA_RECOVERABLE_REVIEW_ONLY")

    if _canonical_transaction_v4(p.get("rent_or_sale"))=="UNKNOWN":
        tx,conf=_v5_infer_transaction(txt)
        if tx:
            suggestions["rent_or_sale"]={"value":tx,"confidence":conf,"reason":"Explicit rent/lease/sale wording in stored text"}
            reasons.append("TRANSACTION_RECOVERABLE")

    high_conf={k:v for k,v in suggestions.items() if int(v.get("confidence") or 0)>=95}
    return {
        "property_id":p.get("property_id"),
        "property_name":p.get("property_name"),
        "source":p.get("source"),
        "before_status":base["data_quality_status"],
        "issues":base["quality_issues"],
        "suggestions":suggestions,
        "high_confidence_suggestions":high_conf,
        "recoverable":bool(high_conf),
        "reasons":reasons
    }

def _v5_duplicate_key(p):
    loc=_norm(p.get("location") or p.get("micro_market"))
    name=_norm(p.get("property_name"))
    area=_property_area(p)
    contacts=sorted(_contact_number_set(
        p.get("owner_contact"),p.get("broker_contact"),p.get("contact_number"),
        p.get("owner_contact_normalized"),p.get("broker_contact_normalized")
    ))
    contact=contacts[0] if contacts else ""
    if not (loc or name or contact): return None
    return hashlib.sha256(f"{loc}|{name}|{round(area or 0,1)}|{contact}".encode()).hexdigest()

def _v5_audit_recovery(limit=1000):
    rows=[]
    duplicate_groups={}
    counts={"total":0,"recoverable":0,"high_confidence_fields":0,"possible_duplicates":0}
    with engine.connect() as c:
        dbrows=c.execute(text("SELECT * FROM pi_properties ORDER BY created_at DESC")).fetchall()
    for row in dbrows:
        p=dict(row._mapping); counts["total"]+=1
        suggestion=_v5_recovery_suggestion(p)
        if suggestion["recoverable"]:
            counts["recoverable"]+=1
            counts["high_confidence_fields"]+=len(suggestion["high_confidence_suggestions"])
            if len(rows)<limit: rows.append(suggestion)
        k=_v5_duplicate_key(p)
        if k: duplicate_groups.setdefault(k,[]).append(p.get("property_id"))
    dups=[ids for ids in duplicate_groups.values() if len(ids)>1]
    counts["possible_duplicates"]=sum(len(x) for x in dups)
    return {"summary":counts,"recoverable":rows,"duplicate_groups":dups[:300]}

def _v5_apply_recovery():
    audit=_v5_audit_recovery(100000)
    updated_records=updated_fields=0
    with engine.begin() as c:
        for rec in audit["recoverable"]:
            sug=rec["high_confidence_suggestions"]
            if not sug: continue
            sets=[]; params={"id":rec["property_id"]}
            for field in ["city","location","property_type","available_area_sqft","rent_or_sale"]:
                if field in sug:
                    sets.append(f"{field}=:{field}")
                    params[field]=sug[field]["value"]
            if sets:
                sets.append("updated_at=NOW()")
                c.execute(text("UPDATE pi_properties SET "+",".join(sets)+" WHERE property_id=:id"),params)
                updated_records+=1; updated_fields+=len(sets)-1
    normalized=_audit_property_database_v4(True)
    return {"updated_records":updated_records,"updated_fields":updated_fields,"post_normalization":normalized["summary"]}

@app.get("/api/v5/data-recovery/audit")
def v5_data_recovery_audit(req:Request):
    need_login(req)
    return {"status":"ok",**_v5_audit_recovery()}

@app.post("/api/v5/data-recovery/apply")
def v5_data_recovery_apply(req:Request):
    need_login(req)
    result=_v5_apply_recovery()
    return {"status":"ok","message":"High-confidence recoveries applied. No phone digits or unsupported facts were invented.",**result}

@app.get("/data-recovery",response_class=HTMLResponse)
def v5_data_recovery_page(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",status_code=303)
    return HTMLResponse("""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Data Recovery V5</title>
<style>*{box-sizing:border-box}body{margin:0;background:#f4f7fb;font-family:Arial;color:#142033}header{background:#0d1d2d;color:#fff;padding:16px 22px}.wrap{max-width:1700px;margin:auto;padding:20px}.card,.kpi{background:#fff;border:1px solid #e4eaf1;border-radius:12px;padding:14px;margin-bottom:12px}.kpis{display:grid;grid-template-columns:repeat(4,1fr);gap:10px}.kpi b{font-size:26px;display:block}.btn{padding:10px 13px;border:0;border-radius:8px;background:#1677ff;color:white;font-weight:700;text-decoration:none;cursor:pointer}.orange{background:#df8b13}.tablewrap{overflow:auto;max-height:65vh}table{width:100%;border-collapse:collapse;font-size:12px}th,td{padding:8px;border-bottom:1px solid #edf1f5;text-align:left;vertical-align:top;white-space:nowrap}th{position:sticky;top:0;background:#f8fafc}</style></head>
<body><header><b>Data Organizer V5 · Intelligent Recovery</b></header><div class="wrap">
<div class="card"><a class="btn" href="/data-quality">Data Quality V4</a> <a class="btn" href="/property-database">Property Database</a> <a class="btn" href="/workspace">Workspace</a></div>
<div class="card"><b>Safety rule:</b> V5 applies only high-confidence facts explicitly recoverable from stored property/source text. It never completes truncated phone numbers and never defaults unknown transaction/location.</div>
<div class="card"><button class="btn" onclick="audit()">Run Recovery Audit</button> <button class="btn orange" onclick="apply()">Apply High-Confidence Recovery</button><span id="msg"></span></div>
<div class="kpis" id="kpis"></div>
<div class="card"><h3>Recoverable Records</h3><div class="tablewrap"><table><thead><tr><th>Property</th><th>Source</th><th>Current Issues</th><th>Suggested Recovery</th><th>Open</th></tr></thead><tbody id="rows"></tbody></table></div></div>
</div><script>
const esc=x=>String(x??'').replace(/[&<>"']/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[m]));
async function api(u,o={}){let r=await fetch(u,o),d=await r.json();if(!r.ok)throw new Error(d.detail||d.message||'Error');return d}
async function audit(){let d=await api('/api/v5/data-recovery/audit'),s=d.summary;document.querySelector('#kpis').innerHTML=[['TOTAL',s.total],['RECOVERABLE',s.recoverable],['FIELDS RECOVERABLE',s.high_confidence_fields],['POSSIBLE DUPLICATE RECORDS',s.possible_duplicates]].map(x=>`<div class="kpi"><span>${x[0]}</span><b>${Number(x[1]||0).toLocaleString()}</b></div>`).join('');document.querySelector('#rows').innerHTML=(d.recoverable||[]).map(x=>`<tr><td><b>${esc(x.property_name||x.property_id)}</b><br>${esc(x.property_id)}</td><td>${esc(x.source||'')}</td><td>${esc((x.issues||[]).join(', '))}</td><td>${Object.entries(x.high_confidence_suggestions||{}).map(([k,v])=>`<b>${esc(k)}</b>: ${esc(v.value)} (${v.confidence}%)`).join('<br>')}</td><td><a target="_blank" href="/property-record/${encodeURIComponent(x.property_id)}">View</a></td></tr>`).join('')}
async function apply(){if(!confirm('Apply only high-confidence recoveries now?'))return;document.querySelector('#msg').textContent=' Working...';let d=await api('/api/v5/data-recovery/apply',{method:'POST'});document.querySelector('#msg').textContent=' '+d.message;await audit()}audit();
</script></body></html>""")

def _ensure_retail_linkedin_campaign():
    name="Retail LinkedIn Requirement Watch"
    with engine.begin() as c:
        row=c.execute(text("SELECT campaign_id FROM ai_requirement_campaigns WHERE campaign_name=:n ORDER BY created_at DESC LIMIT 1"),{"n":name}).first()
        if row:return row[0]
        cid=_new_code("CAM")
        c.execute(text("""INSERT INTO ai_requirement_campaigns(
            campaign_id,campaign_name,property_type,city,location,rent_or_sale,suitable_category,additional_points,status
        ) VALUES(:id,:n,'Retail / Commercial','Delhi NCR','Delhi NCR','Rent','Retail',
        'Auto-created for public/indexed LinkedIn retail leasing requirement signals','ACTIVE')"""),
        {"id":cid,"n":name})
        return cid

def _retail_requirement_score(title,snippet):
    v=_norm((title or "")+" "+(snippet or ""))
    score=0
    intent=["looking for","requirement","required","seeking","need space","needs space","space required","space requirement","looking to lease","want to lease","on lease","for lease"]
    retail=["retail","store","shop","showroom","outlet","brand","qsr","cafe","restaurant"]
    geo=["delhi","delhi ncr","gurgaon","gurugram","noida","faridabad","ghaziabad"]
    role=["business development","bd","expansion","leasing","real estate","property acquisition","projects","store development"]
    if any(x in v for x in intent):score+=40
    if any(x in v for x in retail):score+=20
    if any(x in v for x in geo):score+=20
    if any(x in v for x in role):score+=10
    if any(x in v for x in ["lease","rent"]):score+=10
    return min(100,score)

def _save_retail_linkedin_requirement(item,campaign_id):
    title=item.get("title") or ""; link=item.get("link") or ""; snippet=item.get("snippet") or ""
    if "linkedin.com" not in link.lower():return None
    score=_retail_requirement_score(title,snippet)
    if score<70:return None
    with engine.connect() as c:
        if c.execute(text("SELECT 1 FROM ai_demand_signals WHERE source_url=:u LIMIT 1"),{"u":link}).first():
            return None
    phones,emails=_extract_public_contacts(title+" "+snippet)
    sid=_new_code("DEM")
    with engine.begin() as c:
        c.execute(text("""INSERT INTO ai_demand_signals(
            signal_id,campaign_id,source_type,source_name,source_url,title,excerpt,
            contact_phone,contact_email,location,intent_score,status,match_breakdown,
            contact_verification_status,source_contact_text
        ) VALUES(:sid,:cid,'RETAIL_LINKEDIN_REQUIREMENT','LinkedIn public/indexed',:u,:t,:x,:p,:e,
        'Delhi NCR',:s,'MANUAL_FOLLOWUP_REQUIRED',CAST(:b AS JSONB),:v,:ct)"""),{
            "sid":sid,"cid":campaign_id,"u":link,"t":title,"x":snippet,
            "p":phones[0] if phones else None,"e":emails[0] if emails else None,"s":score,
            "b":json.dumps({"intent_score":score,"signal":"Retail leasing requirement from LinkedIn/public index"}),
            "v":"PUBLIC_SOURCE" if (phones or emails) else "NOT_FOUND","ct":(title+" | "+snippet)[:2500]
        })
    _v6_enrich_retail_signal(sid,title,snippet,link)
    _log_activity("Retail Requirement Bot","DEMAND","LINKEDIN_RETAIL_REQUIREMENT","demand_signal",sid,f"LinkedIn | score {score} | manual follow-up")
    return sid

def _retail_linkedin_requirement_worker(run_id):
    campaign_id=_ensure_retail_linkedin_campaign()
    queries=[
        'site:linkedin.com/posts "looking for" retail space lease Delhi NCR',
        'site:linkedin.com/posts "space requirement" retail lease Delhi NCR',
        'site:linkedin.com/posts "requirement" showroom lease Delhi Gurgaon Noida',
        'site:linkedin.com/posts "looking to lease" store Gurgaon Noida Delhi',
        'site:linkedin.com/posts "business development" "space requirement" retail',
        'site:linkedin.com/posts "expansion" "looking for space" retail Delhi',
        'site:linkedin.com/posts "property acquisition" retail lease Delhi NCR',
        'site:linkedin.com/posts "store development" "requirement" Delhi NCR'
    ]
    found=created=0;errors=[]
    for q in queries:
        try:
            for item in _serper_search(q,10).get("organic",[]):
                found+=1
                sid=_save_retail_linkedin_requirement(item,campaign_id)
                if sid:created+=1
        except Exception as ex:errors.append(str(ex))
    _finish_bot(run_id,"COMPLETED" if created or not errors else "FAILED",found,created,
                f"LinkedIn retail requirement scan: {found} reviewed; {created} manual-follow-up signals saved",
                " | ".join(errors[:5]) or None)

@app.post("/api/v5/retail-linkedin-requirements/start")
def v5_retail_linkedin_requirements(bg:BackgroundTasks,req:Request):
    need_login(req)
    run=_start_bot("Retail LinkedIn Requirement Bot","DEMAND","Scanning public/indexed LinkedIn retail leasing requirements")
    bg.add_task(_retail_linkedin_requirement_worker,run)
    return {"status":"ACCEPTED","run_id":run}


# ============================================================
# V6 ORGANIZED RETAIL REQUIREMENTS
# ============================================================

def _ensure_v6_retail_requirement_columns():
    statements=[
        "ALTER TABLE ai_demand_signals ADD COLUMN IF NOT EXISTS company_name TEXT",
        "ALTER TABLE ai_demand_signals ADD COLUMN IF NOT EXISTS contact_name TEXT",
        "ALTER TABLE ai_demand_signals ADD COLUMN IF NOT EXISTS designation TEXT",
        "ALTER TABLE ai_demand_signals ADD COLUMN IF NOT EXISTS linkedin_profile_url TEXT",
        "ALTER TABLE ai_demand_signals ADD COLUMN IF NOT EXISTS required_area_sqft DOUBLE PRECISION",
        "ALTER TABLE ai_demand_signals ADD COLUMN IF NOT EXISTS required_property_type TEXT",
        "ALTER TABLE ai_demand_signals ADD COLUMN IF NOT EXISTS required_transaction TEXT",
        "ALTER TABLE ai_demand_signals ADD COLUMN IF NOT EXISTS assigned_to TEXT",
        "ALTER TABLE ai_demand_signals ADD COLUMN IF NOT EXISTS followup_status TEXT",
        "ALTER TABLE ai_demand_signals ADD COLUMN IF NOT EXISTS crm_status TEXT",
        "ALTER TABLE ai_demand_signals ADD COLUMN IF NOT EXISTS crm_notes TEXT",
        "ALTER TABLE ai_demand_signals ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP DEFAULT NOW()"
    ]
    with engine.begin() as c:
        for stmt in statements:
            c.execute(text(stmt))

def _v6_extract_retail_requirement_fields(title,snippet):
    raw=(title or "")+" | "+(snippet or "")
    norm=_norm(raw)
    company=_company_guess(title)
    contact_name=None
    designation=None

    m=_re.match(r"^\s*([A-Z][A-Za-z.' -]{2,60})\s*[-|–]\s*(.+)$",(title or "").strip())
    if m:
        tail=_norm(m.group(2))
        if any(role in tail for role in ["business development","expansion","leasing","real estate","property acquisition","store development","projects"]):
            contact_name=m.group(1).strip()

    for key,label in [
        ("business development","Business Development"),
        ("property acquisition","Property Acquisition"),
        ("store development","Store Development"),
        ("expansion","Expansion"),
        ("leasing","Leasing"),
        ("real estate","Real Estate"),
        ("projects","Projects")
    ]:
        if key in norm:
            designation=label
            break

    area=None
    for pat,mult in [
        (r"(?i)(\d[\d,]{2,6})\s*(?:sq\.?\s*ft|sqft|sft|square\s*feet)\b",1.0),
        (r"(?i)(\d[\d,]{2,5})\s*(?:sq\.?\s*yd|sqyd|sq\.?\s*yard|yards?)\b",9.0)
    ]:
        m2=_re.search(pat,raw)
        if m2:
            try:
                area=float(m2.group(1).replace(",",""))*mult
                break
            except Exception:
                pass

    ptype=None
    for terms,label in [
        (["restaurant","cafe","qsr","f&b","food outlet"],"F&B / Restaurant"),
        (["showroom"],"Showroom"),
        (["shop","retail","store","outlet"],"Retail / Store"),
        (["office"],"Office"),
        (["banquet"],"Banquet"),
        (["warehouse"],"Warehouse")
    ]:
        if any(t in norm for t in terms):
            ptype=label
            break

    tx="LEASE" if any(x in norm for x in ["lease","leasing","rent","rental"]) else None
    if company and company.lower().startswith(("linkedin","post by","view")):
        company=None

    return {
        "company_name":company,
        "contact_name":contact_name,
        "designation":designation,
        "required_area_sqft":area,
        "required_property_type":ptype,
        "required_transaction":tx
    }

def _v6_enrich_retail_signal(signal_id,title,snippet,source_url):
    _ensure_v6_retail_requirement_columns()
    f=_v6_extract_retail_requirement_fields(title,snippet)
    profile=None
    u=(source_url or "").lower()
    if "linkedin.com/in/" in u or "linkedin.com/company/" in u:
        profile=source_url

    with engine.begin() as c:
        c.execute(text("""UPDATE ai_demand_signals SET
            company_name=COALESCE(company_name,:company),
            contact_name=COALESCE(contact_name,:contact),
            designation=COALESCE(designation,:designation),
            linkedin_profile_url=COALESCE(linkedin_profile_url,:profile),
            required_area_sqft=COALESCE(required_area_sqft,:area),
            required_property_type=COALESCE(required_property_type,:ptype),
            required_transaction=COALESCE(required_transaction,:tx),
            followup_status=COALESCE(followup_status,'NEW'),
            crm_status=COALESCE(crm_status,'NOT_SENT'),
            updated_at=NOW()
            WHERE signal_id=:id"""),{
            "company":f["company_name"],"contact":f["contact_name"],"designation":f["designation"],
            "profile":profile,"area":f["required_area_sqft"],"ptype":f["required_property_type"],
            "tx":f["required_transaction"],"id":signal_id
        })

def _v6_backfill_retail_signals():
    _ensure_v6_retail_requirement_columns()
    with engine.connect() as c:
        rows=c.execute(text("""SELECT signal_id,title,excerpt,source_url
            FROM ai_demand_signals
            WHERE source_type='RETAIL_LINKEDIN_REQUIREMENT'
               OR source_name ILIKE '%LinkedIn%'""")).fetchall()
    done=0
    for r in rows:
        d=dict(r._mapping)
        _v6_enrich_retail_signal(d.get("signal_id"),d.get("title"),d.get("excerpt"),d.get("source_url"))
        done+=1
    return done

@app.get("/api/v6/retail-requirements")
def v6_retail_requirements(req:Request,limit:int=Query(1000,ge=1,le=3000)):
    need_login(req)
    _ensure_v6_retail_requirement_columns()
    with engine.connect() as c:
        rows=c.execute(text("""SELECT
            signal_id,campaign_id,company_name,contact_name,designation,
            contact_phone,contact_email,linkedin_profile_url,source_url,title,excerpt,
            location,required_area_sqft,required_property_type,required_transaction,
            intent_score,status,contact_verification_status,assigned_to,
            followup_status,crm_status,crm_notes,created_at,updated_at
            FROM ai_demand_signals
            WHERE source_type='RETAIL_LINKEDIN_REQUIREMENT'
               OR source_name ILIKE '%LinkedIn%'
            ORDER BY intent_score DESC,created_at DESC
            LIMIT :n"""),{"n":limit}).fetchall()
    return {"status":"ok","rows":_json_rows(rows)}

@app.post("/api/v6/retail-requirements/backfill")
def v6_retail_requirement_backfill(req:Request):
    need_login(req)
    return {"status":"ok","updated":_v6_backfill_retail_signals()}

@app.post("/api/v6/retail-requirements/{signal_id}/update")
async def v6_update_retail_requirement(signal_id:str,req:Request):
    need_login(req)
    body=await req.json()
    allowed={
        "company_name","contact_name","designation","contact_phone","contact_email",
        "linkedin_profile_url","required_area_sqft","required_property_type",
        "required_transaction","assigned_to","followup_status","crm_status","crm_notes"
    }
    updates={k:v for k,v in body.items() if k in allowed}
    if not updates:
        return {"status":"ok","message":"Nothing to update"}
    sets=[];params={"id":signal_id}
    for k,v in updates.items():
        sets.append(f"{k}=:{k}")
        params[k]=v
    sets.append("updated_at=NOW()")
    with engine.begin() as c:
        c.execute(text("UPDATE ai_demand_signals SET "+",".join(sets)+" WHERE signal_id=:id"),params)
    return {"status":"ok","signal_id":signal_id}

@app.get("/retail-requirements",response_class=HTMLResponse)
def v6_retail_requirements_page(req:Request):
    role=page_role_or_redirect(req)
    if not role:
        return RedirectResponse("/login",status_code=303)
    return HTMLResponse(r"""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Retail Requirement Leads</title>
<style>
*{box-sizing:border-box}body{margin:0;background:#f4f7fb;font-family:Arial;color:#142033}header{background:#0d1d2d;color:#fff;padding:16px 22px}.wrap{padding:18px}.card{background:#fff;border:1px solid #e4eaf1;border-radius:12px;padding:14px;margin-bottom:12px}.toolbar{display:flex;gap:8px;flex-wrap:wrap}.btn{padding:8px 10px;border:0;border-radius:7px;background:#1677ff;color:#fff;font-weight:700;text-decoration:none;cursor:pointer}.gray{background:#edf2f7;color:#24364b}input,select{padding:7px;border:1px solid #ccd7e4;border-radius:6px}.tablewrap{overflow:auto;max-height:72vh}table{width:100%;border-collapse:collapse;font-size:12px}th,td{padding:8px;border-bottom:1px solid #edf1f5;text-align:left;vertical-align:top;white-space:nowrap}th{position:sticky;top:0;background:#f8fafc;z-index:1}.small{font-size:11px;color:#697589}
</style></head><body>
<header><b>Retail Requirement Leads</b><br><small>LinkedIn/public-indexed leasing requirements for manual verification and follow-up</small></header>
<div class="wrap">
<div class="card toolbar">
<a class="btn gray" href="/workspace">Workspace</a><a class="btn gray" href="/data-recovery">Data Recovery</a>
<input id="search" placeholder="Search company, person, phone, email, location, requirement">
<select id="follow"><option value="">All Follow-up</option><option>NEW</option><option>CONTACTED</option><option>QUALIFIED</option><option>NOT_RELEVANT</option></select>
<button class="btn" onclick="load()">Refresh</button><button class="btn gray" onclick="backfill()">Organize Existing Signals</button>
</div>
<div class="card"><div class="tablewrap"><table><thead><tr>
<th>Company / Brand</th><th>Contact Person</th><th>Designation</th><th>Mobile</th><th>Email</th><th>LinkedIn Profile</th><th>Requirement Post</th><th>Requirement</th><th>Location</th><th>Area</th><th>Property Type</th><th>Transaction</th><th>Intent</th><th>Assigned</th><th>Follow-up</th><th>CRM</th><th>Save</th>
</tr></thead><tbody id="rows"></tbody></table></div></div></div>
<script>
const E=s=>String(s??'').replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));
let D=[];
async function A(u,o={}){let r=await fetch(u,o),d=await r.json();if(!r.ok)throw new Error(d.detail||d.message||'Error');return d}
function I(id,k,v){return `<input id="${id}_${k}" value="${E(v||'')}">`}
function S(id,k,v,opts){return `<select id="${id}_${k}">${opts.map(o=>`<option ${o===v?'selected':''}>${o}</option>`).join('')}</select>`}
function R(){let q=(document.querySelector('#search').value||'').toLowerCase(),f=document.querySelector('#follow').value;let rows=D.filter(x=>(!q||JSON.stringify(x).toLowerCase().includes(q))&&(!f||x.followup_status===f));document.querySelector('#rows').innerHTML=rows.map(x=>`<tr>
<td>${I(x.signal_id,'company_name',x.company_name)}</td><td>${I(x.signal_id,'contact_name',x.contact_name)}</td><td>${I(x.signal_id,'designation',x.designation)}</td>
<td>${I(x.signal_id,'contact_phone',x.contact_phone)}</td><td>${I(x.signal_id,'contact_email',x.contact_email)}</td>
<td>${x.linkedin_profile_url?`<a target="_blank" href="${E(x.linkedin_profile_url)}">Profile</a>`:'Not Found'}</td>
<td>${x.source_url?`<a target="_blank" href="${E(x.source_url)}">Open Post</a>`:'Not Found'}</td>
<td style="max-width:320px;white-space:normal"><b>${E(x.title||'')}</b><br><span class="small">${E(x.excerpt||'')}</span></td>
<td>${E(x.location||'')}</td><td>${I(x.signal_id,'required_area_sqft',x.required_area_sqft)}</td><td>${I(x.signal_id,'required_property_type',x.required_property_type)}</td><td>${I(x.signal_id,'required_transaction',x.required_transaction)}</td>
<td>${E(x.intent_score||0)}</td><td>${I(x.signal_id,'assigned_to',x.assigned_to)}</td>
<td>${S(x.signal_id,'followup_status',x.followup_status||'NEW',['NEW','CONTACTED','QUALIFIED','NOT_RELEVANT'])}</td>
<td>${S(x.signal_id,'crm_status',x.crm_status||'NOT_SENT',['NOT_SENT','CRM_READY','ADDED_TO_CRM'])}</td>
<td><button class="btn" onclick="save('${x.signal_id}')">Save</button></td></tr>`).join('')}
async function load(){let d=await A('/api/v6/retail-requirements');D=d.rows||[];R()}
async function backfill(){let d=await A('/api/v6/retail-requirements/backfill',{method:'POST'});alert('Organized '+d.updated+' existing LinkedIn signals');await load()}
async function save(id){let ks=['company_name','contact_name','designation','contact_phone','contact_email','required_area_sqft','required_property_type','required_transaction','assigned_to','followup_status','crm_status'];let b={};ks.forEach(k=>{let e=document.getElementById(id+'_'+k);if(e)b[k]=e.value});await A('/api/v6/retail-requirements/'+encodeURIComponent(id)+'/update',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(b)});alert('Saved')}
document.querySelector('#search').addEventListener('input',R);document.querySelector('#follow').addEventListener('change',R);load();
</script></body></html>""")


# ============================================================
# V7 MASTER PROPERTY DATABASE CLEANER
# ============================================================

def _ensure_v7_master_columns():
    statements=[
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS plot_area_sqft DOUBLE PRECISION",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS builtup_area_sqft DOUBLE PRECISION",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS area_basis TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS occupancy_status TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS offering_transaction TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS duplicate_group_id TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS duplicate_status TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS master_property_id TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS matching_bucket TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v7_quality_score INTEGER",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v7_review_reasons JSONB DEFAULT '[]'::jsonb",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v7_updated_at TIMESTAMP"
    ]
    with engine.begin() as c:
        for stmt in statements:
            c.execute(text(stmt))

def _v7_norm_name(value):
    v=_norm(value)
    v=_re.sub(r'\b(property|prop|available|booking|new|ready|urgent)\b',' ',v)
    v=_re.sub(r'[^a-z0-9]+',' ',v)
    return _re.sub(r'\s+',' ',v).strip()

def _v7_area_facts(p):
    raw=_v5_source_text(p)
    plot=None
    built=None
    basis=None

    m=_re.search(r'(?i)(\d[\d,]{1,5})\s*(?:sq\.?\s*yd|sqyd|sq\.?\s*yard|yards?|yds?|\byd\b)',raw)
    if m:
        try:
            y=float(m.group(1).replace(",",""))
            if 10 <= y <= 100000:
                plot=round(y*9.0,2)
                basis="PLOT_FROM_SQYD"
        except Exception:
            pass

    m2=_re.search(r'(?i)(\d[\d,]{1,6})\s*(?:sq\.?\s*ft|sqft|sft|square\s*feet)\b',raw)
    if m2:
        try:
            f=float(m2.group(1).replace(",",""))
            if 50 <= f <= 500000:
                built=round(f,2)
                basis="BUILTUP_FROM_SQFT"
        except Exception:
            pass

    existing=_property_area(p)
    if existing and not built:
        built=float(existing)
        basis=basis or "EXISTING_STRUCTURED_AREA"

    return {"plot_area_sqft":plot,"builtup_area_sqft":built,"area_basis":basis}

def _v7_transaction_facts(p):
    raw=_norm(_v5_source_text(p))
    occupancy=None
    offered=None
    reasons=[]

    if any(x in raw for x in ["preleased","pre leased","pre-leased"]):
        occupancy="PRELEASED"
        reasons.append("PRELEASED_IS_OCCUPANCY_NOT_RENT_OFFER")
    elif any(x in raw for x in ["rented","tenanted","tenant occupied","leased out"]):
        occupancy="TENANTED"
        reasons.append("RENTED_IS_OCCUPANCY_NOT_RENT_OFFER")
    elif any(x in raw for x in ["vacant","ready possession","ready to move"]):
        occupancy="VACANT_OR_READY"

    rent_offer=any(x in raw for x in [
        "for rent","on rent","to let","available for lease","available on lease",
        "space for lease","offered for lease","lease available"
    ])
    sale_offer=any(x in raw for x in [
        "for sale","on sale","available for sale","asking sale","sale @","resale","sell"
    ])

    if rent_offer and sale_offer:
        offered="RENT/SALE"
    elif rent_offer:
        offered="RENT"
    elif sale_offer:
        offered="SALE"
    else:
        structured=_canonical_transaction_v4(p.get("rent_or_sale"))
        if structured in {"RENT","SALE","RENT/SALE"}:
            offered=structured

    if "lease hold" in raw or "leasehold" in raw:
        reasons.append("LEASEHOLD_IS_TENURE_NOT_RENT_OFFER")

    return {"occupancy_status":occupancy,"offering_transaction":offered,"reasons":reasons}

def _v7_completeness_score(p):
    q=_organize_property_v4(p)
    score=0
    if _canonical_city_v4(p.get("city"))!="UNKNOWN": score+=12
    loc=_dq_text(p.get("location") or p.get("micro_market"))
    if loc and not _dq_unknown(loc): score+=15
    if _canonical_property_type_v4(p.get("property_type"),p.get("suitable_category"),p.get("remarks"))!="UNKNOWN": score+=15
    area=_v7_area_facts(p)
    if area["builtup_area_sqft"]: score+=15
    tx=_v7_transaction_facts(p)
    if tx["offering_transaction"]: score+=15
    if q["contact_ready"]: score+=15
    if _dq_text(p.get("address")): score+=5
    if _dq_text(p.get("property_name")): score+=3
    if str(p.get("verification_status") or "").upper()=="VERIFIED": score+=5
    return min(100,score)

def _v7_duplicate_signature(p):
    name=_v7_norm_name(p.get("property_name"))
    loc=_norm(p.get("location") or p.get("micro_market"))
    city=_norm(p.get("city"))
    area=_v7_area_facts(p)
    a=area["builtup_area_sqft"] or area["plot_area_sqft"] or 0
    contacts=sorted(_contact_number_set(
        p.get("owner_contact"),p.get("broker_contact"),p.get("contact_number"),
        p.get("owner_contact_normalized"),p.get("broker_contact_normalized"),
        p.get("general_contact_normalized")
    ))
    keys=[]
    if contacts:
        for c in contacts:
            keys.append("CONTACT|"+c+"|"+loc)
    if name and len(name)>=3 and a:
        keys.append("NAME_AREA|"+name+"|"+loc+"|"+str(round(float(a),1)))
    if name and loc and city:
        keys.append("NAME_LOC|"+name+"|"+loc+"|"+city)
    return keys

def _v7_duplicate_groups(properties):
    keymap={}
    byid={str(p.get("property_id")):p for p in properties}
    for p in properties:
        pid=str(p.get("property_id"))
        for k in _v7_duplicate_signature(p):
            keymap.setdefault(k,set()).add(pid)

    parent={pid:pid for pid in byid}
    def find(x):
        while parent[x]!=x:
            parent[x]=parent[parent[x]]
            x=parent[x]
        return x
    def union(a,b):
        ra,rb=find(a),find(b)
        if ra!=rb:
            parent[rb]=ra

    for ids in keymap.values():
        ids=list(ids)
        if len(ids)>1:
            for x in ids[1:]:
                union(ids[0],x)

    groups={}
    for pid in byid:
        groups.setdefault(find(pid),[]).append(pid)

    final=[]
    for ids in groups.values():
        if len(ids)>1:
            ranked=sorted(ids,key=lambda pid:(
                _v7_completeness_score(byid[pid]),
                1 if str(byid[pid].get("verification_status") or "").upper()=="VERIFIED" else 0,
                pid
            ),reverse=True)
            master=ranked[0]
            gid="DUP-"+hashlib.sha256("|".join(sorted(ids)).encode()).hexdigest()[:12].upper()
            all_contacts=sorted(set().union(*[
                _contact_number_set(
                    byid[x].get("owner_contact"),byid[x].get("broker_contact"),byid[x].get("contact_number"),
                    byid[x].get("owner_contact_normalized"),byid[x].get("broker_contact_normalized"),
                    byid[x].get("general_contact_normalized")
                ) for x in ids
            ]))
            final.append({"group_id":gid,"master_property_id":master,"property_ids":ranked,"all_contacts":all_contacts})
    return final

def _v7_classify_property(p,is_duplicate=False,is_master=False):
    q=_organize_property_v4(p)
    area=_v7_area_facts(p)
    tx=_v7_transaction_facts(p)
    reasons=list(q["quality_issues"])+list(tx["reasons"])

    core_ok=True
    if _canonical_city_v4(p.get("city"))=="UNKNOWN":
        core_ok=False
    loc=_dq_text(p.get("location") or p.get("micro_market"))
    if not loc or _dq_unknown(loc):
        core_ok=False
    if _canonical_property_type_v4(p.get("property_type"),p.get("suitable_category"),p.get("remarks"))=="UNKNOWN":
        core_ok=False
    if not area["builtup_area_sqft"]:
        core_ok=False
        if area["plot_area_sqft"]:
            reasons.append("PLOT_AREA_KNOWN_BUT_BUILTUP_OR_AVAILABLE_AREA_UNKNOWN")
    if not tx["offering_transaction"]:
        core_ok=False
        reasons.append("OFFERING_TRANSACTION_UNCONFIRMED")

    if is_duplicate and not is_master:
        bucket="DUPLICATE_REVIEW"
        reasons.append("POSSIBLE_DUPLICATE_NON_MASTER")
    elif not core_ok:
        bucket="DATA_REVIEW"
    elif not q["contact_ready"]:
        bucket="CONTACT_REVIEW"
    else:
        bucket="MATCH_READY"

    return {
        "bucket":bucket,
        "quality_score":_v7_completeness_score(p),
        "review_reasons":list(dict.fromkeys(reasons)),
        **area,**tx
    }

def _v7_audit_master_database(limit=1000):
    _ensure_v7_master_columns()
    with engine.connect() as c:
        props=[dict(r._mapping) for r in c.execute(text("SELECT * FROM pi_properties ORDER BY created_at DESC")).fetchall()]

    groups=_v7_duplicate_groups(props)
    duplicate_map={}
    for g in groups:
        for pid in g["property_ids"]:
            duplicate_map[pid]=g

    counts={"total":len(props),"match_ready":0,"contact_review":0,"data_review":0,"duplicate_review":0,
            "duplicate_groups":len(groups),"duplicate_records":sum(len(g["property_ids"]) for g in groups)}
    reviewed=[]
    for p in props:
        pid=str(p.get("property_id"))
        g=duplicate_map.get(pid)
        cl=_v7_classify_property(p,bool(g),bool(g and g["master_property_id"]==pid))
        counts[cl["bucket"].lower()]+=1
        if len(reviewed)<limit and cl["bucket"]!="MATCH_READY":
            reviewed.append({
                "property_id":pid,"property_name":p.get("property_name"),"city":p.get("city"),
                "location":p.get("location"),"property_type":p.get("property_type"),
                "bucket":cl["bucket"],"score":cl["quality_score"],
                "plot_area_sqft":cl["plot_area_sqft"],"builtup_area_sqft":cl["builtup_area_sqft"],
                "occupancy_status":cl["occupancy_status"],"offering_transaction":cl["offering_transaction"],
                "reasons":cl["review_reasons"],
                "duplicate_group_id":g["group_id"] if g else None,
                "master_property_id":g["master_property_id"] if g else None
            })
    return {"summary":counts,"reviewed":reviewed,"duplicate_groups":groups[:500]}

def _v7_apply_classification():
    _ensure_v7_master_columns()
    with engine.connect() as c:
        props=[dict(r._mapping) for r in c.execute(text("SELECT * FROM pi_properties")).fetchall()]
    groups=_v7_duplicate_groups(props)
    duplicate_map={}
    for g in groups:
        for pid in g["property_ids"]:
            duplicate_map[pid]=g

    with engine.begin() as c:
        for p in props:
            pid=str(p.get("property_id"))
            g=duplicate_map.get(pid)
            is_master=bool(g and g["master_property_id"]==pid)
            cl=_v7_classify_property(p,bool(g),is_master)
            c.execute(text("""UPDATE pi_properties SET
                plot_area_sqft=:plot,
                builtup_area_sqft=:built,
                area_basis=:basis,
                occupancy_status=:occ,
                offering_transaction=:tx,
                duplicate_group_id=:gid,
                duplicate_status=:ds,
                master_property_id=:mid,
                matching_bucket=:bucket,
                v7_quality_score=:score,
                v7_review_reasons=CAST(:reasons AS JSONB),
                v7_updated_at=NOW()
                WHERE property_id=:id"""),{
                "plot":cl["plot_area_sqft"],"built":cl["builtup_area_sqft"],"basis":cl["area_basis"],
                "occ":cl["occupancy_status"],"tx":cl["offering_transaction"],
                "gid":g["group_id"] if g else None,
                "ds":("MASTER" if is_master else "DUPLICATE") if g else "UNIQUE",
                "mid":g["master_property_id"] if g else None,
                "bucket":cl["bucket"],"score":cl["quality_score"],
                "reasons":json.dumps(cl["review_reasons"]),"id":pid
            })
    return _v7_audit_master_database(1000)

@app.get("/api/v7/master-cleaner/audit")
def v7_master_cleaner_audit(req:Request):
    need_login(req)
    return {"status":"ok",**_v7_audit_master_database()}

@app.post("/api/v7/master-cleaner/apply")
def v7_master_cleaner_apply(req:Request):
    need_login(req)
    result=_v7_apply_classification()
    return {"status":"ok","message":"V7 classification applied. No records deleted and no phone digits invented.",**result}

@app.get("/master-data-cleaner",response_class=HTMLResponse)
def v7_master_data_cleaner_page(req:Request):
    role=page_role_or_redirect(req)
    if not role:
        return RedirectResponse("/login",status_code=303)
    return HTMLResponse(r"""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>V7 Master Property Database Cleaner</title>
<style>*{box-sizing:border-box}body{margin:0;background:#f4f7fb;font-family:Arial;color:#142033}header{background:#0d1d2d;color:white;padding:16px 22px}.wrap{max-width:1850px;margin:auto;padding:18px}.card,.kpi{background:white;border:1px solid #e4eaf1;border-radius:12px;padding:14px;margin-bottom:12px}.kpis{display:grid;grid-template-columns:repeat(4,minmax(150px,1fr));gap:10px}.kpi b{display:block;font-size:26px;margin-top:5px}.btn{display:inline-block;border:0;border-radius:8px;padding:9px 12px;background:#1677ff;color:white;text-decoration:none;font-weight:700;cursor:pointer}.orange{background:#df8b13}.gray{background:#edf2f7;color:#24364b}.tablewrap{overflow:auto;max-height:62vh}table{width:100%;border-collapse:collapse;font-size:12px}th,td{padding:8px;border-bottom:1px solid #edf1f5;text-align:left;vertical-align:top;white-space:nowrap}th{position:sticky;top:0;background:#f8fafc}.MATCH_READY{color:#08734b;font-weight:700}.DATA_REVIEW,.DUPLICATE_REVIEW{color:#b23b00;font-weight:700}.CONTACT_REVIEW{color:#a36b00;font-weight:700}</style></head><body>
<header><b>V7 Master Property Database Cleaner</b><br><small>Duplicate control · area semantics · occupancy vs transaction · match-ready gating</small></header>
<div class="wrap"><div class="card"><a class="btn gray" href="/workspace">Workspace</a> <a class="btn gray" href="/property-database">Property Database</a> <a class="btn gray" href="/data-recovery">V5 Recovery</a></div>
<div class="card"><b>Safety:</b> No source record is deleted. Square-yard values are plot area, not automatically built-up area. Rented/preleased is occupancy, not automatically a RENT offering.</div>
<div class="card"><button class="btn" onclick="audit()">Run V7 Audit</button> <button class="btn orange" onclick="apply()">Apply Classification</button> <span id="msg"></span></div>
<div class="kpis" id="kpis"></div>
<div class="card"><h3>Records Requiring Review</h3><div class="tablewrap"><table><thead><tr><th>Property</th><th>Bucket</th><th>Score</th><th>City</th><th>Location</th><th>Type</th><th>Plot Area</th><th>Built-up/Available</th><th>Occupancy</th><th>Offer Transaction</th><th>Duplicate Group</th><th>Master</th><th>Reasons</th><th>Open</th></tr></thead><tbody id="rows"></tbody></table></div></div>
<div class="card"><h3>Duplicate Groups</h3><div class="tablewrap"><table><thead><tr><th>Group</th><th>Master</th><th>Records</th><th>All Valid Contacts Across Group</th></tr></thead><tbody id="dups"></tbody></table></div></div></div>
<script>
const E=x=>String(x??'').replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));
async function A(u,o={}){let r=await fetch(u,o),d=await r.json();if(!r.ok)throw new Error(d.detail||d.message||'Error');return d}
function render(d){let s=d.summary;document.querySelector('#kpis').innerHTML=[['TOTAL',s.total],['MATCH READY',s.match_ready],['CONTACT REVIEW',s.contact_review],['DATA REVIEW',s.data_review],['DUPLICATE REVIEW',s.duplicate_review],['DUPLICATE GROUPS',s.duplicate_groups],['DUPLICATE RECORDS',s.duplicate_records]].map(x=>`<div class="kpi"><span>${x[0]}</span><b>${Number(x[1]||0).toLocaleString()}</b></div>`).join('');
document.querySelector('#rows').innerHTML=(d.reviewed||[]).map(x=>`<tr><td><b>${E(x.property_name||x.property_id)}</b><br>${E(x.property_id)}</td><td class="${E(x.bucket)}">${E(x.bucket)}</td><td>${E(x.score)}</td><td>${E(x.city||'')}</td><td>${E(x.location||'')}</td><td>${E(x.property_type||'')}</td><td>${E(x.plot_area_sqft||'')}</td><td>${E(x.builtup_area_sqft||'')}</td><td>${E(x.occupancy_status||'')}</td><td>${E(x.offering_transaction||'')}</td><td>${E(x.duplicate_group_id||'')}</td><td>${E(x.master_property_id||'')}</td><td style="max-width:420px;white-space:normal">${E((x.reasons||[]).join(', '))}</td><td><a target="_blank" href="/property-record/${encodeURIComponent(x.property_id)}">View</a></td></tr>`).join('');
document.querySelector('#dups').innerHTML=(d.duplicate_groups||[]).map(g=>`<tr><td>${E(g.group_id)}</td><td><a target="_blank" href="/property-record/${encodeURIComponent(g.master_property_id)}">${E(g.master_property_id)}</a></td><td>${(g.property_ids||[]).map(id=>`<a target="_blank" href="/property-record/${encodeURIComponent(id)}">${E(id)}</a>`).join('<br>')}</td><td>${E((g.all_contacts||[]).join(', '))}</td></tr>`).join('')}
async function audit(){document.querySelector('#msg').textContent=' Auditing...';let d=await A('/api/v7/master-cleaner/audit');render(d);document.querySelector('#msg').textContent=' Audit complete'}
async function apply(){if(!confirm('Apply V7 classifications? No records will be deleted.'))return;document.querySelector('#msg').textContent=' Classifying...';let d=await A('/api/v7/master-cleaner/apply',{method:'POST'});render(d);document.querySelector('#msg').textContent=' '+d.message}
audit();
</script></body></html>""")


# ============================================================
# V8 SMART MASTER DATA ENGINE
# ============================================================

_V8_GENERIC={"unknown","na","n a","south delhi","new delhi","delhi","4bhk","3bhk","2bhk","1bhk","commercial","residential","shop","office","property","parking","booking","rented","preleased","lease hold","leasehold","unit","bmt","gf","ff","sf"}

def _ensure_v8_columns():
    stmts=[
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v8_identity_key TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v8_locality_key TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v8_duplicate_confidence TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v8_duplicate_evidence JSONB DEFAULT '[]'::jsonb",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v8_master_property_id TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v8_bucket TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v8_match_eligible BOOLEAN",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v8_quality_score INTEGER",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v8_review_reasons JSONB DEFAULT '[]'::jsonb",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v8_updated_at TIMESTAMP"
    ]
    with engine.begin() as c:
        for s in stmts:c.execute(text(s))

def _v8_clean(v):
    v=_norm(v)
    v=_re.sub(r'[^a-z0-9/ -]+',' ',v)
    return _re.sub(r'\s+',' ',v).strip()

def _v8_loc(p):
    v=_v8_clean(p.get("location") or p.get("micro_market"))
    aliases={"kalash colony":"kailash colony","gk 1":"greater kailash 1","gk1":"greater kailash 1","gk 2":"greater kailash 2","gk2":"greater kailash 2","nfc":"new friends colony","gurgaon":"gurugram","safdurjung enclave":"safdarjung enclave","defense colony":"defence colony","chatterpur":"chattarpur"}
    return aliases.get(v,v if v not in {"unknown","na","n a","none","not specified"} else "")

def _v8_identity(p):
    for raw in [p.get("property_name"),p.get("address")]:
        s=_v8_clean(raw)
        if not s or s in _V8_GENERIC:continue
        m=_re.search(r'\b(?:shop\s*no\s*[- ]?\d+[a-z]?|unit\s*no\s*[- ]?\d+[a-z]?|[a-z]{1,3}\s*[-/]\s*\d+[a-z0-9/-]*|\d+[a-z]?\s+[a-z][a-z ]{2,25}(?:complex|tower|house|market|mkt))\b',s,re.I)
        if m:return _v8_clean(m.group(0))
    return ""

def _v8_area(p):
    raw=_v5_source_text(p)
    plot=built=None
    reasons=[]
    y=_re.search(r'(?i)(\d[\d,]{1,5})\s*(?:sq\.?\s*yd|sqyd|sq\.?\s*yard|yards?|yds?|\byd\b)',raw)
    f=_re.search(r'(?i)(\d[\d,]{1,6})\s*(?:sq\.?\s*ft|sqft|sft|square\s*feet)\b',raw)
    if y:
        try:
            yy=float(y.group(1).replace(",","")); plot=round(yy*9,2) if 10<=yy<=100000 else None
        except: pass
    if f:
        try:
            ff=float(f.group(1).replace(",","")); built=round(ff,2) if 50<=ff<=500000 else None
        except: pass
    existing=_property_area(p)
    if plot and not built and existing and abs(float(existing)-plot)<=max(1,plot*.01):
        reasons.append("LEGACY_SQYD_AS_BUILTUP_REJECTED")
        existing=None
    if not built and existing:built=float(existing)
    return {"plot":plot,"built":built,"reasons":reasons}

def _v8_tx(p):
    raw=_norm(_v5_source_text(p))
    reasons=[];occ=None;off=None
    if any(x in raw for x in ["preleased","pre leased","pre-leased"]):occ="PRELEASED"
    elif any(x in raw for x in ["rented","tenanted","leased out"]):occ="TENANTED"
    elif any(x in raw for x in ["vacant","ready possession","ready to move"]):occ="VACANT_OR_READY"

    rent=any(x in raw for x in ["for rent","on rent","to let","available for lease","available on lease","space for lease","offered for lease"])
    sale=any(x in raw for x in ["for sale","on sale","available for sale","asking sale","sale @","resale","sell","selling"])
    if rent and sale:off="RENT/SALE"
    elif rent:off="RENT"
    elif sale:off="SALE"
    else:
        st=_canonical_transaction_v4(p.get("rent_or_sale"))
        bad=(st=="RENT" and any(x in raw for x in ["rented","preleased","pre leased","pre-leased","lease hold","leasehold"]))
        if bad:reasons.append("LEGACY_RENT_INFERENCE_REJECTED")
        elif st in {"RENT","SALE","RENT/SALE"}:off=st
    if "lease hold" in raw or "leasehold" in raw:reasons.append("LEASEHOLD_IS_TENURE")
    return {"occupancy":occ,"offering":off,"reasons":reasons}

def _v8_quality(p):
    q=_organize_property_v4(p); a=_v8_area(p); t=_v8_tx(p); s=0
    if _canonical_city_v4(p.get("city"))!="UNKNOWN":s+=15
    if _v8_loc(p):s+=15
    if _canonical_property_type_v4(p.get("property_type"),p.get("suitable_category"),p.get("remarks"))!="UNKNOWN":s+=15
    if a["built"]:s+=15
    if t["offering"]:s+=15
    if q["contact_ready"]:s+=15
    if _v8_identity(p):s+=5
    if str(p.get("verification_status") or "").upper()=="VERIFIED":s+=5
    return min(100,s)

def _v8_area_ok(a,b):
    if not a or not b:return None
    return abs(float(a)-float(b))/max(float(a),float(b))<=.10

def _v8_pair(a,b):
    ia,ib=_v8_identity(a),_v8_identity(b)
    la,lb=_v8_loc(a),_v8_loc(b)
    if not ia or not ib or ia!=ib or not la or la!=lb:return None,[]
    aa=_v8_area(a)["built"] or _v8_area(a)["plot"]
    ab=_v8_area(b)["built"] or _v8_area(b)["plot"]
    ok=_v8_area_ok(aa,ab)
    if ok is True:return "STRONG",["SAME_SPECIFIC_IDENTITY","SAME_LOCALITY","AREA_COMPATIBLE_10_PERCENT"]
    if ok is None:return "POSSIBLE",["SAME_SPECIFIC_IDENTITY","SAME_LOCALITY","AREA_MISSING"]
    return "POSSIBLE",["SAME_SPECIFIC_IDENTITY","SAME_LOCALITY","AREA_CONFLICT"]

def _v8_duplicates(props):
    byid={str(p.get("property_id")):p for p in props}
    buckets={}
    for p in props:
        i,l=_v8_identity(p),_v8_loc(p)
        if i and l:buckets.setdefault((i,l),[]).append(str(p.get("property_id")))

    strong=[];possible=[]
    for ids in buckets.values():
        ids=list(dict.fromkeys(ids))
        if len(ids)<2 or len(ids)>50:continue
        for i in range(len(ids)):
            for j in range(i+1,len(ids)):
                lvl,ev=_v8_pair(byid[ids[i]],byid[ids[j]])
                if lvl=="STRONG":strong.append((ids[i],ids[j],ev))
                elif lvl=="POSSIBLE":possible.append({"property_ids":[ids[i],ids[j]],"evidence":ev})

    parent={x:x for x in byid}
    def find(x):
        while parent[x]!=x:
            parent[x]=parent[parent[x]];x=parent[x]
        return x
    def union(a,b):
        a,b=find(a),find(b)
        if a!=b:parent[b]=a
    for a,b,_ in strong:union(a,b)

    groups={}
    for pid in byid:groups.setdefault(find(pid),[]).append(pid)

    final=[];member={}
    for ids in groups.values():
        if len(ids)>1:
            ranked=sorted(ids,key=lambda x:(_v8_quality(byid[x]),1 if str(byid[x].get("verification_status") or "").upper()=="VERIFIED" else 0,x),reverse=True)
            gid="V8DUP-"+hashlib.sha256("|".join(sorted(ids)).encode()).hexdigest()[:12].upper()
            ev=[]
            for a,b,e in strong:
                if a in ids and b in ids:ev+=e
            g={"group_id":gid,"master_property_id":ranked[0],"property_ids":ranked,"evidence":list(dict.fromkeys(ev))}
            final.append(g)
            for x in ids:member[x]=g
    return final,possible,member

def _v8_classify(p,g=None,is_master=False,possible=False):
    q=_organize_property_v4(p);a=_v8_area(p);t=_v8_tx(p)
    reasons=list(q["quality_issues"])+a["reasons"]+t["reasons"]
    core=True
    if _canonical_city_v4(p.get("city"))=="UNKNOWN":core=False
    if not _v8_loc(p):core=False
    if _canonical_property_type_v4(p.get("property_type"),p.get("suitable_category"),p.get("remarks"))=="UNKNOWN":core=False
    if not a["built"]:
        core=False
        if a["plot"]:reasons.append("PLOT_AREA_ONLY_BUILTUP_UNKNOWN")
    if not t["offering"]:
        core=False;reasons.append("OFFERING_TRANSACTION_UNCONFIRMED")

    if g and not is_master:bucket="DUPLICATE_REVIEW";reasons.append("STRONG_DUPLICATE_NON_MASTER")
    elif possible:bucket="POSSIBLE_DUPLICATE_REVIEW";reasons.append("POSSIBLE_DUPLICATE_TEAM_REVIEW")
    elif not core:bucket="DATA_REVIEW"
    elif not q["contact_ready"]:bucket="CONTACT_REVIEW"
    else:bucket="MATCH_READY"

    return {"bucket":bucket,"match":bucket=="MATCH_READY","score":_v8_quality(p),"identity":_v8_identity(p),"locality":_v8_loc(p),
            "plot":a["plot"],"built":a["built"],"occupancy":t["occupancy"],"offering":t["offering"],"reasons":list(dict.fromkeys(reasons))}

def _v8_audit(limit=1200):
    _ensure_v8_columns()
    with engine.connect() as c:props=[dict(r._mapping) for r in c.execute(text("SELECT * FROM pi_properties ORDER BY created_at DESC")).fetchall()]
    groups,possible,member=_v8_duplicates(props)
    pmap={}
    for pr in possible:
        for pid in pr["property_ids"]:pmap.setdefault(pid,[]).append(pr)
    s={"total":len(props),"match_ready":0,"contact_review":0,"data_review":0,"duplicate_review":0,"possible_duplicate_review":0,
       "strong_duplicate_groups":len(groups),"strong_duplicate_records":sum(len(g["property_ids"]) for g in groups),"possible_duplicate_pairs":len(possible)}
    rows=[]
    for p in props:
        pid=str(p.get("property_id"));g=member.get(pid);cl=_v8_classify(p,g,bool(g and g["master_property_id"]==pid),bool(pmap.get(pid)))
        s[cl["bucket"].lower()]+=1
        if len(rows)<limit and cl["bucket"]!="MATCH_READY":
            rows.append({"property_id":pid,"property_name":p.get("property_name"),"city":p.get("city"),"location":p.get("location"),
                         "property_type":p.get("property_type"),"bucket":cl["bucket"],"score":cl["score"],"identity":cl["identity"],
                         "locality":cl["locality"],"plot_area_sqft":cl["plot"],"builtup_area_sqft":cl["built"],"occupancy_status":cl["occupancy"],
                         "offering_transaction":cl["offering"],"master_property_id":g["master_property_id"] if g else None,"reasons":cl["reasons"]})
    return {"summary":s,"reviewed":rows,"strong_duplicate_groups":groups[:500],"possible_duplicates":possible[:500]}

def _v8_apply():
    _ensure_v8_columns()
    with engine.connect() as c:props=[dict(r._mapping) for r in c.execute(text("SELECT * FROM pi_properties")).fetchall()]
    groups,possible,member=_v8_duplicates(props)
    pmap={}
    for pr in possible:
        for pid in pr["property_ids"]:pmap.setdefault(pid,[]).append(pr)
    with engine.begin() as c:
        for p in props:
            pid=str(p.get("property_id"));g=member.get(pid);cl=_v8_classify(p,g,bool(g and g["master_property_id"]==pid),bool(pmap.get(pid)))
            ev=g["evidence"] if g else []
            c.execute(text("""UPDATE pi_properties SET
                plot_area_sqft=:plot,builtup_area_sqft=:built,occupancy_status=:occ,offering_transaction=:tx,
                v8_identity_key=:ik,v8_locality_key=:lk,v8_duplicate_confidence=:dc,v8_duplicate_evidence=CAST(:ev AS JSONB),
                v8_master_property_id=:mid,v8_bucket=:bucket,v8_match_eligible=:me,v8_quality_score=:score,
                v8_review_reasons=CAST(:reasons AS JSONB),v8_updated_at=NOW() WHERE property_id=:id"""),
                {"plot":cl["plot"],"built":cl["built"],"occ":cl["occupancy"],"tx":cl["offering"],"ik":cl["identity"],"lk":cl["locality"],
                 "dc":"STRONG" if g else ("POSSIBLE" if pmap.get(pid) else "UNIQUE"),"ev":json.dumps(ev),
                 "mid":g["master_property_id"] if g else None,"bucket":cl["bucket"],"me":cl["match"],"score":cl["score"],
                 "reasons":json.dumps(cl["reasons"]),"id":pid})
    return _v8_audit()

@app.get("/api/v8/master-engine/audit")
def v8_master_engine_audit(req:Request):
    need_login(req);return {"status":"ok",**_v8_audit()}

@app.post("/api/v8/master-engine/apply")
def v8_master_engine_apply(req:Request):
    need_login(req);return {"status":"ok","message":"V8 applied safely. No records deleted; possible duplicates remain review-only.",**_v8_apply()}

@app.get("/smart-master-data",response_class=HTMLResponse)
def v8_smart_master_data_page(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",status_code=303)
    return HTMLResponse(r"""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>V8 Smart Master Data</title>
<style>*{box-sizing:border-box}body{margin:0;background:#f4f7fb;font-family:Arial;color:#142033}header{background:#0d1d2d;color:#fff;padding:16px 22px}.wrap{max-width:1900px;margin:auto;padding:18px}.card,.kpi{background:#fff;border:1px solid #e4eaf1;border-radius:12px;padding:14px;margin-bottom:12px}.kpis{display:grid;grid-template-columns:repeat(4,minmax(150px,1fr));gap:10px}.kpi b{display:block;font-size:25px}.btn{display:inline-block;padding:9px 12px;border:0;border-radius:8px;background:#1677ff;color:#fff;text-decoration:none;font-weight:700;cursor:pointer}.orange{background:#df8b13}.gray{background:#edf2f7;color:#24364b}.tablewrap{overflow:auto;max-height:65vh}table{width:100%;border-collapse:collapse;font-size:12px}th,td{padding:8px;border-bottom:1px solid #edf1f5;text-align:left;vertical-align:top;white-space:nowrap}th{position:sticky;top:0;background:#f8fafc}</style></head><body>
<header><b>V8 Smart Master Data Engine</b><br><small>Conservative duplicates · area correction · transaction correction · matcher-ready inventory</small></header>
<div class="wrap"><div class="card"><a class="btn gray" href="/workspace">Workspace</a> <a class="btn gray" href="/property-database">Property Database</a> <button class="btn" onclick="audit()">Run V8 Audit</button> <button class="btn orange" onclick="apply()">Apply V8</button> <span id="msg"></span></div>
<div class="card"><b>Rules:</b> shared broker phone alone never creates a duplicate. Strong duplicate requires specific property identity + same locality + compatible area. Possible duplicates stay review-only. No records are deleted.</div>
<div class="kpis" id="k"></div><div class="card"><div class="tablewrap"><table><thead><tr><th>Property</th><th>Bucket</th><th>Score</th><th>Identity</th><th>Locality</th><th>City</th><th>Type</th><th>Plot</th><th>Built-up</th><th>Occupancy</th><th>Offer Tx</th><th>Master</th><th>Reasons</th><th>Open</th></tr></thead><tbody id="r"></tbody></table></div></div></div>
<script>
const E=x=>String(x??'').replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));
async function A(u,o={}){let r=await fetch(u,o),d=await r.json();if(!r.ok)throw new Error(d.detail||d.message||'Error');return d}
function R(d){let s=d.summary;document.querySelector('#k').innerHTML=[['TOTAL',s.total],['MATCH READY',s.match_ready],['CONTACT REVIEW',s.contact_review],['DATA REVIEW',s.data_review],['STRONG DUP REVIEW',s.duplicate_review],['POSSIBLE DUP REVIEW',s.possible_duplicate_review],['STRONG DUP GROUPS',s.strong_duplicate_groups],['STRONG DUP RECORDS',s.strong_duplicate_records],['POSSIBLE DUP PAIRS',s.possible_duplicate_pairs]].map(x=>`<div class="kpi"><span>${x[0]}</span><b>${Number(x[1]||0).toLocaleString()}</b></div>`).join('');
document.querySelector('#r').innerHTML=(d.reviewed||[]).map(x=>`<tr><td><b>${E(x.property_name||x.property_id)}</b><br>${E(x.property_id)}</td><td>${E(x.bucket)}</td><td>${E(x.score)}</td><td>${E(x.identity||'')}</td><td>${E(x.locality||'')}</td><td>${E(x.city||'')}</td><td>${E(x.property_type||'')}</td><td>${E(x.plot_area_sqft||'')}</td><td>${E(x.builtup_area_sqft||'')}</td><td>${E(x.occupancy_status||'')}</td><td>${E(x.offering_transaction||'')}</td><td>${E(x.master_property_id||'')}</td><td style="max-width:430px;white-space:normal">${E((x.reasons||[]).join(', '))}</td><td><a target="_blank" href="/property-record/${encodeURIComponent(x.property_id)}">View</a></td></tr>`).join('')}
async function audit(){document.querySelector('#msg').textContent=' Auditing...';R(await A('/api/v8/master-engine/audit'));document.querySelector('#msg').textContent=' Audit complete'}
async function apply(){if(!confirm('Apply V8? No records will be deleted.'))return;document.querySelector('#msg').textContent=' Applying...';R(await A('/api/v8/master-engine/apply',{method:'POST'}));document.querySelector('#msg').textContent=' Applied'}
audit();
</script></body></html>""")


# ============================================================
# V9 INTELLIGENT PROPERTY RECONSTRUCTION ENGINE
# Reconstructs structured property facts from existing raw/source text.
# Conservative, auditable, reversible-by-source. Never invents contacts.
# ============================================================

_V9_FLOOR_MAP = [
    (r'\b(?:bmt|basement|lgf|lower ground)\b',"BASEMENT"),
    (r'\b(?:gf|ground floor)\b',"GROUND FLOOR"),
    (r'\b(?:ff|first floor)\b',"FIRST FLOOR"),
    (r'\b(?:sf|second floor)\b',"SECOND FLOOR"),
    (r'\b(?:tf|third floor)\b',"THIRD FLOOR"),
    (r'\b(?:terr|terrace)\b',"TERRACE"),
]

_V9_TYPE_RULES = [
    (["shop","showroom","retail","store","dda mkt","market shop"],"RETAIL / COMMERCIAL"),
    (["restaurant","cafe","qsr","f&b","food outlet","lounge","club"],"F&B / COMMERCIAL"),
    (["office","business centre","cowork"],"OFFICE"),
    (["banquet","hotel","guest house","hospitality","farmhouse"],"HOSPITALITY"),
    (["warehouse","factory","industrial"],"INDUSTRIAL"),
    (["builder floor","4bhk","3bhk","2bhk","1bhk","apartment","flat","villa","duplex","residential"],"RESIDENTIAL"),
]

_V9_CITY_FROM_LOCALITY = {
    "kailash colony":"New Delhi","greater kailash 1":"New Delhi","greater kailash 2":"New Delhi",
    "new friends colony":"New Delhi","safdarjung enclave":"New Delhi","green park":"New Delhi",
    "vasant vihar":"New Delhi","vasant kunj":"New Delhi","hauz khas":"New Delhi",
    "panchsheel park":"New Delhi","panchsheel enclave":"New Delhi","defence colony":"New Delhi",
    "lajpat nagar":"New Delhi","jangpura":"New Delhi","nizamuddin east":"New Delhi",
    "nizamuddin west":"New Delhi","saket":"New Delhi","malviya nagar":"New Delhi",
    "chittaranjan park":"New Delhi","nehru place":"New Delhi","okhla":"New Delhi",
    "mathura road":"New Delhi","mohan co-operative":"New Delhi","mayapuri":"New Delhi",
    "pitampura":"New Delhi","sainik farm":"New Delhi","gulmohar park":"New Delhi",
    "jasola":"New Delhi","taimoor nagar":"New Delhi","niti bagh":"New Delhi",
    "gurugram":"Gurugram","noida":"Noida","greater noida":"Greater Noida",
    "faridabad":"Faridabad","ghaziabad":"Ghaziabad"
}

def _ensure_v9_columns():
    stmts=[
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v9_property_no TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v9_locality TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v9_city TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v9_floor TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v9_property_type TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v9_plot_area_sqft DOUBLE PRECISION",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v9_builtup_area_sqft DOUBLE PRECISION",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v9_transaction TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v9_occupancy TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v9_recovery_status TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v9_recovery_confidence INTEGER",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v9_recovery_evidence JSONB DEFAULT '{}'::jsonb",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v9_final_bucket TEXT",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v9_match_eligible BOOLEAN",
        "ALTER TABLE pi_properties ADD COLUMN IF NOT EXISTS v9_updated_at TIMESTAMP"
    ]
    with engine.begin() as c:
        for s in stmts:
            c.execute(text(s))

def _v9_raw(p):
    vals=[
        p.get("property_name"),p.get("address"),p.get("location"),p.get("micro_market"),
        p.get("city"),p.get("property_type"),p.get("remarks"),p.get("suitable_category"),
        p.get("source")
    ]
    return " | ".join(str(x or "") for x in vals if str(x or "").strip())

def _v9_property_no(p):
    for raw in [p.get("property_name"),p.get("address")]:
        s=_v8_clean(raw)
        if not s:
            continue
        m=_re.search(r'\b(?:shop\s*no\s*[- ]?\d+[a-z]?|unit\s*no\s*[- ]?\d+[a-z]?|[a-z]{1,3}\s*[-/]\s*\d+[a-z0-9/-]*|\d+[a-z]?\s+[a-z][a-z ]{2,25}(?:complex|tower|house|market|mkt))\b',s,re.I)
        if m:
            return _v8_clean(m.group(0)),98
    return None,0

def _v9_locality(p):
    # Never use property number itself as locality.
    raw_loc=_v8_clean(p.get("location") or p.get("micro_market"))
    prop_no,_=_v9_property_no(p)
    if raw_loc and raw_loc not in {"unknown","na","n a","none","not specified"} and raw_loc!=prop_no:
        return _v8_loc(p),98

    txt=_norm(_v9_raw(p))
    candidates=sorted(_V9_CITY_FROM_LOCALITY.keys(),key=len,reverse=True)
    aliases={"kalash colony":"kailash colony","gk 1":"greater kailash 1","gk1":"greater kailash 1",
             "gk 2":"greater kailash 2","gk2":"greater kailash 2","nfc":"new friends colony",
             "defense colony":"defence colony","safdurjung enclave":"safdarjung enclave",
             "chatterpur":"chattarpur","gurgaon":"gurugram"}
    for a,b in aliases.items():
        if a in txt:
            return b,96
    for loc in candidates:
        if loc in txt:
            return loc,96
    return None,0

def _v9_city(p,locality):
    c=_canonical_city_v4(p.get("city"))
    if c!="UNKNOWN":
        return _dq_text(p.get("city")),99
    if locality and locality in _V9_CITY_FROM_LOCALITY:
        return _V9_CITY_FROM_LOCALITY[locality],96
    txt=_norm(_v9_raw(p))
    if "delhi" in txt:return "New Delhi",95
    if "gurugram" in txt or "gurgaon" in txt:return "Gurugram",96
    if "greater noida" in txt:return "Greater Noida",96
    if "noida" in txt:return "Noida",96
    if "faridabad" in txt:return "Faridabad",96
    if "ghaziabad" in txt:return "Ghaziabad",96
    return None,0

def _v9_floor(p):
    txt=_norm(_v9_raw(p))
    found=[]
    for pat,label in _V9_FLOOR_MAP:
        if _re.search(pat,txt,re.I):
            found.append(label)
    return " + ".join(dict.fromkeys(found)) if found else None

def _v9_property_type(p):
    existing=_canonical_property_type_v4(p.get("property_type"),p.get("suitable_category"),p.get("remarks"))
    if existing!="UNKNOWN":
        # But floor shorthand accidentally stored as type remains invalid.
        raw=_norm(p.get("property_type"))
        if not _re.fullmatch(r'(bmt|basement|lgf|gf|ff|sf|tf|terr|bmt gf|bmt ff|bmt sf|bmt tf|sf terr|tf terr|bmt\+gf|bmt\+ff|bmt\+sf|bmt\+tf|sf\+terr|tf\+terr)',raw):
            return existing,99

    txt=_norm(_v9_raw(p))
    for terms,label in _V9_TYPE_RULES:
        if any(t in txt for t in terms):
            return label,96
    return None,0

def _v9_area(p):
    a=_v8_area(p)
    plot=a["plot"]
    built=a["built"]
    conf=98 if built else (95 if plot else 0)
    return plot,built,conf,a["reasons"]

def _v9_tx(p):
    t=_v8_tx(p)
    return t["offering"],t["occupancy"],(98 if t["offering"] else 0),t["reasons"]

def _v9_reconstruct(p):
    prop_no,pc=_v9_property_no(p)
    loc,lc=_v9_locality(p)
    city,cc=_v9_city(p,loc)
    ptype,tc=_v9_property_type(p)
    floor=_v9_floor(p)
    plot,built,ac,area_reasons=_v9_area(p)
    tx,occ,xc,tx_reasons=_v9_tx(p)
    q=_organize_property_v4(p)

    evidence={
        "property_no_confidence":pc,"locality_confidence":lc,"city_confidence":cc,
        "type_confidence":tc,"area_confidence":ac,"transaction_confidence":xc,
        "area_notes":area_reasons,"transaction_notes":tx_reasons
    }

    recovered=[]
    if (not p.get("location") or _dq_unknown(p.get("location"))) and loc:recovered.append("LOCALITY")
    if _canonical_city_v4(p.get("city"))=="UNKNOWN" and city:recovered.append("CITY")
    if _canonical_property_type_v4(p.get("property_type"),p.get("suitable_category"),p.get("remarks"))=="UNKNOWN" and ptype:recovered.append("PROPERTY_TYPE")
    if _property_area(p) is None and built:recovered.append("BUILTUP_AREA")
    if _canonical_transaction_v4(p.get("rent_or_sale"))=="UNKNOWN" and tx:recovered.append("TRANSACTION")

    core_ok=bool(city and loc and ptype and built and tx)
    if core_ok and q["contact_ready"]:
        final_bucket="MATCH_READY"
    elif core_ok:
        final_bucket="CONTACT_REVIEW"
    else:
        final_bucket="HUMAN_REVIEW"

    if p.get("v8_bucket")=="DUPLICATE_REVIEW":
        final_bucket="CONFIRMED_DUPLICATE"
    elif p.get("v8_bucket")=="POSSIBLE_DUPLICATE_REVIEW":
        final_bucket="POSSIBLE_DUPLICATE"

    confs=[x for x in [pc,lc,cc,tc,ac,xc] if x]
    overall=round(sum(confs)/len(confs)) if confs else 0

    return {
        "property_no":prop_no,"locality":loc,"city":city,"floor":floor,
        "property_type":ptype,"plot_area_sqft":plot,"builtup_area_sqft":built,
        "transaction":tx,"occupancy":occ,
        "recovered_fields":recovered,"evidence":evidence,
        "recovery_status":"AUTO_RECOVERED" if recovered else "NO_AUTO_RECOVERY",
        "recovery_confidence":overall,
        "final_bucket":final_bucket,
        "match_eligible":final_bucket=="MATCH_READY"
    }

def _v9_audit(limit=1500):
    _ensure_v9_columns()
    with engine.connect() as c:
        props=[dict(r._mapping) for r in c.execute(text("SELECT * FROM pi_properties ORDER BY created_at DESC")).fetchall()]
    s={"total":len(props),"auto_recovered":0,"match_ready":0,"contact_review":0,"human_review":0,
       "confirmed_duplicate":0,"possible_duplicate":0}
    rows=[]
    for p in props:
        r=_v9_reconstruct(p)
        if r["recovery_status"]=="AUTO_RECOVERED":s["auto_recovered"]+=1
        s[r["final_bucket"].lower()]+=1
        if len(rows)<limit and (r["recovery_status"]=="AUTO_RECOVERED" or r["final_bucket"]!="MATCH_READY"):
            rows.append({"property_id":str(p.get("property_id")),"property_name":p.get("property_name"),**r})
    return {"summary":s,"rows":rows}

def _v9_apply():
    audit=_v9_audit(1500)
    with engine.connect() as c:
        props=[dict(r._mapping) for r in c.execute(text("SELECT * FROM pi_properties")).fetchall()]
    updated=0
    with engine.begin() as c:
        for p in props:
            r=_v9_reconstruct(p)
            params={
                "pno":r["property_no"],"loc":r["locality"],"city":r["city"],"floor":r["floor"],"ptype":r["property_type"],
                "plot":r["plot_area_sqft"],"built":r["builtup_area_sqft"],"tx":r["transaction"],"occ":r["occupancy"],
                "rs":r["recovery_status"],"rc":r["recovery_confidence"],"ev":json.dumps(r["evidence"]),
                "bucket":r["final_bucket"],"me":r["match_eligible"],"id":p.get("property_id")
            }
            c.execute(text("""UPDATE pi_properties SET
                v9_property_no=:pno,v9_locality=:loc,v9_city=:city,v9_floor=:floor,v9_property_type=:ptype,
                v9_plot_area_sqft=:plot,v9_builtup_area_sqft=:built,v9_transaction=:tx,v9_occupancy=:occ,
                v9_recovery_status=:rs,v9_recovery_confidence=:rc,v9_recovery_evidence=CAST(:ev AS JSONB),
                v9_final_bucket=:bucket,v9_match_eligible=:me,v9_updated_at=NOW()
                WHERE property_id=:id"""),params)

            # Apply only conservative 95+ structured recoveries into legacy fields.
            sets=[];p2={"id":p.get("property_id")}
            ev=r["evidence"]
            if (not p.get("location") or _dq_unknown(p.get("location"))) and r["locality"] and ev["locality_confidence"]>=95:
                sets.append("location=:location");p2["location"]=r["locality"]
            if _canonical_city_v4(p.get("city"))=="UNKNOWN" and r["city"] and ev["city_confidence"]>=95:
                sets.append("city=:city");p2["city"]=r["city"]
            if _canonical_property_type_v4(p.get("property_type"),p.get("suitable_category"),p.get("remarks"))=="UNKNOWN" and r["property_type"] and ev["type_confidence"]>=95:
                sets.append("property_type=:ptype");p2["ptype"]=r["property_type"]
            if _property_area(p) is None and r["builtup_area_sqft"] and ev["area_confidence"]>=95:
                sets.append("available_area_sqft=:area");p2["area"]=r["builtup_area_sqft"]
            if _canonical_transaction_v4(p.get("rent_or_sale"))=="UNKNOWN" and r["transaction"] and ev["transaction_confidence"]>=95:
                sets.append("rent_or_sale=:tx");p2["tx"]=r["transaction"]
            if sets:
                sets.append("updated_at=NOW()")
                c.execute(text("UPDATE pi_properties SET "+",".join(sets)+" WHERE property_id=:id"),p2)
                updated+=1
    return {"updated_records":updated,**_v9_audit(1500)}

@app.get("/api/v9/reconstruction/audit")
def v9_reconstruction_audit(req:Request):
    need_login(req)
    return {"status":"ok",**_v9_audit()}

@app.post("/api/v9/reconstruction/apply")
def v9_reconstruction_apply(req:Request):
    need_login(req)
    result=_v9_apply()
    return {"status":"ok","message":"V9 reconstruction applied conservatively. Original records retained; no phone numbers invented.",**result}

@app.get("/property-reconstruction",response_class=HTMLResponse)
def v9_property_reconstruction_page(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",status_code=303)
    return HTMLResponse(r"""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>V9 Property Reconstruction</title>
<style>*{box-sizing:border-box}body{margin:0;background:#f4f7fb;font-family:Arial;color:#142033}header{background:#0d1d2d;color:#fff;padding:16px 22px}.wrap{max-width:1950px;margin:auto;padding:18px}.card,.kpi{background:#fff;border:1px solid #e4eaf1;border-radius:12px;padding:14px;margin-bottom:12px}.kpis{display:grid;grid-template-columns:repeat(4,minmax(160px,1fr));gap:10px}.kpi b{display:block;font-size:25px}.btn{display:inline-block;padding:9px 12px;border:0;border-radius:8px;background:#1677ff;color:#fff;text-decoration:none;font-weight:700;cursor:pointer}.orange{background:#df8b13}.gray{background:#edf2f7;color:#24364b}.tablewrap{overflow:auto;max-height:68vh}table{width:100%;border-collapse:collapse;font-size:12px}th,td{padding:8px;border-bottom:1px solid #edf1f5;text-align:left;vertical-align:top;white-space:nowrap}th{position:sticky;top:0;background:#f8fafc}.MATCH_READY{color:#08734b;font-weight:700}.HUMAN_REVIEW,.CONFIRMED_DUPLICATE{color:#b23b00;font-weight:700}.CONTACT_REVIEW,.POSSIBLE_DUPLICATE{color:#9b6a00;font-weight:700}</style></head><body>
<header><b>V9 Intelligent Property Reconstruction Engine</b><br><small>Property no · locality · city · floor · type · area · transaction · duplicate-aware reconstruction</small></header>
<div class="wrap"><div class="card"><a class="btn gray" href="/workspace">Workspace</a> <a class="btn gray" href="/property-database">Property Database</a> <a class="btn gray" href="/smart-master-data">V8 Master Data</a> <button class="btn" onclick="audit()">Run V9 Audit</button> <button class="btn orange" onclick="apply()">Apply V9 Reconstruction</button> <span id="msg"></span></div>
<div class="card"><b>Safety:</b> floor abbreviations are separated from property type. Property numbers are never used as locality. Only 95%+ recoveries write back to core fields. Contacts are never invented.</div>
<div class="kpis" id="k"></div>
<div class="card"><div class="tablewrap"><table><thead><tr><th>Property</th><th>Final Bucket</th><th>Recovery</th><th>Confidence</th><th>Property No</th><th>Locality</th><th>City</th><th>Floor</th><th>Type</th><th>Plot</th><th>Built-up</th><th>Transaction</th><th>Occupancy</th><th>Recovered Fields</th><th>Open</th></tr></thead><tbody id="r"></tbody></table></div></div></div>
<script>
const E=x=>String(x??'').replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));
async function A(u,o={}){let r=await fetch(u,o),d=await r.json();if(!r.ok)throw new Error(d.detail||d.message||'Error');return d}
function R(d){let s=d.summary;document.querySelector('#k').innerHTML=[['TOTAL',s.total],['AUTO RECOVERED',s.auto_recovered],['MATCH READY',s.match_ready],['CONTACT REVIEW',s.contact_review],['HUMAN REVIEW',s.human_review],['CONFIRMED DUPLICATE',s.confirmed_duplicate],['POSSIBLE DUPLICATE',s.possible_duplicate]].map(x=>`<div class="kpi"><span>${x[0]}</span><b>${Number(x[1]||0).toLocaleString()}</b></div>`).join('');
document.querySelector('#r').innerHTML=(d.rows||[]).map(x=>`<tr><td><b>${E(x.property_name||x.property_id)}</b><br>${E(x.property_id)}</td><td class="${E(x.final_bucket)}">${E(x.final_bucket)}</td><td>${E(x.recovery_status)}</td><td>${E(x.recovery_confidence)}</td><td>${E(x.property_no||'')}</td><td>${E(x.locality||'')}</td><td>${E(x.city||'')}</td><td>${E(x.floor||'')}</td><td>${E(x.property_type||'')}</td><td>${E(x.plot_area_sqft||'')}</td><td>${E(x.builtup_area_sqft||'')}</td><td>${E(x.transaction||'')}</td><td>${E(x.occupancy||'')}</td><td>${E((x.recovered_fields||[]).join(', '))}</td><td><a target="_blank" href="/property-record/${encodeURIComponent(x.property_id)}">View</a></td></tr>`).join('')}
async function audit(){document.querySelector('#msg').textContent=' Auditing...';R(await A('/api/v9/reconstruction/audit'));document.querySelector('#msg').textContent=' Audit complete'}
async function apply(){if(!confirm('Apply V9 conservative reconstruction?'))return;document.querySelector('#msg').textContent=' Applying...';R(await A('/api/v9/reconstruction/apply',{method:'POST'}));document.querySelector('#msg').textContent=' Applied'}
audit();
</script></body></html>""")


# ============================================================
# V10 UNIVERSAL PROPERTY INTAKE
# Camera/newspaper/handwritten/WhatsApp/PDF ingestion + V9 reconstruction.
# ============================================================

V10_VISION_PROMPT = """You are the Property Intelligence V10 multimodal extraction engine.

INPUT MAY BE:
- a photograph of a newspaper or property magazine
- a phone camera photo of handwritten property notes
- a WhatsApp screenshot
- a printed classified page
- a scan/PDF/image containing multiple property listings or requirements

YOUR JOB:
1. Extract EVERY distinct property listing and EVERY distinct property/retail/hospitality requirement visible.
2. Never merge neighboring advertisements or handwritten rows unless they clearly describe the same property.
3. Read handwritten text conservatively. If a word/digit is unclear, do NOT invent it.
4. Phone numbers:
   - preserve every fully legible phone/mobile number
   - never pad, complete, or guess truncated digits
   - if multiple valid contacts are visible for the same record, preserve them in the relevant contact field separated by " | "
   - short fragments remain only in remarks, prefixed UNCLEAR_CONTACT:
5. Separate concepts correctly:
   - property/unit number is NOT locality
   - BMT=Basement, LGF=Lower Ground Floor, GF=Ground Floor, FF=First Floor, SF=Second Floor, TF=Third Floor, TERR=Terrace
   - these floor abbreviations are NOT property types
   - SQYD/YD is plot area unless the source explicitly says built-up/covered/usable area
   - "rented/preleased" is occupancy, not automatically an offer for rent
   - "leasehold/freehold" is tenure, not transaction
6. Use null/unknown when the source does not support a fact.
7. Keep exact source clues in remarks when they help later reconstruction.
8. For newspaper grids and dense magazine pages, scan the entire image and return all distinct records, not a sample.
9. For handwritten notes, treat each bullet/line/row as a potential separate record.
10. For requirements, capture company/client/contact/location/area/property type/lease-or-sale and the requirement wording.

Return only the required structured schema. No prose outside JSON.
"""

def _ensure_v10_tables():
    with engine.begin() as c:
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_v10_intake_log(
            id BIGSERIAL PRIMARY KEY,
            intake_id UUID UNIQUE NOT NULL,
            source_id BIGINT,
            job_id BIGINT,
            source_type TEXT,
            original_filename TEXT,
            capture_mode TEXT,
            status TEXT DEFAULT 'ACCEPTED',
            file_size BIGINT,
            note TEXT,
            created_by TEXT,
            created_at TIMESTAMPTZ DEFAULT NOW(),
            completed_at TIMESTAMPTZ
        )"""))
        c.execute(text("CREATE INDEX IF NOT EXISTS idx_pi_v10_intake_source ON pi_v10_intake_log(source_id)"))

def _v10_extract_image(path,sid,jid,page_number=1):
    """
    Full image + overlapping tiles using the V10 vision prompt.
    Reuses current exhaustive scan architecture but with handwriting/newspaper rules.
    """
    total={"created":0,"duplicates":0,"property_outputs":0,"requirement_outputs":0,"failed":0}
    units=[("FULL_PAGE",path,False)]+[(label,tp,True) for label,tp in crop_overlapping_tiles(path)]
    for label,img,is_temp in units:
        try:
            state=scan_tile_state(sid,page_number,"V10_"+label)
            if state and state.get("status")=="COMPLETED":
                continue
            mark_scan_tile(sid,page_number,"V10_"+label,"RUNNING")
            prompt=(
                V10_VISION_PROMPT
                + "\nIMAGE REGION: "+label
                + "\nOverlapping crops are intentional. Extract every distinct record visible in this region."
            )
            env=extract_gemini_batch(img,"image/jpeg",prompt)
            cr,du,po,ro=save_scanned_envelope(env,sid,f"V10_PAGE_{page_number}_{label}")
            mark_scan_tile(sid,page_number,"V10_"+label,"COMPLETED",cr,du,None)
            total["created"]+=cr;total["duplicates"]+=du
            total["property_outputs"]+=po;total["requirement_outputs"]+=ro
        except Exception as exc:
            mark_scan_tile(sid,page_number,"V10_"+label,"FAILED",0,0,str(exc))
            total["failed"]+=1
        finally:
            if is_temp:
                try:os.unlink(img)
                except Exception:pass
    return total

def _v10_reconstruct_source_records(sid):
    """Run V9 reconstruction only on properties created from this source."""
    _ensure_v9_columns()
    with engine.connect() as c:
        props=[dict(r._mapping) for r in c.execute(
            text("SELECT * FROM pi_properties WHERE source_id=:sid ORDER BY id"),
            {"sid":sid}
        ).fetchall()]
    updated=0
    with engine.begin() as c:
        for p in props:
            r=_v9_reconstruct(p)
            c.execute(text("""UPDATE pi_properties SET
                v9_property_no=:pno,v9_locality=:loc,v9_city=:city,v9_floor=:floor,v9_property_type=:ptype,
                v9_plot_area_sqft=:plot,v9_builtup_area_sqft=:built,v9_transaction=:tx,v9_occupancy=:occ,
                v9_recovery_status=:rs,v9_recovery_confidence=:rc,v9_recovery_evidence=CAST(:ev AS JSONB),
                v9_final_bucket=:bucket,v9_match_eligible=:me,v9_updated_at=NOW()
                WHERE property_id=:id"""),{
                "pno":r["property_no"],"loc":r["locality"],"city":r["city"],"floor":r["floor"],
                "ptype":r["property_type"],"plot":r["plot_area_sqft"],"built":r["builtup_area_sqft"],
                "tx":r["transaction"],"occ":r["occupancy"],"rs":r["recovery_status"],
                "rc":r["recovery_confidence"],"ev":json.dumps(r["evidence"]),
                "bucket":r["final_bucket"],"me":r["match_eligible"],"id":p.get("property_id")
            })
            updated+=1
    return updated

def _v10_file_worker(sid,jid,path,mime,intake_id,capture_mode):
    try:
        created=duplicates=failed=0
        is_pdf=(mime=="application/pdf" or path.lower().endswith(".pdf"))
        is_image=(mime or "").startswith("image/") or path.lower().endswith((".jpg",".jpeg",".png",".webp"))

        if is_pdf:
            doc=fitz.open(path)
            for i in range(doc.page_count):
                page_img=None
                try:
                    page_img=render_pdf_page(doc,i)
                    r=_v10_extract_image(page_img,sid,jid,i+1)
                    created+=r["created"];duplicates+=r["duplicates"];failed+=r["failed"]
                finally:
                    if page_img:
                        try:os.unlink(page_img)
                        except Exception:pass
            doc.close()
        elif is_image:
            fd,jpg=tempfile.mkstemp(suffix=".jpg");os.close(fd)
            try:
                Image.open(path).convert("RGB").save(jpg,"JPEG",quality=96)
                r=_v10_extract_image(jpg,sid,jid,1)
                created=r["created"];duplicates=r["duplicates"];failed=r["failed"]
            finally:
                try:os.unlink(jpg)
                except Exception:pass
        else:
            env=extract_gemini_batch(path,mime,V10_VISION_PROMPT)
            created,duplicates,_,_=save_scanned_envelope(env,sid,"V10_SOURCE")

        reconstructed=_v10_reconstruct_source_records(sid)
        status="PROCESSED_WITH_ERRORS" if failed else "PROCESSED"
        with engine.begin() as c:
            c.execute(text("""UPDATE pi_sources SET ingestion_status=:st,processed_records=:n,duplicate_records=:d,
                ai_provider='gemini',ai_model=:m,processed_at=NOW() WHERE id=:id"""),
                {"st":status,"n":created,"d":duplicates,"m":GEMINI_MODEL,"id":sid})
            c.execute(text("""UPDATE pi_ai_jobs SET status='COMPLETED',output_summary=:o,completed_at=NOW()
                WHERE id=:id"""),{
                "o":f"V10 intake: {created} new, {duplicates} duplicate/overlap, {failed} failed scan units, {reconstructed} reconstructed",
                "id":jid
            })
            c.execute(text("""UPDATE pi_v10_intake_log SET status=:st,completed_at=NOW()
                WHERE intake_id=CAST(:iid AS UUID)"""),{"st":status,"iid":intake_id})
    except Exception as ex:
        with engine.begin() as c:
            c.execute(text("UPDATE pi_sources SET ingestion_status='FAILED',error_message=:e WHERE id=:id"),{"e":str(ex),"id":sid})
            c.execute(text("UPDATE pi_ai_jobs SET status='FAILED',error_message=:e,completed_at=NOW() WHERE id=:id"),{"e":str(ex),"id":jid})
            c.execute(text("""UPDATE pi_v10_intake_log SET status='FAILED',note=:e,completed_at=NOW()
                WHERE intake_id=CAST(:iid AS UUID)"""),{"e":str(ex),"iid":intake_id})
    finally:
        try:os.unlink(path)
        except Exception:pass

@app.post("/api/v10/intake/file")
async def v10_intake_file(
    bg:BackgroundTasks,
    req:Request,
    file:UploadFile=File(...),
    source_type:str=Form("PHOTO"),
    capture_mode:str=Form("CAMERA"),
    note:Optional[str]=Form(None)
):
    need_login(req)
    _ensure_v10_tables()

    filename=file.filename or "camera-photo.jpg"
    ext=os.path.splitext(filename)[1].lower()
    mime=(file.content_type or "").lower()
    mimemap={
        ".jpg":"image/jpeg",".jpeg":"image/jpeg",".png":"image/png",".webp":"image/webp",
        ".pdf":"application/pdf"
    }
    if not mime or mime=="application/octet-stream":
        mime=mimemap.get(ext,"application/octet-stream")

    allowed={"image/jpeg","image/png","image/webp","application/pdf"}
    if mime not in allowed:
        raise HTTPException(400,"V10 accepts JPG, JPEG, PNG, WEBP or PDF.")

    fd,path=tempfile.mkstemp(suffix=ext or ".bin");os.close(fd)
    total=0
    try:
        with open(path,"wb") as out:
            while True:
                chunk=await file.read(1024*1024)
                if not chunk:break
                total+=len(chunk)
                if total>MAX_UPLOAD_MB*1024*1024:
                    raise HTTPException(413,f"Maximum upload is {MAX_UPLOAD_MB} MB.")
                out.write(chunk)

        sid=source_row(source_type.upper(),filename,filename,mime,reference=note)
        jid=create_job(sid,"V10_MULTIMODAL_EXTRACTION",f"{source_type} | {capture_mode} | {filename}")
        iid=str(uuid.uuid4())
        with engine.begin() as c:
            c.execute(text("""INSERT INTO pi_v10_intake_log(
                intake_id,source_id,job_id,source_type,original_filename,capture_mode,status,file_size,note,created_by
            ) VALUES(CAST(:iid AS UUID),:sid,:jid,:st,:fn,:cm,'ACCEPTED',:sz,:note,:by)"""),{
                "iid":iid,"sid":sid,"jid":jid,"st":source_type.upper(),"fn":filename,
                "cm":capture_mode.upper(),"sz":total,"note":note,"by":actor_name(req)
            })
        bg.add_task(_v10_file_worker,sid,jid,path,mime,iid,capture_mode)
        return {"status":"ACCEPTED","intake_id":iid,"source_id":sid,"job_id":jid,
                "message":"Photo/document received. V10 processing continues in background."}
    except Exception:
        try:os.unlink(path)
        except Exception:pass
        raise

@app.post("/api/v10/intake/text")
async def v10_intake_text(req:Request,bg:BackgroundTasks):
    need_login(req)
    body=await req.json()
    content=str(body.get("text") or "").strip()
    if not content:
        raise HTTPException(400,"Text is required.")
    source_type=str(body.get("source_type") or "HANDWRITTEN_TRANSCRIPTION").upper()
    name=str(body.get("source_name") or "Manual text intake")
    sid=source_row(source_type,name,reference=content)
    jid=create_job(sid,"V10_TEXT_EXTRACTION",name)
    # Existing text extraction is already structured; V10 prompt is prepended.
    bg.add_task(run_text_job,sid,jid,V10_VISION_PROMPT+"\nSOURCE TEXT:\n"+content)
    return {"status":"ACCEPTED","source_id":sid,"job_id":jid}

@app.get("/api/v10/intake/status")
def v10_intake_status(req:Request,limit:int=Query(100,ge=1,le=500)):
    need_login(req);_ensure_v10_tables()
    with engine.connect() as c:
        rows=c.execute(text("""SELECT l.*,s.processed_records,s.duplicate_records,s.error_message,
            j.output_summary FROM pi_v10_intake_log l
            LEFT JOIN pi_sources s ON s.id=l.source_id
            LEFT JOIN pi_ai_jobs j ON j.id=l.job_id
            ORDER BY l.created_at DESC LIMIT :n"""),{"n":limit}).fetchall()
    return {"status":"ok","rows":_json_rows(rows)}

@app.get("/capture-intelligence",response_class=HTMLResponse)
def v10_capture_page(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",status_code=303)
    return HTMLResponse(r"""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>V10 Universal Property Intake</title>
<style>*{box-sizing:border-box}body{margin:0;background:#f4f7fb;font-family:Arial;color:#142033}header{background:#0d1d2d;color:#fff;padding:16px 22px}.wrap{max-width:1250px;margin:auto;padding:18px}.card{background:#fff;border:1px solid #e4eaf1;border-radius:12px;padding:16px;margin-bottom:12px}.grid{display:grid;grid-template-columns:1fr 1fr;gap:12px}.btn{display:inline-block;padding:11px 14px;border:0;border-radius:8px;background:#1677ff;color:#fff;text-decoration:none;font-weight:700;cursor:pointer}.gray{background:#edf2f7;color:#24364b}input,select,textarea{width:100%;padding:11px;border:1px solid #ccd7e4;border-radius:7px;margin:5px 0 11px}textarea{min-height:130px}.drop{padding:25px;border:2px dashed #aebdcd;border-radius:10px;text-align:center;background:#fafcff}.tablewrap{overflow:auto}table{width:100%;border-collapse:collapse;font-size:12px}th,td{padding:8px;border-bottom:1px solid #edf1f5;text-align:left}@media(max-width:760px){.grid{grid-template-columns:1fr}}</style></head><body>
<header><b>V10 Universal Property Intake</b><br><small>Camera · Newspaper · Handwritten Note · WhatsApp Screenshot · PDF</small></header>
<div class="wrap">
<div class="card"><a class="btn gray" href="/workspace">Workspace</a> <a class="btn gray" href="/property-database">Property Database</a> <a class="btn gray" href="/property-reconstruction">V9 Reconstruction</a></div>
<div class="grid">
<div class="card"><h3>Take Photo / Upload Image</h3><p>Use your phone camera for newspaper ads or handwritten notes. Clear, straight photos work best.</p>
<form id="photoForm"><label>Source Type</label><select name="source_type"><option>NEWSPAPER</option><option>HANDWRITTEN</option><option>WHATSAPP_SCREENSHOT</option><option>MAGAZINE</option><option>PHOTO</option><option>PDF</option></select>
<label>Photo / PDF</label><div class="drop"><input name="file" type="file" accept="image/*,.pdf" capture="environment" required></div>
<label>Optional note</label><input name="note" placeholder="e.g. Sunday newspaper page / broker handwritten sheet">
<input type="hidden" name="capture_mode" value="CAMERA_OR_UPLOAD"><button class="btn" type="submit">Upload & Extract</button></form><p id="photoMsg"></p></div>
<div class="card"><h3>Paste Message / Typed Note</h3><select id="textType"><option>WHATSAPP</option><option>HANDWRITTEN_TRANSCRIPTION</option><option>MANUAL_NOTE</option></select>
<textarea id="textContent" placeholder="Paste property message or manually type a handwritten note..."></textarea>
<button class="btn" onclick="sendText()">Extract Text</button><p id="textMsg"></p></div>
</div>
<div class="card"><h3>Recent Intake Jobs</h3><button class="btn gray" onclick="load()">Refresh Status</button><div class="tablewrap"><table><thead><tr><th>Time</th><th>Type</th><th>File</th><th>Status</th><th>New Records</th><th>Duplicates</th><th>Result</th></tr></thead><tbody id="rows"></tbody></table></div></div>
</div>
<script>
const E=x=>String(x??'').replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));
async function load(){let r=await fetch('/api/v10/intake/status'),d=await r.json();document.querySelector('#rows').innerHTML=(d.rows||[]).map(x=>`<tr><td>${E(x.created_at||'')}</td><td>${E(x.source_type||'')}</td><td>${E(x.original_filename||'')}</td><td>${E(x.status||'')}</td><td>${E(x.processed_records||0)}</td><td>${E(x.duplicate_records||0)}</td><td>${E(x.output_summary||x.error_message||'')}</td></tr>`).join('')}
document.querySelector('#photoForm').addEventListener('submit',async e=>{e.preventDefault();document.querySelector('#photoMsg').textContent='Uploading...';let r=await fetch('/api/v10/intake/file',{method:'POST',body:new FormData(e.target)}),d=await r.json();document.querySelector('#photoMsg').textContent=r.ok?'Accepted. AI extraction is running.':(d.detail||d.message||'Upload failed');if(r.ok){e.target.reset();load()}})
async function sendText(){let t=document.querySelector('#textContent').value.trim();if(!t)return;document.querySelector('#textMsg').textContent='Sending...';let r=await fetch('/api/v10/intake/text',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({source_type:document.querySelector('#textType').value,text:t})}),d=await r.json();document.querySelector('#textMsg').textContent=r.ok?'Accepted. AI extraction is running.':(d.detail||d.message||'Failed');if(r.ok){document.querySelector('#textContent').value=''}}
load();setInterval(load,15000);
</script></body></html>""")


# ============================================================
# FINAL TEAM OPERATIONS V2
# Unified Property Contacts -> Manual verification -> Owner/Broker segregation
# ============================================================

_CONTACT_SYNC_RUNNING=False
_CONTACT_SYNC_PENDING=False

def _ensure_unified_contact_directory():
    with engine.begin() as c:
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_contact_directory_v2(
            contact_key TEXT PRIMARY KEY,
            display_name TEXT,
            primary_phone TEXT,
            phones JSONB DEFAULT '[]'::jsonb,
            property_ids JSONB DEFAULT '[]'::jsonb,
            property_names JSONB DEFAULT '[]'::jsonb,
            locations JSONB DEFAULT '[]'::jsonb,
            cities JSONB DEFAULT '[]'::jsonb,
            sources JSONB DEFAULT '[]'::jsonb,
            property_count INTEGER DEFAULT 0,
            verified_property_count INTEGER DEFAULT 0,
            contact_role TEXT DEFAULT 'UNVERIFIED',
            verification_status TEXT DEFAULT 'UNVERIFIED',
            verified_by TEXT,
            verified_at TIMESTAMPTZ,
            remarks TEXT,
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )"""))
        c.execute(text("CREATE INDEX IF NOT EXISTS idx_contact_v2_phone ON pi_contact_directory_v2(primary_phone)"))
        c.execute(text("CREATE INDEX IF NOT EXISTS idx_contact_v2_role ON pi_contact_directory_v2(contact_role)"))

def _unified_contact_key(name,phones):
    phones=sorted(set(phones or []))
    if phones:
        return "PHONE:"+phones[0]
    n=_norm(name)
    if n and n not in {"na","n a","unknown","none","not specified"}:
        return "NAME:"+hashlib.sha256(n.encode()).hexdigest()[:24]
    return None

def _sync_unified_contacts():
    """
    Merge owner, broker and general property contacts into ONE directory first.
    Existing manually verified roles are preserved.
    """
    _ensure_unified_contact_directory()

    with engine.connect() as c:
        existing={
            r._mapping["contact_key"]:dict(r._mapping)
            for r in c.execute(text("SELECT * FROM pi_contact_directory_v2")).fetchall()
        }
        props=[dict(r._mapping) for r in c.execute(text("""SELECT
            property_id,property_name,city,location,source,verification_status,
            owner_name,owner_contact,owner_contact_normalized,
            broker_name,broker_contact,broker_contact_normalized,
            contact_number,general_contact_normalized
            FROM pi_properties""")).fetchall()]

    grouped={}
    for p in props:
        candidates=[
            (p.get("owner_name"),p.get("owner_contact"),p.get("owner_contact_normalized"),"OWNER_SOURCE"),
            (p.get("broker_name"),p.get("broker_contact"),p.get("broker_contact_normalized"),"BROKER_SOURCE"),
            (None,p.get("contact_number"),p.get("general_contact_normalized"),"GENERAL_SOURCE")
        ]
        for name,raw,normed,source_hint in candidates:
            phones=sorted(_contact_number_set(raw,normed))
            key=_unified_contact_key(name,phones)
            if not key:
                continue
            g=grouped.setdefault(key,{
                "contact_key":key,
                "display_name":_dq_text(name) or None,
                "phones":set(),
                "property_ids":set(),
                "property_names":set(),
                "locations":set(),
                "cities":set(),
                "sources":set(),
                "source_hints":set(),
                "verified_property_count":0
            })
            if not g["display_name"] and _dq_text(name):
                g["display_name"]=_dq_text(name)
            g["phones"].update(phones)
            if p.get("property_id"):g["property_ids"].add(str(p.get("property_id")))
            if p.get("property_name"):g["property_names"].add(str(p.get("property_name")))
            if p.get("location") and not _dq_unknown(p.get("location")):g["locations"].add(str(p.get("location")))
            if p.get("city") and not _dq_unknown(p.get("city")):g["cities"].add(str(p.get("city")))
            if p.get("source"):g["sources"].add(str(p.get("source")))
            g["source_hints"].add(source_hint)
            if str(p.get("verification_status") or "").upper()=="VERIFIED":
                g["verified_property_count"]+=1

    with engine.begin() as c:
        for key,g in grouped.items():
            old=existing.get(key,{})
            phones=sorted(g["phones"])
            pids=sorted(g["property_ids"])
            # Manual role always wins. We only show source hints in remarks.
            role=old.get("contact_role") or "UNVERIFIED"
            status=old.get("verification_status") or "UNVERIFIED"
            remarks=old.get("remarks")
            if not remarks:
                remarks="Imported from property fields: "+", ".join(sorted(g["source_hints"]))
            c.execute(text("""INSERT INTO pi_contact_directory_v2(
                contact_key,display_name,primary_phone,phones,property_ids,property_names,
                locations,cities,sources,property_count,verified_property_count,
                contact_role,verification_status,verified_by,verified_at,remarks,updated_at
            ) VALUES(
                :ck,:name,:phone,CAST(:phones AS JSONB),CAST(:pids AS JSONB),CAST(:pnames AS JSONB),
                CAST(:locs AS JSONB),CAST(:cities AS JSONB),CAST(:sources AS JSONB),:pc,:vc,
                :role,:vs,:vby,:vat,:remarks,NOW()
            )
            ON CONFLICT(contact_key) DO UPDATE SET
                display_name=COALESCE(EXCLUDED.display_name,pi_contact_directory_v2.display_name),
                primary_phone=EXCLUDED.primary_phone,
                phones=EXCLUDED.phones,
                property_ids=EXCLUDED.property_ids,
                property_names=EXCLUDED.property_names,
                locations=EXCLUDED.locations,
                cities=EXCLUDED.cities,
                sources=EXCLUDED.sources,
                property_count=EXCLUDED.property_count,
                verified_property_count=EXCLUDED.verified_property_count,
                remarks=COALESCE(pi_contact_directory_v2.remarks,EXCLUDED.remarks),
                updated_at=NOW()"""),{
                "ck":key,"name":g["display_name"],"phone":phones[0] if phones else None,
                "phones":json.dumps(phones),"pids":json.dumps(pids),
                "pnames":json.dumps(sorted(g["property_names"])),
                "locs":json.dumps(sorted(g["locations"])),"cities":json.dumps(sorted(g["cities"])),
                "sources":json.dumps(sorted(g["sources"])),"pc":len(pids),
                "vc":g["verified_property_count"],"role":role,"vs":status,
                "vby":old.get("verified_by"),"vat":old.get("verified_at"),
                "remarks":remarks
            })
    return {"contacts":len(grouped)}

def _contact_sync_worker_v2():
    global _CONTACT_SYNC_RUNNING,_CONTACT_SYNC_PENDING
    import time
    try:
        while True:
            time.sleep(2)
            _CONTACT_SYNC_PENDING=False
            try:_sync_unified_contacts()
            except Exception:pass
            if not _CONTACT_SYNC_PENDING:break
    finally:
        _CONTACT_SYNC_RUNNING=False

def _request_contact_directory_sync():
    global _CONTACT_SYNC_RUNNING,_CONTACT_SYNC_PENDING
    _CONTACT_SYNC_PENDING=True
    if _CONTACT_SYNC_RUNNING:return
    _CONTACT_SYNC_RUNNING=True
    import threading
    threading.Thread(target=_contact_sync_worker_v2,daemon=True).start()

@app.post("/api/final-v2/contact-directory/sync")
def final_v2_sync(req:Request):
    need_login(req)
    return {"status":"ok",**_sync_unified_contacts()}

@app.get("/api/final-v2/contacts")
def final_v2_contacts(req:Request,role:str=Query("ALL"),limit:int=Query(3000,ge=1,le=5000)):
    need_login(req)
    _ensure_unified_contact_directory()
    params={"n":limit}
    where=""
    r=role.upper()
    if r!="ALL":
        if r not in {"UNVERIFIED","OWNER","BROKER","BOTH","OTHER"}:
            raise HTTPException(400,"Invalid role")
        where="WHERE contact_role=:role"
        params["role"]=r
    with engine.connect() as c:
        rows=c.execute(text(f"""SELECT * FROM pi_contact_directory_v2
            {where}
            ORDER BY
                CASE WHEN verification_status='UNVERIFIED' THEN 0 ELSE 1 END,
                property_count DESC,
                display_name NULLS LAST
            LIMIT :n"""),params).fetchall()
    return {"status":"ok","rows":_json_rows(rows)}

@app.post("/api/final-v2/contacts/{contact_key}/verify")
async def final_v2_verify_contact(contact_key:str,req:Request):
    need_login(req)
    body=await req.json()
    role=str(body.get("contact_role") or "").upper().strip()
    if role not in {"OWNER","BROKER","BOTH","OTHER","UNVERIFIED"}:
        raise HTTPException(400,"Role must be OWNER, BROKER, BOTH, OTHER or UNVERIFIED")
    name=str(body.get("display_name") or "").strip() or None
    remarks=str(body.get("remarks") or "").strip() or None
    status="VERIFIED" if role!="UNVERIFIED" else "UNVERIFIED"
    actor=actor_name(req)
    with engine.begin() as c:
        res=c.execute(text("""UPDATE pi_contact_directory_v2 SET
            display_name=COALESCE(:name,display_name),
            contact_role=:role,
            verification_status=:vs,
            verified_by=:by,
            verified_at=CASE WHEN :vs='VERIFIED' THEN NOW() ELSE NULL END,
            remarks=:remarks,
            updated_at=NOW()
            WHERE contact_key=:ck"""),{
                "name":name,"role":role,"vs":status,"by":actor,
                "remarks":remarks,"ck":contact_key
            })
        if res.rowcount==0:
            raise HTTPException(404,"Contact not found")
    return {"status":"ok","contact_key":contact_key,"contact_role":role,"verification_status":status}

@app.get("/api/final-v2/contact-counts")
def final_v2_contact_counts(req:Request):
    need_login(req)
    _ensure_unified_contact_directory()
    with engine.connect() as c:
        rows=c.execute(text("""SELECT contact_role,COUNT(*) n
            FROM pi_contact_directory_v2 GROUP BY contact_role""")).fetchall()
    d={str(r._mapping["contact_role"] or "UNVERIFIED"):int(r._mapping["n"]) for r in rows}
    return {"status":"ok","all":sum(d.values()),"unverified":d.get("UNVERIFIED",0),
            "owners":d.get("OWNER",0)+d.get("BOTH",0),
            "brokers":d.get("BROKER",0)+d.get("BOTH",0),
            "other":d.get("OTHER",0)}

@app.get("/contacts-directory",response_class=HTMLResponse)
def contacts_directory_page(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",status_code=303)
    _request_contact_directory_sync()
    return HTMLResponse(r"""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Property Contacts</title>
<style>*{box-sizing:border-box}body{margin:0;background:#f4f7fb;font-family:Arial;color:#142033}header{background:#0d1d2d;color:#fff;padding:16px 22px}.wrap{max-width:1850px;margin:auto;padding:18px}.card{background:#fff;border:1px solid #e4eaf1;border-radius:12px;padding:14px;margin-bottom:12px}.toolbar{display:flex;gap:8px;flex-wrap:wrap;align-items:center}.btn{padding:9px 12px;border:0;border-radius:8px;background:#1677ff;color:#fff;text-decoration:none;font-weight:700;cursor:pointer}.gray{background:#edf2f7;color:#24364b}input,select{padding:8px;border:1px solid #ccd7e4;border-radius:7px}.tablewrap{overflow:auto;max-height:72vh}table{width:100%;border-collapse:collapse;font-size:12px}th,td{padding:8px;border-bottom:1px solid #edf1f5;text-align:left;vertical-align:top;white-space:nowrap}th{position:sticky;top:0;background:#f8fafc}.UNVERIFIED{color:#a36b00;font-weight:700}.VERIFIED{color:#08734b;font-weight:700}</style></head><body>
<header><b>Unified Property Contacts</b><br><small>First merge all contacts. Then your team verifies and marks Owner / Broker / Both / Other.</small></header>
<div class="wrap">
<div class="card toolbar">
<a class="btn gray" href="/team-dashboard">Team Dashboard</a>
<button class="btn" onclick="sync()">Sync From Properties</button>
<input id="q" placeholder="Search name, phone, property, location">
<select id="filter"><option value="ALL">ALL CONTACTS</option><option value="UNVERIFIED">UNVERIFIED</option><option value="OWNER">OWNER</option><option value="BROKER">BROKER</option><option value="BOTH">BOTH</option><option value="OTHER">OTHER</option></select>
<span id="counts"></span>
</div>
<div class="card"><div class="tablewrap"><table><thead><tr>
<th>Name</th><th>Primary Phone</th><th>All Phones</th><th>Properties</th><th>Locations</th><th>Current Role</th><th>Status</th><th>Edit / Verify</th><th>Property Links</th>
</tr></thead><tbody id="rows"></tbody></table></div></div></div>
<script>
const E=x=>String(x??'').replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));let D=[];
async function api(u,o={}){let r=await fetch(u,o),d=await r.json();if(!r.ok)throw new Error(d.detail||d.message||'Error');return d}
async function load(){let role=document.querySelector('#filter').value;let d=await api('/api/final-v2/contacts?role='+encodeURIComponent(role));D=d.rows||[];render();counts()}
function render(){let q=(document.querySelector('#q').value||'').toLowerCase();document.querySelector('#rows').innerHTML=D.filter(x=>!q||JSON.stringify(x).toLowerCase().includes(q)).map(x=>`<tr>
<td><input id="${E(x.contact_key)}_name" value="${E(x.display_name||'')}" placeholder="Contact name"></td>
<td><b>${E(x.primary_phone||'')}</b></td><td>${E((x.phones||[]).join(', '))}</td>
<td>${E(x.property_count||0)}</td><td>${E((x.locations||[]).join(', '))}</td>
<td><select id="${E(x.contact_key)}_role"><option ${x.contact_role==='UNVERIFIED'?'selected':''}>UNVERIFIED</option><option ${x.contact_role==='OWNER'?'selected':''}>OWNER</option><option ${x.contact_role==='BROKER'?'selected':''}>BROKER</option><option ${x.contact_role==='BOTH'?'selected':''}>BOTH</option><option ${x.contact_role==='OTHER'?'selected':''}>OTHER</option></select></td>
<td class="${E(x.verification_status||'UNVERIFIED')}">${E(x.verification_status||'UNVERIFIED')}<br><small>${E(x.verified_by||'')}</small></td>
<td><input id="${E(x.contact_key)}_remarks" value="${E(x.remarks||'')}" placeholder="Verification note"><br><button class="btn" onclick='save(${JSON.stringify(x.contact_key)})'>Save / Verify</button></td>
<td>${(x.property_ids||[]).slice(0,15).map(id=>`<a target="_blank" href="/property-record/${encodeURIComponent(id)}">${E(id)}</a>`).join('<br>')}</td>
</tr>`).join('')}
async function save(key){let body={display_name:document.getElementById(key+'_name').value,contact_role:document.getElementById(key+'_role').value,remarks:document.getElementById(key+'_remarks').value};await api('/api/final-v2/contacts/'+encodeURIComponent(key)+'/verify',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)});await load()}
async function sync(){let d=await api('/api/final-v2/contact-directory/sync',{method:'POST'});alert('Unified contacts synchronized: '+d.contacts);await load()}
async function counts(){let d=await api('/api/final-v2/contact-counts');document.querySelector('#counts').innerHTML=`All <b>${d.all}</b> · Unverified <b>${d.unverified}</b> · Owners <b>${d.owners}</b> · Brokers <b>${d.brokers}</b>`}
document.querySelector('#q').addEventListener('input',render);document.querySelector('#filter').addEventListener('change',load);setTimeout(load,2000);load();
</script></body></html>""")

@app.get("/owners-directory",response_class=HTMLResponse)
def owners_directory_v2(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",status_code=303)
    return RedirectResponse("/contacts-directory?view=OWNER",status_code=303)

@app.get("/brokers-directory",response_class=HTMLResponse)
def brokers_directory_v2(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",status_code=303)
    return RedirectResponse("/contacts-directory?view=BROKER",status_code=303)

@app.get("/team-dashboard",response_class=HTMLResponse)
def team_dashboard_v2(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",status_code=303)
    _request_contact_directory_sync()
    admin='<a class="nav admin" href="/admin-data-tools">Admin Data Management<small>Technical data tools</small></a>' if role=="admin" else ''
    return HTMLResponse(f"""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>AI Deal Intelligence OS</title>
<style>*{{box-sizing:border-box}}body{{margin:0;background:#f4f7fb;font-family:Arial;color:#142033}}header{{background:#0d1d2d;color:#fff;padding:16px 22px;display:flex;justify-content:space-between;flex-wrap:wrap}}.wrap{{max-width:1650px;margin:auto;padding:18px}}.navs{{display:grid;grid-template-columns:repeat(auto-fit,minmax(210px,1fr));gap:10px}}.nav{{display:block;background:#fff;border:1px solid #e4eaf1;border-radius:11px;padding:13px;text-decoration:none;color:#16324f;font-weight:700}}.nav small{{display:block;color:#68788c;font-weight:400;margin-top:5px}}.admin{{border-color:#df8b13}}.kpis{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:10px;margin:14px 0}}.kpi,.card{{background:#fff;border:1px solid #e4eaf1;border-radius:12px;padding:14px;margin-bottom:12px}}.kpi b{{display:block;font-size:26px}}.btn{{padding:8px 11px;border:0;border-radius:7px;background:#1677ff;color:#fff;font-weight:700;cursor:pointer}}.gray{{background:#edf2f7;color:#24364b}}</style></head><body>
<header><div><b>AI Deal Intelligence OS</b><br><small>Unified Team Operations</small></div><div>{escape(role.upper())} · <a style="color:white" href="/logout">Logout</a></div></header>
<div class="wrap"><div class="navs">
<a class="nav" href="/property-database">Property Database<small>All saved inventory</small></a>
<a class="nav" href="/property-manual">Add Property + Matcher<small>Manual property and matching</small></a>
<a class="nav" href="/capture-intelligence">Capture Property<small>Camera, newspaper, handwritten, WhatsApp, PDF</small></a>
<a class="nav" href="/contacts-directory">Property Contacts<small>Verify then mark Owner / Broker / Both</small></a><a class="nav" href="/data-doctor">V12 Data Doctor<small>Full 5,000+ property reconciliation and contact coverage</small></a>
<a class="nav" href="/retail-requirements">Retail Requirements<small>LinkedIn leasing signals</small></a>
<a class="nav" href="/workspace#hospitality">Hospitality<small>Hospitality intelligence</small></a>
<a class="nav" href="/workspace#contacts">Marketing Contacts<small>WhatsApp contact database</small></a>
<a class="nav" href="/workspace#requirements">Requirement Discovery<small>Demand discovery</small></a>
<a class="nav" href="/workspace#bots">Bot Control Room<small>AI bot actions</small></a>
{admin}
</div>
<div class="card"><b>Contact workflow:</b> All property contacts are merged first. Your team verifies each contact, edits the name if required, and chooses OWNER, BROKER, BOTH or OTHER.</div>
<div class="card"><button class="btn" onclick="counts()">Refresh Contact Counts</button> <button class="btn gray" onclick="sync()">Sync Property Contacts</button> <span id="c"></span></div>
</div>
<script>
async function counts(){{let r=await fetch('/api/final-v2/contact-counts'),d=await r.json();document.querySelector('#c').innerHTML=`All <b>${{d.all}}</b> · Unverified <b>${{d.unverified}}</b> · Owners <b>${{d.owners}}</b> · Brokers <b>${{d.brokers}}</b>`}}
async function sync(){{await fetch('/api/final-v2/contact-directory/sync',{{method:'POST'}});counts()}}setTimeout(counts,2500);counts();
</script></body></html>""")


# ============================================================
# V12 DATA DOCTOR - FULL DATABASE RECONCILIATION
# Audits EVERY property and EVERY recoverable contact evidence.
# Original property records are never deleted or overwritten.
# ============================================================

def _dd_table_exists(name):
    with engine.connect() as c:
        return bool(c.execute(text("""SELECT EXISTS(
            SELECT 1 FROM information_schema.tables
            WHERE table_schema='public' AND table_name=:n
        )"""),{"n":name}).scalar_one())

def _dd_property_columns():
    with engine.connect() as c:
        rows=c.execute(text("""SELECT column_name,data_type
            FROM information_schema.columns
            WHERE table_schema='public' AND table_name='pi_properties'
            ORDER BY ordinal_position""")).fetchall()
    return [(r._mapping["column_name"],r._mapping["data_type"]) for r in rows]

def _dd_text_columns():
    cols=_dd_property_columns()
    result=[]
    for name,dtype in cols:
        if dtype in {"text","character varying","character"}:
            result.append(name)
    return result

def _dd_contact_columns():
    cols=[]
    for name,_ in _dd_property_columns():
        low=name.lower()
        if any(k in low for k in ["contact","phone","mobile","whatsapp","telephone","tel_no","telno"]):
            cols.append(name)
    # Also inspect remarks because many magazine/manual imports preserve source details there.
    for x in ["remarks","source","owner_name","broker_name"]:
        if x not in cols and any(name==x for name,_ in _dd_property_columns()):
            cols.append(x)
    return cols

def _ensure_data_doctor_tables():
    with engine.begin() as c:
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_property_contact_links(
            id BIGSERIAL PRIMARY KEY,
            property_id TEXT NOT NULL,
            normalized_contact TEXT NOT NULL,
            contact_kind TEXT DEFAULT 'PHONE',
            evidence_field TEXT,
            raw_value TEXT,
            role_hint TEXT DEFAULT 'UNVERIFIED',
            confidence INTEGER DEFAULT 100,
            is_primary BOOLEAN DEFAULT FALSE,
            first_seen_at TIMESTAMPTZ DEFAULT NOW(),
            updated_at TIMESTAMPTZ DEFAULT NOW(),
            UNIQUE(property_id,normalized_contact,evidence_field)
        )"""))
        c.execute(text("CREATE INDEX IF NOT EXISTS idx_dd_pcl_property ON pi_property_contact_links(property_id)"))
        c.execute(text("CREATE INDEX IF NOT EXISTS idx_dd_pcl_contact ON pi_property_contact_links(normalized_contact)"))
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_property_health(
            property_id TEXT PRIMARY KEY,
            valid_contact_count INTEGER DEFAULT 0,
            partial_contact_count INTEGER DEFAULT 0,
            contact_status TEXT,
            data_status TEXT,
            data_quality_score INTEGER DEFAULT 0,
            missing_fields JSONB DEFAULT '[]'::jsonb,
            contact_fields_scanned JSONB DEFAULT '[]'::jsonb,
            has_multiple_contacts BOOLEAN DEFAULT FALSE,
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )"""))
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_source_contact_evidence(
            id BIGSERIAL PRIMARY KEY,
            source_id BIGINT,
            normalized_contact TEXT,
            raw_value TEXT,
            source_field TEXT,
            assignable_to_property BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMPTZ DEFAULT NOW(),
            UNIQUE(source_id,normalized_contact,source_field)
        )"""))
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_data_doctor_runs(
            id BIGSERIAL PRIMARY KEY,
            run_id UUID UNIQUE NOT NULL,
            status TEXT DEFAULT 'RUNNING',
            total_properties INTEGER DEFAULT 0,
            properties_with_valid_contact INTEGER DEFAULT 0,
            properties_without_valid_contact INTEGER DEFAULT 0,
            multiple_contact_properties INTEGER DEFAULT 0,
            partial_contact_properties INTEGER DEFAULT 0,
            unique_contacts INTEGER DEFAULT 0,
            property_contact_links INTEGER DEFAULT 0,
            source_orphan_contacts INTEGER DEFAULT 0,
            notes TEXT,
            started_at TIMESTAMPTZ DEFAULT NOW(),
            completed_at TIMESTAMPTZ
        )"""))

def _dd_digits(v):
    return _re.sub(r'\D','',str(v or ''))

def _dd_extract_valid_contacts(value):
    """
    Conservative Indian contact extraction.
    Never pads or invents digits.
    Returns normalized 10-digit Indian mobiles and plausible full landlines.
    """
    raw=str(value or "")
    found=set()

    # Mobile patterns with optional +91/91/0 and separators.
    for m in _re.finditer(r'(?<!\d)(?:\+?91[\s\-]*)?0?([6-9](?:[\s\-]*\d){9})(?!\d)',raw):
        digits=_re.sub(r'\D','',m.group(1))
        if len(digits)==10:
            found.add(digits)

    # Generic contiguous/separated numeric candidates, then validate.
    for m in _re.finditer(r'(?<!\d)(\d(?:[\s\-]*\d){9,11})(?!\d)',raw):
        d=_re.sub(r'\D','',m.group(1))
        if len(d)==12 and d.startswith("91") and d[2] in "6789":
            found.add(d[2:])
        elif len(d)==11 and d.startswith("0") and d[1] in "6789":
            found.add(d[1:])
        elif len(d)==10 and d[0] in "6789":
            found.add(d)
        elif len(d) in {10,11} and d.startswith(("011","0120","0124","0129")):
            found.add(d)
    return sorted(found)

def _dd_partial_contacts(value):
    """
    Finds suspicious short phone-like fragments. They are review evidence only.
    They are NEVER converted into phone numbers.
    """
    raw=str(value or "")
    out=set()
    for m in _re.finditer(r'(?<!\d)([6-9]\d{4,8})(?!\d)',raw):
        d=m.group(1)
        if len(d)<10:
            out.add(d)
    return sorted(out)

def _dd_role_hint(field_name):
    f=(field_name or "").lower()
    if "owner" in f:return "OWNER_SOURCE"
    if "broker" in f:return "BROKER_SOURCE"
    return "UNVERIFIED"

def _dd_data_health(p):
    missing=[]
    score=0

    def known(v):
        s=str(v or "").strip()
        return bool(s) and s.lower() not in {"na","n/a","unknown","none","null","not specified","0"}

    if known(p.get("city")):score+=15
    else:missing.append("CITY")
    if known(p.get("location")):score+=20
    else:missing.append("LOCATION")
    if known(p.get("property_type")) and str(p.get("property_type")).upper() not in {"NA","UNKNOWN"}:score+=15
    else:missing.append("PROPERTY_TYPE")

    area=p.get("available_area_sqft") or p.get("minimum_area_sqft") or p.get("maximum_area_sqft")
    if area not in (None,"",0,"0"):score+=20
    else:missing.append("AREA")

    if known(p.get("rent_or_sale")):score+=15
    else:missing.append("TRANSACTION")
    if known(p.get("property_name")):score+=5
    if known(p.get("floor")):score+=5
    if str(p.get("verification_status") or "").upper()=="VERIFIED":score+=5

    if score>=80:data_status="DATA_STRONG"
    elif score>=60:data_status="DATA_USABLE"
    else:data_status="DATA_REVIEW"
    return score,data_status,missing

def _dd_rebuild_contact_directory_from_links():
    """
    Upgrade the existing unified contact directory from the complete link table.
    Manual verified roles remain untouched.
    """
    if not _dd_table_exists("pi_contact_directory_v2"):
        return {"directory_supported":False,"contacts":0}

    with engine.connect() as c:
        old={r._mapping["contact_key"]:dict(r._mapping) for r in
             c.execute(text("SELECT * FROM pi_contact_directory_v2")).fetchall()}
        rows=[dict(r._mapping) for r in c.execute(text("""SELECT
            l.normalized_contact,
            ARRAY_AGG(DISTINCT l.property_id) property_ids,
            COUNT(DISTINCT l.property_id) property_count,
            ARRAY_AGG(DISTINCT p.property_name) FILTER (WHERE p.property_name IS NOT NULL) property_names,
            ARRAY_AGG(DISTINCT p.location) FILTER (WHERE p.location IS NOT NULL) locations,
            ARRAY_AGG(DISTINCT p.city) FILTER (WHERE p.city IS NOT NULL) cities,
            ARRAY_AGG(DISTINCT p.source) FILTER (WHERE p.source IS NOT NULL) sources,
            ARRAY_AGG(DISTINCT l.role_hint) role_hints
            FROM pi_property_contact_links l
            JOIN pi_properties p ON p.property_id=l.property_id
            GROUP BY l.normalized_contact""")).fetchall()]

    with engine.begin() as c:
        for r in rows:
            phone=r["normalized_contact"]
            key="PHONE:"+phone
            previous=old.get(key,{})
            role=previous.get("contact_role") or "UNVERIFIED"
            vs=previous.get("verification_status") or "UNVERIFIED"
            name=previous.get("display_name")
            hints=[x for x in (r.get("role_hints") or []) if x]
            remark=previous.get("remarks") or ("Data Doctor source hints: "+", ".join(sorted(set(hints))))
            c.execute(text("""INSERT INTO pi_contact_directory_v2(
                contact_key,display_name,primary_phone,phones,property_ids,property_names,
                locations,cities,sources,property_count,verified_property_count,
                contact_role,verification_status,verified_by,verified_at,remarks,updated_at
            ) VALUES(
                :key,:name,:phone,CAST(:phones AS JSONB),CAST(:pids AS JSONB),CAST(:pnames AS JSONB),
                CAST(:locs AS JSONB),CAST(:cities AS JSONB),CAST(:sources AS JSONB),:pc,0,
                :role,:vs,:vby,:vat,:remarks,NOW()
            )
            ON CONFLICT(contact_key) DO UPDATE SET
                primary_phone=EXCLUDED.primary_phone,
                phones=EXCLUDED.phones,
                property_ids=EXCLUDED.property_ids,
                property_names=EXCLUDED.property_names,
                locations=EXCLUDED.locations,
                cities=EXCLUDED.cities,
                sources=EXCLUDED.sources,
                property_count=EXCLUDED.property_count,
                updated_at=NOW()"""),{
                "key":key,"name":name,"phone":phone,"phones":json.dumps([phone]),
                "pids":json.dumps(sorted(r.get("property_ids") or [])),
                "pnames":json.dumps(sorted([x for x in (r.get("property_names") or []) if x])),
                "locs":json.dumps(sorted([x for x in (r.get("locations") or []) if x])),
                "cities":json.dumps(sorted([x for x in (r.get("cities") or []) if x])),
                "sources":json.dumps(sorted([x for x in (r.get("sources") or []) if x])),
                "pc":int(r.get("property_count") or 0),"role":role,"vs":vs,
                "vby":previous.get("verified_by"),"vat":previous.get("verified_at"),
                "remarks":remark
            })
    return {"directory_supported":True,"contacts":len(rows)}

def _data_doctor_rebuild():
    _ensure_data_doctor_tables()
    run_id=str(uuid.uuid4())
    contact_cols=_dd_contact_columns()

    with engine.begin() as c:
        c.execute(text("""INSERT INTO pi_data_doctor_runs(run_id,status,notes)
            VALUES(CAST(:r AS UUID),'RUNNING',:n)"""),{
                "r":run_id,"n":"Full property/contact reconciliation. Originals preserved."
            })
        c.execute(text("DELETE FROM pi_property_contact_links"))
        c.execute(text("DELETE FROM pi_property_health"))
        c.execute(text("DELETE FROM pi_source_contact_evidence"))

    # Fetch full property rows dynamically so legacy/raw columns are included.
    with engine.connect() as c:
        props=[dict(r._mapping) for r in c.execute(text("SELECT * FROM pi_properties ORDER BY id")).fetchall()]

    with engine.begin() as c:
        for p in props:
            pid=str(p.get("property_id"))
            valid_by_contact=set()
            partial=set()

            for field in contact_cols:
                val=p.get(field)
                if val is None:continue
                val_str=str(val)
                for phone in _dd_extract_valid_contacts(val_str):
                    valid_by_contact.add(phone)
                    c.execute(text("""INSERT INTO pi_property_contact_links(
                        property_id,normalized_contact,contact_kind,evidence_field,raw_value,
                        role_hint,confidence,is_primary,updated_at
                    ) VALUES(:pid,:phone,'PHONE',:field,:raw,:hint,100,:primary,NOW())
                    ON CONFLICT(property_id,normalized_contact,evidence_field)
                    DO UPDATE SET raw_value=EXCLUDED.raw_value,role_hint=EXCLUDED.role_hint,updated_at=NOW()"""),{
                        "pid":pid,"phone":phone,"field":field,"raw":val_str[:2000],
                        "hint":_dd_role_hint(field),
                        "primary":field.lower() in {"owner_contact","broker_contact","contact_number"}
                    })
                partial.update(_dd_partial_contacts(val_str))

            score,data_status,missing=_dd_data_health(p)
            if len(valid_by_contact)>=2:contact_status="MULTIPLE_VALID_CONTACTS"
            elif len(valid_by_contact)==1:contact_status="CONTACT_INDEXED"
            elif partial:contact_status="PARTIAL_CONTACT_REVIEW"
            else:contact_status="NO_VALID_CONTACT_FOUND"

            c.execute(text("""INSERT INTO pi_property_health(
                property_id,valid_contact_count,partial_contact_count,contact_status,
                data_status,data_quality_score,missing_fields,contact_fields_scanned,
                has_multiple_contacts,updated_at
            ) VALUES(
                :pid,:vc,:pc,:cs,:ds,:score,CAST(:missing AS JSONB),CAST(:fields AS JSONB),:multi,NOW()
            )"""),{
                "pid":pid,"vc":len(valid_by_contact),"pc":len(partial),"cs":contact_status,
                "ds":data_status,"score":score,"missing":json.dumps(missing),
                "fields":json.dumps(contact_cols),"multi":len(valid_by_contact)>=2
            })

    # Source-level phone evidence is tracked separately; NEVER auto-linked to every property.
    with engine.connect() as c:
        sources=[dict(r._mapping) for r in c.execute(text("""SELECT id,source_reference,source_name,original_filename
            FROM pi_sources WHERE source_reference IS NOT NULL""")).fetchall()]
    with engine.begin() as c:
        for s in sources:
            raw=str(s.get("source_reference") or "")
            for phone in _dd_extract_valid_contacts(raw):
                c.execute(text("""INSERT INTO pi_source_contact_evidence(
                    source_id,normalized_contact,raw_value,source_field,assignable_to_property
                ) VALUES(:sid,:phone,:raw,'source_reference',FALSE)
                ON CONFLICT(source_id,normalized_contact,source_field) DO NOTHING"""),{
                    "sid":s.get("id"),"phone":phone,"raw":raw[:2000]
                })

    directory=_dd_rebuild_contact_directory_from_links()

    with engine.connect() as c:
        one=lambda sql:int(c.execute(text(sql)).scalar_one() or 0)
        total=one("SELECT COUNT(*) FROM pi_property_health")
        with_valid=one("SELECT COUNT(*) FROM pi_property_health WHERE valid_contact_count>0")
        without_valid=one("SELECT COUNT(*) FROM pi_property_health WHERE valid_contact_count=0")
        multi=one("SELECT COUNT(*) FROM pi_property_health WHERE valid_contact_count>1")
        partial=one("SELECT COUNT(*) FROM pi_property_health WHERE partial_contact_count>0")
        unique_contacts=one("SELECT COUNT(DISTINCT normalized_contact) FROM pi_property_contact_links")
        links=one("SELECT COUNT(*) FROM pi_property_contact_links")
        orphan=one("SELECT COUNT(*) FROM pi_source_contact_evidence")

    with engine.begin() as c:
        c.execute(text("""UPDATE pi_data_doctor_runs SET
            status='COMPLETED',total_properties=:t,properties_with_valid_contact=:wv,
            properties_without_valid_contact=:wo,multiple_contact_properties=:m,
            partial_contact_properties=:p,unique_contacts=:u,
            property_contact_links=:l,source_orphan_contacts=:o,completed_at=NOW()
            WHERE run_id=CAST(:r AS UUID)"""),{
                "t":total,"wv":with_valid,"wo":without_valid,"m":multi,"p":partial,
                "u":unique_contacts,"l":links,"o":orphan,"r":run_id
            })
    return {
        "run_id":run_id,"total_properties":total,
        "properties_with_valid_contact":with_valid,
        "properties_without_valid_contact":without_valid,
        "multiple_contact_properties":multi,
        "partial_contact_properties":partial,
        "unique_contacts":unique_contacts,
        "property_contact_links":links,
        "source_orphan_contacts":orphan,
        "directory_contacts":directory.get("contacts",0)
    }

@app.post("/api/v12/data-doctor/rebuild")
def v12_data_doctor_rebuild(req:Request):
    need_login(req)
    return {"status":"ok",**_data_doctor_rebuild()}

@app.get("/api/v12/data-doctor/summary")
def v12_data_doctor_summary(req:Request):
    need_login(req)
    _ensure_data_doctor_tables()
    with engine.connect() as c:
        one=lambda sql:int(c.execute(text(sql)).scalar_one() or 0)
        total=one("SELECT COUNT(*) FROM pi_properties")
        indexed=one("SELECT COUNT(*) FROM pi_property_health WHERE valid_contact_count>0")
        no_valid=one("SELECT COUNT(*) FROM pi_property_health WHERE valid_contact_count=0")
        multi=one("SELECT COUNT(*) FROM pi_property_health WHERE valid_contact_count>1")
        partial=one("SELECT COUNT(*) FROM pi_property_health WHERE partial_contact_count>0")
        unique=one("SELECT COUNT(DISTINCT normalized_contact) FROM pi_property_contact_links")
        links=one("SELECT COUNT(*) FROM pi_property_contact_links")
        orphan=one("SELECT COUNT(*) FROM pi_source_contact_evidence")
        strong=one("SELECT COUNT(*) FROM pi_property_health WHERE data_status='DATA_STRONG'")
        usable=one("SELECT COUNT(*) FROM pi_property_health WHERE data_status='DATA_USABLE'")
        review=one("SELECT COUNT(*) FROM pi_property_health WHERE data_status='DATA_REVIEW'")
    return {"status":"ok","total_properties":total,"contact_indexed_properties":indexed,
            "properties_without_valid_contact":no_valid,"multiple_contact_properties":multi,
            "partial_contact_properties":partial,"unique_contacts":unique,
            "property_contact_links":links,"source_orphan_contacts":orphan,
            "data_strong":strong,"data_usable":usable,"data_review":review}

@app.get("/api/v12/data-doctor/properties")
def v12_data_doctor_properties(
    req:Request,
    contact_status:str=Query("ALL"),
    data_status:str=Query("ALL"),
    q:str=Query(""),
    limit:int=Query(500,ge=1,le=3000)
):
    need_login(req)
    params={"n":limit}
    wh=[]
    if contact_status!="ALL":
        wh.append("h.contact_status=:cs");params["cs"]=contact_status
    if data_status!="ALL":
        wh.append("h.data_status=:ds");params["ds"]=data_status
    if q.strip():
        wh.append("""(p.property_id ILIKE :q OR COALESCE(p.property_name,'') ILIKE :q
            OR COALESCE(p.location,'') ILIKE :q OR COALESCE(p.city,'') ILIKE :q
            OR EXISTS(SELECT 1 FROM pi_property_contact_links l
                WHERE l.property_id=p.property_id AND l.normalized_contact ILIKE :q))""")
        params["q"]="%"+q.strip()+"%"
    where="WHERE "+(" AND ".join(wh)) if wh else ""
    with engine.connect() as c:
        rows=c.execute(text(f"""SELECT p.property_id,p.property_name,p.city,p.location,p.property_type,
            p.available_area_sqft,p.rent_or_sale,p.source,
            h.valid_contact_count,h.partial_contact_count,h.contact_status,h.data_status,
            h.data_quality_score,h.missing_fields,
            COALESCE((SELECT JSONB_AGG(DISTINCT l.normalized_contact)
                FROM pi_property_contact_links l WHERE l.property_id=p.property_id),'[]'::jsonb) contacts
            FROM pi_properties p
            LEFT JOIN pi_property_health h ON h.property_id=p.property_id
            {where}
            ORDER BY h.data_quality_score DESC NULLS LAST,p.id DESC
            LIMIT :n"""),params).fetchall()
    return {"status":"ok","rows":_json_rows(rows)}

@app.get("/api/v12/data-doctor/contact-links")
def v12_contact_links(req:Request,phone:str=Query(""),limit:int=Query(1000,ge=1,le=5000)):
    need_login(req)
    params={"n":limit}
    where=""
    if phone.strip():
        where="WHERE l.normalized_contact ILIKE :q";params["q"]="%"+phone.strip()+"%"
    with engine.connect() as c:
        rows=c.execute(text(f"""SELECT l.normalized_contact,l.property_id,l.evidence_field,l.role_hint,
            p.property_name,p.city,p.location,p.source
            FROM pi_property_contact_links l JOIN pi_properties p ON p.property_id=l.property_id
            {where}
            ORDER BY l.normalized_contact,p.property_id LIMIT :n"""),params).fetchall()
    return {"status":"ok","rows":_json_rows(rows)}

@app.get("/api/v12/data-doctor/export.csv")
def v12_data_doctor_export(req:Request):
    need_login(req)
    import csv as _csv, io as _io
    with engine.connect() as c:
        rows=[dict(r._mapping) for r in c.execute(text("""SELECT
            p.property_id,p.property_name,p.city,p.location,p.property_type,
            p.available_area_sqft,p.rent_or_sale,p.source,
            h.valid_contact_count,h.partial_contact_count,h.contact_status,
            h.data_status,h.data_quality_score,h.missing_fields,
            COALESCE((SELECT STRING_AGG(DISTINCT l.normalized_contact,' | ')
                FROM pi_property_contact_links l WHERE l.property_id=p.property_id),'') contacts
            FROM pi_properties p
            LEFT JOIN pi_property_health h ON h.property_id=p.property_id
            ORDER BY p.id""")).fetchall()]
    out=_io.StringIO()
    if rows:
        w=_csv.DictWriter(out,fieldnames=list(rows[0].keys()))
        w.writeheader()
        for r in rows:
            rr=dict(r)
            if isinstance(rr.get("missing_fields"),(list,dict)):
                rr["missing_fields"]=json.dumps(rr["missing_fields"])
            w.writerow(rr)
    return Response(content=out.getvalue(),media_type="text/csv",
        headers={"Content-Disposition":"attachment; filename=property-data-doctor-export.csv"})

@app.get("/data-doctor",response_class=HTMLResponse)
def v12_data_doctor_page(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",status_code=303)
    _ensure_data_doctor_tables()
    return HTMLResponse(r"""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>V12 Data Doctor</title>
<style>
*{box-sizing:border-box}body{margin:0;background:#f3f6fa;font-family:Arial;color:#142033}header{background:#0d1d2d;color:#fff;padding:17px 22px}.wrap{max-width:1900px;margin:auto;padding:18px}
.card,.kpi{background:#fff;border:1px solid #e1e8f0;border-radius:12px;padding:14px;margin-bottom:12px}.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:9px}.kpi b{display:block;font-size:25px}.toolbar{display:flex;gap:8px;flex-wrap:wrap;align-items:center}.btn{border:0;border-radius:8px;padding:9px 12px;background:#1677ff;color:#fff;text-decoration:none;font-weight:700;cursor:pointer}.gray{background:#edf2f7;color:#24364b}.orange{background:#d98200}input,select{padding:9px;border:1px solid #cad6e2;border-radius:7px}.tablewrap{overflow:auto;max-height:65vh}table{width:100%;border-collapse:collapse;font-size:12px}th,td{padding:8px;border-bottom:1px solid #edf1f5;text-align:left;vertical-align:top;white-space:nowrap}th{position:sticky;top:0;background:#f8fafc}.small{font-size:11px;color:#68788c}</style></head><body>
<header><b>V12 Data Doctor - Full Database Reconciliation</b><br><small>Every property matters. Originals are preserved; contacts are indexed without guessing.</small></header>
<div class="wrap">
<div class="card toolbar"><a class="btn gray" href="/team-dashboard">Team Dashboard</a><a class="btn gray" href="/property-database">Property Database</a><a class="btn gray" href="/contacts-directory">Contacts</a><button class="btn orange" onclick="rebuild()">Rebuild Full Database Index</button><a class="btn" href="/api/v12/data-doctor/export.csv">Export Organized CSV</a><span id="msg"></span></div>
<div class="card"><b>Safety:</b> This engine does not delete or merge property records. It scans all contact/phone/mobile fields, preserves every valid number, keeps short fragments as review-only evidence, builds a Property-to-Contact relationship index, and tracks source-level numbers separately when they cannot be safely assigned to one property.</div>
<div class="kpis" id="kpis"></div>
<div class="card toolbar">
<select id="cs"><option value="ALL">ALL CONTACT STATUS</option><option>CONTACT_INDEXED</option><option>MULTIPLE_VALID_CONTACTS</option><option>PARTIAL_CONTACT_REVIEW</option><option>NO_VALID_CONTACT_FOUND</option></select>
<select id="ds"><option value="ALL">ALL DATA STATUS</option><option>DATA_STRONG</option><option>DATA_USABLE</option><option>DATA_REVIEW</option></select>
<input id="q" placeholder="Search property, location, city or phone"><button class="btn gray" onclick="loadRows()">Search</button>
</div>
<div class="card"><div class="tablewrap"><table><thead><tr><th>Property</th><th>City</th><th>Location</th><th>Type</th><th>Area</th><th>Transaction</th><th>Contacts</th><th>Contact Status</th><th>Data Status</th><th>Score</th><th>Missing</th><th>Open</th></tr></thead><tbody id="rows"></tbody></table></div></div>
</div>
<script>
const E=x=>String(x??'').replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));
async function A(u,o={}){let r=await fetch(u,o),d=await r.json();if(!r.ok)throw new Error(d.detail||d.message||'Error');return d}
async function summary(){let d=await A('/api/v12/data-doctor/summary');let cards=[['TOTAL PROPERTIES',d.total_properties],['CONTACT INDEXED',d.contact_indexed_properties],['NO VALID CONTACT',d.properties_without_valid_contact],['MULTIPLE CONTACTS',d.multiple_contact_properties],['PARTIAL CONTACT',d.partial_contact_properties],['UNIQUE CONTACTS',d.unique_contacts],['PROPERTY-CONTACT LINKS',d.property_contact_links],['SOURCE ORPHAN CONTACTS',d.source_orphan_contacts],['DATA STRONG',d.data_strong],['DATA USABLE',d.data_usable],['DATA REVIEW',d.data_review]];document.querySelector('#kpis').innerHTML=cards.map(x=>`<div class="kpi"><span>${x[0]}</span><b>${Number(x[1]||0).toLocaleString()}</b></div>`).join('')}
async function rebuild(){if(!confirm('Rebuild the full reconciliation index for every property? Original property records will not be deleted or overwritten.'))return;document.querySelector('#msg').textContent=' Rebuilding all property/contact relationships...';let d=await A('/api/v12/data-doctor/rebuild',{method:'POST'});document.querySelector('#msg').textContent=` Complete: ${d.properties_with_valid_contact}/${d.total_properties} properties have valid contacts; ${d.unique_contacts} unique contacts; ${d.property_contact_links} links.`;await summary();await loadRows()}
async function loadRows(){let u='/api/v12/data-doctor/properties?contact_status='+encodeURIComponent(document.querySelector('#cs').value)+'&data_status='+encodeURIComponent(document.querySelector('#ds').value)+'&q='+encodeURIComponent(document.querySelector('#q').value)+'&limit=1000';let d=await A(u);document.querySelector('#rows').innerHTML=(d.rows||[]).map(x=>`<tr><td><b>${E(x.property_name||x.property_id)}</b><br><span class="small">${E(x.property_id)}</span></td><td>${E(x.city||'')}</td><td>${E(x.location||'')}</td><td>${E(x.property_type||'')}</td><td>${E(x.available_area_sqft||'')}</td><td>${E(x.rent_or_sale||'')}</td><td>${E((x.contacts||[]).join(', '))}</td><td>${E(x.contact_status||'NOT INDEXED')}</td><td>${E(x.data_status||'')}</td><td>${E(x.data_quality_score||0)}</td><td>${E((x.missing_fields||[]).join(', '))}</td><td><a target="_blank" href="/property-record/${encodeURIComponent(x.property_id)}">View</a></td></tr>`).join('')}
['cs','ds'].forEach(id=>document.querySelector('#'+id).addEventListener('change',loadRows));document.querySelector('#q').addEventListener('keydown',e=>{if(e.key==='Enter')loadRows()});summary();loadRows();
</script></body></html>""")


@app.get("/api/v4/requirements")
def v4_requirements(req:Request,limit:int=Query(300,ge=1,le=1000)):
    need_login(req)
    with engine.connect() as c:
        rows=c.execute(text("""SELECT requirement_id,client_name,company_name,contact_phone,requirement_type,property_type,
            city,preferred_locations,minimum_area_sqft,maximum_area_sqft,rent_or_sale,status,created_at
            FROM pi_requirements ORDER BY created_at DESC LIMIT :n"""),{"n":limit}).fetchall()
    return {"status":"ok","rows":_json_rows(rows)}

@app.post("/api/v4/match/{rid}")
def v4_match(rid:str,req:Request):
    need_login(req)
    return robust_match_requirement(rid,create_whatsapp=True)

@app.get("/api/v4/bot-runs")
def v4_bot_runs(req:Request,limit:int=Query(100,ge=1,le=300)):
    need_login(req)
    with engine.connect() as c:
        rows=c.execute(text("SELECT * FROM ai_bot_runs ORDER BY started_at DESC LIMIT :n"),{"n":limit}).fetchall()
    return {"status":"ok","rows":_json_rows(rows)}

@app.get("/api/v4/activity")
def v4_activity(req:Request,limit:int=Query(100,ge=1,le=500)):
    need_login(req)
    with engine.connect() as c:
        rows=c.execute(text("SELECT * FROM ai_activity_ledger ORDER BY created_at DESC LIMIT :n"),{"n":limit}).fetchall()
    return {"status":"ok","rows":_json_rows(rows)}

@app.get("/api/v4/overview")
def v4_overview(req:Request):
    need_login(req)
    with engine.connect() as c:
        def one(q,p=None): return c.execute(text(q),p or {}).scalar_one()
        return {"status":"ok","data":{
            "properties":one("SELECT COUNT(*) FROM pi_properties"),
            "requirements":one("SELECT COUNT(*) FROM pi_requirements"),
            "matches":one("SELECT COUNT(*) FROM pi_matches"),
            "owners":one("SELECT COUNT(*) FROM pi_owners"),
            "brokers":one("SELECT COUNT(*) FROM pi_brokers"),
            "hospitality_contacts":one("SELECT COUNT(*) FROM ai_marketing_contacts"),
            "hospitality_companies":one("SELECT COUNT(*) FROM ai_companies WHERE division='HOSPITALITY'"),
            "retail_companies":one("SELECT COUNT(*) FROM ai_companies WHERE division='RETAIL'"),
            "demand_signals":one("SELECT COUNT(*) FROM ai_demand_signals"),
            "bot_runs":one("SELECT COUNT(*) FROM ai_bot_runs")
        }}

_V4_CSS = """
:root{--bg:#f4f7fb;--card:#fff;--ink:#142033;--muted:#6d7b90;--line:#e4eaf1;--nav:#0d1d2d;--blue:#1677ff;--green:#0a9b63;--orange:#df8b13;--red:#d94848}
*{box-sizing:border-box}body{margin:0;background:var(--bg);font-family:Arial,sans-serif;color:var(--ink)}
.shell{display:grid;grid-template-columns:260px 1fr;min-height:100vh}.side{background:var(--nav);color:#fff;padding:18px 13px;position:sticky;top:0;height:100vh;overflow:auto}.brand{font-size:18px;font-weight:800;padding:5px 9px 16px}.brand small{display:block;font-size:11px;color:#90a3b7;margin-top:5px}.group{font-size:10px;letter-spacing:1.2px;color:#7f93a9;margin:17px 9px 6px}.nav{display:block;width:100%;text-align:left;border:0;background:transparent;color:#d9e3ed;padding:10px;border-radius:8px;cursor:pointer;text-decoration:none;font-size:13px}.nav:hover,.nav.active{background:#193653;color:white}
.main{min-width:0}.top{height:66px;background:#fff;border-bottom:1px solid var(--line);padding:0 22px;display:flex;align-items:center;justify-content:space-between}.content{padding:22px;max-width:1650px;margin:auto}.page{display:none}.page.active{display:block}.title{font-size:25px;margin:0;font-weight:800}.sub{font-size:13px;color:var(--muted);margin-top:4px}.kpis{display:grid;grid-template-columns:repeat(5,minmax(120px,1fr));gap:11px;margin:18px 0}.kpi,.card{background:#fff;border:1px solid var(--line);border-radius:12px;box-shadow:0 2px 9px rgba(0,0,0,.03)}.kpi{padding:14px}.kpi span{font-size:10px;color:var(--muted)}.kpi b{display:block;font-size:24px;margin-top:7px}.card{padding:16px;margin:14px 0}.grid2{display:grid;grid-template-columns:1fr 1fr;gap:14px}.grid3{display:grid;grid-template-columns:repeat(3,1fr);gap:14px}
input,select,textarea{width:100%;padding:10px;border:1px solid #d7e0ea;border-radius:8px;background:#fff;font-size:14px}textarea{min-height:76px}.formgrid{display:grid;grid-template-columns:repeat(2,1fr);gap:10px}.full{grid-column:1/-1}.btn{border:0;border-radius:8px;padding:10px 13px;background:var(--blue);color:white;font-weight:700;cursor:pointer}.btn.green{background:var(--green)}.btn.gray{background:#edf2f7;color:#24364b}.btn.orange{background:var(--orange)}.toolbar{display:flex;gap:8px;flex-wrap:wrap;margin:12px 0}
.drop{border:2px dashed #b9c8d8;padding:18px;border-radius:10px;text-align:center;background:#f8fbff}.drop.drag{border-color:var(--blue);background:#eef5ff}.drop input{border:0}
.tablewrap{overflow:auto;max-height:65vh;border:1px solid var(--line);border-radius:10px}table{border-collapse:collapse;width:100%;font-size:12px;background:#fff}th,td{padding:9px;border-bottom:1px solid #edf1f5;text-align:left;white-space:nowrap}th{background:#f8fafc;position:sticky;top:0}.badge{display:inline-block;padding:4px 7px;border-radius:20px;background:#eef3f7;font-size:10px;font-weight:700}.hot{color:var(--red);font-weight:800}.msg{padding:10px;border-radius:8px;margin:8px 0;background:#eef6ff}.good{background:#eaf8f2;color:#086d49}.warn{background:#fff5e6;color:#8a5600}.activity{padding:9px 0;border-bottom:1px solid #edf1f5}.activity small{display:block;color:var(--muted);margin-top:3px}
@media(max-width:1100px){.kpis{grid-template-columns:repeat(3,1fr)}.grid2,.grid3{grid-template-columns:1fr}}@media(max-width:760px){.shell{grid-template-columns:1fr}.side{height:auto;position:relative}.kpis{grid-template-columns:repeat(2,1fr)}.formgrid{grid-template-columns:1fr}.full{grid-column:auto}.content{padding:12px}}
"""

_V4_JS = r"""
const $=s=>document.querySelector(s), $$=s=>[...document.querySelectorAll(s)];
async function api(u,o={}){let r=await fetch(u,o),t=await r.text(),d;try{d=JSON.parse(t)}catch(e){d={message:t}}if(!r.ok)throw new Error(d.detail||d.message||t||('HTTP '+r.status));return d}
function esc(x){return String(x??'').replace(/[&<>"']/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[m]))}
function fmt(x){return x?new Date(x).toLocaleString():'-'} function n(x){return Number(x||0).toLocaleString()}
function nav(id){$$('.page').forEach(x=>x.classList.remove('active'));$$('.nav[data-page]').forEach(x=>x.classList.remove('active'));$('#'+id).classList.add('active');document.querySelector(`[data-page="${id}"]`)?.classList.add('active');loadPage(id)}
$$('.nav[data-page]').forEach(b=>b.onclick=()=>nav(b.dataset.page));
async function loadOverview(){let d=(await api('/api/v4/overview')).data;Object.entries(d).forEach(([k,v])=>$$(`[data-k="${k}"]`).forEach(e=>e.textContent=n(v)));await loadActivity(8)}
async function loadPage(id){if(id==='property')await loadRequirements();if(id==='owners')await loadOwners();if(id==='brokers')await loadBrokers();if(id==='hospitality')await loadHospitality();if(id==='retail')await loadCompanies('RETAIL','retailRows');if(id==='contacts')await loadContacts();if(id==='demand')await loadCampaigns();if(id==='bots')await loadBots();if(id==='activity')await loadActivity(150)}
async function submitProperty(ev){ev.preventDefault();let fd=new FormData(ev.target);let inp=$('#propertyMedia');[...inp.files].forEach(f=>fd.append('media',f));let b=$('#saveProperty');b.disabled=true;try{let d=await api('/api/v4/properties/manual',{method:'POST',body:fd});$('#propertyMsg').innerHTML=`<div class="msg good">Saved ${esc(d.property_id)} · ${d.media.length} media files</div>`;ev.target.reset();await loadOverview()}catch(e){$('#propertyMsg').innerHTML=`<div class="msg warn">${esc(e.message)}</div>`}finally{b.disabled=false}}
function setupDrop(){let d=$('#dropZone'),i=$('#propertyMedia');if(!d||!i)return;['dragenter','dragover'].forEach(x=>d.addEventListener(x,e=>{e.preventDefault();d.classList.add('drag')}));['dragleave','drop'].forEach(x=>d.addEventListener(x,e=>{e.preventDefault();d.classList.remove('drag')}));d.addEventListener('drop',e=>{i.files=e.dataTransfer.files;$('#fileCount').textContent=i.files.length+' file(s) selected'});i.onchange=()=>$('#fileCount').textContent=i.files.length+' file(s) selected'}
async function loadOwners(){let d=await api('/api/v4/owners');$('#ownerRows').innerHTML=d.rows.map(x=>`<tr><td>${esc(x.owner_id)}</td><td><b>${esc(x.owner_name)}</b></td><td>${esc(x.contact_number||'')}</td><td>${esc(x.email||'')}</td><td>${esc(x.city||'')}</td><td>${esc(x.notes||'')}</td></tr>`).join('')}
async function loadBrokers(){let d=await api('/api/v4/brokers');$('#brokerRows').innerHTML=d.rows.map(x=>`<tr><td>${esc(x.broker_id)}</td><td><b>${esc(x.broker_name)}</b></td><td>${esc(x.company_name||'')}</td><td>${esc(x.contact_number||'')}</td><td>${esc(x.email||'')}</td><td>${esc(x.city||'')}</td></tr>`).join('')}
async function addCompany(ev,div){ev.preventDefault();let body=Object.fromEntries(new FormData(ev.target).entries());body.division=div;body.expansion_score=Number(body.expansion_score||0);try{await api('/api/v4/companies',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)});ev.target.reset();await loadCompanies(div,div==='RETAIL'?'retailRows':'hospitalityRows');await loadOverview();alert('Saved')}catch(e){alert(e.message)}}
async function loadCompanies(div,target){let d=await api('/api/v4/companies?division='+div);$('#'+target).innerHTML=d.rows.map(x=>`<tr><td><b>${esc(x.company_name)}</b></td><td>${esc(x.category||'')}</td><td>${esc(x.primary_contact_name||'')}</td><td>${esc(x.primary_contact_phone||'')}</td><td>${esc(x.primary_contact_email||'')}</td><td>${esc(x.target_markets||x.city||'')}</td><td class="${Number(x.expansion_score)>=80?'hot':''}">${Number(x.expansion_score||0).toFixed(0)}</td><td>${esc(x.assigned_to||'')}</td></tr>`).join('')||'<tr><td colspan="8">No records yet.</td></tr>'}
async function loadHospitality(){await loadCompanies('HOSPITALITY','hospitalityRows');await loadContacts()}
async function runHospitality(){let d=await api('/api/v4/hospitality-bot/start',{method:'POST'});alert('Hospitality bot started in background. Run ID: '+d.run_id);setTimeout(loadBots,1500)}
async function runRetail(){let d=await api('/api/v4/retail-bot/start',{method:'POST'});alert('Retail bot started in background. Run ID: '+d.run_id);setTimeout(loadBots,1500)}
async function uploadContacts(ev){ev.preventDefault();let fd=new FormData(ev.target);try{let d=await api('/api/v4/marketing-contacts/upload',{method:'POST',body:fd});alert(`Uploaded: ${d.created} new, ${d.duplicates} duplicates`);await loadContacts();await loadOverview()}catch(e){alert(e.message)}}
async function loadContacts(){let d=await api('/api/v4/marketing-contacts');if($('#contactRows'))$('#contactRows').innerHTML=d.rows.map(x=>`<tr><td>${esc(x.business_type||'')}</td><td><b>${esc(x.brand_name||'')}</b></td><td>${esc(x.contact_name||'')}</td><td>${esc(x.phone||'')}</td><td>${esc(x.email||'')}</td><td>${esc(x.website||'')}</td><td>${esc(x.location||'')}</td><td>${esc(x.consent_status||'UNKNOWN')}</td><td>${esc(x.verification_status||'')}</td></tr>`).join('')}
async function createCampaign(ev){ev.preventDefault();let body=Object.fromEntries(new FormData(ev.target).entries());['area_sqft','monthly_rent'].forEach(k=>body[k]=body[k]?Number(body[k]):null);try{let d=await api('/api/v4/requirement-campaigns',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)});$('#campaignMsg').innerHTML=`<div class="msg good">Campaign ${esc(d.campaign_id)} created.<br><b>Post draft:</b> ${esc(d.post_draft)}</div>`;ev.target.reset();await loadCampaigns()}catch(e){alert(e.message)}}
async function loadCampaigns(){let d=await api('/api/v4/requirement-campaigns');$('#campaignRows').innerHTML=d.rows.map(x=>`<tr><td>${esc(x.campaign_id)}</td><td><b>${esc(x.campaign_name)}</b></td><td>${esc(x.property_type||'')}</td><td>${esc(x.location||x.city||'')}</td><td>${x.area_sqft||''}</td><td>${x.monthly_rent||''}</td><td>${esc(x.status||'')}</td><td><button class="btn green" onclick="startCampaign('${esc(x.campaign_id)}')">Start Requirement Bot</button></td></tr>`).join('');await loadDemandSignals()}
async function startCampaign(id){let d=await api('/api/v4/requirement-campaigns/'+encodeURIComponent(id)+'/start',{method:'POST'});alert('Requirement Discovery Bot started: '+d.run_id)}
async function loadDemandSignals(){let d=await api('/api/v4/demand-signals');$('#demandRows').innerHTML=d.rows.map(x=>`<tr><td>${esc(x.source_name||'')}</td><td><b>${esc(x.title||'')}</b><br><small>${esc(x.excerpt||'')}</small></td><td class="hot">${Number(x.intent_score||0).toFixed(0)}%</td><td><b>${esc(x.contact_phone||'Not found')}</b></td><td>${esc(x.contact_email||'')}</td><td>${esc(x.contact_verification_status||'')}</td><td>${esc(x.location||'')}</td><td>${x.source_url?`<a target="_blank" href="${esc(x.source_url)}">Open source</a>`:''}</td><td>${fmt(x.created_at)}</td></tr>`).join('')||'<tr><td colspan="9">No 90%+ matches found yet.</td></tr>'}
async function loadRequirements(){let d=await api('/api/v4/requirements');let sel=$('#reqSelect');sel.innerHTML='<option value="">Select requirement</option>'+d.rows.map(x=>`<option value="${esc(x.requirement_id)}">${esc(x.requirement_id)} · ${esc(x.company_name||x.client_name||'')} · ${esc(x.city||'')} · ${esc(x.preferred_locations||'')}</option>`).join('');$('#reqRows').innerHTML=d.rows.map(x=>`<tr><td>${esc(x.requirement_id)}</td><td>${esc(x.company_name||x.client_name||'')}</td><td>${esc(x.city||'')}</td><td>${esc(x.preferred_locations||'')}</td><td>${x.minimum_area_sqft||''}-${x.maximum_area_sqft||''}</td><td>${esc(x.rent_or_sale||'')}</td><td><button class="btn" onclick="matchReq('${esc(x.requirement_id)}')">Match</button></td></tr>`).join('')}
async function matchSelected(){let id=$('#reqSelect').value;if(!id)return alert('Select a requirement');await matchReq(id)}
async function verifyProperty(pid){if(!confirm('Confirm: your team called the owner/broker and the property is currently available?'))return;let fd=new FormData();fd.append('status','VERIFIED');let r=await fetch('/api/v4/properties/'+encodeURIComponent(pid)+'/availability-verification',{method:'POST',body:fd});let d=await r.json();if(!r.ok)throw new Error(d.detail||'Verification failed');alert('VERIFIED. Contact numbers stay internal and will NOT be included in client WhatsApp.');await matchSelected()}
async function matchReq(id){try{let d=await api('/api/v4/match/'+encodeURIComponent(id),{method:'POST'});let diag=d.diagnostic||{};let ec=diag.exclusion_counts||{};let exclusions=Object.entries(ec).map(([k,v])=>`${esc(k)}: ${v}`).join(' · ');$('#matchDiag').innerHTML=`<div class="msg ${d.matches.length?'good':'warn'}"><b>${esc(diag.message||'Matching complete')}</b><br><b>Engine:</b> ${esc(d.engine||diag.engine||'MATCHING_V2')} · Database: ${diag.database_properties||0} · Eligible: ${diag.eligible_count||0} · Excluded: ${diag.excluded_count||0} · Returned: ${diag.matches_returned||0}<br><b>Hard-filter exclusions:</b> ${exclusions||'None'}<br><b>Team flow:</b> Call internal contact → verify availability → Mark Verified → then share. Contacts remain internal.</div>`;$('#matchRows').innerHTML=(d.matches||[]).map((x,i)=>{let contacts=[x.owner_contact?('Owner: '+(x.owner_name||'')+' '+x.owner_contact):'',x.broker_contact?('Broker: '+(x.broker_name||'')+' '+x.broker_contact):'',(!x.owner_contact&&!x.broker_contact&&x.contact_number)?('Contact: '+x.contact_number):''].filter(Boolean).join('<br>');let verified=String(x.verification_status||'').toUpperCase()==='VERIFIED';let b=x.score_breakdown||{};let breakdown=`Location ${b.location||0}/30 · Area ${b.area||0}/25 · Type ${b.property_type||0}/15 · Use ${b.suitable_use||0}/15 · Budget ${b.budget||0}/10 · Verification ${b.verification||0}/5`;return `<tr><td>${i+1}</td><td><a href="/property-record/${encodeURIComponent(x.property_id)}" target="_blank"><b>${esc(x.property_name||x.property_id)}</b><br>${esc(x.property_id)}</a><br><a class="btn gray" target="_blank" href="/property-record/${encodeURIComponent(x.property_id)}">View Full Property</a></td><td>${esc(x.city||'')}</td><td>${esc(x.location||'')}</td><td>${x.available_area_sqft||''}</td><td>${x.monthly_rent?Number(x.monthly_rent).toLocaleString():''}</td><td class="${x.score>=80?'hot':''}"><b>${Number(x.score||0).toFixed(0)}%</b></td><td><b>${esc(x.match_band||'')}</b></td><td><b>Internal only</b><br>${contacts||'No contact saved'}</td><td>${verified?'<b>✓ VERIFIED</b>':`<button class="btn green" onclick="verifyProperty('${x.property_id}')">Mark Verified</button>`}</td><td>${esc((x.reasons||[]).join(', '))}</td><td>${esc(breakdown)}</td><td>${esc((x.gaps||[]).join(', '))}</td></tr>`}).join('')||'<tr><td colspan="13">No eligible property passed the hard filters.</td></tr>';$('#excludedRows').innerHTML=(d.excluded||[]).map((x,i)=>`<tr><td>${i+1}</td><td><a href="/property-record/${encodeURIComponent(x.property_id)}" target="_blank"><b>${esc(x.property_name||x.property_id||'')}</b><br>${esc(x.property_id||'')}</a></td><td>${esc(x.city||'')}</td><td>${esc(x.location||'')}</td><td>${esc(x.property_type||'')}</td><td>${x.available_area_sqft||''}</td><td>${esc(x.source||'')}</td><td><b>${esc((x.reasons||[]).join(', '))}</b></td></tr>`).join('')||'<tr><td colspan="8">No inventory excluded.</td></tr>';await loadOverview()}catch(e){alert(e.message)}}
async function loadBots(){let d=await api('/api/v4/bot-runs');$('#botRows').innerHTML=d.rows.map(x=>`<tr><td>${esc(x.bot_name)}</td><td>${esc(x.division||'')}</td><td>${esc(x.status)}</td><td>${x.records_found||0}</td><td>${x.records_created||0}</td><td>${fmt(x.started_at)}</td><td>${esc(x.summary||x.error_message||'')}</td></tr>`).join('')}
async function loadActivity(limit=100){let d=await api('/api/v4/activity?limit='+limit);let h=d.rows.map(x=>`<div class="activity"><b>${esc(x.actor_name)} · ${esc(x.action)}</b><small>${esc(x.division||'')} · ${esc(x.summary||'')} · ${fmt(x.created_at)}</small></div>`).join('')||'No activity yet.';if($('#activityFeed'))$('#activityFeed').innerHTML=h;if($('#activityRows'))$('#activityRows').innerHTML=d.rows.map(x=>`<tr><td>${fmt(x.created_at)}</td><td>${esc(x.actor_name)}</td><td>${esc(x.division||'')}</td><td>${esc(x.action)}</td><td>${esc(x.summary||'')}</td><td>${esc(x.status||'')}</td></tr>`).join('')}
setupDrop();loadOverview();
"""

def _v4_page(role):
    badge="ADMIN" if role=="admin" else "TEAM"
    return f"""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>AI Deal Intelligence OS V4</title><style>{_V4_CSS}</style></head><body><div class="shell">
<aside class="side"><div class="brand">AI Deal Intelligence OS<small>V4 · Property · Hospitality · Retail · Demand</small></div>
<div class="group">COMMAND</div><button class="nav active" data-page="command">▣ Command Centre</button><button class="nav" data-page="activity">◎ AI Activity</button><button class="nav" data-page="bots">⚡ Bot Control Room</button>
<div class="group">PROPERTY</div><button class="nav" data-page="property">⌂ Add Property + Matcher</button><a class="nav" href="/property-database">▦ Full Property Database</a><a class="nav" href="/data-quality">✓ Data Quality / Organizer</a><a class="nav" href="/data-recovery">↻ Intelligent Recovery V5</a><a class="nav" href="/master-data-cleaner">◆ Master Data Cleaner V7</a><a class="nav" href="/smart-master-data">★ Smart Master Data V8</a><a class="nav" href="/property-reconstruction">✦ Property Reconstruction V9</a><a class="nav" href="/capture-intelligence">◉ Camera / Newspaper / Handwritten V10</a><a class="nav" href="/retail-requirements">▤ Retail Requirement Leads</a><button class="nav" data-page="owners">● Owners Database</button><button class="nav" data-page="brokers">● Brokers Database</button><a class="nav" href="/legacy-workspace">Original Upload Workspace</a><a class="nav" href="/database-page">Original Database</a>
<div class="group">LEAD INTELLIGENCE</div><button class="nav" data-page="hospitality">◆ Hospitality</button><button class="nav" data-page="retail">◈ Retail Expansion</button><button class="nav" data-page="contacts">✉ Marketing Contacts</button><button class="nav" data-page="demand">⌕ Requirement Discovery</button>
</aside><main class="main"><header class="top"><div><b>Unified Delhi NCR Deal Intelligence</b><div class="sub">Organized database + AI bots + matching</div></div><div>{badge} · <a href="/logout">Logout</a></div></header><div class="content">

<section class="page active" id="command"><h1 class="title">Command Centre</h1><div class="sub">One view across supply, demand, owners, brokers, hospitality and retail.</div>
<div class="kpis"><div class="kpi"><span>PROPERTIES</span><b data-k="properties">0</b></div><div class="kpi"><span>REQUIREMENTS</span><b data-k="requirements">0</b></div><div class="kpi"><span>MATCHES</span><b data-k="matches">0</b></div><div class="kpi"><span>HOSPITALITY CONTACTS</span><b data-k="hospitality_contacts">0</b></div><div class="kpi"><span>DEMAND SIGNALS</span><b data-k="demand_signals">0</b></div></div>
<div class="grid3"><div class="card"><h3>Property Database</h3><p>Owners: <b data-k="owners">0</b><br>Brokers: <b data-k="brokers">0</b></p><button class="btn" onclick="nav('property')">Open Property</button></div>
<div class="card"><h3>Automated Discovery</h3><p>Hospitality and Retail bots run in the background. Requirement Discovery works from a property/campaign brief.</p><button class="btn green" onclick="nav('bots')">Bot Control Room</button></div>
<div class="card"><h3>Important</h3><p>Public web signals are leads to qualify, not confirmed requirements. Verify before outreach or sharing inventory.</p></div></div>
<div class="card"><h3>Latest AI Activity</h3><div id="activityFeed"></div></div></section>

<section class="page" id="property"><h1 class="title">Property Intelligence</h1><div class="sub">Improved manual property form + drag/drop media + repaired matcher.</div>
<div class="grid2"><div class="card"><h3>Add Property Manually</h3>
<div class="msg good"><b>One master form is now used everywhere.</b><br>
Owner Name · Owner Contact · Broker Name · Broker Contact · Main Contact Number · Monthly Rent in figures · Team Member · Verified/Unverified · Direct Images/Videos.</div>
<a class="btn green" href="/property-manual">Open Correct Add Property Form</a>
</div>
<div class="card"><h3>Property Matching Centre · V2</h3><div class="toolbar"><a class="btn orange" href="/property-database">View Full Property Database</a><span class="badge">Master saved inventory</span></div><select id="reqSelect"><option>Loading requirements...</option></select><div class="toolbar"><button class="btn green" onclick="matchSelected()">Run Smart Match V2</button><a class="btn gray" href="/legacy-workspace">Add Requirement / Upload Source</a></div><div id="matchDiag"></div>
<div class="msg good"><b>Matching order:</b> Self Inventory → Availability → Rent/Sale → Commercial/Residential → Suitable Use → City/Location → 80%-120% Area → 100-point ranking.</div>
<h4>Ranked Eligible Matches</h4>
<div class="tablewrap"><table><thead><tr><th>#</th><th>Property</th><th>City</th><th>Location</th><th>Area</th><th>Rent</th><th>Score</th><th>Band</th><th>Availability Contact (INTERNAL)</th><th>Verification</th><th>Why Matched</th><th>Score Breakdown</th><th>Data Gaps</th></tr></thead><tbody id="matchRows"></tbody></table></div>
<h4 style="margin-top:16px">Excluded Inventory</h4>
<div class="msg warn">These properties failed one or more mandatory rules and are NOT eligible for client sharing.</div>
<div class="tablewrap"><table><thead><tr><th>#</th><th>Property</th><th>City</th><th>Location</th><th>Type</th><th>Area</th><th>Source</th><th>Exclusion Reason</th></tr></thead><tbody id="excludedRows"></tbody></table></div></div>
</div></div>
<div class="card"><h3>Requirements</h3><div class="tablewrap"><table><thead><tr><th>ID</th><th>Client/Company</th><th>City</th><th>Location</th><th>Area</th><th>Rent/Sale</th><th>Action</th></tr></thead><tbody id="reqRows"></tbody></table></div></div></section>

<section class="page" id="owners"><h1 class="title">Owners Database</h1><div class="sub">Owner data is now normalized separately from properties.</div><div class="card"><div class="tablewrap"><table><thead><tr><th>Owner ID</th><th>Name</th><th>Contact</th><th>Email</th><th>City</th><th>Notes</th></tr></thead><tbody id="ownerRows"></tbody></table></div></div></section>
<section class="page" id="brokers"><h1 class="title">Brokers Database</h1><div class="sub">Broker data is now normalized separately from properties.</div><div class="card"><div class="tablewrap"><table><thead><tr><th>Broker ID</th><th>Name</th><th>Company</th><th>Contact</th><th>Email</th><th>City</th></tr></thead><tbody id="brokerRows"></tbody></table></div></div></section>

<section class="page" id="hospitality"><h1 class="title">Hospitality Intelligence</h1><div class="sub">Automatic bot + manual prospect entry. Restaurants, cafes, banquets, hotels, wedding venues and commercial farmhouses.</div>
<div class="toolbar"><button class="btn green" onclick="runHospitality()">Run Delhi NCR Hospitality Data Bot</button><span class="badge">Requires GOOGLE_PLACES_API_KEY</span></div>
<div class="grid2"><div class="card"><h3>Add Hospitality Prospect Manually</h3><form class="formgrid" onsubmit="addCompany(event,'HOSPITALITY')">
<input name="company_name" placeholder="Brand / venue / operator" required><input name="category" placeholder="Restaurant / Cafe / Banquet / Hotel">
<input name="primary_contact_name" placeholder="Contact person name"><input name="primary_contact_phone" placeholder="Contact number">
<input name="primary_contact_email" placeholder="Email ID"><input name="website" placeholder="Website">
<input name="linkedin_url" placeholder="LinkedIn URL"><input name="target_markets" value="Delhi NCR" placeholder="Target market">
<input name="assigned_to" placeholder="Team member"><input name="expansion_score" type="number" min="0" max="100" placeholder="Opportunity score">
<textarea class="full" name="source_excerpt" placeholder="Requirement / availability / notes"></textarea><button class="btn full">Save Hospitality Prospect</button></form></div>
<div class="card"><h3>How automatic mode works</h3><p>The Hospitality Data Bot uses Google Places to discover public business records across Delhi NCR. It stores brand, phone, location and website when available, and attempts to find a public business email on the listed website.</p><p><b>Contact-person names are stored when known; Google Places itself does not normally supply a named decision-maker.</b></p></div></div>
<div class="card"><h3>Hospitality Prospects</h3><div class="tablewrap"><table><thead><tr><th>Brand</th><th>Category</th><th>Contact Name</th><th>Phone</th><th>Email</th><th>Market</th><th>Score</th><th>Team</th></tr></thead><tbody id="hospitalityRows"></tbody></table></div></div></section>

<section class="page" id="retail"><h1 class="title">Retail Expansion Intelligence</h1><div class="sub">Automatic Retail Expansion Bot + manual prospect entry, now with name and contact number.</div>
<div class="toolbar"><button class="btn green" onclick="runRetail()">Run Retail + LinkedIn Requirement Bots</button><span class="badge">Requires SERPER_API_KEY</span></div>
<div class="card"><h3>Add Retail Expansion Prospect Manually</h3><form class="formgrid" onsubmit="addCompany(event,'RETAIL')">
<input name="company_name" placeholder="Retail brand/company" required><input name="category" placeholder="Fashion / QSR / Beauty / Electronics">
<input name="primary_contact_name" placeholder="Contact person name"><input name="primary_contact_phone" placeholder="Contact number">
<input name="primary_contact_email" placeholder="Email ID"><input name="website" placeholder="Website">
<input name="linkedin_url" placeholder="LinkedIn URL"><input name="target_markets" value="Delhi NCR" placeholder="Target market">
<input name="assigned_to" placeholder="Team member"><input name="expansion_score" type="number" min="0" max="100" placeholder="Expansion score">
<textarea class="full" name="source_excerpt" placeholder="Expansion signal / notes"></textarea><button class="btn full">Save Retail Prospect</button></form></div>
<div class="card"><h3>Retail Prospects</h3><div class="tablewrap"><table><thead><tr><th>Brand</th><th>Category</th><th>Contact Name</th><th>Phone</th><th>Email</th><th>Market</th><th>Score</th><th>Team</th></tr></thead><tbody id="retailRows"></tbody></table></div></div></section>

<section class="page" id="contacts"><h1 class="title">WhatsApp / Marketing Contact Database</h1><div class="sub">Hospitality contact database plus your own uploaded contacts. Use only where your outreach is permitted and appropriate.</div>
<div class="grid2"><div class="card"><h3>Upload Existing Contact Database</h3><form onsubmit="uploadContacts(event)"><input name="file" type="file" accept=".csv" required><button class="btn">Upload CSV</button></form><p>Suggested CSV columns: Business Type, Brand Name, Contact Name, Phone, Email, Website, Location, City, Consent Status, Verification Status, Team Member.</p></div>
<div class="card"><h3>Export</h3><p>Export the organized database for approved WhatsApp/CRM workflows.</p><a class="btn green" href="/api/v4/marketing-contacts/export.csv">Export CSV</a></div></div>
<div class="card"><h3>Marketing Contacts</h3><div class="tablewrap"><table><thead><tr><th>Type</th><th>Brand</th><th>Contact Name</th><th>Phone</th><th>Email</th><th>Website</th><th>Location</th><th>Consent</th><th>Verification</th></tr></thead><tbody id="contactRows"></tbody></table></div></div></section>

<section class="page" id="demand"><h1 class="title">Requirement Discovery Bot</h1><div class="sub">Strict mode: only 90%+ campaign matches are saved. Public contact numbers/emails are shown when found on indexed/source pages.</div>
<div class="grid2"><div class="card"><h3>Create Property Demand-Hunt Campaign</h3><form class="formgrid" onsubmit="createCampaign(event)">
<input name="campaign_name" placeholder="Campaign name" required><input name="property_id" placeholder="Existing Property ID (optional)">
<input name="property_type" placeholder="Property type"><input name="city" value="Delhi NCR" placeholder="City">
<input name="location" placeholder="Location"><input name="area_sqft" type="number" placeholder="Area sqft">
<input name="monthly_rent" type="number" placeholder="Monthly rent"><select name="rent_or_sale"><option>Rent</option><option>Sale</option></select>
<input name="suitable_category" placeholder="Suitable for Restaurant / Retail / Banquet"><input name="nearby_brands" placeholder="Nearby brands">
<input name="assigned_to" placeholder="Team member"><textarea class="full" name="additional_points" placeholder="Frontage, floor, parking, special points"></textarea>
<button class="btn full">Create Campaign</button></form><div id="campaignMsg"></div></div>
<div class="card"><h3>Important source rule</h3><p>The current bot searches <b>public/indexed</b> web results, including search-visible LinkedIn, Facebook, Instagram, 99acres and Magicbricks pages. It cannot reliably read arbitrary private/logged-in comments without official authorized API access.</p><p>The form also generates a property post draft. Automatic posting to social accounts should only be activated after the relevant account APIs are connected.</p></div></div>
<div class="card"><h3>Campaigns</h3><div class="tablewrap"><table><thead><tr><th>ID</th><th>Campaign</th><th>Type</th><th>Location</th><th>Area</th><th>Rent</th><th>Status</th><th>Action</th></tr></thead><tbody id="campaignRows"></tbody></table></div></div>
<div class="card"><h3>90%+ Demand Matches Found</h3><div class="msg good"><b>Strict filter:</b> Results below 90% are rejected. Contact details are shown only when found on a public source; missing data is never guessed.</div><div class="tablewrap"><table><thead><tr><th>Source</th><th>Signal</th><th>Match %</th><th>Contact No.</th><th>Email</th><th>Contact Status</th><th>Location</th><th>Source Link</th><th>Found</th></tr></thead><tbody id="demandRows"></tbody></table></div></div></section>

<section class="page" id="bots"><h1 class="title">Bot Control Room</h1><div class="sub">Bots run in the background so dashboard requests do not time out.</div>
<div class="toolbar"><button class="btn green" onclick="runHospitality()">Run Hospitality Bot</button><button class="btn green" onclick="runRetail()">Run Retail Bot</button><button class="btn gray" onclick="loadBots()">Refresh Runs</button></div>
<div class="card"><div class="tablewrap"><table><thead><tr><th>Bot</th><th>Division</th><th>Status</th><th>Found</th><th>Created</th><th>Started</th><th>Summary</th></tr></thead><tbody id="botRows"></tbody></table></div></div></section>

<section class="page" id="activity"><h1 class="title">AI Activity Ledger</h1><div class="sub">Every important bot action recorded for management review.</div><div class="card"><div class="tablewrap"><table><thead><tr><th>Time</th><th>Actor</th><th>Division</th><th>Action</th><th>Summary</th><th>Status</th></tr></thead><tbody id="activityRows"></tbody></table></div></div></section>

</div></main></div><script>{_V4_JS}</script></body></html>"""


def _property_manual_page(role):
    return f"""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Add Property Manually</title>
<style>
*{{box-sizing:border-box}}
body{{margin:0;background:#f4f7fb;font-family:Arial,sans-serif;color:#142033}}
header{{background:#0d1d2d;color:white;padding:18px 24px;display:flex;justify-content:space-between;align-items:center}}
header a{{color:white}}
.wrap{{max-width:1180px;margin:0 auto;padding:22px}}
.card{{background:white;border:1px solid #e4eaf1;border-radius:14px;padding:20px;box-shadow:0 2px 10px rgba(0,0,0,.04);margin-bottom:16px}}
h1{{margin:0 0 5px;font-size:27px}}h2{{font-size:16px;margin:0 0 14px}}.muted{{color:#6d7b90;font-size:13px}}
.grid{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px}}
label{{font-size:12px;font-weight:700;display:block;margin-bottom:5px;color:#42536a}}
input,select,textarea{{width:100%;padding:11px;border:1px solid #d7e0ea;border-radius:8px;background:#fff;font-size:14px}}
textarea{{min-height:90px}}.full{{grid-column:1/-1}}
.section-title{{font-size:12px;font-weight:800;letter-spacing:.8px;color:#1677ff;margin:4px 0 10px;text-transform:uppercase}}
.drop{{border:2px dashed #aebed0;border-radius:12px;padding:24px;text-align:center;background:#f8fbff;cursor:pointer}}
.drop.drag{{border-color:#1677ff;background:#edf5ff}}
.drop input{{margin-top:12px;border:0}}
.btn{{display:inline-block;border:0;border-radius:8px;padding:11px 15px;background:#1677ff;color:white;font-weight:800;cursor:pointer;text-decoration:none}}
.btn.green{{background:#0a9b63}}.btn.gray{{background:#edf2f7;color:#24364b}}
.actions{{display:flex;gap:9px;flex-wrap:wrap}}
.note{{padding:11px;border-radius:8px;background:#eef6ff;margin:10px 0;font-size:13px}}
.good{{background:#eaf8f2;color:#086d49}}.warn{{background:#fff4e6;color:#8a5600}}
#preview{{display:grid;grid-template-columns:repeat(auto-fill,minmax(120px,1fr));gap:9px;margin-top:12px}}
.previewItem{{border:1px solid #e2e8f0;border-radius:8px;padding:6px;background:white;overflow:hidden}}
.previewItem img,.previewItem video{{width:100%;height:95px;object-fit:cover;border-radius:6px}}
@media(max-width:760px){{.grid{{grid-template-columns:1fr}}.full{{grid-column:auto}}}}
</style>
</head>
<body>
<header>
<div><b>AI Deal Intelligence OS</b><br><small>Add Property Manually</small></div>
<div>{role.upper()} · <a href="/workspace">Dashboard</a> · <a href="/logout">Logout</a></div>
</header>
<div class="wrap">
<h1>Add Property Manually</h1>
<div class="muted">This is the only manual property-entry form. All dashboard links should open this page.</div>
<div class="note"><b>No image/video URLs are required.</b> Upload or drag files directly below.</div>

<form id="masterPropertyForm">

<div class="card">
<div class="section-title">A. Property Details</div>
<div class="grid">
<div><label>Property Name / Building</label><input name="property_name" placeholder="Example: ABC Tower"></div>
<div><label>Property Type *</label><input name="property_type" placeholder="Retail / Restaurant / Office / Banquet" required></div>
<div><label>City *</label><input name="city" value="Delhi NCR" required></div>
<div><label>Location / Micro-market *</label><input name="location" placeholder="Golf Course Road / Saket / Noida Sector 18" required></div>
<div><label>Available Area (sqft)</label><input name="available_area_sqft" type="number" step="0.01" placeholder="5000"></div>
<div><label>Minimum Area (sqft)</label><input name="minimum_area_sqft" type="number" step="0.01"></div>
<div><label>Maximum Area (sqft)</label><input name="maximum_area_sqft" type="number" step="0.01"></div>
<div><label>Floor</label><input name="floor" placeholder="Ground / First / Basement"></div>
<div><label>Rent / Sale</label><select name="rent_or_sale"><option value="Rent">Rent</option><option value="Sale">Sale</option></select></div>
<div><label>Monthly Rent - Figures</label><input name="monthly_rent" type="number" step="0.01" placeholder="800000"></div>
<div><label>Nearby Brands</label><input name="nearby_brands" placeholder="Starbucks, Zara, Nike"></div>
<div><label>Suitable Category</label><input name="suitable_category" placeholder="Restaurant / Fashion / Jewellery"></div>
<div><label>Parking</label><input name="parking" placeholder="Available / 10 cars / Valet"></div>
<div><label>Main Contact Number</label><input name="contact_number" placeholder="+91..."></div>
</div>
</div>

<div class="card">
<div class="section-title">B. Follow-up & Verification</div>
<div class="grid">
<div><label>Team Member for Follow-ups *</label><input name="assigned_to" placeholder="Team member name" required></div>
<div><label>Verification Status *</label>
<select name="verification_status" required>
<option value="UNVERIFIED">Unverified</option>
<option value="VERIFIED">Verified</option>
</select></div>
</div>
</div>

<div class="card">
<div class="section-title">C. Owner Details</div>
<div class="grid">
<div><label>Owner Name</label><input name="owner_name" placeholder="Owner full name"></div>
<div><label>Owner Contact Number</label><input name="owner_contact" placeholder="+91..."></div>
<div class="full"><label>Owner Email</label><input name="owner_email" type="email" placeholder="owner@example.com"></div>
</div>
<div class="muted" style="margin-top:8px">Owner information is also stored in the separate Owners Database.</div>
</div>

<div class="card">
<div class="section-title">D. Broker Details</div>
<div class="grid">
<div><label>Broker Name</label><input name="broker_name" placeholder="Broker full name"></div>
<div><label>Broker Contact Number</label><input name="broker_contact" placeholder="+91..."></div>
<div><label>Broker Email</label><input name="broker_email" type="email"></div>
<div><label>Broker Company</label><input name="broker_company" placeholder="Brokerage company"></div>
</div>
<div class="muted" style="margin-top:8px">Broker information is also stored in the separate Brokers Database.</div>
</div>

<div class="card">
<div class="section-title">E. Images & Videos</div>
<div id="dropZoneMaster" class="drop">
<b>Drag & Drop Property Images or Videos Here</b><br>
<span class="muted">or click Choose Files below. Multiple files are allowed.</span>
<input id="masterMedia" name="media" type="file" multiple
accept="image/jpeg,image/png,image/webp,image/gif,video/mp4,video/webm,video/quicktime,video/x-m4v">
<div id="masterFileCount" class="muted" style="margin-top:7px">0 files selected</div>
</div>
<div id="preview"></div>
</div>

<div class="card">
<div class="section-title">F. Remarks</div>
<textarea name="remarks" placeholder="Property details, owner/broker instructions, frontage, power, ceiling height, possession, restrictions, etc."></textarea>
</div>

<div id="saveMessage"></div>
<div class="actions">
<button id="masterSaveBtn" class="btn green" type="submit">Save Property</button>
<a class="btn gray" href="/workspace">Back to Dashboard</a>
</div>
</form>
</div>

<script>
const form=document.getElementById("masterPropertyForm");
const media=document.getElementById("masterMedia");
const drop=document.getElementById("dropZoneMaster");
const count=document.getElementById("masterFileCount");
const preview=document.getElementById("preview");
const msg=document.getElementById("saveMessage");
const btn=document.getElementById("masterSaveBtn");

function renderFiles(files){{
  count.textContent=files.length+" file(s) selected";
  preview.innerHTML="";
  [...files].forEach(file=>{{
    const box=document.createElement("div");
    box.className="previewItem";
    const title=document.createElement("div");
    title.style.fontSize="11px";
    title.style.marginBottom="5px";
    title.textContent=file.name;
    box.appendChild(title);
    const url=URL.createObjectURL(file);
    if(file.type.startsWith("image/")){{
      const im=document.createElement("img");im.src=url;box.appendChild(im);
    }}else if(file.type.startsWith("video/")){{
      const v=document.createElement("video");v.src=url;v.controls=true;box.appendChild(v);
    }}
    preview.appendChild(box);
  }});
}}
media.addEventListener("change",()=>renderFiles(media.files));
["dragenter","dragover"].forEach(ev=>drop.addEventListener(ev,e=>{{e.preventDefault();drop.classList.add("drag")}}));
["dragleave","drop"].forEach(ev=>drop.addEventListener(ev,e=>{{e.preventDefault();drop.classList.remove("drag")}}));
drop.addEventListener("drop",e=>{{media.files=e.dataTransfer.files;renderFiles(media.files)}});

form.addEventListener("submit",async e=>{{
  e.preventDefault();
  btn.disabled=true;
  msg.innerHTML='<div class="note">Saving property and media...</div>';
  try{{
    const fd=new FormData(form);
    const r=await fetch("/api/v4/properties/manual",{{method:"POST",body:fd}});
    const t=await r.text();
    let d;
    try{{d=JSON.parse(t)}}catch(err){{d={{message:t}}}}
    if(!r.ok)throw new Error(d.detail||d.message||t);
    msg.innerHTML='<div class="note good"><b>Property saved successfully.</b><br>Property ID: '+d.property_id+
      '<br>Owner ID: '+(d.owner_id||'-')+' · Broker ID: '+(d.broker_id||'-')+
      '<br>Media uploaded: '+(d.media||[]).length+'</div>';
    form.reset();preview.innerHTML="";count.textContent="0 files selected";
    window.scrollTo({{top:0,behavior:"smooth"}});
  }}catch(err){{
    msg.innerHTML='<div class="note warn"><b>Could not save property:</b> '+String(err.message||err)+'</div>';
  }}finally{{btn.disabled=false}}
}});
</script>
</body>
</html>"""

@app.get("/property-manual",response_class=HTMLResponse)
def property_manual_page(req:Request):
    role=page_role_or_redirect(req)
    if not role:
        return RedirectResponse("/login",status_code=303)
    return HTMLResponse(_property_manual_page(role))

@app.get("/workspace",response_class=HTMLResponse)
def v4_workspace(req:Request):
    role=page_role_or_redirect(req)
    if not role: return RedirectResponse("/login",status_code=303)
    return HTMLResponse(_v4_page(role))

# ============================================================
# V13 FINAL DATA ENGINE + SIMPLIFIED RETAIL REQUIREMENT TABLE
# REPAIRED V13.1
# ============================================================

def _v13_table_exists(name):
    with engine.connect() as c:
        return bool(c.execute(text("""SELECT EXISTS(
            SELECT 1 FROM information_schema.tables
            WHERE table_schema='public' AND table_name=:n
        )"""), {"n": name}).scalar_one())

def _ensure_v13_tables():
    with engine.begin() as c:
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_v13_recovery_suggestions(
            id BIGSERIAL PRIMARY KEY,
            property_id TEXT NOT NULL,
            suggested_contact TEXT,
            matched_property_id TEXT,
            confidence INTEGER DEFAULT 0,
            evidence JSONB DEFAULT '{}'::jsonb,
            decision TEXT DEFAULT 'PENDING',
            decided_by TEXT,
            decided_at TIMESTAMPTZ,
            created_at TIMESTAMPTZ DEFAULT NOW(),
            UNIQUE(property_id,suggested_contact,matched_property_id)
        )"""))
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_v13_master_groups(
            group_key TEXT PRIMARY KEY,
            master_property_id TEXT NOT NULL,
            member_property_ids JSONB DEFAULT '[]'::jsonb,
            group_type TEXT DEFAULT 'SINGLE',
            confidence INTEGER DEFAULT 100,
            review_status TEXT DEFAULT 'AUTO',
            created_at TIMESTAMPTZ DEFAULT NOW(),
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )"""))

def _v13_norm(v):
    return _re.sub(r'[^A-Z0-9]+', ' ', str(v or '').upper()).strip()

def _v13_area(p):
    for k in ("available_area_sqft", "minimum_area_sqft", "maximum_area_sqft"):
        try:
            x = float(p.get(k) or 0)
            if x > 0:
                return x
        except Exception:
            pass
    return 0.0

def _v13_identity(p):
    return (
        _v13_norm(p.get("property_name")),
        _v13_norm(p.get("location")),
        _v13_norm(p.get("city")),
        _v13_norm(p.get("property_type"))
    )

def _v13_contact_recovery():
    _ensure_v13_tables()
    if not _v13_table_exists("pi_property_health") or not _v13_table_exists("pi_property_contact_links"):
        return 0

    with engine.connect() as c:
        props = [dict(r._mapping) for r in c.execute(text("SELECT * FROM pi_properties")).fetchall()]
        links = [dict(r._mapping) for r in c.execute(
            text("SELECT property_id,normalized_contact FROM pi_property_contact_links")
        ).fetchall()]

    bypid = {}
    for x in links:
        bypid.setdefault(str(x["property_id"]), set()).add(x["normalized_contact"])

    indexed = [p for p in props if bypid.get(str(p.get("property_id")))]
    missing = [p for p in props if not bypid.get(str(p.get("property_id")))]
    made = 0

    with engine.begin() as c:
        for p in missing:
            pn, pl, pc, pt = _v13_identity(p)
            pa = _v13_area(p)
            if not pn or not pl:
                continue

            candidates = []
            for q in indexed:
                qn, ql, qc, qt = _v13_identity(q)
                qa = _v13_area(q)

                score = 0
                evidence = []

                if pn == qn:
                    score += 45
                    evidence.append("PROPERTY_IDENTITY")
                elif pn and qn and (pn in qn or qn in pn) and min(len(pn), len(qn)) >= 4:
                    score += 30
                    evidence.append("PROPERTY_NAME_CLOSE")

                if pl == ql:
                    score += 25
                    evidence.append("LOCALITY")

                if pc and qc and pc == qc:
                    score += 8
                    evidence.append("CITY")

                if pt and qt and pt == qt:
                    score += 7
                    evidence.append("TYPE")

                if pa and qa:
                    diff = abs(pa - qa) / max(pa, qa)
                    if diff <= 0.05:
                        score += 15
                        evidence.append("AREA_5_PERCENT")
                    elif diff <= 0.15:
                        score += 8
                        evidence.append("AREA_15_PERCENT")

                # Strict safety: exact property identity + locality required.
                if score >= 85 and "PROPERTY_IDENTITY" in evidence and "LOCALITY" in evidence:
                    for phone in bypid.get(str(q.get("property_id")), []):
                        candidates.append((score, phone, q.get("property_id"), evidence))

            for score, phone, matched_pid, evidence in sorted(candidates, reverse=True)[:5]:
                c.execute(text("""INSERT INTO pi_v13_recovery_suggestions(
                    property_id,suggested_contact,matched_property_id,confidence,evidence
                ) VALUES(:p,:ph,:m,:cf,CAST(:ev AS JSONB))
                ON CONFLICT(property_id,suggested_contact,matched_property_id)
                DO UPDATE SET confidence=EXCLUDED.confidence,evidence=EXCLUDED.evidence"""), {
                    "p": p.get("property_id"),
                    "ph": phone,
                    "m": matched_pid,
                    "cf": score,
                    "ev": __import__("json").dumps(evidence)
                })
                made += 1

    return made

def _v13_duplicate_groups():
    _ensure_v13_tables()

    with engine.connect() as c:
        props = [dict(r._mapping) for r in c.execute(text("SELECT * FROM pi_properties")).fetchall()]

    buckets = {}
    for p in props:
        name, loc, city, ptype = _v13_identity(p)
        area = _v13_area(p)
        if not name or not loc:
            continue
        area_bucket = round(area / 25) * 25 if area else 0
        key = "|".join([name, loc, city, ptype, str(area_bucket)])
        buckets.setdefault(key, []).append(str(p.get("property_id")))

    groups = 0
    with engine.begin() as c:
        for key, ids in buckets.items():
            if len(ids) < 2:
                continue
            master = ids[0]
            c.execute(text("""INSERT INTO pi_v13_master_groups(
                group_key,master_property_id,member_property_ids,group_type,confidence,review_status,updated_at
            ) VALUES(:g,:m,CAST(:ids AS JSONB),'STRONG_DUPLICATE',95,'REVIEW',NOW())
            ON CONFLICT(group_key) DO UPDATE SET
                master_property_id=EXCLUDED.master_property_id,
                member_property_ids=EXCLUDED.member_property_ids,
                updated_at=NOW()"""), {
                    "g": key,
                    "m": master,
                    "ids": __import__("json").dumps(ids)
                })
            groups += 1

    return groups

def _v13_run_final():
    _ensure_v13_tables()
    suggestions = _v13_contact_recovery()
    groups = _v13_duplicate_groups()

    with engine.connect() as c:
        total = int(c.execute(text("SELECT COUNT(*) FROM pi_properties")).scalar() or 0)

        if _v13_table_exists("pi_property_health"):
            indexed = int(c.execute(text(
                "SELECT COUNT(*) FROM pi_property_health WHERE valid_contact_count>0"
            )).scalar() or 0)
            match_ready = int(c.execute(text(
                """SELECT COUNT(*) FROM pi_property_health
                   WHERE data_status IN ('DATA_STRONG','DATA_USABLE')"""
            )).scalar() or 0)
        else:
            indexed = 0
            match_ready = 0

    return {
        "status": "ok",
        "total_properties": total,
        "contact_indexed": indexed,
        "contact_missing": max(0, total - indexed),
        "recovery_suggestions": suggestions,
        "duplicate_groups": groups,
        "match_ready": match_ready
    }

@app.post("/api/v13/final-reconcile")
def v13_final_reconcile(req: Request):
    need_login(req)
    return _v13_run_final()

@app.get("/api/v13/recovery-suggestions")
def v13_recovery_suggestions(req: Request, limit: int = Query(1000, ge=1, le=5000)):
    need_login(req)
    _ensure_v13_tables()

    with engine.connect() as c:
        rows = c.execute(text("""SELECT s.*,p.property_name,p.city,p.location,p.property_type,
            p.available_area_sqft
            FROM pi_v13_recovery_suggestions s
            JOIN pi_properties p ON p.property_id=s.property_id
            WHERE s.decision='PENDING'
            ORDER BY s.confidence DESC,s.created_at DESC LIMIT :n"""), {"n": limit}).fetchall()

    return {"status": "ok", "rows": _json_rows(rows)}

@app.post("/api/v13/recovery-suggestions/{sid}/decision")
async def v13_recovery_decision(sid: int, req: Request):
    need_login(req)
    body = await req.json()

    decision = str(body.get("decision") or "").upper()
    if decision not in {"APPROVED", "REJECTED"}:
        raise HTTPException(400, "decision must be APPROVED or REJECTED")

    who = str(body.get("decided_by") or actor_name(req) or "TEAM")

    with engine.begin() as c:
        row = c.execute(text(
            "SELECT * FROM pi_v13_recovery_suggestions WHERE id=:id"
        ), {"id": sid}).fetchone()

        if not row:
            raise HTTPException(404, "Suggestion not found")

        c.execute(text("""UPDATE pi_v13_recovery_suggestions SET
            decision=:d,decided_by=:w,decided_at=NOW()
            WHERE id=:id"""), {"d": decision, "w": who, "id": sid})

        if decision == "APPROVED":
            r = dict(row._mapping)
            c.execute(text("""INSERT INTO pi_property_contact_links(
                property_id,normalized_contact,contact_kind,evidence_field,raw_value,
                role_hint,confidence,is_primary,updated_at
            ) VALUES(:p,:ph,'PHONE','V13_APPROVED_RECOVERY',:ph,'UNVERIFIED',:cf,FALSE,NOW())
            ON CONFLICT(property_id,normalized_contact,evidence_field)
            DO UPDATE SET confidence=EXCLUDED.confidence,updated_at=NOW()"""), {
                "p": r["property_id"],
                "ph": r["suggested_contact"],
                "cf": r["confidence"]
            })

    return {"status": "ok", "decision": decision}

def _v13_retail_rows():
    with engine.connect() as c:
        cols = {
            r._mapping["column_name"]
            for r in c.execute(text("""SELECT column_name
                FROM information_schema.columns
                WHERE table_schema='public' AND table_name='ai_demand_signals'""")).fetchall()
        }

    def col(name, default_sql="NULL"):
        if name in cols:
            return name
        return f"{default_sql} AS {name}"

    wanted = [
        col("signal_id"),
        col("company_name"),
        col("contact_name"),
        col("designation"),
        col("contact_phone"),
        col("contact_email"),
        col("linkedin_profile_url"),
        col("source_url"),
        col("title"),
        col("excerpt"),
        col("location"),
        col("required_area_sqft"),
        col("required_property_type"),
        col("required_transaction"),
        col("intent_score", "0"),
        col("followup_status", "'NEW'"),
        col("crm_status", "'NOT_SENT'"),
        col("assigned_to"),
        col("created_at", "NOW()")
    ]

    where = []
    if "source_type" in cols:
        where.append("source_type='RETAIL_LINKEDIN_REQUIREMENT'")
    if "source_name" in cols:
        where.append("source_name ILIKE '%LinkedIn%'")
    if not where:
        where = ["1=1"]

    sql = (
        "SELECT " + ",".join(wanted) +
        " FROM ai_demand_signals WHERE (" + " OR ".join(where) + ")" +
        " ORDER BY intent_score DESC,created_at DESC LIMIT 2000"
    )

    with engine.connect() as c:
        return _json_rows(c.execute(text(sql)).fetchall())

@app.get("/api/v13/retail-simple")
def v13_retail_simple_api(req: Request):
    need_login(req)
    return {"status": "ok", "rows": _v13_retail_rows()}

@app.get("/retail-expansion", response_class=HTMLResponse)
def v13_retail_simple_page(req: Request):
    role = page_role_or_redirect(req)
    if not role:
        return RedirectResponse("/login", status_code=303)

    return HTMLResponse("""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Retail Expansion Requirements</title>
<style>
*{box-sizing:border-box}
body{margin:0;background:#f4f7fb;font-family:Arial;color:#172437}
header{padding:18px 22px;background:#102235;color:#fff}
.wrap{padding:18px}
.card{background:#fff;border:1px solid #e2e8f0;border-radius:12px;padding:14px;margin-bottom:12px}
.toolbar{display:flex;gap:8px;flex-wrap:wrap;align-items:center}
.btn{background:#1677ff;color:#fff;border:0;border-radius:7px;padding:9px 12px;text-decoration:none;font-weight:700;cursor:pointer}
.gray{background:#e9eef5;color:#203247}
input,select{padding:9px;border:1px solid #ccd6e2;border-radius:7px}
.stats{font-weight:700}
.tablewrap{overflow:auto;max-height:74vh}
table{width:100%;border-collapse:collapse;font-size:13px}
th,td{padding:9px;border-bottom:1px solid #edf1f5;text-align:left;vertical-align:top}
th{position:sticky;top:0;background:#f8fafc;z-index:1;white-space:nowrap}
.req{min-width:300px;white-space:normal}
.small{font-size:11px;color:#687789}
a{color:#1268d3}
.pill{padding:3px 7px;border-radius:10px;background:#edf4ff;white-space:nowrap}
</style>
</head>
<body>
<header>
<b>Retail Expansion Requirements</b><br>
<small>Simple team table: requirement, person, contact and original LinkedIn/public post</small>
</header>
<div class="wrap">
<div class="card toolbar">
<a class="btn gray" href="/workspace">Workspace</a>
<a class="btn gray" href="/data-command-center">Data Command Center</a>
<input id="q" placeholder="Search brand, person, location, phone, requirement">
<select id="status">
<option value="">All Follow-up</option>
<option>NEW</option>
<option>CONTACTED</option>
<option>QUALIFIED</option>
<option>NOT_RELEVANT</option>
</select>
<button class="btn" onclick="load()">Refresh</button>
<span class="stats" id="count"></span>
</div>
<div class="card">
<div class="tablewrap">
<table>
<thead>
<tr>
<th>Brand / Company</th>
<th>Person</th>
<th>Designation</th>
<th>Contact</th>
<th>Requirement</th>
<th>Location</th>
<th>Area</th>
<th>Type</th>
<th>Deal</th>
<th>Score</th>
<th>LinkedIn / Source</th>
<th>Follow-up</th>
<th>CRM</th>
<th>Assigned</th>
</tr>
</thead>
<tbody id="rows"></tbody>
</table>
</div>
</div>
</div>
<script>
const E=s=>String(s??'').replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));
let D=[];
async function A(u){
  let r=await fetch(u),d=await r.json();
  if(!r.ok)throw Error(d.detail||'Error');
  return d;
}
function render(){
  let q=(document.querySelector('#q').value||'').toLowerCase();
  let st=document.querySelector('#status').value;
  let R=D.filter(x=>
    (!q||JSON.stringify(x).toLowerCase().includes(q)) &&
    (!st||(x.followup_status||'NEW')===st)
  );
  document.querySelector('#count').textContent=R.length+' leads';
  document.querySelector('#rows').innerHTML=R.map(x=>`<tr>
<td><b>${E(x.company_name||'To verify')}</b></td>
<td>${E(x.contact_name||'To verify')}</td>
<td>${E(x.designation||'')}</td>
<td>${x.contact_phone?'<b>'+E(x.contact_phone)+'</b>':'Phone not found'}${x.contact_email?'<br>'+E(x.contact_email):''}</td>
<td class="req"><b>${E(x.title||'')}</b><br><span class="small">${E(x.excerpt||'')}</span></td>
<td>${E(x.location||'')}</td>
<td>${E(x.required_area_sqft||'')}</td>
<td>${E(x.required_property_type||'')}</td>
<td>${E(x.required_transaction||'Lease')}</td>
<td><span class="pill">${E(x.intent_score||0)}</span></td>
<td>${x.linkedin_profile_url?`<a target="_blank" href="${E(x.linkedin_profile_url)}">Profile</a><br>`:''}${x.source_url?`<a target="_blank" href="${E(x.source_url)}">Open requirement post</a>`:'Source unavailable'}</td>
<td>${E(x.followup_status||'NEW')}</td>
<td>${E(x.crm_status||'NOT_SENT')}</td>
<td>${E(x.assigned_to||'')}</td>
</tr>`).join('') || '<tr><td colspan="14">No retail requirement signals found.</td></tr>';
}
async function load(){
  let d=await A('/api/v13/retail-simple');
  D=d.rows||[];
  render();
}
document.querySelector('#q').addEventListener('input',render);
document.querySelector('#status').addEventListener('change',render);
load();
</script>
</body>
</html>""")

@app.get("/data-command-center", response_class=HTMLResponse)
def v13_command_center(req: Request):
    role = page_role_or_redirect(req)
    if not role:
        return RedirectResponse("/login", status_code=303)

    return HTMLResponse("""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Final Data Command Center</title>
<style>
body{font-family:Arial;background:#f4f7fb;margin:0;color:#172437}
header{background:#102235;color:white;padding:20px}
.wrap{padding:20px}
.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:12px}
.card{background:white;border:1px solid #e2e8f0;border-radius:12px;padding:16px}
a{display:block;text-decoration:none;color:#172437;font-weight:bold}
.small{font-size:12px;color:#687789}
</style>
</head>
<body>
<header>
<b>FINAL Data Command Center</b><br>
<small>One simple team entry point</small>
</header>
<div class="wrap">
<div class="grid">
<div class="card"><a href="/property-database">Property Database</a><span class="small">All saved inventory</span></div>
<div class="card"><a href="/data-doctor">Data Doctor</a><span class="small">Full database reconciliation and health</span></div>
<div class="card"><a href="/contacts-directory">Property Contacts</a><span class="small">Verify and classify Owner/Broker/Other</span></div>
<div class="card"><a href="/capture-intelligence">Capture Property</a><span class="small">Camera, newspaper, handwritten, WhatsApp, PDF</span></div>
<div class="card"><a href="/property-manual">Add Property + Matcher</a><span class="small">Manual property and matching</span></div>
<div class="card"><a href="/retail-expansion">Retail Expansion Requirements</a><span class="small">Simple LinkedIn requirement table</span></div>
<div class="card"><a href="/workspace#hospitality">Hospitality</a><span class="small">Hospitality intelligence</span></div>
<div class="card"><a href="/workspace#bots">Bot Control Room</a><span class="small">Run and review AI bots</span></div>
</div>
</div>
</body>
</html>""")

# ============================================================
# V13.2 REFINED MAGAZINE MASTER IMPORT
# Uses the refined Excel inside Property Intelligence Agent
# without overwriting/deleting original pi_properties rows.
# ============================================================

def _ensure_magazine_master_tables():
    with engine.begin() as c:
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_magazine_master(
            source_id TEXT PRIMARY KEY,
            record_status TEXT,
            match_eligible BOOLEAN DEFAULT FALSE,
            category TEXT,
            listing_type TEXT,
            locality TEXT,
            locality_source TEXT,
            plot_block TEXT,
            configuration TEXT,
            area NUMERIC,
            area_unit TEXT,
            floor TEXT,
            price TEXT,
            status_remarks TEXT,
            contact_name_company TEXT,
            valid_mobiles JSONB DEFAULT '[]'::jsonb,
            valid_landlines JSONB DEFAULT '[]'::jsonb,
            partial_contacts JSONB DEFAULT '[]'::jsonb,
            valid_contact_count INTEGER DEFAULT 0,
            quality_issues JSONB DEFAULT '[]'::jsonb,
            original_raw_text TEXT,
            import_batch TEXT,
            imported_at TIMESTAMPTZ DEFAULT NOW(),
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )"""))
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_magazine_contact_links(
            source_id TEXT NOT NULL,
            normalized_contact TEXT NOT NULL,
            contact_type TEXT,
            contact_name_company TEXT,
            locality TEXT,
            property_status TEXT,
            raw_evidence TEXT,
            import_batch TEXT,
            imported_at TIMESTAMPTZ DEFAULT NOW(),
            PRIMARY KEY(source_id,normalized_contact,contact_type)
        )"""))
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_magazine_property_map(
            source_id TEXT PRIMARY KEY,
            property_id TEXT,
            match_method TEXT,
            confidence INTEGER DEFAULT 0,
            map_status TEXT DEFAULT 'UNMATCHED',
            evidence JSONB DEFAULT '{}'::jsonb,
            reviewed_by TEXT,
            reviewed_at TIMESTAMPTZ,
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )"""))

def _xl_list(v):
    if v is None: return []
    vals=[]
    for x in _re.split(r'[,|;/]+',str(v)):
        x=x.strip()
        if x and x not in vals: vals.append(x)
    return vals

def _mag_norm(v):
    return _re.sub(r'[^A-Z0-9]+',' ',str(v or '').upper()).strip()

def _mag_num(v):
    try:return float(v)
    except Exception:return None

def _mag_import_xlsx(path,batch):
    from openpyxl import load_workbook
    _ensure_magazine_master_tables()
    wb=load_workbook(path,read_only=True,data_only=True)

    if "REFINED MASTER" not in wb.sheetnames:
        raise ValueError("REFINED MASTER sheet not found.")
    if "PROPERTY CONTACT LINKS" not in wb.sheetnames:
        raise ValueError("PROPERTY CONTACT LINKS sheet not found.")

    sh=wb["REFINED MASTER"]
    rows=sh.iter_rows(values_only=True)
    headers=[str(x or '').strip() for x in next(rows)]
    idx={h:i for i,h in enumerate(headers)}

    imported=0
    with engine.begin() as c:
        for row in rows:
            def g(name):
                i=idx.get(name)
                return row[i] if i is not None and i<len(row) else None
            sid=str(g("Source ID") or '').strip()
            if not sid: continue
            c.execute(text("""INSERT INTO pi_magazine_master(
                source_id,record_status,match_eligible,category,listing_type,locality,locality_source,
                plot_block,configuration,area,area_unit,floor,price,status_remarks,
                contact_name_company,valid_mobiles,valid_landlines,partial_contacts,
                valid_contact_count,quality_issues,original_raw_text,import_batch,updated_at
            ) VALUES(
                :sid,:rs,:me,:cat,:lt,:loc,:ls,:pb,:cfg,:area,:unit,:floor,:price,:sr,:cn,
                CAST(:mob AS JSONB),CAST(:land AS JSONB),CAST(:part AS JSONB),:vcc,
                CAST(:issues AS JSONB),:raw,:batch,NOW()
            )
            ON CONFLICT(source_id) DO UPDATE SET
                record_status=EXCLUDED.record_status,match_eligible=EXCLUDED.match_eligible,
                category=EXCLUDED.category,listing_type=EXCLUDED.listing_type,locality=EXCLUDED.locality,
                locality_source=EXCLUDED.locality_source,plot_block=EXCLUDED.plot_block,
                configuration=EXCLUDED.configuration,area=EXCLUDED.area,area_unit=EXCLUDED.area_unit,
                floor=EXCLUDED.floor,price=EXCLUDED.price,status_remarks=EXCLUDED.status_remarks,
                contact_name_company=EXCLUDED.contact_name_company,valid_mobiles=EXCLUDED.valid_mobiles,
                valid_landlines=EXCLUDED.valid_landlines,partial_contacts=EXCLUDED.partial_contacts,
                valid_contact_count=EXCLUDED.valid_contact_count,quality_issues=EXCLUDED.quality_issues,
                original_raw_text=EXCLUDED.original_raw_text,import_batch=EXCLUDED.import_batch,updated_at=NOW()
            """),{
                "sid":sid,"rs":str(g("Record Status") or ""),"me":str(g("Match Eligible") or "").upper()=="YES",
                "cat":g("Category"),"lt":g("Listing Type"),"loc":g("Locality"),"ls":g("Locality Source"),
                "pb":g("Plot / Block"),"cfg":g("Configuration"),"area":_mag_num(g("Area")),
                "unit":g("Area Unit"),"floor":g("Floor"),"price":g("Price"),"sr":g("Status / Remarks"),
                "cn":g("Contact Name / Company"),
                "mob":__import__("json").dumps(_xl_list(g("Valid Mobile(s)"))),
                "land":__import__("json").dumps(_xl_list(g("Valid Landline(s)"))),
                "part":__import__("json").dumps(_xl_list(g("Partial Contact Review"))),
                "vcc":int(g("Valid Contact Count") or 0),
                "issues":__import__("json").dumps([x.strip() for x in str(g("Quality Issues") or "").split(";") if x.strip()]),
                "raw":g("Original Raw Text"),"batch":batch
            })
            imported+=1

    sh=wb["PROPERTY CONTACT LINKS"]
    rows=sh.iter_rows(values_only=True)
    headers=[str(x or '').strip() for x in next(rows)]
    idx={h:i for i,h in enumerate(headers)}
    links=0
    with engine.begin() as c:
        for row in rows:
            def g(name):
                i=idx.get(name)
                return row[i] if i is not None and i<len(row) else None
            sid=str(g("Property ID") or '').strip()
            phone=str(g("Normalized Contact") or '').strip()
            ctype=str(g("Contact Type") or '').strip()
            if not sid or not phone: continue
            c.execute(text("""INSERT INTO pi_magazine_contact_links(
                source_id,normalized_contact,contact_type,contact_name_company,locality,
                property_status,raw_evidence,import_batch
            ) VALUES(:sid,:ph,:ct,:cn,:loc,:ps,:raw,:batch)
            ON CONFLICT(source_id,normalized_contact,contact_type) DO UPDATE SET
                contact_name_company=EXCLUDED.contact_name_company,locality=EXCLUDED.locality,
                property_status=EXCLUDED.property_status,raw_evidence=EXCLUDED.raw_evidence,
                import_batch=EXCLUDED.import_batch,imported_at=NOW()"""),{
                "sid":sid,"ph":phone,"ct":ctype,"cn":g("Contact Name / Company"),
                "loc":g("Locality"),"ps":g("Property Status"),"raw":g("Raw Evidence"),"batch":batch
            })
            links+=1
    wb.close()
    return imported,links

def _mag_existing_property_contacts():
    bypid={}
    if not _v13_table_exists("pi_property_contact_links"):
        return bypid
    with engine.connect() as c:
        for r in c.execute(text("SELECT property_id,normalized_contact FROM pi_property_contact_links")).fetchall():
            bypid.setdefault(str(r._mapping["property_id"]),set()).add(str(r._mapping["normalized_contact"]))
    return bypid

def _mag_reconcile():
    """
    V13.2.1 optimized reconciliation.
    Important:
    - compares only properties in the same normalized locality
    - converts SqYd to SqFt for area comparison
    - phone alone never maps a property
    - original pi_properties rows are never changed
    """
    _ensure_magazine_master_tables()

    with engine.connect() as c:
        mags=[dict(r._mapping) for r in c.execute(text(
            "SELECT * FROM pi_magazine_master WHERE record_status<>'EXCLUDE_NON_PROPERTY'"
        )).fetchall()]
        props=[dict(r._mapping) for r in c.execute(text("SELECT * FROM pi_properties")).fetchall()]
        link_rows=[dict(r._mapping) for r in c.execute(text(
            "SELECT source_id,normalized_contact FROM pi_magazine_contact_links"
        )).fetchall()]

    mag_contacts={}
    for r in link_rows:
        mag_contacts.setdefault(str(r["source_id"]),set()).add(str(r["normalized_contact"]))

    existing_contacts=_mag_existing_property_contacts()

    by_locality={}
    for p in props:
        pl=_mag_norm(p.get("location"))
        if pl:
            by_locality.setdefault(pl,[]).append(p)

    def magazine_area_sqft(m):
        try:
            a=float(m.get("area") or 0)
        except Exception:
            return 0.0
        unit=_mag_norm(m.get("area_unit"))
        if not a:
            return 0.0
        if unit in {"SQYD","SQ YD","SQYDS","SQ YDS","YARD","YARDS"}:
            return a*9.0
        return a

    mapped=review=unmatched=0

    with engine.begin() as c:
        for m in mags:
            sid=str(m.get("source_id"))
            mn=_mag_norm(m.get("plot_block"))
            ml=_mag_norm(m.get("locality"))
            mc=_mag_norm(m.get("configuration"))
            ma=magazine_area_sqft(m)
            mcontacts=mag_contacts.get(sid,set())

            candidates=by_locality.get(ml,[])
            scored=[]

            for p in candidates:
                pn=_mag_norm(p.get("property_name"))
                pc=_mag_norm(p.get("property_type"))

                try:
                    pa=float(
                        p.get("available_area_sqft")
                        or p.get("minimum_area_sqft")
                        or p.get("maximum_area_sqft")
                        or 0
                    )
                except Exception:
                    pa=0.0

                score=0
                evidence=["LOCALITY"]

                identity=False
                if mn and pn and mn==pn:
                    score+=50
                    identity=True
                    evidence.append("EXACT_IDENTITY")
                elif mn and pn and (mn in pn or pn in mn) and min(len(mn),len(pn))>=3:
                    score+=38
                    identity=True
                    evidence.append("CLOSE_IDENTITY")

                if not identity:
                    continue

                score+=25

                if mc and pc and (mc==pc or mc in pc or pc in mc):
                    score+=7
                    evidence.append("CONFIGURATION_TYPE")

                if ma and pa:
                    diff=abs(ma-pa)/max(ma,pa)
                    if diff<=0.05:
                        score+=15
                        evidence.append("AREA_5_PERCENT")
                    elif diff<=0.15:
                        score+=8
                        evidence.append("AREA_15_PERCENT")

                shared=mcontacts & existing_contacts.get(str(p.get("property_id")),set())
                if shared:
                    score+=15
                    evidence.append("SHARED_CONTACT")

                scored.append((score,str(p.get("property_id")),evidence))

            scored.sort(reverse=True,key=lambda x:x[0])

            if scored:
                top_score,pid,ev=scored[0]
                margin=top_score-(scored[1][0] if len(scored)>1 else 0)

                if top_score>=90 and margin>=10:
                    status="AUTO_MAPPED"
                    mapped+=1
                else:
                    status="REVIEW"
                    review+=1

                c.execute(text("""INSERT INTO pi_magazine_property_map(
                    source_id,property_id,match_method,confidence,map_status,evidence,updated_at
                ) VALUES(:sid,:pid,'V13_2_1_LOCALITY_INDEXED',:cf,:st,CAST(:ev AS JSONB),NOW())
                ON CONFLICT(source_id) DO UPDATE SET
                    property_id=EXCLUDED.property_id,
                    match_method=EXCLUDED.match_method,
                    confidence=EXCLUDED.confidence,
                    map_status=EXCLUDED.map_status,
                    evidence=EXCLUDED.evidence,
                    updated_at=NOW()"""),{
                    "sid":sid,"pid":pid,"cf":top_score,"st":status,
                    "ev":__import__("json").dumps(ev)
                })
            else:
                unmatched+=1
                c.execute(text("""INSERT INTO pi_magazine_property_map(
                    source_id,property_id,match_method,confidence,map_status,evidence,updated_at
                ) VALUES(:sid,NULL,'NO_SAFE_MATCH',0,'UNMATCHED','{}'::jsonb,NOW())
                ON CONFLICT(source_id) DO UPDATE SET
                    property_id=NULL,match_method='NO_SAFE_MATCH',
                    confidence=0,map_status='UNMATCHED',
                    evidence='{}'::jsonb,updated_at=NOW()"""),{"sid":sid})

    return {"auto_mapped":mapped,"review":review,"unmatched":unmatched}

def _mag_sync_contacts_to_agent():
    if not _v13_table_exists("pi_property_contact_links"):
        return 0
    _ensure_magazine_master_tables()
    added=0
    with engine.begin() as c:
        rows=c.execute(text("""SELECT mp.property_id,l.source_id,l.normalized_contact,l.contact_type,
            l.contact_name_company
            FROM pi_magazine_property_map mp
            JOIN pi_magazine_contact_links l ON l.source_id=mp.source_id
            WHERE mp.map_status IN ('AUTO_MAPPED','MANUALLY_CONFIRMED')
              AND mp.property_id IS NOT NULL""")).fetchall()
        for rr in rows:
            r=rr._mapping
            c.execute(text("""INSERT INTO pi_property_contact_links(
                property_id,normalized_contact,contact_kind,evidence_field,raw_value,
                role_hint,confidence,is_primary,updated_at
            ) VALUES(:pid,:ph,:kind,'REFINED_MAGAZINE_IMPORT',:raw,'UNVERIFIED',100,FALSE,NOW())
            ON CONFLICT(property_id,normalized_contact,evidence_field)
            DO UPDATE SET updated_at=NOW()"""),{
                "pid":r["property_id"],"ph":r["normalized_contact"],
                "kind":r["contact_type"] or "PHONE",
                "raw":(r["contact_name_company"] or "")+" | "+r["source_id"]
            })
            added+=1
    return added

@app.post("/api/v13-2/magazine/import")
async def v132_magazine_import(req:Request,file:UploadFile=File(...)):
    need_login(req)
    filename=file.filename or "refined-magazine.xlsx"
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(400,"Please upload Delhi_Property_Magazine_REFINED_FINAL.xlsx.")

    fd,path=tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)

    try:
        total=0
        with open(path,"wb") as out:
            while True:
                chunk=await file.read(1024*1024)
                if not chunk:
                    break
                total+=len(chunk)
                if total>50*1024*1024:
                    raise HTTPException(413,"Maximum workbook size is 50 MB.")
                out.write(chunk)

        if total==0:
            raise HTTPException(400,"Uploaded workbook is empty.")

        batch="MAG-"+datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")

        master_rows,links=_mag_import_xlsx(path,batch)
        reconciliation=_mag_reconcile()
        synced=_mag_sync_contacts_to_agent()

        return {
            "status":"ok",
            "batch":batch,
            "master_rows":master_rows,
            "contact_links":links,
            "agent_contact_links_synced":synced,
            **reconciliation
        }

    except HTTPException:
        raise
    except Exception as ex:
        import traceback
        print("V13.2.1 MAGAZINE IMPORT ERROR")
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"{type(ex).__name__}: {str(ex)}"
        )
    finally:
        try:
            os.unlink(path)
        except Exception:
            pass

@app.post("/api/v13-2/magazine/reconcile-existing")
def v132_reconcile_existing(req:Request):
    need_login(req)
    try:
        reconciliation=_mag_reconcile()
        synced=_mag_sync_contacts_to_agent()
        return {"status":"ok","agent_contact_links_synced":synced,**reconciliation}
    except Exception as ex:
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"{type(ex).__name__}: {str(ex)}"
        )

@app.get("/api/v13-2/magazine/summary")
def v132_magazine_summary(req:Request):
    need_login(req)
    _ensure_magazine_master_tables()
    with engine.connect() as c:
        one=lambda q:int(c.execute(text(q)).scalar_one() or 0)
        return {"status":"ok",
            "master_rows":one("SELECT COUNT(*) FROM pi_magazine_master"),
            "match_ready":one("SELECT COUNT(*) FROM pi_magazine_master WHERE record_status='MATCH_READY'"),
            "data_review":one("SELECT COUNT(*) FROM pi_magazine_master WHERE record_status='DATA_REVIEW'"),
            "excluded":one("SELECT COUNT(*) FROM pi_magazine_master WHERE record_status='EXCLUDE_NON_PROPERTY'"),
            "contact_links":one("SELECT COUNT(*) FROM pi_magazine_contact_links"),
            "auto_mapped":one("SELECT COUNT(*) FROM pi_magazine_property_map WHERE map_status='AUTO_MAPPED'"),
            "review":one("SELECT COUNT(*) FROM pi_magazine_property_map WHERE map_status='REVIEW'"),
            "unmatched":one("SELECT COUNT(*) FROM pi_magazine_property_map WHERE map_status='UNMATCHED'")
        }

@app.get("/magazine-master-import",response_class=HTMLResponse)
def magazine_master_import_page(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",status_code=303)
    return HTMLResponse("""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Refined Magazine Master Import</title></head>
<body style="font-family:Arial;background:#f4f7fb;margin:0;color:#172437">
<div style="background:#102235;color:white;padding:20px"><b>Refined Magazine Master Import</b><br><small>Safe reconciliation into Property Intelligence Agent</small></div>
<div style="padding:20px;max-width:1200px;margin:auto">
<div style="background:white;padding:16px;border-radius:12px;margin-bottom:12px">
<a href="/data-command-center">Data Command Center</a> · <a href="/data-doctor">Data Doctor</a> · <a href="/property-database">Property Database</a>
</div>
<div style="background:white;padding:16px;border-radius:12px">
<p><b>Upload Delhi_Property_Magazine_REFINED_FINAL.xlsx</b></p>
<form id="f"><input type="file" name="file" accept=".xlsx" required> <button type="submit">Import + Reconcile</button></form><p><button type="button" onclick="reconcileExisting()">Reconcile Already Imported Data</button></p>
<p id="msg"></p><div id="summary"></div>
</div></div>
<script>
async function load(){let r=await fetch('/api/v13-2/magazine/summary'),d=await r.json();summary.innerHTML=`Master rows <b>${d.master_rows}</b> · Match ready <b>${d.match_ready}</b> · Data review <b>${d.data_review}</b> · Excluded <b>${d.excluded}</b> · Contact links <b>${d.contact_links}</b> · Auto mapped <b>${d.auto_mapped}</b> · Review <b>${d.review}</b> · Unmatched <b>${d.unmatched}</b>`}
async function reconcileExisting(){msg.textContent='Reconciling staged data...';let r=await fetch('/api/v13-2/magazine/reconcile-existing',{method:'POST'}),d=await r.json();if(!r.ok){msg.textContent='RECONCILE ERROR: '+(d.detail||'Unknown error');return}msg.textContent=`Reconciled. Auto mapped ${d.auto_mapped}, review ${d.review}, unmatched ${d.unmatched}, contact links synced ${d.agent_contact_links_synced}.`;load()}
f.addEventListener('submit',async e=>{e.preventDefault();msg.textContent='Importing and reconciling...';let r=await fetch('/api/v13-2/magazine/import',{method:'POST',body:new FormData(f)}),d=await r.json();if(!r.ok){msg.textContent='IMPORT ERROR: '+(d.detail||d.message||'Unknown error');return}msg.textContent=`Imported ${d.master_rows} rows, ${d.contact_links} contact links. Auto mapped ${d.auto_mapped}, review ${d.review}, unmatched ${d.unmatched}.`;load()});load();
</script></body></html>""")

# ============================================================
# V13.3 UNMATCHED INVENTORY ACTIVATION SYSTEM
# Turns refined magazine rows into a controlled team workflow:
# CREATE NEW PROPERTY / LINK EXISTING / KEEP IN REVIEW
# Originals remain preserved.
# ============================================================

def _ensure_v133_tables():
    _ensure_magazine_master_tables()
    with engine.begin() as c:
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_magazine_activation_log(
            id BIGSERIAL PRIMARY KEY,
            source_id TEXT NOT NULL,
            action TEXT NOT NULL,
            property_id TEXT,
            previous_map_status TEXT,
            new_map_status TEXT,
            notes TEXT,
            acted_by TEXT,
            acted_at TIMESTAMPTZ DEFAULT NOW()
        )"""))
        c.execute(text("CREATE INDEX IF NOT EXISTS idx_mag_activation_source ON pi_magazine_activation_log(source_id)"))

def _v133_area_fields(m):
    """
    Preserve area meaning safely.
    SqFt can populate available_area_sqft.
    SqYd remains plot_area_sqyd and is NOT silently treated as available_area_sqft.
    """
    try:
        area=float(m.get("area") or 0)
    except Exception:
        area=0.0

    unit=_mag_norm(m.get("area_unit"))
    out={}
    if area<=0:
        return out
    if unit in {"SQFT","SQ FT","FT","SQUARE FEET","SQUARE FOOT"}:
        out["available_area_sqft"]=area
        out["minimum_area_sqft"]=area
        out["maximum_area_sqft"]=area
    elif unit in {"SQYD","SQ YD","SQYDS","SQ YDS","YARD","YARDS"}:
        out["plot_area_sqyd"]=area
    else:
        out["area_raw"]=area
        out["area_unit_raw"]=m.get("area_unit")
    return out

def _v133_contact_for_source(source_id):
    with engine.connect() as c:
        rows=c.execute(text("""SELECT normalized_contact,contact_type
            FROM pi_magazine_contact_links WHERE source_id=:sid
            ORDER BY CASE WHEN contact_type='MOBILE' THEN 0 ELSE 1 END, normalized_contact"""),
            {"sid":source_id}).fetchall()
    vals=[str(r._mapping["normalized_contact"]) for r in rows if r._mapping["normalized_contact"]]
    return vals

def _v133_create_property_from_magazine(source_id,req):
    _ensure_v133_tables()

    with engine.connect() as c:
        mrow=c.execute(text("SELECT * FROM pi_magazine_master WHERE source_id=:sid"),{"sid":source_id}).fetchone()
        maprow=c.execute(text("SELECT * FROM pi_magazine_property_map WHERE source_id=:sid"),{"sid":source_id}).fetchone()

    if not mrow:
        raise HTTPException(404,"Magazine record not found.")

    m=dict(mrow._mapping)
    mp=dict(maprow._mapping) if maprow else {}

    if m.get("record_status")=="EXCLUDE_NON_PROPERTY":
        raise HTTPException(400,"This row is classified as non-property and cannot be activated.")

    if mp.get("map_status") in {"AUTO_MAPPED","MANUALLY_CONFIRMED","CREATED_NEW"} and mp.get("property_id"):
        raise HTTPException(409,f"This row is already connected to property {mp.get('property_id')}.")

    # Require team confirmation for data-review rows.
    body={}
    try:
        body=req
    except Exception:
        body={}

    contacts=_v133_contact_for_source(source_id)

    raw_parts=[
        f"REFINED MAGAZINE SOURCE ID: {source_id}",
        f"ORIGINAL AREA: {m.get('area') or ''} {m.get('area_unit') or ''}",
        f"QUALITY ISSUES: {', '.join(m.get('quality_issues') or []) if isinstance(m.get('quality_issues'),list) else m.get('quality_issues') or ''}",
        f"RAW: {m.get('original_raw_text') or ''}"
    ]

    payload={
        "property_name": (m.get("plot_block") or m.get("source_id")),
        "city": body.get("city") or "New Delhi",
        "location": body.get("location") or m.get("locality"),
        "property_type": body.get("property_type") or m.get("configuration") or m.get("category") or "Property",
        "floor": body.get("floor") or m.get("floor"),
        "rent_or_sale": body.get("rent_or_sale") or m.get("listing_type"),
        "source": "REFINED_MAGAZINE_V13_3",
        "remarks": " | ".join([x for x in raw_parts if x]),
        "contact_number": " | ".join(contacts),
        "owner_name": None,
        "broker_name": None,
        "entry_status": "UNVERIFIED",
        "verification_status": "UNVERIFIED",
        "availability_status": "UNVERIFIED",
        "suitable_category": m.get("category"),
        "status_remarks": m.get("status_remarks"),
        "price": m.get("price")
    }
    payload.update(_v133_area_fields(m))

    # Existing save_property remains the canonical creator.
    try:
        result=save_property(payload)
    except TypeError:
        result=save_property(payload,None)

    if not isinstance(result,dict) or not result.get("property_id"):
        raise HTTPException(500,f"Property creation did not return a property_id: {result}")

    pid=result["property_id"]

    with engine.begin() as c:
        c.execute(text("""INSERT INTO pi_magazine_property_map(
            source_id,property_id,match_method,confidence,map_status,evidence,reviewed_by,reviewed_at,updated_at
        ) VALUES(:sid,:pid,'TEAM_CREATE_NEW',100,'CREATED_NEW','{}'::jsonb,:by,NOW(),NOW())
        ON CONFLICT(source_id) DO UPDATE SET
            property_id=EXCLUDED.property_id,match_method='TEAM_CREATE_NEW',
            confidence=100,map_status='CREATED_NEW',reviewed_by=EXCLUDED.reviewed_by,
            reviewed_at=NOW(),updated_at=NOW()"""),{
                "sid":source_id,"pid":pid,"by":body.get("acted_by") or "TEAM"
            })
        c.execute(text("""INSERT INTO pi_magazine_activation_log(
            source_id,action,property_id,previous_map_status,new_map_status,notes,acted_by
        ) VALUES(:sid,'CREATE_NEW',:pid,:old,'CREATED_NEW',:notes,:by)"""),{
            "sid":source_id,"pid":pid,"old":mp.get("map_status"),
            "notes":body.get("notes"),"by":body.get("acted_by") or "TEAM"
        })

    # Add all refined contacts as property-contact relationships.
    if _v13_table_exists("pi_property_contact_links"):
        with engine.begin() as c:
            for ph in contacts:
                c.execute(text("""INSERT INTO pi_property_contact_links(
                    property_id,normalized_contact,contact_kind,evidence_field,raw_value,
                    role_hint,confidence,is_primary,updated_at
                ) VALUES(:pid,:ph,'PHONE','REFINED_MAGAZINE_NEW_PROPERTY',:raw,'UNVERIFIED',100,FALSE,NOW())
                ON CONFLICT(property_id,normalized_contact,evidence_field)
                DO UPDATE SET updated_at=NOW()"""),{
                    "pid":pid,"ph":ph,"raw":source_id
                })

    return {"status":"ok","property_id":pid,"source_id":source_id}

def _v133_link_existing(source_id,property_id,actor,notes=None):
    _ensure_v133_tables()
    with engine.begin() as c:
        exists=c.execute(text("SELECT 1 FROM pi_properties WHERE property_id=:pid"),{"pid":property_id}).fetchone()
        if not exists:
            raise HTTPException(404,"Existing property ID not found.")

        old=c.execute(text("SELECT map_status FROM pi_magazine_property_map WHERE source_id=:sid"),{"sid":source_id}).fetchone()
        old_status=old._mapping["map_status"] if old else None

        c.execute(text("""INSERT INTO pi_magazine_property_map(
            source_id,property_id,match_method,confidence,map_status,evidence,reviewed_by,reviewed_at,updated_at
        ) VALUES(:sid,:pid,'TEAM_LINK_EXISTING',100,'MANUALLY_CONFIRMED','{}'::jsonb,:by,NOW(),NOW())
        ON CONFLICT(source_id) DO UPDATE SET
            property_id=EXCLUDED.property_id,match_method='TEAM_LINK_EXISTING',
            confidence=100,map_status='MANUALLY_CONFIRMED',reviewed_by=EXCLUDED.reviewed_by,
            reviewed_at=NOW(),updated_at=NOW()"""),{
                "sid":source_id,"pid":property_id,"by":actor
            })

        c.execute(text("""INSERT INTO pi_magazine_activation_log(
            source_id,action,property_id,previous_map_status,new_map_status,notes,acted_by
        ) VALUES(:sid,'LINK_EXISTING',:pid,:old,'MANUALLY_CONFIRMED',:notes,:by)"""),{
            "sid":source_id,"pid":property_id,"old":old_status,"notes":notes,"by":actor
        })

    synced=_mag_sync_contacts_to_agent()
    return {"status":"ok","property_id":property_id,"contact_links_synced":synced}

def _v133_keep_review(source_id,actor,notes=None):
    _ensure_v133_tables()
    with engine.begin() as c:
        old=c.execute(text("SELECT map_status FROM pi_magazine_property_map WHERE source_id=:sid"),{"sid":source_id}).fetchone()
        old_status=old._mapping["map_status"] if old else None

        c.execute(text("""INSERT INTO pi_magazine_property_map(
            source_id,property_id,match_method,confidence,map_status,evidence,reviewed_by,reviewed_at,updated_at
        ) VALUES(:sid,NULL,'TEAM_KEEP_REVIEW',0,'KEEP_REVIEW','{}'::jsonb,:by,NOW(),NOW())
        ON CONFLICT(source_id) DO UPDATE SET
            match_method='TEAM_KEEP_REVIEW',map_status='KEEP_REVIEW',
            reviewed_by=EXCLUDED.reviewed_by,reviewed_at=NOW(),updated_at=NOW()"""),{
                "sid":source_id,"by":actor
            })

        c.execute(text("""INSERT INTO pi_magazine_activation_log(
            source_id,action,property_id,previous_map_status,new_map_status,notes,acted_by
        ) VALUES(:sid,'KEEP_REVIEW',NULL,:old,'KEEP_REVIEW',:notes,:by)"""),{
            "sid":source_id,"old":old_status,"notes":notes,"by":actor
        })

    return {"status":"ok"}

@app.get("/api/v13-3/inventory/summary")
def v133_inventory_summary(req:Request):
    need_login(req)
    _ensure_v133_tables()
    with engine.connect() as c:
        one=lambda q:int(c.execute(text(q)).scalar_one() or 0)
        return {"status":"ok",
            "total_magazine":one("SELECT COUNT(*) FROM pi_magazine_master"),
            "unmatched":one("SELECT COUNT(*) FROM pi_magazine_property_map WHERE map_status='UNMATCHED'"),
            "unmatched_match_ready":one("""SELECT COUNT(*) FROM pi_magazine_property_map mp
                JOIN pi_magazine_master m ON m.source_id=mp.source_id
                WHERE mp.map_status='UNMATCHED' AND m.record_status='MATCH_READY'"""),
            "unmatched_data_review":one("""SELECT COUNT(*) FROM pi_magazine_property_map mp
                JOIN pi_magazine_master m ON m.source_id=mp.source_id
                WHERE mp.map_status='UNMATCHED' AND m.record_status='DATA_REVIEW'"""),
            "review":one("SELECT COUNT(*) FROM pi_magazine_property_map WHERE map_status='REVIEW'"),
            "auto_mapped":one("SELECT COUNT(*) FROM pi_magazine_property_map WHERE map_status='AUTO_MAPPED'"),
            "created_new":one("SELECT COUNT(*) FROM pi_magazine_property_map WHERE map_status='CREATED_NEW'"),
            "manually_linked":one("SELECT COUNT(*) FROM pi_magazine_property_map WHERE map_status='MANUALLY_CONFIRMED'"),
            "keep_review":one("SELECT COUNT(*) FROM pi_magazine_property_map WHERE map_status='KEEP_REVIEW'")
        }

@app.get("/api/v13-3/inventory/queue")
def v133_inventory_queue(
    req:Request,
    status:str=Query("UNMATCHED"),
    data_status:str=Query("MATCH_READY"),
    q:str=Query(""),
    page:int=Query(1,ge=1),
    page_size:int=Query(50,ge=10,le=200)
):
    need_login(req)
    _ensure_v133_tables()
    params={"lim":page_size,"off":(page-1)*page_size}
    wh=[]

    st=status.upper()
    if st!="ALL":
        wh.append("COALESCE(mp.map_status,'UNMATCHED')=:st")
        params["st"]=st

    ds=data_status.upper()
    if ds!="ALL":
        wh.append("m.record_status=:ds")
        params["ds"]=ds

    if q.strip():
        wh.append("""(
            m.source_id ILIKE :q OR COALESCE(m.locality,'') ILIKE :q OR
            COALESCE(m.plot_block,'') ILIKE :q OR COALESCE(m.configuration,'') ILIKE :q OR
            COALESCE(m.contact_name_company,'') ILIKE :q OR
            EXISTS(SELECT 1 FROM pi_magazine_contact_links l
                WHERE l.source_id=m.source_id AND l.normalized_contact ILIKE :q)
        )""")
        params["q"]="%"+q.strip()+"%"

    where="WHERE "+(" AND ".join(wh)) if wh else ""

    with engine.connect() as c:
        total=int(c.execute(text("""SELECT COUNT(*) FROM pi_magazine_master m
            LEFT JOIN pi_magazine_property_map mp ON mp.source_id=m.source_id """+where),params).scalar_one() or 0)

        rows=c.execute(text("""SELECT
            m.source_id,m.record_status,m.match_eligible,m.category,m.listing_type,
            m.locality,m.plot_block,m.configuration,m.area,m.area_unit,m.floor,m.price,
            m.status_remarks,m.contact_name_company,m.valid_mobiles,m.valid_landlines,
            m.quality_issues,m.original_raw_text,
            COALESCE(mp.map_status,'UNMATCHED') map_status,
            mp.property_id suggested_property_id,mp.confidence,mp.evidence
            FROM pi_magazine_master m
            LEFT JOIN pi_magazine_property_map mp ON mp.source_id=m.source_id
            """+where+"""
            ORDER BY
                CASE WHEN m.record_status='MATCH_READY' THEN 0 ELSE 1 END,
                COALESCE(mp.confidence,0) DESC,m.source_id
            LIMIT :lim OFFSET :off"""),params).fetchall()

    return {"status":"ok","total":total,"page":page,"page_size":page_size,"rows":_json_rows(rows)}

@app.post("/api/v13-3/inventory/{source_id}/create-new")
async def v133_create_new(source_id:str,req:Request):
    need_login(req)
    body=await req.json()
    body["acted_by"]=actor_name(req)
    return _v133_create_property_from_magazine(source_id,body)

@app.post("/api/v13-3/inventory/{source_id}/link-existing")
async def v133_link_existing_api(source_id:str,req:Request):
    need_login(req)
    body=await req.json()
    pid=str(body.get("property_id") or "").strip()
    if not pid:
        raise HTTPException(400,"property_id is required.")
    return _v133_link_existing(source_id,pid,actor_name(req),body.get("notes"))

@app.post("/api/v13-3/inventory/{source_id}/keep-review")
async def v133_keep_review_api(source_id:str,req:Request):
    need_login(req)
    body=await req.json()
    return _v133_keep_review(source_id,actor_name(req),body.get("notes"))

@app.get("/inventory-activation",response_class=HTMLResponse)
def v133_inventory_activation_page(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",status_code=303)

    return HTMLResponse("""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Inventory Activation</title><style>
*{box-sizing:border-box}body{margin:0;background:#f4f7fb;font-family:Arial;color:#172437}header{background:#102235;color:#fff;padding:18px 22px}.wrap{max-width:1750px;margin:auto;padding:18px}
.card,.kpi{background:#fff;border:1px solid #e2e8f0;border-radius:12px;padding:14px;margin-bottom:12px}.toolbar{display:flex;gap:8px;flex-wrap:wrap;align-items:center}.btn{background:#1677ff;color:#fff;border:0;border-radius:7px;padding:8px 11px;text-decoration:none;font-weight:700;cursor:pointer}.green{background:#08734b}.orange{background:#d98200}.gray{background:#e9eef5;color:#203247}.red{background:#a83b32}
.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:9px}.kpi b{display:block;font-size:24px}.queue{display:grid;grid-template-columns:1.3fr 1fr;gap:12px}.raw{background:#f8fafc;border-radius:8px;padding:8px;white-space:pre-wrap;font-size:12px}.small{font-size:11px;color:#687789}.actions{display:flex;gap:6px;flex-wrap:wrap;margin-top:10px}input,select{padding:8px;border:1px solid #ccd6e2;border-radius:7px}.pager{display:flex;justify-content:space-between;align-items:center}
@media(max-width:900px){.queue{grid-template-columns:1fr}}</style></head><body>
<header><b>Unmatched Inventory Activation</b><br><small>Create New Property · Link Existing Property · Keep in Review</small></header>
<div class="wrap">
<div class="card toolbar"><a class="btn gray" href="/data-command-center">Data Command Center</a><a class="btn gray" href="/magazine-master-import">Magazine Master</a><a class="btn gray" href="/property-database">Property Database</a>
<select id="status"><option>UNMATCHED</option><option>REVIEW</option><option>KEEP_REVIEW</option><option>ALL</option></select>
<select id="ds"><option>MATCH_READY</option><option>DATA_REVIEW</option><option>ALL</option></select>
<input id="q" placeholder="Search property, locality, contact"><button class="btn" onclick="reload()">Search</button></div>
<div class="kpis" id="kpis"></div>
<div id="cards"></div>
<div class="card pager"><button class="btn gray" onclick="prev()">Previous</button><span id="pageTxt"></span><button class="btn gray" onclick="next()">Next</button></div>
</div><script>
const E=x=>String(x??'').replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));
let page=1,size=25,total=0,rows=[];
async function A(u,o={}){let r=await fetch(u,o),d=await r.json();if(!r.ok)throw Error(d.detail||d.message||'Error');return d}
async function summary(){let d=await A('/api/v13-3/inventory/summary');let a=[['UNMATCHED',d.unmatched],['UNMATCHED MATCH READY',d.unmatched_match_ready],['UNMATCHED DATA REVIEW',d.unmatched_data_review],['REVIEW',d.review],['AUTO MAPPED',d.auto_mapped],['CREATED NEW',d.created_new],['MANUALLY LINKED',d.manually_linked],['KEEP REVIEW',d.keep_review]];document.querySelector('#kpis').innerHTML=a.map(x=>`<div class="kpi"><span>${x[0]}</span><b>${Number(x[1]||0).toLocaleString()}</b></div>`).join('')}
function render(){document.querySelector('#pageTxt').textContent=`Page ${page} of ${Math.max(1,Math.ceil(total/size))}`;document.querySelector('#cards').innerHTML=rows.map((x,i)=>`<div class="card queue">
<div><h3>${E(x.plot_block||x.source_id)} <span class="small">${E(x.source_id)}</span></h3><p><b>${E(x.locality||'Locality missing')}</b> · ${E(x.configuration||x.category||'')} · ${E(x.area||'')} ${E(x.area_unit||'')} · ${E(x.listing_type||'')}</p><p>Contact: <b>${E((x.valid_mobiles||[]).join(', '))}</b> ${E((x.valid_landlines||[]).join(', '))}</p><p>Status: <b>${E(x.record_status)}</b> · Map: <b>${E(x.map_status)}</b></p><div class="raw">${E(x.original_raw_text||'')}</div></div>
<div><p><b>Quality Issues</b><br>${E((x.quality_issues||[]).join(', ')||'None')}</p>${x.suggested_property_id?`<p>Suggested existing: <a target="_blank" href="/property-record/${encodeURIComponent(x.suggested_property_id)}">${E(x.suggested_property_id)}</a> · Confidence ${E(x.confidence||0)}</p>`:''}
<label>Existing Property ID</label><br><input id="pid_${i}" value="${E(x.suggested_property_id||'')}" placeholder="PROP-..."><div class="actions">
<button class="btn green" onclick="createNew(${i})">Create New Property</button><button class="btn orange" onclick="linkExisting(${i})">Link Existing</button><button class="btn gray" onclick="keepReview(${i})">Keep in Review</button></div><p class="small">Create New always creates UNVERIFIED inventory. SqYd remains plot area and is not silently treated as available SqFt.</p></div>
</div>`).join('')||'<div class="card">No records in this queue.</div>'}
async function load(){let u='/api/v13-3/inventory/queue?status='+encodeURIComponent(status.value)+'&data_status='+encodeURIComponent(ds.value)+'&q='+encodeURIComponent(q.value)+'&page='+page+'&page_size='+size;let d=await A(u);total=d.total;rows=d.rows||[];render();summary()}
async function createNew(i){let x=rows[i];if(!confirm('Create a NEW UNVERIFIED property from '+x.source_id+'?'))return;try{let d=await A('/api/v13-3/inventory/'+encodeURIComponent(x.source_id)+'/create-new',{method:'POST',headers:{'Content-Type':'application/json'},body:'{}'});alert('Created '+d.property_id);load()}catch(e){alert(e.message)}}
async function linkExisting(i){let x=rows[i],pid=document.querySelector('#pid_'+i).value.trim();if(!pid){alert('Enter existing Property ID');return}if(!confirm('Link '+x.source_id+' to '+pid+'?'))return;try{await A('/api/v13-3/inventory/'+encodeURIComponent(x.source_id)+'/link-existing',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({property_id:pid})});load()}catch(e){alert(e.message)}}
async function keepReview(i){let x=rows[i];await A('/api/v13-3/inventory/'+encodeURIComponent(x.source_id)+'/keep-review',{method:'POST',headers:{'Content-Type':'application/json'},body:'{}'});load()}
function reload(){page=1;load()}function next(){if(page<Math.ceil(total/size)){page++;load()}}function prev(){if(page>1){page--;load()}}
status.addEventListener('change',reload);ds.addEventListener('change',reload);q.addEventListener('keydown',e=>{if(e.key==='Enter')reload()});load();
</script></body></html>""")

# ============================================================
# V13.4 CLEAN TEAM DASHBOARD
# Keeps only operational links for the team.
# Old technical workspace remains available at /legacy-workspace.
# ============================================================

@app.get("/team-workspace-clean", response_class=HTMLResponse)
def v134_clean_team_workspace(req: Request):
    role = page_role_or_redirect(req)
    if not role:
        return RedirectResponse("/login", status_code=303)

    admin_tools = ""
    if role == "admin":
        admin_tools = """
        <div class="adminbox">
          <div><b>Admin tools</b><span>Only for data maintenance and system control.</span></div>
          <a href="/data-command-center">Data Command Center</a>
          <a href="/data-doctor">Data Doctor</a>
          <a href="/magazine-master-import">Magazine Master Import</a>
        </div>
        """

    return HTMLResponse(f"""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>AI Deal Intelligence OS</title>
<style>
*{{box-sizing:border-box}}
body{{margin:0;background:#f4f7fb;font-family:Arial,sans-serif;color:#172437}}
header{{background:#102235;color:#fff;padding:20px 24px;display:flex;justify-content:space-between;gap:12px;align-items:center;flex-wrap:wrap}}
.brand b{{font-size:20px}} .brand small{{display:block;color:#cbd8e5;margin-top:4px}}
.logout{{color:#fff;text-decoration:none;font-size:13px}}
.wrap{{max-width:1450px;margin:auto;padding:22px}}
.intro{{margin-bottom:16px}}
.intro h1{{margin:0 0 5px;font-size:25px}} .intro p{{margin:0;color:#66788c}}
.section{{margin-top:20px}} .section h2{{font-size:14px;letter-spacing:.08em;color:#526579;margin:0 0 10px}}
.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:12px}}
.card{{display:block;background:#fff;border:1px solid #e1e8f0;border-radius:12px;padding:16px;text-decoration:none;color:#172437;min-height:108px}}
.card:hover{{border-color:#9bb9d8;box-shadow:0 4px 14px rgba(16,34,53,.06)}}
.card b{{display:block;font-size:16px;margin-bottom:7px}}
.card span{{display:block;color:#68798c;font-size:13px;line-height:1.4}}
.tag{{display:inline-block;margin-top:10px;padding:3px 7px;border-radius:10px;background:#edf4ff;color:#275b91;font-size:11px}}
.adminbox{{margin-top:22px;background:#fff8e8;border:1px solid #efcf8b;border-radius:12px;padding:14px;display:flex;gap:10px;align-items:center;flex-wrap:wrap}}
.adminbox div{{margin-right:auto}} .adminbox span{{display:block;font-size:12px;color:#7b6a47;margin-top:3px}}
.adminbox a{{background:#fff;color:#704e05;border:1px solid #e6c36f;padding:8px 10px;border-radius:7px;text-decoration:none;font-size:12px;font-weight:700}}
.note{{margin-top:20px;background:#edf7f2;border:1px solid #c8e5d6;border-radius:10px;padding:12px;color:#355d48;font-size:13px}}
@media(max-width:650px){{.wrap{{padding:14px}}.grid{{grid-template-columns:1fr}}}}
</style>
</head>
<body>
<header>
  <div class="brand"><b>AI Deal Intelligence OS</b><small>Property · Hospitality · Retail · Demand</small></div>
  <div>{escape(role.upper())} · <a class="logout" href="/logout">Logout</a></div>
</header>

<div class="wrap">
  <div class="intro">
    <h1>Team Workspace</h1>
    <p>Only the pages required for daily work.</p>
  </div>

  <div class="section">
    <h2>PROPERTY</h2>
    <div class="grid">
      <a class="card" href="/property-database">
        <b>Property Database</b>
        <span>Search and open all saved properties and complete property records.</span>
        <span class="tag">Main database</span>
      </a>

      <a class="card" href="/property-manual">
        <b>Add Property + Matcher</b>
        <span>Add a property manually, create requirements and run property matching.</span>
        <span class="tag">Daily use</span>
      </a>

      <a class="card" href="/capture-intelligence">
        <b>Capture Property</b>
        <span>Upload camera photos, newspapers, handwritten notes, WhatsApp screenshots and PDFs.</span>
        <span class="tag">AI intake</span>
      </a>

      <a class="card" href="/inventory-activation">
        <b>Inventory Activation</b>
        <span>Review unmatched magazine inventory: Create New, Link Existing or Keep in Review.</span>
        <span class="tag">Refined magazine</span>
      </a>

      <a class="card" href="/contacts-directory">
        <b>Property Contacts</b>
        <span>Verify contacts and classify them as Owner, Broker, Both or Other.</span>
        <span class="tag">Verification</span>
      </a>
    </div>
  </div>

  <div class="section">
    <h2>LEAD INTELLIGENCE</h2>
    <div class="grid">
      <a class="card" href="/retail-expansion">
        <b>Retail Expansion Requirements</b>
        <span>Simple table of brand requirements, person, contact, location and source post.</span>
        <span class="tag">Retail</span>
      </a>

      <a class="card" href="/legacy-workspace#hospitality">
        <b>Hospitality Intelligence</b>
        <span>Restaurants, cafes, clubs, lounges, banquets, hotels and guest houses.</span>
        <span class="tag">Hospitality</span>
      </a>

      <a class="card" href="/legacy-workspace#requirements">
        <b>Requirement Discovery</b>
        <span>Review demand signals and property requirements discovered by AI.</span>
        <span class="tag">Demand</span>
      </a>

      <a class="card" href="/legacy-workspace#contacts">
        <b>Marketing Contacts</b>
        <span>Contact database for approved outreach and team follow-up.</span>
        <span class="tag">Marketing</span>
      </a>

      <a class="card" href="/legacy-workspace#bots">
        <b>Bot Control Room</b>
        <span>Run and review Hospitality, Retail and Requirement Discovery bots.</span>
        <span class="tag">AI bots</span>
      </a>
    </div>
  </div>

  {admin_tools}

  <div class="note">
    Public web signals are leads to qualify. Verify requirements and availability before outreach or sharing inventory.
  </div>
</div>
</body>
</html>""")

# Make /workspace the clean daily team workspace.
# The original technical V4 workspace is still available at /legacy-workspace.
@app.middleware("http")
async def v134_clean_workspace_router(request, call_next):
    if request.url.path == "/workspace":
        return RedirectResponse(url="/team-workspace-clean", status_code=307)
    return await call_next(request)

# ============================================================
# V13.7.2 UNIVERSAL REQUIREMENTS SYSTEM
# No dependency on V13.5/V13.6 markers.
# Retail + Hospitality, AI + Manual, with unified matching.
# ============================================================

def _v1372_table_columns(table_name):
    with engine.connect() as c:
        rows=c.execute(text("""SELECT column_name
            FROM information_schema.columns
            WHERE table_schema='public' AND table_name=:t"""),{"t":table_name}).fetchall()
    return {r._mapping["column_name"] for r in rows}

def _v1372_ensure_tables():
    with engine.begin() as c:
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_retail_manual_requirements(
            id BIGSERIAL PRIMARY KEY,
            requirement_id TEXT UNIQUE NOT NULL,
            company_name TEXT,
            contact_name TEXT,
            designation TEXT,
            contact_phone TEXT,
            contact_email TEXT,
            linkedin_profile_url TEXT,
            requirement_text TEXT,
            location TEXT,
            required_area_sqft NUMERIC,
            required_property_type TEXT,
            required_transaction TEXT DEFAULT 'LEASE',
            priority TEXT DEFAULT 'NORMAL',
            followup_status TEXT DEFAULT 'NEW',
            crm_status TEXT DEFAULT 'NOT_SENT',
            assigned_to TEXT,
            notes TEXT,
            created_by TEXT,
            created_at TIMESTAMPTZ DEFAULT NOW(),
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )"""))
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_hospitality_manual_requirements(
            id BIGSERIAL PRIMARY KEY,
            requirement_id TEXT UNIQUE NOT NULL,
            company_name TEXT,
            contact_name TEXT,
            designation TEXT,
            contact_phone TEXT,
            contact_email TEXT,
            requirement_text TEXT,
            location TEXT,
            required_area_sqft NUMERIC,
            required_property_type TEXT,
            required_transaction TEXT DEFAULT 'LEASE',
            priority TEXT DEFAULT 'NORMAL',
            followup_status TEXT DEFAULT 'NEW',
            assigned_to TEXT,
            notes TEXT,
            created_by TEXT,
            created_at TIMESTAMPTZ DEFAULT NOW(),
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )"""))
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_requirement_bridge(
            source_key TEXT PRIMARY KEY,
            source_division TEXT NOT NULL,
            source_type TEXT NOT NULL,
            source_id TEXT NOT NULL,
            pi_requirement_id TEXT,
            created_at TIMESTAMPTZ DEFAULT NOW(),
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )"""))

def _v1372_ai_classify(title, excerpt):
    txt=(" "+str(title or "")+" "+str(excerpt or "")+" ").lower()
    bad=[
        "job opportunity","hiring","vacancy","salary","store manager",
        "assistant store manager","web development","wordpress","full stack developer",
        "market report","market growth","rental yield","commercial vs residential",
        "office leasing market","retail leasing up","research says"
    ]
    if any(x in txt for x in bad):
        return "NOT_REQUIREMENT",15
    score=20
    for x in [
        "space requirement","looking for space","looking for retail space",
        "seeking retail space","actively seeking","looking to lease",
        "rental spaces","space required","immediate leasing opportunity"
    ]:
        if x in txt: score+=20
    if any(x in txt for x in ["sq ft","sqft","square feet","carpet area"]): score+=15
    if any(x in txt for x in ["delhi","gurgaon","gurugram","noida","south delhi","delhi ncr"]): score+=10
    if any(x in txt for x in ["lease","leasing","rent","rental"]): score+=10
    if score>=70:return "LIKELY_REQUIREMENT",min(100,score)
    if score>=50:return "POSSIBLE_REQUIREMENT",min(100,score)
    return "NOT_REQUIREMENT",min(100,score)

def _v1372_ai_rows(division):
    cols=_v1372_table_columns("ai_demand_signals")
    if not cols:
        return []
    with engine.connect() as c:
        rows=[dict(r._mapping) for r in c.execute(text("SELECT * FROM ai_demand_signals ORDER BY created_at DESC LIMIT 3000")).fetchall()]
    out=[]
    for r in rows:
        blob=" ".join(str(v or "") for v in r.values()).lower()
        if division=="RETAIL":
            if not any(x in blob for x in ["retail","store","shop"]): continue
            title=r.get("title") or r.get("signal_title") or ""
            excerpt=r.get("excerpt") or r.get("source_excerpt") or r.get("requirement_text") or ""
            cls,score=_v1372_ai_classify(title,excerpt)
        else:
            if not any(x in blob for x in ["hospitality","restaurant","cafe","banquet","hotel","guest house","lounge","club"]): continue
            cls="AI_SIGNAL"
            score=int(r.get("intent_score") or r.get("score") or 0)
        out.append({
            "division":division,
            "source_type":"AI",
            "source_id":str(r.get("signal_id") or r.get("id") or ""),
            "company_name":r.get("company_name") or r.get("company") or r.get("brand_name"),
            "contact_name":r.get("contact_name") or r.get("person_name"),
            "contact_phone":r.get("contact_phone") or r.get("phone") or r.get("mobile"),
            "contact_email":r.get("contact_email") or r.get("email"),
            "requirement_text":r.get("excerpt") or r.get("source_excerpt") or r.get("requirement_text") or r.get("title"),
            "location":r.get("location"),
            "required_area_sqft":r.get("required_area_sqft") or r.get("area_sqft"),
            "required_property_type":r.get("required_property_type") or r.get("category"),
            "required_transaction":r.get("required_transaction") or "LEASE",
            "system_classification":cls,
            "system_score":score,
            "source_url":r.get("source_url") or r.get("linkedin_post_url") or r.get("url")
        })
    return out

def _v1372_manual_rows(division):
    _v1372_ensure_tables()
    table="pi_retail_manual_requirements" if division=="RETAIL" else "pi_hospitality_manual_requirements"
    with engine.connect() as c:
        rows=[dict(r._mapping) for r in c.execute(text(f"SELECT * FROM {table} ORDER BY created_at DESC")).fetchall()]
    out=[]
    for r in rows:
        out.append({
            "division":division,
            "source_type":"MANUAL",
            "source_id":r.get("requirement_id"),
            "company_name":r.get("company_name"),
            "contact_name":r.get("contact_name"),
            "contact_phone":r.get("contact_phone"),
            "contact_email":r.get("contact_email"),
            "requirement_text":r.get("requirement_text"),
            "location":r.get("location"),
            "required_area_sqft":r.get("required_area_sqft"),
            "required_property_type":r.get("required_property_type"),
            "required_transaction":r.get("required_transaction"),
            "system_classification":"MANUAL_CONFIRMED",
            "system_score":100
        })
    return out

def _v1372_promote(payload):
    _v1372_ensure_tables()
    division=str(payload.get("division") or "").upper()
    source_type=str(payload.get("source_type") or "").upper()
    source_id=str(payload.get("source_id") or "")
    source_key=f"{division}|{source_type}|{source_id}"

    with engine.connect() as c:
        old=c.execute(text("SELECT pi_requirement_id FROM pi_requirement_bridge WHERE source_key=:k"),{"k":source_key}).fetchone()
        if old and old._mapping["pi_requirement_id"]:
            return str(old._mapping["pi_requirement_id"])

    cols=_v1372_table_columns("pi_requirements")
    if not cols:
        raise HTTPException(500,"pi_requirements table missing")

    rid="REQ-"+division[:3]+"-"+datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")+"-"+uuid.uuid4().hex[:5].upper()
    vals={
        "requirement_id":rid,"id":rid,
        "retailer_name":payload.get("company_name") or division.title()+" Requirement",
        "client_company":payload.get("company_name"),"company_name":payload.get("company_name"),
        "contact_name":payload.get("contact_name"),"contact_number":payload.get("contact_phone"),
        "contact_phone":payload.get("contact_phone"),"email":payload.get("contact_email"),
        "contact_email":payload.get("contact_email"),"city":"Delhi NCR","location":payload.get("location"),
        "requirement_sqft":payload.get("required_area_sqft"),"required_area_sqft":payload.get("required_area_sqft"),
        "minimum_area_sqft":payload.get("required_area_sqft"),"maximum_area_sqft":payload.get("required_area_sqft"),
        "retailers_purpose":payload.get("required_transaction") or "LEASE",
        "transaction_type":payload.get("required_transaction") or "LEASE",
        "rent_or_sale":payload.get("required_transaction") or "LEASE",
        "retailers_category":payload.get("required_property_type") or division,
        "required_property_type":payload.get("required_property_type"),
        "property_type":payload.get("required_property_type"),
        "nearby_brands":"",
        "additional_points":payload.get("requirement_text") or "",
        "remarks":payload.get("requirement_text") or "",
        "source":source_type,"division":division,"status":"NEW"
    }
    insert_cols=[];params={}
    for k,v in vals.items():
        if k in cols and k not in insert_cols:
            insert_cols.append(k);params[k]=v
    if "requirement_id" not in insert_cols and "id" not in insert_cols:
        raise HTTPException(500,"No compatible requirement ID field in pi_requirements")

    with engine.begin() as c:
        c.execute(text("INSERT INTO pi_requirements("+",".join(insert_cols)+") VALUES("+",".join(":"+x for x in insert_cols)+")"),params)
        c.execute(text("""INSERT INTO pi_requirement_bridge(
            source_key,source_division,source_type,source_id,pi_requirement_id,updated_at
        ) VALUES(:k,:d,:t,:sid,:rid,NOW())
        ON CONFLICT(source_key) DO UPDATE SET pi_requirement_id=EXCLUDED.pi_requirement_id,updated_at=NOW()"""),
        {"k":source_key,"d":division,"t":source_type,"sid":source_id,"rid":rid})
    return rid

@app.post("/api/v13-7-2/manual/{division}")
async def v1372_manual_add(division:str,req:Request):
    need_login(req);_v1372_ensure_tables()
    div=division.upper()
    if div not in {"RETAIL","HOSPITALITY"}:raise HTTPException(400,"Invalid division")
    body=await req.json()
    rid=("RMR-" if div=="RETAIL" else "HMR-")+datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")+"-"+uuid.uuid4().hex[:5].upper()
    table="pi_retail_manual_requirements" if div=="RETAIL" else "pi_hospitality_manual_requirements"
    with engine.begin() as c:
        c.execute(text(f"""INSERT INTO {table}(
            requirement_id,company_name,contact_name,designation,contact_phone,contact_email,
            requirement_text,location,required_area_sqft,required_property_type,
            required_transaction,priority,followup_status,assigned_to,notes,created_by
        ) VALUES(:rid,:co,:cn,:des,:ph,:em,:req,:loc,:area,:pt,:tr,:pri,'NEW',:asgn,:notes,:by)"""),
        {"rid":rid,"co":body.get("company_name"),"cn":body.get("contact_name"),
         "des":body.get("designation"),"ph":body.get("contact_phone"),"em":body.get("contact_email"),
         "req":body.get("requirement_text"),"loc":body.get("location"),
         "area":body.get("required_area_sqft") or None,"pt":body.get("required_property_type"),
         "tr":body.get("required_transaction") or "LEASE","pri":body.get("priority") or "NORMAL",
         "asgn":body.get("assigned_to"),"notes":body.get("notes"),"by":actor_name(req)})
    return {"status":"ok","requirement_id":rid}

@app.get("/api/v13-7-2/requirements")
def v1372_requirements(req:Request,division:str=Query("ALL"),source_type:str=Query("ALL")):
    need_login(req);_v1372_ensure_tables()
    division=division.upper();source_type=source_type.upper();rows=[]
    for div in ["RETAIL","HOSPITALITY"]:
        if division not in {"ALL",div}:continue
        if source_type in {"ALL","MANUAL"}:rows.extend(_v1372_manual_rows(div))
        if source_type in {"ALL","AI"}:rows.extend(_v1372_ai_rows(div))
    return {"status":"ok","rows":rows}

@app.post("/api/v13-7-2/match")
async def v1372_match(req:Request):
    need_login(req)
    body=await req.json()
    rid=_v1372_promote(body)
    result=robust_match_requirement(rid,create_whatsapp=False)
    return {"status":"ok","pi_requirement_id":rid,**result}

@app.get("/requirements-match-center",response_class=HTMLResponse)
def v1372_center(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",status_code=303)
    return HTMLResponse("""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Requirements + Matches</title><style>
body{font-family:Arial;margin:0;background:#f4f7fb;color:#172437}header{background:#102235;color:#fff;padding:18px}.wrap{padding:18px}.nav{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px}.nav a{background:#e9eef5;padding:8px 10px;border-radius:7px;text-decoration:none;color:#203247;font-weight:bold}.card{background:#fff;border:1px solid #e2e8f0;border-radius:12px;padding:12px;margin-bottom:10px}.toolbar{display:flex;gap:8px;flex-wrap:wrap}.pill{padding:3px 7px;border-radius:10px;background:#edf4ff}.manual{background:#dcfce7}.ai{background:#fef3c7}.matches{margin-top:8px;background:#f8fafc;padding:8px;border-radius:7px}</style></head><body>
<header><b>Requirements + Matches</b><br><small>Retail + Hospitality · AI + Manual</small></header><div class="wrap">
<div class="nav"><a href="/workspace">← Team Workspace</a><a href="/requirements-entry?division=HOSPITALITY">Hospitality Entry</a><a href="/requirements-entry?division=RETAIL">Retail Entry</a><a href="/property-database">Property Database</a></div>
<div class="toolbar"><select id="division"><option>ALL</option><option>RETAIL</option><option>HOSPITALITY</option></select><select id="source"><option>ALL</option><option>MANUAL</option><option>AI</option></select><input id="q" placeholder="Search"><button onclick="load()">Refresh</button></div><div id="rows"></div></div>
<script>
const E=x=>String(x??'').replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));let D=[];
async function A(u,o={}){let r=await fetch(u,o),d=await r.json();if(!r.ok)throw Error(d.detail||'Error');return d}
function render(){let s=(q.value||'').toLowerCase(),R=D.filter(x=>!s||JSON.stringify(x).toLowerCase().includes(s));rows.innerHTML=R.map((x,i)=>`<div class="card"><b>${E(x.division)} · <span class="pill ${x.source_type==='MANUAL'?'manual':'ai'}">${E(x.source_type)}</span> · ${E(x.company_name||'To verify')}</b><br>${E(x.requirement_text||'')}<br>${E(x.location||'')} · ${E(x.required_area_sqft||'')} · ${E(x.required_property_type||'')}<br><button onclick="match(${i})">Run Match</button><div class="matches" id="m_${i}">Not run</div></div>`).join('')||'No requirements found.'}
async function load(){let d=await A('/api/v13-7-2/requirements?division='+division.value+'&source_type='+source.value);D=d.rows||[];render()}
async function match(i){let x=D[i],box=document.getElementById('m_'+i);box.textContent='Matching...';try{let d=await A('/api/v13-7-2/match',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(x)});let ms=d.matches||[];box.innerHTML=ms.slice(0,5).map((m,j)=>`${j+1}. <a target="_blank" href="/property-record/${encodeURIComponent(m.property_id)}">${E(m.property_name||m.property_id)}</a> · Score ${E(m.score||'')}`).join('<br>')||'No matches';}catch(e){box.textContent=e.message}}
division.addEventListener('change',load);source.addEventListener('change',load);q.addEventListener('input',render);load();
</script></body></html>""")

@app.get("/requirements-entry",response_class=HTMLResponse)
def v1372_entry(req:Request,division:str=Query("RETAIL")):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",status_code=303)
    div=division.upper()
    if div not in {"RETAIL","HOSPITALITY"}:div="RETAIL"
    return HTMLResponse(f"""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{div.title()} Requirement Entry</title></head>
<body style="font-family:Arial;background:#f4f7fb;margin:0;color:#172437"><div style="background:#102235;color:white;padding:18px"><b>{div.title()} Requirement Entry</b></div><div style="padding:18px;max-width:1000px;margin:auto">
<p><a href="/workspace">← Team Workspace</a> · <a href="/requirements-match-center">Requirements + Matches</a></p>
<div style="background:white;padding:15px;border-radius:12px"><form id="f">
<p><input name="company_name" placeholder="Company / Brand *" required> <input name="contact_name" placeholder="Contact Person"> <input name="contact_phone" placeholder="Mobile"></p>
<p><input name="location" placeholder="Location *" required> <input name="required_area_sqft" type="number" placeholder="Area SqFt"> <input name="required_property_type" placeholder="Property Type"></p>
<p><textarea name="requirement_text" style="width:100%;min-height:90px" placeholder="Requirement details *" required></textarea></p><button>Save Manual Requirement</button></form><p id="msg"></p></div></div>
<script>f.addEventListener('submit',async e=>{{e.preventDefault();let b=Object.fromEntries(new FormData(f));b.required_area_sqft=b.required_area_sqft?Number(b.required_area_sqft):null;let r=await fetch('/api/v13-7-2/manual/{div}',{{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify(b)}}),d=await r.json();if(!r.ok){{msg.textContent=d.detail||'Error';return}}msg.innerHTML='Saved <b>'+d.requirement_id+'</b>. <a href="/requirements-match-center">Open Requirements + Matches</a>';f.reset()}})</script></body></html>""")

# ============================================================
# V13.8 FINAL SIMPLE DASHBOARD + REQUIREMENT MATCHER FIX
# - Simple daily dashboard
# - Add Property is separate from Property Matcher
# - Requirements page has TWO columns: AI Generated / Manual
# - Retail and Hospitality are separate filters
# - Match creation introspects pi_requirements safely
# ============================================================

def _v138_cols_meta(table_name):
    with engine.connect() as c:
        rows=c.execute(text("""SELECT column_name,data_type,is_nullable,column_default
            FROM information_schema.columns
            WHERE table_schema='public' AND table_name=:t
            ORDER BY ordinal_position"""),{"t":table_name}).fetchall()
    return [dict(r._mapping) for r in rows]

def _v138_table_exists(table_name):
    with engine.connect() as c:
        return bool(c.execute(text("""SELECT EXISTS(
            SELECT 1 FROM information_schema.tables
            WHERE table_schema='public' AND table_name=:t
        )"""),{"t":table_name}).scalar_one())

def _v138_ensure_tables():
    with engine.begin() as c:
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_unified_manual_requirements(
            requirement_id TEXT PRIMARY KEY,
            division TEXT NOT NULL,
            company_name TEXT,
            contact_name TEXT,
            contact_phone TEXT,
            contact_email TEXT,
            location TEXT,
            required_area_sqft NUMERIC,
            required_property_type TEXT,
            required_transaction TEXT DEFAULT 'LEASE',
            requirement_text TEXT,
            assigned_to TEXT,
            status TEXT DEFAULT 'NEW',
            created_by TEXT,
            created_at TIMESTAMPTZ DEFAULT NOW(),
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )"""))
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_unified_requirement_bridge(
            source_key TEXT PRIMARY KEY,
            division TEXT NOT NULL,
            source_type TEXT NOT NULL,
            source_id TEXT NOT NULL,
            pi_requirement_id TEXT,
            last_error TEXT,
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )"""))

def _v138_classify_ai(title,excerpt):
    txt = (" " + str(title or "") + " " + str(excerpt or "") + " ").lower()

    noise = [
        "job opportunity","job vacancy","hiring","salary","store manager",
        "assistant store manager","full stack developer","web development",
        "market report","market growth","rental yield","commercial vs residential",
        "office leasing market","research says","real estate leader",
        "modern trade network","luxury sales"
    ]
    if any(x in txt for x in noise):
        return "NOT_REQUIREMENT", 10

    supply = [
        "property for sale","commercial property for sale","available for rent",
        "available on rent","property available","shop available","space available",
        "for sale in","for rent in","per sqft","/ sqft","suitable for restaurant",
        "suitable for cafe","suitable for salon","suitable for saloon"
    ]
    if sum(1 for x in supply if x in txt) >= 2:
        return "PROPERTY_SUPPLY_NOT_REQUIREMENT", 10

    score = 20
    for x in [
        "space requirement","looking for space","looking for retail space",
        "seeking retail space","actively seeking","looking to lease",
        "rental spaces","space required","immediate leasing opportunity",
        "looking to lease hotels","requirement: retail space"
    ]:
        if x in txt:
            score += 20
    if any(x in txt for x in ["sq ft","sqft","square feet","carpet area"]):
        score += 15
    if any(x in txt for x in ["lease","leasing","rent","rental"]):
        score += 10

    if score >= 70:
        return "LIKELY_REQUIREMENT", min(score,100)
    if score >= 50:
        return "POSSIBLE_REQUIREMENT", min(score,100)
    return "LOW_CONFIDENCE", min(score,100)

def _v138_ai_rows(division):
    if not _v138_table_exists("ai_demand_signals"):
        return []
    with engine.connect() as c:
        rows=[dict(r._mapping) for r in c.execute(text(
            "SELECT * FROM ai_demand_signals ORDER BY created_at DESC LIMIT 2500"
        )).fetchall()]
    out=[]
    for r in rows:
        blob=" ".join(str(v or "") for v in r.values()).lower()
        if division=="RETAIL":
            if not any(x in blob for x in ["retail","store","shop","pharmacy"]):continue
        else:
            if not any(x in blob for x in ["hospitality","restaurant","cafe","banquet","hotel","guest house","lounge","club"]):continue
        title=r.get("title") or r.get("signal_title") or ""
        excerpt=r.get("excerpt") or r.get("source_excerpt") or r.get("requirement_text") or title
        cls,score=_v138_classify_ai(title,excerpt)
        out.append({
            "division":division,"source_type":"AI",
            "source_id":str(r.get("signal_id") or r.get("id") or ""),
            "company_name":r.get("company_name") or r.get("brand_name") or r.get("company"),
            "contact_name":r.get("contact_name") or r.get("person_name"),
            "contact_phone":r.get("contact_phone") or r.get("phone") or r.get("mobile"),
            "contact_email":r.get("contact_email") or r.get("email"),
            "location":r.get("location") or "Delhi NCR",
            "required_area_sqft":r.get("required_area_sqft") or r.get("area_sqft"),
            "required_property_type":r.get("required_property_type") or r.get("category"),
            "required_transaction":r.get("required_transaction") or "LEASE",
            "requirement_text":excerpt,
            "source_url":r.get("source_url") or r.get("linkedin_post_url") or r.get("url"),
            "classification":cls,"confidence":score
        })
    return out

def _v138_manual_rows(division):
    _v138_ensure_tables()
    with engine.connect() as c:
        rows=c.execute(text("""SELECT * FROM pi_unified_manual_requirements
            WHERE division=:d ORDER BY created_at DESC"""),{"d":division}).fetchall()
    return [dict(r._mapping)|{"source_type":"MANUAL","source_id":r._mapping["requirement_id"],
            "classification":"MANUAL","confidence":100} for r in rows]

def _v138_requirement_insert(payload):
    if not _v138_table_exists("pi_requirements"):
        raise RuntimeError("pi_requirements table is missing")

    meta = _v138_cols_meta("pi_requirements")
    rid = "REQ-" + str(payload.get("division") or "GEN")[:3].upper() + "-" + datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S") + "-" + uuid.uuid4().hex[:5].upper()

    aliases = {
        "requirement_id": rid,
        "retailer_name": payload.get("company_name") or "Requirement",
        "client_company": payload.get("company_name") or "Requirement",
        "company_name": payload.get("company_name") or "Requirement",
        "contact_name": payload.get("contact_name"),
        "contact_number": payload.get("contact_phone"),
        "contact_phone": payload.get("contact_phone"),
        "email": payload.get("contact_email"),
        "contact_email": payload.get("contact_email"),
        "city": payload.get("city") or "Delhi NCR",
        "location": payload.get("location") or "Delhi NCR",
        "requirement_sqft": payload.get("required_area_sqft") or 0,
        "required_area_sqft": payload.get("required_area_sqft") or 0,
        "minimum_area_sqft": payload.get("required_area_sqft") or 0,
        "maximum_area_sqft": payload.get("required_area_sqft") or 0,
        "retailers_purpose": payload.get("required_transaction") or "LEASE",
        "transaction_type": payload.get("required_transaction") or "LEASE",
        "rent_or_sale": payload.get("required_transaction") or "LEASE",
        "retailers_category": payload.get("required_property_type") or payload.get("division") or "COMMERCIAL",
        "required_property_type": payload.get("required_property_type") or payload.get("division") or "COMMERCIAL",
        "property_type": payload.get("required_property_type") or payload.get("division") or "COMMERCIAL",
        "nearby_brands": "",
        "additional_points": payload.get("requirement_text") or "",
        "remarks": payload.get("requirement_text") or "",
        "source": payload.get("source_type") or "MANUAL",
        "division": payload.get("division"),
        "status": "NEW"
    }

    values = {}
    for m in meta:
        name = m["column_name"]
        dt = (m.get("data_type") or "").lower()
        default = m.get("column_default")
        required = m.get("is_nullable") == "NO"
        numeric = any(x in dt for x in ["bigint","integer","smallint","numeric","double precision","real"])

        # Critical fix: never put REQ-... text into numeric pi_requirements.id.
        if name == "id" and numeric:
            if default:
                continue
            with engine.connect() as c:
                values[name] = int(c.execute(text("SELECT COALESCE(MAX(id),0)+1 FROM pi_requirements")).scalar_one() or 1)
            continue

        if name in aliases:
            val = aliases[name]
            if numeric:
                try:
                    val = 0 if val in (None, "") else float(val)
                    if any(x in dt for x in ["bigint","integer","smallint"]):
                        val = int(val)
                except Exception:
                    val = 0
            values[name] = val
        elif required and not default:
            if numeric:
                values[name] = 0
            elif "boolean" in dt:
                values[name] = False
            elif "timestamp" in dt or dt == "date":
                continue
            else:
                values[name] = ""

    if not values:
        raise RuntimeError("No compatible pi_requirements columns found")

    names = list(values)
    sql = "INSERT INTO pi_requirements(" + ",".join(names) + ") VALUES(" + ",".join(":" + x for x in names) + ") RETURNING *"
    with engine.begin() as c:
        row = c.execute(text(sql), values).fetchone()

    if row:
        d = dict(row._mapping)
        return str(d.get("requirement_id") or d.get("id") or rid)
    return rid

def _v138_promote(payload):
    _v138_ensure_tables()
    key=f"{payload.get('division')}|{payload.get('source_type')}|{payload.get('source_id')}"
    with engine.connect() as c:
        old=c.execute(text("SELECT pi_requirement_id FROM pi_unified_requirement_bridge WHERE source_key=:k"),{"k":key}).fetchone()
        if old and old._mapping["pi_requirement_id"]:
            return str(old._mapping["pi_requirement_id"])
    try:
        rid=_v138_requirement_insert(payload)
        err=None
    except Exception as ex:
        with engine.begin() as c:
            c.execute(text("""INSERT INTO pi_unified_requirement_bridge(
                source_key,division,source_type,source_id,last_error,updated_at
            ) VALUES(:k,:d,:t,:sid,:e,NOW())
            ON CONFLICT(source_key) DO UPDATE SET last_error=EXCLUDED.last_error,updated_at=NOW()"""),
            {"k":key,"d":payload.get("division"),"t":payload.get("source_type"),
             "sid":payload.get("source_id"),"e":f"{type(ex).__name__}: {ex}"})
        raise
    with engine.begin() as c:
        c.execute(text("""INSERT INTO pi_unified_requirement_bridge(
            source_key,division,source_type,source_id,pi_requirement_id,last_error,updated_at
        ) VALUES(:k,:d,:t,:sid,:rid,NULL,NOW())
        ON CONFLICT(source_key) DO UPDATE SET pi_requirement_id=EXCLUDED.pi_requirement_id,
            last_error=NULL,updated_at=NOW()"""),
        {"k":key,"d":payload.get("division"),"t":payload.get("source_type"),
         "sid":payload.get("source_id"),"rid":rid})
    return rid

@app.post("/api/v13-8/manual")
async def v138_manual_add(req:Request):
    need_login(req);_v138_ensure_tables()
    body=await req.json()
    div=str(body.get("division") or "").upper()
    if div not in {"RETAIL","HOSPITALITY"}:raise HTTPException(400,"Choose RETAIL or HOSPITALITY")
    rid=("RMR-" if div=="RETAIL" else "HMR-")+datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")+"-"+uuid.uuid4().hex[:5].upper()
    with engine.begin() as c:
        c.execute(text("""INSERT INTO pi_unified_manual_requirements(
            requirement_id,division,company_name,contact_name,contact_phone,contact_email,
            location,required_area_sqft,required_property_type,required_transaction,
            requirement_text,assigned_to,created_by
        ) VALUES(:rid,:d,:co,:cn,:ph,:em,:loc,:area,:pt,:tr,:req,:asgn,:by)"""),
        {"rid":rid,"d":div,"co":body.get("company_name"),"cn":body.get("contact_name"),
         "ph":body.get("contact_phone"),"em":body.get("contact_email"),"loc":body.get("location"),
         "area":body.get("required_area_sqft") or None,"pt":body.get("required_property_type"),
         "tr":body.get("required_transaction") or "LEASE","req":body.get("requirement_text"),
         "asgn":body.get("assigned_to"),"by":actor_name(req)})
    return {"status":"ok","requirement_id":rid}

@app.get("/api/v13-8/requirements")
def v138_requirements(req:Request,division:str=Query("RETAIL")):
    need_login(req);div=division.upper()
    if div not in {"RETAIL","HOSPITALITY"}:div="RETAIL"
    return {"status":"ok","ai":_v138_ai_rows(div),"manual":_v138_manual_rows(div)}

@app.post("/api/v13-8/match")
async def v138_match(req:Request):
    need_login(req)
    payload=await req.json()
    try:
        rid=_v138_promote(payload)
        result=robust_match_requirement(rid,create_whatsapp=False)
        return {"status":"ok","requirement_id":rid,**result}
    except HTTPException:
        raise
    except Exception as ex:
        import traceback
        traceback.print_exc()
        raise HTTPException(500,f"{type(ex).__name__}: {str(ex)}")

@app.get("/simple-dashboard",response_class=HTMLResponse)
def v138_dashboard(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",status_code=303)
    return HTMLResponse("""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>AI Deal Intelligence OS</title>
<style>*{box-sizing:border-box}body{margin:0;background:#f4f7fb;font-family:Arial;color:#172437}header{background:#102235;color:white;padding:20px}.wrap{max-width:1300px;margin:auto;padding:20px}.grid{display:grid;grid-template-columns:repeat(3,1fr);gap:12px}.card{display:block;background:white;border:1px solid #e2e8f0;border-radius:12px;padding:18px;text-decoration:none;color:#172437;min-height:120px}.card b{font-size:17px}.card span{display:block;color:#687789;margin-top:8px;line-height:1.4}.main{border:2px solid #8eb9e6}.section{margin:20px 0 10px;font-size:13px;color:#63768a;font-weight:bold;letter-spacing:.08em}@media(max-width:800px){.grid{grid-template-columns:1fr}}</style></head>
<body><header><b>AI Deal Intelligence OS</b><br><small>Simple Team Dashboard</small></header><div class="wrap">
<div class="section">DAILY WORK</div><div class="grid">
<a class="card main" href="/requirements-workbench"><b>Property Matcher</b><span>Create/select Retail or Hospitality requirement, see AI vs Manual separately and run matching.</span></a>
<a class="card" href="/property-manual"><b>Add Property Manually</b><span>Use only when your team wants to add a new property manually.</span></a>
<a class="card" href="/property-database"><b>Property Database</b><span>Search and open all saved property inventory.</span></a>
<a class="card" href="/capture-intelligence"><b>Capture Property</b><span>Camera, newspaper, handwritten note, WhatsApp screenshot or PDF.</span></a>
<a class="card" href="/inventory-activation"><b>Inventory Activation</b><span>Review unmatched refined magazine properties.</span></a>
<a class="card" href="/contacts-directory"><b>Property Contacts</b><span>Verify contact and mark Owner, Broker, Both or Other.</span></a>
</div>
<div class="section">LEADS & AI</div><div class="grid">
<a class="card" href="/requirements-workbench?division=RETAIL"><b>Retail Requirements</b><span>Two columns: AI Generated and Manual.</span></a>
<a class="card" href="/requirements-workbench?division=HOSPITALITY"><b>Hospitality Requirements</b><span>Two columns: AI Generated and Manual.</span></a>
<a class="card" href="/legacy-workspace#bots"><b>Bot Control Room</b><span>Run background discovery bots.</span></a>
</div></div></body></html>""")

@app.get("/requirements-workbench",response_class=HTMLResponse)
def v138_workbench(req:Request,division:str=Query("RETAIL")):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",status_code=303)
    div=division.upper() if division.upper() in {"RETAIL","HOSPITALITY"} else "RETAIL"
    return HTMLResponse(f"""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Requirements Workbench</title>
<style>*{{box-sizing:border-box}}body{{margin:0;background:#f4f7fb;font-family:Arial;color:#172437}}header{{background:#102235;color:white;padding:18px}}.wrap{{padding:18px;max-width:1800px;margin:auto}}.nav{{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px}}.nav a,.btn{{background:#e9eef5;color:#203247;border:0;padding:8px 10px;border-radius:7px;text-decoration:none;font-weight:bold;cursor:pointer}}.primary{{background:#1677ff;color:white}}.columns{{display:grid;grid-template-columns:1fr 1fr;gap:12px}}.col{{background:white;border:1px solid #e2e8f0;border-radius:12px;padding:12px}}.item{{border:1px solid #e5ebf1;border-radius:10px;padding:10px;margin-bottom:8px}}.good{{border-left:5px solid #30936b}}.bad{{opacity:.6}}.matches{{margin-top:8px;background:#f8fafc;padding:8px;border-radius:7px}}input,select,textarea{{padding:8px;border:1px solid #ccd6e2;border-radius:7px}}textarea{{width:100%;min-height:70px}}.formgrid{{display:grid;grid-template-columns:repeat(2,1fr);gap:7px}}.full{{grid-column:1/-1}}@media(max-width:950px){{.columns{{grid-template-columns:1fr}}}}</style></head>
<body><header><b>{div.title()} Requirements + Property Matcher</b><br><small>AI Generated and Manual are kept separate</small></header><div class="wrap">
<div class="nav"><a href="/simple-dashboard">← Dashboard</a><a href="/requirements-workbench?division=RETAIL">Retail</a><a href="/requirements-workbench?division=HOSPITALITY">Hospitality</a><a href="/property-database">Property Database</a><a href="/property-manual">Add Property Manually</a></div>
<div class="columns">
<div class="col"><h2>AI Generated Requirements</h2><p>AI/public-web signals. Ignore NOT_REQUIREMENT items.</p><div id="ai"></div></div>
<div class="col"><h2>Manual Requirements</h2>
<form id="f" class="formgrid"><input name="company_name" placeholder="Company / Brand *" required><input name="contact_name" placeholder="Contact Person"><input name="contact_phone" placeholder="Mobile"><input name="contact_email" placeholder="Email"><input name="location" placeholder="Location *" required><input name="required_area_sqft" type="number" placeholder="Area SqFt"><input name="required_property_type" placeholder="Property Type"><select name="required_transaction"><option>LEASE</option><option>RENT</option><option>SALE</option></select><textarea class="full" name="requirement_text" placeholder="Requirement details *" required></textarea><input name="assigned_to" placeholder="Assigned Team Member"><button class="btn primary">Save Manual Requirement</button></form><hr><div id="manual"></div></div>
</div></div>
<script>
const DIV='{div}',E=x=>String(x??'').replace(/[&<>"]/g,c=>({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}}[c]));let AI=[],MAN=[];
async function A(u,o={{}}){{let r=await fetch(u,o),d=await r.json();if(!r.ok)throw Error(d.detail||'Error');return d}}
function item(x,i,src){{let bad=x.classification==='NOT_REQUIREMENT';return `<div class="item ${{bad?'bad':'good'}}"><b>${{E(x.company_name||'To verify')}}</b> · ${{E(x.classification||src)}}<br>${{E(x.requirement_text||'')}}<br>${{E(x.location||'')}} · ${{E(x.required_area_sqft||'')}} · ${{E(x.required_property_type||'')}}<br>${{x.source_url?`<a target="_blank" href="${{E(x.source_url)}}">Open source</a> · `:''}}<button class="btn primary" onclick="runMatch('${{src}}',${{i}})">Run Match</button><div class="matches" id="${{src}}_${{i}}">Not run</div></div>`}}
function render(){{ai.innerHTML=AI.map((x,i)=>item(x,i,'AI')).join('')||'No AI requirements.';manual.innerHTML=MAN.map((x,i)=>item(x,i,'MANUAL')).join('')||'No manual requirements.'}}
async function load(){{let d=await A('/api/v13-8/requirements?division='+DIV);AI=d.ai||[];MAN=d.manual||[];render()}}
async function runMatch(src,i){{let x=(src==='AI'?AI:MAN)[i],box=document.getElementById(src+'_'+i);if(['NOT_REQUIREMENT','PROPERTY_SUPPLY_NOT_REQUIREMENT'].includes(x.classification)&&!confirm('AI classified this as '+x.classification+'. Run anyway?'))return;box.textContent='Matching...';try{{let d=await A('/api/v13-8/match',{{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify(x)}});let ms=d.matches||[];box.innerHTML=ms.slice(0,10).map((m,j)=>`${{j+1}}. <a target="_blank" href="/property-record/${{encodeURIComponent(m.property_id)}}">${{E(m.property_name||m.property_id)}}</a> · Score <b>${{E(m.score||'')}}</b>`).join('<br>')||'No matches';}}catch(e){{box.innerHTML='<b>ERROR:</b> '+E(e.message)}}}}
f.addEventListener('submit',async e=>{{e.preventDefault();let b=Object.fromEntries(new FormData(f));b.division=DIV;b.required_area_sqft=b.required_area_sqft?Number(b.required_area_sqft):null;let d=await A('/api/v13-8/manual',{{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify(b)}});alert('Saved '+d.requirement_id);f.reset();load()}});load();
</script></body></html>""")

@app.middleware("http")
async def v138_workspace_redirect(request,call_next):
    if request.url.path=="/workspace":
        return RedirectResponse("/simple-dashboard",status_code=307)
    return await call_next(request)

# V13.8.2 CORRECTED BIGINT MATCHER FIX

# ============================================================
# V14 FRESH VERIFIED INVENTORY + CONFIRMED REQUIREMENT OS
# Magazine/legacy data is NOT used by this matcher.
# Fresh properties -> verify -> activate -> match.
# AI signals -> human confirm -> manual confirmed requirement -> match.
# Pictures stored in PostgreSQL so Railway redeploys do not lose them.
# ============================================================

def _v14_setup():
    with engine.begin() as c:
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_v14_properties(
            id BIGSERIAL PRIMARY KEY,
            property_code TEXT UNIQUE NOT NULL,
            property_name TEXT,
            city TEXT NOT NULL,
            location TEXT NOT NULL,
            micro_market TEXT,
            full_address TEXT,
            property_type TEXT NOT NULL,
            suitable_category TEXT,
            transaction_type TEXT NOT NULL,
            availability_status TEXT DEFAULT 'UNVERIFIED',
            matcher_eligible BOOLEAN DEFAULT FALSE,
            available_area_sqft NUMERIC NOT NULL,
            floor TEXT,
            frontage_ft NUMERIC,
            monthly_rent NUMERIC,
            rent_psf NUMERIC,
            sale_price NUMERIC,
            cam_psf NUMERIC,
            security_deposit_months NUMERIC,
            possession TEXT,
            parking TEXT,
            ceiling_height_ft NUMERIC,
            power_load_kw NUMERIC,
            nearby_brands TEXT,
            main_road BOOLEAN DEFAULT FALSE,
            corner_property BOOLEAN DEFAULT FALSE,
            food_use_allowed TEXT,
            exhaust_possible TEXT,
            notes TEXT,
            contact_name TEXT,
            contact_phone TEXT,
            contact_role TEXT DEFAULT 'UNVERIFIED',
            contact_verified BOOLEAN DEFAULT FALSE,
            verified_by TEXT,
            verified_at TIMESTAMPTZ,
            created_by TEXT,
            created_at TIMESTAMPTZ DEFAULT NOW(),
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )"""))
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_v14_property_media(
            id BIGSERIAL PRIMARY KEY,
            property_code TEXT NOT NULL,
            filename TEXT,
            content_type TEXT,
            content BYTEA NOT NULL,
            created_at TIMESTAMPTZ DEFAULT NOW()
        )"""))
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_v14_requirements(
            id BIGSERIAL PRIMARY KEY,
            requirement_code TEXT UNIQUE NOT NULL,
            division TEXT NOT NULL,
            company_name TEXT NOT NULL,
            contact_name TEXT,
            contact_phone TEXT,
            contact_email TEXT,
            source_type TEXT DEFAULT 'MANUAL_CONFIRMED',
            source_url TEXT,
            city TEXT NOT NULL,
            locations TEXT NOT NULL,
            property_type TEXT NOT NULL,
            suitable_category TEXT,
            transaction_type TEXT NOT NULL,
            area_min_sqft NUMERIC NOT NULL,
            area_max_sqft NUMERIC NOT NULL,
            ideal_area_sqft NUMERIC,
            floor_preference TEXT,
            ground_floor_mandatory BOOLEAN DEFAULT FALSE,
            max_monthly_rent NUMERIC,
            max_rent_psf NUMERIC,
            max_sale_budget NUMERIC,
            min_frontage_ft NUMERIC,
            possession TEXT,
            parking_required TEXT,
            nearby_brands TEXT,
            main_road_required BOOLEAN DEFAULT FALSE,
            corner_preferred BOOLEAN DEFAULT FALSE,
            food_use_required BOOLEAN DEFAULT FALSE,
            exhaust_required BOOLEAN DEFAULT FALSE,
            must_have TEXT,
            preferred TEXT,
            reject_if TEXT,
            assigned_to TEXT,
            confirmation_status TEXT DEFAULT 'CONFIRMED',
            created_by TEXT,
            created_at TIMESTAMPTZ DEFAULT NOW(),
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )"""))

def _v14_s(v):
    return str(v or "").strip()

def _v14_n(v):
    try: return float(v) if v not in (None,"") else None
    except: return None

def _v14_b(v):
    return str(v or "").lower() in {"true","1","yes","on","y"}

def _v14_norm(v):
    return _re.sub(r"[^a-z0-9]+"," ",_v14_s(v).lower()).strip()

def _v14_actor(req):
    try: return actor_name(req)
    except: return "TEAM"

def _v14_score(p,q):
    reasons=[]; penalties=[]
    # Hard filters
    if _v14_norm(p.get("city")) != _v14_norm(q.get("city")):
        return None,["City mismatch"]
    qlocs=[_v14_norm(x) for x in _v14_s(q.get("locations")).split(",") if _v14_norm(x)]
    ploc=_v14_norm(p.get("location"))
    if qlocs and not any(x in ploc or ploc in x for x in qlocs):
        return None,["Location mismatch"]
    if _v14_norm(p.get("transaction_type")) != _v14_norm(q.get("transaction_type")):
        return None,["Transaction mismatch"]
    area=_v14_n(p.get("available_area_sqft")) or 0
    amin=_v14_n(q.get("area_min_sqft")) or 0
    amax=_v14_n(q.get("area_max_sqft")) or 10**12
    if area < amin or area > amax:
        return None,["Area outside required range"]
    if not p.get("matcher_eligible") or _v14_norm(p.get("availability_status"))!="verified active":
        return None,["Not verified active"]

    if q.get("ground_floor_mandatory"):
        fl=_v14_norm(p.get("floor"))
        if not any(x in fl for x in ["ground","gf","upper ground","ug"]):
            return None,["Ground floor mandatory"]

    score=60
    reasons += ["City","Location","Transaction","Area","Verified active"]

    # Property/category
    qtype=_v14_norm(q.get("property_type")); ptype=_v14_norm(p.get("property_type"))
    if qtype and ptype and (qtype in ptype or ptype in qtype):
        score+=10; reasons.append("Property type")
    else:
        penalties.append("Property type differs"); score-=8

    qcat=_v14_norm(q.get("suitable_category")); pcat=_v14_norm(p.get("suitable_category"))
    if qcat and pcat and (qcat in pcat or pcat in qcat):
        score+=6; reasons.append("Suitable category")

    ideal=_v14_n(q.get("ideal_area_sqft"))
    if ideal and area:
        diff=abs(area-ideal)/ideal
        if diff<=.10: score+=8; reasons.append("Near ideal area")
        elif diff<=.20: score+=4

    # Floor preference
    qfloor=_v14_norm(q.get("floor_preference")); pfloor=_v14_norm(p.get("floor"))
    if qfloor and pfloor:
        if qfloor in pfloor or pfloor in qfloor:
            score+=6; reasons.append("Floor")
        else:
            score-=4; penalties.append("Floor preference differs")

    # Commercial limits
    maxrent=_v14_n(q.get("max_monthly_rent")); prent=_v14_n(p.get("monthly_rent"))
    if maxrent and prent:
        if prent<=maxrent: score+=5; reasons.append("Rent within budget")
        else: score-=15; penalties.append("Rent above budget")
    maxpsf=_v14_n(q.get("max_rent_psf")); ppsf=_v14_n(p.get("rent_psf"))
    if maxpsf and ppsf:
        if ppsf<=maxpsf: score+=4
        else: score-=10; penalties.append("Rent/sqft above budget")
    maxsale=_v14_n(q.get("max_sale_budget")); psale=_v14_n(p.get("sale_price"))
    if maxsale and psale and psale>maxsale:
        score-=15; penalties.append("Sale price above budget")

    minfront=_v14_n(q.get("min_frontage_ft")); front=_v14_n(p.get("frontage_ft"))
    if minfront:
        if front and front>=minfront: score+=5; reasons.append("Frontage")
        elif front: score-=8; penalties.append("Frontage below requirement")
        else: penalties.append("Frontage unknown")

    if q.get("main_road_required"):
        if p.get("main_road"): score+=3
        else: score-=8; penalties.append("Main road not confirmed")
    if q.get("corner_preferred") and p.get("corner_property"):
        score+=2; reasons.append("Corner")
    if q.get("food_use_required"):
        if _v14_norm(p.get("food_use_allowed")) in {"yes","allowed","true"}: score+=4
        else: score-=10; penalties.append("Food use not confirmed")
    if q.get("exhaust_required"):
        if _v14_norm(p.get("exhaust_possible")) in {"yes","possible","true"}: score+=4
        else: score-=10; penalties.append("Exhaust not confirmed")

    return max(0,min(100,round(score))), reasons+penalties

@app.post("/api/v14/property")
async def v14_add_property(req:Request):
    need_login(req); _v14_setup()
    form=await req.form()
    code="FP-"+datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")+"-"+uuid.uuid4().hex[:4].upper()
    required=["city","location","property_type","transaction_type","available_area_sqft","contact_phone"]
    missing=[x for x in required if not _v14_s(form.get(x))]
    if missing: raise HTTPException(400,"Required: "+", ".join(missing))
    vals={k:form.get(k) for k in form.keys() if k!="pictures"}
    with engine.begin() as c:
        c.execute(text("""INSERT INTO pi_v14_properties(
          property_code,property_name,city,location,micro_market,full_address,property_type,suitable_category,
          transaction_type,available_area_sqft,floor,frontage_ft,monthly_rent,rent_psf,sale_price,cam_psf,
          security_deposit_months,possession,parking,ceiling_height_ft,power_load_kw,nearby_brands,
          main_road,corner_property,food_use_allowed,exhaust_possible,notes,contact_name,contact_phone,
          contact_role,created_by
        ) VALUES(:code,:property_name,:city,:location,:micro_market,:full_address,:property_type,:suitable_category,
          :transaction_type,:area,:floor,:frontage,:monthly_rent,:rent_psf,:sale_price,:cam,:deposit,:possession,
          :parking,:ceiling,:power,:nearby,:mainroad,:corner,:food,:exhaust,:notes,:contact_name,:contact_phone,
          :contact_role,:by)"""),{
          "code":code,"property_name":form.get("property_name"),"city":form.get("city"),"location":form.get("location"),
          "micro_market":form.get("micro_market"),"full_address":form.get("full_address"),"property_type":form.get("property_type"),
          "suitable_category":form.get("suitable_category"),"transaction_type":form.get("transaction_type"),
          "area":_v14_n(form.get("available_area_sqft")),"floor":form.get("floor"),"frontage":_v14_n(form.get("frontage_ft")),
          "monthly_rent":_v14_n(form.get("monthly_rent")),"rent_psf":_v14_n(form.get("rent_psf")),
          "sale_price":_v14_n(form.get("sale_price")),"cam":_v14_n(form.get("cam_psf")),
          "deposit":_v14_n(form.get("security_deposit_months")),"possession":form.get("possession"),
          "parking":form.get("parking"),"ceiling":_v14_n(form.get("ceiling_height_ft")),
          "power":_v14_n(form.get("power_load_kw")),"nearby":form.get("nearby_brands"),
          "mainroad":_v14_b(form.get("main_road")),"corner":_v14_b(form.get("corner_property")),
          "food":form.get("food_use_allowed"),"exhaust":form.get("exhaust_possible"),"notes":form.get("notes"),
          "contact_name":form.get("contact_name"),"contact_phone":form.get("contact_phone"),
          "contact_role":form.get("contact_role") or "UNVERIFIED","by":_v14_actor(req)
        })
    pics=form.getlist("pictures")
    for pic in pics[:8]:
        try:
            data=await pic.read()
            if data and len(data)<=8*1024*1024:
                with engine.begin() as c:
                    c.execute(text("""INSERT INTO pi_v14_property_media(property_code,filename,content_type,content)
                    VALUES(:p,:f,:ct,:b)"""),{"p":code,"f":pic.filename,"ct":pic.content_type or "image/jpeg","b":data})
        except: pass
    return {"status":"ok","property_code":code,"message":"Saved UNVERIFIED. Verify before matching."}

@app.post("/api/v14/property/{code}/verify")
async def v14_verify_property(code:str,req:Request):
    need_login(req); _v14_setup()
    body=await req.json()
    role=_v14_s(body.get("contact_role") or "OTHER").upper()
    if role not in {"OWNER","BROKER","BOTH","OTHER"}: role="OTHER"
    with engine.begin() as c:
        r=c.execute(text("""UPDATE pi_v14_properties SET availability_status='VERIFIED_ACTIVE',
          matcher_eligible=TRUE,contact_verified=TRUE,contact_role=:role,verified_by=:by,
          verified_at=NOW(),updated_at=NOW() WHERE property_code=:code"""),
          {"role":role,"by":_v14_actor(req),"code":code})
        if r.rowcount==0: raise HTTPException(404,"Property not found")
    return {"status":"ok"}

@app.post("/api/v14/requirement")
async def v14_add_requirement(req:Request):
    need_login(req); _v14_setup()
    b=await req.json()
    required=["division","company_name","city","locations","property_type","transaction_type","area_min_sqft","area_max_sqft"]
    missing=[x for x in required if not _v14_s(b.get(x))]
    if missing: raise HTTPException(400,"Required: "+", ".join(missing))
    code="CR-"+datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")+"-"+uuid.uuid4().hex[:4].upper()
    with engine.begin() as c:
        c.execute(text("""INSERT INTO pi_v14_requirements(
          requirement_code,division,company_name,contact_name,contact_phone,contact_email,source_type,source_url,
          city,locations,property_type,suitable_category,transaction_type,area_min_sqft,area_max_sqft,ideal_area_sqft,
          floor_preference,ground_floor_mandatory,max_monthly_rent,max_rent_psf,max_sale_budget,min_frontage_ft,
          possession,parking_required,nearby_brands,main_road_required,corner_preferred,food_use_required,
          exhaust_required,must_have,preferred,reject_if,assigned_to,created_by
        ) VALUES(:code,:division,:company,:contact,:phone,:email,'MANUAL_CONFIRMED',:url,:city,:locations,:ptype,
          :category,:tx,:amin,:amax,:ideal,:floor,:gf,:maxrent,:maxpsf,:maxsale,:front,:poss,:parking,:brands,
          :mainroad,:corner,:food,:exhaust,:must,:pref,:reject,:assigned,:by)"""),{
          "code":code,"division":b.get("division"),"company":b.get("company_name"),"contact":b.get("contact_name"),
          "phone":b.get("contact_phone"),"email":b.get("contact_email"),"url":b.get("source_url"),"city":b.get("city"),
          "locations":b.get("locations"),"ptype":b.get("property_type"),"category":b.get("suitable_category"),
          "tx":b.get("transaction_type"),"amin":_v14_n(b.get("area_min_sqft")),"amax":_v14_n(b.get("area_max_sqft")),
          "ideal":_v14_n(b.get("ideal_area_sqft")),"floor":b.get("floor_preference"),"gf":_v14_b(b.get("ground_floor_mandatory")),
          "maxrent":_v14_n(b.get("max_monthly_rent")),"maxpsf":_v14_n(b.get("max_rent_psf")),
          "maxsale":_v14_n(b.get("max_sale_budget")),"front":_v14_n(b.get("min_frontage_ft")),
          "poss":b.get("possession"),"parking":b.get("parking_required"),"brands":b.get("nearby_brands"),
          "mainroad":_v14_b(b.get("main_road_required")),"corner":_v14_b(b.get("corner_preferred")),
          "food":_v14_b(b.get("food_use_required")),"exhaust":_v14_b(b.get("exhaust_required")),
          "must":b.get("must_have"),"pref":b.get("preferred"),"reject":b.get("reject_if"),
          "assigned":b.get("assigned_to"),"by":_v14_actor(req)
        })
    return {"status":"ok","requirement_code":code}

@app.get("/api/v14/match/{code}")
def v14_match(code:str,req:Request):
    need_login(req); _v14_setup()
    with engine.connect() as c:
        qr=c.execute(text("SELECT * FROM pi_v14_requirements WHERE requirement_code=:x"),{"x":code}).fetchone()
        if not qr: raise HTTPException(404,"Requirement not found")
        q=dict(qr._mapping)
        props=[dict(r._mapping) for r in c.execute(text("""SELECT * FROM pi_v14_properties
          WHERE matcher_eligible=TRUE AND availability_status='VERIFIED_ACTIVE'""")).fetchall()]
    out=[]
    for p in props:
        score,reasons=_v14_score(p,q)
        if score is None: continue
        p["match_score"]=score;p["match_reasons"]=reasons
        out.append(p)
    out.sort(key=lambda x:x["match_score"],reverse=True)
    return {"status":"ok","requirement":q,"matches":out[:100],"total":len(out)}

@app.get("/api/v14/properties")
def v14_properties(req:Request,status:str=Query("ALL")):
    need_login(req); _v14_setup()
    sql="SELECT * FROM pi_v14_properties"
    params={}
    if status=="ACTIVE": sql+=" WHERE matcher_eligible=TRUE"
    elif status=="UNVERIFIED": sql+=" WHERE matcher_eligible=FALSE"
    sql+=" ORDER BY created_at DESC"
    with engine.connect() as c:
        rows=[dict(r._mapping) for r in c.execute(text(sql),params).fetchall()]
    return {"status":"ok","rows":rows}

@app.get("/api/v14/requirements")
def v14_requirements(req:Request):
    need_login(req); _v14_setup()
    with engine.connect() as c:
        rows=[dict(r._mapping) for r in c.execute(text("SELECT * FROM pi_v14_requirements ORDER BY created_at DESC")).fetchall()]
    return {"status":"ok","rows":rows}

@app.get("/api/v14/media/{mid}")
def v14_media(mid:int,req:Request):
    need_login(req); _v14_setup()
    with engine.connect() as c:
        r=c.execute(text("SELECT content,content_type FROM pi_v14_property_media WHERE id=:i"),{"i":mid}).fetchone()
    if not r: raise HTTPException(404,"Image not found")
    from fastapi.responses import Response
    return Response(content=bytes(r._mapping["content"]),media_type=r._mapping["content_type"] or "image/jpeg")

@app.get("/api/v14/property/{code}/media")
def v14_media_list(code:str,req:Request):
    need_login(req); _v14_setup()
    with engine.connect() as c:
        rows=[dict(r._mapping) for r in c.execute(text(
          "SELECT id,filename,content_type FROM pi_v14_property_media WHERE property_code=:p ORDER BY id"
        ),{"p":code}).fetchall()]
    return {"status":"ok","rows":rows}

@app.get("/v14-dashboard",response_class=HTMLResponse)
def v14_dashboard(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",status_code=303)
    return HTMLResponse("""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Deal Intelligence</title>
<style>body{font-family:Arial;margin:0;background:#f5f7fb;color:#172437}header{background:#102235;color:#fff;padding:20px}.w{padding:20px;max-width:1200px;margin:auto}.g{display:grid;grid-template-columns:repeat(2,1fr);gap:14px}.c{background:white;border:1px solid #dfe6ee;border-radius:12px;padding:20px;text-decoration:none;color:#172437}.c b{font-size:19px}.c span{display:block;color:#687789;margin-top:8px}.archive{margin-top:25px;padding-top:15px;border-top:1px solid #ccd6e2}.archive a{margin-right:15px}@media(max-width:700px){.g{grid-template-columns:1fr}}</style></head>
<body><header><b>AI Deal Intelligence OS</b><br><small>Fresh verified inventory + confirmed requirements</small></header><div class=w>
<div class=g>
<a class=c href="/v14-requirement-leads"><b>1. AI Requirement Leads</b><span>Review discovered demand. Team confirms before matching.</span></a>
<a class=c href="/v14-requirement-form"><b>2. Add Confirmed Requirement</b><span>Enter complete confirmed requirement and run matching.</span></a>
<a class=c href="/v14-property-form"><b>3. Add Fresh Property</b><span>Complete inventory form with pictures and contact.</span></a>
<a class=c href="/v14-inventory"><b>4. Verify / Active Inventory</b><span>Verify fresh properties. Only VERIFIED ACTIVE properties match.</span></a>
<a class=c href="/v14-matcher"><b>5. Property Matcher</b><span>Confirmed requirements matched only against fresh verified inventory.</span></a>
<a class=c href="/contacts-directory"><b>6. Contacts</b><span>Existing contact verification directory.</span></a>
</div><div class=archive><b>Archive / Research only — excluded from V14 matching</b><p>
<a href="/magazine-import">Magazine Data</a><a href="/data-doctor">Legacy Data Doctor</a><a href="/capture-intelligence">Camera / Newspaper / PDF</a><a href="/property-database">Legacy Property Database</a></p></div>
</div></body></html>""")

@app.get("/v14-property-form",response_class=HTMLResponse)
def v14_property_form(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    return HTMLResponse("""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Add Fresh Property</title>
<style>body{font-family:Arial;background:#f5f7fb;margin:0}header{background:#102235;color:white;padding:18px}.w{max-width:1100px;margin:auto;padding:18px}.card{background:white;padding:18px;border-radius:12px}.g{display:grid;grid-template-columns:repeat(3,1fr);gap:10px}input,select,textarea{width:100%;padding:10px;box-sizing:border-box}textarea{min-height:80px}.full{grid-column:1/-1}.req:after{content:" *";color:red}button{padding:11px 18px;background:#1677ff;color:white;border:0;border-radius:8px}@media(max-width:800px){.g{grid-template-columns:1fr}.full{grid-column:auto}}</style></head>
<body><header><b>Add Fresh Property</b> · <a style="color:white" href="/v14-dashboard">← Dashboard</a></header><div class=w><div class=card>
<p>New property starts <b>UNVERIFIED</b>. It cannot match until your team verifies it.</p>
<form id=f enctype=multipart/form-data><div class=g>
<label class=req>Property Name/No.<input name=property_name></label>
<label class=req>City<input name=city required></label><label class=req>Location<input name=location required></label>
<label>Micro Market<input name=micro_market></label><label class=full>Full Address<input name=full_address></label>
<label class=req>Property Type<select name=property_type required><option></option><option>High Street Retail</option><option>Retail Store</option><option>Restaurant/F&B</option><option>Office</option><option>Showroom</option><option>Hotel</option><option>Banquet</option><option>Commercial Land</option><option>Other</option></select></label>
<label>Suitable Category<input name=suitable_category placeholder="Retail, Restaurant, Cafe..."></label>
<label class=req>Transaction<select name=transaction_type required><option>LEASE</option><option>SALE</option></select></label>
<label class=req>Available Area SqFt<input type=number name=available_area_sqft required></label>
<label>Floor<input name=floor placeholder="Ground Floor"></label><label>Frontage Ft<input type=number step=any name=frontage_ft></label>
<label>Monthly Rent ₹<input type=number step=any name=monthly_rent></label><label>Rent / SqFt ₹<input type=number step=any name=rent_psf></label>
<label>Sale Price ₹<input type=number step=any name=sale_price></label><label>CAM / SqFt<input type=number step=any name=cam_psf></label>
<label>Security Deposit Months<input type=number step=any name=security_deposit_months></label><label>Possession<input name=possession></label>
<label>Parking<input name=parking></label><label>Ceiling Height Ft<input type=number step=any name=ceiling_height_ft></label>
<label>Power Load KW<input type=number step=any name=power_load_kw></label><label>Nearby Brands<input name=nearby_brands></label>
<label>Main Road<select name=main_road><option value=false>No/Unknown</option><option value=true>Yes</option></select></label>
<label>Corner Property<select name=corner_property><option value=false>No/Unknown</option><option value=true>Yes</option></select></label>
<label>Food Use Allowed<select name=food_use_allowed><option>Unknown</option><option>Yes</option><option>No</option></select></label>
<label>Exhaust Possible<select name=exhaust_possible><option>Unknown</option><option>Yes</option><option>No</option></select></label>
<label class=req>Contact Number<input name=contact_phone required></label><label>Contact Name<input name=contact_name></label>
<label>Contact Role<select name=contact_role><option>UNVERIFIED</option><option>OWNER</option><option>BROKER</option><option>BOTH</option><option>OTHER</option></select></label>
<label class="full req">Property Pictures (up to 8)<input type=file name=pictures accept="image/*" multiple required></label>
<label class=full>Notes<textarea name=notes></textarea></label></div><p><button>Save Fresh Property</button></p></form><div id=m></div></div></div>
<script>f.onsubmit=async e=>{e.preventDefault();m.textContent='Saving...';let r=await fetch('/api/v14/property',{method:'POST',body:new FormData(f)}),d=await r.json();m.textContent=r.ok?'Saved '+d.property_code+' — now verify it in Active Inventory.':(d.detail||'Error');if(r.ok)f.reset()}</script></body></html>""")

@app.get("/v14-requirement-form",response_class=HTMLResponse)
def v14_requirement_form(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    return HTMLResponse("""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Confirmed Requirement</title>
<style>body{font-family:Arial;background:#f5f7fb;margin:0}header{background:#102235;color:white;padding:18px}.w{max-width:1100px;margin:auto;padding:18px}.card{background:white;padding:18px;border-radius:12px}.g{display:grid;grid-template-columns:repeat(3,1fr);gap:10px}input,select,textarea{width:100%;padding:10px;box-sizing:border-box}textarea{min-height:75px}.full{grid-column:1/-1}.req:after{content:" *";color:red}button{padding:11px 18px;background:#1677ff;color:white;border:0;border-radius:8px}@media(max-width:800px){.g{grid-template-columns:1fr}.full{grid-column:auto}}</style></head>
<body><header><b>Add Confirmed Requirement</b> · <a style="color:white" href="/v14-dashboard">← Dashboard</a></header><div class=w><div class=card>
<p>Use this form <b>after your team confirms the requirement</b>. AI-discovered signals should not match directly.</p><form id=f><div class=g>
<label class=req>Division<select name=division required><option>RETAIL</option><option>HOSPITALITY</option></select></label>
<label class=req>Company / Brand<input name=company_name required></label><label>Contact Person<input name=contact_name></label>
<label>Mobile<input name=contact_phone></label><label>Email<input name=contact_email></label><label>Source URL<input name=source_url></label>
<label class=req>City<input name=city required></label><label class="full req">Locations — comma separated<input name=locations required placeholder="GK1, GK2, Defence Colony"></label>
<label class=req>Property Type<input name=property_type required placeholder="High Street Retail"></label><label>Suitable Category<input name=suitable_category placeholder="Restaurant / Cafe"></label>
<label class=req>Transaction<select name=transaction_type required><option>LEASE</option><option>SALE</option></select></label>
<label class=req>Minimum Area SqFt<input type=number name=area_min_sqft required></label><label class=req>Maximum Area SqFt<input type=number name=area_max_sqft required></label>
<label>Ideal Area SqFt<input type=number name=ideal_area_sqft></label><label>Floor Preference<input name=floor_preference></label>
<label>Ground Floor Mandatory<select name=ground_floor_mandatory><option value=false>No</option><option value=true>Yes</option></select></label>
<label>Max Monthly Rent ₹<input type=number name=max_monthly_rent></label><label>Max Rent/SqFt ₹<input type=number name=max_rent_psf></label>
<label>Max Sale Budget ₹<input type=number name=max_sale_budget></label><label>Minimum Frontage Ft<input type=number name=min_frontage_ft></label>
<label>Possession<input name=possession></label><label>Parking Required<input name=parking_required></label><label>Nearby Brands<input name=nearby_brands></label>
<label>Main Road Mandatory<select name=main_road_required><option value=false>No</option><option value=true>Yes</option></select></label>
<label>Corner Preferred<select name=corner_preferred><option value=false>No</option><option value=true>Yes</option></select></label>
<label>Food Use Required<select name=food_use_required><option value=false>No</option><option value=true>Yes</option></select></label>
<label>Exhaust Required<select name=exhaust_required><option value=false>No</option><option value=true>Yes</option></select></label>
<label>Assigned Team Member<input name=assigned_to></label>
<label class=full>Must Have<textarea name=must_have></textarea></label><label class=full>Preferred<textarea name=preferred></textarea></label>
<label class=full>Reject If<textarea name=reject_if></textarea></label></div><p><button>Save Confirmed Requirement</button></p></form><div id=m></div></div></div>
<script>f.onsubmit=async e=>{e.preventDefault();let b=Object.fromEntries(new FormData(f));['ground_floor_mandatory','main_road_required','corner_preferred','food_use_required','exhaust_required'].forEach(k=>b[k]=b[k]==='true');let r=await fetch('/api/v14/requirement',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(b)}),d=await r.json();m.innerHTML=r.ok?'Saved <b>'+d.requirement_code+'</b>. <a href="/v14-matcher">Open Matcher</a>':(d.detail||'Error');if(r.ok)f.reset()}</script></body></html>""")

@app.get("/v14-inventory",response_class=HTMLResponse)
def v14_inventory_page(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    return HTMLResponse("""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Fresh Inventory</title>
<style>body{font-family:Arial;background:#f5f7fb;margin:0}header{background:#102235;color:white;padding:18px}.w{padding:18px;overflow:auto}table{width:100%;border-collapse:collapse;background:white}th,td{padding:8px;border-bottom:1px solid #ddd;text-align:left;white-space:nowrap}button{padding:7px}</style></head>
<body><header><b>Fresh Inventory Verification</b> · <a style="color:white" href="/v14-dashboard">← Dashboard</a></header><div class=w><p><button onclick="load('UNVERIFIED')">Unverified</button> <button onclick="load('ACTIVE')">Verified Active</button> <button onclick="load('ALL')">All Fresh</button></p><div id=x></div></div>
<script>async function load(s){let d=await (await fetch('/api/v14/properties?status='+s)).json(),r=d.rows||[];x.innerHTML='<table><tr><th>Status</th><th>Property</th><th>Location</th><th>Type</th><th>Area</th><th>Floor</th><th>Rent</th><th>Frontage</th><th>Contact</th><th>Role</th><th>Action</th></tr>'+r.map(p=>`<tr><td>${p.availability_status}</td><td>${p.property_code}<br>${p.property_name||''}</td><td>${p.city} / ${p.location}</td><td>${p.property_type}</td><td>${p.available_area_sqft}</td><td>${p.floor||''}</td><td>${p.monthly_rent||''}</td><td>${p.frontage_ft||''}</td><td><b>${p.contact_phone||''}</b><br>${p.contact_name||''}</td><td>${p.contact_role||''}</td><td>${p.matcher_eligible?'MATCH ACTIVE':`<select id="r_${p.property_code}"><option>OWNER</option><option>BROKER</option><option>BOTH</option><option>OTHER</option></select> <button onclick="verify('${p.property_code}')">Verify + Activate</button>`}</td></tr>`).join('')+'</table>'}async function verify(c){let role=document.getElementById('r_'+c).value,r=await fetch('/api/v14/property/'+c+'/verify',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({contact_role:role})});if(r.ok)load('UNVERIFIED');else alert('Verification failed')}load('UNVERIFIED')</script></body></html>""")

@app.get("/v14-matcher",response_class=HTMLResponse)
def v14_matcher_page(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    return HTMLResponse("""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Verified Property Matcher</title>
<style>body{font-family:Arial;background:#f5f7fb;margin:0}header{background:#102235;color:white;padding:18px}.w{padding:18px;overflow:auto}.req{background:white;padding:12px;border-radius:10px;margin-bottom:10px}table{width:100%;border-collapse:collapse;background:white}th,td{padding:8px;border-bottom:1px solid #ddd;text-align:left;vertical-align:top}.score{font-size:20px;font-weight:bold}</style></head>
<body><header><b>Confirmed Requirement → Verified Property Matcher</b> · <a style="color:white" href="/v14-dashboard">← Dashboard</a></header><div class=w><div id=reqs></div><div id=res></div></div>
<script>const E=x=>String(x??'').replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));async function load(){let d=await(await fetch('/api/v14/requirements')).json();reqs.innerHTML=(d.rows||[]).map(q=>`<div class=req><b>${E(q.company_name)} · ${E(q.requirement_code)}</b><br>${E(q.locations)} · ${E(q.area_min_sqft)}-${E(q.area_max_sqft)} sqft · ${E(q.property_type)} · ${E(q.transaction_type)} <button onclick="match('${q.requirement_code}')">Run Match</button></div>`).join('')||'No confirmed requirements.'}async function match(c){res.innerHTML='Matching fresh verified inventory only...';let d=await(await fetch('/api/v14/match/'+c)).json(),r=d.matches||[];res.innerHTML='<h3>'+r.length+' qualified matches</h3><table><tr><th>Match</th><th>Property</th><th>Location</th><th>Area/Floor</th><th>Commercial</th><th>Contact to Verify</th><th>Why</th></tr>'+r.map(p=>`<tr><td class=score>${p.match_score}%</td><td><b>${E(p.property_code)}</b><br>${E(p.property_name)}</td><td>${E(p.city)}<br>${E(p.location)}<br>${E(p.full_address)}</td><td>${E(p.available_area_sqft)} sqft<br>${E(p.floor)}<br>Frontage ${E(p.frontage_ft)} ft</td><td>Rent ₹${E(p.monthly_rent)}<br>₹/sqft ${E(p.rent_psf)}<br>${E(p.transaction_type)}</td><td><b>${E(p.contact_phone)}</b><br>${E(p.contact_name)}<br>${E(p.contact_role)}<br>Verified: ${p.contact_verified?'YES':'NO'}</td><td>${(p.match_reasons||[]).map(E).join('<br>')}</td></tr>`).join('')+'</table>'}load()</script></body></html>""")

@app.get("/v14-requirement-leads",response_class=HTMLResponse)
def v14_requirement_leads(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    return HTMLResponse("""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>AI Requirement Leads</title>
<style>body{font-family:Arial;background:#f5f7fb;margin:0}header{background:#102235;color:white;padding:18px}.w{padding:18px}.note{background:#fff3cd;padding:12px;border-radius:8px}</style></head>
<body><header><b>AI Requirement Leads</b> · <a style="color:white" href="/v14-dashboard">← Dashboard</a></header><div class=w><div class=note><b>Rule:</b> AI/public-web signals are leads only. Contact and confirm the requirement first. Then enter the confirmed details in the V14 Confirmed Requirement form. AI signals do not run directly against property inventory.</div><p><a href="/requirements-workbench?division=RETAIL">Open Retail AI Signals</a></p><p><a href="/requirements-workbench?division=HOSPITALITY">Open Hospitality AI Signals</a></p><p><a href="/v14-requirement-form"><b>Requirement confirmed? → Enter Confirmed Requirement</b></a></p></div></body></html>""")

# ============================================================
# V15 DYNAMIC DEAL INTELLIGENCE DASHBOARD
# Clean team interface. No version numbers shown to users.
# ============================================================

def _v15_safe_count(sql, params=None):
    try:
        with engine.connect() as c:
            return int(c.execute(text(sql), params or {}).scalar_one() or 0)
    except Exception:
        return 0

@app.get("/api/v15/dashboard/summary")
def v15_dashboard_summary(req:Request):
    need_login(req)
    try:_v14_setup()
    except Exception:pass

    return {
        "status":"ok",

        "delhi_verified_properties":_v15_safe_count(
            "SELECT COUNT(*) FROM pi_v14_properties WHERE matcher_eligible=TRUE AND availability_status='VERIFIED_ACTIVE'"
        ),
        "delhi_unverified_properties":_v15_safe_count(
            "SELECT COUNT(*) FROM pi_v14_properties WHERE matcher_eligible=FALSE"
        ),
        "delhi_confirmed_requirements":_v15_safe_count(
            "SELECT COUNT(*) FROM pi_v14_requirements WHERE confirmation_status='CONFIRMED'"
        ),

        "retail_ai_signals":_v15_safe_count(
            """SELECT COUNT(*) FROM ai_demand_signals
               WHERE LOWER(COALESCE(source_type,'')) LIKE '%retail%'
                  OR LOWER(COALESCE(source_name,'')) LIKE '%retail%'"""
        ),
        "hospitality_ai_signals":_v15_safe_count(
            """SELECT COUNT(*) FROM ai_demand_signals
               WHERE LOWER(COALESCE(source_type,'')) LIKE '%hospital%'
                  OR LOWER(COALESCE(source_name,'')) LIKE '%hospital%'"""
        ),

        "contacts":_v15_safe_count(
            "SELECT COUNT(*) FROM pi_property_contact_links"
        ),

        "goa_properties":_v15_safe_count(
            "SELECT COUNT(*) FROM pi_goa_properties"
        ),
        "goa_requirements":_v15_safe_count(
            "SELECT COUNT(*) FROM pi_goa_requirements"
        ),

        "archived_magazine":_v15_safe_count(
            "SELECT COUNT(*) FROM pi_magazine_master"
        )
    }

@app.get("/v15-dashboard",response_class=HTMLResponse)
def v15_dashboard(req:Request):
    role=page_role_or_redirect(req)
    if not role:
        return RedirectResponse("/login",status_code=303)

    return HTMLResponse("""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>AI Deal Intelligence OS</title>
<style>
:root{
  --nav:#102235;--bg:#f4f7fb;--card:#fff;--line:#e1e8f0;
  --text:#172437;--muted:#68798c;--blue:#1677ff;--green:#138a63;
  --orange:#d98200;--red:#b5473d;--purple:#6d5bd0
}
*{box-sizing:border-box}
body{margin:0;background:var(--bg);font-family:Arial,sans-serif;color:var(--text)}
header{background:var(--nav);color:#fff;padding:18px 24px;display:flex;justify-content:space-between;align-items:center;gap:16px;flex-wrap:wrap}
.brand b{font-size:21px}.brand small{display:block;color:#cbd7e3;margin-top:4px}
.toplinks{display:flex;gap:8px;align-items:center}.toplinks a{color:#fff;text-decoration:none;font-size:13px}
.wrap{max-width:1500px;margin:auto;padding:22px}
.hero{background:linear-gradient(135deg,#102235,#1f4365);color:#fff;border-radius:16px;padding:24px;margin-bottom:18px}
.hero h1{margin:0 0 6px;font-size:28px}.hero p{margin:0;color:#dce8f4}
.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:10px;margin-top:18px}
.kpi{background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.15);padding:13px;border-radius:12px}
.kpi b{display:block;font-size:24px}.kpi span{font-size:11px;color:#d9e5ef}
.section{margin-top:24px}.section h2{font-size:14px;letter-spacing:.08em;color:#526579;margin:0 0 10px}
.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(245px,1fr));gap:12px}
.card{display:block;background:var(--card);border:1px solid var(--line);border-radius:14px;padding:17px;text-decoration:none;color:var(--text);min-height:120px;position:relative;transition:.16s}
.card:hover{transform:translateY(-2px);box-shadow:0 8px 24px rgba(16,34,53,.08);border-color:#b8cae0}
.card b{display:block;font-size:17px;margin-bottom:7px}.card span{display:block;font-size:13px;line-height:1.45;color:var(--muted);padding-right:20px}
.icon{font-size:25px;margin-bottom:10px}.tag{display:inline-block;margin-top:10px;padding:4px 8px;border-radius:11px;background:#edf4ff;color:#285d92;font-size:10px;font-weight:bold}
.green .tag{background:#e8f7f1;color:#176248}.orange .tag{background:#fff4df;color:#815716}.purple .tag{background:#f1edff;color:#5744ad}
.alert{margin-top:20px;background:#fff8e8;border:1px solid #efd08d;padding:12px;border-radius:10px;color:#6b572e;font-size:13px}
.archive{margin-top:25px;background:#fff;border:1px dashed #b9c7d7;border-radius:12px;padding:14px}
.archive summary{cursor:pointer;font-weight:bold}.archive-links{display:flex;gap:10px;flex-wrap:wrap;margin-top:12px}.archive-links a{background:#eef2f6;color:#35485b;text-decoration:none;padding:8px 10px;border-radius:8px;font-size:12px}
@media(max-width:650px){.wrap{padding:14px}.hero{padding:18px}.hero h1{font-size:23px}}
</style>
</head>
<body>
<header>
  <div class="brand"><b>AI Deal Intelligence OS</b><small>Property · Requirements · AI Leads · Contacts · Goa</small></div>
  <div class="toplinks"><span id="role">TEAM</span><a href="/logout">Logout</a></div>
</header>

<div class="wrap">
  <div class="hero">
    <h1>Team Command Center</h1>
    <p>Fresh verified inventory, confirmed requirements and controlled AI lead discovery.</p>
    <div class="kpis">
      <div class="kpi"><b id="pv">0</b><span>VERIFIED DELHI PROPERTIES</span></div>
      <div class="kpi"><b id="pu">0</b><span>UNVERIFIED PROPERTIES</span></div>
      <div class="kpi"><b id="rq">0</b><span>CONFIRMED REQUIREMENTS</span></div>
      <div class="kpi"><b id="rt">0</b><span>RETAIL AI SIGNALS</span></div>
      <div class="kpi"><b id="hs">0</b><span>HOSPITALITY AI SIGNALS</span></div>
      <div class="kpi"><b id="ct">0</b><span>CONTACT RELATIONSHIPS</span></div>
    </div>
  </div>

  <div class="section">
    <h2>DELHI NCR · DAILY OPERATIONS</h2>
    <div class="grid">
      <a class="card green" href="/v14-property-form"><div class="icon">＋</div><b>Add Property Manually</b><span>Complete standardized property form with pictures and contact details. New entries start unverified.</span><span class="tag">Fresh Inventory</span></a>
      <a class="card green" href="/v14-requirement-form"><div class="icon">◎</div><b>Add Requirement Manually</b><span>Enter a confirmed client requirement using the same fields used by the matcher.</span><span class="tag">Confirmed Demand</span></a>
      <a class="card" href="/v14-inventory"><div class="icon">✓</div><b>Verify Properties</b><span>Review unverified inventory and activate only properties confirmed by your team.</span><span class="tag">Verified / Unverified</span></a>
      <a class="card" href="/v14-matcher"><div class="icon">◆</div><b>Property Matcher</b><span>Match confirmed requirements only against fresh VERIFIED ACTIVE inventory.</span><span class="tag">Match Engine</span></a>
    </div>
  </div>

  <div class="section">
    <h2>DATABASE & CAPTURE</h2>
    <div class="grid">
      <a class="card" href="/v14-inventory"><div class="icon">▦</div><b>Fresh Property Database</b><span>Search and review active and unverified fresh inventory separately from legacy data.</span><span class="tag">Searchable Inventory</span></a>
      <a class="card orange" href="/capture-intelligence"><div class="icon">◉</div><b>Capture / Import Property</b><span>Camera, screenshot, handwritten note, WhatsApp screenshot, newspaper, magazine or PDF.</span><span class="tag">Staging Only</span></a>
      <a class="card" href="/contacts-directory"><div class="icon">☎</div><b>Property Contacts</b><span>Verify contacts and classify Owner, Broker, Both or Other.</span><span class="tag">Verification</span></a>
    </div>
  </div>

  <div class="section">
    <h2>AI REQUIREMENT DISCOVERY</h2>
    <div class="grid">
      <a class="card purple" href="/requirements-workbench?division=RETAIL"><div class="icon">◈</div><b>AI Retail</b><span>Review retailer expansion signals. Team confirms first, then converts the genuine requirement into the manual requirement form.</span><span class="tag">AI Lead Discovery</span></a>
      <a class="card purple" href="/ai-hospitality-master-final"><div class="icon">◆</div><b>AI Hospitality</b><span>Restaurants, cafes, lounges, clubs, hotels, guest houses and banquet requirement discovery.</span><span class="tag">AI Lead Discovery</span></a>
      <a class="card" href="/v14-requirement-leads"><div class="icon">✓</div><b>Requirement Verification</b><span>Human-review rule: AI signals do not match directly. Confirm first, then enter as a manual requirement.</span><span class="tag">Human Confirmation</span></a>
    </div>
  </div>

  <div class="section">
    <h2>MARKETING & CONTACTS</h2>
    <div class="grid">
      <a class="card" href="/marketing-contacts-final"><div class="icon">✉</div><b>Marketing Contacts</b><span>Property contacts, AI-generated hospitality/retail contacts and verified database contacts for approved outreach.</span><span class="tag">WhatsApp Ready</span></a>
      <a class="card" href="/legacy-workspace#contacts"><div class="icon">⇧</div><b>Upload Contact List</b><span>Add external contact databases for marketing workflows without mixing them into property inventory.</span><span class="tag">Contact Import</span></a>
      <a class="card" href="/legacy-workspace#bots"><div class="icon">⚡</div><b>Bot Control Room</b><span>Run and review discovery bots and system activity.</span><span class="tag">Automation</span></a>
    </div>
  </div>

  <div class="section">
    <h2>GOA PROPERTY</h2>
    <div class="grid">
      <a class="card green" href="/goa-property-form"><div class="icon">＋</div><b>Add Goa Property</b><span>Fresh Goa property inventory with photos, commercial details and verification workflow.</span><span class="tag">Goa Inventory</span></a>
      <a class="card green" href="/goa-requirement-form"><div class="icon">◎</div><b>Add Goa Requirement</b><span>Buyer/investor requirement entry with Goa-specific fields.</span><span class="tag">Goa Demand</span></a>
      <a class="card" href="/goa-matcher"><div class="icon">◆</div><b>Goa Matcher</b><span>Match verified Goa requirements against verified Goa inventory only.</span><span class="tag">Goa Matching</span></a>
      <a class="card" href="/goa-database"><div class="icon">▦</div><b>Goa Database</b><span>Search, verify and manage Goa inventory separately from Delhi NCR.</span><span class="tag">Separate Database</span></a>
    </div>
  </div>

  <div class="alert"><b>Operating rule:</b> AI-discovered requirements are leads only. Human verification is required before they enter the manual requirement list and property matcher.</div>

  <details class="archive">
    <summary>Archive / Technical Data Sources</summary>
    <div class="archive-links">
      <a href="/property-database">Legacy Property Database</a>
      <a href="/magazine-master-import">Magazine Archive</a>
      <a href="/data-doctor">Data Doctor</a>
      <a href="/data-command-center">Admin Data Tools</a>
      <a href="/legacy-workspace">Legacy Workspace</a>
    </div>
  </details>
</div>

<script>
async function load(){
  try{
    const d=await (await fetch('/api/v15/dashboard/summary')).json();
    pv.textContent=(d.delhi_verified_properties||0).toLocaleString();
    pu.textContent=(d.delhi_unverified_properties||0).toLocaleString();
    rq.textContent=(d.delhi_confirmed_requirements||0).toLocaleString();
    rt.textContent=(d.retail_ai_signals||0).toLocaleString();
    hs.textContent=(d.hospitality_ai_signals||0).toLocaleString();
    ct.textContent=(d.contacts||0).toLocaleString();
  }catch(e){}
}
load();
</script>
</body></html>""")

# Clean daily entry point.
@app.middleware("http")
async def v15_workspace_router(request,call_next):
    if request.url.path=="/workspace":
        return RedirectResponse("/v15-dashboard",status_code=307)
    return await call_next(request)

# Friendly placeholder routes for Goa until the dedicated Goa module is installed.
def _v15_goa_placeholder(title,desc):
    return HTMLResponse(f"""<!doctype html><html><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>{title}</title></head>
    <body style='font-family:Arial;background:#f4f7fb;margin:0;color:#172437'>
    <div style='background:#102235;color:white;padding:18px'><b>{title}</b></div>
    <div style='max-width:900px;margin:auto;padding:20px'><p><a href='/v15-dashboard'>← Dashboard</a></p>
    <div style='background:white;border:1px solid #e1e8f0;border-radius:12px;padding:18px'>
    <b>Goa module</b><p>{desc}</p><p>This page is reserved so the dashboard navigation remains stable. The Goa inventory/matcher can be installed as a separate database module without affecting Delhi NCR.</p>
    </div></div></body></html>""")

@app.get("/goa-property-form",response_class=HTMLResponse)
def v15_goa_property_form(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    return _v15_goa_placeholder("Add Goa Property","Fresh Goa property entry with pictures and verification.")

@app.get("/goa-requirement-form",response_class=HTMLResponse)
def v15_goa_requirement_form(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    return _v15_goa_placeholder("Add Goa Requirement","Buyer/investor requirement entry for Goa properties.")

@app.get("/goa-matcher",response_class=HTMLResponse)
def v15_goa_matcher(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    return _v15_goa_placeholder("Goa Matcher","Goa-specific verified inventory and requirement matching.")

@app.get("/goa-database",response_class=HTMLResponse)
def v15_goa_database(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    return _v15_goa_placeholder("Goa Database","Separate searchable Goa property database.")

# ============================================================
# V15.1 MARKETING CONTACTS + DASHBOARD CACHE/NAVIGATION FIX
# ============================================================

def _v151_setup():
    with engine.begin() as c:
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_marketing_contacts(
            id BIGSERIAL PRIMARY KEY,
            contact_key TEXT UNIQUE NOT NULL,
            contact_name TEXT,
            primary_phone TEXT NOT NULL,
            all_phones JSONB DEFAULT '[]'::jsonb,
            company_brand TEXT,
            category TEXT DEFAULT 'OTHER',
            subcategory TEXT,
            city TEXT,
            location TEXT,
            email TEXT,
            website TEXT,
            source TEXT,
            source_detail TEXT,
            verified_status TEXT DEFAULT 'UNVERIFIED',
            whatsapp_status TEXT DEFAULT 'NOT_CONTACTED',
            opt_out BOOLEAN DEFAULT FALSE,
            linked_property_count INTEGER DEFAULT 0,
            linked_properties JSONB DEFAULT '[]'::jsonb,
            notes TEXT,
            date_added TIMESTAMPTZ DEFAULT NOW(),
            last_contacted_at TIMESTAMPTZ,
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )"""))
        c.execute(text("CREATE INDEX IF NOT EXISTS idx_marketing_contacts_phone ON pi_marketing_contacts(primary_phone)"))
        c.execute(text("CREATE INDEX IF NOT EXISTS idx_marketing_contacts_category ON pi_marketing_contacts(category)"))
        c.execute(text("CREATE INDEX IF NOT EXISTS idx_marketing_contacts_source ON pi_marketing_contacts(source)"))
        c.execute(text("CREATE INDEX IF NOT EXISTS idx_marketing_contacts_verified ON pi_marketing_contacts(verified_status)"))
        c.execute(text("CREATE INDEX IF NOT EXISTS idx_marketing_contacts_whatsapp ON pi_marketing_contacts(whatsapp_status)"))

def _v151_digits(v):
    d=_re.sub(r"\D","",str(v or ""))
    if d.startswith("91") and len(d)==12:
        d=d[2:]
    if len(d)==10:
        return d
    return ""

def _v151_category_from_text(*parts):
    txt=" ".join(str(x or "") for x in parts).lower()
    rules=[
        ("CAFE",["cafe","coffee","bakery"]),
        ("RESTAURANT",["restaurant","restro","diner"]),
        ("BANQUET",["banquet","wedding venue"]),
        ("HOTEL",["hotel","resort"]),
        ("GUEST_HOUSE",["guest house","guesthouse"]),
        ("LOUNGE",["lounge"]),
        ("CLUB",["club"]),
        ("BAR",["bar","pub"]),
        ("FARMHOUSE",["farmhouse","farm house"]),
        ("RETAILER",["retail","store","brand"]),
        ("BROKER",["broker","property dealer"]),
        ("OWNER",["owner","landlord","builder"])
    ]
    for cat,terms in rules:
        if any(t in txt for t in terms):
            return cat
    return "OTHER"

def _v151_upsert_contact(phone, name=None, company=None, category=None, city=None, location=None,
                          email=None, website=None, source=None, source_detail=None,
                          property_ids=None, notes=None):
    _v151_setup()
    ph=_v151_digits(phone)
    if not ph:
        return False
    key="91"+ph
    props=list(dict.fromkeys([str(x) for x in (property_ids or []) if x]))
    cat=(category or "OTHER").upper()
    with engine.begin() as c:
        old=c.execute(text("SELECT * FROM pi_marketing_contacts WHERE contact_key=:k"),{"k":key}).fetchone()
        if old:
            o=dict(old._mapping)
            old_props=o.get("linked_properties") or []
            merged=list(dict.fromkeys([str(x) for x in old_props]+props))
            old_sources=[x.strip() for x in str(o.get("source") or "").split(",") if x.strip()]
            if source and source not in old_sources:
                old_sources.append(source)
            c.execute(text("""UPDATE pi_marketing_contacts SET
                contact_name=COALESCE(NULLIF(:name,''),contact_name),
                company_brand=COALESCE(NULLIF(:company,''),company_brand),
                category=CASE WHEN category='OTHER' AND :cat<>'OTHER' THEN :cat ELSE category END,
                city=COALESCE(NULLIF(:city,''),city),
                location=COALESCE(NULLIF(:location,''),location),
                email=COALESCE(NULLIF(:email,''),email),
                website=COALESCE(NULLIF(:website,''),website),
                source=:source,
                source_detail=COALESCE(NULLIF(:detail,''),source_detail),
                linked_property_count=:cnt,
                linked_properties=CAST(:props AS jsonb),
                notes=COALESCE(NULLIF(:notes,''),notes),
                updated_at=NOW()
                WHERE contact_key=:k"""),{
                    "name":name or "","company":company or "","cat":cat,"city":city or "",
                    "location":location or "","email":email or "","website":website or "",
                    "source":", ".join(old_sources),"detail":source_detail or "",
                    "cnt":len(merged),"props":json.dumps(merged),"notes":notes or "","k":key
                })
        else:
            c.execute(text("""INSERT INTO pi_marketing_contacts(
                contact_key,contact_name,primary_phone,all_phones,company_brand,category,city,location,
                email,website,source,source_detail,linked_property_count,linked_properties,notes
            ) VALUES(:k,:name,:ph,CAST(:phones AS jsonb),:company,:cat,:city,:location,:email,:website,
                :source,:detail,:cnt,CAST(:props AS jsonb),:notes)"""),{
                    "k":key,"name":name,"ph":ph,"phones":json.dumps([ph]),
                    "company":company,"cat":cat,"city":city,"location":location,"email":email,
                    "website":website,"source":source,"detail":source_detail,"cnt":len(props),
                    "props":json.dumps(props),"notes":notes
                })
    return True

def _v151_sync_property_contacts():
    """
    Property contacts remain owner/broker verification data.
    This copies a marketing-safe representation into the Marketing Contacts DB,
    while preserving source='PROPERTY_DATABASE' or 'MAGAZINE'.
    """
    _v151_setup()
    added=0

    # Prefer full relationship table if present.
    if _v138_table_exists("pi_property_contact_links"):
        with engine.connect() as c:
            rows=c.execute(text("""SELECT pcl.property_id,pcl.normalized_contact,pcl.role_hint,
                p.property_name,p.city,p.location,p.source,p.owner_name,p.broker_name
                FROM pi_property_contact_links pcl
                LEFT JOIN pi_properties p ON p.property_id=pcl.property_id
                ORDER BY pcl.property_id""")).fetchall()
        grouped={}
        for r in rows:
            d=dict(r._mapping)
            ph=_v151_digits(d.get("normalized_contact"))
            if not ph: continue
            g=grouped.setdefault(ph,{"props":[],"names":[],"cities":[],"locations":[],"sources":[],"roles":[]})
            if d.get("property_id"):g["props"].append(d["property_id"])
            for k,outk in [("owner_name","names"),("broker_name","names"),("city","cities"),("location","locations"),("source","sources"),("role_hint","roles")]:
                if d.get(k):g[outk].append(str(d[k]))
        for ph,g in grouped.items():
            source_blob=" ".join(g["sources"]).lower()
            source="MAGAZINE" if "magazine" in source_blob else "PROPERTY_DATABASE"
            role_blob=" ".join(g["roles"])
            cat=_v151_category_from_text(role_blob," ".join(g["names"]))
            name=(g["names"][0] if g["names"] else None)
            city=(g["cities"][0] if g["cities"] else None)
            loc=", ".join(list(dict.fromkeys(g["locations"]))[:5])
            if _v151_upsert_contact(ph,name=name,category=cat,city=city,location=loc,source=source,
                                    source_detail="Property contact sync",property_ids=list(dict.fromkeys(g["props"]))):
                added+=1

    # AI demand/contact sources.
    if _v138_table_exists("ai_demand_signals"):
        with engine.connect() as c:
            ai=[dict(r._mapping) for r in c.execute(text("SELECT * FROM ai_demand_signals ORDER BY created_at DESC LIMIT 5000")).fetchall()]
        for r in ai:
            ph=_v151_digits(r.get("contact_phone") or r.get("phone") or r.get("mobile"))
            if not ph: continue
            blob=" ".join(str(v or "") for v in r.values())
            cat=_v151_category_from_text(blob)
            src="AI_HOSPITALITY" if cat in {"CAFE","RESTAURANT","BANQUET","HOTEL","GUEST_HOUSE","LOUNGE","CLUB","BAR","FARMHOUSE"} else "AI_RETAIL"
            _v151_upsert_contact(
                ph,
                name=r.get("contact_name") or r.get("person_name"),
                company=r.get("company_name") or r.get("brand_name"),
                category=cat,
                city=r.get("city"),
                location=r.get("location"),
                email=r.get("contact_email") or r.get("email"),
                website=r.get("website"),
                source=src,
                source_detail=r.get("source_url") or r.get("linkedin_post_url"),
                notes=r.get("excerpt") or r.get("requirement_text")
            )
            added+=1
    return added

@app.post("/api/v15-1/marketing-contacts/sync")
def v151_sync_contacts(req:Request):
    need_login(req)
    try:
        n=_v151_sync_property_contacts()
        return {"status":"ok","processed":n}
    except Exception as ex:
        raise HTTPException(500,f"{type(ex).__name__}: {ex}")

@app.get("/api/v15-1/marketing-contacts")
def v151_contacts(req:Request,category:str=Query("ALL"),source:str=Query("ALL"),
                  verified:str=Query("ALL"),whatsapp:str=Query("ALL"),q:str=Query("")):
    need_login(req);_v151_setup()
    wh=[];params={}
    if category!="ALL":
        wh.append("category=:cat");params["cat"]=category
    if source!="ALL":
        wh.append("source ILIKE :src");params["src"]="%"+source+"%"
    if verified!="ALL":
        wh.append("verified_status=:ver");params["ver"]=verified
    if whatsapp!="ALL":
        wh.append("whatsapp_status=:wa");params["wa"]=whatsapp
    if q.strip():
        wh.append("""(
            COALESCE(contact_name,'') ILIKE :q OR COALESCE(primary_phone,'') ILIKE :q OR
            COALESCE(company_brand,'') ILIKE :q OR COALESCE(location,'') ILIKE :q OR
            COALESCE(city,'') ILIKE :q OR COALESCE(email,'') ILIKE :q
        )""");params["q"]="%"+q.strip()+"%"
    sql="SELECT * FROM pi_marketing_contacts"
    if wh:sql+=" WHERE "+" AND ".join(wh)
    sql+=" ORDER BY date_added DESC LIMIT 5000"
    with engine.connect() as c:
        rows=[dict(r._mapping) for r in c.execute(text(sql),params).fetchall()]
    return {"status":"ok","rows":rows}

@app.post("/api/v15-1/marketing-contacts/{cid}/update")
async def v151_update_contact(cid:int,req:Request):
    need_login(req);_v151_setup()
    b=await req.json()
    allowed={"contact_name","company_brand","category","subcategory","city","location","email","website",
             "verified_status","whatsapp_status","opt_out","notes"}
    vals={k:v for k,v in b.items() if k in allowed}
    if not vals:raise HTTPException(400,"No fields to update")
    sets=[];params={"id":cid}
    for i,(k,v) in enumerate(vals.items()):
        key="v"+str(i);sets.append(f"{k}=:{key}");params[key]=v
    with engine.begin() as c:
        c.execute(text("UPDATE pi_marketing_contacts SET "+",".join(sets)+",updated_at=NOW() WHERE id=:id"),params)
    return {"status":"ok"}

@app.get("/marketing-contacts",response_class=HTMLResponse)
def v151_marketing_contacts_page(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    return HTMLResponse("""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Marketing Contacts</title>
<style>*{box-sizing:border-box}body{font-family:Arial;margin:0;background:#f4f7fb;color:#172437}header{background:#102235;color:white;padding:18px}.w{padding:18px}.bar{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px}.bar a,.bar button{padding:8px 10px;border-radius:7px;border:0;background:#e9eef5;color:#203247;text-decoration:none;font-weight:bold}.bar button.primary{background:#1677ff;color:white}select,input{padding:8px;border:1px solid #ccd6e2;border-radius:7px}.stats{margin-left:auto}.tablewrap{overflow:auto;background:white;border-radius:10px}table{width:100%;border-collapse:collapse;font-size:12px}th,td{padding:8px;border-bottom:1px solid #edf1f5;text-align:left;vertical-align:top;white-space:nowrap}th{position:sticky;top:0;background:#f8fafc}.small{font-size:11px;color:#687789}.pill{padding:3px 6px;border-radius:8px;background:#edf4ff}.ready{background:#dcfce7}</style></head>
<body><header><b>Marketing Contacts Database</b><br><small>Cafe · Restaurant · Banquet · Hotel · Retail · Property · Magazine · AI</small></header><div class=w>
<div class=bar><a href="/workspace">← Dashboard</a><button class=primary onclick="sync()">Sync Sources</button>
<select id=category><option>ALL</option><option>CAFE</option><option>RESTAURANT</option><option>BANQUET</option><option>HOTEL</option><option>GUEST_HOUSE</option><option>LOUNGE</option><option>CLUB</option><option>BAR</option><option>FARMHOUSE</option><option>RETAILER</option><option>BROKER</option><option>OWNER</option><option>OTHER</option></select>
<select id=source><option>ALL</option><option>AI_HOSPITALITY</option><option>AI_RETAIL</option><option>PROPERTY_DATABASE</option><option>MAGAZINE</option></select>
<select id=verified><option>ALL</option><option>UNVERIFIED</option><option>VERIFIED</option></select>
<select id=whatsapp><option>ALL</option><option>NOT_CONTACTED</option><option>READY</option><option>SENT</option><option>REPLIED</option><option>OPT_OUT</option></select>
<input id=q placeholder="Search name, phone, brand, location"><span class=stats id=stats></span></div>
<div class=tablewrap><table><thead><tr><th>Select</th><th>Name</th><th>Phone</th><th>Brand</th><th>Category</th><th>Location</th><th>Source</th><th>Verified</th><th>WhatsApp</th><th>Properties</th><th>Date Added</th></tr></thead><tbody id=rows></tbody></table></div>
</div>
<script>
const E=x=>String(x??'').replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));let D=[];
async function A(u,o={}){let r=await fetch(u,o),d=await r.json();if(!r.ok)throw Error(d.detail||'Error');return d}
async function load(){let u='/api/v15-1/marketing-contacts?category='+category.value+'&source='+source.value+'&verified='+verified.value+'&whatsapp='+whatsapp.value+'&q='+encodeURIComponent(q.value);let d=await A(u);D=d.rows||[];stats.textContent=D.length+' contacts';rows.innerHTML=D.map(x=>`<tr><td><input type=checkbox ${x.opt_out?'disabled':''}></td><td>${E(x.contact_name||'')}</td><td><b>${E(x.primary_phone)}</b><br><span class=small>${E((x.all_phones||[]).join(', '))}</span></td><td>${E(x.company_brand||'')}</td><td><span class="pill">${E(x.category)}</span></td><td>${E(x.city||'')}<br>${E(x.location||'')}</td><td>${E(x.source||'')}</td><td><select onchange="upd(${x.id},'verified_status',this.value)"><option ${x.verified_status==='UNVERIFIED'?'selected':''}>UNVERIFIED</option><option ${x.verified_status==='VERIFIED'?'selected':''}>VERIFIED</option></select></td><td><select onchange="upd(${x.id},'whatsapp_status',this.value)"><option ${x.whatsapp_status==='NOT_CONTACTED'?'selected':''}>NOT_CONTACTED</option><option ${x.whatsapp_status==='READY'?'selected':''}>READY</option><option ${x.whatsapp_status==='SENT'?'selected':''}>SENT</option><option ${x.whatsapp_status==='REPLIED'?'selected':''}>REPLIED</option><option ${x.whatsapp_status==='OPT_OUT'?'selected':''}>OPT_OUT</option></select></td><td>${E(x.linked_property_count||0)}</td><td>${E((x.date_added||'').slice(0,10))}</td></tr>`).join('')||'<tr><td colspan=11>No contacts found.</td></tr>'}
async function upd(id,k,v){let b={};b[k]=v;await A('/api/v15-1/marketing-contacts/'+id+'/update',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(b)});}
async function sync(){stats.textContent='Syncing...';try{let d=await A('/api/v15-1/marketing-contacts/sync',{method:'POST'});stats.textContent='Synced '+d.processed;load()}catch(e){alert(e.message)}}
[category,source,verified,whatsapp].forEach(x=>x.onchange=load);q.oninput=load;load();
</script></body></html>""")

# Update dashboard Marketing Contacts link to dedicated page.
# Runtime middleware below also ensures no stale team dashboard is cached.

@app.middleware("http")
async def v151_no_cache_and_dashboard_router(request,call_next):
    old_dashboard_paths={
        "/team-workspace-clean","/simple-dashboard","/v14-dashboard","/data-command-center"
    }
    # Only redirect former team dashboards. Admin data-command-center remains available
    # when explicitly requested with ?admin=1.
    if request.url.path in {"/team-workspace-clean","/simple-dashboard","/v14-dashboard"}:
        return RedirectResponse("/v15-dashboard",status_code=307)

    response=await call_next(request)

    if (
        request.url.path.startswith("/v15")
        or request.url.path.startswith("/v14")
        or request.url.path in {"/workspace","/marketing-contacts","/contacts-directory",
                                "/requirements-workbench","/capture-intelligence"}
    ):
        response.headers["Cache-Control"]="no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"]="no-cache"
        response.headers["Expires"]="0"

    return response

# ============================================================
# V15.2 AI HOSPITALITY CONTACT SYNC FIX
# Robust schema introspection for ai_demand_signals / hospitality tables.
# ============================================================

def _v152_cols(table_name):
    try:
        with engine.connect() as c:
            rows=c.execute(text("""SELECT column_name
                FROM information_schema.columns
                WHERE table_schema='public' AND table_name=:t"""),{"t":table_name}).fetchall()
        return {r._mapping["column_name"] for r in rows}
    except Exception:
        return set()

def _v152_table_exists(table_name):
    try:
        with engine.connect() as c:
            return bool(c.execute(text("""SELECT EXISTS(
                SELECT 1 FROM information_schema.tables
                WHERE table_schema='public' AND table_name=:t
            )"""),{"t":table_name}).scalar_one())
    except Exception:
        return False

def _v152_pick(row,*names):
    for n in names:
        if n in row and row.get(n) not in (None,""):
            return row.get(n)
    return None

def _v152_phone_candidates(row):
    vals=[]
    for k,v in row.items():
        lk=str(k).lower()
        if any(x in lk for x in ["phone","mobile","contact_no","contactnumber","contact_number","whatsapp"]):
            if isinstance(v,list):
                vals.extend(v)
            else:
                vals.append(v)
    out=[]
    for v in vals:
        if not v: continue
        # Split common separators but do not guess digits.
        parts=_re.split(r"[,;/| ]+",str(v))
        for p in parts:
            ph=_v151_digits(p)
            if ph and ph not in out:
                out.append(ph)
    return out

def _v152_detect_hospitality_category(row):
    blob=" ".join(str(v or "") for v in row.values()).lower()
    if "cafe" in blob or "coffee" in blob or "bakery" in blob:
        return "CAFE"
    if "restaurant" in blob or "restro" in blob or "diner" in blob:
        return "RESTAURANT"
    if "banquet" in blob or "wedding venue" in blob:
        return "BANQUET"
    if "guest house" in blob or "guesthouse" in blob:
        return "GUEST_HOUSE"
    if "lounge" in blob:
        return "LOUNGE"
    if "club" in blob:
        return "CLUB"
    if "bar" in blob or "pub" in blob:
        return "BAR"
    if "farmhouse" in blob or "farm house" in blob:
        return "FARMHOUSE"
    if "hotel" in blob or "resort" in blob:
        return "HOTEL"
    return None

def _v152_all_hospitality_source_rows():
    """
    Scan likely hospitality source tables dynamically.
    This avoids assuming one fixed bot schema.
    """
    candidate_tables=[
        "ai_demand_signals",
        "hospitality_prospects",
        "pi_hospitality_prospects",
        "hospitality_contacts",
        "ai_hospitality_contacts",
        "hospitality_leads"
    ]
    rows=[]
    for table in candidate_tables:
        if not _v152_table_exists(table):
            continue
        try:
            with engine.connect() as c:
                data=[dict(r._mapping) for r in c.execute(text(f"SELECT * FROM {table} ORDER BY 1 DESC LIMIT 10000")).fetchall()]
            for r in data:
                cat=_v152_detect_hospitality_category(r)
                blob=" ".join(str(v or "") for v in r.values()).lower()
                # ai_demand_signals can contain non-hospitality data, so require category or hospitality keyword.
                if table=="ai_demand_signals" and not cat and "hospitality" not in blob:
                    continue
                if cat or "hospitality" in blob:
                    r["_v152_source_table"]=table
                    r["_v152_category"]=cat or "OTHER"
                    rows.append(r)
        except Exception:
            continue
    return rows

def _v152_sync_ai_hospitality():
    _v151_setup()
    rows=_v152_all_hospitality_source_rows()
    processed=0
    skipped_no_phone=0
    errors=[]

    for r in rows:
        phones=_v152_phone_candidates(r)
        if not phones:
            skipped_no_phone+=1
            continue

        name=_v152_pick(r,"contact_name","person_name","name","owner_name","manager_name")
        company=_v152_pick(r,"company_name","brand_name","brand","venue_name","business_name","company")
        city=_v152_pick(r,"city","target_city")
        location=_v152_pick(r,"location","address","target_market","area","locality")
        email=_v152_pick(r,"contact_email","email","email_id")
        website=_v152_pick(r,"website","website_url")
        source_url=_v152_pick(r,"source_url","linkedin_post_url","linkedin_url","url","website")
        notes=_v152_pick(r,"excerpt","requirement_text","notes","description","summary")
        cat=r.get("_v152_category") or _v152_detect_hospitality_category(r) or "OTHER"
        src_table=r.get("_v152_source_table") or "UNKNOWN"

        for ph in phones:
            try:
                ok=_v151_upsert_contact(
                    ph,
                    name=name,
                    company=company,
                    category=cat,
                    city=city,
                    location=location,
                    email=email,
                    website=website,
                    source="AI_HOSPITALITY",
                    source_detail=f"{src_table}: {source_url or ''}".strip(),
                    notes=notes
                )
                if ok:
                    processed+=1
            except Exception as ex:
                errors.append(f"{ph}: {type(ex).__name__}: {ex}")

    return {
        "processed":processed,
        "source_rows":len(rows),
        "skipped_no_phone":skipped_no_phone,
        "errors":errors[:20]
    }

@app.post("/api/v15-2/marketing-contacts/sync-ai-hospitality")
def v152_sync_ai_hospitality_api(req:Request):
    need_login(req)
    try:
        result=_v152_sync_ai_hospitality()
        return {"status":"ok",**result}
    except Exception as ex:
        raise HTTPException(500,f"{type(ex).__name__}: {ex}")

@app.get("/api/v15-2/marketing-contacts/debug-ai-hospitality")
def v152_debug_ai_hospitality(req:Request):
    need_login(req)
    rows=_v152_all_hospitality_source_rows()
    sample=[]
    for r in rows[:20]:
        sample.append({
            "source_table":r.get("_v152_source_table"),
            "category":r.get("_v152_category"),
            "phones":_v152_phone_candidates(r),
            "company":_v152_pick(r,"company_name","brand_name","brand","venue_name","business_name","company"),
            "name":_v152_pick(r,"contact_name","person_name","name"),
            "location":_v152_pick(r,"location","address","target_market","area","locality")
        })
    return {"status":"ok","source_rows":len(rows),"sample":sample}

@app.get("/marketing-contacts-v2",response_class=HTMLResponse)
def v152_marketing_contacts_v2(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    return HTMLResponse("""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Marketing Contacts</title>
<style>body{font-family:Arial;margin:0;background:#f4f7fb;color:#172437}header{background:#102235;color:#fff;padding:18px}.w{padding:18px}.bar{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px}.btn,a.btn{padding:8px 10px;border:0;border-radius:7px;background:#1677ff;color:#fff;text-decoration:none;font-weight:bold}.gray{background:#e9eef5!important;color:#203247!important}.msg{background:#fff;border:1px solid #e2e8f0;border-radius:10px;padding:10px;margin-bottom:12px}.tablewrap{overflow:auto;background:#fff;border-radius:10px}table{width:100%;border-collapse:collapse;font-size:12px}th,td{padding:8px;border-bottom:1px solid #edf1f5;text-align:left;white-space:nowrap}th{background:#f8fafc}</style></head>
<body><header><b>Marketing Contacts</b><br><small>Dedicated AI Hospitality sync</small></header><div class=w>
<div class=bar><a class="btn gray" href="/workspace">← Dashboard</a><button class=btn onclick="syncHosp()">Sync AI Hospitality</button><a class="btn gray" href="/marketing-contacts">Open Full Marketing Contacts</a></div>
<div class=msg id=msg>Click <b>Sync AI Hospitality</b>. This scans all known hospitality source tables dynamically.</div>
<div class=tablewrap><table><thead><tr><th>Source</th><th>Category</th><th>Phone</th><th>Company</th><th>Name</th><th>Location</th></tr></thead><tbody id=rows></tbody></table></div></div>
<script>
async function syncHosp(){msg.textContent='Syncing AI hospitality contacts...';let r=await fetch('/api/v15-2/marketing-contacts/sync-ai-hospitality',{method:'POST'}),d=await r.json();if(!r.ok){msg.textContent='ERROR: '+(d.detail||'Sync failed');return}msg.textContent=`Done. ${d.processed} contacts synced from ${d.source_rows} hospitality source rows. ${d.skipped_no_phone} rows had no valid 10-digit mobile.`;debug()}
async function debug(){let d=await(await fetch('/api/v15-2/marketing-contacts/debug-ai-hospitality')).json();rows.innerHTML=(d.sample||[]).map(x=>`<tr><td>${x.source_table||''}</td><td>${x.category||''}</td><td>${(x.phones||[]).join(', ')}</td><td>${x.company||''}</td><td>${x.name||''}</td><td>${x.location||''}</td></tr>`).join('')||'<tr><td colspan=6>No hospitality source rows detected.</td></tr>'}debug()
</script></body></html>""")

# ============================================================
# V15.3 FULL HOSPITALITY CONTACT REBUILD
# Universal schema discovery: finds hospitality/contact/prospect/lead
# tables dynamically and rebuilds Marketing Contacts from real source rows.
# ============================================================

def _v153_candidate_tables():
    with engine.connect() as c:
        rows=c.execute(text("""SELECT table_name
            FROM information_schema.tables
            WHERE table_schema='public'
              AND table_type='BASE TABLE'
              AND (
                   table_name ILIKE '%hospital%'
                OR table_name ILIKE '%contact%'
                OR table_name ILIKE '%prospect%'
                OR table_name ILIKE '%lead%'
              )
              AND table_name NOT IN ('pi_marketing_contacts')
            ORDER BY table_name""")).fetchall()
    return [r._mapping["table_name"] for r in rows]

def _v153_columns(table):
    with engine.connect() as c:
        rows=c.execute(text("""SELECT column_name,data_type
            FROM information_schema.columns
            WHERE table_schema='public' AND table_name=:t
            ORDER BY ordinal_position"""),{"t":table}).fetchall()
    return [dict(r._mapping) for r in rows]

def _v153_phone_cols(cols):
    out=[]
    for c in cols:
        n=c["column_name"].lower()
        if any(x in n for x in ["phone","mobile","contact_no","contactnumber","contact_number","whatsapp","telephone","tel_no"]):
            out.append(c["column_name"])
    return out

def _v153_is_hospitality_row(row, table):
    blob=(" "+table+" "+" ".join(str(v or "") for v in row.values())+" ").lower()
    terms=["hospitality","restaurant","cafe","coffee","banquet","hotel","guest house","guesthouse","lounge","club","bar","pub","farmhouse","farm house","resort"]
    return any(t in blob for t in terms)

def _v153_extract_phones(row, phone_cols):
    found=[]
    for col in phone_cols:
        v=row.get(col)
        if v is None: continue
        vals=v if isinstance(v,(list,tuple)) else _re.split(r"[,;/|]+",str(v))
        for raw in vals:
            candidates=[str(raw)] + _re.findall(r"(?:\+?91[\s\-]?)?[6-9]\d{9}",str(raw))
            for cand in candidates:
                ph=_v151_digits(cand)
                if ph and ph not in found:
                    found.append(ph)
    return found

def _v153_first(row,*keys):
    for k in keys:
        if k in row and row.get(k) not in (None,""):
            return row.get(k)
    return None

def _v153_rebuild_hospitality():
    _v151_setup()
    tables=_v153_candidate_tables()
    source_rows=0
    valid_phone_rows=0
    synced=0
    table_stats=[]
    errors=[]

    for table in tables:
        try:
            cols=_v153_columns(table)
            phone_cols=_v153_phone_cols(cols)
            if not phone_cols:
                table_stats.append({"table":table,"rows":0,"phones":0,"synced":0,"note":"no phone-like columns"})
                continue

            with engine.connect() as c:
                rows=[dict(r._mapping) for r in c.execute(text(f'SELECT * FROM "{table}" LIMIT 20000')).fetchall()]

            t_rows=t_phone=t_sync=0
            for r in rows:
                if not _v153_is_hospitality_row(r,table):
                    continue
                t_rows+=1
                source_rows+=1
                phones=_v153_extract_phones(r,phone_cols)
                if not phones:
                    continue
                t_phone+=1
                valid_phone_rows+=1

                name=_v153_first(r,"contact_name","person_name","name","owner_name","manager_name","contact_person")
                company=_v153_first(r,"company_name","brand_name","brand","venue_name","business_name","company","restaurant_name","hotel_name")
                city=_v153_first(r,"city","target_city")
                location=_v153_first(r,"location","address","target_market","area","locality","market")
                email=_v153_first(r,"contact_email","email","email_id")
                website=_v153_first(r,"website","website_url")
                src_url=_v153_first(r,"source_url","linkedin_post_url","linkedin_url","url")
                notes=_v153_first(r,"excerpt","requirement_text","notes","description","summary","remarks")
                cat=_v152_detect_hospitality_category(r) or "OTHER"

                for ph in phones:
                    try:
                        if _v151_upsert_contact(
                            ph,
                            name=name,
                            company=company,
                            category=cat,
                            city=city,
                            location=location,
                            email=email,
                            website=website,
                            source="AI_HOSPITALITY",
                            source_detail=f"{table}: {src_url or ''}".strip(),
                            notes=notes
                        ):
                            synced+=1
                            t_sync+=1
                    except Exception as ex:
                        errors.append(f"{table}/{ph}: {type(ex).__name__}: {ex}")

            table_stats.append({"table":table,"rows":t_rows,"phones":t_phone,"synced":t_sync})
        except Exception as ex:
            table_stats.append({"table":table,"rows":0,"phones":0,"synced":0,"error":f"{type(ex).__name__}: {ex}"})

    total_contacts=_v15_safe_count("SELECT COUNT(*) FROM pi_marketing_contacts WHERE source ILIKE '%AI_HOSPITALITY%'")
    return {
        "candidate_tables":len(tables),
        "source_rows":source_rows,
        "rows_with_valid_phone":valid_phone_rows,
        "synced":synced,
        "total_ai_hospitality_contacts":total_contacts,
        "table_stats":table_stats,
        "errors":errors[:30]
    }

@app.post("/api/v15-3/marketing-contacts/rebuild-hospitality")
def v153_rebuild_hospitality_api(req:Request):
    need_login(req)
    try:
        return {"status":"ok",**_v153_rebuild_hospitality()}
    except Exception as ex:
        raise HTTPException(500,f"{type(ex).__name__}: {ex}")

@app.get("/marketing-contacts-v3",response_class=HTMLResponse)
def v153_marketing_contacts_page(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    return HTMLResponse("""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Marketing Contacts</title>
<style>*{box-sizing:border-box}body{font-family:Arial;margin:0;background:#f4f7fb;color:#172437}header{background:#102235;color:#fff;padding:18px}.w{padding:18px}.bar{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px}.btn,a.btn{padding:8px 10px;border:0;border-radius:7px;background:#1677ff;color:#fff;text-decoration:none;font-weight:bold;cursor:pointer}.gray{background:#e9eef5!important;color:#203247!important}.msg{background:white;border:1px solid #e2e8f0;border-radius:10px;padding:12px;margin-bottom:12px}.kpis{display:grid;grid-template-columns:repeat(4,minmax(150px,1fr));gap:8px;margin-bottom:12px}.k{background:white;border:1px solid #e2e8f0;border-radius:10px;padding:12px}.k b{font-size:22px;display:block}.tablewrap{overflow:auto;background:white;border-radius:10px}table{width:100%;border-collapse:collapse;font-size:12px}th,td{padding:8px;border-bottom:1px solid #edf1f5;text-align:left;white-space:nowrap}th{background:#f8fafc}@media(max-width:700px){.kpis{grid-template-columns:1fr 1fr}}</style></head>
<body><header><b>Marketing Contacts Database</b><br><small>Full AI Hospitality rebuild + category filters</small></header><div class=w>
<div class=bar><a class="btn gray" href="/workspace">← Dashboard</a><button class=btn onclick="rebuild()">Rebuild Full AI Hospitality Contacts</button></div>
<div class=msg id=msg>Click <b>Rebuild Full AI Hospitality Contacts</b>. This scans the real database schema instead of assuming one table name.</div>
<div class=kpis><div class=k><b id=tables>0</b><span>Source tables found</span></div><div class=k><b id=srows>0</b><span>Hospitality source rows</span></div><div class=k><b id=prows>0</b><span>Rows with valid mobile</span></div><div class=k><b id=total>0</b><span>AI Hospitality contacts</span></div></div>
<div class=tablewrap><table><thead><tr><th>Source Table</th><th>Hospitality Rows</th><th>Rows With Phone</th><th>Contacts Synced</th><th>Status</th></tr></thead><tbody id=rows></tbody></table></div>
</div><script>
async function rebuild(){msg.textContent='Scanning hospitality/contact/prospect/lead tables and rebuilding contacts...';let r=await fetch('/api/v15-3/marketing-contacts/rebuild-hospitality',{method:'POST'}),d=await r.json();if(!r.ok){msg.textContent='ERROR: '+(d.detail||'Rebuild failed');return}tables.textContent=d.candidate_tables||0;srows.textContent=d.source_rows||0;prows.textContent=d.rows_with_valid_phone||0;total.textContent=d.total_ai_hospitality_contacts||0;rows.innerHTML=(d.table_stats||[]).map(x=>`<tr><td>${x.table||''}</td><td>${x.rows||0}</td><td>${x.phones||0}</td><td>${x.synced||0}</td><td>${x.error||x.note||'OK'}</td></tr>`).join('');msg.textContent=`Rebuild complete. Marketing database now has ${d.total_ai_hospitality_contacts||0} AI Hospitality contacts.`;}
</script></body></html>""")

@app.middleware("http")
async def v153_marketing_route_fix(request,call_next):
    if request.url.path in {"/marketing-contacts","/marketing-contacts-v2"}:
        return RedirectResponse("/marketing-contacts-v3",status_code=307)
    return await call_next(request)

# ============================================================
# V15.5.1 UNIVERSAL FINAL MARKETING CONTACTS INTERFACE
# No V15.4 dependency. Uses existing V15.1 contact database APIs.
# ============================================================

@app.get("/api/v15-5-1/marketing-contacts/summary")
def v1551_marketing_summary(req:Request):
    need_login(req)
    _v151_setup()

    def one(sql,params=None):
        try:
            with engine.connect() as c:
                return int(c.execute(text(sql),params or {}).scalar_one() or 0)
        except Exception:
            return 0

    categories={}
    for cat in ["CAFE","RESTAURANT","BANQUET","HOTEL","GUEST_HOUSE","LOUNGE","CLUB","BAR","FARMHOUSE","RETAILER","BROKER","OWNER","OTHER"]:
        categories[cat]=one("SELECT COUNT(*) FROM pi_marketing_contacts WHERE category=:c",{"c":cat})

    return {
        "status":"ok",
        "total":one("SELECT COUNT(*) FROM pi_marketing_contacts"),
        "ai_hospitality":one("SELECT COUNT(*) FROM pi_marketing_contacts WHERE source ILIKE '%AI_HOSPITALITY%'"),
        "ai_retail":one("SELECT COUNT(*) FROM pi_marketing_contacts WHERE source ILIKE '%AI_RETAIL%'"),
        "property_database":one("SELECT COUNT(*) FROM pi_marketing_contacts WHERE source ILIKE '%PROPERTY_DATABASE%'"),
        "magazine":one("SELECT COUNT(*) FROM pi_marketing_contacts WHERE source ILIKE '%MAGAZINE%'"),
        "verified":one("SELECT COUNT(*) FROM pi_marketing_contacts WHERE verified_status='VERIFIED'"),
        "ready":one("SELECT COUNT(*) FROM pi_marketing_contacts WHERE whatsapp_status='READY'"),
        "categories":categories
    }

@app.get("/marketing-contacts-final",response_class=HTMLResponse)
def v1551_marketing_contacts_final(req:Request):
    role=page_role_or_redirect(req)
    if not role:
        return RedirectResponse("/login",303)

    return HTMLResponse("""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Marketing Contacts Database</title>
<style>
*{box-sizing:border-box}body{margin:0;background:#f4f7fb;font-family:Arial;color:#172437}
header{background:#102235;color:white;padding:18px 22px}
.wrap{max-width:1800px;margin:auto;padding:18px}
.nav,.tabs,.filters{display:flex;gap:8px;flex-wrap:wrap;align-items:center;margin-bottom:12px}
.btn,a.btn{padding:8px 11px;border:0;border-radius:8px;background:#1677ff;color:white;text-decoration:none;font-weight:700;cursor:pointer}
.gray{background:#e9eef5!important;color:#203247!important}.green{background:#08734b!important}
.tab{padding:9px 12px;border:1px solid #dbe3ec;background:#fff;border-radius:9px;cursor:pointer;font-weight:700}
.tab.active{background:#102235;color:#fff}
.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:9px;margin-bottom:14px}
.k{background:#fff;border:1px solid #e2e8f0;border-radius:11px;padding:12px}.k b{display:block;font-size:22px}.k span{font-size:11px;color:#687789}
.card{background:#fff;border:1px solid #e2e8f0;border-radius:12px;padding:12px;margin-bottom:12px}
select,input{padding:8px;border:1px solid #ccd6e2;border-radius:7px}input.search{min-width:260px}
.tablewrap{overflow:auto;max-height:68vh;background:#fff;border-radius:10px}
table{width:100%;border-collapse:collapse;font-size:12px}th,td{padding:8px;border-bottom:1px solid #edf1f5;text-align:left;vertical-align:top;white-space:nowrap}
th{position:sticky;top:0;background:#f8fafc;z-index:2}.pill{display:inline-block;padding:3px 7px;border-radius:10px;background:#edf4ff}
.small{font-size:11px;color:#687789}.source{font-weight:700}.catgrid{display:flex;gap:6px;flex-wrap:wrap}.cchip{padding:5px 8px;background:#eef3f8;border-radius:10px;font-size:11px;cursor:pointer}.cchip b{margin-left:4px}
.msg{background:#fff8e8;border:1px solid #eed18f;border-radius:9px;padding:10px;margin-bottom:12px}
</style>
</head>
<body>
<header><b>Marketing Contacts Database</b><br><small>Segregated by Category · Source · Verification · WhatsApp Status</small></header>
<div class="wrap">

<div class="nav">
<a class="btn gray" href="/workspace">← Dashboard</a>
<button class="btn green" onclick="syncHosp()">Sync AI Hospitality</button>
<button class="btn" onclick="syncAll()">Sync All Sources</button>
</div>

<div class="kpis">
<div class="k"><b id="ktotal">0</b><span>ALL CONTACTS</span></div>
<div class="k"><b id="khosp">0</b><span>AI HOSPITALITY</span></div>
<div class="k"><b id="kret">0</b><span>AI RETAIL</span></div>
<div class="k"><b id="kprop">0</b><span>PROPERTY DATABASE</span></div>
<div class="k"><b id="kmag">0</b><span>MAGAZINE</span></div>
<div class="k"><b id="kver">0</b><span>VERIFIED</span></div>
<div class="k"><b id="kready">0</b><span>WHATSAPP READY</span></div>
</div>

<div class="tabs">
<button class="tab active" data-source="ALL" onclick="setSource(this)">All</button>
<button class="tab" data-source="AI_HOSPITALITY" onclick="setSource(this)">AI Hospitality</button>
<button class="tab" data-source="AI_RETAIL" onclick="setSource(this)">AI Retail</button>
<button class="tab" data-source="PROPERTY_DATABASE" onclick="setSource(this)">Property Database</button>
<button class="tab" data-source="MAGAZINE" onclick="setSource(this)">Magazine</button>
</div>

<div class="card">
<div><b>Category Segregation</b></div>
<div id="catgrid" class="catgrid"></div>
</div>

<div class="filters">
<select id="category">
<option>ALL</option><option>CAFE</option><option>RESTAURANT</option><option>BANQUET</option><option>HOTEL</option>
<option>GUEST_HOUSE</option><option>LOUNGE</option><option>CLUB</option><option>BAR</option><option>FARMHOUSE</option>
<option>RETAILER</option><option>BROKER</option><option>OWNER</option><option>OTHER</option>
</select>
<select id="verified"><option>ALL</option><option>UNVERIFIED</option><option>VERIFIED</option></select>
<select id="whatsapp"><option>ALL</option><option>NOT_CONTACTED</option><option>READY</option><option>SENT</option><option>REPLIED</option><option>OPT_OUT</option></select>
<input class="search" id="q" placeholder="Search name, mobile, brand, location, email">
<button class="btn" onclick="loadContacts()">Search</button>
<span id="count" class="small"></span>
</div>

<div id="msg" class="msg">Use source tabs and category filters. Example: <b>AI Hospitality + CAFE</b>.</div>

<div class="tablewrap">
<table>
<thead><tr>
<th>Select</th><th>Name</th><th>Mobile</th><th>Brand / Company</th><th>Category</th><th>Location</th><th>Source</th><th>Verified</th><th>WhatsApp</th><th>Linked Properties</th><th>Date Added</th><th>Notes</th>
</tr></thead>
<tbody id="rows"></tbody>
</table>
</div>
</div>

<script>
const E=x=>String(x??'').replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));
let source='ALL';

async function A(u,o={}){
  let r=await fetch(u,o),d=await r.json();
  if(!r.ok) throw Error(d.detail||'Error');
  return d;
}
function setSource(el){
  document.querySelectorAll('.tab').forEach(x=>x.classList.remove('active'));
  el.classList.add('active'); source=el.dataset.source; loadContacts();
}
async function summary(){
  let d=await A('/api/v15-5-1/marketing-contacts/summary');
  ktotal.textContent=d.total||0;khosp.textContent=d.ai_hospitality||0;kret.textContent=d.ai_retail||0;
  kprop.textContent=d.property_database||0;kmag.textContent=d.magazine||0;kver.textContent=d.verified||0;kready.textContent=d.ready||0;
  catgrid.innerHTML=Object.entries(d.categories||{}).map(([k,v])=>`<span class="cchip" onclick="pickCat('${k}')">${k}<b>${v}</b></span>`).join('');
}
function pickCat(c){category.value=c;loadContacts();}
async function loadContacts(){
  let u='/api/v15-1/marketing-contacts?category='+encodeURIComponent(category.value)
      +'&source='+encodeURIComponent(source)
      +'&verified='+encodeURIComponent(verified.value)
      +'&whatsapp='+encodeURIComponent(whatsapp.value)
      +'&q='+encodeURIComponent(q.value||'');
  let d=await A(u),r=d.rows||[];
  count.textContent=r.length+' contacts';
  rows.innerHTML=r.map(x=>`<tr>
    <td><input type="checkbox" ${x.opt_out?'disabled':''}></td>
    <td>${E(x.contact_name||'')}</td>
    <td><b>${E(x.primary_phone||'')}</b><br><span class="small">${E((x.all_phones||[]).join(', '))}</span></td>
    <td>${E(x.company_brand||'')}</td>
    <td><span class="pill">${E(x.category||'OTHER')}</span></td>
    <td>${E(x.city||'')}<br>${E(x.location||'')}</td>
    <td><span class="source">${E(x.source||'')}</span><br><span class="small">${E(x.source_detail||'')}</span></td>
    <td><select onchange="upd(${x.id},'verified_status',this.value)">
      <option ${x.verified_status==='UNVERIFIED'?'selected':''}>UNVERIFIED</option>
      <option ${x.verified_status==='VERIFIED'?'selected':''}>VERIFIED</option>
    </select></td>
    <td><select onchange="upd(${x.id},'whatsapp_status',this.value)">
      <option ${x.whatsapp_status==='NOT_CONTACTED'?'selected':''}>NOT_CONTACTED</option>
      <option ${x.whatsapp_status==='READY'?'selected':''}>READY</option>
      <option ${x.whatsapp_status==='SENT'?'selected':''}>SENT</option>
      <option ${x.whatsapp_status==='REPLIED'?'selected':''}>REPLIED</option>
      <option ${x.whatsapp_status==='OPT_OUT'?'selected':''}>OPT_OUT</option>
    </select></td>
    <td>${E(x.linked_property_count||0)}</td>
    <td>${E((x.date_added||'').slice(0,10))}</td>
    <td>${E(x.notes||'')}</td>
  </tr>`).join('') || '<tr><td colspan="12">No contacts found for this filter.</td></tr>';
}
async function upd(id,k,v){
  let b={};b[k]=v;
  await A('/api/v15-1/marketing-contacts/'+id+'/update',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(b)});
  summary();
}
async function syncHosp(){
  msg.textContent='Syncing AI Hospitality contacts...';
  try{
    let endpoint='/api/v15-3/marketing-contacts/rebuild-hospitality';
    let r=await fetch(endpoint,{method:'POST'});
    if(r.status===404){endpoint='/api/v15-2/marketing-contacts/sync-ai-hospitality';r=await fetch(endpoint,{method:'POST'});}
    let d=await r.json();
    if(!r.ok) throw Error(d.detail||'Sync failed');
    msg.textContent='AI Hospitality sync complete.';
    await summary();await loadContacts();
  }catch(e){msg.textContent='ERROR: '+e.message}
}
async function syncAll(){
  msg.textContent='Syncing all sources...';
  try{
    let d=await A('/api/v15-1/marketing-contacts/sync',{method:'POST'});
    msg.textContent='Source sync complete.';
    await summary();await loadContacts();
  }catch(e){msg.textContent='ERROR: '+e.message}
}
[category,verified,whatsapp].forEach(x=>x.onchange=loadContacts);
q.onkeydown=e=>{if(e.key==='Enter')loadContacts()};
summary();loadContacts();
</script>
</body>
</html>""")

@app.middleware("http")
async def v1551_marketing_final_router(request,call_next):
    if request.url.path in {"/marketing-contacts","/marketing-contacts-v2","/marketing-contacts-v3","/marketing-contacts-v4"}:
        return RedirectResponse("/marketing-contacts-final",status_code=307)
    return await call_next(request)

# ============================================================
# V15.6 RECOVER EXISTING AI HOSPITALITY MASTER
# Recovers ALL existing ai_marketing_contacts rows first.
# No Hospitality Bot rerun required.
# ============================================================

def _v156_setup():
    with engine.begin() as c:
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_ai_hospitality_master(
            id BIGSERIAL PRIMARY KEY,
            source_row_key TEXT UNIQUE NOT NULL,
            source_table TEXT DEFAULT 'ai_marketing_contacts',
            source_row_id TEXT,
            business_name TEXT,
            category TEXT DEFAULT 'OTHER',
            subcategory TEXT,
            city TEXT,
            location TEXT,
            full_address TEXT,
            contact_name TEXT,
            primary_phone TEXT,
            all_phones JSONB DEFAULT '[]'::jsonb,
            email TEXT,
            website TEXT,
            source_url TEXT,
            raw_text TEXT,
            raw_payload JSONB,
            contact_status TEXT DEFAULT 'NEEDS_ENRICHMENT',
            verified_status TEXT DEFAULT 'UNVERIFIED',
            enrichment_status TEXT DEFAULT 'NOT_STARTED',
            date_fetched TIMESTAMPTZ,
            recovered_at TIMESTAMPTZ DEFAULT NOW(),
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )"""))
        c.execute(text("CREATE INDEX IF NOT EXISTS idx_ai_hosp_master_category ON pi_ai_hospitality_master(category)"))
        c.execute(text("CREATE INDEX IF NOT EXISTS idx_ai_hosp_master_contact_status ON pi_ai_hospitality_master(contact_status)"))
        c.execute(text("CREATE INDEX IF NOT EXISTS idx_ai_hosp_master_phone ON pi_ai_hospitality_master(primary_phone)"))

def _v156_flatten(value, prefix=""):
    out=[]
    if value is None:
        return out
    if isinstance(value,dict):
        for k,v in value.items():
            p=f"{prefix}.{k}" if prefix else str(k)
            out.extend(_v156_flatten(v,p))
        return out
    if isinstance(value,(list,tuple)):
        for i,v in enumerate(value):
            out.extend(_v156_flatten(v,f"{prefix}[{i}]"))
        return out
    if isinstance(value,str):
        s=value.strip()
        if len(s)>=2 and ((s.startswith("{") and s.endswith("}")) or (s.startswith("[") and s.endswith("]"))):
            try:
                return _v156_flatten(json.loads(s),prefix)
            except Exception:
                pass
    out.append((prefix,value))
    return out

def _v156_pairs(row):
    out=[]
    for k,v in row.items():
        out.extend(_v156_flatten(v,str(k)))
    return out

def _v156_find(row,keywords,max_len=500):
    pairs=_v156_pairs(row)
    # exact-ish path preference
    for path,val in pairs:
        lp=path.lower()
        if any(k in lp for k in keywords):
            s=str(val or "").strip()
            if s and len(s)<=max_len:
                return val
    return None

def _v156_phones(row):
    found=[]
    for path,val in _v156_pairs(row):
        txt=str(val or "")
        for m in _re.findall(r"(?<!\d)(?:\+?91[\s\-]?)?([6-9]\d{9})(?!\d)",txt):
            ph=_v151_digits(m)
            if ph and ph not in found:
                found.append(ph)
    return found

def _v156_email(row):
    direct=_v156_find(row,["contact_email","email_id","email"])
    if direct:
        return str(direct).strip()
    blob=" ".join(str(v or "") for _,v in _v156_pairs(row))
    m=_re.search(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}",blob,_re.I)
    return m.group(0) if m else None

def _v156_url(row):
    direct=_v156_find(row,["website_url","website","source_url","linkedin_url","google_url","url"])
    if direct:
        return str(direct).strip()
    blob=" ".join(str(v or "") for _,v in _v156_pairs(row))
    m=_re.search(r"https?://[^\s<>\"]+",blob)
    return m.group(0) if m else None

def _v156_category(row):
    blob=" ".join(str(v or "") for _,v in _v156_pairs(row)).lower()
    mapping=[
        ("CAFE",["cafe","coffee","bakery"]),
        ("RESTAURANT",["restaurant","restro","diner"]),
        ("BANQUET",["banquet","wedding venue"]),
        ("HOTEL",["hotel","resort"]),
        ("GUEST_HOUSE",["guest house","guesthouse"]),
        ("LOUNGE",["lounge"]),
        ("CLUB",["club"]),
        ("BAR",["bar","pub"]),
        ("FARMHOUSE",["farmhouse","farm house"])
    ]
    for cat,terms in mapping:
        if any(t in blob for t in terms):
            return cat
    return "OTHER"

def _v156_source_id(row,idx):
    for k in ["id","contact_id","lead_id","prospect_id","record_id","uuid"]:
        if row.get(k) not in (None,""):
            return str(row.get(k))
    return str(idx)

def _v156_recover_all():
    _v156_setup()
    if not _v152_table_exists("ai_marketing_contacts"):
        raise RuntimeError("ai_marketing_contacts table not found")

    with engine.connect() as c:
        rows=[dict(r._mapping) for r in c.execute(text('SELECT * FROM "ai_marketing_contacts" LIMIT 50000')).fetchall()]

    recovered=0
    contact_ready=0
    needs_enrichment=0
    categories={}
    promoted=0
    samples=[]

    for idx,row in enumerate(rows,1):
        sid=_v156_source_id(row,idx)
        key="ai_marketing_contacts:"+sid

        phones=_v156_phones(row)
        business=_v156_find(row,["business_name","company_name","brand_name","venue_name","restaurant_name","hotel_name","brand","company","name"])
        contact=_v156_find(row,["contact_name","contact_person","person_name","manager_name","owner_name"])
        city=_v156_find(row,["city","target_city"])
        location=_v156_find(row,["location","locality","address","target_market","market","area"])
        full_address=_v156_find(row,["full_address","address"])
        email=_v156_email(row)
        website=_v156_find(row,["website_url","website"])
        source_url=_v156_url(row)
        cat=_v156_category(row)
        categories[cat]=categories.get(cat,0)+1

        # preserve readable raw text plus entire source row
        raw_text=" | ".join(str(v) for v in row.values() if v not in (None,""))[:12000]
        status="CONTACT_READY" if phones or email else "NEEDS_ENRICHMENT"
        if status=="CONTACT_READY": contact_ready+=1
        else: needs_enrichment+=1

        date_fetched=None
        for k in ["created_at","date_fetched","fetched_at","discovered_at","updated_at"]:
            if row.get(k):
                date_fetched=row.get(k)
                break

        payload=json.dumps(row,default=str)

        with engine.begin() as c:
            c.execute(text("""INSERT INTO pi_ai_hospitality_master(
                source_row_key,source_table,source_row_id,business_name,category,city,location,full_address,
                contact_name,primary_phone,all_phones,email,website,source_url,raw_text,raw_payload,
                contact_status,date_fetched,recovered_at,updated_at
            ) VALUES(
                :key,'ai_marketing_contacts',:sid,:business,:cat,:city,:loc,:addr,
                :contact,:phone,CAST(:phones AS jsonb),:email,:website,:url,:raw,CAST(:payload AS jsonb),
                :status,:df,NOW(),NOW()
            )
            ON CONFLICT(source_row_key) DO UPDATE SET
                business_name=EXCLUDED.business_name,category=EXCLUDED.category,city=EXCLUDED.city,
                location=EXCLUDED.location,full_address=EXCLUDED.full_address,contact_name=EXCLUDED.contact_name,
                primary_phone=EXCLUDED.primary_phone,all_phones=EXCLUDED.all_phones,email=EXCLUDED.email,
                website=EXCLUDED.website,source_url=EXCLUDED.source_url,raw_text=EXCLUDED.raw_text,
                raw_payload=EXCLUDED.raw_payload,contact_status=EXCLUDED.contact_status,
                date_fetched=EXCLUDED.date_fetched,updated_at=NOW()"""),{
                "key":key,"sid":sid,"business":business,"cat":cat,"city":city,"loc":location,"addr":full_address,
                "contact":contact,"phone":phones[0] if phones else None,"phones":json.dumps(phones),
                "email":email,"website":website,"url":source_url,"raw":raw_text,"payload":payload,
                "status":status,"df":date_fetched
            })
        recovered+=1

        # Promote valid phone records to Marketing Contacts, but preserve master row regardless.
        if phones:
            for ph in phones:
                try:
                    if _v151_upsert_contact(
                        ph,name=contact,company=business,category=cat,city=city,location=location,
                        email=email,website=website,source="AI_HOSPITALITY",
                        source_detail=f"ai_marketing_contacts:{sid}",notes=raw_text[:1500]
                    ):
                        promoted+=1
                except Exception:
                    pass

        if len(samples)<25:
            samples.append({
                "id":sid,"business":business,"category":cat,"location":location,
                "contact_name":contact,"phones":phones,"email":email,"status":status
            })

    return {
        "source_rows":len(rows),
        "recovered":recovered,
        "contact_ready":contact_ready,
        "needs_enrichment":needs_enrichment,
        "promoted_to_marketing_contacts":promoted,
        "categories":categories,
        "sample":samples
    }

@app.post("/api/v15-6/recover-ai-hospitality")
def v156_recover_api(req:Request):
    need_login(req)
    try:
        return {"status":"ok",**_v156_recover_all()}
    except Exception as ex:
        raise HTTPException(500,f"{type(ex).__name__}: {ex}")

@app.get("/api/v15-6/ai-hospitality-master")
def v156_master_list(req:Request,category:str=Query("ALL"),status:str=Query("ALL"),q:str=Query("")):
    need_login(req);_v156_setup()
    wh=[];p={}
    if category!="ALL":
        wh.append("category=:cat");p["cat"]=category
    if status!="ALL":
        wh.append("contact_status=:st");p["st"]=status
    if q.strip():
        wh.append("""(
            COALESCE(business_name,'') ILIKE :q OR COALESCE(location,'') ILIKE :q OR
            COALESCE(primary_phone,'') ILIKE :q OR COALESCE(email,'') ILIKE :q OR
            COALESCE(contact_name,'') ILIKE :q
        )""");p["q"]="%"+q.strip()+"%"
    sql="SELECT id,source_row_id,business_name,category,city,location,contact_name,primary_phone,all_phones,email,website,source_url,contact_status,verified_status,enrichment_status,date_fetched,recovered_at FROM pi_ai_hospitality_master"
    if wh: sql+=" WHERE "+" AND ".join(wh)
    sql+=" ORDER BY recovered_at DESC,id DESC LIMIT 5000"
    with engine.connect() as c:
        rows=[dict(r._mapping) for r in c.execute(text(sql),p).fetchall()]
    return {"status":"ok","rows":rows}

@app.get("/ai-hospitality-master",response_class=HTMLResponse)
def v156_master_page(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    return HTMLResponse("""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>AI Hospitality Master</title>
<style>*{box-sizing:border-box}body{font-family:Arial;margin:0;background:#f4f7fb;color:#172437}header{background:#102235;color:#fff;padding:18px}.w{padding:18px}.bar{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px}.btn,a.btn{padding:8px 10px;border:0;border-radius:7px;background:#1677ff;color:#fff;text-decoration:none;font-weight:bold;cursor:pointer}.gray{background:#e9eef5!important;color:#203247!important}.kpis{display:grid;grid-template-columns:repeat(4,minmax(150px,1fr));gap:8px;margin-bottom:12px}.k{background:#fff;padding:12px;border-radius:10px;border:1px solid #e2e8f0}.k b{display:block;font-size:22px}.msg{background:#fff8e8;border:1px solid #eed18f;border-radius:9px;padding:10px;margin-bottom:12px}select,input{padding:8px;border:1px solid #ccd6e2;border-radius:7px}.tablewrap{overflow:auto;background:white;border-radius:10px;max-height:70vh}table{width:100%;border-collapse:collapse;font-size:12px}th,td{padding:8px;border-bottom:1px solid #edf1f5;text-align:left;vertical-align:top;white-space:nowrap}th{position:sticky;top:0;background:#f8fafc}.pill{padding:3px 6px;background:#edf4ff;border-radius:8px}.good{background:#dcfce7}.warn{background:#fef3c7}</style></head>
<body><header><b>AI Hospitality Master Database</b><br><small>Recover all existing AI-generated cafe / restaurant / banquet / hotel records first</small></header><div class=w>
<div class=bar><a class="btn gray" href="/workspace">← Dashboard</a><button class=btn onclick="recover()">Recover Existing AI Hospitality Data</button><a class="btn gray" href="/marketing-contacts-final">Marketing Contacts</a>
<select id=category><option>ALL</option><option>CAFE</option><option>RESTAURANT</option><option>BANQUET</option><option>HOTEL</option><option>GUEST_HOUSE</option><option>LOUNGE</option><option>CLUB</option><option>BAR</option><option>FARMHOUSE</option><option>OTHER</option></select>
<select id=status><option>ALL</option><option>CONTACT_READY</option><option>NEEDS_ENRICHMENT</option></select><input id=q placeholder="Search business, location, phone, email"><button onclick="load()">Search</button></div>
<div class=kpis><div class=k><b id=src>0</b><span>SOURCE ROWS</span></div><div class=k><b id=rec>0</b><span>RECOVERED</span></div><div class=k><b id=ready>0</b><span>CONTACT READY</span></div><div class=k><b id=need>0</b><span>NEEDS ENRICHMENT</span></div></div>
<div id=msg class=msg>This recovers the existing AI Hospitality database. It does not rerun the Hospitality Bot.</div>
<div class=tablewrap><table><thead><tr><th>Business</th><th>Category</th><th>Location</th><th>Contact Person</th><th>Mobile</th><th>Email</th><th>Website / Source</th><th>Status</th><th>Date Fetched</th></tr></thead><tbody id=rows></tbody></table></div></div>
<script>
const E=x=>String(x??'').replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));
async function recover(){msg.textContent='Recovering existing AI Hospitality rows...';let r=await fetch('/api/v15-6/recover-ai-hospitality',{method:'POST'}),d=await r.json();if(!r.ok){msg.textContent='ERROR: '+(d.detail||'Recovery failed');return}src.textContent=d.source_rows||0;rec.textContent=d.recovered||0;ready.textContent=d.contact_ready||0;need.textContent=d.needs_enrichment||0;msg.textContent=`Recovered ${d.recovered||0} existing AI Hospitality records. ${d.contact_ready||0} are contact-ready; ${d.needs_enrichment||0} need enrichment.`;load()}
async function load(){let u='/api/v15-6/ai-hospitality-master?category='+category.value+'&status='+status.value+'&q='+encodeURIComponent(q.value||'');let d=await(await fetch(u)).json(),r=d.rows||[];rows.innerHTML=r.map(x=>`<tr><td><b>${E(x.business_name||'Unknown business')}</b></td><td><span class=pill>${E(x.category||'OTHER')}</span></td><td>${E(x.city||'')}<br>${E(x.location||'')}</td><td>${E(x.contact_name||'')}</td><td><b>${E(x.primary_phone||'')}</b><br>${E((x.all_phones||[]).join(', '))}</td><td>${E(x.email||'')}</td><td>${x.website?`<a target=_blank href="${E(x.website)}">Website</a>`:''} ${x.source_url?`<a target=_blank href="${E(x.source_url)}">Source</a>`:''}</td><td><span class="pill ${x.contact_status==='CONTACT_READY'?'good':'warn'}">${E(x.contact_status)}</span></td><td>${E((x.date_fetched||'').slice(0,10))}</td></tr>`).join('')||'<tr><td colspan=9>No recovered records for this filter.</td></tr>'}
category.onchange=load;status.onchange=load;q.onkeydown=e=>{if(e.key==='Enter')load()};load()
</script></body></html>""")

# ============================================================
# V15.7 HOSPITALITY MASTER SEPARATION
# Keeps AI Hospitality records separate from Owner/Broker/Property contacts.
# ============================================================

@app.get("/api/v15-7/hospitality-master/summary")
def v157_hospitality_summary(req:Request):
    need_login(req); _v156_setup()
    def one(sql,params=None):
        try:
            with engine.connect() as c:
                return int(c.execute(text(sql),params or {}).scalar_one() or 0)
        except Exception:
            return 0
    cats={}
    for cat in ["CAFE","RESTAURANT","BANQUET","HOTEL","GUEST_HOUSE","LOUNGE","CLUB","BAR","FARMHOUSE","OTHER"]:
        cats[cat]=one("SELECT COUNT(*) FROM pi_ai_hospitality_master WHERE category=:c",{"c":cat})
    return {
        "status":"ok",
        "total":one("SELECT COUNT(*) FROM pi_ai_hospitality_master"),
        "contact_ready":one("SELECT COUNT(*) FROM pi_ai_hospitality_master WHERE contact_status='CONTACT_READY'"),
        "needs_enrichment":one("SELECT COUNT(*) FROM pi_ai_hospitality_master WHERE contact_status='NEEDS_ENRICHMENT'"),
        "verified":one("SELECT COUNT(*) FROM pi_ai_hospitality_master WHERE verified_status='VERIFIED'"),
        "categories":cats
    }

@app.get("/ai-hospitality-master-only",response_class=HTMLResponse)
def v157_hospitality_master_only(req:Request):
    role=page_role_or_redirect(req)
    if not role:
        return RedirectResponse("/login",303)
    return HTMLResponse("""<!doctype html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>AI Hospitality Master</title>
<style>
*{box-sizing:border-box}body{margin:0;background:#f4f7fb;font-family:Arial;color:#172437}
header{background:#102235;color:white;padding:18px 22px}.wrap{max-width:1700px;margin:auto;padding:18px}
.nav,.filters{display:flex;gap:8px;flex-wrap:wrap;align-items:center;margin-bottom:12px}
.btn,a.btn{padding:8px 10px;border:0;border-radius:8px;background:#1677ff;color:#fff;text-decoration:none;font-weight:bold;cursor:pointer}
.gray{background:#e9eef5!important;color:#203247!important}
.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:9px;margin-bottom:12px}
.k{background:#fff;border:1px solid #e2e8f0;border-radius:11px;padding:12px}.k b{display:block;font-size:22px}.k span{font-size:11px;color:#687789}
.card{background:#fff;border:1px solid #e2e8f0;border-radius:11px;padding:12px;margin-bottom:12px}
.catgrid{display:flex;gap:6px;flex-wrap:wrap}.chip{padding:5px 8px;border-radius:10px;background:#eef3f8;font-size:11px;cursor:pointer}.chip b{margin-left:4px}
select,input{padding:8px;border:1px solid #ccd6e2;border-radius:7px}.tablewrap{overflow:auto;max-height:70vh;background:#fff;border-radius:10px}
table{width:100%;border-collapse:collapse;font-size:12px}th,td{padding:8px;border-bottom:1px solid #edf1f5;text-align:left;vertical-align:top;white-space:nowrap}
th{position:sticky;top:0;background:#f8fafc}.pill{padding:3px 7px;border-radius:10px;background:#edf4ff}.good{background:#dcfce7}.warn{background:#fef3c7}
.msg{background:#fff8e8;border:1px solid #eed18f;border-radius:9px;padding:10px;margin-bottom:12px}
</style></head>
<body>
<header><b>AI Hospitality Master</b><br><small>Only Hospitality Bot records. Owner/Broker/Property contacts are excluded.</small></header>
<div class="wrap">
<div class="nav">
<a class="btn gray" href="/workspace">← Dashboard</a>
<button class="btn" onclick="recover()">Recover Existing Hospitality Bot Data</button>
<a class="btn gray" href="/marketing-contacts-final">Marketing Contacts</a>
</div>

<div class="kpis">
<div class="k"><b id="total">0</b><span>HOSPITALITY MASTER ROWS</span></div>
<div class="k"><b id="ready">0</b><span>CONTACT READY</span></div>
<div class="k"><b id="need">0</b><span>NEEDS ENRICHMENT</span></div>
<div class="k"><b id="ver">0</b><span>VERIFIED</span></div>
</div>

<div class="card"><b>Hospitality Categories</b><div id="cats" class="catgrid"></div></div>

<div class="filters">
<select id="category">
<option>ALL</option><option>CAFE</option><option>RESTAURANT</option><option>BANQUET</option><option>HOTEL</option>
<option>GUEST_HOUSE</option><option>LOUNGE</option><option>CLUB</option><option>BAR</option><option>FARMHOUSE</option><option>OTHER</option>
</select>
<select id="status"><option>ALL</option><option>CONTACT_READY</option><option>NEEDS_ENRICHMENT</option></select>
<input id="q" placeholder="Search business, location, mobile, email">
<button class="btn" onclick="load()">Search</button>
<span id="count"></span>
</div>

<div id="msg" class="msg">This page does not show owners or brokers. It shows only records recovered from the AI Hospitality source table.</div>

<div class="tablewrap"><table>
<thead><tr><th>Business</th><th>Category</th><th>Location</th><th>Contact Person</th><th>Mobile</th><th>Email</th><th>Website / Source</th><th>Status</th><th>Date Fetched</th></tr></thead>
<tbody id="rows"></tbody></table></div>
</div>

<script>
const E=x=>String(x??'').replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));
async function A(u,o={}){let r=await fetch(u,o),d=await r.json();if(!r.ok)throw Error(d.detail||'Error');return d}
async function summary(){
  let d=await A('/api/v15-7/hospitality-master/summary');
  total.textContent=d.total||0;ready.textContent=d.contact_ready||0;need.textContent=d.needs_enrichment||0;ver.textContent=d.verified||0;
  cats.innerHTML=Object.entries(d.categories||{}).map(([k,v])=>`<span class="chip" onclick="pick('${k}')">${k}<b>${v}</b></span>`).join('');
}
function pick(c){category.value=c;load();}
async function recover(){
  msg.textContent='Recovering existing Hospitality Bot records...';
  try{
    let d=await A('/api/v15-6/recover-ai-hospitality',{method:'POST'});
    msg.textContent=`Recovered ${d.recovered||0} Hospitality Bot rows. ${d.contact_ready||0} contact-ready; ${d.needs_enrichment||0} need enrichment.`;
    await summary();await load();
  }catch(e){msg.textContent='ERROR: '+e.message}
}
async function load(){
  let u='/api/v15-6/ai-hospitality-master?category='+encodeURIComponent(category.value)+'&status='+encodeURIComponent(status.value)+'&q='+encodeURIComponent(q.value||'');
  let d=await A(u),r=d.rows||[];count.textContent=r.length+' records';
  rows.innerHTML=r.map(x=>`<tr>
    <td><b>${E(x.business_name||'Unknown business')}</b></td>
    <td><span class="pill">${E(x.category||'OTHER')}</span></td>
    <td>${E(x.city||'')}<br>${E(x.location||'')}</td>
    <td>${E(x.contact_name||'')}</td>
    <td><b>${E(x.primary_phone||'')}</b><br>${E((x.all_phones||[]).join(', '))}</td>
    <td>${E(x.email||'')}</td>
    <td>${x.website?`<a target="_blank" href="${E(x.website)}">Website</a>`:''} ${x.source_url?`<a target="_blank" href="${E(x.source_url)}">Source</a>`:''}</td>
    <td><span class="pill ${x.contact_status==='CONTACT_READY'?'good':'warn'}">${E(x.contact_status)}</span></td>
    <td>${E((x.date_fetched||'').slice(0,10))}</td>
  </tr>`).join('')||'<tr><td colspan="9">No Hospitality Bot records for this filter.</td></tr>';
}
category.onchange=load;status.onchange=load;q.onkeydown=e=>{if(e.key==='Enter')load()};
summary();load();
</script></body></html>""")

@app.middleware("http")
async def v157_hospitality_nav_fix(request,call_next):
    if request.url.path=="/ai-hospitality-master":
        return RedirectResponse("/ai-hospitality-master-only",status_code=307)
    return await call_next(request)

# ============================================================
# V15.7.1 HOSPITALITY LIST DISPLAY FIX
# Fixes browser window.status collision that caused ALL records
# to query as status=undefined and display 0 rows.
# ============================================================

@app.get("/ai-hospitality-master-final", response_class=HTMLResponse)
def v1571_hospitality_master_final(req:Request):
    role=page_role_or_redirect(req)
    if not role:
        return RedirectResponse("/login",303)

    return HTMLResponse("""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>AI Hospitality Master</title>
<style>
*{box-sizing:border-box}
body{margin:0;background:#f4f7fb;font-family:Arial;color:#172437}
header{background:#102235;color:#fff;padding:18px 22px}
.wrap{max-width:1750px;margin:auto;padding:18px}
.nav,.filters{display:flex;gap:8px;flex-wrap:wrap;align-items:center;margin-bottom:12px}
.btn,a.btn{padding:8px 11px;border:0;border-radius:8px;background:#1677ff;color:#fff;text-decoration:none;font-weight:700;cursor:pointer}
.gray{background:#e9eef5!important;color:#203247!important}
.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:9px;margin-bottom:12px}
.k{background:#fff;border:1px solid #e2e8f0;border-radius:11px;padding:12px}.k b{display:block;font-size:22px}.k span{font-size:11px;color:#687789}
.card{background:#fff;border:1px solid #e2e8f0;border-radius:11px;padding:12px;margin-bottom:12px}
.catgrid{display:flex;gap:6px;flex-wrap:wrap}.chip{padding:5px 8px;border-radius:10px;background:#eef3f8;font-size:11px;cursor:pointer}.chip b{margin-left:4px}
select,input{padding:8px;border:1px solid #ccd6e2;border-radius:7px}
.tablewrap{overflow:auto;max-height:70vh;background:#fff;border-radius:10px}
table{width:100%;border-collapse:collapse;font-size:12px}
th,td{padding:8px;border-bottom:1px solid #edf1f5;text-align:left;vertical-align:top;white-space:nowrap}
th{position:sticky;top:0;background:#f8fafc;z-index:2}
.pill{padding:3px 7px;border-radius:10px;background:#edf4ff}.good{background:#dcfce7}.warn{background:#fef3c7}
.msg{background:#fff8e8;border:1px solid #eed18f;border-radius:9px;padding:10px;margin-bottom:12px}
.small{font-size:11px;color:#687789}
</style>
</head>
<body>
<header><b>AI Hospitality Master</b><br><small>Only Hospitality Bot records · Property owners/brokers excluded</small></header>
<div class="wrap">

<div class="nav">
<a class="btn gray" href="/workspace">← Dashboard</a>
<button class="btn" onclick="recoverHospitality()">Recover Existing Hospitality Bot Data</button>
<a class="btn gray" href="/marketing-contacts-final">Marketing Contacts</a>
</div>

<div class="kpis">
<div class="k"><b id="kTotal">0</b><span>HOSPITALITY MASTER ROWS</span></div>
<div class="k"><b id="kReady">0</b><span>CONTACT READY</span></div>
<div class="k"><b id="kNeed">0</b><span>NEEDS ENRICHMENT</span></div>
<div class="k"><b id="kVerified">0</b><span>VERIFIED</span></div>
</div>

<div class="card">
<b>Hospitality Categories</b>
<div id="categoryChips" class="catgrid"></div>
</div>

<div class="filters">
<select id="categoryFilter">
<option value="ALL">ALL</option><option>CAFE</option><option>RESTAURANT</option><option>BANQUET</option><option>HOTEL</option>
<option>GUEST_HOUSE</option><option>LOUNGE</option><option>CLUB</option><option>BAR</option><option>FARMHOUSE</option><option>OTHER</option>
</select>

<select id="contactStatusFilter">
<option value="ALL">ALL</option>
<option value="CONTACT_READY">CONTACT_READY</option>
<option value="NEEDS_ENRICHMENT">NEEDS_ENRICHMENT</option>
</select>

<input id="searchBox" placeholder="Search business, location, mobile, email">
<button class="btn" onclick="loadHospitality()">Search</button>
<span id="recordCount" class="small"></span>
</div>

<div id="message" class="msg">
Recovered data is stored separately from property owners/brokers. Use category and contact-status filters to review all 1,298 Hospitality Bot records.
</div>

<div class="tablewrap">
<table>
<thead>
<tr>
<th>Business</th><th>Category</th><th>Location</th><th>Contact Person</th>
<th>Mobile</th><th>Email</th><th>Website / Source</th><th>Status</th><th>Date Fetched</th>
</tr>
</thead>
<tbody id="hospitalityRows"></tbody>
</table>
</div>
</div>

<script>
const E=x=>String(x??'').replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));

async function api(url,opts={}){
  const r=await fetch(url,opts);
  let d={};
  try{d=await r.json()}catch(e){}
  if(!r.ok) throw Error(d.detail||('HTTP '+r.status));
  return d;
}

async function loadSummary(){
  const d=await api('/api/v15-7/hospitality-master/summary');
  document.getElementById('kTotal').textContent=d.total||0;
  document.getElementById('kReady').textContent=d.contact_ready||0;
  document.getElementById('kNeed').textContent=d.needs_enrichment||0;
  document.getElementById('kVerified').textContent=d.verified||0;

  document.getElementById('categoryChips').innerHTML=
    Object.entries(d.categories||{}).map(([k,v]) =>
      `<span class="chip" onclick="selectCategory('${k}')">${E(k)}<b>${v}</b></span>`
    ).join('');
}

function selectCategory(cat){
  document.getElementById('categoryFilter').value=cat;
  loadHospitality();
}

async function recoverHospitality(){
  const msg=document.getElementById('message');
  msg.textContent='Recovering existing Hospitality Bot data...';
  try{
    const d=await api('/api/v15-6/recover-ai-hospitality',{method:'POST'});
    msg.textContent=`Recovered ${d.recovered||0} Hospitality Bot rows. ${d.contact_ready||0} contact-ready; ${d.needs_enrichment||0} need enrichment.`;
    await loadSummary();
    await loadHospitality();
  }catch(e){
    msg.textContent='ERROR: '+e.message;
  }
}

async function loadHospitality(){
  const cat=document.getElementById('categoryFilter').value || 'ALL';
  const contactState=document.getElementById('contactStatusFilter').value || 'ALL';
  const query=document.getElementById('searchBox').value || '';

  const url='/api/v15-6/ai-hospitality-master?category='+
      encodeURIComponent(cat)+'&status='+encodeURIComponent(contactState)+
      '&q='+encodeURIComponent(query);

  const msg=document.getElementById('message');

  try{
    const d=await api(url);
    const list=d.rows||[];
    document.getElementById('recordCount').textContent=list.length+' records';

    document.getElementById('hospitalityRows').innerHTML=list.map(x=>`<tr>
      <td><b>${E(x.business_name||'Unknown business')}</b></td>
      <td><span class="pill">${E(x.category||'OTHER')}</span></td>
      <td>${E(x.city||'')}<br>${E(x.location||'')}</td>
      <td>${E(x.contact_name||'')}</td>
      <td><b>${E(x.primary_phone||'')}</b><br><span class="small">${E((x.all_phones||[]).join(', '))}</span></td>
      <td>${E(x.email||'')}</td>
      <td>
        ${x.website?`<a target="_blank" href="${E(x.website)}">Website</a>`:''}
        ${x.source_url?` <a target="_blank" href="${E(x.source_url)}">Source</a>`:''}
      </td>
      <td><span class="pill ${x.contact_status==='CONTACT_READY'?'good':'warn'}">${E(x.contact_status||'')}</span></td>
      <td>${E((x.date_fetched||'').slice(0,10))}</td>
    </tr>`).join('') || '<tr><td colspan="9">No Hospitality Bot records for this filter.</td></tr>';

    msg.textContent='Showing '+list.length+' Hospitality Bot records. Owner/Broker property contacts are excluded.';
  }catch(e){
    document.getElementById('recordCount').textContent='0 records';
    document.getElementById('hospitalityRows').innerHTML='<tr><td colspan="9">Unable to load Hospitality records.</td></tr>';
    msg.textContent='LOAD ERROR: '+e.message;
  }
}

document.getElementById('categoryFilter').addEventListener('change',loadHospitality);
document.getElementById('contactStatusFilter').addEventListener('change',loadHospitality);
document.getElementById('searchBox').addEventListener('keydown',e=>{if(e.key==='Enter')loadHospitality()});

loadSummary();
loadHospitality();
</script>
</body></html>""")

@app.middleware("http")
async def v1571_hospitality_final_router(request,call_next):
    if request.url.path in {"/ai-hospitality-master","/ai-hospitality-master-only"}:
        return RedirectResponse("/ai-hospitality-master-final",status_code=307)
    return await call_next(request)

# ============================================================
# V15.8 HISTORICAL HOSPITALITY PHONE RECOVERY
# Purpose: recover previously-fetched phone numbers for WhatsApp marketing.
# Scans historical database tables and raw payloads, then links phones
# back to AI Hospitality Master using business/location/source evidence.
# Never invents phone digits.
# ============================================================

def _v158_setup():
    _v156_setup()
    with engine.begin() as c:
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_hospitality_phone_evidence(
            id BIGSERIAL PRIMARY KEY,
            hospitality_master_id BIGINT,
            business_name TEXT,
            recovered_phone TEXT,
            source_table TEXT,
            source_row_ref TEXT,
            match_method TEXT,
            confidence INTEGER,
            evidence_text TEXT,
            applied BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMPTZ DEFAULT NOW()
        )"""))
        c.execute(text("CREATE INDEX IF NOT EXISTS idx_hosp_phone_evidence_master ON pi_hospitality_phone_evidence(hospitality_master_id)"))
        c.execute(text("CREATE INDEX IF NOT EXISTS idx_hosp_phone_evidence_phone ON pi_hospitality_phone_evidence(recovered_phone)"))

def _v158_norm(s):
    return _re.sub(r"[^a-z0-9]+"," ",str(s or "").lower()).strip()

def _v158_valid_phones_from_text(txt):
    found=[]
    for m in _re.findall(r"(?<!\d)(?:\+?91[\s\-]?)?([6-9]\d{9})(?!\d)",str(txt or "")):
        ph=_v151_digits(m)
        if ph and ph not in found:
            found.append(ph)
    return found

def _v158_table_names():
    with engine.connect() as c:
        rows=c.execute(text("""SELECT table_name FROM information_schema.tables
            WHERE table_schema='public' AND table_type='BASE TABLE'
            ORDER BY table_name""")).fetchall()
    excluded={
        "pi_hospitality_phone_evidence",
        "pi_ai_hospitality_master",
        "pi_marketing_contacts"
    }
    return [r._mapping["table_name"] for r in rows if r._mapping["table_name"] not in excluded]

def _v158_table_columns(table):
    with engine.connect() as c:
        rows=c.execute(text("""SELECT column_name,data_type FROM information_schema.columns
            WHERE table_schema='public' AND table_name=:t ORDER BY ordinal_position"""),{"t":table}).fetchall()
    return [dict(r._mapping) for r in rows]

def _v158_relevant_table(table, cols):
    name=table.lower()
    ctext=" ".join(c["column_name"].lower() for c in cols)
    keywords=["hospital","marketing","contact","lead","prospect","ai_","bot","source","raw","business"]
    return any(k in name for k in keywords) or any(k in ctext for k in ["phone","mobile","whatsapp","contact","payload","raw"])

def _v158_row_blob(row):
    parts=[]
    for k,v in row.items():
        if v is None: continue
        try:
            if isinstance(v,(dict,list,tuple)):
                parts.append(json.dumps(v,default=str))
            else:
                parts.append(str(v))
        except Exception:
            parts.append(str(v))
    return " | ".join(parts)

def _v158_master_rows():
    with engine.connect() as c:
        return [dict(r._mapping) for r in c.execute(text("""SELECT id,source_row_id,business_name,category,city,location,
            primary_phone,email,website,source_url,raw_text,contact_status
            FROM pi_ai_hospitality_master ORDER BY id""")).fetchall()]

def _v158_match_candidate(master, blob, table, row):
    bnorm=_v158_norm(blob)
    business=_v158_norm(master.get("business_name"))
    location=_v158_norm(master.get("location"))
    source_id=_v158_norm(master.get("source_row_id"))
    src_url=_v158_norm(master.get("source_url"))
    score=0
    methods=[]

    if business and len(business)>=4 and business in bnorm:
        score+=60;methods.append("BUSINESS")
    if location and len(location)>=8:
        # use first meaningful 2-3 tokens rather than exact giant address
        toks=[x for x in location.split() if len(x)>=4][:4]
        hits=sum(1 for x in toks if x in bnorm)
        if hits>=2:
            score+=20;methods.append("LOCATION")
        elif hits==1:
            score+=8;methods.append("LOCATION_PARTIAL")
    if source_id and source_id in bnorm:
        score+=30;methods.append("SOURCE_ROW_ID")
    if src_url:
        # compare host/path fragments conservatively
        toks=[x for x in src_url.split() if len(x)>=8][:3]
        if any(x in bnorm for x in toks):
            score+=20;methods.append("SOURCE_URL")
    # If source row id equals a visible row id-like value.
    for key in ["id","contact_id","lead_id","prospect_id","record_id","source_id"]:
        if key in row and source_id and _v158_norm(row.get(key))==source_id:
            score+=35;methods.append("ID_EXACT")
            break

    return min(score,100),"+".join(methods) if methods else ""

def _v158_recover_historical_phones():
    _v158_setup()
    masters=_v158_master_rows()
    tables=[]
    for t in _v158_table_names():
        try:
            cols=_v158_table_columns(t)
            if _v158_relevant_table(t,cols):
                tables.append(t)
        except Exception:
            pass

    scanned_rows=0
    evidence_count=0
    applied=0
    master_with_new_phone=set()
    table_stats=[]
    errors=[]

    # Cache master records for simpler matching.
    for table in tables:
        trows=tphones=tevidence=tapplied=0
        try:
            with engine.connect() as c:
                rows=[dict(r._mapping) for r in c.execute(text(f'SELECT * FROM "{table}" LIMIT 30000')).fetchall()]
            trows=len(rows);scanned_rows+=trows

            for row in rows:
                blob=_v158_row_blob(row)
                phones=_v158_valid_phones_from_text(blob)
                if not phones:
                    continue
                tphones+=1

                # Find best hospitality master match.
                best=None
                for m in masters:
                    score,method=_v158_match_candidate(m,blob,table,row)
                    if score>=70 and (best is None or score>best[0]):
                        best=(score,method,m)

                if not best:
                    continue

                score,method,m=best
                sid=str(row.get("id") or row.get("contact_id") or row.get("lead_id") or row.get("prospect_id") or "")
                for ph in phones:
                    # Don't duplicate same evidence.
                    exists=False
                    with engine.connect() as c:
                        ex=c.execute(text("""SELECT 1 FROM pi_hospitality_phone_evidence
                            WHERE hospitality_master_id=:mid AND recovered_phone=:ph AND source_table=:t LIMIT 1"""),
                            {"mid":m["id"],"ph":ph,"t":table}).first()
                        exists=bool(ex)
                    if exists:
                        continue

                    with engine.begin() as c:
                        c.execute(text("""INSERT INTO pi_hospitality_phone_evidence(
                            hospitality_master_id,business_name,recovered_phone,source_table,source_row_ref,
                            match_method,confidence,evidence_text,applied
                        ) VALUES(:mid,:bn,:ph,:t,:ref,:method,:conf,:ev,FALSE)"""),{
                            "mid":m["id"],"bn":m.get("business_name"),"ph":ph,"t":table,"ref":sid,
                            "method":method,"conf":score,"ev":blob[:2500]
                        })
                    evidence_count+=1;tevidence+=1

                    # Auto-apply only very strong matches (90+), and preserve all phones.
                    if score>=90:
                        with engine.begin() as c:
                            current=c.execute(text("""SELECT all_phones,primary_phone FROM pi_ai_hospitality_master WHERE id=:id"""),
                                {"id":m["id"]}).fetchone()
                            if current:
                                cur=dict(current._mapping)
                                arr=cur.get("all_phones") or []
                                if isinstance(arr,str):
                                    try:arr=json.loads(arr)
                                    except:arr=[]
                                arr=[str(x) for x in arr if x]
                                if ph not in arr:arr.append(ph)
                                primary=cur.get("primary_phone") or ph
                                c.execute(text("""UPDATE pi_ai_hospitality_master
                                    SET primary_phone=:primary,all_phones=CAST(:phones AS jsonb),
                                        contact_status='CONTACT_READY',updated_at=NOW()
                                    WHERE id=:id"""),{
                                    "primary":primary,"phones":json.dumps(arr),"id":m["id"]
                                })
                                c.execute(text("""UPDATE pi_hospitality_phone_evidence
                                    SET applied=TRUE WHERE hospitality_master_id=:mid AND recovered_phone=:ph AND source_table=:t"""),
                                    {"mid":m["id"],"ph":ph,"t":table})
                                applied+=1;tapplied+=1;master_with_new_phone.add(m["id"])
                                # Promote to WhatsApp marketing contact DB.
                                try:
                                    _v151_upsert_contact(
                                        ph,
                                        company=m.get("business_name"),
                                        category=m.get("category"),
                                        city=m.get("city"),
                                        location=m.get("location"),
                                        email=m.get("email"),
                                        website=m.get("website"),
                                        source="AI_HOSPITALITY",
                                        source_detail=f"HISTORICAL_RECOVERY:{table}",
                                        notes=f"Recovered historical phone. Confidence {score}. Match {method}."
                                    )
                                except Exception:
                                    pass

            table_stats.append({"table":table,"rows":trows,"rows_with_phone":tphones,"evidence":tevidence,"applied":tapplied})
        except Exception as ex:
            errors.append(f"{table}: {type(ex).__name__}: {ex}")
            table_stats.append({"table":table,"error":f"{type(ex).__name__}: {ex}"})

    total_ready=_v15_safe_count("SELECT COUNT(*) FROM pi_ai_hospitality_master WHERE contact_status='CONTACT_READY'")
    total_with_phone=_v15_safe_count("SELECT COUNT(*) FROM pi_ai_hospitality_master WHERE primary_phone IS NOT NULL AND primary_phone<>''")
    return {
        "candidate_tables":len(tables),
        "rows_scanned":scanned_rows,
        "phone_evidence_found":evidence_count,
        "auto_applied":applied,
        "hospitality_records_improved":len(master_with_new_phone),
        "total_contact_ready":total_ready,
        "total_with_phone":total_with_phone,
        "table_stats":table_stats,
        "errors":errors[:30]
    }

@app.post("/api/v15-8/recover-historical-hospitality-phones")
def v158_recover_api(req:Request):
    need_login(req)
    try:
        return {"status":"ok",**_v158_recover_historical_phones()}
    except Exception as ex:
        raise HTTPException(500,f"{type(ex).__name__}: {ex}")

@app.get("/api/v15-8/phone-evidence")
def v158_evidence_api(req:Request,applied:str=Query("ALL")):
    need_login(req);_v158_setup()
    sql="""SELECT id,hospitality_master_id,business_name,recovered_phone,source_table,source_row_ref,
        match_method,confidence,applied,created_at FROM pi_hospitality_phone_evidence"""
    p={}
    if applied=="YES":sql+=" WHERE applied=TRUE"
    elif applied=="NO":sql+=" WHERE applied=FALSE"
    sql+=" ORDER BY confidence DESC,id DESC LIMIT 5000"
    with engine.connect() as c:
        rows=[dict(r._mapping) for r in c.execute(text(sql),p).fetchall()]
    return {"status":"ok","rows":rows}

@app.get("/hospitality-phone-recovery",response_class=HTMLResponse)
def v158_recovery_page(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)

    return HTMLResponse("""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Hospitality Phone Recovery</title>
<style>*{box-sizing:border-box}body{font-family:Arial;margin:0;background:#f4f7fb;color:#172437}header{background:#102235;color:white;padding:18px}.w{padding:18px;max-width:1700px;margin:auto}.bar{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px}.btn,a.btn{padding:8px 10px;border:0;border-radius:8px;background:#1677ff;color:#fff;text-decoration:none;font-weight:bold;cursor:pointer}.gray{background:#e9eef5!important;color:#203247!important}.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:8px;margin-bottom:12px}.k{background:#fff;border:1px solid #e2e8f0;border-radius:10px;padding:12px}.k b{display:block;font-size:22px}.msg{background:#fff8e8;border:1px solid #eed18f;padding:10px;border-radius:9px;margin-bottom:12px}.tablewrap{overflow:auto;background:#fff;border-radius:10px;max-height:65vh}table{width:100%;border-collapse:collapse;font-size:12px}th,td{padding:8px;border-bottom:1px solid #edf1f5;text-align:left;white-space:nowrap}th{position:sticky;top:0;background:#f8fafc}.good{color:#08734b;font-weight:bold}</style></head>
<body><header><b>Historical Hospitality Phone Recovery</b><br><small>Recover previously fetched phone numbers for WhatsApp marketing</small></header><div class=w>
<div class=bar><a class="btn gray" href="/ai-hospitality-master-final">← Hospitality Master</a><button class=btn onclick="recover()">Recover Historical Phone Numbers</button><a class="btn gray" href="/marketing-contacts-final">Marketing Contacts</a></div>
<div class=kpis><div class=k><b id=tables>0</b><span>TABLES SCANNED</span></div><div class=k><b id=scan>0</b><span>ROWS SCANNED</span></div><div class=k><b id=ev>0</b><span>PHONE EVIDENCE</span></div><div class=k><b id=apply>0</b><span>AUTO APPLIED</span></div><div class=k><b id=improved>0</b><span>HOSPITALITY RECORDS IMPROVED</span></div><div class=k><b id=withphone>0</b><span>MASTER ROWS WITH PHONE</span></div></div>
<div id=msg class=msg>This does not rerun the Hospitality Bot. It searches historical database tables/raw payloads for phone numbers already fetched earlier and links them back to the matching hospitality business.</div>
<div class=tablewrap><table><thead><tr><th>Table</th><th>Rows</th><th>Rows With Phone</th><th>Matched Phone Evidence</th><th>Auto Applied</th><th>Status</th></tr></thead><tbody id=rows></tbody></table></div>
</div><script>
async function recover(){msg.textContent='Scanning historical data for previously fetched Hospitality phones...';let r=await fetch('/api/v15-8/recover-historical-hospitality-phones',{method:'POST'}),d=await r.json();if(!r.ok){msg.textContent='ERROR: '+(d.detail||'Recovery failed');return}tables.textContent=d.candidate_tables||0;scan.textContent=d.rows_scanned||0;ev.textContent=d.phone_evidence_found||0;apply.textContent=d.auto_applied||0;improved.textContent=d.hospitality_records_improved||0;withphone.textContent=d.total_with_phone||0;rows.innerHTML=(d.table_stats||[]).map(x=>`<tr><td>${x.table||''}</td><td>${x.rows||0}</td><td>${x.rows_with_phone||0}</td><td>${x.evidence||0}</td><td>${x.applied||0}</td><td>${x.error||'OK'}</td></tr>`).join('');msg.textContent=`Recovery complete. ${d.phone_evidence_found||0} matched phone records found; ${d.auto_applied||0} high-confidence phones applied.`}
</script></body></html>""")

# ============================================================
# V15.8.1 BACKGROUND HOSPITALITY PHONE RECOVERY
# Fixes V15.8 long-running synchronous scan.
# Runs recovery in background with progress + indexed matching.
# ============================================================

def _v1581_setup():
    _v158_setup()
    with engine.begin() as c:
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_hospitality_phone_recovery_jobs(
            id BIGSERIAL PRIMARY KEY,
            status TEXT DEFAULT 'QUEUED',
            current_table TEXT,
            tables_total INTEGER DEFAULT 0,
            tables_done INTEGER DEFAULT 0,
            rows_scanned BIGINT DEFAULT 0,
            rows_with_phone BIGINT DEFAULT 0,
            phone_evidence BIGINT DEFAULT 0,
            auto_applied BIGINT DEFAULT 0,
            improved_records BIGINT DEFAULT 0,
            error_message TEXT,
            started_at TIMESTAMPTZ DEFAULT NOW(),
            finished_at TIMESTAMPTZ,
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )"""))

def _v1581_candidate_tables():
    preferred=[
        "ai_marketing_contacts",
        "ai_contacts",
        "hospitality_contacts",
        "ai_hospitality_contacts",
        "hospitality_prospects",
        "pi_hospitality_prospects",
        "hospitality_leads",
        "ai_demand_signals",
        "pi_contact_directory_v2",
        "pi_contacts"
    ]
    with engine.connect() as c:
        existing={r._mapping["table_name"] for r in c.execute(text(
            "SELECT table_name FROM information_schema.tables WHERE table_schema='public' AND table_type='BASE TABLE'"
        )).fetchall()}
    return [t for t in preferred if t in existing]

def _v1581_master_index():
    masters=_v158_master_rows()
    exact={}
    token_index={}
    for m in masters:
        name=_v158_norm(m.get("business_name"))
        if name:
            exact.setdefault(name,[]).append(m)
            for tok in [x for x in name.split() if len(x)>=5][:4]:
                token_index.setdefault(tok,[]).append(m)
    return masters,exact,token_index

def _v1581_candidates_from_blob(blob, exact, token_index):
    bnorm=_v158_norm(blob)
    cands={}
    # exact business name hit
    for name,rows in exact.items():
        if len(name)>=4 and name in bnorm:
            for m in rows:cands[m["id"]]=m
    # token narrowing only if exact did not find
    if not cands:
        for tok,rows in token_index.items():
            if tok in bnorm:
                for m in rows:cands[m["id"]]=m
    return list(cands.values())

def _v1581_update_job(job_id, **vals):
    if not vals:return
    sets=[];p={"id":job_id}
    for i,(k,v) in enumerate(vals.items()):
        key=f"v{i}";sets.append(f"{k}=:{key}");p[key]=v
    with engine.begin() as c:
        c.execute(text("UPDATE pi_hospitality_phone_recovery_jobs SET "+",".join(sets)+",updated_at=NOW() WHERE id=:id"),p)

def _v1581_worker(job_id):
    try:
        _v1581_setup()
        masters,exact,token_index=_v1581_master_index()
        tables=_v1581_candidate_tables()
        _v1581_update_job(job_id,status="RUNNING",tables_total=len(tables),tables_done=0)

        improved=set()
        total_rows=total_phone_rows=total_evidence=total_applied=0

        for ti,table in enumerate(tables,1):
            _v1581_update_job(job_id,current_table=table,tables_done=ti-1)

            with engine.connect() as c:
                rows=[dict(r._mapping) for r in c.execute(text(f'SELECT * FROM "{table}" LIMIT 20000')).fetchall()]

            for row in rows:
                total_rows+=1
                blob=_v158_row_blob(row)
                phones=_v158_valid_phones_from_text(blob)
                if not phones:
                    if total_rows % 500 == 0:
                        _v1581_update_job(job_id,rows_scanned=total_rows,rows_with_phone=total_phone_rows,
                                         phone_evidence=total_evidence,auto_applied=total_applied,
                                         improved_records=len(improved))
                    continue

                total_phone_rows+=1
                candidates=_v1581_candidates_from_blob(blob,exact,token_index)
                if not candidates:
                    continue

                best=None
                for m in candidates:
                    score,method=_v158_match_candidate(m,blob,table,row)
                    if score>=70 and (best is None or score>best[0]):
                        best=(score,method,m)
                if not best:
                    continue

                score,method,m=best
                row_ref=str(row.get("id") or row.get("contact_id") or row.get("lead_id") or row.get("prospect_id") or "")

                for ph in phones:
                    with engine.connect() as c:
                        exists=c.execute(text("""SELECT 1 FROM pi_hospitality_phone_evidence
                            WHERE hospitality_master_id=:mid AND recovered_phone=:ph AND source_table=:t LIMIT 1"""),
                            {"mid":m["id"],"ph":ph,"t":table}).first()
                    if exists:
                        continue

                    with engine.begin() as c:
                        c.execute(text("""INSERT INTO pi_hospitality_phone_evidence(
                            hospitality_master_id,business_name,recovered_phone,source_table,source_row_ref,
                            match_method,confidence,evidence_text,applied
                        ) VALUES(:mid,:bn,:ph,:t,:ref,:method,:conf,:ev,FALSE)"""),{
                            "mid":m["id"],"bn":m.get("business_name"),"ph":ph,"t":table,"ref":row_ref,
                            "method":method,"conf":score,"ev":blob[:2500]
                        })
                    total_evidence+=1

                    if score>=90:
                        with engine.begin() as c:
                            current=c.execute(text(
                                "SELECT all_phones,primary_phone FROM pi_ai_hospitality_master WHERE id=:id"
                            ),{"id":m["id"]}).fetchone()
                            if current:
                                cur=dict(current._mapping)
                                arr=cur.get("all_phones") or []
                                if isinstance(arr,str):
                                    try:arr=json.loads(arr)
                                    except:arr=[]
                                arr=[str(x) for x in arr if x]
                                if ph not in arr:arr.append(ph)
                                primary=cur.get("primary_phone") or ph
                                c.execute(text("""UPDATE pi_ai_hospitality_master
                                    SET primary_phone=:p,all_phones=CAST(:phones AS jsonb),
                                        contact_status='CONTACT_READY',updated_at=NOW()
                                    WHERE id=:id"""),{
                                    "p":primary,"phones":json.dumps(arr),"id":m["id"]
                                })
                                c.execute(text("""UPDATE pi_hospitality_phone_evidence
                                    SET applied=TRUE WHERE hospitality_master_id=:mid
                                    AND recovered_phone=:ph AND source_table=:t"""),
                                    {"mid":m["id"],"ph":ph,"t":table})
                        total_applied+=1
                        improved.add(m["id"])
                        try:
                            _v151_upsert_contact(
                                ph,
                                company=m.get("business_name"),
                                category=m.get("category"),
                                city=m.get("city"),
                                location=m.get("location"),
                                email=m.get("email"),
                                website=m.get("website"),
                                source="AI_HOSPITALITY",
                                source_detail=f"HISTORICAL_RECOVERY:{table}",
                                notes=f"Recovered historical phone. Confidence {score}. Match {method}."
                            )
                        except Exception:
                            pass

                if total_rows % 250 == 0:
                    _v1581_update_job(job_id,rows_scanned=total_rows,rows_with_phone=total_phone_rows,
                                     phone_evidence=total_evidence,auto_applied=total_applied,
                                     improved_records=len(improved))

            _v1581_update_job(job_id,tables_done=ti,rows_scanned=total_rows,rows_with_phone=total_phone_rows,
                             phone_evidence=total_evidence,auto_applied=total_applied,
                             improved_records=len(improved))

        _v1581_update_job(job_id,status="COMPLETED",current_table=None,tables_done=len(tables),
                         rows_scanned=total_rows,rows_with_phone=total_phone_rows,
                         phone_evidence=total_evidence,auto_applied=total_applied,
                         improved_records=len(improved),finished_at=datetime.now(timezone.utc))
    except Exception as ex:
        try:
            _v1581_update_job(job_id,status="FAILED",error_message=f"{type(ex).__name__}: {ex}",
                             finished_at=datetime.now(timezone.utc))
        except Exception:
            pass

@app.post("/api/v15-8-1/phone-recovery/start")
def v1581_start(req:Request, background_tasks:BackgroundTasks):
    need_login(req);_v1581_setup()
    with engine.begin() as c:
        existing=c.execute(text("""SELECT id FROM pi_hospitality_phone_recovery_jobs
            WHERE status IN ('QUEUED','RUNNING') ORDER BY id DESC LIMIT 1""")).first()
        if existing:
            return {"status":"already_running","job_id":existing._mapping["id"]}
        row=c.execute(text("""INSERT INTO pi_hospitality_phone_recovery_jobs(status,started_at,updated_at)
            VALUES('QUEUED',NOW(),NOW()) RETURNING id""")).first()
        job_id=row._mapping["id"]
    background_tasks.add_task(_v1581_worker,job_id)
    return {"status":"started","job_id":job_id}

@app.get("/api/v15-8-1/phone-recovery/status")
def v1581_status(req:Request):
    need_login(req);_v1581_setup()
    with engine.connect() as c:
        row=c.execute(text("""SELECT * FROM pi_hospitality_phone_recovery_jobs
            ORDER BY id DESC LIMIT 1""")).first()
    if not row:
        return {"status":"none"}
    d=dict(row._mapping)
    d["master_rows_with_phone"]=_v15_safe_count(
        "SELECT COUNT(*) FROM pi_ai_hospitality_master WHERE primary_phone IS NOT NULL AND primary_phone<>''"
    )
    d["contact_ready"]=_v15_safe_count(
        "SELECT COUNT(*) FROM pi_ai_hospitality_master WHERE contact_status='CONTACT_READY'"
    )
    return d

@app.get("/hospitality-phone-recovery-v2",response_class=HTMLResponse)
def v1581_page(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    return HTMLResponse("""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Hospitality Phone Recovery</title>
<style>*{box-sizing:border-box}body{font-family:Arial;margin:0;background:#f4f7fb;color:#172437}header{background:#102235;color:white;padding:18px}.w{max-width:1500px;margin:auto;padding:18px}.bar{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px}.btn,a.btn{padding:8px 10px;border:0;border-radius:8px;background:#1677ff;color:white;text-decoration:none;font-weight:bold;cursor:pointer}.gray{background:#e9eef5!important;color:#203247!important}.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:8px;margin-bottom:12px}.k{background:white;border:1px solid #e2e8f0;border-radius:10px;padding:12px}.k b{display:block;font-size:22px}.msg{background:#fff8e8;border:1px solid #eed18f;padding:11px;border-radius:9px}.progress{height:18px;background:#e8edf3;border-radius:9px;overflow:hidden;margin:12px 0}.fill{height:100%;background:#1677ff;width:0}.small{font-size:12px;color:#687789}</style></head>
<body><header><b>Hospitality Phone Recovery</b><br><small>Background recovery of phone numbers already fetched earlier</small></header><div class=w>
<div class=bar><a class="btn gray" href="/ai-hospitality-master-final">← Hospitality Master</a><button class=btn onclick="start()">Start Historical Phone Recovery</button><a class="btn gray" href="/marketing-contacts-final">Marketing Contacts</a></div>
<div class=kpis>
<div class=k><b id=tables>0</b><span>TABLES DONE / TOTAL</span></div>
<div class=k><b id=rows>0</b><span>ROWS SCANNED</span></div>
<div class=k><b id=evidence>0</b><span>PHONE EVIDENCE</span></div>
<div class=k><b id=applied>0</b><span>AUTO APPLIED</span></div>
<div class=k><b id=improved>0</b><span>RECORDS IMPROVED</span></div>
<div class=k><b id=masterphones>0</b><span>MASTER ROWS WITH PHONE</span></div>
</div>
<div class=progress><div id=fill class=fill></div></div>
<div id=msg class=msg>Ready. This runs in the background, so the page will not freeze or time out.</div>
<p class=small>Current table: <b id=current>-</b></p>
</div>
<script>
async function start(){msg.textContent='Starting recovery...';let r=await fetch('/api/v15-8-1/phone-recovery/start',{method:'POST'}),d=await r.json();if(!r.ok){msg.textContent='ERROR';return}msg.textContent=d.status==='already_running'?'Recovery is already running.':'Recovery started in background.';poll()}
async function poll(){try{let d=await(await fetch('/api/v15-8-1/phone-recovery/status')).json();if(d.status==='none')return;tables.textContent=(d.tables_done||0)+' / '+(d.tables_total||0);rows.textContent=d.rows_scanned||0;evidence.textContent=d.phone_evidence||0;applied.textContent=d.auto_applied||0;improved.textContent=d.improved_records||0;masterphones.textContent=d.master_rows_with_phone||0;current.textContent=d.current_table||'-';let pct=d.tables_total?Math.round((d.tables_done||0)*100/d.tables_total):0;fill.style.width=pct+'%';msg.textContent=d.status+(d.error_message?': '+d.error_message:'');if(['QUEUED','RUNNING'].includes(d.status))setTimeout(poll,3000)}catch(e){msg.textContent='Status error: '+e.message}}poll()
</script></body></html>""")

@app.middleware("http")
async def v1581_route_fix(request,call_next):
    if request.url.path=="/hospitality-phone-recovery":
        return RedirectResponse("/hospitality-phone-recovery-v2",status_code=307)
    return await call_next(request)

# ============================================================
# V16 FINAL DEAL INTELLIGENCE OS
# Single dashboard + Hospitality Contact Enrichment + WhatsApp-ready export
# ============================================================

def _v16_setup():
    _v156_setup()
    with engine.begin() as c:
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_hospitality_enrichment_jobs(
            id BIGSERIAL PRIMARY KEY,
            status TEXT DEFAULT 'QUEUED',
            category TEXT DEFAULT 'ALL',
            total_records INTEGER DEFAULT 0,
            processed INTEGER DEFAULT 0,
            mobile_found INTEGER DEFAULT 0,
            email_found INTEGER DEFAULT 0,
            website_checked INTEGER DEFAULT 0,
            google_places_checked INTEGER DEFAULT 0,
            failed INTEGER DEFAULT 0,
            current_business TEXT,
            error_message TEXT,
            started_at TIMESTAMPTZ DEFAULT NOW(),
            finished_at TIMESTAMPTZ,
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )"""))
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_hospitality_enrichment_evidence(
            id BIGSERIAL PRIMARY KEY,
            hospitality_master_id BIGINT NOT NULL,
            provider TEXT,
            source_url TEXT,
            recovered_phone TEXT,
            recovered_email TEXT,
            recovered_website TEXT,
            confidence INTEGER DEFAULT 0,
            evidence_text TEXT,
            applied BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMPTZ DEFAULT NOW()
        )"""))
        c.execute(text("CREATE INDEX IF NOT EXISTS idx_v16_evidence_master ON pi_hospitality_enrichment_evidence(hospitality_master_id)"))
        c.execute(text("CREATE INDEX IF NOT EXISTS idx_v16_evidence_phone ON pi_hospitality_enrichment_evidence(recovered_phone)"))

def _v16_valid_mobile(text_value):
    vals=[]
    for m in _re.findall(r"(?<!\d)(?:\+?91[\s\-\(\)]*)?([6-9]\d{9})(?!\d)", str(text_value or "")):
        ph=_v151_digits(m)
        if ph and ph not in vals:
            vals.append(ph)
    return vals

def _v16_all_indian_phones(text_value):
    mobiles=_v16_valid_mobile(text_value)
    landlines=[]
    # conservative Delhi/NCR-ish landline capture, stored separately only
    for m in _re.findall(r"(?<!\d)(?:\+?91[\s\-]?)?0?([1-9][0-9]{1,3})[\s\-]?([2-9][0-9]{5,7})(?!\d)", str(text_value or "")):
        num="0"+m[0]+m[1]
        if num not in landlines:
            landlines.append(num)
    return mobiles, landlines

def _v16_extract_email(text_value):
    found=[]
    for m in _re.findall(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}",str(text_value or ""),_re.I):
        e=m.lower()
        if e not in found and not e.endswith((".png",".jpg",".jpeg",".webp")):
            found.append(e)
    return found

def _v16_http_get(url, timeout=15):
    if not url or not str(url).startswith(("http://","https://")):
        return None, None
    try:
        import urllib.request
        req=urllib.request.Request(str(url),headers={"User-Agent":"Mozilla/5.0 PropertyIntelligenceBot/1.0"})
        with urllib.request.urlopen(req,timeout=timeout) as resp:
            body=resp.read(1500000)
            ctype=resp.headers.get("Content-Type","")
        return body.decode("utf-8","ignore"),ctype
    except Exception:
        return None,None

def _v16_google_places_lookup(business, location):
    key=os.getenv("GOOGLE_PLACES_API_KEY") or os.getenv("GOOGLE_MAPS_API_KEY")
    if not key:
        return None
    try:
        import urllib.request, urllib.error
        q=(str(business or "")+" "+str(location or "")).strip()
        payload=json.dumps({"textQuery":q,"maxResultCount":3}).encode("utf-8")
        req=urllib.request.Request(
            "https://places.googleapis.com/v1/places:searchText",
            data=payload,
            method="POST",
            headers={
                "Content-Type":"application/json",
                "X-Goog-Api-Key":key,
                "X-Goog-FieldMask":"places.displayName,places.formattedAddress,places.nationalPhoneNumber,places.internationalPhoneNumber,places.websiteUri,places.googleMapsUri"
            }
        )
        with urllib.request.urlopen(req,timeout=20) as resp:
            d=json.loads(resp.read().decode("utf-8","ignore"))
        places=d.get("places") or []
        if not places:
            return None
        return places[0]
    except Exception:
        return None

def _v16_apply_contact(master_id, phone=None, email=None, website=None, provider=None, source_url=None, confidence=0, evidence=None):
    _v16_setup()
    applied=False
    with engine.begin() as c:
        row=c.execute(text("""SELECT * FROM pi_ai_hospitality_master WHERE id=:id"""),{"id":master_id}).first()
        if not row:
            return False
        m=dict(row._mapping)

        phones=m.get("all_phones") or []
        if isinstance(phones,str):
            try: phones=json.loads(phones)
            except: phones=[]
        phones=[str(x) for x in phones if x]

        if phone and phone not in phones:
            phones.append(phone)

        primary=m.get("primary_phone") or (phone if phone else None)
        new_email=m.get("email") or email
        new_website=m.get("website") or website
        contact_status="CONTACT_READY" if primary or new_email else "NEEDS_ENRICHMENT"
        enrichment_status="ENRICHED" if primary else ("PARTIAL" if new_email or new_website else "NOT_FOUND")

        c.execute(text("""UPDATE pi_ai_hospitality_master
            SET primary_phone=:p,
                all_phones=CAST(:phones AS jsonb),
                email=:email,
                website=:website,
                contact_status=:cs,
                enrichment_status=:es,
                updated_at=NOW()
            WHERE id=:id"""),{
                "p":primary,"phones":json.dumps(phones),"email":new_email,"website":new_website,
                "cs":contact_status,"es":enrichment_status,"id":master_id
            })

        c.execute(text("""INSERT INTO pi_hospitality_enrichment_evidence(
            hospitality_master_id,provider,source_url,recovered_phone,recovered_email,
            recovered_website,confidence,evidence_text,applied
        ) VALUES(:mid,:provider,:source,:ph,:email,:website,:conf,:ev,:applied)"""),{
            "mid":master_id,"provider":provider,"source":source_url,"ph":phone,"email":email,
            "website":website,"conf":confidence,"ev":(evidence or "")[:3000],"applied":bool(phone or email or website)
        })
        applied=True

    if phone:
        try:
            _v151_upsert_contact(
                phone,
                company=m.get("business_name"),
                category=m.get("category"),
                city=m.get("city"),
                location=m.get("location"),
                email=(m.get("email") or email),
                website=(m.get("website") or website),
                source="AI_HOSPITALITY",
                source_detail=f"V16_ENRICHMENT:{provider or 'UNKNOWN'}",
                notes="Recovered by Hospitality Contact Enrichment"
            )
        except Exception:
            pass
    return applied

def _v16_enrich_one(m):
    result={"phone":None,"email":None,"website":None,"provider":None,"source_url":None,"confidence":0,"evidence":"","website_checked":0,"places_checked":0}

    # 1) Existing raw payload/source fields first.
    raw=" ".join([
        str(m.get("raw_text") or ""),
        str(m.get("source_url") or ""),
        str(m.get("website") or ""),
        str(m.get("location") or ""),
        str(m.get("business_name") or "")
    ])
    phones=_v16_valid_mobile(raw)
    emails=_v16_extract_email(raw)
    if phones:
        result.update(phone=phones[0],email=(emails[0] if emails else None),provider="EXISTING_RAW",source_url=m.get("source_url"),confidence=95,evidence=raw[:3000])
        return result
    if emails:
        result.update(email=emails[0],provider="EXISTING_RAW",source_url=m.get("source_url"),confidence=80,evidence=raw[:3000])

    # 2) Website/source page scrape.
    candidates=[]
    for u in [m.get("website"),m.get("source_url")]:
        if u and str(u).startswith(("http://","https://")) and u not in candidates:
            candidates.append(u)

    for u in candidates[:2]:
        html,_=_v16_http_get(u)
        result["website_checked"]+=1
        if not html:
            continue
        p=_v16_valid_mobile(html)
        e=_v16_extract_email(html)
        if p:
            result.update(phone=p[0],email=(e[0] if e else result.get("email")),website=(m.get("website") or u),
                          provider="WEBSITE",source_url=u,confidence=92,evidence=html[:3000])
            return result
        if e and not result.get("email"):
            result.update(email=e[0],website=(m.get("website") or u),provider="WEBSITE",source_url=u,confidence=78,evidence=html[:3000])

    # 3) Google Places Text Search if API key is configured.
    place=_v16_google_places_lookup(m.get("business_name"),m.get("location"))
    result["places_checked"]+=1 if (os.getenv("GOOGLE_PLACES_API_KEY") or os.getenv("GOOGLE_MAPS_API_KEY")) else 0
    if place:
        phs=_v16_valid_mobile(place.get("nationalPhoneNumber") or place.get("internationalPhoneNumber") or "")
        website=place.get("websiteUri")
        gm=place.get("googleMapsUri")
        ev=json.dumps(place,default=str)
        if phs:
            result.update(phone=phs[0],website=(website or result.get("website")),provider="GOOGLE_PLACES",
                          source_url=(gm or website),confidence=98,evidence=ev[:3000])
            return result
        if website and not result.get("website"):
            result.update(website=website,provider="GOOGLE_PLACES",source_url=(gm or website),confidence=75,evidence=ev[:3000])

    return result

def _v16_job_update(job_id, **vals):
    if not vals:return
    sets=[];p={"id":job_id}
    for i,(k,v) in enumerate(vals.items()):
        kk=f"v{i}";sets.append(f"{k}=:{kk}");p[kk]=v
    with engine.begin() as c:
        c.execute(text("UPDATE pi_hospitality_enrichment_jobs SET "+",".join(sets)+",updated_at=NOW() WHERE id=:id"),p)

def _v16_enrichment_worker(job_id, category="ALL", limit=1000):
    try:
        _v16_setup()
        wh=["(primary_phone IS NULL OR primary_phone='')"]
        p={}
        if category!="ALL":
            wh.append("category=:cat");p["cat"]=category
        sql="""SELECT * FROM pi_ai_hospitality_master WHERE """+" AND ".join(wh)+" ORDER BY id LIMIT :lim"
        p["lim"]=int(limit)
        with engine.connect() as c:
            rows=[dict(r._mapping) for r in c.execute(text(sql),p).fetchall()]

        _v16_job_update(job_id,status="RUNNING",total_records=len(rows),processed=0)

        processed=mobile_found=email_found=website_checked=places_checked=failed=0
        for m in rows:
            _v16_job_update(job_id,current_business=str(m.get("business_name") or "")[:200])
            try:
                r=_v16_enrich_one(m)
                website_checked+=r.get("website_checked",0)
                places_checked+=r.get("places_checked",0)
                if r.get("phone"):
                    mobile_found+=1
                if r.get("email"):
                    email_found+=1
                if r.get("phone") or r.get("email") or r.get("website"):
                    _v16_apply_contact(
                        m["id"],r.get("phone"),r.get("email"),r.get("website"),
                        r.get("provider"),r.get("source_url"),r.get("confidence",0),r.get("evidence")
                    )
            except Exception:
                failed+=1
            processed+=1
            if processed % 10 == 0:
                _v16_job_update(job_id,processed=processed,mobile_found=mobile_found,email_found=email_found,
                                website_checked=website_checked,google_places_checked=places_checked,failed=failed)

        _v16_job_update(job_id,status="COMPLETED",processed=processed,mobile_found=mobile_found,email_found=email_found,
                        website_checked=website_checked,google_places_checked=places_checked,failed=failed,
                        current_business=None,finished_at=datetime.now(timezone.utc))
    except Exception as ex:
        _v16_job_update(job_id,status="FAILED",error_message=f"{type(ex).__name__}: {ex}",
                        finished_at=datetime.now(timezone.utc))

@app.post("/api/v16/hospitality-enrichment/start")
async def v16_start_enrichment(req:Request, background_tasks:BackgroundTasks):
    need_login(req);_v16_setup()
    body=await req.json()
    category=str(body.get("category") or "ALL").upper()
    limit=max(1,min(int(body.get("limit") or 1000),5000))
    with engine.begin() as c:
        c.execute(text("""
            UPDATE pi_hospitality_enrichment_jobs
            SET status='FAILED',
                error_message=COALESCE(error_message,'') || ' | Auto-reset stale job before new enrichment run',
                finished_at=NOW(),
                updated_at=NOW(),
                current_business=NULL
            WHERE status IN ('QUEUED','RUNNING')
              AND updated_at < NOW() - INTERVAL '15 minutes'
        """))
        active=c.execute(text("""SELECT id FROM pi_hospitality_enrichment_jobs
            WHERE status IN ('QUEUED','RUNNING') ORDER BY id DESC LIMIT 1""")).first()
        if active:
            return {"status":"already_running","job_id":active._mapping["id"]}
        row=c.execute(text("""INSERT INTO pi_hospitality_enrichment_jobs(status,category,started_at,updated_at)
            VALUES('QUEUED',:cat,NOW(),NOW()) RETURNING id"""),{"cat":category}).first()
        jid=row._mapping["id"]
    background_tasks.add_task(_v16_enrichment_worker,jid,category,limit)
    return {"status":"started","job_id":jid}

@app.get("/api/v16/hospitality-enrichment/status")
def v16_enrichment_status(req:Request):
    need_login(req);_v16_setup()
    with engine.connect() as c:
        row=c.execute(text("SELECT * FROM pi_hospitality_enrichment_jobs ORDER BY id DESC LIMIT 1")).first()
    if not row:return {"status":"none"}
    d=dict(row._mapping)
    d["master_total"]=_v15_safe_count("SELECT COUNT(*) FROM pi_ai_hospitality_master")
    d["master_with_mobile"]=_v15_safe_count("SELECT COUNT(*) FROM pi_ai_hospitality_master WHERE primary_phone IS NOT NULL AND primary_phone<>''")
    d["needs_enrichment"]=_v15_safe_count("SELECT COUNT(*) FROM pi_ai_hospitality_master WHERE primary_phone IS NULL OR primary_phone=''")
    d["whatsapp_ready"]=_v15_safe_count("""SELECT COUNT(*) FROM pi_marketing_contacts
        WHERE source ILIKE '%AI_HOSPITALITY%' AND primary_phone IS NOT NULL
        AND primary_phone<>'' AND whatsapp_status IN ('READY','NOT_CONTACTED') AND COALESCE(opt_out,FALSE)=FALSE""")
    d["google_places_configured"]=bool(os.getenv("GOOGLE_PLACES_API_KEY") or os.getenv("GOOGLE_MAPS_API_KEY"))
    return d

@app.post("/api/v16/hospitality/{hid}/verify")
async def v16_verify_hospitality(hid:int, req:Request):
    need_login(req)
    body=await req.json()
    status=str(body.get("verified_status") or "VERIFIED").upper()
    if status not in {"VERIFIED","UNVERIFIED"}: status="UNVERIFIED"
    with engine.begin() as c:
        c.execute(text("UPDATE pi_ai_hospitality_master SET verified_status=:s,updated_at=NOW() WHERE id=:id"),{"s":status,"id":hid})
    return {"status":"ok"}

@app.get("/api/v16/whatsapp-ready.csv")
def v16_whatsapp_ready_csv(req:Request,category:str=Query("ALL"),verified:str=Query("ALL")):
    need_login(req)
    wh=["primary_phone IS NOT NULL","primary_phone<>''"]
    p={}
    if category!="ALL":
        wh.append("category=:cat");p["cat"]=category
    if verified!="ALL":
        wh.append("verified_status=:ver");p["ver"]=verified
    sql="""SELECT business_name,category,city,location,contact_name,primary_phone,email,website,source_url,verified_status
        FROM pi_ai_hospitality_master WHERE """+" AND ".join(wh)+" ORDER BY category,business_name"
    with engine.connect() as c:
        rows=[dict(r._mapping) for r in c.execute(text(sql),p).fetchall()]
    import io,csv
    s=io.StringIO()
    w=csv.writer(s)
    w.writerow(["Business Name","Category","City","Location","Contact Name","Mobile","Email","Website","Source URL","Verified"])
    for r in rows:
        w.writerow([r.get("business_name"),r.get("category"),r.get("city"),r.get("location"),r.get("contact_name"),
                    r.get("primary_phone"),r.get("email"),r.get("website"),r.get("source_url"),r.get("verified_status")])
    return Response(content=s.getvalue(),media_type="text/csv",headers={"Content-Disposition":"attachment; filename=hospitality_whatsapp_ready.csv"})

@app.get("/hospitality-enrichment",response_class=HTMLResponse)
def v16_hospitality_enrichment(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    return HTMLResponse("""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Hospitality Contact Enrichment</title>
<style>*{box-sizing:border-box}body{font-family:Arial;margin:0;background:#f4f7fb;color:#172437}header{background:#102235;color:#fff;padding:18px}.w{max-width:1500px;margin:auto;padding:18px}.bar{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px}.btn,a.btn{padding:8px 10px;border:0;border-radius:8px;background:#1677ff;color:#fff;text-decoration:none;font-weight:bold;cursor:pointer}.gray{background:#e9eef5!important;color:#203247!important}.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:8px;margin-bottom:12px}.k{background:#fff;border:1px solid #e2e8f0;border-radius:10px;padding:12px}.k b{display:block;font-size:22px}.msg{background:#fff8e8;border:1px solid #eed18f;padding:11px;border-radius:9px}.progress{height:18px;background:#e8edf3;border-radius:9px;overflow:hidden;margin:12px 0}.fill{height:100%;background:#1677ff;width:0}select,input{padding:8px;border:1px solid #ccd6e2;border-radius:7px}.small{font-size:12px;color:#687789}</style></head>
<body><header><b>Hospitality Contact Enrichment</b><br><small>Phone-first enrichment for WhatsApp marketing</small></header><div class=w>
<div class=bar><a class="btn gray" href="/final-dashboard">← Final Dashboard</a>
<select id=cat><option>ALL</option><option>CAFE</option><option>RESTAURANT</option><option>BANQUET</option><option>HOTEL</option><option>GUEST_HOUSE</option><option>LOUNGE</option><option>CLUB</option><option>BAR</option><option>FARMHOUSE</option></select>
<input id=limit type=number min=1 max=5000 value=1000>
<button class=btn onclick="start()">Start Enrichment</button>
<a class="btn gray" href="/api/v16/whatsapp-ready.csv">Export WhatsApp CSV</a></div>
<div class=kpis>
<div class=k><b id=total>0</b><span>MASTER BUSINESSES</span></div>
<div class=k><b id=mobiles>0</b><span>WITH MOBILE</span></div>
<div class=k><b id=need>0</b><span>NEEDS MOBILE</span></div>
<div class=k><b id=processed>0</b><span>JOB PROCESSED</span></div>
<div class=k><b id=found>0</b><span>MOBILES FOUND THIS JOB</span></div>
<div class=k><b id=wa>0</b><span>WHATSAPP READY</span></div>
</div>
<div class=progress><div id=fill class=fill></div></div>
<div id=msg class=msg>Ready. Existing 1,298 Hospitality businesses are preserved. Enrichment only works on records missing a valid mobile.</div>
<p class=small>Current business: <b id=current>-</b> · Google Places API configured: <b id=gp>NO</b></p>
</div>
<script>
async function start(){msg.textContent='Starting enrichment...';let r=await fetch('/api/v16/hospitality-enrichment/start',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({category:cat.value,limit:Number(limit.value||1000)})}),d=await r.json();msg.textContent=d.status==='already_running'?'An enrichment job is already running.':'Enrichment started.';poll()}
async function poll(){try{let d=await(await fetch('/api/v16/hospitality-enrichment/status')).json();if(d.status==='none')return;total.textContent=d.master_total||0;mobiles.textContent=d.master_with_mobile||0;need.textContent=d.needs_enrichment||0;processed.textContent=d.processed||0;found.textContent=d.mobile_found||0;wa.textContent=d.whatsapp_ready||0;current.textContent=d.current_business||'-';gp.textContent=d.google_places_configured?'YES':'NO';let pct=d.total_records?Math.round((d.processed||0)*100/d.total_records):0;fill.style.width=pct+'%';msg.textContent=d.status+(d.error_message?': '+d.error_message:'');if(['QUEUED','RUNNING'].includes(d.status))setTimeout(poll,3000)}catch(e){msg.textContent='Status error: '+e.message}}poll()
</script></body></html>""")

@app.get("/final-dashboard",response_class=HTMLResponse)
def v16_final_dashboard(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    return HTMLResponse("""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>AI Deal Intelligence OS</title>
<style>*{box-sizing:border-box}body{font-family:Arial;margin:0;background:#f4f7fb;color:#172437}header{background:#102235;color:white;padding:22px}.w{max-width:1500px;margin:auto;padding:20px}.section{margin-bottom:22px}.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:12px}.card{background:#fff;border:1px solid #e2e8f0;border-radius:12px;padding:16px;text-decoration:none;color:#172437;min-height:110px;display:block}.card b{font-size:16px}.card p{font-size:12px;color:#687789}.tag{display:inline-block;margin-top:7px;padding:3px 7px;border-radius:10px;background:#edf4ff;font-size:10px}.primary{border:2px solid #1677ff}.danger{border-color:#e5e7eb}.small{color:#9db0c5;font-size:12px}</style></head>
<body><header><b>AI Deal Intelligence OS</b><br><span class=small>Final Team Dashboard · Property · Requirements · Hospitality · Retail · WhatsApp Marketing</span></header><div class=w>
<div class=section><h2>Daily Operations</h2><div class=grid>
<a class="card primary" href="/v14-property-form"><b>Add Property Manually</b><p>Fresh inventory with required fields and verification status.</p><span class=tag>DAILY</span></a>
<a class="card primary" href="/v14-requirement-form"><b>Add Requirement Manually</b><p>Confirmed requirement entry for accurate matching.</p><span class=tag>DAILY</span></a>
<a class="card primary" href="/v14-matcher"><b>Property Matcher</b><p>Match only fresh/manual verified inventory.</p><span class=tag>DAILY</span></a>
<a class="card" href="/v14-inventory"><b>Fresh Inventory Database</b><p>Search and review clean inventory separately.</p></a>
</div></div>

<div class=section><h2>AI Hospitality</h2><div class=grid>
<a class="card primary" href="/ai-hospitality-master-final"><b>Hospitality Master</b><p>1,298 cafes, restaurants, banquets, hotels, guest houses, lounges, clubs and bars.</p><span class=tag>MASTER</span></a>
<a class="card primary" href="/hospitality-enrichment"><b>Find Missing Contact Numbers</b><p>Phone-first enrichment for WhatsApp marketing. Existing records are preserved.</p><span class=tag>ENRICH</span></a>
<a class="card" href="/marketing-contacts-final"><b>Marketing Contacts</b><p>Segregated contact database with verification and WhatsApp status.</p></a>
<a class="card" href="/api/v16/whatsapp-ready.csv"><b>Export WhatsApp Contacts</b><p>Download hospitality records that contain usable mobile numbers.</p><span class=tag>CSV</span></a>
</div></div>

<div class=section><h2>AI Demand</h2><div class=grid>
<a class="card" href="/retail-expansion"><b>AI Retail Expansion</b><p>Retail demand signals and expansion intelligence.</p></a>
<a class="card" href="/requirements-match-center"><b>Requirements Centre</b><p>Manual and AI requirements, kept separate for verification.</p></a>
<a class="card" href="/requirements-entry?division=RETAIL"><b>Add Retail Requirement</b><p>Confirmed retail requirement entry.</p></a>
<a class="card" href="/requirements-entry?division=HOSPITALITY"><b>Add Hospitality Requirement</b><p>Confirmed hospitality requirement for matching.</p></a>
</div></div>

<div class=section><h2>Database & Capture</h2><div class=grid>
<a class="card" href="/capture-intelligence"><b>Camera / Screenshot / Newspaper / PDF</b><p>Capture property information from images, handwritten notes, WhatsApp and documents.</p></a>
<a class="card" href="/property-database"><b>Full Property Database</b><p>Legacy/master property archive. Not used by the fresh V14 matcher.</p></a>
<a class="card" href="/contacts-directory"><b>Owner / Broker Contacts</b><p>Property contact verification only. Separate from marketing contacts.</p></a>
<a class="card" href="/data-doctor"><b>Data Doctor</b><p>Admin reconciliation and database health.</p><span class=tag>ADMIN</span></a>
</div></div>

<div class=section><h2>Goa</h2><div class=grid>
<a class="card" href="/goa-property"><b>Goa Property</b><p>Dedicated Goa workflow when enabled.</p></a>
<a class="card" href="/goa-requirement"><b>Goa Requirement</b><p>Goa demand entry when enabled.</p></a>
<a class="card" href="/goa-matcher"><b>Goa Matcher</b><p>Separate Goa matching workflow.</p></a>
</div></div>
</div></body></html>""")

@app.middleware("http")
async def v16_final_navigation(request,call_next):
    if request.url.path in {"/workspace","/v15-dashboard","/simple-dashboard","/team-workspace-clean"}:
        return RedirectResponse("/final-dashboard",status_code=307)
    response=await call_next(request)
    if request.url.path.startswith(("/final-dashboard","/hospitality-enrichment","/ai-hospitality-master-final","/marketing-contacts-final")):
        response.headers["Cache-Control"]="no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"]="no-cache"
        response.headers["Expires"]="0"
    return response

# ============================================================
# V16.1 FINAL BOT CONTROL DASHBOARD
# Adds one-click Hospitality Bot + Retail Bot run controls
# to the single final dashboard.
# Existing V4 bot endpoints are reused.
# ============================================================

@app.get("/api/v16-1/bots/status")
def v161_bot_status(req:Request):
    need_login(req)
    try:
        with engine.connect() as c:
            rows=[dict(r._mapping) for r in c.execute(text("""
                SELECT * FROM ai_bot_runs
                WHERE division IN ('HOSPITALITY','RETAIL','DEMAND')
                   OR bot_name ILIKE '%Hospitality%'
                   OR bot_name ILIKE '%Retail%'
                ORDER BY id DESC
                LIMIT 30
            """)).fetchall()]
        return {"status":"ok","rows":rows}
    except Exception as ex:
        return {"status":"error","message":f"{type(ex).__name__}: {ex}","rows":[]}

@app.get("/final-dashboard-v2",response_class=HTMLResponse)
def v161_final_dashboard_v2(req:Request):
    role=page_role_or_redirect(req)
    if not role:
        return RedirectResponse("/login",303)

    return HTMLResponse("""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>AI Deal Intelligence OS</title>
<style>
*{box-sizing:border-box}
body{font-family:Arial;margin:0;background:#f4f7fb;color:#172437}
header{background:#102235;color:#fff;padding:22px}
.wrap{max-width:1550px;margin:auto;padding:20px}
.section{margin-bottom:24px}
.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:12px}
.card{background:#fff;border:1px solid #e2e8f0;border-radius:12px;padding:16px;text-decoration:none;color:#172437;min-height:110px;display:block}
.card b{font-size:16px}.card p{font-size:12px;color:#687789;line-height:1.4}
.primary{border:2px solid #1677ff}
.bot{border:2px solid #14a673}
.tag{display:inline-block;padding:3px 7px;border-radius:10px;background:#edf4ff;font-size:10px;margin-top:6px}
.btn{display:inline-block;padding:9px 11px;border:0;border-radius:8px;background:#1677ff;color:white;font-weight:bold;cursor:pointer;text-decoration:none}
.green{background:#08734b}.gray{background:#e9eef5;color:#203247}
.botbar{display:flex;gap:8px;flex-wrap:wrap;margin-top:10px}
.status{margin-top:10px;padding:8px;border-radius:8px;background:#f6f8fb;font-size:12px;min-height:30px}
table{width:100%;border-collapse:collapse;background:white;font-size:12px}
th,td{padding:8px;border-bottom:1px solid #edf1f5;text-align:left}
.tablewrap{overflow:auto;background:white;border-radius:10px;border:1px solid #e2e8f0}
.small{color:#9db0c5;font-size:12px}
</style>
</head>
<body>
<header>
<b>AI Deal Intelligence OS</b><br>
<span class="small">Final Team Dashboard · Property · Hospitality · Retail · Requirements · WhatsApp Marketing</span>
</header>

<div class="wrap">

<div class="section">
<h2>Run AI Bots</h2>
<div class="grid">

<div class="card bot">
<b>◆ Hospitality Bot</b>
<p>Fetch fresh Cafe, Restaurant, Banquet, Wedding Venue, Hotel, Guest House, Lounge, Club and Bar business contacts/signals.</p>
<div class="botbar">
<button class="btn green" onclick="runBot('hospitality')">▶ Run Hospitality Bot</button>
<a class="btn gray" href="/ai-hospitality-master-final">Open Hospitality Data</a>
</div>
<div id="hospitalityMsg" class="status">Ready to run.</div>
</div>

<div class="card bot">
<b>◈ Retail Expansion Bot</b>
<p>Run fresh retail expansion discovery and public retail leasing requirement discovery using the existing Retail Bot.</p>
<div class="botbar">
<button class="btn green" onclick="runBot('retail')">▶ Run Retail Bot</button>
<a class="btn gray" href="/retail-expansion">Open Retail Results</a>
</div>
<div id="retailMsg" class="status">Ready to run.</div>
</div>

<div class="card">
<b>Bot Run History</b>
<p>Check whether Hospitality and Retail bots are RUNNING, COMPLETED or FAILED.</p>
<div class="botbar">
<button class="btn" onclick="loadStatus()">Refresh Bot Status</button>
</div>
<div id="overallMsg" class="status">Status will refresh automatically.</div>
</div>

<div class="card">
<b>Hospitality Contact Enrichment</b>
<p>Use this after discovery to find missing mobile numbers for WhatsApp marketing without rerunning the full bot.</p>
<div class="botbar">
<a class="btn" href="/hospitality-enrichment">Find Missing Contacts</a>
</div>
</div>

</div>
</div>

<div class="section">
<h2>Daily Property Operations</h2>
<div class="grid">
<a class="card primary" href="/v14-property-form"><b>Add Property Manually</b><p>Add fresh structured inventory.</p><span class=tag>DAILY</span></a>
<a class="card primary" href="/v14-requirement-form"><b>Add Requirement Manually</b><p>Add confirmed client requirement.</p><span class=tag>DAILY</span></a>
<a class="card primary" href="/v14-matcher"><b>Property Matcher</b><p>Run matching against fresh verified inventory.</p><span class=tag>DAILY</span></a>
<a class="card" href="/v14-inventory"><b>Fresh Inventory Database</b><p>Search and manage fresh inventory.</p></a>
</div>
</div>

<div class="section">
<h2>Hospitality & Marketing</h2>
<div class="grid">
<a class="card" href="/ai-hospitality-master-final"><b>Hospitality Master</b><p>All recovered/generated Hospitality businesses, segregated by category.</p></a>
<a class="card" href="/hospitality-enrichment"><b>Find Missing Contact Numbers</b><p>Phone-first enrichment for WhatsApp marketing.</p></a>
<a class="card" href="/marketing-contacts-final"><b>Marketing Contacts</b><p>Segregated contacts with verification and WhatsApp status.</p></a>
<a class="card" href="/api/v16/whatsapp-ready.csv"><b>Export WhatsApp Contacts</b><p>Download mobile-ready Hospitality contacts.</p></a>
</div>
</div>

<div class="section">
<h2>Retail & Requirements</h2>
<div class="grid">
<a class="card" href="/retail-expansion"><b>AI Retail Expansion</b><p>Retail expansion signals and discovered opportunities.</p></a>
<a class="card" href="/requirements-match-center"><b>Requirements Centre</b><p>AI-generated and manually confirmed requirements.</p></a>
<a class="card" href="/requirements-entry?division=RETAIL"><b>Add Retail Requirement</b><p>Enter a confirmed retail requirement.</p></a>
<a class="card" href="/requirements-entry?division=HOSPITALITY"><b>Add Hospitality Requirement</b><p>Enter a confirmed Hospitality requirement.</p></a>
</div>
</div>

<div class="section">
<h2>Database & Capture</h2>
<div class="grid">
<a class="card" href="/capture-intelligence"><b>Capture Property</b><p>Camera, screenshot, handwritten note, newspaper, magazine and PDF.</p></a>
<a class="card" href="/property-database"><b>Full Property Database</b><p>Master/legacy property archive.</p></a>
<a class="card" href="/contacts-directory"><b>Owner / Broker Contacts</b><p>Property contacts only, separate from marketing contacts.</p></a>
<a class="card" href="/data-doctor"><b>Data Doctor</b><p>Admin reconciliation and database health.</p><span class=tag>ADMIN</span></a>
</div>
</div>

<div class="section">
<h2>Recent Bot Runs</h2>
<div class="tablewrap">
<table>
<thead><tr><th>Bot</th><th>Division</th><th>Status</th><th>Summary</th><th>Run ID</th><th>Started / Created</th></tr></thead>
<tbody id="botRows"><tr><td colspan="6">Loading...</td></tr></tbody>
</table>
</div>
</div>

</div>

<script>
const E=x=>String(x??'').replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));

async function runBot(type){
  const isHosp=type==='hospitality';
  const box=document.getElementById(isHosp?'hospitalityMsg':'retailMsg');
  const url=isHosp?'/api/v4/hospitality-bot/start':'/api/v4/retail-bot/start';
  box.textContent='Starting '+(isHosp?'Hospitality':'Retail')+' Bot...';
  try{
    const r=await fetch(url,{method:'POST'});
    let d={};
    try{d=await r.json()}catch(e){}
    if(!r.ok) throw Error(d.detail||d.message||('HTTP '+r.status));
    let ids=[];
    if(d.run_id)ids.push(d.run_id);
    if(d.requirement_run_id)ids.push(d.requirement_run_id);
    box.textContent='Started successfully'+(ids.length?' · Run ID: '+ids.join(', '):'')+'. Running in background.';
    setTimeout(loadStatus,1000);
  }catch(e){
    box.textContent='ERROR: '+e.message;
  }
}

async function loadStatus(){
  try{
    const r=await fetch('/api/v16-1/bots/status');
    const d=await r.json();
    const rows=d.rows||[];
    document.getElementById('botRows').innerHTML=rows.map(x=>`<tr>
      <td><b>${E(x.bot_name||'')}</b></td>
      <td>${E(x.division||'')}</td>
      <td>${E(x.status||'')}</td>
      <td>${E(x.summary||x.output_summary||'')}</td>
      <td>${E(x.run_id||'')}</td>
      <td>${E(x.started_at||x.created_at||'')}</td>
    </tr>`).join('')||'<tr><td colspan="6">No Hospitality/Retail bot runs found.</td></tr>';

    const running=rows.filter(x=>String(x.status||'').toUpperCase()==='RUNNING').length;
    document.getElementById('overallMsg').textContent=running?running+' bot run(s) currently RUNNING.':'No Hospitality/Retail bots currently running.';
  }catch(e){
    document.getElementById('overallMsg').textContent='Status error: '+e.message;
  }
}
loadStatus();
setInterval(loadStatus,10000);
</script>
</body>
</html>""")

@app.middleware("http")
async def v161_dashboard_router(request,call_next):
    if request.url.path in {"/workspace","/final-dashboard","/v15-dashboard","/simple-dashboard","/team-workspace-clean"}:
        return RedirectResponse("/final-dashboard-v2",status_code=307)
    response=await call_next(request)
    if request.url.path.startswith("/final-dashboard-v2"):
        response.headers["Cache-Control"]="no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"]="no-cache"
        response.headers["Expires"]="0"
    return response

# ============================================================
# V16.2 ADMIN TOOLS + PHONE CONTACT UPLOAD
# Fixes Admin Data Tools and adds VCF/CSV/XLSX phone-contact import
# for WhatsApp marketing.
# ============================================================

def _v162_setup():
    _v151_setup()
    with engine.begin() as c:
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_phone_contact_uploads(
            id BIGSERIAL PRIMARY KEY,
            upload_id TEXT UNIQUE NOT NULL,
            filename TEXT,
            default_category TEXT,
            rows_read INTEGER DEFAULT 0,
            valid_mobile_rows INTEGER DEFAULT 0,
            contacts_created INTEGER DEFAULT 0,
            contacts_updated INTEGER DEFAULT 0,
            duplicates INTEGER DEFAULT 0,
            invalid_rows INTEGER DEFAULT 0,
            uploaded_by TEXT,
            created_at TIMESTAMPTZ DEFAULT NOW()
        )"""))
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_phone_contact_upload_evidence(
            id BIGSERIAL PRIMARY KEY,
            upload_id TEXT NOT NULL,
            contact_name TEXT,
            raw_phone TEXT,
            normalized_phone TEXT,
            email TEXT,
            company_brand TEXT,
            category TEXT,
            city TEXT,
            location TEXT,
            status TEXT,
            reason TEXT,
            created_at TIMESTAMPTZ DEFAULT NOW()
        )"""))

def _v162_mobile(v):
    if v is None:
        return None
    digits=_re.sub(r"\D","",str(v))
    if len(digits)>10 and digits.startswith("91"):
        digits=digits[-10:]
    if len(digits)==10 and digits[0] in "6789":
        return digits
    return None

def _v162_split_phones(v):
    if v is None:return []
    vals=[]
    if isinstance(v,(list,tuple)):
        chunks=v
    else:
        chunks=_re.split(r"[,;/|]+",str(v))
    for x in chunks:
        ph=_v162_mobile(x)
        if ph and ph not in vals:vals.append(ph)
    return vals

def _v162_upsert_phone_contact(phone,name=None,email=None,company=None,category=None,city=None,location=None,upload_id=None):
    """
    Uses the existing V15 marketing contact upsert layer so phone uploads
    dedupe against the existing WhatsApp marketing database.
    """
    before=0
    try:
        with engine.connect() as c:
            before=int(c.execute(text("SELECT COUNT(*) FROM pi_marketing_contacts WHERE primary_phone=:p"),{"p":phone}).scalar_one() or 0)
    except Exception:
        pass

    _v151_upsert_contact(
        phone,
        name=name,
        company=company,
        category=(category or "OTHER"),
        city=city,
        location=location,
        email=email,
        website=None,
        source="PHONE_UPLOAD",
        source_detail=f"PHONE_UPLOAD:{upload_id or ''}",
        notes="Imported from user's phone contacts for WhatsApp marketing; verify/consent status before outreach."
    )

    return "updated" if before else "created"

def _v162_parse_vcf(raw_text, default_category):
    records=[]
    current=None
    for raw in raw_text.splitlines():
        line=raw.strip()
        if line.upper()=="BEGIN:VCARD":
            current={"phones":[]}
            continue
        if line.upper()=="END:VCARD":
            if current is not None:
                records.append(current)
            current=None
            continue
        if current is None:continue
        left,sep,val=line.partition(":")
        if not sep:continue
        key=left.upper()
        val=val.strip()
        if key.startswith("FN"):
            current["name"]=val
        elif key.startswith("ORG"):
            current["company"]=val.replace(";"," ").strip()
        elif key.startswith("TEL"):
            current.setdefault("phones",[]).append(val)
        elif key.startswith("EMAIL"):
            current["email"]=val
        elif key.startswith("ADR"):
            current["location"]=val.replace(";"," ").strip()
        elif key.startswith("CATEGORIES"):
            current["category"]=val.split(",")[0].strip().upper()
    for r in records:
        r["category"]=r.get("category") or default_category
    return records

def _v162_parse_csv_bytes(data, default_category):
    import io,csv
    text_data=data.decode("utf-8-sig","ignore")
    reader=csv.DictReader(io.StringIO(text_data))
    out=[]
    for row in reader:
        low={str(k or "").strip().lower():v for k,v in row.items()}
        def first(*keys):
            for k in keys:
                if low.get(k) not in (None,""):return low.get(k)
            return None
        out.append({
            "name":first("name","contact name","full name","contact_name"),
            "phones":[first("phone","mobile","mobile number","phone number","contact number","whatsapp")],
            "email":first("email","email id","email_id"),
            "company":first("company","brand","business","organization","organisation"),
            "category":(first("category","type","segment") or default_category),
            "city":first("city"),
            "location":first("location","address","area")
        })
    return out

def _v162_parse_xlsx(path,default_category):
    from openpyxl import load_workbook
    wb=load_workbook(path,read_only=True,data_only=True)
    ws=wb.active
    rows=ws.iter_rows(values_only=True)
    try:headers=[str(x or "").strip().lower() for x in next(rows)]
    except StopIteration:
        wb.close();return []
    idx={h:i for i,h in enumerate(headers)}
    def gi(row,*keys):
        for k in keys:
            i=idx.get(k)
            if i is not None and i<len(row) and row[i] not in (None,""):
                return row[i]
        return None
    out=[]
    for row in rows:
        out.append({
            "name":gi(row,"name","contact name","full name","contact_name"),
            "phones":[gi(row,"phone","mobile","mobile number","phone number","contact number","whatsapp")],
            "email":gi(row,"email","email id","email_id"),
            "company":gi(row,"company","brand","business","organization","organisation"),
            "category":gi(row,"category","type","segment") or default_category,
            "city":gi(row,"city"),
            "location":gi(row,"location","address","area")
        })
    wb.close()
    return out

@app.post("/api/v16-2/phone-contacts/upload")
async def v162_phone_contacts_upload(
    req:Request,
    file:UploadFile=File(...),
    default_category:str=Form("OTHER")
):
    need_login(req);_v162_setup()
    fn=file.filename or "contacts"
    ext=fn.lower().rsplit(".",1)[-1] if "." in fn else ""
    if ext not in {"vcf","csv","xlsx"}:
        raise HTTPException(400,"Upload .vcf, .csv or .xlsx")

    data=await file.read()
    if len(data)>25*1024*1024:
        raise HTTPException(413,"Maximum file size is 25 MB.")

    default_category=str(default_category or "OTHER").strip().upper()
    upload_id="PHONE-"+datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")+"-"+uuid.uuid4().hex[:6].upper()

    tmp=None
    try:
        if ext=="vcf":
            records=_v162_parse_vcf(data.decode("utf-8","ignore"),default_category)
        elif ext=="csv":
            records=_v162_parse_csv_bytes(data,default_category)
        else:
            fd,tmp=tempfile.mkstemp(suffix=".xlsx");os.close(fd)
            with open(tmp,"wb") as f:f.write(data)
            records=_v162_parse_xlsx(tmp,default_category)
    finally:
        if tmp:
            try:os.unlink(tmp)
            except:pass

    rows_read=len(records)
    valid_mobile_rows=created=updated=duplicates=invalid=0

    with engine.begin() as c:
        c.execute(text("""INSERT INTO pi_phone_contact_uploads(
            upload_id,filename,default_category,rows_read,uploaded_by
        ) VALUES(:id,:fn,:cat,:rows,:by)"""),{
            "id":upload_id,"fn":fn,"cat":default_category,"rows":rows_read,"by":actor_name(req)
        })

    seen=set()
    for r in records:
        phones=[]
        for rawp in (r.get("phones") or []):
            phones.extend(_v162_split_phones(rawp))
        phones=list(dict.fromkeys(phones))
        if not phones:
            invalid+=1
            with engine.begin() as c:
                c.execute(text("""INSERT INTO pi_phone_contact_upload_evidence(
                    upload_id,contact_name,raw_phone,email,company_brand,category,city,location,status,reason
                ) VALUES(:u,:n,:raw,:e,:co,:cat,:city,:loc,'INVALID','No valid 10-digit Indian mobile')"""),{
                    "u":upload_id,"n":r.get("name"),"raw":" | ".join(str(x or "") for x in (r.get("phones") or [])),
                    "e":r.get("email"),"co":r.get("company"),"cat":r.get("category") or default_category,
                    "city":r.get("city"),"loc":r.get("location")
                })
            continue

        valid_mobile_rows+=1
        for ph in phones:
            if ph in seen:
                duplicates+=1
                continue
            seen.add(ph)
            try:
                outcome=_v162_upsert_phone_contact(
                    ph,name=r.get("name"),email=r.get("email"),company=r.get("company"),
                    category=str(r.get("category") or default_category).upper(),
                    city=r.get("city"),location=r.get("location"),upload_id=upload_id
                )
                if outcome=="created":created+=1
                else:updated+=1
                with engine.begin() as c:
                    c.execute(text("""INSERT INTO pi_phone_contact_upload_evidence(
                        upload_id,contact_name,raw_phone,normalized_phone,email,company_brand,category,
                        city,location,status,reason
                    ) VALUES(:u,:n,:raw,:ph,:e,:co,:cat,:city,:loc,'IMPORTED',:reason)"""),{
                        "u":upload_id,"n":r.get("name"),"raw":" | ".join(str(x or "") for x in (r.get("phones") or [])),
                        "ph":ph,"e":r.get("email"),"co":r.get("company"),
                        "cat":str(r.get("category") or default_category).upper(),
                        "city":r.get("city"),"loc":r.get("location"),"reason":outcome.upper()
                    })
            except Exception as ex:
                invalid+=1

    with engine.begin() as c:
        c.execute(text("""UPDATE pi_phone_contact_uploads SET
            valid_mobile_rows=:v,contacts_created=:c,contacts_updated=:u,
            duplicates=:d,invalid_rows=:i WHERE upload_id=:id"""),{
            "v":valid_mobile_rows,"c":created,"u":updated,"d":duplicates,"i":invalid,"id":upload_id
        })

    return {
        "status":"ok","upload_id":upload_id,"rows_read":rows_read,
        "valid_mobile_rows":valid_mobile_rows,"contacts_created":created,
        "contacts_updated":updated,"duplicates":duplicates,"invalid_rows":invalid
    }

@app.get("/api/v16-2/phone-contacts/uploads")
def v162_phone_uploads(req:Request):
    need_login(req);_v162_setup()
    with engine.connect() as c:
        rows=[dict(r._mapping) for r in c.execute(text(
            "SELECT * FROM pi_phone_contact_uploads ORDER BY id DESC LIMIT 50"
        )).fetchall()]
    return {"status":"ok","rows":rows}

@app.get("/phone-contact-upload",response_class=HTMLResponse)
def v162_phone_upload_page(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    return HTMLResponse("""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Upload Phone Contacts</title>
<style>*{box-sizing:border-box}body{font-family:Arial;margin:0;background:#f4f7fb;color:#172437}header{background:#102235;color:#fff;padding:18px}.w{max-width:1200px;margin:auto;padding:18px}.card{background:#fff;border:1px solid #e2e8f0;border-radius:12px;padding:16px;margin-bottom:14px}.bar{display:flex;gap:8px;flex-wrap:wrap}.btn,a.btn{padding:9px 11px;border:0;border-radius:8px;background:#1677ff;color:white;text-decoration:none;font-weight:bold;cursor:pointer}.gray{background:#e9eef5!important;color:#203247!important}select,input{padding:9px;border:1px solid #ccd6e2;border-radius:7px}.msg{margin-top:12px;padding:10px;border-radius:8px;background:#fff8e8;border:1px solid #eed18f}table{width:100%;border-collapse:collapse;font-size:12px}th,td{padding:8px;border-bottom:1px solid #edf1f5;text-align:left}.small{font-size:12px;color:#687789}</style></head>
<body><header><b>Upload Phone Contacts for WhatsApp Marketing</b><br><small>VCF · CSV · XLSX · Dedupe by mobile number</small></header><div class=w>
<div class=card>
<div class=bar><a class="btn gray" href="/final-dashboard-v3">← Final Dashboard</a><a class="btn gray" href="/marketing-contacts-final">Marketing Contacts</a></div>
<h3>Upload contacts from your phone</h3>
<p class=small>Android and iPhone can export contacts as a <b>.VCF</b> file. You can also upload CSV/XLSX. Existing mobile numbers are updated/deduplicated instead of duplicated.</p>
<form id=f>
<label>Default Category</label><br>
<select name=default_category>
<option>OTHER</option><option>CAFE</option><option>RESTAURANT</option><option>BANQUET</option><option>HOTEL</option>
<option>GUEST_HOUSE</option><option>LOUNGE</option><option>CLUB</option><option>BAR</option><option>FARMHOUSE</option>
<option>RETAILER</option><option>BROKER</option><option>OWNER</option>
</select><br><br>
<input type=file name=file accept=".vcf,.csv,.xlsx" required>
<button class=btn type=submit>Upload Contacts</button>
</form>
<div id=msg class=msg>Phone-uploaded contacts enter Marketing Contacts with source <b>PHONE_UPLOAD</b>. Verify before WhatsApp outreach.</div>
</div>
<div class=card><h3>Recent Uploads</h3><table><thead><tr><th>File</th><th>Rows</th><th>Valid Mobile Rows</th><th>Created</th><th>Updated</th><th>Duplicates</th><th>Invalid</th><th>Date</th></tr></thead><tbody id=rows></tbody></table></div>
</div>
<script>
f.onsubmit=async e=>{e.preventDefault();msg.textContent='Uploading and organizing contacts...';let r=await fetch('/api/v16-2/phone-contacts/upload',{method:'POST',body:new FormData(f)}),d=await r.json();if(!r.ok){msg.textContent='ERROR: '+(d.detail||d.message||'Upload failed');return}msg.textContent=`Done. ${d.contacts_created} created, ${d.contacts_updated} updated, ${d.duplicates} duplicates skipped, ${d.invalid_rows} invalid rows.`;load()}
async function load(){let d=await(await fetch('/api/v16-2/phone-contacts/uploads')).json();rows.innerHTML=(d.rows||[]).map(x=>`<tr><td>${x.filename||''}</td><td>${x.rows_read||0}</td><td>${x.valid_mobile_rows||0}</td><td>${x.contacts_created||0}</td><td>${x.contacts_updated||0}</td><td>${x.duplicates||0}</td><td>${x.invalid_rows||0}</td><td>${String(x.created_at||'').slice(0,16)}</td></tr>`).join('')||'<tr><td colspan=8>No uploads yet.</td></tr>'}load()
</script></body></html>""")

def _v162_table_count(table):
    try:
        with engine.connect() as c:
            return int(c.execute(text(f'SELECT COUNT(*) FROM "{table}"')).scalar_one() or 0)
    except Exception:
        return None

@app.get("/api/v16-2/admin/health")
def v162_admin_health(req:Request):
    role=need_login(req)
    if role!="admin":raise HTTPException(403,"Admin only")
    tables=["pi_properties","pi_requirements","pi_matches","pi_ai_hospitality_master","pi_marketing_contacts",
            "ai_marketing_contacts","ai_bot_runs","pi_sources","pi_phone_contact_uploads"]
    counts={t:_v162_table_count(t) for t in tables}
    try:
        with engine.connect() as c:
            c.execute(text("SELECT 1")).scalar_one()
        db="OK"
    except Exception as ex:
        db=f"ERROR: {ex}"
    return {"status":"ok","database":db,"counts":counts}

@app.get("/admin-data-tools-v2",response_class=HTMLResponse)
def v162_admin_tools(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    if role!="admin":
        return HTMLResponse("<h2>Admin only</h2><p>Please log in with the admin code.</p>",status_code=403)
    return HTMLResponse("""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Admin Data Tools</title>
<style>*{box-sizing:border-box}body{font-family:Arial;margin:0;background:#f4f7fb;color:#172437}header{background:#102235;color:#fff;padding:18px}.w{max-width:1300px;margin:auto;padding:18px}.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:12px}.card{display:block;background:#fff;border:1px solid #e2e8f0;border-radius:12px;padding:15px;text-decoration:none;color:#172437}.card p{font-size:12px;color:#687789}.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:8px;margin-bottom:14px}.k{background:#fff;border:1px solid #e2e8f0;border-radius:10px;padding:12px}.k b{display:block;font-size:20px}.btn{display:inline-block;padding:8px 10px;border-radius:8px;background:#e9eef5;color:#203247;text-decoration:none;font-weight:bold;margin-bottom:12px}</style></head>
<body><header><b>Admin Data Tools V2</b><br><small>Safe database health and maintenance links</small></header><div class=w>
<a class=btn href="/final-dashboard-v3">← Final Dashboard</a>
<div class=kpis id=kpis><div class=k><b>Loading...</b><span>Database health</span></div></div>
<div class=grid>
<a class=card href="/data-doctor"><b>Data Doctor</b><p>Full property/contact reconciliation and database health.</p></a>
<a class=card href="/property-database"><b>Full Property Database</b><p>Inspect master property data.</p></a>
<a class=card href="/contacts-directory"><b>Owner / Broker Contacts</b><p>Verify property contact roles.</p></a>
<a class=card href="/ai-hospitality-master-final"><b>Hospitality Master</b><p>Inspect AI Hospitality business data.</p></a>
<a class=card href="/marketing-contacts-final"><b>Marketing Contacts</b><p>WhatsApp marketing contact database.</p></a>
<a class=card href="/phone-contact-upload"><b>Phone Contacts Upload</b><p>Import VCF/CSV/XLSX contacts.</p></a>
<a class=card href="/hospitality-enrichment"><b>Hospitality Enrichment</b><p>Find missing business mobile numbers.</p></a>
<a class=card href="/capture-intelligence"><b>Capture Intelligence</b><p>Upload and extract source documents/images.</p></a>
</div></div>
<script>
(async()=>{let r=await fetch('/api/v16-2/admin/health'),d=await r.json();if(!r.ok){kpis.innerHTML='<div class=k><b>ERROR</b><span>'+(d.detail||'Unable to load')+'</span></div>';return}let h=`<div class=k><b>${d.database}</b><span>DATABASE</span></div>`;for(const [k,v] of Object.entries(d.counts||{})){h+=`<div class=k><b>${v===null?'N/A':v}</b><span>${k}</span></div>`}kpis.innerHTML=h})()
</script></body></html>""")

@app.get("/final-dashboard-v3",response_class=HTMLResponse)
def v162_final_dashboard(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    admin_card = """<a class="card" href="/admin-data-tools-v2"><b>Admin Data Tools</b><p>Database health, reconciliation and maintenance links.</p><span class=tag>ADMIN</span></a>""" if role=="admin" else ""
    return HTMLResponse(f"""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>AI Deal Intelligence OS</title>
<style>*{{box-sizing:border-box}}body{{font-family:Arial;margin:0;background:#f4f7fb;color:#172437}}header{{background:#102235;color:#fff;padding:22px}}.w{{max-width:1550px;margin:auto;padding:20px}}.section{{margin-bottom:24px}}.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:12px}}.card{{background:#fff;border:1px solid #e2e8f0;border-radius:12px;padding:16px;text-decoration:none;color:#172437;min-height:105px;display:block}}.card b{{font-size:16px}}.card p{{font-size:12px;color:#687789;line-height:1.4}}.primary{{border:2px solid #1677ff}}.bot{{border:2px solid #14a673}}.tag{{display:inline-block;padding:3px 7px;border-radius:10px;background:#edf4ff;font-size:10px}}.btn{{display:inline-block;padding:9px 11px;border:0;border-radius:8px;background:#08734b;color:#fff;font-weight:bold;cursor:pointer}}.status{{margin-top:8px;padding:7px;background:#f6f8fb;border-radius:7px;font-size:12px}}</style></head>
<body><header><b>AI Deal Intelligence OS</b><br><small>Final Team Dashboard · One entry point</small></header><div class=w>

<div class=section><h2>Run AI Bots</h2><div class=grid>
<a class="card" href="/property-discovery?v=17.4"><b>Find Property by Demand</b><p>Search public web sources from a natural-language requirement, review structured individual properties, verify availability, then add selected inventory to the Property Database.</p><span class=tag>AI SEARCH</span></a>

<div class="card bot"><b>Hospitality Bot</b><p>Fetch fresh Hospitality business contacts and signals.</p><button class=btn onclick="runBot('hospitality')">▶ Run Hospitality Bot</button><div id=hmsg class=status>Ready</div></div>
<div class="card bot"><b>Retail Bot</b><p>Fetch fresh Retail expansion and leasing signals.</p><button class=btn onclick="runBot('retail')">▶ Run Retail Bot</button><div id=rmsg class=status>Ready</div></div>
<a class=card href="/hospitality-enrichment"><b>Find Missing Hospitality Contacts</b><p>Phone-first enrichment for existing businesses.</p></a>
<a class=card href="/ai-hospitality-master-final"><b>Hospitality Master</b><p>Review Hospitality database by category.</p></a>
</div></div>

<div class=section><h2>Property & Requirements</h2><div class=grid>
<a class="card primary" href="/v14-property-form"><b>Add Property Manually</b><p>Fresh structured inventory.</p></a>
<a class="card primary" href="/v14-requirement-form"><b>Add Requirement Manually</b><p>Confirmed requirement entry.</p></a>
<a class="card primary" href="/v14-matcher"><b>Property Matcher</b><p>Match fresh/verified inventory.</p></a>
<a class=card href="/v14-inventory"><b>Fresh Inventory Database</b><p>Search current working inventory.</p></a>
<a class=card href="/requirements-match-center"><b>Requirements Centre</b><p>AI + manual requirements separated.</p></a>
<a class=card href="/retail-expansion"><b>Retail Expansion</b><p>Retail AI results.</p></a>
</div></div>

<div class=section><h2>WhatsApp Marketing</h2><div class=grid>
<a class="card primary" href="/phone-contact-upload"><b>Upload Contacts From Phone</b><p>Upload iPhone/Android .VCF, CSV or XLSX. Dedupe automatically.</p><span class=tag>NEW</span></a>
<a class=card href="/marketing-contacts-final"><b>Marketing Contacts</b><p>All WhatsApp marketing contacts with source/category filters.</p></a>
<a class=card href="/api/v16/whatsapp-ready.csv"><b>Export Hospitality WhatsApp CSV</b><p>Download Hospitality contacts with usable mobile numbers.</p></a>
<a class=card href="/contacts-directory"><b>Owner / Broker Contacts</b><p>Property contacts remain separate.</p></a>
</div></div>


<div class=section><h2>Contacts & Enrichment</h2><div class=grid>
<a class="card primary" href="/contacts-control-center"><b>All Contacts Control Center</b><p>One place for marketing, hospitality, owner/broker, uploaded and recovered contacts.</p><span class=tag>ALL CONTACTS</span></a>
<a class="card primary" href="/hospitality-enrichment"><b>Contact Enrichment</b><p>Find missing Hospitality mobile numbers. Stale jobs auto-reset safely.</p><span class=tag>ENRICH</span></a>
<a class=card href="/marketing-contacts-final"><b>Marketing Contacts</b><p>All WhatsApp marketing contacts with source/category filters.</p></a>
<a class=card href="/ai-hospitality-master-final"><b>Hospitality Master</b><p>All AI Hospitality businesses, including records still needing enrichment.</p></a>
<a class=card href="/contacts-directory"><b>Owner / Broker Contacts</b><p>Property owner and broker contacts remain preserved separately.</p></a>
<a class=card href="/phone-contact-upload"><b>Uploaded Phone Contacts</b><p>VCF, CSV and XLSX contacts already imported or ready to import.</p></a>
</div></div>


<div class=section><h2>Goa Property Intelligence</h2><div class=grid>
<a class="card primary" href="/goa-property-form"><b>Add Goa Property</b><p>Goa-specific property inventory entry and verification.</p><span class=tag>GOA INVENTORY</span></a>
<a class="card primary" href="/goa-requirement-form"><b>Add Goa Requirement</b><p>Buyer/investor demand entry for Goa properties.</p><span class=tag>GOA DEMAND</span></a>
<a class=card href="/goa-matcher"><b>Goa Matcher</b><p>Match verified Goa requirements against Goa inventory.</p></a>
<a class=card href="/goa-database"><b>Goa Database</b><p>Search and manage Goa inventory separately.</p></a>
</div></div>

<div class=section><h2>Database & Admin</h2><div class=grid>
<a class=card href="/capture-intelligence"><b>Capture Property</b><p>Camera, screenshot, newspaper, magazine, handwritten note, PDF.</p></a>
<a class=card href="/property-database"><b>Full Property Database</b><p>Master property archive.</p></a>
<a class=card href="/data-doctor"><b>Data Doctor</b><p>Database reconciliation and health.</p></a>
{admin_card}
</div></div>

</div><script>
async function runBot(t){{let box=document.getElementById(t==='hospitality'?'hmsg':'rmsg'),u=t==='hospitality'?'/api/v4/hospitality-bot/start':'/api/v4/retail-bot/start';box.textContent='Starting...';try{{let r=await fetch(u,{{method:'POST'}}),d=await r.json();if(!r.ok)throw Error(d.detail||d.message||'Failed');box.textContent='Started in background'+(d.run_id?' · '+d.run_id:'')}}catch(e){{box.textContent='ERROR: '+e.message}}}}
</script></body></html>""")

@app.middleware("http")
async def v162_final_router(request,call_next):
    if request.url.path in {"/workspace","/final-dashboard","/final-dashboard-v2","/admin-data-tools"}:
        if request.url.path=="/admin-data-tools":
            return RedirectResponse("/admin-data-tools-v2",status_code=307)
        return RedirectResponse("/final-dashboard-v3",status_code=307)
    response=await call_next(request)
    if request.url.path.startswith(("/final-dashboard-v3","/admin-data-tools-v2","/phone-contact-upload","/property-discovery","/api/discovery","/contacts-control-center","/api/v17-5","/goa-")):
        response.headers["Cache-Control"]="no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"]="no-cache"
        response.headers["Expires"]="0"
    return response





# ============================================================
# V17.5 GOA + CONTACTS RECOVERY CONTROL CENTER
# ============================================================

@app.get("/api/v17-5/contacts/summary")
def v175_contacts_summary(req:Request):
    need_login(req)
    def one(sql):
        try:
            with engine.connect() as c:
                return int(c.execute(text(sql)).scalar_one() or 0)
        except Exception:
            return 0
    return {
        "status":"ok",
        "marketing_contacts":one("SELECT COUNT(*) FROM pi_marketing_contacts"),
        "marketing_with_phone":one("SELECT COUNT(*) FROM pi_marketing_contacts WHERE primary_phone IS NOT NULL AND primary_phone<>''"),
        "hospitality_master":one("SELECT COUNT(*) FROM pi_ai_hospitality_master"),
        "hospitality_with_phone":one("SELECT COUNT(*) FROM pi_ai_hospitality_master WHERE primary_phone IS NOT NULL AND primary_phone<>''"),
        "hospitality_needs_enrichment":one("SELECT COUNT(*) FROM pi_ai_hospitality_master WHERE primary_phone IS NULL OR primary_phone=''"),
        "property_owner_contacts":one("SELECT COUNT(*) FROM pi_properties WHERE owner_contact IS NOT NULL AND owner_contact<>''"),
        "property_broker_contacts":one("SELECT COUNT(*) FROM pi_properties WHERE broker_contact IS NOT NULL AND broker_contact<>''"),
        "phone_upload_batches":one("SELECT COUNT(*) FROM pi_phone_contact_uploads")
    }

@app.post("/api/v17-5/hospitality-enrichment/reset-stuck")
def v175_reset_stuck_enrichment(req:Request):
    need_login(req)
    _v16_setup()
    with engine.begin() as c:
        rows=c.execute(text("""
            UPDATE pi_hospitality_enrichment_jobs
            SET status='FAILED',
                error_message=COALESCE(error_message,'') || ' | Auto-reset stale job after application/Railway restart',
                finished_at=NOW(),
                updated_at=NOW(),
                current_business=NULL
            WHERE status IN ('QUEUED','RUNNING')
              AND updated_at < NOW() - INTERVAL '15 minutes'
            RETURNING id
        """)).fetchall()
    return {"status":"ok","stale_jobs_reset":len(rows),"job_ids":[r[0] for r in rows]}

@app.get("/contacts-control-center",response_class=HTMLResponse)
def v175_contacts_control_center(req:Request):
    role=page_role_or_redirect(req)
    if not role:
        return RedirectResponse("/login",303)
    return HTMLResponse("""<!doctype html>
<html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1">
<title>All Contacts Control Center</title>
<style>
*{box-sizing:border-box}body{font-family:Arial;margin:0;background:#f4f7fb;color:#172437}
header{background:#102235;color:#fff;padding:20px}.w{max-width:1500px;margin:auto;padding:20px}
.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:12px;margin-bottom:18px}
.card{background:#fff;border:1px solid #e2e8f0;border-radius:12px;padding:15px;text-decoration:none;color:#172437}
.card b{display:block;font-size:20px}.card span,.small{font-size:12px;color:#687789}
.actions{display:flex;gap:8px;flex-wrap:wrap}.btn,a.btn{padding:9px 11px;border:0;border-radius:8px;background:#1677ff;color:#fff;text-decoration:none;font-weight:bold;cursor:pointer}
.green{background:#08734b}.gray{background:#e9eef5!important;color:#203247!important}.warn{background:#b66a00}
.msg{margin-top:12px;background:#fff8e8;border:1px solid #eed18f;padding:11px;border-radius:9px}
</style></head>
<body><header><b>All Contacts Control Center</b><br><small>Restore, enrich, verify and export existing contacts without deleting data.</small></header>
<div class=w>
<div class=actions>
<a class="btn gray" href="/final-dashboard-v3">← Dashboard</a>
<a class=btn href="/marketing-contacts-final">Marketing Contacts</a>
<a class=btn href="/ai-hospitality-master-final">Hospitality Master</a>
<a class=btn href="/contacts-directory">Owner / Broker Contacts</a>
<a class=btn href="/phone-contact-upload">Phone Uploads</a>
<a class=btn href="/hospitality-enrichment">Enrichment</a>
<a class=btn href="/api/v16/whatsapp-ready.csv">Export WhatsApp CSV</a>
</div>

<h2>Contact Inventory</h2>
<div class=grid>
<div class=card><b id=mkt>0</b><span>Marketing Contacts</span></div>
<div class=card><b id=mktph>0</b><span>Marketing Contacts With Phone</span></div>
<div class=card><b id=hosp>0</b><span>Hospitality Master</span></div>
<div class=card><b id=hospph>0</b><span>Hospitality With Phone</span></div>
<div class=card><b id=need>0</b><span>Hospitality Needs Enrichment</span></div>
<div class=card><b id=owners>0</b><span>Property Owner Contacts</span></div>
<div class=card><b id=brokers>0</b><span>Property Broker Contacts</span></div>
<div class=card><b id=uploads>0</b><span>Phone Upload Batches</span></div>
</div>

<h2>Recovery & Repair</h2>
<div class=actions>
<button class="btn green" onclick="recoverHospitality()">Recover Existing AI Hospitality Contacts</button>
<button class="btn green" onclick="recoverPhones()">Recover Historical Phone Numbers</button>
<button class="btn warn" onclick="resetStuck()">Reset Stuck Enrichment Job</button>
<a class=btn href="/hospitality-enrichment">Start / Continue Enrichment</a>
</div>
<div id=msg class=msg>No records are deleted by these controls.</div>
</div>
<script>
async function summary(){
 let r=await fetch('/api/v17-5/contacts/summary'),d=await r.json();
 mkt.textContent=d.marketing_contacts||0;mktph.textContent=d.marketing_with_phone||0;
 hosp.textContent=d.hospitality_master||0;hospph.textContent=d.hospitality_with_phone||0;
 need.textContent=d.hospitality_needs_enrichment||0;owners.textContent=d.property_owner_contacts||0;
 brokers.textContent=d.property_broker_contacts||0;uploads.textContent=d.phone_upload_batches||0;
}
async function recoverHospitality(){
 msg.textContent='Recovering existing AI Hospitality records...';
 let r=await fetch('/api/v15-6/recover-ai-hospitality',{method:'POST'}),d=await r.json();
 msg.textContent=r.ok?`Recovered ${d.recovered||0}; contact-ready ${d.contact_ready||0}; needs enrichment ${d.needs_enrichment||0}.`:`ERROR: ${d.detail||d.message||'Recovery failed'}`;
 await summary();
}
async function recoverPhones(){
 msg.textContent='Scanning historical tables for phone numbers already fetched earlier...';
 let r=await fetch('/api/v15-8/recover-historical-hospitality-phones',{method:'POST'}),d=await r.json();
 msg.textContent=r.ok?`Phone recovery complete. Evidence ${d.phone_evidence_found||0}; auto-applied ${d.auto_applied||0}; master with phone ${d.total_with_phone||0}.`:`ERROR: ${d.detail||d.message||'Phone recovery failed'}`;
 await summary();
}
async function resetStuck(){
 msg.textContent='Checking stale enrichment jobs...';
 let r=await fetch('/api/v17-5/hospitality-enrichment/reset-stuck',{method:'POST'}),d=await r.json();
 msg.textContent=r.ok?`Reset ${d.stale_jobs_reset||0} stale job(s). You can start enrichment again.`:`ERROR: ${d.detail||d.message||'Reset failed'}`;
}
summary();
</script></body></html>""")


# === PROPERTY DISCOVERY DASHBOARD ALIAS V17.4 ===
@app.get("/final-dashboard-v3/property-discovery")
def property_discovery_dashboard_alias(req: Request):
    if not page_role_or_redirect(req):
        return RedirectResponse("/login", status_code=303)
    return RedirectResponse("/property-discovery?v=17.4", status_code=307)
# === END PROPERTY DISCOVERY DASHBOARD ALIAS V17.4 ===



# === V18.1 RELIABLE PROPERTY ENTRY INTEGRATION ===
from reliable_property_entry import install_reliable_property_entry as _install_reliable_property_entry
_install_reliable_property_entry(
    app=app,
    engine=engine,
    need_login=need_login,
    page_role_or_redirect=page_role_or_redirect,
    actor_name=actor_name,
)

@app.middleware("http")
async def v181_manual_property_redirect(request, call_next):
    if request.url.path in {
        "/v14-property-form",
        "/add-property-manual",
        "/manual-property-entry",
        "/add-property"
    }:
        return RedirectResponse("/property-entry-reliable", status_code=307)
    return await call_next(request)
# === END V18.1 RELIABLE PROPERTY ENTRY INTEGRATION ===

# === PROPERTY DISCOVERY V17 INTEGRATION ===
from property_discovery import install_property_discovery as _install_property_discovery
_install_property_discovery(
    app=app,
    engine=engine,
    need_login=need_login,
    save_property=save_property,
    actor_name=actor_name,
)
# === END PROPERTY DISCOVERY V17 INTEGRATION ===

# ============================================================
# V17 OPERATIONAL FORMS + GOA + MATCHER FIX
# ============================================================
def _v17_setup():
    with engine.begin() as c:
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_operational_properties(
            id BIGSERIAL PRIMARY KEY, property_code TEXT UNIQUE NOT NULL, division TEXT NOT NULL,
            property_name TEXT, property_types JSONB DEFAULT '[]'::jsonb, city TEXT, location TEXT NOT NULL,
            google_location TEXT, area_sqft NUMERIC(14,2) NOT NULL, rent_amount NUMERIC(16,2) NOT NULL,
            rent_unit TEXT DEFAULT 'MONTH', transaction_type TEXT DEFAULT 'LEASE', floor TEXT, frontage TEXT,
            parking TEXT, possession TEXT, suitable_for TEXT, nearby_brands TEXT, owner_broker_name TEXT,
            contact_number TEXT, contact_role TEXT DEFAULT 'UNVERIFIED', verification_status TEXT DEFAULT 'UNVERIFIED',
            remarks TEXT, created_by TEXT, created_at TIMESTAMPTZ DEFAULT NOW(), updated_at TIMESTAMPTZ DEFAULT NOW()
        )"""))
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_operational_property_media(
            id BIGSERIAL PRIMARY KEY, property_code TEXT NOT NULL, media_type TEXT NOT NULL,
            filename TEXT, mime_type TEXT, file_size BIGINT, content BYTEA NOT NULL, created_at TIMESTAMPTZ DEFAULT NOW()
        )"""))
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_operational_requirements(
            id BIGSERIAL PRIMARY KEY, requirement_code TEXT UNIQUE NOT NULL, division TEXT NOT NULL,
            client_name TEXT, company_name TEXT, contact_number TEXT, requirement_types JSONB DEFAULT '[]'::jsonb,
            city TEXT, preferred_locations TEXT NOT NULL, minimum_area_sqft NUMERIC(14,2) NOT NULL,
            maximum_area_sqft NUMERIC(14,2) NOT NULL, maximum_rent NUMERIC(16,2), transaction_type TEXT DEFAULT 'LEASE',
            additional_points TEXT, verification_status TEXT DEFAULT 'VERIFIED', created_by TEXT,
            created_at TIMESTAMPTZ DEFAULT NOW(), updated_at TIMESTAMPTZ DEFAULT NOW()
        )"""))
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_operational_matches(
            id BIGSERIAL PRIMARY KEY, requirement_code TEXT NOT NULL, property_code TEXT NOT NULL,
            score NUMERIC(5,2), match_band TEXT, reasons JSONB DEFAULT '[]'::jsonb,
            created_at TIMESTAMPTZ DEFAULT NOW(), UNIQUE(requirement_code,property_code)
        )"""))

def _v17_code(prefix):
    return f"{prefix}-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:5].upper()}"

def _v17_posnum(v,label):
    try:
        x=float(str(v).replace(",","").strip())
        if x<=0: raise ValueError
        return x
    except Exception:
        raise HTTPException(400,f"{label} must be greater than 0.")

def _v17_arr(v):
    if isinstance(v,list): return v
    try:
        x=json.loads(v or "[]")
        return x if isinstance(x,list) else []
    except Exception:return []

def _v17_tokens(v):
    return {x for x in _re.sub(r"[^a-z0-9]+"," ",str(v or "").lower()).split() if len(x)>=3}

@app.post("/api/v17/property/save")
async def v17_property_save(req:Request, division:str=Form("DELHI_NCR"), property_name:str=Form(""),
    property_types:str=Form(...), city:str=Form(""), location:str=Form(...), google_location:str=Form(""),
    area_sqft:str=Form(...), rent_amount:str=Form(...), rent_unit:str=Form("MONTH"),
    transaction_type:str=Form("LEASE"), floor:str=Form(""), frontage:str=Form(""), parking:str=Form(""),
    possession:str=Form(""), suitable_for:str=Form(""), nearby_brands:str=Form(""),
    owner_broker_name:str=Form(""), contact_number:str=Form(""), contact_role:str=Form("UNVERIFIED"),
    verification_status:str=Form("UNVERIFIED"), remarks:str=Form(""),
    images:list[UploadFile]=File(default=[]), videos:list[UploadFile]=File(default=[])):
    need_login(req); _v17_setup()
    pts=[x.strip() for x in property_types.split("|") if x.strip()]
    if not pts: raise HTTPException(400,"Select at least one Property Type.")
    if not location.strip(): raise HTTPException(400,"Location is required.")
    area=_v17_posnum(area_sqft,"Area"); rent=_v17_posnum(rent_amount,"Rent")
    code=_v17_code("PROP")
    with engine.begin() as c:
        c.execute(text("""INSERT INTO pi_operational_properties(
            property_code,division,property_name,property_types,city,location,google_location,area_sqft,rent_amount,
            rent_unit,transaction_type,floor,frontage,parking,possession,suitable_for,nearby_brands,
            owner_broker_name,contact_number,contact_role,verification_status,remarks,created_by
        ) VALUES(:code,:div,:name,CAST(:types AS jsonb),:city,:loc,:google,:area,:rent,:ru,:tt,:floor,:front,
            :park,:poss,:suitable,:nearby,:person,:phone,:role,:ver,:remarks,:by)"""),{
            "code":code,"div":division.upper(),"name":property_name or None,"types":json.dumps(pts),"city":city or None,
            "loc":location,"google":google_location or None,"area":area,"rent":rent,"ru":rent_unit,"tt":transaction_type,
            "floor":floor or None,"front":frontage or None,"park":parking or None,"poss":possession or None,
            "suitable":suitable_for or None,"nearby":nearby_brands or None,"person":owner_broker_name or None,
            "phone":contact_number or None,"role":contact_role,"ver":verification_status,"remarks":remarks or None,
            "by":actor_name(req)})
    saved=0
    for typ,files,limit in [("IMAGE",images or [],12*1024*1024),("VIDEO",videos or [],80*1024*1024)]:
        for f in files:
            if not f or not f.filename: continue
            data=await f.read()
            if not data: continue
            if len(data)>limit: raise HTTPException(413,f"{f.filename} is too large.")
            with engine.begin() as c:
                c.execute(text("""INSERT INTO pi_operational_property_media(property_code,media_type,filename,mime_type,file_size,content)
                    VALUES(:p,:t,:f,:m,:s,:b)"""),{"p":code,"t":typ,"f":f.filename,"m":f.content_type,"s":len(data),"b":data})
            saved+=1
    return {"status":"ok","property_code":code,"media_saved":saved}

@app.post("/api/v17/requirement/save")
async def v17_requirement_save(req:Request):
    need_login(req); _v17_setup(); b=await req.json()
    loc=str(b.get("preferred_locations") or "").strip()
    if not loc: raise HTTPException(400,"Preferred Locations are required.")
    mina=_v17_posnum(b.get("minimum_area_sqft"),"Minimum Area")
    maxa=_v17_posnum(b.get("maximum_area_sqft"),"Maximum Area")
    if maxa<mina: mina,maxa=maxa,mina
    maxr=None
    if str(b.get("maximum_rent") or "").strip(): maxr=_v17_posnum(b.get("maximum_rent"),"Maximum Rent")
    code=_v17_code("REQ")
    with engine.begin() as c:
        c.execute(text("""INSERT INTO pi_operational_requirements(
            requirement_code,division,client_name,company_name,contact_number,requirement_types,city,
            preferred_locations,minimum_area_sqft,maximum_area_sqft,maximum_rent,transaction_type,
            additional_points,verification_status,created_by)
            VALUES(:code,:div,:client,:company,:phone,CAST(:types AS jsonb),:city,:loc,:mina,:maxa,:rent,:tt,:pts,:ver,:by)"""),{
            "code":code,"div":str(b.get("division") or "DELHI_NCR").upper(),"client":b.get("client_name"),
            "company":b.get("company_name"),"phone":b.get("contact_number"),"types":json.dumps(b.get("requirement_types") or []),
            "city":b.get("city"),"loc":loc,"mina":mina,"maxa":maxa,"rent":maxr,
            "tt":b.get("transaction_type") or "LEASE","pts":b.get("additional_points"),
            "ver":b.get("verification_status") or "VERIFIED","by":actor_name(req)})
    return {"status":"ok","requirement_code":code}

def _v17_match(code):
    _v17_setup()
    with engine.connect() as c:
        rr=c.execute(text("SELECT * FROM pi_operational_requirements WHERE requirement_code=:r"),{"r":code}).first()
        if not rr: raise HTTPException(404,"Requirement not found.")
        q=dict(rr._mapping)
        props=[dict(x._mapping) for x in c.execute(text("SELECT * FROM pi_operational_properties WHERE division=:d ORDER BY id DESC"),
            {"d":q["division"]}).fetchall()]
    rt={str(x).lower() for x in _v17_arr(q.get("requirement_types"))}
    qloc=_v17_tokens(q.get("preferred_locations")); mina=float(q["minimum_area_sqft"]); maxa=float(q["maximum_area_sqft"])
    maxr=float(q.get("maximum_rent") or 0); out=[]
    for p in props:
        area=float(p["area_sqft"]); rent=float(p["rent_amount"]); pt={str(x).lower() for x in _v17_arr(p.get("property_types"))}
        if area < mina*0.75 or area > maxa*1.25: continue
        score=0; why=[]; ploc=_v17_tokens(p.get("location")); overlap=len(qloc & ploc)
        if overlap: score+=min(30,10+10*overlap); why.append("Location match")
        if mina<=area<=maxa: score+=30; why.append("Area within requirement")
        else: score+=15; why.append("Area within tolerance")
        if rt and pt and rt&pt: score+=20; why.append("Property type match")
        if maxr:
            if rent<=maxr: score+=15; why.append("Rent within budget")
            elif rent<=maxr*1.15: score+=7; why.append("Rent slightly above budget")
        else: score+=8
        if str(p.get("verification_status")).upper()=="VERIFIED": score+=5; why.append("Verified")
        score=min(100,score); band="EXCELLENT" if score>=85 else "STRONG" if score>=70 else "GOOD" if score>=55 else "POSSIBLE"
        out.append({"property_code":p["property_code"],"property_name":p.get("property_name"),"property_types":_v17_arr(p.get("property_types")),
            "location":p.get("location"),"google_location":p.get("google_location"),"area_sqft":area,"rent_amount":rent,
            "owner_broker_name":p.get("owner_broker_name"),"contact_number":p.get("contact_number"),
            "verification_status":p.get("verification_status"),"score":score,"match_band":band,"reasons":why})
    out.sort(key=lambda x:x["score"],reverse=True)
    with engine.begin() as c:
        c.execute(text("DELETE FROM pi_operational_matches WHERE requirement_code=:r"),{"r":code})
        for x in out[:100]:
            c.execute(text("""INSERT INTO pi_operational_matches(requirement_code,property_code,score,match_band,reasons)
                VALUES(:r,:p,:s,:b,CAST(:why AS jsonb)) ON CONFLICT(requirement_code,property_code)
                DO UPDATE SET score=EXCLUDED.score,match_band=EXCLUDED.match_band,reasons=EXCLUDED.reasons,created_at=NOW()"""),
                {"r":code,"p":x["property_code"],"s":x["score"],"b":x["match_band"],"why":json.dumps(x["reasons"])})
    return out[:100]

@app.post("/api/v17/match/{requirement_code}")
def v17_match_api(requirement_code:str,req:Request):
    need_login(req); out=_v17_match(requirement_code); return {"status":"ok","count":len(out),"matches":out}

@app.get("/api/v17/requirements")
def v17_requirements_api(req:Request,division:str=Query("DELHI_NCR")):
    need_login(req); _v17_setup()
    with engine.connect() as c:
        rows=[dict(x._mapping) for x in c.execute(text("SELECT * FROM pi_operational_requirements WHERE division=:d ORDER BY id DESC"),
            {"d":division.upper()}).fetchall()]
    return {"status":"ok","rows":rows}

def _v17_types():
    return ["Retail Shop","High Street Retail","Mall Retail","Office","Restaurant","Cafe","Banquet / Wedding Venue","Hotel",
            "Guest House","Lounge","Club","Bar","Farmhouse","Warehouse","Industrial","Land","Mixed Use","Residential / Villa"]

@app.get("/operational-property-form",response_class=HTMLResponse)
def v17_property_form(req:Request,division:str=Query("DELHI_NCR")):
    if not page_role_or_redirect(req): return RedirectResponse("/login",303)
    d=division.upper(); checks="".join(f'<label><input type=checkbox name=ptype value="{escape(x)}"> {escape(x)}</label>' for x in _v17_types())
    return HTMLResponse(f"""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Property Form</title>
<style>body{{font-family:Arial;background:#f4f7fb;margin:0;color:#172437}}header{{background:#102235;color:#fff;padding:18px}}.w{{max-width:1150px;margin:auto;padding:18px}}.card{{background:#fff;padding:15px;border-radius:12px;margin-bottom:12px}}.g{{display:grid;grid-template-columns:1fr 1fr;gap:10px}}input,select,textarea{{width:100%;padding:9px;border:1px solid #ccd6e2;border-radius:7px}}.checks{{display:grid;grid-template-columns:repeat(3,1fr);gap:6px}}.checks label{{background:#f6f8fb;padding:6px;border-radius:6px}}.checks input{{width:auto}}.drop{{border:2px dashed #9eb6cf;padding:18px;border-radius:10px;text-align:center;cursor:pointer}}.btn{{padding:9px 12px;background:#1677ff;color:#fff;border:0;border-radius:8px;text-decoration:none;cursor:pointer}}.msg{{margin-top:10px;background:#fff8e8;padding:9px}}@media(max-width:800px){{.g,.checks{{grid-template-columns:1fr}}}}</style></head>
<body><header><b>{'Goa' if d=='GOA' else 'Delhi NCR'} Manual Property Form</b><br><small>Area + Rent required · Images/Videos optional</small></header><div class=w>
<a class=btn href="/final-dashboard-v4">← Dashboard</a><br><br><form id=f enctype=multipart/form-data><input type=hidden name=division value="{d}">
<div class=card><div class=g><input name=property_name placeholder="Property Name"><input name=city value="{'Goa' if d=='GOA' else 'Delhi NCR'}">
<input name=location placeholder="Location *" required><input name=google_location placeholder="Google Maps location/link">
<input name=area_sqft placeholder="Area sq ft *" required><input name=rent_amount placeholder="Rent amount *" required>
<select name=rent_unit><option>MONTH</option><option>SQFT_MONTH</option></select><select name=transaction_type><option>LEASE</option><option>SALE</option><option>LEASE_OR_SALE</option></select>
<input name=floor placeholder="Floor"><input name=frontage placeholder="Frontage"><input name=parking placeholder="Parking"><input name=possession placeholder="Possession">
<input name=owner_broker_name placeholder="Owner/Broker/Contact Name"><input name=contact_number placeholder="Contact Number">
<select name=contact_role><option>UNVERIFIED</option><option>OWNER</option><option>BROKER</option><option>BOTH</option></select>
<select name=verification_status><option>UNVERIFIED</option><option>VERIFIED</option></select></div></div>
<div class=card><b>Property Type — select multiple *</b><div class=checks>{checks}</div></div>
<div class=card><input name=suitable_for placeholder="Suitable For"><br><br><input name=nearby_brands placeholder="Nearby Brands"><br><br><textarea name=remarks placeholder="Remarks"></textarea></div>
<div class=card><b>Photos optional</b><div class=drop id=idrop>Drag photos or click<input type=file id=images name=images accept="image/*" multiple hidden></div><small id=ip>No photos selected</small></div>
<div class=card><b>Videos optional</b><div class=drop id=vdrop>Drag videos or click<input type=file id=videos name=videos accept="video/*" multiple hidden></div><small id=vp>No videos selected</small></div>
<button class=btn>Save Property</button><div id=msg class=msg>Ready.</div></form></div>
<script>
function dz(box,input,p){{box.onclick=()=>input.click();['dragover','drop'].forEach(n=>box.addEventListener(n,e=>e.preventDefault()));box.addEventListener('drop',e=>{{input.files=e.dataTransfer.files;p.textContent=input.files.length+' file(s) selected'}});input.onchange=()=>p.textContent=input.files.length+' file(s) selected'}}dz(idrop,images,ip);dz(vdrop,videos,vp);
f.onsubmit=async e=>{{e.preventDefault();let pts=[...document.querySelectorAll('[name=ptype]:checked')].map(x=>x.value);if(!pts.length){{msg.textContent='Select at least one Property Type.';return}}let fd=new FormData(f);fd.set('property_types',pts.join('|'));msg.textContent='Saving...';let r=await fetch('/api/v17/property/save',{{method:'POST',body:fd}}),d=await r.json();msg.textContent=r.ok?'Saved '+d.property_code+' · media '+d.media_saved:'ERROR: '+(d.detail||d.message||'Save failed')}};</script></body></html>""")

@app.get("/operational-requirement-form",response_class=HTMLResponse)
def v17_requirement_form(req:Request,division:str=Query("DELHI_NCR")):
    if not page_role_or_redirect(req): return RedirectResponse("/login",303)
    d=division.upper(); checks="".join(f'<label><input type=checkbox name=rtype value="{escape(x)}"> {escape(x)}</label>' for x in _v17_types())
    return HTMLResponse(f"""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Requirement Form</title>
<style>body{{font-family:Arial;background:#f4f7fb;margin:0;color:#172437}}header{{background:#102235;color:#fff;padding:18px}}.w{{max-width:1050px;margin:auto;padding:18px}}.card{{background:#fff;padding:15px;border-radius:12px;margin-bottom:12px}}.g{{display:grid;grid-template-columns:1fr 1fr;gap:10px}}input,select,textarea{{width:100%;padding:9px;border:1px solid #ccd6e2;border-radius:7px}}.checks{{display:grid;grid-template-columns:repeat(3,1fr);gap:6px}}.checks label{{background:#f6f8fb;padding:6px;border-radius:6px}}.checks input{{width:auto}}.btn{{padding:9px 12px;background:#1677ff;color:#fff;border:0;border-radius:8px;text-decoration:none;cursor:pointer}}.msg{{margin-top:10px;background:#fff8e8;padding:9px}}@media(max-width:800px){{.g,.checks{{grid-template-columns:1fr}}}}</style></head><body>
<header><b>{'Goa' if d=='GOA' else 'Delhi NCR'} Requirement Form</b></header><div class=w><a class=btn href="/final-dashboard-v4">← Dashboard</a><br><br>
<div class=card><div class=g><input id=client placeholder="Client Name"><input id=company placeholder="Company"><input id=phone placeholder="Contact Number"><input id=city value="{'Goa' if d=='GOA' else 'Delhi NCR'}">
<input id=loc placeholder="Preferred Locations *"><input id=mina placeholder="Minimum Area *"><input id=maxa placeholder="Maximum Area *"><input id=rent placeholder="Maximum Rent">
<select id=tt><option>LEASE</option><option>SALE</option><option>LEASE_OR_SALE</option></select><select id=ver><option>VERIFIED</option><option>UNVERIFIED</option></select></div></div>
<div class=card><b>Property Types</b><div class=checks>{checks}</div></div><div class=card><textarea id=pts placeholder="Additional Points"></textarea></div>
<button class=btn onclick=save()>Save Requirement</button><div id=msg class=msg>Ready.</div></div>
<script>async function save(){{let types=[...document.querySelectorAll('[name=rtype]:checked')].map(x=>x.value);let b={{division:'{d}',client_name:client.value,company_name:company.value,contact_number:phone.value,requirement_types:types,city:city.value,preferred_locations:loc.value,minimum_area_sqft:mina.value,maximum_area_sqft:maxa.value,maximum_rent:rent.value,transaction_type:tt.value,additional_points:pts.value,verification_status:ver.value}};let r=await fetch('/api/v17/requirement/save',{{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify(b)}}),d=await r.json();msg.innerHTML=r.ok?'Saved '+d.requirement_code+' · <a href="/operational-matcher?division={d}">Open Matcher</a>':'ERROR: '+(d.detail||d.message||'Save failed')}};</script></body></html>""")

@app.get("/operational-matcher",response_class=HTMLResponse)
def v17_matcher_page(req:Request,division:str=Query("DELHI_NCR")):
    if not page_role_or_redirect(req): return RedirectResponse("/login",303)
    d=division.upper()
    return HTMLResponse(f"""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Matcher</title>
<style>body{{font-family:Arial;background:#f4f7fb;margin:0;color:#172437}}header{{background:#102235;color:#fff;padding:18px}}.w{{padding:18px}}.btn{{padding:9px 11px;background:#1677ff;color:#fff;border:0;border-radius:8px;text-decoration:none}}select{{padding:9px;min-width:380px}}table{{width:100%;border-collapse:collapse;background:#fff;font-size:12px;margin-top:12px}}th,td{{padding:8px;border-bottom:1px solid #eee;text-align:left}}.msg{{padding:9px;background:#fff8e8;margin-top:10px}}</style></head><body>
<header><b>{'Goa' if d=='GOA' else 'Delhi NCR'} Property Matcher</b><br><small>Uses only clean V17 operational data</small></header><div class=w>
<a class=btn href="/final-dashboard-v4">← Dashboard</a> <a class=btn href="/operational-property-form?division={d}">Add Property</a> <a class=btn href="/operational-requirement-form?division={d}">Add Requirement</a><br><br>
<select id=reqs><option>Loading...</option></select> <button class=btn onclick=run()>Run Match</button><div id=msg class=msg>Select a requirement.</div>
<table><thead><tr><th>Score</th><th>Property</th><th>Type</th><th>Location</th><th>Area</th><th>Rent</th><th>Contact</th><th>Verified</th><th>Google</th><th>Reasons</th></tr></thead><tbody id=rows></tbody></table></div>
<script>async function load(){{let d=await(await fetch('/api/v17/requirements?division={d}')).json();reqs.innerHTML='<option value="">Select requirement</option>'+(d.rows||[]).map(x=>`<option value="${{x.requirement_code}}">${{x.requirement_code}} · ${{x.company_name||x.client_name||''}} · ${{x.preferred_locations}}</option>`).join('')}}async function run(){{if(!reqs.value)return;let r=await fetch('/api/v17/match/'+reqs.value,{{method:'POST'}}),d=await r.json();msg.textContent=r.ok?d.count+' matches found':'ERROR: '+(d.detail||d.message||'Matcher failed');rows.innerHTML=(d.matches||[]).map(x=>`<tr><td><b>${{x.score}}</b><br>${{x.match_band}}</td><td>${{x.property_name||x.property_code}}</td><td>${{(x.property_types||[]).join(', ')}}</td><td>${{x.location||''}}</td><td>${{x.area_sqft}}</td><td>${{x.rent_amount}}</td><td>${{x.owner_broker_name||''}}<br><b>${{x.contact_number||''}}</b></td><td>${{x.verification_status||''}}</td><td>${{x.google_location?`<a target=_blank href="${{x.google_location}}">Map</a>`:''}}</td><td>${{(x.reasons||[]).join(' · ')}}</td></tr>`).join('')}}load()</script></body></html>""")

@app.get("/final-dashboard-v4",response_class=HTMLResponse)
def v17_dashboard(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    admin='<a class=card href="/admin-data-tools-v2"><b>Admin Data Tools</b><p>Database health and tools.</p></a>' if role=="admin" else ""
    return HTMLResponse(f"""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Final Dashboard</title>
<style>body{{font-family:Arial;background:#f4f7fb;margin:0;color:#172437}}header{{background:#102235;color:#fff;padding:22px}}.w{{max-width:1450px;margin:auto;padding:20px}}.g{{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:12px;margin-bottom:24px}}.card{{display:block;background:#fff;border:1px solid #e2e8f0;border-radius:12px;padding:15px;text-decoration:none;color:#172437}}.card p{{font-size:12px;color:#687789}}.primary{{border:2px solid #1677ff}}</style></head><body><header><b>AI Deal Intelligence OS</b><br><small>Final Operational Dashboard</small></header><div class=w>
<h2>Delhi NCR</h2><div class=g><a class="card primary" href="/operational-property-form?division=DELHI_NCR"><b>Add Property</b><p>Area + Rent required, multi-type, Google Location, optional photo/video.</p></a><a class="card primary" href="/operational-requirement-form?division=DELHI_NCR"><b>Add Requirement</b><p>Confirmed manual demand.</p></a><a class="card primary" href="/operational-matcher?division=DELHI_NCR"><b>Run Matcher</b><p>Clean V17 matcher.</p></a></div>
<h2>Goa</h2><div class=g><a class="card primary" href="/operational-property-form?division=GOA"><b>Add Goa Property</b><p>Working Goa form.</p></a><a class="card primary" href="/operational-requirement-form?division=GOA"><b>Add Goa Requirement</b><p>Working Goa requirement form.</p></a><a class="card primary" href="/operational-matcher?division=GOA"><b>Goa Matcher</b><p>Matches only Goa operational data.</p></a></div>
<h2>AI & Marketing</h2><div class=g><a class=card href="/final-dashboard-v3"><b>Bot Controls</b><p>Hospitality and Retail bot controls.</p></a><a class=card href="/ai-hospitality-master-final"><b>Hospitality Master</b><p>Hospitality contacts/data.</p></a><a class=card href="/marketing-contacts-final"><b>Marketing Contacts</b><p>WhatsApp marketing database.</p></a><a class=card href="/phone-contact-upload"><b>Upload Phone Contacts</b><p>VCF/CSV/XLSX.</p></a></div>
<h2>Database</h2><div class=g><a class=card href="/capture-intelligence"><b>Capture Property</b><p>Camera, screenshot, PDF, magazine.</p></a><a class=card href="/property-database"><b>Legacy Property Database</b><p>Historical data, not used by V17 matcher.</p></a><a class=card href="/contacts-directory"><b>Owner/Broker Contacts</b><p>Property contacts.</p></a>{admin}</div>
</div></body></html>""")

@app.middleware("http")
async def v17_router(request,call_next):
    p=request.url.path
    if p in {"/workspace","/final-dashboard","/final-dashboard-v2","/final-dashboard-v3"}:
        return RedirectResponse("/final-dashboard-v4",307)
    if p in {"/property-manual","/v14-property-form"}:
        return RedirectResponse("/operational-property-form?division=DELHI_NCR",307)
    if p=="/v14-requirement-form":
        return RedirectResponse("/operational-requirement-form?division=DELHI_NCR",307)
    if p=="/v14-matcher":
        return RedirectResponse("/operational-matcher?division=DELHI_NCR",307)
    if p in {"/goa-property","/goa-property-form"}:
        return RedirectResponse("/operational-property-form?division=GOA",307)
    if p in {"/goa-requirement","/goa-requirement-form"}:
        return RedirectResponse("/operational-requirement-form?division=GOA",307)
    if p=="/goa-matcher":
        return RedirectResponse("/operational-matcher?division=GOA",307)
    response=await call_next(request)
    if p.startswith(("/final-dashboard-v4","/operational-property-form","/operational-requirement-form","/operational-matcher")):
        response.headers["Cache-Control"]="no-store, no-cache, must-revalidate, max-age=0"
    return response

# ============================================================
# V17.2 FINAL EXECUTION LAYER
# Unified dashboard + final property forms + Goa brochure upload
# Uses existing V17 operational tables and matcher.
# ============================================================

def _v172_route_exists(path:str)->bool:
    try:
        return any(getattr(r,"path",None)==path for r in app.routes)
    except Exception:
        return False

@app.get("/api/v17-2/property/{property_code}/media")
def v172_property_media(property_code:str, req:Request):
    need_login(req); _v17_setup()
    with engine.connect() as c:
        rows=[dict(r._mapping) for r in c.execute(text("""
            SELECT id,media_type,filename,mime_type,file_size,created_at
            FROM pi_operational_property_media
            WHERE property_code=:p
            ORDER BY id
        """),{"p":property_code}).fetchall()]
    return {"status":"ok","rows":rows}

@app.get("/api/v17-2/property-media/{media_id}")
def v172_property_media_download(media_id:int, req:Request):
    need_login(req); _v17_setup()
    with engine.connect() as c:
        row=c.execute(text("""
            SELECT filename,mime_type,content FROM pi_operational_property_media WHERE id=:id
        """),{"id":media_id}).first()
    if not row:
        raise HTTPException(404,"Media not found.")
    d=dict(row._mapping)
    return Response(
        content=d["content"],
        media_type=d.get("mime_type") or "application/octet-stream",
        headers={"Content-Disposition":f'inline; filename="{d.get("filename") or "file"}"'}
    )

@app.post("/api/v17-2/property/save")
async def v172_property_save(
    req:Request,
    division:str=Form("DELHI_NCR"),
    property_name:str=Form(""),
    property_types:str=Form(...),
    city:str=Form(""),
    location:str=Form(...),
    google_location:str=Form(""),
    area_sqft:str=Form(...),
    rent_amount:str=Form(...),
    rent_unit:str=Form("MONTH"),
    transaction_type:str=Form("LEASE"),
    floor:str=Form(""),
    frontage:str=Form(""),
    parking:str=Form(""),
    possession:str=Form(""),
    suitable_for:str=Form(""),
    nearby_brands:str=Form(""),
    owner_broker_name:str=Form(""),
    contact_number:str=Form(""),
    contact_role:str=Form("UNVERIFIED"),
    verification_status:str=Form("UNVERIFIED"),
    remarks:str=Form(""),
    images:list[UploadFile]=File(default=[]),
    videos:list[UploadFile]=File(default=[]),
    brochure:UploadFile|None=File(default=None)
):
    need_login(req); _v17_setup()
    pts=[x.strip() for x in str(property_types or "").split("|") if x.strip()]
    if not pts:
        raise HTTPException(400,"Select at least one Property Type.")
    if not str(location or "").strip():
        raise HTTPException(400,"Location is required.")
    area=_v17_posnum(area_sqft,"Area")
    rent=_v17_posnum(rent_amount,"Rent")

    code=_v17_code("PROP")
    with engine.begin() as c:
        c.execute(text("""INSERT INTO pi_operational_properties(
            property_code,division,property_name,property_types,city,location,google_location,area_sqft,rent_amount,
            rent_unit,transaction_type,floor,frontage,parking,possession,suitable_for,nearby_brands,
            owner_broker_name,contact_number,contact_role,verification_status,remarks,created_by
        ) VALUES(
            :code,:div,:name,CAST(:types AS jsonb),:city,:loc,:google,:area,:rent,:ru,:tt,:floor,:front,
            :park,:poss,:suitable,:nearby,:person,:phone,:role,:ver,:remarks,:by
        )"""),{
            "code":code,"div":division.upper(),"name":property_name or None,"types":json.dumps(pts),
            "city":city or None,"loc":location.strip(),"google":google_location or None,
            "area":area,"rent":rent,"ru":rent_unit,"tt":transaction_type,
            "floor":floor or None,"front":frontage or None,"park":parking or None,"poss":possession or None,
            "suitable":suitable_for or None,"nearby":nearby_brands or None,
            "person":owner_broker_name or None,"phone":contact_number or None,
            "role":contact_role,"ver":verification_status,"remarks":remarks or None,"by":actor_name(req)
        })

    saved={"IMAGE":0,"VIDEO":0,"BROCHURE":0}
    async def save_file(f,typ,limit):
        if not f or not getattr(f,"filename",None):
            return
        data=await f.read()
        if not data:
            return
        if len(data)>limit:
            raise HTTPException(413,f"{f.filename} is too large.")
        with engine.begin() as c:
            c.execute(text("""INSERT INTO pi_operational_property_media(
                property_code,media_type,filename,mime_type,file_size,content
            ) VALUES(:p,:t,:f,:m,:s,:b)"""),{
                "p":code,"t":typ,"f":f.filename,"m":f.content_type,"s":len(data),"b":data
            })
        saved[typ]+=1

    for f in images or []:
        await save_file(f,"IMAGE",12*1024*1024)
    for f in videos or []:
        await save_file(f,"VIDEO",80*1024*1024)
    if brochure and brochure.filename:
        allowed={"application/pdf","application/msword","application/vnd.openxmlformats-officedocument.wordprocessingml.document"}
        if brochure.content_type not in allowed and not brochure.filename.lower().endswith((".pdf",".doc",".docx")):
            raise HTTPException(400,"Brochure must be PDF, DOC or DOCX.")
        await save_file(brochure,"BROCHURE",30*1024*1024)

    return {
        "status":"ok","property_code":code,
        "images_saved":saved["IMAGE"],"videos_saved":saved["VIDEO"],"brochure_saved":saved["BROCHURE"]
    }

def _v172_types_html(name):
    return "".join(
        f'<label><input type="checkbox" name="{name}" value="{escape(x)}"> {escape(x)}</label>'
        for x in _v17_types()
    )

@app.get("/property-form-final",response_class=HTMLResponse)
def v172_property_form(req:Request,division:str=Query("DELHI_NCR")):
    if not page_role_or_redirect(req):
        return RedirectResponse("/login",303)
    d=division.upper()
    title="Goa Manual Property Form" if d=="GOA" else "Delhi NCR Manual Property Form"
    brochure_html = """
      <div class="card"><b>Brochure — optional</b>
      <div class="drop" id="bdrop">Drag brochure here or click
      <input type="file" id="brochure" name="brochure" accept=".pdf,.doc,.docx,application/pdf" hidden></div>
      <small id="bp">No brochure selected</small></div>
    """ if d=="GOA" else """
      <div class="card"><b>Brochure — optional</b>
      <div class="drop" id="bdrop">Drag brochure here or click
      <input type="file" id="brochure" name="brochure" accept=".pdf,.doc,.docx,application/pdf" hidden></div>
      <small id="bp">No brochure selected</small></div>
    """
    return HTMLResponse(f"""<!doctype html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{title}</title>
<style>
*{{box-sizing:border-box}}body{{font-family:Arial;background:#f4f7fb;margin:0;color:#172437}}
header{{background:#102235;color:#fff;padding:18px}}.w{{max-width:1180px;margin:auto;padding:18px}}
.card{{background:#fff;padding:15px;border-radius:12px;margin-bottom:12px;border:1px solid #e2e8f0}}
.g{{display:grid;grid-template-columns:1fr 1fr;gap:10px}}input,select,textarea{{width:100%;padding:9px;border:1px solid #ccd6e2;border-radius:7px}}
.checks{{display:grid;grid-template-columns:repeat(3,1fr);gap:6px}}.checks label{{background:#f6f8fb;padding:6px;border-radius:6px}}.checks input{{width:auto}}
.drop{{border:2px dashed #9eb6cf;padding:18px;border-radius:10px;text-align:center;cursor:pointer;background:#fafcff}}
.btn{{padding:9px 12px;background:#1677ff;color:#fff;border:0;border-radius:8px;text-decoration:none;cursor:pointer;display:inline-block}}
.msg{{margin-top:10px;background:#fff8e8;padding:9px;border-radius:8px}}.req{{color:#b42318}}
@media(max-width:800px){{.g,.checks{{grid-template-columns:1fr}}}}
</style></head>
<body><header><b>{title}</b><br><small>Area + Rent required · Multiple property types · Photos/Videos/Brochure optional</small></header>
<div class="w">
<a class="btn" href="/final-dashboard-v6">← Dashboard</a>
<a class="btn" href="/operational-matcher?division={d}">Matcher</a><br><br>
<form id="f" enctype="multipart/form-data"><input type="hidden" name="division" value="{d}">
<div class="card"><div class="g">
<input name="property_name" placeholder="Property Name">
<input name="city" value="{'Goa' if d=='GOA' else 'Delhi NCR'}">
<input name="location" placeholder="Location *" required>
<input name="google_location" placeholder="Google Maps location/link">
<input name="area_sqft" placeholder="Area sq ft *" required inputmode="decimal">
<input name="rent_amount" placeholder="Rent amount *" required inputmode="decimal">
<select name="rent_unit"><option>MONTH</option><option>SQFT_MONTH</option></select>
<select name="transaction_type"><option>LEASE</option><option>SALE</option><option>LEASE_OR_SALE</option></select>
<input name="floor" placeholder="Floor"><input name="frontage" placeholder="Frontage">
<input name="parking" placeholder="Parking"><input name="possession" placeholder="Possession">
<input name="owner_broker_name" placeholder="Owner/Broker/Contact Name"><input name="contact_number" placeholder="Contact Number">
<select name="contact_role"><option>UNVERIFIED</option><option>OWNER</option><option>BROKER</option><option>BOTH</option><option>OTHER</option></select>
<select name="verification_status"><option>UNVERIFIED</option><option>VERIFIED</option></select>
</div></div>
<div class="card"><b>Property Type — select one or more *</b><div class="checks">{_v172_types_html("ptype")}</div></div>
<div class="card"><input name="suitable_for" placeholder="Suitable For"><br><br><input name="nearby_brands" placeholder="Nearby Brands"><br><br><textarea name="remarks" placeholder="Remarks"></textarea></div>
<div class="card"><b>Photos — optional</b><div class="drop" id="idrop">Drag photos or click<input type="file" id="images" name="images" accept="image/*" multiple hidden></div><small id="ip">No photos selected</small></div>
<div class="card"><b>Videos — optional</b><div class="drop" id="vdrop">Drag videos or click<input type="file" id="videos" name="videos" accept="video/*" multiple hidden></div><small id="vp">No videos selected</small></div>
{brochure_html}
<button class="btn">Save Property</button><div id="msg" class="msg">Ready.</div>
</form></div>
<script>
function dz(box,input,p){{box.onclick=()=>input.click();['dragover','drop'].forEach(n=>box.addEventListener(n,e=>e.preventDefault()));box.addEventListener('drop',e=>{{input.files=e.dataTransfer.files;p.textContent=input.files.length+' file(s) selected'}});input.onchange=()=>p.textContent=input.files.length+' file(s) selected'}}
dz(idrop,images,ip);dz(vdrop,videos,vp);dz(bdrop,brochure,bp);
f.onsubmit=async e=>{{e.preventDefault();let pts=[...document.querySelectorAll('[name=ptype]:checked')].map(x=>x.value);if(!pts.length){{msg.textContent='Select at least one Property Type.';return}}let fd=new FormData(f);fd.set('property_types',pts.join('|'));msg.textContent='Saving...';let r=await fetch('/api/v17-2/property/save',{{method:'POST',body:fd}}),d=await r.json();msg.textContent=r.ok?`Saved ${{d.property_code}} · photos ${{d.images_saved}} · videos ${{d.videos_saved}} · brochure ${{d.brochure_saved}}`:'ERROR: '+(d.detail||d.message||'Save failed')}}
</script></body></html>""")

@app.get("/matcher-final",response_class=HTMLResponse)
def v172_matcher(req:Request,division:str=Query("DELHI_NCR")):
    if not page_role_or_redirect(req):
        return RedirectResponse("/login",303)
    d=division.upper()
    return HTMLResponse(f"""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Matcher</title>
<style>body{{font-family:Arial;background:#f4f7fb;margin:0;color:#172437}}header{{background:#102235;color:#fff;padding:18px}}.w{{padding:18px}}.btn{{padding:9px 11px;background:#1677ff;color:#fff;border:0;border-radius:8px;text-decoration:none;cursor:pointer}}select{{padding:9px;min-width:380px}}table{{width:100%;border-collapse:collapse;background:#fff;font-size:12px;margin-top:12px}}th,td{{padding:8px;border-bottom:1px solid #eee;text-align:left;vertical-align:top}}.msg{{padding:9px;background:#fff8e8;margin-top:10px}}</style></head><body>
<header><b>{'Goa' if d=='GOA' else 'Delhi NCR'} Property Matcher</b><br><small>Clean V17 operational matcher</small></header><div class=w>
<a class=btn href="/final-dashboard-v6">← Dashboard</a> <a class=btn href="/property-form-final?division={d}">Add Property</a> <a class=btn href="/operational-requirement-form?division={d}">Add Requirement</a><br><br>
<select id=reqs><option>Loading...</option></select> <button class=btn onclick=run()>Run Match</button><div id=msg class=msg>Select a requirement.</div>
<table><thead><tr><th>Score</th><th>Property</th><th>Type</th><th>Location</th><th>Area</th><th>Rent</th><th>Contact</th><th>Verified</th><th>Google</th><th>Brochure</th><th>Reasons</th></tr></thead><tbody id=rows></tbody></table></div>
<script>
async function load(){{let d=await(await fetch('/api/v17/requirements?division={d}')).json();reqs.innerHTML='<option value="">Select requirement</option>'+(d.rows||[]).map(x=>`<option value="${{x.requirement_code}}">${{x.requirement_code}} · ${{x.company_name||x.client_name||''}} · ${{x.preferred_locations}}</option>`).join('')}}
async function brochure(code){{let d=await(await fetch('/api/v17-2/property/'+encodeURIComponent(code)+'/media')).json();let b=(d.rows||[]).find(x=>x.media_type==='BROCHURE');return b?`<a target="_blank" href="/api/v17-2/property-media/${{b.id}}">View Brochure</a>`:''}}
async function run(){{if(!reqs.value)return;let r=await fetch('/api/v17/match/'+reqs.value,{{method:'POST'}}),d=await r.json();msg.textContent=r.ok?d.count+' matches found':'ERROR: '+(d.detail||d.message||'Matcher failed');let html='';for(const x of (d.matches||[])){{html+=`<tr><td><b>${{x.score}}</b><br>${{x.match_band}}</td><td>${{x.property_name||x.property_code}}</td><td>${{(x.property_types||[]).join(', ')}}</td><td>${{x.location||''}}</td><td>${{x.area_sqft}}</td><td>${{x.rent_amount}}</td><td>${{x.owner_broker_name||''}}<br><b>${{x.contact_number||''}}</b></td><td>${{x.verification_status||''}}</td><td>${{x.google_location?`<a target="_blank" href="${{x.google_location}}">Map</a>`:''}}</td><td>${{await brochure(x.property_code)}}</td><td>${{(x.reasons||[]).join(' · ')}}</td></tr>`}}rows.innerHTML=html||'<tr><td colspan=11>No matches found.</td></tr>'}}
load()
</script></body></html>""")

@app.get("/final-dashboard-v6",response_class=HTMLResponse)
def v172_dashboard(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    admin = '<a class="card" href="/admin-data-tools-v2"><b>Admin Data Tools</b><p>Database health and maintenance.</p></a>' if role=="admin" else ""
    search = "/property-discovery" if _v172_route_exists("/property-discovery") else "/final-dashboard-v3/property-discovery"
    return HTMLResponse(f"""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Final Dashboard</title>
<style>body{{font-family:Arial;background:#f4f7fb;margin:0;color:#172437}}header{{background:#102235;color:#fff;padding:22px}}.w{{max-width:1550px;margin:auto;padding:20px}}.g{{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:12px;margin-bottom:24px}}.card{{display:block;background:#fff;border:1px solid #e2e8f0;border-radius:12px;padding:15px;text-decoration:none;color:#172437}}.card p{{font-size:12px;color:#687789}}.primary{{border:2px solid #1677ff}}.search{{border:2px solid #7c3aed}}</style></head><body>
<header><b>AI Deal Intelligence OS</b><br><small>FINAL EXECUTION DASHBOARD</small></header><div class=w>

<h2>Delhi NCR</h2><div class=g>
<a class="card primary" href="/property-form-final?division=DELHI_NCR"><b>Add Property</b><p>Area + Rent required, multiple property types, Google Location, optional photo/video/brochure.</p></a>
<a class="card primary" href="/operational-requirement-form?division=DELHI_NCR"><b>Add Requirement</b><p>Confirmed manual demand.</p></a>
<a class="card primary" href="/matcher-final?division=DELHI_NCR"><b>Run Matcher</b><p>Clean operational matcher with contacts and brochure link.</p></a>
<a class=card href="/operational-inventory?division=DELHI_NCR"><b>Fresh Inventory</b><p>Operational inventory.</p></a>
</div>

<h2>Goa</h2><div class=g>
<a class="card primary" href="/property-form-final?division=GOA"><b>Add Goa Property</b><p>Includes optional brochure upload.</p></a>
<a class="card primary" href="/operational-requirement-form?division=GOA"><b>Add Goa Requirement</b><p>Manual Goa requirement.</p></a>
<a class="card primary" href="/matcher-final?division=GOA"><b>Goa Matcher</b><p>Contacts, map and brochure shown with results.</p></a>
<a class=card href="/operational-inventory?division=GOA"><b>Goa Inventory</b><p>Goa operational inventory.</p></a>
</div>

<h2>Search & Discovery</h2><div class=g>
<a class="card search" href="{search}"><b>Property Discovery / Search Engine</b><p>Existing V17 property discovery engine.</p></a>
<a class=card href="/requirements-match-center"><b>Requirements Centre</b><p>AI and manual requirements.</p></a>
<a class=card href="/retail-expansion"><b>Retail Expansion</b><p>Retail AI signals.</p></a>
<a class=card href="/capture-intelligence"><b>Capture Property</b><p>Camera, screenshot, handwritten, magazine and PDF.</p></a>
</div>

<h2>AI Bots & Marketing</h2><div class=g>
<a class=card href="/final-dashboard-v3"><b>Bot Controls</b><p>Hospitality and Retail bot controls.</p></a>
<a class=card href="/ai-hospitality-master-final"><b>Hospitality Master</b><p>Hospitality database.</p></a>
<a class=card href="/hospitality-enrichment"><b>Find Missing Hospitality Contacts</b><p>Phone-first enrichment.</p></a>
<a class=card href="/marketing-contacts-final"><b>Marketing Contacts</b><p>WhatsApp marketing database.</p></a>
<a class=card href="/phone-contact-upload"><b>Upload Phone Contacts</b><p>VCF/CSV/XLSX.</p></a>
<a class=card href="/api/v16/whatsapp-ready.csv"><b>Export WhatsApp CSV</b><p>Export usable Hospitality mobiles.</p></a>
</div>

<h2>Database & Admin</h2><div class=g>
<a class=card href="/property-database"><b>Full Property Database</b><p>Legacy/master archive.</p></a>
<a class=card href="/contacts-directory"><b>Owner / Broker Contacts</b><p>Property contact verification.</p></a>
<a class=card href="/data-doctor"><b>Data Doctor</b><p>Reconciliation and database health.</p></a>
{admin}
</div>

</div></body></html>""")

@app.get("/api/v17-2/system-audit")
def v172_system_audit(req:Request):
    need_login(req)
    checks={
        "property_form":_v172_route_exists("/property-form-final"),
        "property_save":_v172_route_exists("/api/v17-2/property/save"),
        "requirement_form":_v172_route_exists("/operational-requirement-form"),
        "matcher":_v172_route_exists("/matcher-final"),
        "search_engine":_v172_route_exists("/property-discovery") or _v172_route_exists("/final-dashboard-v3/property-discovery"),
        "hospitality_master":_v172_route_exists("/ai-hospitality-master-final"),
        "marketing_contacts":_v172_route_exists("/marketing-contacts-final"),
        "phone_upload":_v172_route_exists("/phone-contact-upload"),
        "admin_tools":_v172_route_exists("/admin-data-tools-v2"),
        "data_doctor":_v172_route_exists("/data-doctor")
    }
    return {"status":"ok","all_ready":all(checks.values()),"checks":checks}

@app.middleware("http")
async def v172_final_router(request,call_next):
    p=request.url.path
    if p in {"/workspace","/final-dashboard","/final-dashboard-v2","/final-dashboard-v3","/final-dashboard-v4","/final-dashboard-v5"}:
        return RedirectResponse("/final-dashboard-v6",status_code=307)
    if p in {"/operational-property-form","/property-manual","/v14-property-form"}:
        q=request.url.query
        div="GOA" if "division=GOA" in q.upper() else "DELHI_NCR"
        return RedirectResponse(f"/property-form-final?division={div}",status_code=307)
    if p in {"/operational-matcher","/v14-matcher","/goa-matcher"}:
        q=request.url.query
        div="GOA" if ("division=GOA" in q.upper() or p=="/goa-matcher") else "DELHI_NCR"
        return RedirectResponse(f"/matcher-final?division={div}",status_code=307)
    response=await call_next(request)
    if p.startswith(("/final-dashboard-v6","/property-form-final","/matcher-final")):
        response.headers["Cache-Control"]="no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"]="no-cache"
        response.headers["Expires"]="0"
    return response

# ============================================================
# V17.4 FINAL TEMPLATE
# Final consolidated repair:
# - Fresh Inventory working for Delhi NCR + Goa
# - Manual vs AI source clearly separated
# - Manual data saved in clean operational DB
# - Entry Source / Entry Date / Entered By / Verification shown
# - AI requirements must be manually confirmed before matching
# - Manual requirements can run matcher
# - Rent field starts blank for manual team entry
# - Goa brochure upload retained
# - Final unified dashboard
# ============================================================

def _v174_setup():
    _v17_setup()
    with engine.begin() as c:
        # Add source metadata to operational properties.
        for stmt in [
            "ALTER TABLE pi_operational_properties ADD COLUMN IF NOT EXISTS entry_source TEXT DEFAULT 'MANUAL'",
            "ALTER TABLE pi_operational_properties ADD COLUMN IF NOT EXISTS entered_by TEXT",
            "ALTER TABLE pi_operational_properties ADD COLUMN IF NOT EXISTS entry_date TIMESTAMPTZ DEFAULT NOW()",
            "ALTER TABLE pi_operational_requirements ADD COLUMN IF NOT EXISTS entry_source TEXT DEFAULT 'MANUAL'",
            "ALTER TABLE pi_operational_requirements ADD COLUMN IF NOT EXISTS entered_by TEXT",
            "ALTER TABLE pi_operational_requirements ADD COLUMN IF NOT EXISTS entry_date TIMESTAMPTZ DEFAULT NOW()",
        ]:
            c.execute(text(stmt))

def _v174_num(v):
    if v in (None,""): return None
    try:
        x=float(str(v).replace(",","").strip())
        return x if x>0 else None
    except Exception:
        return None

@app.post("/api/v17-4/property/save")
async def v174_property_save(
    req:Request,
    division:str=Form("DELHI_NCR"),
    property_name:str=Form(""),
    property_types:str=Form(...),
    city:str=Form(""),
    location:str=Form(...),
    google_location:str=Form(""),
    area_sqft:str=Form(...),
    rent_amount:str=Form(...),
    rent_unit:str=Form("MONTH"),
    transaction_type:str=Form("LEASE"),
    floor:str=Form(""),
    frontage:str=Form(""),
    parking:str=Form(""),
    possession:str=Form(""),
    suitable_for:str=Form(""),
    nearby_brands:str=Form(""),
    owner_broker_name:str=Form(""),
    contact_number:str=Form(""),
    contact_role:str=Form("UNVERIFIED"),
    verification_status:str=Form("UNVERIFIED"),
    remarks:str=Form(""),
    images:list[UploadFile]=File(default=[]),
    videos:list[UploadFile]=File(default=[]),
    brochure:UploadFile|None=File(default=None)
):
    need_login(req); _v174_setup()
    pts=[x.strip() for x in str(property_types or "").split("|") if x.strip()]
    if not pts:
        raise HTTPException(400,"Select at least one Property Type.")
    if not str(location or "").strip():
        raise HTTPException(400,"Location is required.")
    area=_v17_posnum(area_sqft,"Area")
    rent=_v17_posnum(rent_amount,"Rent")
    code=_v17_code("PROP")
    who=actor_name(req)

    with engine.begin() as c:
        c.execute(text("""INSERT INTO pi_operational_properties(
            property_code,division,property_name,property_types,city,location,google_location,area_sqft,rent_amount,
            rent_unit,transaction_type,floor,frontage,parking,possession,suitable_for,nearby_brands,
            owner_broker_name,contact_number,contact_role,verification_status,remarks,created_by,
            entry_source,entered_by,entry_date
        ) VALUES(
            :code,:div,:name,CAST(:types AS jsonb),:city,:loc,:google,:area,:rent,:ru,:tt,:floor,:front,
            :park,:poss,:suitable,:nearby,:person,:phone,:role,:ver,:remarks,:by,
            'MANUAL',:entered_by,NOW()
        )"""),{
            "code":code,"div":division.upper(),"name":property_name or None,"types":json.dumps(pts),
            "city":city or None,"loc":location.strip(),"google":google_location or None,
            "area":area,"rent":rent,"ru":rent_unit,"tt":transaction_type,"floor":floor or None,
            "front":frontage or None,"park":parking or None,"poss":possession or None,
            "suitable":suitable_for or None,"nearby":nearby_brands or None,
            "person":owner_broker_name or None,"phone":contact_number or None,
            "role":contact_role,"ver":verification_status,"remarks":remarks or None,
            "by":who,"entered_by":who
        })

    saved={"IMAGE":0,"VIDEO":0,"BROCHURE":0}
    async def save_file(f,typ,limit):
        if not f or not getattr(f,"filename",None): return
        data=await f.read()
        if not data: return
        if len(data)>limit:
            raise HTTPException(413,f"{f.filename} is too large.")
        with engine.begin() as c:
            c.execute(text("""INSERT INTO pi_operational_property_media(
                property_code,media_type,filename,mime_type,file_size,content
            ) VALUES(:p,:t,:f,:m,:s,:b)"""),{
                "p":code,"t":typ,"f":f.filename,"m":f.content_type,"s":len(data),"b":data
            })
        saved[typ]+=1

    for f in images or []: await save_file(f,"IMAGE",12*1024*1024)
    for f in videos or []: await save_file(f,"VIDEO",80*1024*1024)
    if brochure and brochure.filename:
        if brochure.content_type not in {
            "application/pdf",
            "application/msword",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        } and not brochure.filename.lower().endswith((".pdf",".doc",".docx")):
            raise HTTPException(400,"Brochure must be PDF, DOC or DOCX.")
        await save_file(brochure,"BROCHURE",30*1024*1024)

    return {"status":"ok","property_code":code,"saved":saved}

@app.post("/api/v17-4/requirement/save")
async def v174_requirement_save(req:Request):
    need_login(req); _v174_setup()
    b=await req.json()
    loc=str(b.get("preferred_locations") or "").strip()
    if not loc: raise HTTPException(400,"Preferred Locations are required.")
    mina=_v17_posnum(b.get("minimum_area_sqft"),"Minimum Area")
    maxa=_v17_posnum(b.get("maximum_area_sqft"),"Maximum Area")
    if maxa<mina: mina,maxa=maxa,mina
    maxr=None
    if str(b.get("maximum_rent") or "").strip():
        maxr=_v17_posnum(b.get("maximum_rent"),"Maximum Rent")
    code=_v17_code("REQ")
    who=actor_name(req)
    with engine.begin() as c:
        c.execute(text("""INSERT INTO pi_operational_requirements(
            requirement_code,division,client_name,company_name,contact_number,requirement_types,city,
            preferred_locations,minimum_area_sqft,maximum_area_sqft,maximum_rent,transaction_type,
            additional_points,verification_status,created_by,entry_source,entered_by,entry_date
        ) VALUES(
            :code,:div,:client,:company,:phone,CAST(:types AS jsonb),:city,:loc,:mina,:maxa,:rent,:tt,:pts,:ver,:by,
            'MANUAL',:entered_by,NOW()
        )"""),{
            "code":code,"div":str(b.get("division") or "DELHI_NCR").upper(),
            "client":b.get("client_name"),"company":b.get("company_name"),"phone":b.get("contact_number"),
            "types":json.dumps(b.get("requirement_types") or []),"city":b.get("city"),
            "loc":loc,"mina":mina,"maxa":maxa,"rent":maxr,
            "tt":b.get("transaction_type") or "LEASE","pts":b.get("additional_points"),
            "ver":b.get("verification_status") or "VERIFIED","by":who,"entered_by":who
        })
    return {"status":"ok","requirement_code":code}

@app.get("/api/v17-4/properties")
def v174_properties(req:Request,division:str=Query("DELHI_NCR"),source:str=Query("ALL"),verified:str=Query("ALL"),q:str=Query("")):
    need_login(req); _v174_setup()
    wh=["p.division=:d"]; p={"d":division.upper()}
    if source!="ALL":
        wh.append("COALESCE(p.entry_source,'MANUAL')=:source"); p["source"]=source
    if verified!="ALL":
        wh.append("p.verification_status=:verified"); p["verified"]=verified
    if q.strip():
        wh.append("""(
            COALESCE(p.property_code,'') ILIKE :q OR COALESCE(p.property_name,'') ILIKE :q OR
            COALESCE(p.location,'') ILIKE :q OR COALESCE(p.city,'') ILIKE :q OR
            COALESCE(p.contact_number,'') ILIKE :q OR COALESCE(p.owner_broker_name,'') ILIKE :q
        )"""); p["q"]="%"+q.strip()+"%"
    sql="""SELECT p.*,
        COALESCE(p.entry_source,'MANUAL') AS entry_source,
        COALESCE(p.entered_by,p.created_by) AS entered_by,
        COALESCE(p.entry_date,p.created_at) AS entry_date,
        (SELECT COUNT(*) FROM pi_operational_property_media m WHERE m.property_code=p.property_code AND m.media_type='IMAGE') image_count,
        (SELECT COUNT(*) FROM pi_operational_property_media m WHERE m.property_code=p.property_code AND m.media_type='VIDEO') video_count,
        (SELECT COUNT(*) FROM pi_operational_property_media m WHERE m.property_code=p.property_code AND m.media_type='BROCHURE') brochure_count
        FROM pi_operational_properties p WHERE """+" AND ".join(wh)+" ORDER BY p.id DESC LIMIT 5000"
    with engine.connect() as c:
        rows=[dict(r._mapping) for r in c.execute(text(sql),p).fetchall()]
    return {"status":"ok","rows":rows}

@app.get("/fresh-inventory-final",response_class=HTMLResponse)
def v174_inventory(req:Request,division:str=Query("DELHI_NCR")):
    if not page_role_or_redirect(req): return RedirectResponse("/login",303)
    d=division.upper()
    title="Goa Fresh Inventory" if d=="GOA" else "Delhi NCR Fresh Inventory"
    return HTMLResponse(f"""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1">
<title>{title}</title><style>
body{{font-family:Arial;background:#f4f7fb;margin:0;color:#172437}}header{{background:#102235;color:#fff;padding:18px}}
.w{{padding:18px}}.bar{{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px}}.btn{{padding:9px 11px;background:#1677ff;color:white;text-decoration:none;border-radius:8px;border:0;cursor:pointer}}
input,select{{padding:9px;border:1px solid #ccd6e2;border-radius:7px}}input{{min-width:300px}}.tablewrap{{overflow:auto;background:white;border-radius:10px}}
table{{width:100%;border-collapse:collapse;font-size:12px}}th,td{{padding:8px;border-bottom:1px solid #eee;text-align:left;white-space:nowrap}}th{{background:#f8fafc;position:sticky;top:0}}
</style></head><body><header><b>{title}</b><br><small>Manual data clearly marked in Entry Source column</small></header><div class=w>
<div class=bar><a class=btn href="/final-dashboard-v8">← Dashboard</a><a class=btn href="/manual-property-final?division={d}">Add Property</a><a class=btn href="/matcher-final?division={d}">Matcher</a></div>
<div class=bar><input id=q placeholder="Search property, location, contact"><select id=source><option>ALL</option><option>MANUAL</option><option>AI</option><option>PROPERTY_DISCOVERY</option><option>IMPORTED</option><option>MAGAZINE</option></select><select id=verified><option>ALL</option><option>VERIFIED</option><option>UNVERIFIED</option></select><button class=btn onclick=load()>Search</button><span id=count></span></div>
<div class=tablewrap><table><thead><tr><th>Code</th><th>Property</th><th>Entry Source</th><th>Entry Date</th><th>Entered By</th><th>Verification</th><th>Types</th><th>Location</th><th>Area</th><th>Rent</th><th>Contact</th><th>Google</th><th>Photos</th><th>Videos</th><th>Brochure</th></tr></thead><tbody id=rows></tbody></table></div>
</div><script>
const E=x=>String(x??'').replace(/[&<>"]/g,c=>({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}}[c]));
async function load(){{let u='/api/v17-4/properties?division={d}&source='+source.value+'&verified='+verified.value+'&q='+encodeURIComponent(q.value||'');let d=await(await fetch(u)).json();let a=d.rows||[];count.textContent=a.length+' properties';let out='';for(const x of a){{let bd='';if((x.brochure_count||0)>0){{let m=await(await fetch('/api/v17-2/property/'+encodeURIComponent(x.property_code)+'/media')).json();let b=(m.rows||[]).find(y=>y.media_type==='BROCHURE');if(b)bd=`<a target=_blank href="/api/v17-2/property-media/${{b.id}}">View</a>`}}out+=`<tr><td>${{E(x.property_code)}}</td><td>${{E(x.property_name||'')}}</td><td><b>${{E(x.entry_source||'MANUAL')}}</b></td><td>${{E(String(x.entry_date||'').slice(0,16))}}</td><td>${{E(x.entered_by||'')}}</td><td>${{E(x.verification_status||'')}}</td><td>${{E((x.property_types||[]).join(', '))}}</td><td>${{E(x.location||'')}}</td><td>${{E(x.area_sqft)}}</td><td>${{E(x.rent_amount)}}</td><td>${{E(x.owner_broker_name||'')}}<br><b>${{E(x.contact_number||'')}}</b></td><td>${{x.google_location?`<a target=_blank href="${{E(x.google_location)}}">Map</a>`:''}}</td><td>${{x.image_count||0}}</td><td>${{x.video_count||0}}</td><td>${{bd}}</td></tr>`}}rows.innerHTML=out||'<tr><td colspan=15>No properties found.</td></tr>'}}
source.onchange=load;verified.onchange=load;load()
</script></body></html>""")

@app.get("/manual-property-final",response_class=HTMLResponse)
def v174_property_form(req:Request,division:str=Query("DELHI_NCR")):
    if not page_role_or_redirect(req): return RedirectResponse("/login",303)
    d=division.upper()
    checks="".join(f'<label><input type=checkbox name=ptype value="{escape(x)}"> {escape(x)}</label>' for x in _v17_types())
    return HTMLResponse(f"""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Property Form</title>
<style>body{{font-family:Arial;background:#f4f7fb;margin:0;color:#172437}}header{{background:#102235;color:#fff;padding:18px}}.w{{max-width:1150px;margin:auto;padding:18px}}.card{{background:#fff;padding:15px;border-radius:12px;margin-bottom:12px}}.g{{display:grid;grid-template-columns:1fr 1fr;gap:10px}}input,select,textarea{{width:100%;padding:9px;border:1px solid #ccd6e2;border-radius:7px}}.checks{{display:grid;grid-template-columns:repeat(3,1fr);gap:6px}}.checks label{{background:#f6f8fb;padding:6px;border-radius:6px}}.checks input{{width:auto}}.drop{{border:2px dashed #9eb6cf;padding:18px;border-radius:10px;text-align:center;cursor:pointer}}.btn{{padding:9px 12px;background:#1677ff;color:#fff;border:0;border-radius:8px;text-decoration:none;cursor:pointer}}.msg{{margin-top:10px;background:#fff8e8;padding:9px}}@media(max-width:800px){{.g,.checks{{grid-template-columns:1fr}}}}</style></head><body>
<header><b>{'Goa' if d=='GOA' else 'Delhi NCR'} Manual Property Form</b><br><small>Saved with Entry Source = MANUAL · Rent starts blank for team entry</small></header><div class=w>
<a class=btn href="/final-dashboard-v8">← Dashboard</a><br><br><form id=f enctype=multipart/form-data autocomplete=off><input type=hidden name=division value="{d}">
<div class=card><div class=g><input name=property_name placeholder="Property Name"><input name=city value="{'Goa' if d=='GOA' else 'Delhi NCR'}">
<input name=location placeholder="Location *" required><input name=google_location placeholder="Google Maps location/link">
<input name=area_sqft value="" autocomplete=off placeholder="Area sq ft * — type manually" required><input name=rent_amount value="" autocomplete=off placeholder="Rent amount * — type manually" required>
<select name=rent_unit><option>MONTH</option><option>SQFT_MONTH</option></select><select name=transaction_type><option>LEASE</option><option>SALE</option><option>LEASE_OR_SALE</option></select>
<input name=floor placeholder="Floor"><input name=frontage placeholder="Frontage"><input name=parking placeholder="Parking"><input name=possession placeholder="Possession">
<input name=owner_broker_name placeholder="Owner/Broker/Contact Name"><input name=contact_number placeholder="Contact Number">
<select name=contact_role><option>UNVERIFIED</option><option>OWNER</option><option>BROKER</option><option>BOTH</option></select><select name=verification_status><option>UNVERIFIED</option><option>VERIFIED</option></select></div></div>
<div class=card><b>Property Type — select multiple *</b><div class=checks>{checks}</div></div>
<div class=card><input name=suitable_for placeholder="Suitable For"><br><br><input name=nearby_brands placeholder="Nearby Brands"><br><br><textarea name=remarks placeholder="Remarks"></textarea></div>
<div class=card><b>Photos optional</b><div class=drop id=idrop>Drag photos or click<input type=file id=images name=images accept="image/*" multiple hidden></div><small id=ip>No photos selected</small></div>
<div class=card><b>Videos optional</b><div class=drop id=vdrop>Drag videos or click<input type=file id=videos name=videos accept="video/*" multiple hidden></div><small id=vp>No videos selected</small></div>
<div class=card><b>Brochure optional</b><div class=drop id=bdrop>Drag PDF/DOC/DOCX or click<input type=file id=brochure name=brochure accept=".pdf,.doc,.docx,application/pdf" hidden></div><small id=bp>No brochure selected</small></div>
<button class=btn>Save Property</button><div id=msg class=msg>Ready.</div></form></div>
<script>
function dz(box,input,p){{box.onclick=()=>input.click();['dragover','drop'].forEach(n=>box.addEventListener(n,e=>e.preventDefault()));box.addEventListener('drop',e=>{{input.files=e.dataTransfer.files;p.textContent=input.files.length+' file(s) selected'}});input.onchange=()=>p.textContent=input.files.length+' file(s) selected'}}dz(idrop,images,ip);dz(vdrop,videos,vp);dz(bdrop,brochure,bp);
f.onsubmit=async e=>{{e.preventDefault();let pts=[...document.querySelectorAll('[name=ptype]:checked')].map(x=>x.value);if(!pts.length){{msg.textContent='Select at least one Property Type.';return}}let fd=new FormData(f);fd.set('property_types',pts.join('|'));msg.textContent='Saving...';let r=await fetch('/api/v17-4/property/save',{{method:'POST',body:fd}}),d=await r.json();msg.textContent=r.ok?'Saved '+d.property_code+' · Entry Source MANUAL':'ERROR: '+(d.detail||d.message||'Save failed')}};</script></body></html>""")

@app.get("/manual-requirement-final",response_class=HTMLResponse)
def v174_requirement_form(req:Request,division:str=Query("DELHI_NCR")):
    if not page_role_or_redirect(req): return RedirectResponse("/login",303)
    d=division.upper(); checks="".join(f'<label><input type=checkbox name=rtype value="{escape(x)}"> {escape(x)}</label>' for x in _v17_types())
    return HTMLResponse(f"""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Requirement Form</title>
<style>body{{font-family:Arial;background:#f4f7fb;margin:0;color:#172437}}header{{background:#102235;color:#fff;padding:18px}}.w{{max-width:1050px;margin:auto;padding:18px}}.card{{background:#fff;padding:15px;border-radius:12px;margin-bottom:12px}}.g{{display:grid;grid-template-columns:1fr 1fr;gap:10px}}input,select,textarea{{width:100%;padding:9px;border:1px solid #ccd6e2;border-radius:7px}}.checks{{display:grid;grid-template-columns:repeat(3,1fr);gap:6px}}.checks label{{background:#f6f8fb;padding:6px;border-radius:6px}}.checks input{{width:auto}}.btn{{padding:9px 12px;background:#1677ff;color:#fff;border:0;border-radius:8px;text-decoration:none;cursor:pointer}}.msg{{margin-top:10px;background:#fff8e8;padding:9px}}@media(max-width:800px){{.g,.checks{{grid-template-columns:1fr}}}}</style></head><body>
<header><b>{'Goa' if d=='GOA' else 'Delhi NCR'} Manual Requirement Form</b><br><small>Saved with Entry Source = MANUAL</small></header><div class=w><a class=btn href="/final-dashboard-v8">← Dashboard</a><br><br>
<div class=card><div class=g><input id=client placeholder="Client Name"><input id=company placeholder="Company"><input id=phone placeholder="Contact Number"><input id=city value="{'Goa' if d=='GOA' else 'Delhi NCR'}">
<input id=loc placeholder="Preferred Locations *"><input id=mina placeholder="Minimum Area *"><input id=maxa placeholder="Maximum Area *"><input id=rent value="" autocomplete=off placeholder="Maximum Rent — type manually if applicable">
<select id=tt><option>LEASE</option><option>SALE</option><option>LEASE_OR_SALE</option></select><select id=ver><option>VERIFIED</option><option>UNVERIFIED</option></select></div></div>
<div class=card><b>Property Types</b><div class=checks>{checks}</div></div><div class=card><textarea id=pts placeholder="Additional Points"></textarea></div>
<button class=btn onclick=save()>Save Requirement</button><div id=msg class=msg>Ready.</div></div>
<script>async function save(){{let types=[...document.querySelectorAll('[name=rtype]:checked')].map(x=>x.value);let b={{division:'{d}',client_name:client.value,company_name:company.value,contact_number:phone.value,requirement_types:types,city:city.value,preferred_locations:loc.value,minimum_area_sqft:mina.value,maximum_area_sqft:maxa.value,maximum_rent:rent.value,transaction_type:tt.value,additional_points:pts.value,verification_status:ver.value}};let r=await fetch('/api/v17-4/requirement/save',{{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify(b)}}),d=await r.json();msg.innerHTML=r.ok?'Saved '+d.requirement_code+' · Entry Source MANUAL · <a href="/matcher-final?division={d}">Open Matcher</a>':'ERROR: '+(d.detail||d.message||'Save failed')}};</script></body></html>""")

@app.get("/final-dashboard-v8",response_class=HTMLResponse)
def v174_dashboard(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    admin='<a class=card href="/admin-data-tools-v2"><b>Admin Data Tools</b><p>Database health and tools.</p></a>' if role=="admin" else ""
    return HTMLResponse(f"""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Final Dashboard</title>
<style>body{{font-family:Arial;background:#f4f7fb;margin:0;color:#172437}}header{{background:#102235;color:#fff;padding:22px}}.w{{max-width:1500px;margin:auto;padding:20px}}.g{{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:12px;margin-bottom:24px}}.card{{display:block;background:#fff;border:1px solid #e2e8f0;border-radius:12px;padding:15px;text-decoration:none;color:#172437}}.card p{{font-size:12px;color:#687789}}.primary{{border:2px solid #1677ff}}</style></head><body><header><b>AI Deal Intelligence OS</b><br><small>FINAL EXECUTION DASHBOARD</small></header><div class=w>
<h2>Delhi NCR</h2><div class=g><a class="card primary" href="/manual-property-final?division=DELHI_NCR"><b>Add Property Manually</b><p>Saved to clean operational database with Entry Source = MANUAL.</p></a><a class="card primary" href="/manual-requirement-final?division=DELHI_NCR"><b>Add Requirement Manually</b><p>Saved with Entry Source = MANUAL.</p></a><a class="card primary" href="/matcher-final?division=DELHI_NCR"><b>Run Matcher</b><p>Clean operational matcher.</p></a><a class=card href="/fresh-inventory-final?division=DELHI_NCR"><b>Fresh Inventory</b><p>Shows Entry Source, Entry Date, Entered By and Verification.</p></a></div>
<h2>Goa</h2><div class=g><a class="card primary" href="/manual-property-final?division=GOA"><b>Add Goa Property</b><p>Includes optional brochure upload.</p></a><a class="card primary" href="/manual-requirement-final?division=GOA"><b>Add Goa Requirement</b><p>Saved with Entry Source = MANUAL.</p></a><a class="card primary" href="/matcher-final?division=GOA"><b>Goa Matcher</b><p>Clean Goa matcher.</p></a><a class=card href="/fresh-inventory-final?division=GOA"><b>Goa Fresh Inventory</b><p>Shows source metadata.</p></a></div>
<h2>Requirements & Search</h2><div class=g><a class=card href="/requirements-center-final"><b>Requirements Centre</b><p>AI and manual separated. AI must be manually confirmed before match.</p></a><a class=card href="/property-discovery"><b>Property Discovery / Search Engine</b><p>Existing search engine.</p></a><a class=card href="/retail-expansion"><b>Retail Expansion</b><p>Retail AI signals.</p></a><a class=card href="/capture-intelligence"><b>Capture Property</b><p>Screenshot, camera, magazine, PDF.</p></a></div>
<h2>AI & Marketing</h2><div class=g><a class=card href="/final-dashboard-v3"><b>Bot Controls</b><p>Hospitality and Retail bots.</p></a><a class=card href="/ai-hospitality-master-final"><b>Hospitality Master</b><p>Hospitality data.</p></a><a class=card href="/hospitality-enrichment"><b>Find Missing Contacts</b><p>Phone-first enrichment.</p></a><a class=card href="/marketing-contacts-final"><b>Marketing Contacts</b><p>WhatsApp marketing database.</p></a><a class=card href="/phone-contact-upload"><b>Upload Phone Contacts</b><p>VCF/CSV/XLSX.</p></a></div>
<h2>Database</h2><div class=g><a class=card href="/property-database"><b>Full Property Database</b><p>Legacy archive.</p></a><a class=card href="/contacts-directory"><b>Owner / Broker Contacts</b><p>Contact verification.</p></a><a class=card href="/data-doctor"><b>Data Doctor</b><p>Database health.</p></a>{admin}</div>
</div></body></html>""")

@app.middleware("http")
async def v174_router(request,call_next):
    p=request.url.path
    if p in {"/workspace","/final-dashboard","/final-dashboard-v2","/final-dashboard-v4","/final-dashboard-v5","/final-dashboard-v6","/final-dashboard-v7"}:
        return RedirectResponse("/final-dashboard-v8",status_code=307)
    if p=="/requirements-match-center":
        return RedirectResponse("/requirements-center-final",status_code=307)
    if p in {"/operational-property-form","/property-form-final","/property-form-v17-3","/property-manual","/v14-property-form"}:
        div="GOA" if "DIVISION=GOA" in request.url.query.upper() else "DELHI_NCR"
        return RedirectResponse(f"/manual-property-final?division={div}",status_code=307)
    if p in {"/operational-requirement-form","/v14-requirement-form"}:
        div="GOA" if "DIVISION=GOA" in request.url.query.upper() else "DELHI_NCR"
        return RedirectResponse(f"/manual-requirement-final?division={div}",status_code=307)
    if p in {"/operational-inventory"}:
        div="GOA" if "DIVISION=GOA" in request.url.query.upper() else "DELHI_NCR"
        return RedirectResponse(f"/fresh-inventory-final?division={div}",status_code=307)
    response=await call_next(request)
    if p.startswith(("/final-dashboard-v8","/manual-property-final","/manual-requirement-final","/fresh-inventory-final")):
        response.headers["Cache-Control"]="no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"]="no-cache"
        response.headers["Expires"]="0"
    return response

# ============================================================
# V17.5 DATA RECOVERY DOCTOR + IMMUTABLE DATA VAULT
# Purpose:
# 1) Recover manual property/requirement records that were saved into older tables.
# 2) Make Fresh Inventory read the clean operational DB after recovery.
# 3) Restore a working Requirements Centre.
# 4) Protect core property/requirement data from hard DELETE/TRUNCATE.
# 5) Keep immutable JSON snapshots of every insert/update/delete attempt.
# ============================================================

def _v175_setup():
    _v174_setup()
    with engine.begin() as c:
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_data_vault_events(
            id BIGSERIAL PRIMARY KEY,
            entity_type TEXT NOT NULL,
            entity_key TEXT,
            action TEXT NOT NULL,
            source_table TEXT,
            snapshot JSONB,
            actor TEXT,
            created_at TIMESTAMPTZ DEFAULT NOW()
        )"""))
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_recovery_log(
            id BIGSERIAL PRIMARY KEY,
            entity_type TEXT NOT NULL,
            source_table TEXT,
            source_row_id TEXT,
            recovered_key TEXT,
            fingerprint TEXT,
            status TEXT,
            detail TEXT,
            created_at TIMESTAMPTZ DEFAULT NOW(),
            UNIQUE(entity_type,source_table,source_row_id,fingerprint)
        )"""))
        c.execute(text("""CREATE INDEX IF NOT EXISTS idx_vault_entity ON pi_data_vault_events(entity_type,entity_key)"""))
        c.execute(text("""CREATE INDEX IF NOT EXISTS idx_recovery_status ON pi_recovery_log(status)"""))

        # generic audit trigger function
        c.execute(text("""
        CREATE OR REPLACE FUNCTION pi_core_data_vault_trigger()
        RETURNS trigger AS $$
        DECLARE snap jsonb;
        DECLARE k text;
        BEGIN
            IF TG_OP='DELETE' THEN
                snap=to_jsonb(OLD);
                k=COALESCE(OLD.property_code::text, OLD.requirement_code::text, OLD.id::text);
                INSERT INTO pi_data_vault_events(entity_type,entity_key,action,source_table,snapshot,actor)
                VALUES(TG_ARGV[0],k,'DELETE_BLOCKED',TG_TABLE_NAME,snap,current_user);
                RAISE EXCEPTION 'Hard delete blocked for protected core data. Use status/trash workflow instead.';
            ELSIF TG_OP='UPDATE' THEN
                snap=to_jsonb(OLD);
                k=COALESCE(OLD.property_code::text, OLD.requirement_code::text, OLD.id::text);
                INSERT INTO pi_data_vault_events(entity_type,entity_key,action,source_table,snapshot,actor)
                VALUES(TG_ARGV[0],k,'BEFORE_UPDATE',TG_TABLE_NAME,snap,current_user);
                RETURN NEW;
            ELSE
                snap=to_jsonb(NEW);
                k=COALESCE(NEW.property_code::text, NEW.requirement_code::text, NEW.id::text);
                INSERT INTO pi_data_vault_events(entity_type,entity_key,action,source_table,snapshot,actor)
                VALUES(TG_ARGV[0],k,'INSERT',TG_TABLE_NAME,snap,current_user);
                RETURN NEW;
            END IF;
        END;
        $$ LANGUAGE plpgsql;
        """))

        for table,entity in [
            ("pi_operational_properties","PROPERTY"),
            ("pi_operational_requirements","REQUIREMENT")
        ]:
            c.execute(text(f"DROP TRIGGER IF EXISTS trg_{table}_vault ON {table}"))
            c.execute(text(f"""CREATE TRIGGER trg_{table}_vault
                BEFORE INSERT OR UPDATE OR DELETE ON {table}
                FOR EACH ROW EXECUTE FUNCTION pi_core_data_vault_trigger('{entity}')"""))

        c.execute(text("""
        CREATE OR REPLACE FUNCTION pi_block_core_truncate()
        RETURNS trigger AS $$
        BEGIN
            RAISE EXCEPTION 'TRUNCATE blocked for protected core data.';
        END;
        $$ LANGUAGE plpgsql;
        """))
        for table in ["pi_operational_properties","pi_operational_requirements"]:
            c.execute(text(f"DROP TRIGGER IF EXISTS trg_{table}_no_truncate ON {table}"))
            c.execute(text(f"""CREATE TRIGGER trg_{table}_no_truncate
                BEFORE TRUNCATE ON {table}
                FOR EACH STATEMENT EXECUTE FUNCTION pi_block_core_truncate()"""))

def _v175_exists(table):
    try:
        with engine.connect() as c:
            return bool(c.execute(text("""SELECT 1 FROM information_schema.tables
              WHERE table_schema='public' AND table_name=:t"""),{"t":table}).first())
    except Exception:
        return False

def _v175_cols(table):
    try:
        with engine.connect() as c:
            return [r._mapping["column_name"] for r in c.execute(text("""SELECT column_name
              FROM information_schema.columns WHERE table_schema='public' AND table_name=:t"""),{"t":table}).fetchall()]
    except Exception:
        return []

def _v175_pick(row,*names):
    low={str(k).lower():v for k,v in row.items()}
    for name in names:
        v=low.get(name.lower())
        if v not in (None,""):
            return v
    return None

def _v175_num(v):
    if v in (None,""): return None
    try:
        x=float(str(v).replace(",","").strip())
        return x if x>0 else None
    except Exception:
        return None

def _v175_fingerprint(parts):
    import hashlib
    raw="|".join(str(x or "").strip().lower() for x in parts)
    return hashlib.sha256(raw.encode("utf-8","ignore")).hexdigest()

def _v175_candidate_property_tables():
    preferred=[
        "pi_properties",
        "properties",
        "property_inventory",
        "manual_properties",
        "pi_manual_properties",
        "v14_properties",
        "pi_property_inventory"
    ]
    return [t for t in preferred if _v175_exists(t)]

def _v175_candidate_requirement_tables():
    preferred=[
        "pi_requirements",
        "requirements",
        "pi_retail_manual_requirements",
        "pi_hospitality_manual_requirements",
        "manual_requirements",
        "v14_requirements"
    ]
    return [t for t in preferred if _v175_exists(t)]

def _v175_recover_properties():
    _v175_setup()
    recovered=0; skipped=0; reviewed=0; tables=[]
    for table in _v175_candidate_property_tables():
        cols=_v175_cols(table)
        try:
            with engine.connect() as c:
                rows=[dict(r._mapping) for r in c.execute(text(f'SELECT * FROM "{table}" ORDER BY 1 DESC LIMIT 10000')).fetchall()]
        except Exception:
            continue
        t_rec=t_skip=t_review=0
        for idx,r in enumerate(rows):
            rowid=str(_v175_pick(r,"id","property_id","property_code","record_id") or idx)
            location=_v175_pick(r,"location","locality","area","micro_market")
            area=_v175_num(_v175_pick(r,"area_sqft","available_area","area","size_sqft","super_area","builtup_area"))
            rent=_v175_num(_v175_pick(r,"rent_amount","rent","monthly_rent","asking_rent"))
            if not location or not area or not rent:
                t_review+=1; reviewed+=1
                continue
            name=_v175_pick(r,"property_name","name","building_name","title")
            city=_v175_pick(r,"city") or ("Goa" if "goa" in str(location).lower() else "Delhi NCR")
            division="GOA" if ("goa" in str(city).lower() or "goa" in str(location).lower()) else "DELHI_NCR"
            phone=_v175_pick(r,"contact_number","owner_contact","broker_contact","phone","mobile")
            person=_v175_pick(r,"owner_name","broker_name","contact_name","owner_broker_name")
            ptype=_v175_pick(r,"property_type","type","category")
            fp=_v175_fingerprint([division,name,location,area,rent,phone])

            with engine.connect() as c:
                dup=c.execute(text("""SELECT property_code FROM pi_operational_properties
                    WHERE division=:d AND lower(COALESCE(location,''))=lower(:loc)
                    AND area_sqft=:area AND rent_amount=:rent
                    AND COALESCE(contact_number,'')=COALESCE(:phone,'')
                    LIMIT 1"""),{"d":division,"loc":str(location),"area":area,"rent":rent,"phone":phone}).first()
            if dup:
                t_skip+=1; skipped+=1
                continue

            code=_v17_code("RECOVERED-PROP")
            with engine.begin() as c:
                c.execute(text("""INSERT INTO pi_operational_properties(
                    property_code,division,property_name,property_types,city,location,area_sqft,rent_amount,
                    transaction_type,owner_broker_name,contact_number,verification_status,remarks,created_by,
                    entry_source,entered_by,entry_date
                ) VALUES(:code,:div,:name,CAST(:types AS jsonb),:city,:loc,:area,:rent,'LEASE',
                    :person,:phone,'UNVERIFIED',:remarks,'RECOVERY_DOCTOR',
                    'RECOVERED_MANUAL','RECOVERY_DOCTOR',NOW())"""),{
                    "code":code,"div":division,"name":name,"types":json.dumps([str(ptype)] if ptype else []),
                    "city":city,"loc":str(location),"area":area,"rent":rent,
                    "person":person,"phone":phone,
                    "remarks":f"Recovered safely from historical table {table}; original row preserved."
                })
                c.execute(text("""INSERT INTO pi_recovery_log(
                    entity_type,source_table,source_row_id,recovered_key,fingerprint,status,detail
                ) VALUES('PROPERTY',:t,:rid,:key,:fp,'RECOVERED',:detail)
                ON CONFLICT DO NOTHING"""),{
                    "t":table,"rid":rowid,"key":code,"fp":fp,"detail":"Copied into clean operational property database; source row not modified."
                })
            t_rec+=1; recovered+=1
        tables.append({"table":table,"recovered":t_rec,"skipped_duplicates":t_skip,"needs_review":t_review})
    return {"recovered":recovered,"skipped_duplicates":skipped,"needs_review":reviewed,"tables":tables}

def _v175_recover_requirements():
    _v175_setup()
    recovered=0; skipped=0; reviewed=0; tables=[]
    for table in _v175_candidate_requirement_tables():
        try:
            with engine.connect() as c:
                rows=[dict(r._mapping) for r in c.execute(text(f'SELECT * FROM "{table}" ORDER BY 1 DESC LIMIT 10000')).fetchall()]
        except Exception:
            continue
        t_rec=t_skip=t_review=0
        for idx,r in enumerate(rows):
            rowid=str(_v175_pick(r,"id","requirement_id","record_id") or idx)
            loc=_v175_pick(r,"preferred_locations","location","locality","area","city")
            mina=_v175_num(_v175_pick(r,"minimum_area_sqft","min_area_sqft","minimum_area","min_area"))
            maxa=_v175_num(_v175_pick(r,"maximum_area_sqft","max_area_sqft","maximum_area","max_area"))
            one=_v175_num(_v175_pick(r,"area_sqft","requirement_sqft","required_area_sqft"))
            if mina is None and one: mina=one
            if maxa is None and one: maxa=one
            if not loc or not mina or not maxa:
                t_review+=1; reviewed+=1
                continue
            if maxa<mina: mina,maxa=maxa,mina
            company=_v175_pick(r,"company_name","brand_name","company","client_name","name")
            city=_v175_pick(r,"city") or ("Goa" if "goa" in str(loc).lower() else "Delhi NCR")
            division="GOA" if ("goa" in str(city).lower() or "goa" in str(loc).lower()) else "DELHI_NCR"
            phone=_v175_pick(r,"contact_number","phone","mobile","contact_phone")
            ptype=_v175_pick(r,"property_type","requirement_type","category")
            maxrent=_v175_num(_v175_pick(r,"maximum_rent","max_rent","rent","budget"))
            textv=_v175_pick(r,"additional_points","requirement_text","description","remarks")
            fp=_v175_fingerprint([division,company,loc,mina,maxa,maxrent,phone])

            with engine.connect() as c:
                dup=c.execute(text("""SELECT requirement_code FROM pi_operational_requirements
                    WHERE division=:d AND lower(COALESCE(preferred_locations,''))=lower(:loc)
                    AND minimum_area_sqft=:mina AND maximum_area_sqft=:maxa
                    AND COALESCE(company_name,'')=COALESCE(:company,'')
                    LIMIT 1"""),{"d":division,"loc":str(loc),"mina":mina,"maxa":maxa,"company":company}).first()
            if dup:
                t_skip+=1; skipped+=1
                continue

            code=_v17_code("RECOVERED-REQ")
            with engine.begin() as c:
                c.execute(text("""INSERT INTO pi_operational_requirements(
                    requirement_code,division,company_name,contact_number,requirement_types,city,
                    preferred_locations,minimum_area_sqft,maximum_area_sqft,maximum_rent,
                    transaction_type,additional_points,verification_status,created_by,
                    entry_source,entered_by,entry_date
                ) VALUES(:code,:div,:company,:phone,CAST(:types AS jsonb),:city,:loc,:mina,:maxa,:rent,
                    'LEASE',:pts,'UNVERIFIED','RECOVERY_DOCTOR',
                    'RECOVERED_MANUAL','RECOVERY_DOCTOR',NOW())"""),{
                    "code":code,"div":division,"company":company,"phone":phone,
                    "types":json.dumps([str(ptype)] if ptype else []),"city":city,"loc":str(loc),
                    "mina":mina,"maxa":maxa,"rent":maxrent,"pts":textv
                })
                c.execute(text("""INSERT INTO pi_recovery_log(
                    entity_type,source_table,source_row_id,recovered_key,fingerprint,status,detail
                ) VALUES('REQUIREMENT',:t,:rid,:key,:fp,'RECOVERED',:detail)
                ON CONFLICT DO NOTHING"""),{
                    "t":table,"rid":rowid,"key":code,"fp":fp,"detail":"Copied into clean operational requirement database; source row not modified."
                })
            t_rec+=1; recovered+=1
        tables.append({"table":table,"recovered":t_rec,"skipped_duplicates":t_skip,"needs_review":t_review})
    return {"recovered":recovered,"skipped_duplicates":skipped,"needs_review":reviewed,"tables":tables}

@app.post("/api/v17-5/recovery/run")
def v175_recovery(req:Request):
    need_login(req)
    props=_v175_recover_properties()
    reqs=_v175_recover_requirements()
    return {"status":"ok","properties":props,"requirements":reqs}

@app.get("/api/v17-5/security/status")
def v175_security_status(req:Request):
    need_login(req); _v175_setup()
    with engine.connect() as c:
        props=int(c.execute(text("SELECT COUNT(*) FROM pi_operational_properties")).scalar_one() or 0)
        reqs=int(c.execute(text("SELECT COUNT(*) FROM pi_operational_requirements")).scalar_one() or 0)
        vault=int(c.execute(text("SELECT COUNT(*) FROM pi_data_vault_events")).scalar_one() or 0)
        recovered=int(c.execute(text("SELECT COUNT(*) FROM pi_recovery_log WHERE status='RECOVERED'")).scalar_one() or 0)
    return {
        "status":"ok","properties":props,"requirements":reqs,
        "vault_snapshots":vault,"recovered_records":recovered,
        "hard_delete_protection":True,"truncate_protection":True
    }

@app.get("/requirements-center-secure",response_class=HTMLResponse)
def v175_requirements_center(req:Request,division:str=Query("DELHI_NCR")):
    if not page_role_or_redirect(req): return RedirectResponse("/login",303)
    d=division.upper()
    _v175_setup()
    with engine.connect() as c:
        rows=[dict(r._mapping) for r in c.execute(text("""SELECT * FROM pi_operational_requirements
            WHERE division=:d ORDER BY id DESC LIMIT 5000"""),{"d":d}).fetchall()]
    data=json.dumps(rows,default=str)
    return HTMLResponse(f"""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Requirements Centre</title>
<style>body{{font-family:Arial;background:#f4f7fb;margin:0;color:#172437}}header{{background:#102235;color:white;padding:18px}}.w{{padding:18px}}.bar{{display:flex;gap:8px;flex-wrap:wrap}}.btn{{padding:9px 11px;background:#1677ff;color:white;border:0;border-radius:8px;text-decoration:none;cursor:pointer}}table{{width:100%;border-collapse:collapse;background:white;font-size:12px;margin-top:12px}}th,td{{padding:8px;border-bottom:1px solid #eee;text-align:left;vertical-align:top}}.small{{font-size:11px;color:#687789}}</style></head><body>
<header><b>{'Goa' if d=='GOA' else 'Delhi NCR'} Requirements Centre</b><br><small>Clean operational requirements only · protected by Data Vault</small></header><div class=w>
<div class=bar><a class=btn href="/final-dashboard-v9">← Dashboard</a><a class=btn href="/manual-requirement-final?division={d}">Add Requirement</a><a class=btn href="/matcher-final?division={d}">Open Matcher</a></div>
<table><thead><tr><th>Requirement</th><th>Source</th><th>Entry Date</th><th>Entered By</th><th>Company</th><th>Location</th><th>Area</th><th>Max Rent</th><th>Verification</th><th>Action</th></tr></thead><tbody id=rows></tbody></table></div>
<script>
const data={data};const E=x=>String(x??'').replace(/[&<>"]/g,c=>({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}}[c]));
rows.innerHTML=data.map(x=>`<tr><td>${{E(x.requirement_code)}}</td><td><b>${{E(x.entry_source||'MANUAL')}}</b></td><td>${{E(String(x.entry_date||x.created_at||'').slice(0,16))}}</td><td>${{E(x.entered_by||x.created_by||'')}}</td><td>${{E(x.company_name||x.client_name||'')}}</td><td>${{E(x.preferred_locations||'')}}</td><td>${{E(x.minimum_area_sqft)}} - ${{E(x.maximum_area_sqft)}}</td><td>${{E(x.maximum_rent||'')}}</td><td>${{E(x.verification_status||'')}}</td><td><a href="/matcher-final?division={d}">Run Match</a></td></tr>`).join('')||'<tr><td colspan=10>No requirements found.</td></tr>';
</script></body></html>""")

@app.get("/data-recovery-doctor",response_class=HTMLResponse)
def v175_recovery_page(req:Request):
    if not page_role_or_redirect(req): return RedirectResponse("/login",303)
    return HTMLResponse("""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Data Recovery Doctor</title>
<style>body{font-family:Arial;background:#f4f7fb;margin:0;color:#172437}header{background:#102235;color:white;padding:18px}.w{max-width:1300px;margin:auto;padding:18px}.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:10px}.k{background:white;border:1px solid #e2e8f0;border-radius:10px;padding:12px}.k b{display:block;font-size:22px}.btn{padding:9px 11px;background:#1677ff;color:white;border:0;border-radius:8px;text-decoration:none;cursor:pointer}.msg{margin-top:12px;background:#fff8e8;border:1px solid #eed18f;padding:10px;border-radius:8px}</style></head><body>
<header><b>Data Recovery Doctor + Immutable Vault</b><br><small>Recover historical manual data without deleting or modifying source rows</small></header><div class=w>
<p><a class=btn href="/final-dashboard-v9">← Dashboard</a> <button class=btn onclick=recover()>Recover Historical Manual Data</button></p>
<div class=grid><div class=k><b id=props>0</b><span>Operational Properties</span></div><div class=k><b id=reqs>0</b><span>Operational Requirements</span></div><div class=k><b id=vault>0</b><span>Immutable Vault Snapshots</span></div><div class=k><b id=rec>0</b><span>Recovered Records</span></div><div class=k><b>ON</b><span>Hard Delete Protection</span></div><div class=k><b>ON</b><span>Truncate Protection</span></div></div>
<div id=msg class=msg>Ready. Recovery copies valid historical records into the clean operational DB. Original records are preserved.</div></div>
<script>
async function status(){let d=await(await fetch('/api/v17-5/security/status')).json();props.textContent=d.properties||0;reqs.textContent=d.requirements||0;vault.textContent=d.vault_snapshots||0;rec.textContent=d.recovered_records||0}
async function recover(){msg.textContent='Scanning historical manual tables...';let r=await fetch('/api/v17-5/recovery/run',{method:'POST'}),d=await r.json();if(!r.ok){msg.textContent='ERROR: '+(d.detail||d.message||'Recovery failed');return}msg.textContent=`Recovery complete. Properties recovered ${{d.properties.recovered||0}}, duplicates skipped ${{d.properties.skipped_duplicates||0}}, review ${{d.properties.needs_review||0}}. Requirements recovered ${{d.requirements.recovered||0}}, duplicates skipped ${{d.requirements.skipped_duplicates||0}}, review ${{d.requirements.needs_review||0}}.`;status()}status()
</script></body></html>""")

@app.get("/final-dashboard-v9",response_class=HTMLResponse)
def v175_dashboard(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    admin='<a class=card href="/admin-data-tools-v2"><b>Admin Data Tools</b><p>Database health and tools.</p></a>' if role=="admin" else ""
    return HTMLResponse(f"""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Final Dashboard</title>
<style>body{{font-family:Arial;background:#f4f7fb;margin:0;color:#172437}}header{{background:#102235;color:white;padding:22px}}.w{{max-width:1500px;margin:auto;padding:20px}}.g{{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:12px;margin-bottom:24px}}.card{{display:block;background:white;border:1px solid #e2e8f0;border-radius:12px;padding:15px;text-decoration:none;color:#172437}}.card p{{font-size:12px;color:#687789}}.primary{{border:2px solid #1677ff}}.secure{{border:2px solid #08734b}}</style></head><body><header><b>AI Deal Intelligence OS</b><br><small>SECURE FINAL DASHBOARD</small></header><div class=w>
<h2>Data Security & Recovery</h2><div class=g><a class="card secure" href="/data-recovery-doctor"><b>Data Recovery Doctor</b><p>Recover older manual records and verify immutable vault protection.</p></a><a class="card secure" href="/fresh-inventory-final?division=DELHI_NCR"><b>Delhi NCR Fresh Inventory</b><p>Clean operational inventory with source metadata.</p></a><a class="card secure" href="/fresh-inventory-final?division=GOA"><b>Goa Fresh Inventory</b><p>Clean Goa operational inventory.</p></a><a class="card secure" href="/requirements-center-secure?division=DELHI_NCR"><b>Requirements Centre</b><p>Working secure requirements database.</p></a></div>
<h2>Delhi NCR</h2><div class=g><a class="card primary" href="/manual-property-final?division=DELHI_NCR"><b>Add Property</b><p>Manual record saved to protected core DB.</p></a><a class="card primary" href="/manual-requirement-final?division=DELHI_NCR"><b>Add Requirement</b><p>Manual requirement saved to protected core DB.</p></a><a class="card primary" href="/matcher-final?division=DELHI_NCR"><b>Run Matcher</b><p>Clean operational matcher.</p></a></div>
<h2>Goa</h2><div class=g><a class="card primary" href="/manual-property-final?division=GOA"><b>Add Goa Property</b><p>Protected core DB + brochure/media.</p></a><a class="card primary" href="/manual-requirement-final?division=GOA"><b>Add Goa Requirement</b><p>Protected Goa requirement.</p></a><a class="card primary" href="/matcher-final?division=GOA"><b>Goa Matcher</b><p>Clean Goa matcher.</p></a><a class=card href="/requirements-center-secure?division=GOA"><b>Goa Requirements Centre</b><p>Goa requirements database.</p></a></div>
<h2>Search, AI & Marketing</h2><div class=g><a class=card href="/property-discovery"><b>Property Discovery</b><p>Search engine.</p></a><a class=card href="/retail-expansion"><b>Retail Expansion</b><p>Retail AI.</p></a><a class=card href="/ai-hospitality-master-final"><b>Hospitality Master</b><p>Hospitality data.</p></a><a class=card href="/marketing-contacts-final"><b>Marketing Contacts</b><p>WhatsApp marketing.</p></a><a class=card href="/phone-contact-upload"><b>Upload Phone Contacts</b><p>VCF/CSV/XLSX.</p></a><a class=card href="/capture-intelligence"><b>Capture Property</b><p>Camera/screenshot/PDF/magazine.</p></a></div>
<h2>Database</h2><div class=g><a class=card href="/property-database"><b>Full Property Database</b><p>Legacy archive.</p></a><a class=card href="/contacts-directory"><b>Owner / Broker Contacts</b><p>Verification.</p></a><a class=card href="/data-doctor"><b>Data Doctor</b><p>Database health.</p></a>{admin}</div>
</div></body></html>""")

@app.middleware("http")
async def v175_router(request,call_next):
    p=request.url.path
    if p in {"/workspace","/final-dashboard","/final-dashboard-v2","/final-dashboard-v4","/final-dashboard-v5","/final-dashboard-v6","/final-dashboard-v7","/final-dashboard-v8"}:
        return RedirectResponse("/final-dashboard-v9",status_code=307)
    if p in {"/requirements-match-center","/requirements-center-final"}:
        div="GOA" if "DIVISION=GOA" in request.url.query.upper() else "DELHI_NCR"
        return RedirectResponse(f"/requirements-center-secure?division={div}",status_code=307)
    response=await call_next(request)
    if p.startswith(("/final-dashboard-v9","/data-recovery-doctor","/requirements-center-secure","/fresh-inventory-final")):
        response.headers["Cache-Control"]="no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"]="no-cache"
        response.headers["Expires"]="0"
    return response

# ============================================================
# V17.6 UNIVERSAL RECOVERY + DATA SECURITY REPAIR
# ============================================================

def _v176_setup():
    _v174_setup()
    with engine.begin() as c:
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_recovery_staging(
            id BIGSERIAL PRIMARY KEY,
            entity_type TEXT NOT NULL,
            source_table TEXT NOT NULL,
            source_row_id TEXT,
            detected_at TIMESTAMPTZ DEFAULT NOW(),
            confidence INTEGER DEFAULT 0,
            recovery_status TEXT DEFAULT 'DISCOVERED',
            reason TEXT,
            raw_snapshot JSONB NOT NULL,
            UNIQUE(entity_type,source_table,source_row_id)
        )"""))
        c.execute(text("""CREATE TABLE IF NOT EXISTS pi_core_data_history(
            id BIGSERIAL PRIMARY KEY,
            entity_type TEXT NOT NULL,
            entity_key TEXT,
            action TEXT NOT NULL,
            snapshot JSONB NOT NULL,
            changed_at TIMESTAMPTZ DEFAULT NOW(),
            changed_by TEXT DEFAULT current_user
        )"""))
        c.execute(text("""
        CREATE OR REPLACE FUNCTION pi_property_history_guard()
        RETURNS trigger AS $$
        BEGIN
            IF TG_OP='INSERT' THEN
                INSERT INTO pi_core_data_history(entity_type,entity_key,action,snapshot)
                VALUES('PROPERTY',NEW.property_code,'INSERT',to_jsonb(NEW)); RETURN NEW;
            ELSIF TG_OP='UPDATE' THEN
                INSERT INTO pi_core_data_history(entity_type,entity_key,action,snapshot)
                VALUES('PROPERTY',OLD.property_code,'BEFORE_UPDATE',to_jsonb(OLD)); RETURN NEW;
            ELSE
                INSERT INTO pi_core_data_history(entity_type,entity_key,action,snapshot)
                VALUES('PROPERTY',OLD.property_code,'DELETE_BLOCKED',to_jsonb(OLD)); RETURN NULL;
            END IF;
        END; $$ LANGUAGE plpgsql;
        """))
        c.execute(text("""
        CREATE OR REPLACE FUNCTION pi_requirement_history_guard()
        RETURNS trigger AS $$
        BEGIN
            IF TG_OP='INSERT' THEN
                INSERT INTO pi_core_data_history(entity_type,entity_key,action,snapshot)
                VALUES('REQUIREMENT',NEW.requirement_code,'INSERT',to_jsonb(NEW)); RETURN NEW;
            ELSIF TG_OP='UPDATE' THEN
                INSERT INTO pi_core_data_history(entity_type,entity_key,action,snapshot)
                VALUES('REQUIREMENT',OLD.requirement_code,'BEFORE_UPDATE',to_jsonb(OLD)); RETURN NEW;
            ELSE
                INSERT INTO pi_core_data_history(entity_type,entity_key,action,snapshot)
                VALUES('REQUIREMENT',OLD.requirement_code,'DELETE_BLOCKED',to_jsonb(OLD)); RETURN NULL;
            END IF;
        END; $$ LANGUAGE plpgsql;
        """))
        for t in ["pi_operational_properties","pi_operational_requirements"]:
            c.execute(text(f"DROP TRIGGER IF EXISTS trg_{t}_vault ON {t}"))
        c.execute(text("""CREATE TRIGGER trg_pi_operational_properties_vault
            BEFORE INSERT OR UPDATE OR DELETE ON pi_operational_properties
            FOR EACH ROW EXECUTE FUNCTION pi_property_history_guard()"""))
        c.execute(text("""CREATE TRIGGER trg_pi_operational_requirements_vault
            BEFORE INSERT OR UPDATE OR DELETE ON pi_operational_requirements
            FOR EACH ROW EXECUTE FUNCTION pi_requirement_history_guard()"""))

def _v176_tables():
    with engine.connect() as c:
        return [r._mapping["table_name"] for r in c.execute(text("""
            SELECT table_name FROM information_schema.tables
            WHERE table_schema='public' AND table_type='BASE TABLE' ORDER BY table_name
        """)).fetchall()]

def _v176_cols(t):
    with engine.connect() as c:
        return [r._mapping["column_name"] for r in c.execute(text("""
            SELECT column_name FROM information_schema.columns
            WHERE table_schema='public' AND table_name=:t
        """),{"t":t}).fetchall()]

def _v176_pick(row,*names):
    low={str(k).lower():v for k,v in row.items()}
    for n in names:
        if low.get(n.lower()) not in (None,""): return low[n.lower()]
    return None

def _v176_num(v):
    if v in (None,""): return None
    try:
        return float(_re.sub(r"[^\d.]","",str(v)))
    except Exception:
        return None

def _v176_recent_sql(table,cols):
    for c in ["created_at","updated_at","entry_date","date_added","created_on"]:
        if c in cols:
            return f'SELECT * FROM "{table}" WHERE "{c}" >= NOW() - INTERVAL \'48 hours\' ORDER BY "{c}" DESC LIMIT 10000'
    return f'SELECT * FROM "{table}" LIMIT 3000'

def _v176_classify(table,cols,row):
    tl=table.lower(); cset=set(x.lower() for x in cols)
    pkeys={"property_name","property_type","rent","rent_amount","owner_name","broker_name","available_area","area_sqft","google_location"}
    rkeys={"requirement_text","minimum_area_sqft","maximum_area_sqft","requirement_sqft","preferred_locations","client_name","company_name"}
    src=str(_v176_pick(row,"source","entry_source","source_type") or "").upper()
    manual=("manual" in tl or src=="MANUAL")
    ps=len(cset&pkeys)+(2 if "property" in tl else 0)+(2 if manual else 0)
    rs=len(cset&rkeys)+(2 if "requirement" in tl else 0)+(2 if manual else 0)
    if ps>=3 and ps>=rs:return "PROPERTY",min(100,40+ps*8)
    if rs>=3:return "REQUIREMENT",min(100,40+rs*8)
    return None,0

def _v176_stage(entity,table,row,rowid,confidence,reason):
    with engine.begin() as c:
        c.execute(text("""INSERT INTO pi_recovery_staging(
            entity_type,source_table,source_row_id,confidence,recovery_status,reason,raw_snapshot
        ) VALUES(:e,:t,:rid,:conf,'DISCOVERED',:reason,CAST(:snap AS jsonb))
        ON CONFLICT(entity_type,source_table,source_row_id)
        DO UPDATE SET confidence=EXCLUDED.confidence,reason=EXCLUDED.reason,raw_snapshot=EXCLUDED.raw_snapshot
        """),{"e":entity,"t":table,"rid":str(rowid),"conf":confidence,"reason":reason,"snap":json.dumps(row,default=str)})

def _v176_recover_property(table,row,rowid):
    loc=_v176_pick(row,"location","locality","micro_market","area","address")
    area=_v176_num(_v176_pick(row,"area_sqft","available_area","area","size_sqft","super_area"))
    rent=_v176_num(_v176_pick(row,"rent_amount","rent","monthly_rent","asking_rent"))
    if not loc or not area or not rent:return False,"Missing location/area/rent"
    name=_v176_pick(row,"property_name","name","building_name","title")
    city=_v176_pick(row,"city") or ("Goa" if "goa" in str(loc).lower() else "Delhi NCR")
    division="GOA" if ("goa" in str(city).lower() or "goa" in str(loc).lower()) else "DELHI_NCR"
    phone=_v176_pick(row,"contact_number","owner_contact","broker_contact","phone","mobile")
    person=_v176_pick(row,"owner_name","broker_name","contact_name","owner_broker_name")
    ptype=_v176_pick(row,"property_type","type","category")
    with engine.connect() as c:
        dup=c.execute(text("""SELECT property_code FROM pi_operational_properties
            WHERE division=:d AND lower(COALESCE(location,''))=lower(:loc)
            AND area_sqft=:a AND rent_amount=:r AND COALESCE(contact_number,'')=COALESCE(:p,'') LIMIT 1"""),
            {"d":division,"loc":str(loc),"a":area,"r":rent,"p":phone}).first()
    if dup:return False,"Duplicate"
    code=_v17_code("RECOVERED-PROP")
    with engine.begin() as c:
        c.execute(text("""INSERT INTO pi_operational_properties(
            property_code,division,property_name,property_types,city,location,area_sqft,rent_amount,
            transaction_type,owner_broker_name,contact_number,verification_status,remarks,created_by,
            entry_source,entered_by,entry_date
        ) VALUES(:code,:div,:name,CAST(:types AS jsonb),:city,:loc,:area,:rent,'LEASE',
            :person,:phone,'UNVERIFIED',:remarks,'RECOVERY_DOCTOR','RECOVERED_MANUAL','RECOVERY_DOCTOR',NOW())"""),
            {"code":code,"div":division,"name":name,"types":json.dumps([str(ptype)] if ptype else []),
             "city":city,"loc":str(loc),"area":area,"rent":rent,"person":person,"phone":phone,
             "remarks":f"Recovered from {table} row {rowid}; original preserved."})
    return True,code

def _v176_recover_requirement(table,row,rowid):
    loc=_v176_pick(row,"preferred_locations","location","locality","area","city")
    mina=_v176_num(_v176_pick(row,"minimum_area_sqft","min_area_sqft","minimum_area","min_area"))
    maxa=_v176_num(_v176_pick(row,"maximum_area_sqft","max_area_sqft","maximum_area","max_area"))
    one=_v176_num(_v176_pick(row,"area_sqft","requirement_sqft","required_area_sqft"))
    if mina is None and one:mina=one
    if maxa is None and one:maxa=one
    if not loc or not mina or not maxa:return False,"Missing location/area range"
    if maxa<mina:mina,maxa=maxa,mina
    company=_v176_pick(row,"company_name","brand_name","company","client_name","name")
    city=_v176_pick(row,"city") or ("Goa" if "goa" in str(loc).lower() else "Delhi NCR")
    division="GOA" if ("goa" in str(city).lower() or "goa" in str(loc).lower()) else "DELHI_NCR"
    ptype=_v176_pick(row,"property_type","requirement_type","category")
    rent=_v176_num(_v176_pick(row,"maximum_rent","max_rent","rent","budget"))
    textv=_v176_pick(row,"additional_points","requirement_text","description","remarks")
    with engine.connect() as c:
        dup=c.execute(text("""SELECT requirement_code FROM pi_operational_requirements
            WHERE division=:d AND lower(COALESCE(preferred_locations,''))=lower(:loc)
            AND minimum_area_sqft=:mina AND maximum_area_sqft=:maxa
            AND COALESCE(company_name,'')=COALESCE(:company,'') LIMIT 1"""),
            {"d":division,"loc":str(loc),"mina":mina,"maxa":maxa,"company":company}).first()
    if dup:return False,"Duplicate"
    code=_v17_code("RECOVERED-REQ")
    with engine.begin() as c:
        c.execute(text("""INSERT INTO pi_operational_requirements(
            requirement_code,division,company_name,requirement_types,city,preferred_locations,
            minimum_area_sqft,maximum_area_sqft,maximum_rent,transaction_type,additional_points,
            verification_status,created_by,entry_source,entered_by,entry_date
        ) VALUES(:code,:div,:company,CAST(:types AS jsonb),:city,:loc,:mina,:maxa,:rent,'LEASE',:pts,
            'UNVERIFIED','RECOVERY_DOCTOR','RECOVERED_MANUAL','RECOVERY_DOCTOR',NOW())"""),
            {"code":code,"div":division,"company":company,"types":json.dumps([str(ptype)] if ptype else []),
             "city":city,"loc":str(loc),"mina":mina,"maxa":maxa,"rent":rent,"pts":textv})
    return True,code

@app.post("/api/v17-6/universal-recovery")
def v176_universal_recovery(req:Request):
    need_login(req);_v176_setup()
    protected={"pi_operational_properties","pi_operational_requirements","pi_operational_property_media",
               "pi_core_data_history","pi_recovery_staging","pi_data_vault_events","pi_recovery_log"}
    totals={"tables_scanned":0,"rows_scanned":0,"staged":0,"properties_recovered":0,"requirements_recovered":0,"needs_review":0}
    report=[]
    for table in _v176_tables():
        if table in protected:continue
        try:
            cols=_v176_cols(table)
            with engine.connect() as c:
                rows=[dict(r._mapping) for r in c.execute(text(_v176_recent_sql(table,cols))).fetchall()]
        except Exception as ex:
            continue
        totals["tables_scanned"]+=1
        rec=rev=stage=0
        for idx,row in enumerate(rows):
            totals["rows_scanned"]+=1
            entity,conf=_v176_classify(table,cols,row)
            if not entity:continue
            rowid=_v176_pick(row,"id","property_id","requirement_id","record_id","lead_id") or idx
            _v176_stage(entity,table,row,rowid,conf,"Universal recent/manual scan")
            totals["staged"]+=1;stage+=1
            src=str(_v176_pick(row,"source","entry_source","source_type") or "").upper()
            if entity=="PROPERTY":
                ok,reason=_v176_recover_property(table,row,rowid)
            else:
                if not ("manual" in table.lower() or src=="MANUAL"):
                    ok,reason=False,"AI/public requirement requires manual confirmation"
                else:
                    ok,reason=_v176_recover_requirement(table,row,rowid)
            with engine.begin() as c:
                c.execute(text("""UPDATE pi_recovery_staging SET recovery_status=:s,reason=:r
                    WHERE entity_type=:e AND source_table=:t AND source_row_id=:rid"""),
                    {"s":"RECOVERED" if ok else "NEEDS_REVIEW","r":reason,"e":entity,"t":table,"rid":str(rowid)})
            if ok:
                rec+=1
                totals["properties_recovered" if entity=="PROPERTY" else "requirements_recovered"]+=1
            else:
                rev+=1;totals["needs_review"]+=1
        if rows or stage:
            report.append({"table":table,"rows_scanned":len(rows),"staged":stage,"recovered":rec,"needs_review":rev})
    totals["report"]=report
    return {"status":"ok",**totals}

@app.get("/api/v17-6/security-status")
def v176_security_status(req:Request):
    need_login(req);_v176_setup()
    with engine.connect() as c:
        props=int(c.execute(text("SELECT COUNT(*) FROM pi_operational_properties")).scalar_one() or 0)
        reqs=int(c.execute(text("SELECT COUNT(*) FROM pi_operational_requirements")).scalar_one() or 0)
        hist=int(c.execute(text("SELECT COUNT(*) FROM pi_core_data_history")).scalar_one() or 0)
        stage=int(c.execute(text("SELECT COUNT(*) FROM pi_recovery_staging")).scalar_one() or 0)
        review=int(c.execute(text("SELECT COUNT(*) FROM pi_recovery_staging WHERE recovery_status='NEEDS_REVIEW'")).scalar_one() or 0)
    return {"status":"ok","operational_properties":props,"operational_requirements":reqs,
            "history_snapshots":hist,"staged_records":stage,"needs_review":review}

@app.get("/requirements-center-v176",response_class=HTMLResponse)
def v176_requirements_center(req:Request,division:str=Query("DELHI_NCR")):
    if not page_role_or_redirect(req):return RedirectResponse("/login",303)
    d=division.upper();_v176_setup()
    return HTMLResponse(f"""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Requirements Centre</title>
<style>body{{font-family:Arial;background:#f4f7fb;margin:0;color:#172437}}header{{background:#102235;color:#fff;padding:18px}}.w{{padding:18px}}.btn{{padding:9px 11px;background:#1677ff;color:#fff;text-decoration:none;border-radius:8px}}table{{width:100%;border-collapse:collapse;background:#fff;font-size:12px;margin-top:12px}}th,td{{padding:8px;border-bottom:1px solid #eee;text-align:left}}</style></head><body>
<header><b>{'Goa' if d=='GOA' else 'Delhi NCR'} Requirements Centre</b></header><div class=w><a class=btn href="/final-dashboard-v10">← Dashboard</a> <a class=btn href="/manual-requirement-final?division={d}">Add / Confirm Requirement</a> <a class=btn href="/matcher-final?division={d}">Matcher</a>
<table><thead><tr><th>Code</th><th>Source</th><th>Company</th><th>Location</th><th>Area</th><th>Rent</th><th>Verification</th></tr></thead><tbody id=rows></tbody></table></div>
<script>async function load(){{let d=await(await fetch('/api/v17/requirements?division={d}')).json();rows.innerHTML=(d.rows||[]).map(x=>`<tr><td>${{x.requirement_code}}</td><td>${{x.entry_source||'MANUAL'}}</td><td>${{x.company_name||x.client_name||''}}</td><td>${{x.preferred_locations||''}}</td><td>${{x.minimum_area_sqft}} - ${{x.maximum_area_sqft}}</td><td>${{x.maximum_rent||''}}</td><td>${{x.verification_status||''}}</td></tr>`).join('')||'<tr><td colspan=7>No confirmed requirements.</td></tr>'}}load()</script>
</body></html>""")

@app.get("/universal-recovery-doctor",response_class=HTMLResponse)
def v176_recovery_page(req:Request):
    if not page_role_or_redirect(req):return RedirectResponse("/login",303)
    return HTMLResponse("""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Recovery Doctor</title>
<style>body{font-family:Arial;background:#f4f7fb;margin:0;color:#172437}header{background:#102235;color:#fff;padding:18px}.w{max-width:1350px;margin:auto;padding:18px}.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(170px,1fr));gap:10px}.k{background:#fff;border:1px solid #e2e8f0;border-radius:10px;padding:12px}.k b{display:block;font-size:22px}.btn{padding:9px 11px;background:#1677ff;color:#fff;border:0;border-radius:8px;text-decoration:none;cursor:pointer}.msg{margin-top:12px;background:#fff8e8;padding:10px;border-radius:8px}table{width:100%;border-collapse:collapse;background:#fff;font-size:12px;margin-top:12px}th,td{padding:7px;border-bottom:1px solid #eee;text-align:left}</style></head><body>
<header><b>Universal Data Recovery Doctor</b><br><small>Scans all public tables for recent/manual records</small></header><div class=w><p><a class=btn href="/final-dashboard-v10">← Dashboard</a> <button class=btn onclick=run()>Scan & Recover Recent Manual Data</button></p>
<div class=grid><div class=k><b id=p>0</b><span>Operational Properties</span></div><div class=k><b id=r>0</b><span>Operational Requirements</span></div><div class=k><b id=h>0</b><span>History Snapshots</span></div><div class=k><b id=s>0</b><span>Staged Records</span></div><div class=k><b id=rv>0</b><span>Needs Review</span></div></div>
<div id=msg class=msg>Ready.</div><table><thead><tr><th>Table</th><th>Rows Scanned</th><th>Staged</th><th>Recovered</th><th>Needs Review</th></tr></thead><tbody id=rows></tbody></table></div>
<script>
async function status(){let d=await(await fetch('/api/v17-6/security-status')).json();p.textContent=d.operational_properties||0;r.textContent=d.operational_requirements||0;h.textContent=d.history_snapshots||0;s.textContent=d.staged_records||0;rv.textContent=d.needs_review||0}
async function run(){msg.textContent='Scanning all tables...';let q=await fetch('/api/v17-6/universal-recovery',{method:'POST'}),d=await q.json();if(!q.ok){msg.textContent='ERROR: '+(d.detail||d.message||'Recovery failed');return}msg.textContent=`Done. Tables ${d.tables_scanned}; rows ${d.rows_scanned}; properties recovered ${d.properties_recovered}; requirements recovered ${d.requirements_recovered}; review ${d.needs_review}.`;rows.innerHTML=(d.report||[]).map(x=>`<tr><td>${x.table}</td><td>${x.rows_scanned||0}</td><td>${x.staged||0}</td><td>${x.recovered||0}</td><td>${x.needs_review||0}</td></tr>`).join('');status()}status()
</script></body></html>""")

@app.get("/final-dashboard-v10",response_class=HTMLResponse)
def v176_dashboard(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    return HTMLResponse("""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Final Dashboard</title>
<style>body{font-family:Arial;background:#f4f7fb;margin:0;color:#172437}header{background:#102235;color:#fff;padding:22px}.w{max-width:1500px;margin:auto;padding:20px}.g{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:12px;margin-bottom:24px}.card{display:block;background:#fff;border:1px solid #e2e8f0;border-radius:12px;padding:15px;text-decoration:none;color:#172437}.card p{font-size:12px;color:#687789}.primary{border:2px solid #1677ff}.secure{border:2px solid #08734b}</style></head><body><header><b>AI Deal Intelligence OS</b><br><small>SECURE EXECUTION DASHBOARD</small></header><div class=w>
<h2>Recovery & Security</h2><div class=g><a class="card secure" href="/universal-recovery-doctor"><b>Universal Recovery Doctor</b><p>Scan all public tables for recent/manual records.</p></a><a class="card secure" href="/fresh-inventory-final?division=DELHI_NCR"><b>Delhi NCR Fresh Inventory</b></a><a class="card secure" href="/fresh-inventory-final?division=GOA"><b>Goa Fresh Inventory</b></a><a class="card secure" href="/requirements-center-v176?division=DELHI_NCR"><b>Requirements Centre</b></a></div>
<h2>Delhi NCR</h2><div class=g><a class="card primary" href="/manual-property-final?division=DELHI_NCR"><b>Add Property</b></a><a class="card primary" href="/manual-requirement-final?division=DELHI_NCR"><b>Add Requirement</b></a><a class="card primary" href="/matcher-final?division=DELHI_NCR"><b>Run Matcher</b></a></div>
<h2>Goa</h2><div class=g><a class="card primary" href="/manual-property-final?division=GOA"><b>Add Goa Property</b></a><a class="card primary" href="/manual-requirement-final?division=GOA"><b>Add Goa Requirement</b></a><a class="card primary" href="/matcher-final?division=GOA"><b>Goa Matcher</b></a><a class=card href="/requirements-center-v176?division=GOA"><b>Goa Requirements Centre</b></a></div>
<h2>Search, AI & Marketing</h2><div class=g><a class=card href="/property-discovery"><b>Property Discovery</b></a><a class=card href="/retail-expansion"><b>Retail Expansion</b></a><a class=card href="/ai-hospitality-master-final"><b>Hospitality Master</b></a><a class=card href="/marketing-contacts-final"><b>Marketing Contacts</b></a><a class=card href="/phone-contact-upload"><b>Upload Phone Contacts</b></a><a class=card href="/capture-intelligence"><b>Capture Property</b></a></div>
</div></body></html>""")

@app.middleware("http")
async def v176_router(request,call_next):
    p=request.url.path
    if p in {"/workspace","/final-dashboard-v9"}: return RedirectResponse("/final-dashboard-v10",status_code=307)
    if p in {"/requirements-match-center","/requirements-center-secure"}:
        div="GOA" if "DIVISION=GOA" in request.url.query.upper() else "DELHI_NCR"
        return RedirectResponse(f"/requirements-center-v176?division={div}",status_code=307)
    if p=="/data-recovery-doctor": return RedirectResponse("/universal-recovery-doctor",status_code=307)
    return await call_next(request)

# ============================================================
# V17.7 MANUAL PROPERTY DATABASE + FULL DETAIL VIEW
# ============================================================

def _v177_setup():
    _v174_setup()

def _v177_media_rows(property_code):
    with engine.connect() as c:
        return [dict(r._mapping) for r in c.execute(text("""
            SELECT id,property_code,media_type,filename,mime_type,file_size,created_at
            FROM pi_operational_property_media
            WHERE property_code=:p
            ORDER BY id
        """),{"p":property_code}).fetchall()]

@app.get("/api/v17-7/manual-properties")
def v177_manual_properties(req:Request, division:str=Query("ALL"), source:str=Query("MANUAL"),
                           verified:str=Query("ALL"), q:str=Query("")):
    need_login(req); _v177_setup()
    wh=["1=1"]; params={}
    d=division.upper()
    if d!="ALL":
        wh.append("p.division=:division"); params["division"]=d
    source=source.upper()
    if source=="MANUAL":
        wh.append("COALESCE(p.entry_source,'MANUAL') IN ('MANUAL','RECOVERED_MANUAL')")
    elif source!="ALL":
        wh.append("COALESCE(p.entry_source,'MANUAL')=:source"); params["source"]=source
    if verified.upper()!="ALL":
        wh.append("COALESCE(p.verification_status,'UNVERIFIED')=:verified")
        params["verified"]=verified.upper()
    if q.strip():
        wh.append("""(
          COALESCE(p.property_code,'') ILIKE :q OR COALESCE(p.property_name,'') ILIKE :q OR
          COALESCE(p.location,'') ILIKE :q OR COALESCE(p.city,'') ILIKE :q OR
          COALESCE(p.owner_broker_name,'') ILIKE :q OR COALESCE(p.contact_number,'') ILIKE :q
        )""")
        params["q"]="%"+q.strip()+"%"

    sql="""SELECT p.*,
      COALESCE(p.entry_source,'MANUAL') AS display_source,
      COALESCE(p.entered_by,p.created_by,'') AS display_entered_by,
      COALESCE(p.entry_date,p.created_at) AS display_entry_date,
      (SELECT COUNT(*) FROM pi_operational_property_media m WHERE m.property_code=p.property_code AND m.media_type='IMAGE') image_count,
      (SELECT COUNT(*) FROM pi_operational_property_media m WHERE m.property_code=p.property_code AND m.media_type='VIDEO') video_count,
      (SELECT COUNT(*) FROM pi_operational_property_media m WHERE m.property_code=p.property_code AND m.media_type='BROCHURE') brochure_count
      FROM pi_operational_properties p
      WHERE """+" AND ".join(wh)+"""
      ORDER BY COALESCE(p.entry_date,p.created_at) DESC,p.id DESC LIMIT 10000"""
    with engine.connect() as c:
        rows=[dict(r._mapping) for r in c.execute(text(sql),params).fetchall()]

    today=date.today().isoformat()
    summary={
        "total":len(rows),
        "added_today":sum(1 for x in rows if str(x.get("display_entry_date") or "")[:10]==today),
        "verified":sum(1 for x in rows if str(x.get("verification_status") or "").upper()=="VERIFIED"),
        "unverified":sum(1 for x in rows if str(x.get("verification_status") or "").upper()!="VERIFIED"),
        "photos":sum(int(x.get("image_count") or 0) for x in rows),
        "videos":sum(int(x.get("video_count") or 0) for x in rows),
        "brochures":sum(int(x.get("brochure_count") or 0) for x in rows)
    }
    return {"status":"ok","rows":rows,"summary":summary}

@app.get("/api/v17-7/property/{property_code}")
def v177_property_detail_api(property_code:str,req:Request):
    need_login(req); _v177_setup()
    with engine.connect() as c:
        row=c.execute(text("""
            SELECT p.*,
              COALESCE(p.entry_source,'MANUAL') AS display_source,
              COALESCE(p.entered_by,p.created_by,'') AS display_entered_by,
              COALESCE(p.entry_date,p.created_at) AS display_entry_date
            FROM pi_operational_properties p WHERE p.property_code=:p
        """),{"p":property_code}).first()
    if not row: raise HTTPException(404,"Property not found.")
    return {"status":"ok","property":dict(row._mapping),"media":_v177_media_rows(property_code)}

@app.get("/manual-property-database",response_class=HTMLResponse)
def v177_manual_property_database(req:Request,division:str=Query("ALL")):
    if not page_role_or_redirect(req): return RedirectResponse("/login",303)
    d=division.upper()
    return HTMLResponse(f"""<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Manual Property Database</title>
<style>
*{{box-sizing:border-box}}body{{font-family:Arial;margin:0;background:#f4f7fb;color:#172437}}header{{background:#102235;color:#fff;padding:20px}}
.w{{max-width:1650px;margin:auto;padding:18px}}.bar{{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px;align-items:center}}
.btn,a.btn{{display:inline-block;padding:9px 11px;border:0;border-radius:8px;background:#1677ff;color:#fff;text-decoration:none;font-weight:bold;cursor:pointer}}
.gray{{background:#e9eef5!important;color:#203247!important}}input,select{{padding:9px;border:1px solid #ccd6e2;border-radius:7px}}input{{min-width:280px}}
.kpis{{display:grid;grid-template-columns:repeat(auto-fit,minmax(145px,1fr));gap:9px;margin:14px 0}}.k{{background:#fff;border:1px solid #e2e8f0;border-radius:10px;padding:12px}}.k b{{display:block;font-size:24px}}
.tablewrap{{overflow:auto;background:#fff;border:1px solid #e2e8f0;border-radius:10px}}table{{width:100%;border-collapse:collapse;font-size:12px}}
th,td{{padding:9px;border-bottom:1px solid #edf1f5;text-align:left;vertical-align:top;white-space:nowrap}}th{{background:#f8fafc;position:sticky;top:0;z-index:2}}
.manualrow td{{font-weight:700}}.small{{font-size:11px;color:#687789;font-weight:normal}}.badge{{display:inline-block;padding:3px 7px;border-radius:10px;background:#dcfce7;color:#166534;font-size:10px;font-weight:bold}}
.unv{{background:#fef3c7;color:#92400e}}.today{{background:#dbeafe;color:#1d4ed8}}.media{{font-weight:bold}}.no{{font-size:16px;font-weight:800}}
</style></head><body>
<header><b>Manual Property Database</b><br><small>Saved operational properties with full details, photos, videos and brochures</small></header><div class=w>
<div class=bar>
<a class="btn gray" href="/final-dashboard-v11">← Dashboard</a>
<a class=btn href="/manual-property-final?division=DELHI_NCR">Add Delhi NCR Property</a>
<a class=btn href="/manual-property-final?division=GOA">Add Goa Property</a>
<select id=division><option value="ALL">ALL AREAS</option><option value="DELHI_NCR">DELHI NCR</option><option value="GOA">GOA</option></select>
<select id=source><option value="MANUAL">MANUAL + RECOVERED MANUAL</option><option value="ALL">ALL OPERATIONAL SOURCES</option><option value="RECOVERED_MANUAL">RECOVERED MANUAL ONLY</option></select>
<select id=verified><option value="ALL">ALL VERIFICATION</option><option>VERIFIED</option><option>UNVERIFIED</option></select>
<input id=q placeholder="Search property, location, name or contact"><button class=btn onclick=load()>Search</button>
</div>
<div class=kpis>
<div class=k><b id=kTotal>0</b><span>MANUAL PROPERTIES</span></div><div class=k><b id=kToday>0</b><span>ADDED TODAY</span></div>
<div class=k><b id=kVerified>0</b><span>VERIFIED</span></div><div class=k><b id=kUnverified>0</b><span>UNVERIFIED</span></div>
<div class=k><b id=kPhotos>0</b><span>PHOTOS</span></div><div class=k><b id=kVideos>0</b><span>VIDEOS</span></div><div class=k><b id=kBrochures>0</b><span>BROCHURES</span></div>
</div>
<div class=tablewrap><table><thead><tr><th>S.No.</th><th>Property</th><th>Source</th><th>Date</th><th>Verification</th><th>Location</th><th>Area</th><th>Rent</th><th>Contact</th><th>Media</th><th>Action</th></tr></thead><tbody id=rows></tbody></table></div>
</div>
<script>
const initialDivision={json.dumps(d)};const E=x=>String(x??'').replace(/[&<>"]/g,c=>({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}}[c]));
if(['ALL','DELHI_NCR','GOA'].includes(initialDivision)) division.value=initialDivision;
function money(x){{if(x===null||x===undefined||x==='')return '';let n=Number(x);return Number.isFinite(n)?n.toLocaleString('en-IN'):E(x)}}
async function load(){{
 let u='/api/v17-7/manual-properties?division='+division.value+'&source='+source.value+'&verified='+verified.value+'&q='+encodeURIComponent(q.value||'');
 let r=await fetch(u),d=await r.json();if(!r.ok){{rows.innerHTML='<tr><td colspan=11><b>ERROR: '+E(d.detail||d.message||'Unable to load')+'</b></td></tr>';return}}
 const s=d.summary||{{}};kTotal.textContent=s.total||0;kToday.textContent=s.added_today||0;kVerified.textContent=s.verified||0;kUnverified.textContent=s.unverified||0;kPhotos.textContent=s.photos||0;kVideos.textContent=s.videos||0;kBrochures.textContent=s.brochures||0;
 const today=new Date().toISOString().slice(0,10);
 rows.innerHTML=(d.rows||[]).map((x,i)=>{{const dt=String(x.display_entry_date||x.created_at||'');const isToday=dt.slice(0,10)===today;const ver=String(x.verification_status||'UNVERIFIED').toUpperCase();
 return `<tr class="manualrow"><td class=no>${{i+1}}</td><td><b>${{E(x.property_name||x.property_code)}}</b><br><span class=small>${{E(x.property_code)}}</span></td>
 <td><span class=badge>${{E(x.display_source||'MANUAL')}}</span>${{isToday?'<br><span class="badge today">TODAY</span>':''}}</td>
 <td><b>${{E(dt.slice(0,16).replace('T',' '))}}</b><br><span class=small>${{E(x.display_entered_by||'')}}</span></td>
 <td><span class="badge ${{ver==='VERIFIED'?'':'unv'}}">${{E(ver)}}</span></td><td><b>${{E(x.location||'')}}</b><br><span class=small>${{E(x.city||'')}}</span></td>
 <td><b>${{money(x.area_sqft)}} sq ft</b></td><td><b>₹${{money(x.rent_amount)}}</b><br><span class=small>${{E(x.rent_unit||'')}}</span></td>
 <td><b>${{E(x.owner_broker_name||'')}}</b><br><b>${{E(x.contact_number||'')}}</b></td>
 <td class=media>Photos ${{x.image_count||0}} | Videos ${{x.video_count||0}} | Brochure ${{x.brochure_count||0}}</td>
 <td><a class=btn href="/property-detail-final/${{encodeURIComponent(x.property_code)}}">View Full Property</a></td></tr>`}}).join('')||'<tr><td colspan=11><b>No manual operational properties found for this filter.</b></td></tr>';
}}
division.onchange=load;source.onchange=load;verified.onchange=load;q.addEventListener('keydown',e=>{{if(e.key==='Enter')load()}});load();
</script></body></html>""")

@app.get("/property-detail-final/{property_code}",response_class=HTMLResponse)
def v177_property_detail_page(property_code:str,req:Request):
    if not page_role_or_redirect(req): return RedirectResponse("/login",303)
    _v177_setup()
    with engine.connect() as c:
        row=c.execute(text("""SELECT p.*,COALESCE(p.entry_source,'MANUAL') AS display_source,
            COALESCE(p.entered_by,p.created_by,'') AS display_entered_by,
            COALESCE(p.entry_date,p.created_at) AS display_entry_date
            FROM pi_operational_properties p WHERE p.property_code=:p"""),{"p":property_code}).first()
    if not row: return HTMLResponse("<h2>Property not found.</h2>",status_code=404)
    prop=dict(row._mapping); media=_v177_media_rows(property_code)
    imgs=[m for m in media if str(m.get("media_type")).upper()=="IMAGE"]
    vids=[m for m in media if str(m.get("media_type")).upper()=="VIDEO"]
    bros=[m for m in media if str(m.get("media_type")).upper()=="BROCHURE"]

    def val(x): return escape(str(x if x not in (None,"") else "-"))
    def arr(v):
        x=_v17_arr(v); return ", ".join(str(z) for z in x) if x else "-"

    image_html="".join(
        '<a class="thumb" target="_blank" href="/api/v17-2/property-media/{id}"><img src="/api/v17-2/property-media/{id}" alt="{fn}"><span>{fn}</span></a>'.format(
            id=m["id"],fn=escape(str(m.get("filename") or "Property photo")))
        for m in imgs
    ) or '<div class="empty">No photos uploaded.</div>'

    video_html="".join(
        '<div class="video"><video controls preload="metadata" src="/api/v17-2/property-media/{id}"></video><div><b>{fn}</b> | <a target="_blank" href="/api/v17-2/property-media/{id}">Open Video</a></div></div>'.format(
            id=m["id"],fn=escape(str(m.get("filename") or "Property video")))
        for m in vids
    ) or '<div class="empty">No videos uploaded.</div>'

    brochure_html="".join(
        '<a class="brochure" target="_blank" href="/api/v17-2/property-media/{id}">View Brochure | {fn}</a>'.format(
            id=m["id"],fn=escape(str(m.get("filename") or "Brochure")))
        for m in bros
    ) or '<div class="empty">No brochure uploaded.</div>'

    google_html=('-' if not prop.get("google_location") else '<a target="_blank" href="'+escape(str(prop.get("google_location")))+'">Open Google Location</a>')

    return HTMLResponse(f"""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1">
<title>{val(prop.get("property_name") or prop.get("property_code"))}</title>
<style>
*{{box-sizing:border-box}}body{{font-family:Arial;margin:0;background:#f4f7fb;color:#172437}}header{{background:#102235;color:#fff;padding:20px}}.w{{max-width:1450px;margin:auto;padding:18px}}
.actions{{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px}}.btn{{padding:9px 11px;border-radius:8px;background:#1677ff;color:white;text-decoration:none;font-weight:bold}}
.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:10px}}.field{{background:#fff;border:1px solid #e2e8f0;border-radius:10px;padding:12px;min-height:80px}}
.field span{{display:block;font-size:11px;color:#687789;text-transform:uppercase;margin-bottom:6px}}.field b{{font-size:16px;word-break:break-word}}
.section{{background:#fff;border:1px solid #e2e8f0;border-radius:12px;padding:15px;margin-top:14px}}.photos{{display:grid;grid-template-columns:repeat(auto-fit,minmax(190px,1fr));gap:10px}}
.thumb{{text-decoration:none;color:#172437;border:1px solid #e2e8f0;border-radius:9px;overflow:hidden;background:#fafafa}}.thumb img{{display:block;width:100%;height:180px;object-fit:cover}}.thumb span{{display:block;padding:7px;font-size:11px}}
.videos{{display:grid;grid-template-columns:repeat(auto-fit,minmax(320px,1fr));gap:12px}}.video{{border:1px solid #e2e8f0;border-radius:9px;padding:8px}}.video video{{width:100%;max-height:420px;background:#000}}
.brochure{{display:inline-block;padding:11px 13px;background:#eef4ff;border-radius:8px;text-decoration:none;font-weight:bold;margin:4px}}.big{{font-size:20px!important}}.contact{{font-size:20px!important;color:#102235}}.empty{{color:#687789;padding:10px}}
</style></head><body>
<header><b>{val(prop.get("property_name") or prop.get("property_code"))}</b><br><small>{val(prop.get("property_code"))} | {val(prop.get("division"))} | {val(prop.get("display_source"))}</small></header>
<div class=w><div class=actions><a class=btn href="/manual-property-database?division={val(prop.get("division"))}">← Manual Property Database</a><a class=btn href="/final-dashboard-v11">Dashboard</a></div>
<div class=grid>
<div class=field><span>Property Code</span><b>{val(prop.get("property_code"))}</b></div><div class=field><span>Property Name</span><b class=big>{val(prop.get("property_name"))}</b></div>
<div class=field><span>Entry Source</span><b>{val(prop.get("display_source"))}</b></div><div class=field><span>Entry Date</span><b>{val(prop.get("display_entry_date"))}</b></div>
<div class=field><span>Entered By</span><b>{val(prop.get("display_entered_by"))}</b></div><div class=field><span>Verification</span><b>{val(prop.get("verification_status"))}</b></div>
<div class=field><span>Property Types</span><b>{val(arr(prop.get("property_types")))}</b></div><div class=field><span>City</span><b>{val(prop.get("city"))}</b></div>
<div class=field><span>Location</span><b class=big>{val(prop.get("location"))}</b></div><div class=field><span>Area</span><b class=big>{val(prop.get("area_sqft"))} sq ft</b></div>
<div class=field><span>Rent</span><b class=big>₹{val(prop.get("rent_amount"))} | {val(prop.get("rent_unit"))}</b></div><div class=field><span>Transaction</span><b>{val(prop.get("transaction_type"))}</b></div>
<div class=field><span>Floor</span><b>{val(prop.get("floor"))}</b></div><div class=field><span>Frontage</span><b>{val(prop.get("frontage"))}</b></div>
<div class=field><span>Parking</span><b>{val(prop.get("parking"))}</b></div><div class=field><span>Possession</span><b>{val(prop.get("possession"))}</b></div>
<div class=field><span>Suitable For</span><b>{val(prop.get("suitable_for"))}</b></div><div class=field><span>Nearby Brands</span><b>{val(prop.get("nearby_brands"))}</b></div>
<div class=field><span>Owner / Broker / Contact</span><b class=contact>{val(prop.get("owner_broker_name"))}</b></div><div class=field><span>Contact Number</span><b class=contact>{val(prop.get("contact_number"))}</b></div>
<div class=field><span>Contact Role</span><b>{val(prop.get("contact_role"))}</b></div><div class=field><span>Google Location</span><b>{google_html}</b></div>
</div>
<div class=section><h2>Remarks</h2><b>{val(prop.get("remarks"))}</b></div>
<div class=section><h2>Photos ({len(imgs)})</h2><div class=photos>{image_html}</div></div>
<div class=section><h2>Videos ({len(vids)})</h2><div class=videos>{video_html}</div></div>
<div class=section><h2>Brochure ({len(bros)})</h2>{brochure_html}</div>
</div></body></html>""")

@app.get("/final-dashboard-v11",response_class=HTMLResponse)
def v177_dashboard(req:Request):
    role=page_role_or_redirect(req)
    if not role:return RedirectResponse("/login",303)
    admin='<a class=card href="/admin-data-tools-v2"><b>Admin Data Tools</b><p>Database health and maintenance.</p></a>' if role=="admin" else ""
    return HTMLResponse(f"""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>AI Deal Intelligence OS</title>
<style>body{{font-family:Arial;margin:0;background:#f4f7fb;color:#172437}}header{{background:#102235;color:white;padding:22px}}.w{{max-width:1500px;margin:auto;padding:20px}}.g{{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:12px;margin-bottom:24px}}.card{{display:block;background:white;border:1px solid #e2e8f0;border-radius:12px;padding:15px;text-decoration:none;color:#172437}}.card p{{font-size:12px;color:#687789}}.primary{{border:2px solid #1677ff}}.manual{{border:3px solid #08734b}}</style></head><body>
<header><b>AI Deal Intelligence OS</b><br><small>Operational Dashboard | Manual Property Database with Media</small></header><div class=w>
<h2>Manual Property Data</h2><div class=g>
<a class="card manual" href="/manual-property-database"><b>Manual Property Database</b><p>Bold numbered records. Open any property to see all details, photos, videos and brochure.</p></a>
<a class="card primary" href="/manual-property-final?division=DELHI_NCR"><b>Add Delhi NCR Property</b></a><a class="card primary" href="/manual-property-final?division=GOA"><b>Add Goa Property</b></a>
<a class=card href="/universal-recovery-doctor"><b>Universal Recovery Doctor</b><p>Recovery staging stays separate from clean manual inventory.</p></a>
</div>
<h2>Requirements & Matching</h2><div class=g><a class=card href="/manual-requirement-final?division=DELHI_NCR"><b>Add Delhi NCR Requirement</b></a><a class=card href="/matcher-final?division=DELHI_NCR"><b>Delhi NCR Matcher</b></a><a class=card href="/requirements-center-v176?division=DELHI_NCR"><b>Requirements Centre</b></a><a class=card href="/manual-requirement-final?division=GOA"><b>Add Goa Requirement</b></a><a class=card href="/matcher-final?division=GOA"><b>Goa Matcher</b></a></div>
<h2>AI, Search & Marketing</h2><div class=g><a class=card href="/property-discovery"><b>Property Discovery / Search Engine</b></a><a class=card href="/retail-expansion"><b>Retail Expansion</b></a><a class=card href="/ai-hospitality-master-final"><b>Hospitality Master</b></a><a class=card href="/hospitality-enrichment"><b>Hospitality Enrichment</b></a><a class=card href="/marketing-contacts-final"><b>Marketing Contacts</b></a><a class=card href="/phone-contact-upload"><b>Upload Phone Contacts</b></a><a class=card href="/capture-intelligence"><b>Capture Property</b></a></div>
<h2>Database & Admin</h2><div class=g><a class=card href="/property-database"><b>Full Legacy Property Database</b></a><a class=card href="/contacts-directory"><b>Owner / Broker Contacts</b></a><a class=card href="/data-doctor"><b>Data Doctor</b></a>{admin}</div>
</div></body></html>""")

@app.middleware("http")
async def v177_router(request,call_next):
    p=request.url.path
    if p in {"/workspace","/final-dashboard-v10"}:
        return RedirectResponse("/final-dashboard-v11",status_code=307)
    if p=="/fresh-inventory-final":
        q=request.url.query.upper()
        div="GOA" if "DIVISION=GOA" in q else ("DELHI_NCR" if "DIVISION=DELHI_NCR" in q else "ALL")
        return RedirectResponse(f"/manual-property-database?division={div}",status_code=307)
    response=await call_next(request)
    if p.startswith(("/final-dashboard-v11","/manual-property-database","/property-detail-final","/api/v17-7")):
        response.headers["Cache-Control"]="no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"]="no-cache";response.headers["Expires"]="0"
    return response

# ============================================================
# V17.8 SAFE MANUAL PROPERTY EDIT
# ============================================================

@app.get("/api/v17-8/property/{property_code}")
def v178_get_property(property_code:str,req:Request):
    need_login(req)
    _v177_setup()
    with engine.connect() as c:
        row=c.execute(text("SELECT * FROM pi_operational_properties WHERE property_code=:p"),{"p":property_code}).first()
    if not row: raise HTTPException(404,"Property not found.")
    return {"status":"ok","property":dict(row._mapping),"media":_v177_media_rows(property_code)}

@app.post("/api/v17-8/property/{property_code}/edit")
async def v178_edit_property(
    property_code:str, req:Request,
    property_name:str=Form(""), property_types:list[str]=Form([]),
    city:str=Form(""), location:str=Form(""), area_sqft:str=Form(""),
    rent_amount:str=Form(""), rent_unit:str=Form("MONTH"),
    transaction_type:str=Form("LEASE"), floor:str=Form(""),
    frontage:str=Form(""), parking:str=Form(""), possession:str=Form(""),
    suitable_for:str=Form(""), nearby_brands:str=Form(""),
    owner_broker_name:str=Form(""), contact_number:str=Form(""),
    contact_role:str=Form(""), verification_status:str=Form("UNVERIFIED"),
    google_location:str=Form(""), remarks:str=Form(""), entered_by:str=Form(""),
    images:list[UploadFile]=File([]), videos:list[UploadFile]=File([]),
    brochure:UploadFile|None=File(None)
):
    need_login(req); _v177_setup()
    with engine.connect() as c:
        old=c.execute(text("SELECT property_code FROM pi_operational_properties WHERE property_code=:p"),{"p":property_code}).first()
    if not old: raise HTTPException(404,"Property not found.")

    def num(v,label,required=False):
        s=str(v or "").replace(",","").strip()
        if not s:
            if required: raise HTTPException(400,f"{label} is required.")
            return None
        try:return float(s)
        except:raise HTTPException(400,f"{label} must be numeric.")

    area=num(area_sqft,"Area",True)
    rent=num(rent_amount,"Rent",False)
    if area<=0: raise HTTPException(400,"Area must be greater than 0.")
    if not location.strip(): raise HTTPException(400,"Location is required.")
    types=[x.strip() for x in property_types if x.strip()]
    if not types: raise HTTPException(400,"Select at least one Property Type.")

    with engine.begin() as c:
        c.execute(text("""UPDATE pi_operational_properties SET
          property_name=:pn,property_types=CAST(:pt AS jsonb),city=:city,location=:loc,
          area_sqft=:area,rent_amount=:rent,rent_unit=:ru,transaction_type=:tt,
          floor=:floor,frontage=:frontage,parking=:parking,possession=:possession,
          suitable_for=:sf,nearby_brands=:nb,owner_broker_name=:ob,
          contact_number=:phone,contact_role=:cr,verification_status=:vs,
          google_location=:gl,remarks=:remarks,
          entered_by=COALESCE(NULLIF(:eb,''),entered_by),updated_at=NOW()
          WHERE property_code=:pc"""),{
          "pn":property_name.strip(),"pt":json.dumps(types),"city":city.strip(),"loc":location.strip(),
          "area":area,"rent":rent,"ru":rent_unit.strip() or "MONTH","tt":transaction_type.strip() or "LEASE",
          "floor":floor.strip(),"frontage":frontage.strip(),"parking":parking.strip(),"possession":possession.strip(),
          "sf":suitable_for.strip(),"nb":nearby_brands.strip(),"ob":owner_broker_name.strip(),
          "phone":contact_number.strip(),"cr":contact_role.strip(),"vs":verification_status.upper(),
          "gl":google_location.strip(),"remarks":remarks.strip(),"eb":entered_by.strip(),"pc":property_code})

    async def add_media(files,kind):
        for f in files or []:
            if not f or not f.filename: continue
            data=await f.read()
            if not data: continue
            with engine.begin() as c:
                c.execute(text("""INSERT INTO pi_operational_property_media
                (property_code,media_type,filename,mime_type,file_size,file_data,created_at)
                VALUES(:pc,:mt,:fn,:mime,:sz,:data,NOW())"""),
                {"pc":property_code,"mt":kind,"fn":f.filename,
                 "mime":f.content_type or "application/octet-stream","sz":len(data),"data":data})
    await add_media(images,"IMAGE")
    await add_media(videos,"VIDEO")
    if brochure and brochure.filename: await add_media([brochure],"BROCHURE")
    return {"status":"ok","property_code":property_code,"message":"Property updated safely."}

@app.delete("/api/v17-8/property/{property_code}/media/{media_id}")
def v178_remove_media(property_code:str,media_id:int,req:Request):
    need_login(req)
    with engine.begin() as c:
        found=c.execute(text("SELECT id FROM pi_operational_property_media WHERE id=:id AND property_code=:pc"),
                        {"id":media_id,"pc":property_code}).first()
        if not found: raise HTTPException(404,"Media not found.")
        c.execute(text("DELETE FROM pi_operational_property_media WHERE id=:id AND property_code=:pc"),
                  {"id":media_id,"pc":property_code})
    return {"status":"ok"}

@app.get("/edit-property/{property_code}",response_class=HTMLResponse)
def v178_edit_page(property_code:str,req:Request):
    if not page_role_or_redirect(req): return RedirectResponse("/login",303)
    _v177_setup()
    with engine.connect() as c:
        row=c.execute(text("SELECT * FROM pi_operational_properties WHERE property_code=:p"),{"p":property_code}).first()
    if not row:return HTMLResponse("<h2>Property not found.</h2>",404)
    p=dict(row._mapping); media=_v177_media_rows(property_code)
    current=set(_v17_arr(p.get("property_types")))
    choices=["SHOP","RETAIL","RESTAURANT","CAFE","BANQUET","HOTEL","OFFICE","SHOWROOM","WAREHOUSE","INDUSTRIAL","COMMERCIAL","VILLA","APARTMENT","LAND","FARMHOUSE","OTHER"]
    checks="".join(f'<label><input type=checkbox name=property_types value="{escape(x)}" {"checked" if x in current else ""}> {escape(x)}</label>' for x in choices)
    media_html="".join(f'<div class=media><b>{escape(str(m["media_type"]))}</b> · {escape(str(m["filename"]))} · <a target=_blank href="/api/v17-2/property-media/{m["id"]}">View</a> · <button type=button onclick="rm({m["id"]})">Remove</button></div>' for m in media) or "No existing media."
    def V(k):
        x=p.get(k); return escape(str(x if x not in (None,"") else ""))
    ver=str(p.get("verification_status") or "UNVERIFIED").upper()
    tx=str(p.get("transaction_type") or "LEASE").upper()
    return HTMLResponse(f"""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Edit Property</title>
<style>*{{box-sizing:border-box}}body{{font-family:Arial;margin:0;background:#f4f7fb;color:#172437}}header{{background:#102235;color:white;padding:20px}}.w{{max-width:1250px;margin:auto;padding:18px}}form,.box{{background:white;padding:16px;border-radius:12px;margin-bottom:14px}}.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:12px}}input,select,textarea{{width:100%;padding:10px;border:1px solid #ccd6e2;border-radius:7px}}textarea{{min-height:90px}}label small{{display:block;font-weight:bold;margin-bottom:5px}}.types{{display:flex;gap:12px;flex-wrap:wrap}}.types input{{width:auto}}.btn,button{{padding:9px 12px;background:#1677ff;color:white;border:0;border-radius:7px;text-decoration:none;font-weight:bold;cursor:pointer}}.media{{padding:8px;border-bottom:1px solid #eee}}.notice{{background:#eefbf4;padding:10px;border-radius:8px;margin-bottom:12px}}</style></head><body>
<header><b>Edit Property</b><br><small>{V("property_code")} · same record, no duplicate</small></header><div class=w>
<div class=notice><b>Safe Edit:</b> Existing photos/videos/brochure stay attached unless Remove is clicked. Previous property values remain protected by the existing history trigger.</div>
<form id=f enctype=multipart/form-data><div class=grid>
<label><small>Property Name</small><input name=property_name value="{V("property_name")}"></label>
<label><small>City</small><input name=city value="{V("city")}"></label>
<label><small>Location *</small><input required name=location value="{V("location")}"></label>
<label><small>Google Location</small><input name=google_location value="{V("google_location")}"></label>
<label><small>Area sq ft *</small><input required name=area_sqft value="{V("area_sqft")}"></label>
<label><small>Rent</small><input name=rent_amount value="{V("rent_amount")}" placeholder="Can remain blank"></label>
<label><small>Rent Unit</small><input name=rent_unit value="{V("rent_unit")}"></label>
<label><small>Transaction</small><select name=transaction_type><option {"selected" if tx=="LEASE" else ""}>LEASE</option><option {"selected" if tx=="SALE" else ""}>SALE</option></select></label>
<label><small>Floor</small><input name=floor value="{V("floor")}"></label>
<label><small>Frontage</small><input name=frontage value="{V("frontage")}"></label>
<label><small>Parking</small><input name=parking value="{V("parking")}"></label>
<label><small>Possession</small><input name=possession value="{V("possession")}"></label>
<label><small>Suitable For</small><input name=suitable_for value="{V("suitable_for")}"></label>
<label><small>Nearby Brands</small><input name=nearby_brands value="{V("nearby_brands")}"></label>
<label><small>Owner/Broker/Contact</small><input name=owner_broker_name value="{V("owner_broker_name")}"></label>
<label><small>Contact Number</small><input name=contact_number value="{V("contact_number")}"></label>
<label><small>Contact Role</small><input name=contact_role value="{V("contact_role")}"></label>
<label><small>Verification</small><select name=verification_status><option {"selected" if ver=="UNVERIFIED" else ""}>UNVERIFIED</option><option {"selected" if ver=="VERIFIED" else ""}>VERIFIED</option></select></label>
<label><small>Updated By</small><input name=entered_by value="{V("entered_by")}"></label>
</div><h3>Property Type *</h3><div class=types>{checks}</div><h3>Remarks</h3><textarea name=remarks>{V("remarks")}</textarea>
<h3>Add More Media</h3><div class=grid><label><small>Photos</small><input type=file name=images accept="image/*" multiple></label><label><small>Videos</small><input type=file name=videos accept="video/*" multiple></label><label><small>Brochure</small><input type=file name=brochure accept=".pdf,application/pdf"></label></div>
<p><button type=submit>Save Changes</button> <a class=btn href="/property-detail-final/{escape(property_code)}">Cancel</a> <b id=msg></b></p></form>
<div class=box><h2>Existing Media</h2>{media_html}</div></div>
<script>const code={json.dumps(property_code)};f.onsubmit=async e=>{{e.preventDefault();msg.textContent='Saving...';let r=await fetch('/api/v17-8/property/'+encodeURIComponent(code)+'/edit',{{method:'POST',body:new FormData(f)}}),d=await r.json();if(!r.ok){{msg.textContent='ERROR: '+(d.detail||'Save failed');return}}location.href='/property-detail-final/'+encodeURIComponent(code)}};async function rm(id){{if(!confirm('Remove this media file?'))return;let r=await fetch('/api/v17-8/property/'+encodeURIComponent(code)+'/media/'+id,{{method:'DELETE'}});if(r.ok)location.reload();else alert('Unable to remove media')}};</script></body></html>""")

@app.get("/manual-property-database-v178",response_class=HTMLResponse)
def v178_database(req:Request,division:str=Query("ALL")):
    if not page_role_or_redirect(req):return RedirectResponse("/login",303)
    d=division.upper()
    return HTMLResponse(f"""<!doctype html><html><head><meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1"><title>Manual Property Database</title>
<style>body{{font-family:Arial;margin:0;background:#f4f7fb;color:#172437}}header{{background:#102235;color:white;padding:20px}}.w{{padding:18px}}.btn{{padding:8px 10px;background:#1677ff;color:white;text-decoration:none;border-radius:7px;font-weight:bold}}.edit{{background:#08734b}}table{{width:100%;border-collapse:collapse;background:white;margin-top:15px}}th,td{{padding:9px;border-bottom:1px solid #eee;text-align:left}}td{{font-weight:bold}}select{{padding:8px}}</style></head><body><header><b>Manual Property Database · View / Edit</b></header><div class=w><a class=btn href="/final-dashboard-v12">← Dashboard</a> <select id=division><option value=ALL>ALL</option><option value=DELHI_NCR>DELHI NCR</option><option value=GOA>GOA</option></select><table><thead><tr><th>No.</th><th>Property</th><th>Location</th><th>Area</th><th>Rent</th><th>Contact</th><th>Media</th><th>Actions</th></tr></thead><tbody id=rows></tbody></table></div>
<script>const init={json.dumps(d)};if(['ALL','DELHI_NCR','GOA'].includes(init))division.value=init;const E=x=>String(x??'');async function load(){{let d=await(await fetch('/api/v17-7/manual-properties?division='+division.value+'&source=MANUAL&verified=ALL&q=')).json();rows.innerHTML=(d.rows||[]).map((x,i)=>`<tr><td>${{i+1}}</td><td>${{E(x.property_name||x.property_code)}}<br>${{E(x.property_code)}}</td><td>${{E(x.location)}}</td><td>${{E(x.area_sqft)}} sq ft</td><td>${{x.rent_amount==null?'':('₹'+E(x.rent_amount))}}</td><td>${{E(x.owner_broker_name)}}<br>${{E(x.contact_number)}}</td><td>Photos ${{x.image_count||0}} | Videos ${{x.video_count||0}} | Brochure ${{x.brochure_count||0}}</td><td><a class=btn href="/property-detail-final/${{encodeURIComponent(x.property_code)}}">View</a> <a class="btn edit" href="/edit-property/${{encodeURIComponent(x.property_code)}}">Edit Property</a></td></tr>`).join('')||'<tr><td colspan=8>No properties found.</td></tr>'}}division.onchange=load;load()</script></body></html>""")

@app.get("/final-dashboard-v12",response_class=HTMLResponse)
def v178_dashboard(req:Request):
    if not page_role_or_redirect(req):return RedirectResponse("/login",303)
    return HTMLResponse("""<!doctype html><html><head><meta charset=utf-8><title>AI Deal Intelligence OS</title><style>body{font-family:Arial;margin:0;background:#f4f7fb;color:#172437}header{background:#102235;color:white;padding:22px}.w{padding:20px}.g{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:12px}.card{display:block;background:white;padding:15px;border-radius:12px;text-decoration:none;color:#172437;border:1px solid #ddd}.main{border:3px solid #08734b}</style></head><body><header><b>AI Deal Intelligence OS · V17.8</b></header><div class=w><div class=g><a class="card main" href="/manual-property-database-v178"><b>Manual Property Database · View / Edit</b></a><a class=card href="/manual-property-final?division=DELHI_NCR"><b>Add Delhi NCR Property</b></a><a class=card href="/manual-property-final?division=GOA"><b>Add Goa Property</b></a><a class=card href="/requirements-center-v176?division=DELHI_NCR"><b>Requirements Centre</b></a><a class=card href="/matcher-final?division=DELHI_NCR"><b>Property Matcher</b></a><a class=card href="/property-discovery"><b>Property Discovery / Search Engine</b></a><a class=card href="/retail-expansion"><b>Retail Bot</b></a><a class=card href="/ai-hospitality-master-final"><b>Hospitality Bot / Master</b></a><a class=card href="/marketing-contacts-final"><b>Marketing Contacts</b></a><a class=card href="/phone-contact-upload"><b>Upload Phone Contacts</b></a><a class=card href="/universal-recovery-doctor"><b>Recovery Doctor</b></a></div></div></body></html>""")

@app.middleware("http")
async def v178_router(request,call_next):
    p=request.url.path
    if p in {"/workspace","/final-dashboard-v11"}:return RedirectResponse("/final-dashboard-v12",307)
    if p=="/manual-property-database":return RedirectResponse("/manual-property-database-v178"+(("?"+request.url.query) if request.url.query else ""),307)
    response=await call_next(request)
    if p.startswith(("/final-dashboard-v12","/manual-property-database-v178","/edit-property","/api/v17-8")):
        response.headers["Cache-Control"]="no-store, no-cache, must-revalidate, max-age=0"
    return response

# ============================================================
# V17.8.2 FINAL UI REPAIR
# Restore full V17.7 dashboard.
# Repair ONLY Manual Property Database.
# Keep existing V17.8 Edit Property route/functions.
# ============================================================

@app.get("/manual-property-database-v1782", response_class=HTMLResponse)
def v1782_manual_property_database(req:Request, division:str=Query("ALL")):
    if not page_role_or_redirect(req):
        return RedirectResponse("/login",303)

    d=division.upper()

    return HTMLResponse(f"""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Manual Property Database</title>
<style>
*{{box-sizing:border-box}}
body{{font-family:Arial;margin:0;background:#f4f7fb;color:#172437}}
header{{background:#102235;color:#fff;padding:20px}}
.w{{width:100%;padding:18px}}
.toolbar{{display:flex;gap:8px;align-items:center;flex-wrap:wrap;margin-bottom:12px}}
.btn{{display:inline-block;padding:9px 11px;border:0;border-radius:8px;background:#1677ff;color:#fff;text-decoration:none;font-weight:700;cursor:pointer;white-space:nowrap}}
.btn.edit{{background:#08734b}}
.btn.gray{{background:#e8edf3;color:#243447}}
input,select{{padding:9px;border:1px solid #ccd6e2;border-radius:7px}}
input{{min-width:280px}}
.notice{{background:#eef4ff;border-radius:9px;padding:10px;margin:10px 0;font-size:12px}}
.kpis{{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:9px;margin:14px 0}}
.kpi{{background:#fff;border:1px solid #e2e8f0;border-radius:10px;padding:12px}}
.kpi b{{font-size:23px;display:block}}
.tablebox{{width:100%;overflow:auto;background:#fff;border:1px solid #e2e8f0;border-radius:10px;max-height:73vh}}
table{{border-collapse:collapse;min-width:2650px;width:100%;font-size:12px}}
th,td{{padding:9px;border-bottom:1px solid #edf1f5;text-align:left;vertical-align:top;white-space:nowrap}}
th{{background:#f8fafc;position:sticky;top:0;z-index:4}}
td{{font-weight:700}}
.sno{{font-size:16px;font-weight:900}}
.wrap{{white-space:normal;min-width:170px;max-width:280px}}
.small{{font-size:11px;color:#687789;font-weight:400}}
.badge{{display:inline-block;border-radius:10px;padding:3px 7px;background:#dcfce7;color:#166534;font-size:10px;font-weight:800}}
.badge.unv{{background:#fef3c7;color:#92400e}}
.badge.today{{background:#dbeafe;color:#1d4ed8}}
.actions{{display:flex;gap:6px;min-width:220px}}
.actions .btn{{padding:7px 9px}}
.sticky1{{position:sticky;left:0;background:#fff;z-index:2}}
.sticky2{{position:sticky;left:58px;background:#fff;z-index:2}}
thead .sticky1,thead .sticky2{{background:#f8fafc;z-index:6}}
</style>
</head>
<body>

<header>
<b>Manual Property Database · View / Edit</b><br>
<small>Complete manual property fields · bold values · full media view · safe editing</small>
</header>

<div class="w">

<div class="toolbar">
<a class="btn gray" href="/final-dashboard-v11">← Full Dashboard</a>
<a class="btn" href="/manual-property-final?division=DELHI_NCR">Add Delhi NCR Property</a>
<a class="btn" href="/manual-property-final?division=GOA">Add Goa Property</a>

<select id="division">
<option value="ALL">ALL AREAS</option>
<option value="DELHI_NCR">DELHI NCR</option>
<option value="GOA">GOA</option>
</select>

<select id="source">
<option value="MANUAL">MANUAL + RECOVERED MANUAL</option>
<option value="ALL">ALL OPERATIONAL SOURCES</option>
<option value="RECOVERED_MANUAL">RECOVERED MANUAL ONLY</option>
</select>

<select id="verified">
<option value="ALL">ALL VERIFICATION</option>
<option value="VERIFIED">VERIFIED</option>
<option value="UNVERIFIED">UNVERIFIED</option>
</select>

<input id="q" placeholder="Search property, location, contact or name">
<button class="btn" onclick="load()">Search</button>
</div>

<div class="notice">
<b>All fields are preserved.</b> Scroll horizontally for the complete record.
Use <b>View Full Property</b> for pictures/videos/brochure and <b>Edit Property</b> to change the same property record.
</div>

<div class="kpis">
<div class="kpi"><b id="kTotal">0</b>Total Manual Properties</div>
<div class="kpi"><b id="kToday">0</b>Added Today</div>
<div class="kpi"><b id="kVerified">0</b>Verified</div>
<div class="kpi"><b id="kUnverified">0</b>Unverified</div>
<div class="kpi"><b id="kPhotos">0</b>Photos</div>
<div class="kpi"><b id="kVideos">0</b>Videos</div>
<div class="kpi"><b id="kBrochures">0</b>Brochures</div>
</div>

<div class="tablebox">
<table>
<thead>
<tr>
<th class="sticky1">S.No.</th>
<th class="sticky2">Property / Code</th>
<th>Entry Source</th>
<th>Entry Date</th>
<th>Entered By</th>
<th>Verification</th>
<th>Property Type</th>
<th>City</th>
<th>Location</th>
<th>Area</th>
<th>Rent</th>
<th>Rent Unit</th>
<th>Transaction</th>
<th>Floor</th>
<th>Frontage</th>
<th>Parking</th>
<th>Possession</th>
<th>Suitable For</th>
<th>Nearby Brands</th>
<th>Owner / Broker / Contact</th>
<th>Contact No.</th>
<th>Contact Role</th>
<th>Google Location</th>
<th>Photos</th>
<th>Videos</th>
<th>Brochure</th>
<th>Remarks</th>
<th>Actions</th>
</tr>
</thead>
<tbody id="rows"></tbody>
</table>
</div>
</div>

<script>
const initialDivision={json.dumps(d)};
if(['ALL','DELHI_NCR','GOA'].includes(initialDivision)) division.value=initialDivision;

const E=x=>String(x??'').replace(/[&<>"]/g,c=>({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}}[c]));

const M=x=>{{
  if(x===null||x===undefined||x==='') return '';
  const n=Number(x);
  return Number.isFinite(n)?n.toLocaleString('en-IN'):E(x);
}};

const TYPES=x=>{{
  if(Array.isArray(x)) return x.join(', ');
  if(!x) return '';
  if(typeof x==='string'){{
    try{{
      const v=JSON.parse(x);
      if(Array.isArray(v)) return v.join(', ');
    }}catch(e){{}}
  }}
  return String(x);
}};

async function load(){{
  const u='/api/v17-7/manual-properties'
    +'?division='+encodeURIComponent(division.value)
    +'&source='+encodeURIComponent(source.value)
    +'&verified='+encodeURIComponent(verified.value)
    +'&q='+encodeURIComponent(q.value||'');

  const r=await fetch(u,{{cache:'no-store'}});
  const d=await r.json();

  if(!r.ok){{
    rows.innerHTML='<tr><td colspan="28"><b>ERROR: '+E(d.detail||d.message||'Unable to load records')+'</b></td></tr>';
    return;
  }}

  const s=d.summary||{{}};
  kTotal.textContent=s.total||0;
  kToday.textContent=s.added_today||0;
  kVerified.textContent=s.verified||0;
  kUnverified.textContent=s.unverified||0;
  kPhotos.textContent=s.photos||0;
  kVideos.textContent=s.videos||0;
  kBrochures.textContent=s.brochures||0;

  const today=new Date().toISOString().slice(0,10);

  rows.innerHTML=(d.rows||[]).map((x,i)=>{{
    const dt=String(x.display_entry_date||x.created_at||'');
    const isToday=dt.slice(0,10)===today;
    const ver=String(x.verification_status||'UNVERIFIED').toUpperCase();

    const mapLink=x.google_location
      ? `<a target="_blank" href="${{E(x.google_location)}}">Open Map</a>`
      : '';

    return `<tr>
      <td class="sticky1 sno">${{i+1}}</td>

      <td class="sticky2 wrap">
        <b>${{E(x.property_name||x.property_code)}}</b><br>
        <span class="small">${{E(x.property_code)}}</span>
      </td>

      <td>
        <span class="badge">${{E(x.display_source||'MANUAL')}}</span>
        ${{isToday?'<br><span class="badge today">TODAY</span>':''}}
      </td>

      <td><b>${{E(dt.slice(0,16).replace('T',' '))}}</b></td>
      <td>${{E(x.display_entered_by||'')}}</td>

      <td>
        <span class="badge ${{ver==='VERIFIED'?'':'unv'}}">${{E(ver)}}</span>
      </td>

      <td class="wrap">${{E(TYPES(x.property_types))}}</td>
      <td>${{E(x.city||'')}}</td>
      <td class="wrap"><b>${{E(x.location||'')}}</b></td>
      <td><b>${{M(x.area_sqft)}} sq ft</b></td>
      <td><b>${{x.rent_amount===null||x.rent_amount===''?'':'₹'+M(x.rent_amount)}}</b></td>
      <td>${{E(x.rent_unit||'')}}</td>
      <td>${{E(x.transaction_type||'')}}</td>
      <td>${{E(x.floor||'')}}</td>
      <td>${{E(x.frontage||'')}}</td>
      <td>${{E(x.parking||'')}}</td>
      <td>${{E(x.possession||'')}}</td>
      <td class="wrap">${{E(x.suitable_for||'')}}</td>
      <td class="wrap">${{E(x.nearby_brands||'')}}</td>
      <td class="wrap"><b>${{E(x.owner_broker_name||'')}}</b></td>
      <td><b>${{E(x.contact_number||'')}}</b></td>
      <td>${{E(x.contact_role||'')}}</td>
      <td>${{mapLink}}</td>
      <td><b>${{x.image_count||0}}</b></td>
      <td><b>${{x.video_count||0}}</b></td>
      <td><b>${{x.brochure_count||0}}</b></td>
      <td class="wrap">${{E(x.remarks||'')}}</td>

      <td>
        <div class="actions">
          <a class="btn" href="/property-detail-final/${{encodeURIComponent(x.property_code)}}">View Full Property</a>
          <a class="btn edit" href="/edit-property/${{encodeURIComponent(x.property_code)}}">Edit Property</a>
        </div>
      </td>
    </tr>`;
  }}).join('') || '<tr><td colspan="28"><b>No manual properties found for this filter.</b></td></tr>';
}}

division.onchange=load;
source.onchange=load;
verified.onchange=load;
q.addEventListener('keydown',e=>{{if(e.key==='Enter') load();}});
load();
</script>

</body>
</html>""")

@app.middleware("http")
async def v1782_final_ui_router(request,call_next):
    p=request.url.path

    # Restore the complete V17.7 dashboard, not the simplified V17.8 dashboard.
    if p in {"/workspace","/final-dashboard-v12"}:
        return RedirectResponse("/final-dashboard-v11",status_code=307)

    # Replace only Manual Property Database.
    if p in {"/manual-property-database","/manual-property-database-v178","/manual-property-database-v1781"}:
        suffix=("?"+request.url.query) if request.url.query else ""
        return RedirectResponse("/manual-property-database-v1782"+suffix,status_code=307)

    response=await call_next(request)

    if p.startswith("/manual-property-database-v1782"):
        response.headers["Cache-Control"]="no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"]="no-cache"
        response.headers["Expires"]="0"

    return response

# ============================================================
# V17.8.3 REDIRECT LOOP FIX
# Breaks V17.7 <-> V17.8 dashboard redirect loop.
# Uses the full existing V17.7 dashboard content via v177_dashboard().
# Keeps V17.8.2 Manual Property Database repair.
# ============================================================

@app.get("/final-dashboard-v13", response_class=HTMLResponse)
def v1783_final_dashboard(req:Request):
    # Reuse the complete V17.7 dashboard renderer without redirecting through v11/v12.
    return v177_dashboard(req)

@app.middleware("http")
async def v1783_redirect_loop_fix(request, call_next):
    p=request.url.path

    # Send every old dashboard entry to one neutral final route
    # that older middleware does not know about.
    if p in {
        "/workspace",
        "/final-dashboard",
        "/final-dashboard-v11",
        "/final-dashboard-v12"
    }:
        return RedirectResponse("/final-dashboard-v13", status_code=307)

    # Ensure every old Manual Property Database entry reaches the repaired V17.8.2 page
    # before older middleware can redirect it elsewhere.
    if p in {
        "/manual-property-database",
        "/manual-property-database-v178",
        "/manual-property-database-v1781"
    }:
        suffix=("?"+request.url.query) if request.url.query else ""
        return RedirectResponse("/manual-property-database-v1782"+suffix, status_code=307)

    response=await call_next(request)

    if p.startswith(("/final-dashboard-v13","/manual-property-database-v1782")):
        response.headers["Cache-Control"]="no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"]="no-cache"
        response.headers["Expires"]="0"

    return response
