import io
import json
import re
import hashlib
from datetime import datetime
from html import escape
from typing import Any

from fastapi import APIRouter, Request, UploadFile, File, Form, Query, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, Response, RedirectResponse
from sqlalchemy import text
from PIL import Image, ImageOps
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

VERSION = "8.1-INTEGRATED-POSTGRES"

SYSTEM_PROMPT = r'''
You are a specialist Indian real-estate newspaper data extractor.
Build a CLEAN PROPERTY DATABASE from the WHOLE newspaper image.

COVERAGE RULES
1. Scan the ENTIRE image from top-left to bottom-right. Do not stop after large ads.
2. Capture EVERY independently advertised property, requirement, wanted listing, commercial listing, farm/land listing, builder-floor listing, apartment listing, plot listing, lease listing, and useful rent-roll market-data item.
3. ONE PROPERTY OR DISTINCT PROPERTY OFFER = ONE JSON OBJECT.
4. If one advertisement lists multiple different properties, create one object per property and repeat the advertiser/contact details on each.
5. The same phone number does NOT mean duplicate. Brokers can advertise many properties.
6. Do not omit a property because some fields are unreadable. Use Unknown / Price on request / - and mark completeness Partial.
7. Keep all clearly associated phone numbers together separated by " / ".
8. Preserve useful details: floor, bedrooms, terrace, park facing, road width, corner, plot size, rent, ROI, possession, collaboration, booking, etc.
9. Do not invent. If uncertain, preserve only readable information and mark Partial or Needs Review.
10. Ignore non-property news/articles.
11. Output ONLY valid JSON. No markdown fences. No explanation.

OUTPUT
Return a JSON array. Every object must contain exactly these keys:
{
  "lead_type": "Available - Sale | Available - Rent | Available - Lease | Available - Commercial | Booking - Floor | New/Booking - Floor | Wanted | Wanted & Available | Available/Wanted | Requirement - Buy | Requirement - Rent | Market Data - Rent Roll | Unknown",
  "locality": "specific locality/project/sector/property name",
  "area": "area exactly as readable",
  "configuration_details": "clean concise property description",
  "price": "price or Price on request",
  "agency_brand": "agency/company or -",
  "contact_person": "person name or -",
  "phone_numbers": "phone(s) or -",
  "notes": "important additional information or -",
  "source": "Newspaper - Property/<section or locality group if clear>",
  "completeness": "Complete | Partial | Partial - area not specified | Partial - price not specified | Partial - cropped | Needs Review"
}

QUALITY TARGET
The result must look like a manually cleaned Excel property database, not OCR text.
Missing a real property is worse than returning a Partial row.
'''

LEAD_TYPES = [
    "Available - Sale","Available - Rent","Available - Lease","Available - Commercial",
    "Booking - Floor","New/Booking - Floor","Wanted","Wanted & Available","Available/Wanted",
    "Requirement - Buy","Requirement - Rent","Market Data - Rent Roll","Unknown"
]
VERIFICATION_OPTIONS = ["Unverified","Verified","Not Reachable","Archived"]
COMPLETENESS_OPTIONS = ["Complete","Partial","Partial - area not specified","Partial - price not specified","Partial - cropped","Needs Review"]

SCHEMA = r'''
CREATE TABLE IF NOT EXISTS pi_newspaper_sources(
 id BIGSERIAL PRIMARY KEY,
 source_hash VARCHAR(64) UNIQUE NOT NULL,
 original_filename TEXT,
 mime_type TEXT,
 image_content BYTEA NOT NULL,
 source_label TEXT DEFAULT 'Newspaper - Property',
 ai_model TEXT,
 extraction_status VARCHAR(50) DEFAULT 'RECEIVED',
 extracted_records INTEGER DEFAULT 0,
 duplicate_records INTEGER DEFAULT 0,
 error_message TEXT,
 raw_ai_text TEXT,
 created_at TIMESTAMPTZ DEFAULT NOW(),
 updated_at TIMESTAMPTZ DEFAULT NOW()
);
CREATE TABLE IF NOT EXISTS pi_newspaper_properties(
 id BIGSERIAL PRIMARY KEY,
 record_id VARCHAR(50) UNIQUE NOT NULL,
 source_id BIGINT NOT NULL REFERENCES pi_newspaper_sources(id) ON DELETE CASCADE,
 fingerprint VARCHAR(64) NOT NULL,
 date_captured DATE DEFAULT CURRENT_DATE,
 lead_type TEXT,
 locality TEXT,
 area TEXT,
 configuration_details TEXT,
 price TEXT,
 agency_brand TEXT,
 contact_person TEXT,
 phone_numbers TEXT,
 notes TEXT,
 source TEXT,
 completeness TEXT DEFAULT 'Partial',
 verification TEXT DEFAULT 'Unverified',
 team_member TEXT,
 created_at TIMESTAMPTZ DEFAULT NOW(),
 updated_at TIMESTAMPTZ DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_pi_newspaper_fp ON pi_newspaper_properties(fingerprint);
CREATE INDEX IF NOT EXISTS idx_pi_newspaper_locality ON pi_newspaper_properties(locality);
CREATE INDEX IF NOT EXISTS idx_pi_newspaper_phone ON pi_newspaper_properties(phone_numbers);
CREATE INDEX IF NOT EXISTS idx_pi_newspaper_source ON pi_newspaper_properties(source_id);
'''


def _clean(value, fallback="-"):
    if value is None:
        return fallback
    s = str(value).strip()
    if not s or s.lower() in {"null","none","n/a","na"}:
        return fallback
    return s


def _norm(value):
    s = str(value or "").lower().strip()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _norm_phone(value):
    d = re.sub(r"\D", "", str(value or ""))
    if len(d) > 10 and d.startswith("91"):
        d = d[-10:]
    return d


def canonical(item: dict, source_label: str):
    return {
        "lead_type": _clean(item.get("lead_type") or item.get("Lead Type"), "Unknown"),
        "locality": _clean(item.get("locality") or item.get("Locality"), "Unknown"),
        "area": _clean(item.get("area") or item.get("Area (Sq Yds)"), "Unknown"),
        "configuration_details": _clean(item.get("configuration_details") or item.get("configuration") or item.get("Configuration/Details"), "Details not fully readable"),
        "price": _clean(item.get("price") or item.get("Price"), "Price on request"),
        "agency_brand": _clean(item.get("agency_brand") or item.get("Agency/Brand"), "-"),
        "contact_person": _clean(item.get("contact_person") or item.get("Contact Person"), "-"),
        "phone_numbers": _clean(item.get("phone_numbers") or item.get("Phone Number(s)"), "-"),
        "notes": _clean(item.get("notes") or item.get("Notes"), "-"),
        "source": _clean(item.get("source") or item.get("Source"), source_label),
        "completeness": _clean(item.get("completeness") or item.get("Completeness"), "Partial"),
    }


def fingerprint(row: dict):
    phones = sorted({_norm_phone(x) for x in re.split(r"[/,;|]", str(row.get("phone_numbers") or "")) if len(_norm_phone(x)) >= 8})
    parts = [
        _norm(row.get("lead_type")), _norm(row.get("locality")), _norm(row.get("area")),
        _norm(row.get("configuration_details"))[:220], _norm(row.get("price")), "|".join(phones)
    ]
    return hashlib.sha256("||".join(parts).encode()).hexdigest()


def parse_json_array(raw: str):
    t = (raw or "").strip()
    t = re.sub(r"^```(?:json)?\s*", "", t, flags=re.I)
    t = re.sub(r"\s*```$", "", t)
    a, b = t.find("["), t.rfind("]")
    if a >= 0 and b > a:
        t = t[a:b+1]
    data = json.loads(t)
    if isinstance(data, dict):
        for key in ("properties","records","data","results"):
            if isinstance(data.get(key), list):
                data = data[key]
                break
    if not isinstance(data, list):
        raise ValueError("AI response was not a JSON array")
    return data


def _resp_text(resp):
    if getattr(resp, "text", None):
        return resp.text
    try:
        return "".join(getattr(p, "text", "") for p in resp.candidates[0].content.parts)
    except Exception:
        return ""


def _generate(client, requested_model: str, contents):
    candidates=[]
    for name in [requested_model, "gemini-3.6-flash", "gemini-3.5-flash", "gemini-3.5-flash-lite", "gemini-3.1-flash-lite"]:
        if name and name not in candidates:
            candidates.append(name)
    errors=[]
    for name in candidates:
        try:
            resp=client.models.generate_content(model=name, contents=contents)
            return resp, name
        except Exception as exc:
            errors.append(f"{name}: {exc}")
    raise RuntimeError("No Gemini Vision model was available. " + " | ".join(errors))


def extract_all(client, image: Image.Image, requested_model: str, high_accuracy=True):
    first_resp, active = _generate(client, requested_model, [SYSTEM_PROMPT, image])
    first_raw = _resp_text(first_resp)
    first = parse_json_array(first_raw)
    if not high_accuracy:
        return first, first_raw, active
    existing = json.dumps(first, ensure_ascii=False)[:28000]
    audit_prompt = f'''Perform a SECOND independent coverage audit of the SAME newspaper image.
Records already extracted in pass 1:
{existing}
Scan the whole image again, especially small boxes, edges, narrow columns, and ads sharing broker phone numbers.
Return ONLY genuinely missed property records using exactly the same JSON schema.
Do not repeat records already represented above. Same phone with a different property/locality is NOT a duplicate.
If nothing was missed, return [].'''
    audit_resp, audit_model = _generate(client, active, [audit_prompt, image])
    audit_raw = _resp_text(audit_resp)
    missed = parse_json_array(audit_raw)
    return first + missed, first_raw + "\n\n--- COVERAGE AUDIT ---\n" + audit_raw, audit_model


def _dedupe(rows, source_label):
    out=[]; seen=set()
    for item in rows:
        row=canonical(item, source_label)
        fp=fingerprint(row)
        if fp in seen:
            continue
        seen.add(fp); out.append((row, fp))
    return out


def _safe_json(row):
    return {k: (v.isoformat() if hasattr(v, "isoformat") else v) for k,v in dict(row).items()}


def _html_page():
    return r'''<!doctype html><html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Newspaper Property Capture</title><style>
:root{--bg:#f5f7fb;--card:#fff;--text:#172033;--muted:#6b7280;--brand:#b42318;--line:#e5e7eb;--green:#027a48}
*{box-sizing:border-box}body{margin:0;font-family:Arial,sans-serif;background:var(--bg);color:var(--text)}header{background:#111827;color:#fff;padding:16px 22px;display:flex;justify-content:space-between;align-items:center}header a{color:#fff;text-decoration:none}.wrap{max-width:1500px;margin:0 auto;padding:20px}.grid{display:grid;grid-template-columns:1fr 1fr;gap:16px}.card{background:#fff;border:1px solid var(--line);border-radius:14px;padding:18px;margin-bottom:16px;box-shadow:0 1px 2px #00000008}h1,h2,h3{margin-top:0}.muted{color:var(--muted)}input,select,textarea{width:100%;padding:10px 11px;border:1px solid #d1d5db;border-radius:8px;margin:5px 0 10px;background:white}button,.btn{border:0;border-radius:8px;padding:10px 14px;background:var(--brand);color:white;font-weight:700;cursor:pointer;text-decoration:none;display:inline-block}.btn.secondary{background:#374151}.btn.green{background:var(--green)}table{width:100%;border-collapse:collapse;font-size:13px}th,td{padding:9px;border-bottom:1px solid var(--line);text-align:left;vertical-align:top}th{position:sticky;top:0;background:#f9fafb}.tablewrap{overflow:auto;max-height:600px;border:1px solid var(--line);border-radius:10px}.status{padding:10px;border-radius:8px;background:#eef2ff;margin:10px 0;white-space:pre-wrap}.rowactions button{padding:6px 9px;margin-right:4px}.cols3{display:grid;grid-template-columns:repeat(3,1fr);gap:10px}@media(max-width:900px){.grid,.cols3{grid-template-columns:1fr}}
</style></head><body><header><div><b>Property Intelligence · Newspaper Capture</b><br><small>Full-page AI Vision → clean records → persistent PostgreSQL</small></div><div><a href="/workspace">← Dashboard</a></div></header><div class="wrap">
<div class="grid"><div class="card"><h2>📷 Capture Newspaper</h2><p class="muted">Upload the full newspaper photo. High Accuracy runs a second pass for missed small ads. One property = one database row.</p><form id="uploadForm"><label>Newspaper image</label><input type="file" name="file" accept="image/*" capture="environment" required><label>Source label</label><input name="source_label" value="Newspaper - Property"><label><input style="width:auto" type="checkbox" name="high_accuracy" checked> High Accuracy second coverage pass</label><br><label><input style="width:auto" type="checkbox" name="replace_same" checked> Replace earlier unverified extraction when reprocessing same newspaper</label><br><br><button>Process Full Newspaper</button></form><div id="status" class="status">Ready.</div></div>
<div class="card"><h2>🔎 Search Database</h2><input id="q" placeholder="Search locality, phone, broker, price, details..."><div><button onclick="loadRows()">Search</button> <a class="btn secondary" href="/api/newspaper/export.xlsx">Download Excel</a></div><p class="muted" id="count"></p></div></div>
<div class="card"><h2>📰 Newspaper Property Database</h2><div class="tablewrap"><table><thead><tr><th>Date</th><th>Lead Type</th><th>Locality</th><th>Area</th><th>Configuration / Details</th><th>Price</th><th>Agency / Brand</th><th>Contact</th><th>Phone</th><th>Verification</th><th>Actions</th></tr></thead><tbody id="rows"></tbody></table></div></div>
<div class="card" id="editCard" style="display:none"><h2>Edit Property</h2><form id="editForm"><input type="hidden" name="record_id" id="record_id"><div class="cols3"><div><label>Lead Type</label><input name="lead_type" id="lead_type"><label>Locality</label><input name="locality" id="locality"><label>Area</label><input name="area" id="area"><label>Price</label><input name="price" id="price"></div><div><label>Agency / Brand</label><input name="agency_brand" id="agency_brand"><label>Contact Person</label><input name="contact_person" id="contact_person"><label>Phone Number(s)</label><input name="phone_numbers" id="phone_numbers"><label>Verification</label><select name="verification" id="verification"><option>Unverified</option><option>Verified</option><option>Not Reachable</option><option>Archived</option></select></div><div><label>Completeness</label><input name="completeness" id="completeness"><label>Team Member</label><input name="team_member" id="team_member"><label>Source</label><input name="source" id="source"></div></div><label>Configuration / Details</label><textarea name="configuration_details" id="configuration_details" rows="4"></textarea><label>Notes</label><textarea name="notes" id="notes" rows="3"></textarea><button>Save Changes</button> <button type="button" class="btn secondary" onclick="document.getElementById('editCard').style.display='none'">Cancel</button></form></div>
</div><script>
const $=id=>document.getElementById(id); async function J(url,opt){let r=await fetch(url,opt);let t=await r.text();let d;try{d=JSON.parse(t)}catch{d={message:t}}if(!r.ok)throw new Error(d.message||d.detail||t);return d}
$('uploadForm').onsubmit=async e=>{e.preventDefault();$('status').textContent='Processing full page with AI Vision...';let fd=new FormData(e.target);fd.set('high_accuracy',e.target.high_accuracy.checked?'true':'false');fd.set('replace_same',e.target.replace_same.checked?'true':'false');try{let d=await J('/api/newspaper/process',{method:'POST',body:fd});$('status').textContent=`Saved ${d.inserted} unique properties. ${d.skipped} duplicates skipped. Model: ${d.model}`;await loadRows()}catch(err){$('status').textContent='ERROR: '+err.message}}
function esc(s){return String(s??'').replace(/[&<>"']/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[m]))}
let CACHE={};async function loadRows(){let d=await J('/api/newspaper/records?q='+encodeURIComponent($('q').value));CACHE=Object.fromEntries(d.rows.map(x=>[x.record_id,x]));$('count').textContent=d.rows.length+' property record(s)';$('rows').innerHTML=d.rows.map(x=>`<tr><td>${esc(x.date_captured||'')}</td><td>${esc(x.lead_type||'')}</td><td><b>${esc(x.locality||'')}</b></td><td>${esc(x.area||'')}</td><td>${esc(x.configuration_details||'')}</td><td>${esc(x.price||'')}</td><td>${esc(x.agency_brand||'')}</td><td>${esc(x.contact_person||'')}</td><td><b>${esc(x.phone_numbers||'')}</b></td><td>${esc(x.verification||'')}</td><td class=rowactions><button class="btn green" onclick="editRowById('${x.record_id}')">Edit</button><button onclick="delRow('${x.record_id}')">Delete</button></td></tr>`).join('')}
function editRowById(id){let x=CACHE[id];if(!x)return;for(let k of ['record_id','lead_type','locality','area','price','agency_brand','contact_person','phone_numbers','verification','completeness','team_member','source','configuration_details','notes'])if($(k))$(k).value=x[k]||'';$('editCard').style.display='block';$('editCard').scrollIntoView({behavior:'smooth'})}
$('editForm').onsubmit=async e=>{e.preventDefault();let fd=new FormData(e.target);let id=fd.get('record_id');await J('/api/newspaper/records/'+encodeURIComponent(id),{method:'POST',body:fd});$('editCard').style.display='none';await loadRows()}
async function delRow(id){if(!confirm('Delete '+id+'?'))return;await J('/api/newspaper/records/'+encodeURIComponent(id)+'/delete',{method:'POST'});await loadRows()}
$('q').addEventListener('keydown',e=>{if(e.key==='Enter')loadRows()});loadRows();
</script></body></html>'''


def register(core):
    app = core.app
    engine = core.engine
    need_login = core.need_login
    page_role_or_redirect = core.page_role_or_redirect
    actor_name = core.actor_name
    router = APIRouter()

    def init_schema():
        with engine.begin() as c:
            for stmt in [x.strip() for x in SCHEMA.split(';') if x.strip()]:
                c.execute(text(stmt))

    @app.on_event("startup")
    def newspaper_startup():
        init_schema()

    @router.get('/newspaper', response_class=HTMLResponse)
    def newspaper_page(req: Request):
        if not page_role_or_redirect(req):
            return RedirectResponse('/login', status_code=303)
        return HTMLResponse(_html_page())

    @router.post('/api/newspaper/process')
    async def process_newspaper(req: Request, file: UploadFile = File(...), source_label: str = Form('Newspaper - Property'), high_accuracy: str = Form('true'), replace_same: str = Form('true')):
        need_login(req)
        if not getattr(core, 'GEMINI_API_KEY', ''):
            raise HTTPException(500, 'GEMINI_API_KEY is not configured on the server.')
        content = await file.read()
        if not content:
            raise HTTPException(400, 'Empty image')
        if len(content) > 20 * 1024 * 1024:
            raise HTTPException(413, 'Newspaper image is too large. Maximum 20 MB.')
        try:
            img = Image.open(io.BytesIO(content))
            img = ImageOps.exif_transpose(img).convert('RGB')
        except Exception:
            raise HTTPException(400, 'Unsupported or invalid image file')
        sha = hashlib.sha256(content).hexdigest()
        high = str(high_accuracy).lower() in {'1','true','yes','on'}
        replace = str(replace_same).lower() in {'1','true','yes','on'}
        client = getattr(core, 'client', None)
        if client is None:
            client = core.genai.Client(api_key=core.GEMINI_API_KEY)
        requested = getattr(core, 'NEWSPAPER_GEMINI_MODEL', None) or 'gemini-3.6-flash'
        filename = (file.filename or 'newspaper.jpg')[:500]
        mime = (file.content_type or 'image/jpeg')[:150]
        with engine.begin() as c:
            src = c.execute(text('SELECT id FROM pi_newspaper_sources WHERE source_hash=:h'), {'h': sha}).first()
            if src:
                sid = src[0]
                c.execute(text('UPDATE pi_newspaper_sources SET image_content=:b, original_filename=:f, mime_type=:m, source_label=:s, extraction_status=\'PROCESSING\', error_message=NULL, updated_at=NOW() WHERE id=:id'), {'b':content,'f':filename,'m':mime,'s':source_label,'id':sid})
                if replace:
                    c.execute(text("DELETE FROM pi_newspaper_properties WHERE source_id=:sid AND COALESCE(verification,'Unverified')!='Verified'"), {'sid':sid})
            else:
                sid = c.execute(text('INSERT INTO pi_newspaper_sources(source_hash,original_filename,mime_type,image_content,source_label,extraction_status) VALUES(:h,:f,:m,:b,:s,\'PROCESSING\') RETURNING id'), {'h':sha,'f':filename,'m':mime,'b':content,'s':source_label}).scalar_one()
        try:
            items, raw, active_model = extract_all(client, img, requested, high_accuracy=high)
            unique = _dedupe(items, source_label)
            inserted=0; skipped=0
            with engine.begin() as c:
                existing = {r[0] for r in c.execute(text('SELECT fingerprint FROM pi_newspaper_properties')).fetchall()}
                for row, fp in unique:
                    if fp in existing:
                        skipped += 1; continue
                    rid = 'NEWS-' + hashlib.sha1((fp + str(sid) + str(inserted)).encode()).hexdigest()[:10].upper()
                    c.execute(text('''INSERT INTO pi_newspaper_properties(record_id,source_id,fingerprint,lead_type,locality,area,configuration_details,price,agency_brand,contact_person,phone_numbers,notes,source,completeness,verification,team_member) VALUES(:rid,:sid,:fp,:lead_type,:locality,:area,:configuration_details,:price,:agency_brand,:contact_person,:phone_numbers,:notes,:source,:completeness,'Unverified','')'''), {'rid':rid,'sid':sid,'fp':fp,**row})
                    existing.add(fp); inserted += 1
                c.execute(text("UPDATE pi_newspaper_sources SET ai_model=:model, extraction_status='COMPLETED', extracted_records=:n, duplicate_records=:d, raw_ai_text=:raw, updated_at=NOW() WHERE id=:sid"), {'model':active_model,'n':inserted,'d':skipped,'raw':raw[:200000],'sid':sid})
            return {'status':'ok','inserted':inserted,'skipped':skipped,'model':active_model,'source_id':sid}
        except Exception as exc:
            with engine.begin() as c:
                c.execute(text("UPDATE pi_newspaper_sources SET extraction_status='FAILED', error_message=:e, updated_at=NOW() WHERE id=:sid"), {'e':str(exc)[:4000],'sid':sid})
            raise HTTPException(500, str(exc))

    @router.get('/api/newspaper/records')
    def get_records(req: Request, q: str = Query('')):
        need_login(req)
        sql='''SELECT record_id,date_captured,lead_type,locality,area,configuration_details,price,agency_brand,contact_person,phone_numbers,notes,source,completeness,verification,team_member,created_at,updated_at FROM pi_newspaper_properties'''
        params={}
        if q.strip():
            sql += " WHERE CONCAT_WS(' ',record_id,lead_type,locality,area,configuration_details,price,agency_brand,contact_person,phone_numbers,notes,source,completeness,verification,team_member) ILIKE :q"
            params['q']='%'+q.strip()+'%'
        sql += ' ORDER BY created_at DESC, locality ASC LIMIT 5000'
        with engine.begin() as c:
            rows=[_safe_json(r._mapping) for r in c.execute(text(sql),params)]
        return {'status':'ok','rows':rows}

    @router.post('/api/newspaper/records/{record_id}')
    async def update_record(record_id: str, req: Request):
        need_login(req)
        form = await req.form()
        allowed=['lead_type','locality','area','configuration_details','price','agency_brand','contact_person','phone_numbers','notes','source','completeness','verification','team_member']
        data={k:str(form.get(k,'')).strip() for k in allowed}
        fp=fingerprint(data)
        with engine.begin() as c:
            exists=c.execute(text('SELECT 1 FROM pi_newspaper_properties WHERE record_id=:id'),{'id':record_id}).first()
            if not exists: raise HTTPException(404,'Record not found')
            data.update({'id':record_id,'fp':fp})
            c.execute(text('''UPDATE pi_newspaper_properties SET fingerprint=:fp,lead_type=:lead_type,locality=:locality,area=:area,configuration_details=:configuration_details,price=:price,agency_brand=:agency_brand,contact_person=:contact_person,phone_numbers=:phone_numbers,notes=:notes,source=:source,completeness=:completeness,verification=:verification,team_member=:team_member,updated_at=NOW() WHERE record_id=:id'''),data)
        return {'status':'ok','record_id':record_id}

    @router.post('/api/newspaper/records/{record_id}/delete')
    def delete_record(record_id: str, req: Request):
        need_login(req)
        with engine.begin() as c:
            result=c.execute(text('DELETE FROM pi_newspaper_properties WHERE record_id=:id'),{'id':record_id})
        return {'status':'ok','deleted':result.rowcount}

    @router.get('/api/newspaper/export.xlsx')
    def export_xlsx(req: Request):
        need_login(req)
        with engine.begin() as c:
            rows=[dict(r._mapping) for r in c.execute(text('''SELECT date_captured AS "Date Captured", lead_type AS "Lead Type", locality AS "Locality", area AS "Area (Sq Yds)", configuration_details AS "Configuration/Details", price AS "Price", agency_brand AS "Agency/Brand", contact_person AS "Contact Person", phone_numbers AS "Phone Number(s)", notes AS "Notes", source AS "Source", completeness AS "Completeness" FROM pi_newspaper_properties ORDER BY created_at DESC, locality ASC'''))]
        wb=Workbook(); ws=wb.active; ws.title='Clean Database'
        headers=['Date Captured','Lead Type','Locality','Area (Sq Yds)','Configuration/Details','Price','Agency/Brand','Contact Person','Phone Number(s)','Notes','Source','Completeness']
        ws.append(headers)
        for r in rows: ws.append([r.get(h) for h in headers])
        for cell in ws[1]:
            cell.font=Font(name='Arial',bold=True,color='FFFFFF'); cell.fill=PatternFill('solid',fgColor='C0392B'); cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        ws.freeze_panes='A2'; widths=[14,20,28,20,58,22,25,22,25,42,34,24]
        for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
        for row in ws.iter_rows(min_row=2):
            for cell in row: cell.alignment=Alignment(vertical='top',wrap_text=True)
        ws.auto_filter.ref=ws.dimensions
        bio=io.BytesIO(); wb.save(bio)
        return Response(bio.getvalue(),media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',headers={'Content-Disposition':'attachment; filename=newspaper_clean_database.xlsx'})

    @router.get('/api/newspaper/source/{source_id}/image')
    def source_image(source_id: int, req: Request):
        need_login(req)
        with engine.begin() as c:
            row=c.execute(text('SELECT image_content,mime_type FROM pi_newspaper_sources WHERE id=:id'),{'id':source_id}).first()
        if not row: raise HTTPException(404,'Source image not found')
        return Response(row[0],media_type=row[1] or 'image/jpeg')

    app.include_router(router)
    return router
